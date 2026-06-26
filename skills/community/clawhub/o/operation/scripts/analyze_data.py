#!/usr/bin/env python3
"""
Excel Data Analyzer - Analyze data in an Excel file and generate a comprehensive report.

Usage:
    python analyze_data.py <input_file> [--sheet <sheet_name>] [--output <output_file>] [--focus <columns>]

Arguments:
    input_file       Path to the Excel file (.xlsx, .xls, .csv)

Options:
    --sheet          Sheet name or index (0-based). Default: first sheet
    --output         Output Excel file path. Default: <input>_analysis_report.xlsx
    --focus          Comma-separated column names/letters to focus analysis on (optional)

Examples:
    python analyze_data.py data.xlsx
    python analyze_data.py data.xlsx --sheet "Sheet1"
    python analyze_data.py data.xlsx --focus "姓名,年龄,部门"
    python analyze_data.py data.xlsx --output report.xlsx
"""

import argparse
import sys
import os
import json
from collections import Counter

try:
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Required packages not found. Installing dependencies...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "numpy", "-q"])
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter


def resolve_column(df, col_ref):
    """Resolve a column reference (name, letter, or 0-based index) to a column name."""
    if col_ref in df.columns:
        return col_ref
    col_upper = col_ref.upper()
    if col_upper.isalpha() and len(col_upper) <= 3:
        col_idx = 0
        for ch in col_upper:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        col_idx -= 1
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]
    try:
        col_idx = int(col_ref)
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]
    except ValueError:
        pass
    return None


def detect_column_type(series):
    """Detect the semantic type of a column."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return "empty"

    # Check numeric
    try:
        numeric = pd.to_numeric(non_null, errors='coerce')
        if numeric.notna().sum() / len(non_null) > 0.8:
            return "numeric"
    except Exception:
        pass

    # Check datetime
    try:
        datetime = pd.to_datetime(non_null, errors='coerce', infer_datetime_format=True)
        if datetime.notna().sum() / len(non_null) > 0.8:
            return "datetime"
    except Exception:
        pass

    # Check categorical (low cardinality relative to row count)
    unique_ratio = non_null.nunique() / len(non_null) if len(non_null) > 0 else 0
    if unique_ratio < 0.05 or non_null.nunique() <= 20:
        return "categorical"

    # High cardinality text
    if unique_ratio > 0.5:
        return "text"

    return "categorical"


def analyze_numeric(series, col_name):
    """Analyze a numeric column."""
    non_null = series.dropna()
    numeric = pd.to_numeric(non_null, errors='coerce').dropna()

    stats = {
        "column": col_name,
        "type": "numeric",
        "count": len(numeric),
        "null_count": series.isna().sum(),
        "null_pct": f"{series.isna().sum() / len(series) * 100:.1f}%" if len(series) > 0 else "0%",
        "mean": round(numeric.mean(), 4) if len(numeric) > 0 else None,
        "median": round(numeric.median(), 4) if len(numeric) > 0 else None,
        "std": round(numeric.std(), 4) if len(numeric) > 0 else None,
        "min": round(numeric.min(), 4) if len(numeric) > 0 else None,
        "max": round(numeric.max(), 4) if len(numeric) > 0 else None,
        "q25": round(numeric.quantile(0.25), 4) if len(numeric) > 0 else None,
        "q75": round(numeric.quantile(0.75), 4) if len(numeric) > 0 else None,
        "skewness": round(numeric.skew(), 4) if len(numeric) > 1 else None,
        "kurtosis": round(numeric.kurtosis(), 4) if len(numeric) > 1 else None,
        "zero_count": int((numeric == 0).sum()),
        "negative_count": int((numeric < 0).sum()),
        "outlier_count": None,
    }

    # Detect outliers using IQR
    if len(numeric) > 0:
        q1 = numeric.quantile(0.25)
        q3 = numeric.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        stats["outlier_count"] = int(((numeric < lower) | (numeric > upper)).sum())
        stats["outlier_lower"] = round(lower, 4)
        stats["outlier_upper"] = round(upper, 4)

    # Distribution bins for histogram
    if len(numeric) > 0:
        try:
            hist, bin_edges = np.histogram(numeric, bins=min(20, max(5, int(np.sqrt(len(numeric))))))
            stats["hist_bins"] = [
                {"range": f"{round(bin_edges[i], 2)}~{round(bin_edges[i+1], 2)}", "count": int(hist[i])}
                for i in range(len(hist))
            ]
        except Exception:
            stats["hist_bins"] = []

    # Top 10 values
    top_vals = numeric.value_counts().head(10)
    stats["top_values"] = [{"value": round(v, 4), "count": int(c)} for v, c in top_vals.items()]

    return stats


def analyze_categorical(series, col_name):
    """Analyze a categorical/text column."""
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]

    stats = {
        "column": col_name,
        "type": "categorical",
        "count": len(non_null),
        "null_count": int(series.isna().sum()),
        "null_pct": f"{series.isna().sum() / len(series) * 100:.1f}%" if len(series) > 0 else "0%",
        "unique_count": non_null.nunique(),
        "unique_ratio": f"{non_null.nunique() / len(non_null) * 100:.1f}%" if len(non_null) > 0 else "0%",
        "most_frequent": non_null.mode().iloc[0] if len(non_null) > 0 else None,
        "most_frequent_count": int(non_null.value_counts().iloc[0]) if len(non_null) > 0 else 0,
        "least_frequent": non_null.value_counts().index[-1] if len(non_null) > 0 and non_null.value_counts().iloc[-1] == 1 else None,
        "duplicate_count": int((non_null.value_counts() > 1).sum()),
        "single_occurrence": int((non_null.value_counts() == 1).sum()),
    }

    # Top values
    top_vals = non_null.value_counts().head(20)
    stats["top_values"] = [{"value": v, "count": int(c)} for v, c in top_vals.items()]

    # String length stats (if text-like)
    lengths = non_null.str.len()
    if len(lengths) > 0:
        stats["avg_length"] = round(lengths.mean(), 1)
        stats["min_length"] = int(lengths.min())
        stats["max_length"] = int(lengths.max())

    return stats


def analyze_datetime(series, col_name):
    """Analyze a datetime column."""
    non_null = series.dropna()
    try:
        dt_series = pd.to_datetime(non_null, errors='coerce').dropna()
    except Exception:
        dt_series = pd.Series(dtype='datetime64[ns]')

    stats = {
        "column": col_name,
        "type": "datetime",
        "count": len(dt_series),
        "null_count": int(series.isna().sum()),
        "null_pct": f"{series.isna().sum() / len(series) * 100:.1f}%" if len(series) > 0 else "0%",
        "earliest": str(dt_series.min()) if len(dt_series) > 0 else None,
        "latest": str(dt_series.max()) if len(dt_series) > 0 else None,
        "range_days": (dt_series.max() - dt_series.min()).days if len(dt_series) > 0 else None,
    }

    if len(dt_series) > 0:
        # Year distribution
        year_counts = dt_series.dt.year.value_counts().sort_index()
        stats["year_distribution"] = [{"year": int(y), "count": int(c)} for y, c in year_counts.items()]

        # Month distribution
        month_counts = dt_series.dt.month.value_counts().sort_index()
        stats["month_distribution"] = [{"month": int(m), "count": int(c)} for m, c in month_counts.items()]

        # Day of week distribution
        dow_counts = dt_series.dt.dayofweek.value_counts().sort_index()
        dow_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        stats["dow_distribution"] = [
            {"day": dow_names[int(d)], "count": int(c)} for d, c in dow_counts.items()
        ]

    return stats


def analyze_file(input_file, sheet=None, output=None, focus=None):
    """Analyze an Excel/CSV file and generate a comprehensive report."""

    # Read the file
    ext = os.path.splitext(input_file)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(input_file)
    else:
        sheet_param = sheet
        if sheet_param is not None:
            try:
                sheet_param = int(sheet_param)
            except ValueError:
                pass
        df = pd.read_excel(input_file, sheet_name=sheet_param if sheet_param is not None else 0)

    # Determine columns to analyze
    if focus:
        focus_cols = []
        for ref in focus.split(","):
            ref = ref.strip()
            resolved = resolve_column(df, ref)
            if resolved:
                focus_cols.append(resolved)
            else:
                print(f"Warning: Column '{ref}' not found. Skipping.")
        if not focus_cols:
            print("Error: No valid focus columns found. Available columns: {}".format(list(df.columns)))
            sys.exit(1)
        analyze_cols = focus_cols
    else:
        analyze_cols = list(df.columns)

    # Determine output path
    if output is None:
        base, ext = os.path.splitext(input_file)
        output = f"{base}_analysis_report.xlsx"

    # ============================
    # Perform Analysis
    # ============================

    all_results = {}
    for col in analyze_cols:
        series = df[col]
        col_type = detect_column_type(series)

        if col_type == "numeric":
            all_results[col] = analyze_numeric(series, col)
        elif col_type == "datetime":
            all_results[col] = analyze_datetime(series, col)
        elif col_type == "empty":
            all_results[col] = {"column": col, "type": "empty", "count": 0, "null_count": len(series)}
        else:
            all_results[col] = analyze_categorical(series, col)

    # ============================
    # Print Console Summary
    # ============================

    print(f"\n{'='*60}")
    print(f"  Excel Data Analysis Report")
    print(f"{'='*60}")
    print(f"  File:      {input_file}")
    print(f"  Rows:      {len(df)}")
    print(f"  Columns:   {len(df.columns)}")
    print(f"  Analyzed:  {len(analyze_cols)} columns")
    print(f"{'─'*60}")

    for col, stats in all_results.items():
        col_type = stats["type"]
        print(f"\n  [{col}] ({col_type})")
        if col_type == "numeric":
            print(f"    Count={stats['count']}, Null={stats['null_pct']}")
            print(f"    Mean={stats['mean']}, Median={stats['median']}, Std={stats['std']}")
            print(f"    Range: [{stats['min']}, {stats['max']}]")
            print(f"    IQR: [{stats.get('q25')}, {stats.get('q75')}], Outliers={stats.get('outlier_count')}")
        elif col_type == "categorical":
            print(f"    Count={stats['count']}, Unique={stats['unique_count']}, Null={stats['null_pct']}")
            print(f"    Most frequent: {stats['most_frequent']} ({stats['most_frequent_count']} times)")
        elif col_type == "datetime":
            print(f"    Count={stats['count']}, Null={stats['null_pct']}")
            print(f"    Range: {stats['earliest']} ~ {stats['latest']} ({stats.get('range_days')} days)")

    print(f"\n{'='*60}\n")

    # ============================
    # Generate Excel Report
    # ============================

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_dark = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_fill_orange = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    header_fill_purple = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_fill_teal = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sub_header_font = Font(bold=True, size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    num_alignment = Alignment(vertical="center", horizontal="right")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, row, col, text, fill=header_fill_dark):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = cell_alignment
        cell.border = thin_border

    def write_cell(ws, row, col, value, is_number=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = num_alignment if is_number else cell_alignment
        cell.border = thin_border

    # ── Sheet 1: Overview ──
    ws_overview = wb.active
    ws_overview.title = "Overview"

    overview_data = [
        ("Metric", "Value"),
        ("Source File", input_file),
        ("Sheet", str(sheet if sheet else "First sheet")),
        ("Total Rows", len(df)),
        ("Total Columns", len(df.columns)),
        ("Analyzed Columns", len(analyze_cols)),
        ("Total Cells", len(df) * len(df.columns)),
        ("Missing Cells", int(df.isna().sum().sum())),
        ("Missing %", f"{df.isna().sum().sum() / (len(df) * len(df.columns)) * 100:.1f}%"),
        ("Duplicate Rows", int(df.duplicated().sum())),
        ("Duplicate Row %", f"{df.duplicated().sum() / len(df) * 100:.1f}%" if len(df) > 0 else "0%"),
        ("Memory Usage", f"{df.memory_usage(deep=True).sum() / 1024:.1f} KB"),
    ]
    for r, (label, value) in enumerate(overview_data, 1):
        if r == 1:
            write_header(ws_overview, r, 1, label, header_fill_dark)
            write_header(ws_overview, r, 2, value, header_fill_dark)
        else:
            write_cell(ws_overview, r, 1, label)
            write_cell(ws_overview, r, 2, value, isinstance(value, (int, float)))
    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 35

    # Column type summary on overview
    type_row = len(overview_data) + 2
    write_header(ws_overview, type_row, 1, "Column", header_fill_teal)
    write_header(ws_overview, type_row, 2, "Detected Type", header_fill_teal)
    write_header(ws_overview, type_row, 3, "Non-Null", header_fill_teal)
    write_header(ws_overview, type_row, 4, "Null %", header_fill_teal)
    for i, col in enumerate(analyze_cols, 1):
        r = type_row + i
        stats = all_results[col]
        write_cell(ws_overview, r, 1, col)
        write_cell(ws_overview, r, 2, stats["type"])
        write_cell(ws_overview, r, 3, stats.get("count", 0), is_number=True)
        write_cell(ws_overview, r, 4, stats.get("null_pct", "0%"))
    ws_overview.column_dimensions['C'].width = 12
    ws_overview.column_dimensions['D'].width = 12

    # ── Sheet 2: Numeric Summary ──
    numeric_cols = [c for c in analyze_cols if all_results[c]["type"] == "numeric"]
    if numeric_cols:
        ws_numeric = wb.create_sheet(title="Numeric Summary")
        metrics = ["count", "null_count", "null_pct", "mean", "median", "std",
                    "min", "q25", "q75", "max", "skewness", "kurtosis",
                    "zero_count", "negative_count", "outlier_count"]
        metric_labels = ["Count", "Null Count", "Null %", "Mean", "Median", "Std Dev",
                         "Min", "Q1 (25%)", "Q3 (75%)", "Max", "Skewness", "Kurtosis",
                         "Zero Count", "Negative Count", "Outlier Count (IQR)"]

        write_header(ws_numeric, 1, 1, "Metric", header_fill_green)
        for j, col in enumerate(numeric_cols, 2):
            write_header(ws_numeric, 1, j, col, header_fill_green)
        for i, (metric, label) in enumerate(zip(metrics, metric_labels), 2):
            write_cell(ws_numeric, i, 1, label)
            for j, col in enumerate(numeric_cols, 2):
                val = all_results[col].get(metric)
                write_cell(ws_numeric, i, j, val if val is not None else "N/A", is_number=isinstance(val, (int, float)))
        ws_numeric.column_dimensions['A'].width = 22
        for j in range(2, len(numeric_cols) + 2):
            ws_numeric.column_dimensions[get_column_letter(j)].width = 16

        # Histogram per numeric column (top 5 only to keep it manageable)
        for col_idx, col in enumerate(numeric_cols[:5]):
            hist_data = all_results[col].get("hist_bins", [])
            if hist_data:
                ws_hist = wb.create_sheet(title=f"Hist-{col[:26]}")
                write_header(ws_hist, 1, 1, "Range", header_fill_green)
                write_header(ws_hist, 1, 2, "Count", header_fill_green)
                for i, bin_info in enumerate(hist_data, 2):
                    write_cell(ws_hist, i, 1, bin_info["range"])
                    write_cell(ws_hist, i, 2, bin_info["count"], is_number=True)
                ws_hist.column_dimensions['A'].width = 25
                ws_hist.column_dimensions['B'].width = 12

                # Add bar chart
                chart = BarChart()
                chart.title = f"Distribution: {col}"
                chart.x_axis.title = "Value Range"
                chart.y_axis.title = "Frequency"
                chart.style = 10
                chart.width = 20
                chart.height = 12
                data_ref = Reference(ws_hist, min_col=2, min_row=1, max_row=len(hist_data) + 1)
                cats_ref = Reference(ws_hist, min_col=1, min_row=2, max_row=len(hist_data) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.shape = 4
                ws_hist.add_chart(chart, "D2")

    # ── Sheet 3: Categorical Summary ──
    cat_cols = [c for c in analyze_cols if all_results[c]["type"] == "categorical"]
    if cat_cols:
        ws_cat = wb.create_sheet(title="Categorical Summary")
        metrics = ["count", "null_count", "null_pct", "unique_count", "unique_ratio",
                    "most_frequent", "most_frequent_count", "duplicate_count", "single_occurrence"]
        metric_labels = ["Count", "Null Count", "Null %", "Unique Values", "Unique Ratio",
                         "Most Frequent", "Freq Count", "Has Duplicates", "Single Occurrence"]

        write_header(ws_cat, 1, 1, "Metric", header_fill_orange)
        for j, col in enumerate(cat_cols, 2):
            write_header(ws_cat, 1, j, col, header_fill_orange)
        for i, (metric, label) in enumerate(zip(metrics, metric_labels), 2):
            write_cell(ws_cat, i, 1, label)
            for j, col in enumerate(cat_cols, 2):
                val = all_results[col].get(metric)
                write_cell(ws_cat, i, j, val if val is not None else "N/A")
        ws_cat.column_dimensions['A'].width = 18
        for j in range(2, len(cat_cols) + 2):
            ws_cat.column_dimensions[get_column_letter(j)].width = 20

        # Top values per categorical column
        for col in cat_cols:
            top_vals = all_results[col].get("top_values", [])
            if top_vals:
                ws_top = wb.create_sheet(title=f"Top-{col[:26]}")
                write_header(ws_top, 1, 1, "Value", header_fill_orange)
                write_header(ws_top, 1, 2, "Count", header_fill_orange)
                write_header(ws_top, 1, 3, "Percentage", header_fill_orange)
                total = sum(v["count"] for v in top_vals)
                for i, item in enumerate(top_vals, 2):
                    write_cell(ws_top, i, 1, str(item["value"]))
                    write_cell(ws_top, i, 2, item["count"], is_number=True)
                    pct = f"{item['count'] / all_results[col]['count'] * 100:.1f}%" if all_results[col]['count'] > 0 else "0%"
                    write_cell(ws_top, i, 3, pct)
                ws_top.column_dimensions['A'].width = 30
                ws_top.column_dimensions['B'].width = 12
                ws_top.column_dimensions['C'].width = 14

                # Add pie chart for top 10
                if len(top_vals) >= 2:
                    chart = PieChart()
                    chart.title = f"Top Values: {col}"
                    chart.style = 10
                    chart.width = 18
                    chart.height = 12
                    data_ref = Reference(ws_top, min_col=2, min_row=1, max_row=min(len(top_vals), 10) + 1)
                    cats_ref = Reference(ws_top, min_col=1, min_row=2, max_row=min(len(top_vals), 10) + 1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    ws_top.add_chart(chart, "E2")

    # ── Sheet 4: Datetime Summary ──
    dt_cols = [c for c in analyze_cols if all_results[c]["type"] == "datetime"]
    if dt_cols:
        ws_dt = wb.create_sheet(title="Datetime Summary")
        write_header(ws_dt, 1, 1, "Metric", header_fill_purple)
        for j, col in enumerate(dt_cols, 2):
            write_header(ws_dt, 1, j, col, header_fill_purple)
        metrics = ["count", "null_count", "null_pct", "earliest", "latest", "range_days"]
        metric_labels = ["Count", "Null Count", "Null %", "Earliest", "Latest", "Range (Days)"]
        for i, (metric, label) in enumerate(zip(metrics, metric_labels), 2):
            write_cell(ws_dt, i, 1, label)
            for j, col in enumerate(dt_cols, 2):
                val = all_results[col].get(metric)
                write_cell(ws_dt, i, j, val if val is not None else "N/A")
        ws_dt.column_dimensions['A'].width = 18
        for j in range(2, len(dt_cols) + 2):
            ws_dt.column_dimensions[get_column_letter(j)].width = 25

    # ── Sheet 5: Data Quality ──
    ws_quality = wb.create_sheet(title="Data Quality")
    quality_headers = ["Column", "Type", "Total Rows", "Non-Null", "Null Count", "Null %",
                       "Unique", "Duplicate Values", "Issues"]
    for j, h in enumerate(quality_headers, 1):
        write_header(ws_quality, 1, j, h, header_fill_dark)

    for i, col in enumerate(analyze_cols, 2):
        stats = all_results[col]
        issues = []
        if stats.get("null_count", 0) > 0:
            null_pct = float(stats.get("null_pct", "0%").replace("%", ""))
            if null_pct > 50:
                issues.append("High null rate")
            elif null_pct > 20:
                issues.append("Moderate null rate")
        if stats["type"] == "numeric" and stats.get("outlier_count", 0) > 0:
            outlier_pct = stats["outlier_count"] / max(stats["count"], 1) * 100
            if outlier_pct > 5:
                issues.append(f"Many outliers ({stats['outlier_count']})")
        if stats["type"] == "categorical" and stats.get("single_occurrence", 0) > stats.get("unique_count", 0) * 0.8:
            issues.append("Mostly unique values (possible ID/text)")
        if stats["type"] == "empty":
            issues.append("Empty column")

        write_cell(ws_quality, i, 1, col)
        write_cell(ws_quality, i, 2, stats["type"])
        write_cell(ws_quality, i, 3, len(df), is_number=True)
        write_cell(ws_quality, i, 4, stats.get("count", 0), is_number=True)
        write_cell(ws_quality, i, 5, stats.get("null_count", 0), is_number=True)
        write_cell(ws_quality, i, 6, stats.get("null_pct", "0%"))
        write_cell(ws_quality, i, 7, stats.get("unique_count", stats.get("count", "N/A")), is_number=True)
        if stats["type"] == "categorical":
            write_cell(ws_quality, i, 8, stats.get("duplicate_count", "N/A"), is_number=True)
        else:
            write_cell(ws_quality, i, 8, "N/A")
        write_cell(ws_quality, i, 9, "; ".join(issues) if issues else "OK")

    for j, width in enumerate([20, 12, 12, 12, 12, 10, 12, 16, 30], 1):
        ws_quality.column_dimensions[get_column_letter(j)].width = width

    # ── Sheet 6: Correlation (numeric only) ──
    if len(numeric_cols) >= 2:
        ws_corr = wb.create_sheet(title="Correlation")
        numeric_df = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        corr_matrix = numeric_df.corr()

        write_header(ws_corr, 1, 1, "", header_fill_teal)
        for j, col in enumerate(numeric_cols, 2):
            write_header(ws_corr, 1, j, col, header_fill_teal)
        for i, row_col in enumerate(numeric_cols, 2):
            write_header(ws_corr, i, 1, row_col, header_fill_teal)
            for j, col_col in enumerate(numeric_cols, 2):
                val = corr_matrix.loc[row_col, col_col]
                cell = ws_corr.cell(row=i, column=j, value=round(val, 4) if not pd.isna(val) else "N/A")
                cell.alignment = num_alignment
                cell.border = thin_border
                # Color code: strong correlation
                if not pd.isna(val):
                    if abs(val) > 0.7:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif abs(val) < 0.3:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws_corr.column_dimensions['A'].width = 20
        for j in range(2, len(numeric_cols) + 2):
            ws_corr.column_dimensions[get_column_letter(j)].width = 16

    wb.save(output)
    print(f"Analysis report saved to: {output}")

    return all_results


def main():
    parser = argparse.ArgumentParser(
        description="Analyze data in an Excel file and generate a comprehensive report."
    )
    parser.add_argument("input_file", help="Path to the Excel/CSV file")
    parser.add_argument("--sheet", default=None, help="Sheet name or index (default: first sheet)")
    parser.add_argument("--output", default=None, help="Output file path (default: <input>_analysis_report.xlsx)")
    parser.add_argument("--focus", default=None, help="Comma-separated column names to focus on (optional)")

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Error: File not found: {args.input_file}")
        sys.exit(1)

    analyze_file(
        input_file=args.input_file,
        sheet=args.sheet,
        output=args.output,
        focus=args.focus,
    )


if __name__ == "__main__":
    main()
