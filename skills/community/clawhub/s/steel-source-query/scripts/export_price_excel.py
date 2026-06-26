#!/usr/bin/env python3
"""
钢材价格查询并导出Excel
支持批量查询多个品种和地区
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# 添加脚本目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from scrape_price import query_price

# 默认查询配置
DEFAULT_QUERIES = [
    {"type": "螺纹钢", "region": "唐山"},
    {"type": "热轧板卷", "region": "唐山"},
    {"type": "冷轧板卷", "region": "上海"},
    {"type": "中厚板", "region": "唐山"},
    {"type": "H型钢", "region": "唐山"},
]


def query_multiple(queries: List[Dict]) -> List[Dict]:
    """批量查询价格"""
    results = []
    
    for query in queries:
        steel_type = query.get("type", "螺纹钢")
        region = query.get("region", "唐山")
        
        print(f"查询: {region} {steel_type}...")
        
        # 使用找钢网查询
        data = query_price("zhaogang", steel_type, region, cache_hours=2)
        
        if data and data.get("price"):
            results.append({
                "序号": len(results) + 1,
                "地区": region,
                "品种": steel_type,
                "价格(元/吨)": data["price"],
                "数据来源": data.get("source", "找钢网"),
                "查询时间": data.get("date", datetime.now().strftime("%Y-%m-%d")),
                "备注": ""
            })
        else:
            results.append({
                "序号": len(results) + 1,
                "地区": region,
                "品种": steel_type,
                "价格(元/吨)": "查询失败",
                "数据来源": "-",
                "查询时间": datetime.now().strftime("%Y-%m-%d"),
                "备注": "请手动查询"
            })
    
    return results


def export_to_excel(results: List[Dict], output_path: str):
    """导出到Excel，带样式"""
    if not EXCEL_SUPPORT:
        print("错误: 请先安装依赖: pip install pandas openpyxl")
        return False
    
    # 创建DataFrame
    df = pd.DataFrame(results)
    
    # 使用openpyxl创建带样式的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "钢材价格"
    
    # 写入标题
    headers = list(results[0].keys()) if results else []
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据
    for row_idx, record in enumerate(results, 2):
        for col_idx, key in enumerate(headers, 1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 价格列加粗
            if key == "价格(元/吨)" and isinstance(value, (int, float)):
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # 调整列宽
    column_widths = {
        "A": 8,   # 序号
        "B": 12,  # 地区
        "C": 15,  # 品种
        "D": 15,  # 价格
        "E": 15,  # 数据来源
        "F": 15,  # 查询时间
        "G": 20,  # 备注
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(results)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # 保存
    wb.save(output_path)
    print(f"✓ Excel文件已保存: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description="钢材价格查询并导出Excel")
    parser.add_argument("--types", help="钢材品种，逗号分隔，如：螺纹钢,热轧板卷")
    parser.add_argument("--region", default="唐山", help="地区")
    parser.add_argument("--output", default="钢材价格查询表.xlsx", help="输出文件路径")
    parser.add_argument("--default", action="store_true", help="使用默认查询配置")
    
    args = parser.parse_args()
    
    # 构建查询列表
    if args.default:
        queries = DEFAULT_QUERIES
    elif args.types:
        types = [t.strip() for t in args.types.split(",")]
        queries = [{"type": t, "region": args.region} for t in types]
    else:
        queries = DEFAULT_QUERIES
    
    print(f"开始查询 {len(queries)} 个品种...")
    print("-" * 40)
    
    # 查询价格
    results = query_multiple(queries)
    
    print("-" * 40)
    print(f"查询完成: {len([r for r in results if r['价格(元/吨)'] != '查询失败'])} 个成功")
    
    # 导出Excel
    export_to_excel(results, args.output)
    
    return results


if __name__ == "__main__":
    main()
