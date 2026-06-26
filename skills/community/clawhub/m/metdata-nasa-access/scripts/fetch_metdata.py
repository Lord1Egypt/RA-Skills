#!/usr/bin/env python3
"""
MetData-NASA-Access: Fetch NASA POWER meteorological data for wind/solar energy.
Generates Excel files with monthly, daily, and climatology data.
"""

import argparse
import json
import os
import sys
import time
import urllib.request
import urllib.error

# Default parameters for wind and solar energy analysis
DEFAULT_PARAMS = [
    # Solar/Radiation
    "ALLSKY_SFC_SW_DWN",  # 水平面总辐射量
    "CLRSKY_SFC_SW_DWN",  # 晴空辐射量
    "ALLSKY_TOA_SW_DWN",  # 大气顶层辐射量
    "ALLSKY_SFC_LW_DWN",  # 向下长波辐射
    "KT",                  # 日照清晰度指数
    "KT_CLEAR",            # 晴空清晰度指数
    # Temperature
    "T2M",                 # 2米气温
    "T2M_MAX",             # 2米最高温
    "T2M_MIN",             # 2米最低温
    "T10M",                # 10米气温
    "T10M_MAX",            # 10米最高温
    "T10M_MIN",            # 10米最低温
    "TS",                  # 地表温度
    "TS_MAX",              # 地表最高温
    "TS_MIN",              # 地表最低温
    # Humidity
    "RH2M",                # 2米相对湿度
    "QV2M",                # 2米绝对湿度
    "T2MDEW",              # 2米露点温度
    # Wind
    "WSC",                 # 修正风速(海拔校正)
    "WS50M",               # 50米风速
    "WS50M_MAX",           # 50米最大风速
    "WS50M_MIN",           # 50米最小风速
    "WS10M",               # 10米风速
    "WS10M_MAX",           # 10米最大风速
    "WS10M_MIN",           # 10米最小风速
    "WD50M",               # 50米风向
    "WD10M",               # 10米风向
    # Pressure
    "PSC",                 # 修正气压(海拔校正)
    "PS",                  # 地表气压
    # Other
    "PRECTOT",             # 降水量
    "TQV",                 # 大气可降水量
    "FROST_DAYS",          # 霜冻天数
]

REQUEST_DELAY = 2  # seconds between requests


def fetch_single_param(lat, lon, granularity, param, start_year=None, end_year=None):
    """Fetch a single parameter from NASA POWER API."""
    if granularity == "climatology":
        url = (
            f"https://power.larc.nasa.gov/api/temporal/climatology/point?"
            f"parameters={param}&community=RE"
            f"&longitude={lon}&latitude={lat}&format=JSON"
        )
    else:
        url = (
            f"https://power.larc.nasa.gov/api/temporal/{granularity}/point?"
            f"parameters={param}&community=RE"
            f"&longitude={lon}&latitude={lat}&format=JSON"
            f"&start={start_year}&end={end_year}"
        )

    req = urllib.request.Request(url)
    req.add_header("User-Agent", "MetData-NASA-Access/1.0")

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8")), None
    except urllib.error.HTTPError as e:
        return None, f"HTTP {e.code}"
    except Exception as e:
        return None, str(e)


def fetch_params(lat, lon, granularity, parameters, start_year, end_year):
    """Fetch all parameters one by one and merge results."""
    merged_feature = None
    success_count = 0
    failed_params = []

    total = len(parameters)
    for i, param in enumerate(parameters):
        print(f"  [{i+1}/{total}] {param}...", end=" ", flush=True)

        data, error = fetch_single_param(lat, lon, granularity, param, start_year, end_year)

        if data and "properties" in data:
            batch_params = data.get("properties", {}).get("parameter", {})
            if param in batch_params and batch_params[param]:
                if merged_feature is None:
                    merged_feature = data
                else:
                    merged_feature["properties"]["parameter"][param] = batch_params[param]
                success_count += 1
                print(f"OK ({len(batch_params[param])} points)")
            else:
                print("NO DATA")
                failed_params.append(f"{param} (no data)")
        else:
            print(f"FAIL ({error})")
            failed_params.append(f"{param} ({error})")

        time.sleep(REQUEST_DELAY)

    print(f"\n  Result: {success_count}/{total} params fetched successfully.")
    if failed_params:
        print(f"  Failed: {', '.join(failed_params)}")
    return merged_feature


def extract_table(feature, granularity):
    """Extract parameter data into [Date, param1, param2, ...] rows."""
    if not feature:
        return [], []

    properties = feature.get("properties", {})
    parameters = properties.get("parameter", {})
    if not parameters:
        return [], []

    param_names = list(parameters.keys())
    first_param = param_names[0]
    time_keys = sorted(parameters[first_param].keys())

    headers = ["日期"] + param_names
    rows = []
    for tk in time_keys:
        row = [tk]
        for pn in param_names:
            val = parameters[pn].get(tk)
            if val is not None and val != -999.0:
                row.append(round(val, 2))
            else:
                row.append(None)
        rows.append(row)

    return headers, rows


def write_excel(output_path, monthly_feature, daily_feature, climatology_feature):
    """Write data to Excel with openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    sheets_created = 0

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_sheet(ws, headers, rows, title):
        ws.title = title
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_data in rows:
                if col_idx - 1 < len(row_data) and row_data[col_idx - 1] is not None:
                    max_len = max(max_len, len(str(row_data[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Monthly
    if monthly_feature:
        headers, rows = extract_table(monthly_feature, "monthly")
        if headers and rows:
            ws = wb.active
            ws.title = "月度数据"
            write_sheet(ws, headers, rows, "月度数据")
            sheets_created += 1

    # Daily
    if daily_feature:
        headers, rows = extract_table(daily_feature, "daily")
        if headers and rows:
            ws = wb.create_sheet("日均数据")
            write_sheet(ws, headers, rows, "日均数据")
            sheets_created += 1

    # Climatology
    if climatology_feature:
        headers, rows = extract_table(climatology_feature, "climatology")
        if headers and rows:
            ws = wb.create_sheet("气候平均数据")
            write_sheet(ws, headers, rows, "气候平均数据")
            sheets_created += 1

    if sheets_created == 0:
        wb.remove(wb.active)

    wb.save(output_path)
    print(f"\nExcel saved: {output_path} ({sheets_created} sheets)")


def main():
    parser = argparse.ArgumentParser(description="Fetch NASA POWER meteorological data")
    parser.add_argument("--lat", type=float, required=True, help="Latitude")
    parser.add_argument("--lon", type=float, required=True, help="Longitude")
    parser.add_argument("--start", type=int, default=2016, help="Start year")
    parser.add_argument("--end", type=int, default=2017, help="End year")
    parser.add_argument("--output", type=str, default=None, help="Output Excel path")
    parser.add_argument("--params", type=str, default=None,
                        help="Comma-separated parameter list (default: wind+ solar defaults)")
    parser.add_argument("--granularity", type=str, default="all",
                        choices=["monthly", "daily", "climatology", "all"],
                        help="Data granularity (default: all)")
    args = parser.parse_args()

    lat, lon = args.lat, args.lon
    if not (-90 <= lat <= 90):
        print(f"Error: latitude {lat} out of range [-90, 90]", file=sys.stderr)
        sys.exit(1)
    if not (-180 <= lon <= 180):
        print(f"Error: longitude {lon} out of range [-180, 180]", file=sys.stderr)
        sys.exit(1)

    parameters = [p.strip() for p in args.params.split(",")] if args.params else DEFAULT_PARAMS

    print(f"Location: lat={lat}, lon={lon}")
    print(f"Period: {args.start} - {args.end}")
    print(f"Parameters: {len(parameters)}")
    print(f"Granularity: {args.granularity}")
    print()

    monthly_data = daily_data = climatology_data = None

    if args.granularity in ("monthly", "all"):
        print("=== Fetching Monthly Data ===")
        monthly_data = fetch_params(lat, lon, "monthly", parameters, args.start, args.end)
        time.sleep(REQUEST_DELAY)

    if args.granularity in ("daily", "all"):
        print("\n=== Fetching Daily Data ===")
        daily_data = fetch_params(lat, lon, "daily", parameters, args.start, args.end)
        time.sleep(REQUEST_DELAY)

    if args.granularity in ("climatology", "all"):
        print("\n=== Fetching Climatology Data ===")
        climatology_data = fetch_params(lat, lon, "climatology", parameters, None, None)

    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.expanduser("~/.openclaw/workspace/output/metdata")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(
            output_dir,
            f"metdata_nasa_{lat}_{lon}_{args.start}_{args.end}.xlsx"
        )

    print(f"\nGenerating Excel: {output_path}")
    write_excel(output_path, monthly_data, daily_data, climatology_data)
    print(f"\nDone! File saved to: {output_path}")


if __name__ == "__main__":
    main()
