#!/usr/bin/env python3
"""
Solar-Tilt-Calculator: Calculate solar radiation on tilted surfaces.
Based on Hay's anisotropic model and Duffie & Beckman formulas.

References:
- Duffie, J. A., & Beckman, W. A. (2013). Solar Engineering of Thermal Processes.
- 杨金焕等。不同方位倾斜面上太阳辐射量及最佳倾角的计算。上海交通大学学报，2002.
"""

import argparse
import math
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Constants
MONTH_NAMES = ["1 月", "2 月", "3 月", "4 月", "5 月", "6 月", "7 月", "8 月", "9 月", "10 月", "11 月", "12 月", "年度均值"]
MONTH_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "ANN"]

# Solar declination for each month (degrees) - approximate values for middle of month
# Source: Duffie & Beckman, Solar Engineering of Thermal Processes
SOLAR_DECLINATION = [-20.0, -12.0, 0.0, 12.0, 20.0, 23.4, 21.0, 12.0, 0.0, -12.0, -20.0, -23.0]

# Ground reflectance (typical value for most surfaces)
DEFAULT_ALBEDO = 0.2


def read_input_excel(input_path):
    """Read input Excel file and extract radiation data."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    
    if "气象数据" not in wb.sheetnames:
        print(f"Error: '气象数据' sheet not found in {input_path}", file=sys.stderr)
        sys.exit(1)
    
    ws = wb["气象数据"]
    
    data = {}
    for row in ws.iter_rows(min_row=4, max_row=16, values_only=True):
        if row[0] in MONTH_NAMES:
            month = row[0]
            data[month] = {
                "T2M": row[1],
                "T10M": row[2],
                "ALBEDO": row[3],
                "DNI": row[4],   # 直接法向辐射 (kWh/m²/天)
                "DIFF": row[5],  # 散射辐射 (kWh/m²/天)
                "DWN": row[6],   # 水平面总辐射 (kWh/m²/天)
            }
    
    # Extract latitude
    lat = None
    if "基本信息" in wb.sheetnames:
        ws_info = wb["基本信息"]
        for row in ws_info.iter_rows(min_row=3, max_row=9, values_only=True):
            if row[0] == "纬度" and row[1]:
                lat_str = str(row[1]).replace("°N", "").replace("°", "").strip()
                try:
                    lat = float(lat_str)
                except ValueError:
                    pass
    
    return data, lat


def sunset_hour_angle(latitude, declination):
    """Calculate sunset hour angle on horizontal surface (degrees)."""
    lat_rad = math.radians(latitude)
    dec_rad = math.radians(declination)
    
    cos_omega_s = -math.tan(lat_rad) * math.tan(dec_rad)
    cos_omega_s = max(-1.0, min(1.0, cos_omega_s))
    
    return math.degrees(math.acos(cos_omega_s))


def calculate_extra_terrestrial_radiation(latitude, declination):
    """Calculate extra-terrestrial radiation on horizontal surface (MJ/m²/day)."""
    lat_rad = math.radians(latitude)
    dec_rad = math.radians(declination)
    
    omega_s = sunset_hour_angle(latitude, declination)
    omega_s_rad = math.radians(omega_s)
    
    G_sc = 1367  # W/m² (solar constant)
    
    H0 = (24 * 3600 / math.pi) * G_sc * (
        omega_s_rad * math.sin(lat_rad) * math.sin(dec_rad) +
        math.cos(lat_rad) * math.cos(dec_rad) * math.sin(omega_s_rad)
    ) / 1000000
    
    return max(0.0, H0)


def calculate_Rb(latitude, declination, tilt, azimuth):
    """
    Calculate monthly mean ratio of direct radiation on tilted surface to horizontal.
    
    For equator-facing surfaces (γ = 0), uses Duffie & Beckman formula:
    Rb = [cos(φ-β)cosδsinωs + (ωsπ/180)sin(φ-β)sinδ] / [cosφcosδsinωs + (ωsπ/180)sinφsinδ]
    
    For non-zero azimuth, uses simplified approach with azimuth correction factor.
    
    Parameters:
    - latitude: degrees (positive for North)
    - declination: degrees (positive for summer)
    - tilt: degrees (0-90, 0 = horizontal)
    - azimuth: degrees (-180 to +180, 0 = equator-facing, negative = east, positive = west)
    
    Returns:
    - Rb: ratio (typically 0.5-2.0 for most conditions)
    """
    lat_rad = math.radians(latitude)
    dec_rad = math.radians(declination)
    tilt_rad = math.radians(tilt)
    az_rad = math.radians(azimuth)
    
    omega_s = sunset_hour_angle(latitude, declination)
    omega_s_rad = math.radians(omega_s)
    
    # For equator-facing surfaces (γ ≈ 0)
    if abs(azimuth) < 1.0:
        # Duffie & Beckman formula for monthly mean Rb
        numerator = (math.cos(lat_rad - tilt_rad) * math.cos(dec_rad) * math.sin(omega_s_rad) +
                     (omega_s * math.pi / 180) * math.sin(lat_rad - tilt_rad) * math.sin(dec_rad))
        
        denominator = (math.cos(lat_rad) * math.cos(dec_rad) * math.sin(omega_s_rad) +
                       (omega_s * math.pi / 180) * math.sin(lat_rad) * math.sin(dec_rad))
        
        if abs(denominator) < 1e-10:
            return 1.0
        
        Rb = numerator / denominator
    else:
        # For non-zero azimuth, use simplified approach
        # Based on empirical correlation: Rb(γ) ≈ Rb(0) × cos(γ) for small angles
        # This is a simplification; full calculation requires iterative solution
        
        # First calculate Rb for γ = 0
        numerator_0 = (math.cos(lat_rad - tilt_rad) * math.cos(dec_rad) * math.sin(omega_s_rad) +
                       (omega_s * math.pi / 180) * math.sin(lat_rad - tilt_rad) * math.sin(dec_rad))
        
        denominator = (math.cos(lat_rad) * math.cos(dec_rad) * math.sin(omega_s_rad) +
                       (omega_s * math.pi / 180) * math.sin(lat_rad) * math.sin(dec_rad))
        
        if abs(denominator) < 1e-10:
            Rb_0 = 1.0
        else:
            Rb_0 = numerator_0 / denominator
        
        # Apply azimuth correction factor
        # For |γ| < 45°, use cos(γ) approximation
        # For |γ| >= 45°, use more aggressive reduction
        abs_azimuth = abs(azimuth)
        
        if abs_azimuth < 45:
            # cos(γ) approximation
            azimuth_factor = math.cos(az_rad)
        else:
            # More aggressive reduction for large azimuth angles
            # Linear interpolation from cos(45°) at 45° to 0.5 at 90°
            azimuth_factor = math.cos(math.radians(45)) - (abs_azimuth - 45) * (math.cos(math.radians(45)) - 0.5) / 45
        
        Rb = Rb_0 * azimuth_factor
    
    # Clamp to reasonable range (0.0 to 3.0)
    # Typical values: 0.5-2.0 for most conditions
    return max(0.0, min(3.0, Rb))


def calculate_tilted_radiation(Hb, Hd, H0, Rb, tilt, H_total, albedo=DEFAULT_ALBEDO):
    """
    Calculate total radiation on tilted surface using Hay's anisotropic model.
    
    HT = Hb×Rb + Hd×[(Hb/H0)×Rb + 0.5×(1-Hb/H0)(1+cosβ)] + 0.5×ρ×H×(1-cosβ)
    
    Parameters:
    - Hb: Direct radiation on horizontal surface (MJ/m²/day)
    - Hd: Diffuse radiation on horizontal surface (MJ/m²/day)
    - H0: Extra-terrestrial radiation (MJ/m²/day)
    - Rb: Ratio of direct radiation (tilted/horizontal)
    - tilt: Tilt angle (degrees)
    - H_total: Total radiation on horizontal surface (MJ/m²/day)
    - albedo: Ground reflectance (default 0.2)
    
    Returns:
    - Hb_tilt, Hd_tilt, Hr, HT (all in MJ/m²/day)
    """
    # Direct radiation on tilted surface
    Hb_tilt = Hb * Rb
    
    # Diffuse radiation on tilted surface (Hay's anisotropic model)
    if H0 > 0:
        ratio = Hb / H0
    else:
        ratio = 0.0
    
    tilt_rad = math.radians(tilt)
    Hd_tilt = Hd * (ratio * Rb + 0.5 * (1 - ratio) * (1 + math.cos(tilt_rad)))
    
    # Ground reflected radiation
    Hr = 0.5 * albedo * H_total * (1 - math.cos(tilt_rad))
    
    # Total radiation
    HT = Hb_tilt + Hd_tilt + Hr
    
    return Hb_tilt, Hd_tilt, Hr, HT


def calculate_radiation_for_month(data, month, tilt, azimuth, latitude):
    """Calculate radiation on tilted surface for a given month."""
    if month == "年度均值":
        # Use average declination for annual mean
        dec = 0.0
    else:
        month_idx = MONTH_NAMES.index(month)
        dec = SOLAR_DECLINATION[month_idx]
    
    # Get input data (kWh/m²/天)
    DWN = data[month]["DWN"]   # 水平面总辐射
    DIFF = data[month]["DIFF"] # 散射辐射
    
    if DWN is None or DIFF is None:
        return None
    
    # Convert to MJ/m²/天
    H = DWN * 3.6      # 水平面总辐射
    Hd = DIFF * 3.6    # 散射辐射
    Hb = H - Hd        # 直接辐射（水平面）= 总辐射 - 散射辐射
    
    # Calculate extra-terrestrial radiation
    H0 = calculate_extra_terrestrial_radiation(latitude, dec)
    
    # Calculate Rb
    Rb = calculate_Rb(latitude, dec, tilt, azimuth)
    
    # Get albedo
    albedo = data[month]["ALBEDO"] if data[month]["ALBEDO"] else DEFAULT_ALBEDO
    
    # Calculate tilted surface radiation
    Hb_tilt, Hd_tilt, Hr, HT = calculate_tilted_radiation(Hb, Hd, H0, Rb, tilt, H, albedo)
    
    # Convert back to kWh/m²/天
    return {
        "Hb_tilt": Hb_tilt / 3.6,
        "Hd_tilt": Hd_tilt / 3.6,
        "Hr": Hr / 3.6,
        "HT": HT / 3.6,
        "H": H / 3.6,
        "Rb": Rb,
    }


def write_output_excel(output_path, data, tilt, azimuth, latitude, results):
    """Write calculation results to Excel."""
    wb = openpyxl.Workbook()
    
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    
    # Sheet 1: 基本信息
    ws1 = wb.active
    ws1.title = "基本信息"
    
    ws1.merge_cells("A1:C1")
    ws1.cell(row=1, column=1, value="倾斜面太阳辐射计算").font = title_font
    ws1.cell(row=1, column=1).fill = title_fill
    ws1.cell(row=1, column=1).alignment = center_align
    
    info_data = [
        ["项目", "值"],
        ["当地纬度", f"{latitude:.2f}°"],
        ["倾斜角度", f"{tilt:.1f}°"],
        ["方位角度", f"{azimuth:.1f}° (0°=正南，负=东，正=西)"],
        ["计算模型", "Hay 各向异性模型"],
        ["地面反射率", f"{DEFAULT_ALBEDO:.2f}"],
    ]
    
    for row_idx, row_data in enumerate(info_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
    
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 35
    
    # Sheet 2: 倾斜面辐射数据
    ws2 = wb.create_sheet("倾斜面辐射数据")
    
    ws2.merge_cells("A1:H1")
    ws2.cell(row=1, column=1, value="倾斜面各月日均辐射量").font = title_font
    ws2.cell(row=1, column=1).fill = title_fill
    ws2.cell(row=1, column=1).alignment = center_align
    
    headers = ["月份", "直射辐射\n(kWh/m²/天)", "散射辐射\n(kWh/m²/天)", 
               "地面反射\n(kWh/m²/天)", "倾斜面总辐射\n(kWh/m²/天)", 
               "水平面总辐射\n(kWh/m²/天)", "增益/损失\n(%)"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, month in enumerate(MONTH_NAMES, 4):
        ws2.cell(row=row_idx, column=1, value=month).border = thin_border
        ws2.cell(row=row_idx, column=1).alignment = center_align
        
        if month in results:
            r = results[month]
            gain_loss = (r["HT"] - r["H"]) / r["H"] * 100 if r["H"] > 0 else 0
            
            values = [r["Hb_tilt"], r["Hd_tilt"], r["Hr"], r["HT"], r["H"], gain_loss]
            for col_idx, value in enumerate(values, 2):
                cell = ws2.cell(row=row_idx, column=col_idx, value=round(value, 4))
                cell.border = thin_border
                cell.alignment = center_align
    
    for col_idx in range(1, 8):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Sheet 3: 参数说明
    ws3 = wb.create_sheet("参数说明")
    
    ws3.merge_cells("A1:C1")
    ws3.cell(row=1, column=1, value="参数说明与计算公式").font = title_font
    ws3.cell(row=1, column=1).fill = title_fill
    ws3.cell(row=1, column=1).alignment = center_align
    
    param_info = [
        ["参数", "说明", "单位"],
        ["倾斜面总辐射", "倾斜面上接收到的太阳辐射总量", "kWh/m²/天"],
        ["直射辐射", "直接辐射在倾斜面上的分量", "kWh/m²/天"],
        ["散射辐射", "天空散射辐射在倾斜面上的分量", "kWh/m²/天"],
        ["地面反射", "地面反射辐射", "kWh/m²/天"],
        ["增益/损失", "相对于水平面的辐射增益或损失百分比", "%"],
        ["", "", ""],
        ["计算公式", "", ""],
        ["HT = Hb×Rb + Hd×[(Hb/H0)×Rb + 0.5×(1-Hb/H0)(1+cosβ)] + 0.5×ρ×H×(1-cosβ)", "", ""],
        ["", "", ""],
        ["符号说明", "", ""],
        ["HT - 倾斜面总辐射", "Hb - 水平面直接辐射", "Hd - 水平面散射辐射"],
        ["H0 - 大气层外辐射", "Rb - 直接辐射比", "β - 倾角", "ρ - 地面反射率"],
        ["", "", ""],
        ["参考论文", "", ""],
        ["杨金焕等。不同方位倾斜面上太阳辐射量及最佳倾角的计算 [J].", "", ""],
        ["上海交通大学学报，2002, 36(7): 1032-1035.", "", ""],
    ]
    
    for row_idx, row_data in enumerate(param_info, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if row_idx in [3, 11, 16]:
                cell.font = header_font
                cell.fill = header_fill
    
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 20
    
    wb.save(output_path)
    print(f"Excel saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Calculate solar radiation on tilted surfaces")
    parser.add_argument("--input", type=str, required=True, help="Input Excel file path")
    parser.add_argument("--tilt", type=float, required=True, help="Tilt angle (0-90°)")
    parser.add_argument("--azimuth", type=float, default=0.0, help="Azimuth angle (-90° to +90°, 0°=equator, negative=east, positive=west)")
    parser.add_argument("--output", type=str, default=None, help="Output Excel path")
    args = parser.parse_args()
    
    if not (0 <= args.tilt <= 90):
        print(f"Error: tilt angle {args.tilt} out of range [0, 90]", file=sys.stderr)
        sys.exit(1)
    if not (-90 <= args.azimuth <= 90):
        print(f"Error: azimuth angle {args.azimuth} out of range [-90, 90]", file=sys.stderr)
        sys.exit(1)
    
    print(f"Input: {args.input}")
    print(f"Tilt angle: {args.tilt}°")
    print(f"Azimuth angle: {args.azimuth}°")
    print()
    
    data, lat = read_input_excel(args.input)
    if lat is None:
        print("Error: Could not extract latitude from input file", file=sys.stderr)
        sys.exit(1)
    
    print(f"Latitude: {lat:.2f}°")
    print(f"Months found: {len(data)}")
    print()
    
    results = {}
    for month in MONTH_NAMES:
        if month in data:
            result = calculate_radiation_for_month(data, month, args.tilt, args.azimuth, lat)
            if result:
                results[month] = result
                print(f"{month}: HT = {result['HT']:.4f} kWh/m²/天")
    
    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.expanduser("~/.openclaw/workspace/output/solar-tilt")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"tilt_{args.tilt}_az_{args.azimuth}.xlsx")
    
    print(f"\nGenerating Excel: {output_path}")
    write_output_excel(output_path, data, args.tilt, args.azimuth, lat, results)
    print(f"\nDone! File saved to: {output_path}")


if __name__ == "__main__":
    main()
