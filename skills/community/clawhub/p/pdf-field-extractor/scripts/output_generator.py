#!/usr/bin/env python3
"""
Output generation for PDF Field Extractor.
Supports Excel (.xlsx) and JSON output formats.
Also builds Feishu-compatible message content.
"""

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Default Styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
BORDER_SIDE = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_excel(
    results: List[Dict[str, Any]],
    output_path: str,
    sheet_name: str = "Sheet1",
    include_metadata: bool = True,
) -> str:
    """
    Generate an Excel file from extraction results.

    Args:
        results: List of result dictionaries, one per PDF.
        output_path: Path to save the Excel file.
        sheet_name: Name of the worksheet.
        include_metadata: Whether to include filename and timestamp metadata columns.

    Returns:
        Path to the generated Excel file.
    """
    if not results:
        # Create empty workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(output_path)
        return output_path

    # Collect all unique field keys across all results
    all_keys = set()
    for result in results:
        all_keys.update(result.keys())

    # Filter out internal fields
    internal_fields = {"_filename", "_timestamp", "_doc_type", "_page_count", "_is_scanned"}
    display_keys = sorted([k for k in all_keys if k not in internal_fields])

    # Build column headers
    if include_metadata:
        headers = ["文件名", "提取时间", "文档类型"] + display_keys
    else:
        headers = display_keys

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER

    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        if include_metadata:
            ws.cell(row=row_idx, column=1, value=result.get("_filename", "")).font = CELL_FONT
            ws.cell(row=row_idx, column=1).border = CELL_BORDER
            ws.cell(row=row_idx, column=1).alignment = CELL_ALIGNMENT

            ws.cell(row=row_idx, column=2, value=result.get("_timestamp", "")).font = CELL_FONT
            ws.cell(row=row_idx, column=2).border = CELL_BORDER
            ws.cell(row=row_idx, column=2).alignment = CELL_ALIGNMENT

            doc_type_display = {
                "invoice": "发票",
                "contract": "合同",
                "receipt": "收据",
                "bank_statement": "银行对账单",
                "license": "营业执照",
                "id_card": "身份证/护照",
                "express": "快递单",
                "generic": "通用文档",
            }.get(result.get("_doc_type", ""), result.get("_doc_type", ""))
            ws.cell(row=row_idx, column=3, value=doc_type_display).font = CELL_FONT
            ws.cell(row=row_idx, column=3).border = CELL_BORDER
            ws.cell(row=row_idx, column=3).alignment = CELL_ALIGNMENT

            col_offset = 4
        else:
            col_offset = 1

        for col_idx, key in enumerate(display_keys, start=col_offset):
            value = result.get(key, "")
            # Convert complex types to string
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = CELL_BORDER
            cell.alignment = CELL_ALIGNMENT

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row_idx in range(2, len(results) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return output_path


def generate_json(
    results: List[Dict[str, Any]],
    output_path: str,
    pretty: bool = True,
) -> str:
    """
    Generate a JSON file from extraction results.

    Args:
        results: List of result dictionaries.
        output_path: Path to save the JSON file.
        pretty: Whether to use pretty printing.

    Returns:
        Path to the generated JSON file.
    """
    # Clean up internal fields for output
    cleaned_results = []
    internal_fields = {"_filename", "_timestamp", "_doc_type", "_page_count", "_is_scanned"}

    for result in results:
        cleaned = {k: v for k, v in result.items() if k not in internal_fields}
        cleaned_results.append(cleaned)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        if pretty:
            json.dump(cleaned_results, f, ensure_ascii=False, indent=2)
        else:
            json.dump(cleaned_results, f, ensure_ascii=False)

    return output_path


def build_feishu_message(
    results: List[Dict[str, Any]],
    doc_type: str = "generic",
    max_items: int = 10,
) -> Dict[str, Any]:
    """
    Build Feishu interactive card message content for extraction results.

    Args:
        results: List of result dictionaries.
        doc_type: Document type for display.
        max_items: Maximum number of items to show in the summary.

    Returns:
        Feishu card content dictionary.
    """
    if not results:
        return {
            "msg_type": "text",
            "content": {"text": "未提取到任何数据"},
        }

    doc_type_display = {
        "invoice": "发票",
        "contract": "合同",
        "receipt": "收据",
        "bank_statement": "银行对账单",
        "license": "营业执照",
        "id_card": "身份证/护照",
        "express": "快递单",
        "generic": "通用文档",
    }.get(doc_type, doc_type)

    total = len(results)
    # Show summary of first few items
    items = []
    for i, result in enumerate(results[:max_items]):
        filename = result.get("_filename", f"文档{i+1}")
        # Show key fields summary
        key_fields = {k: v for k, v in result.items() if not k.startswith("_") and v}
        if key_fields:
            first_field = next(iter(key_fields.items()), None)
            if first_field:
                summary = f"**{first_field[0]}**: {str(first_field[1])[:50]}"
        else:
            summary = "无提取数据"
        items.append({"tag": "div", "text": {"tag": "lark_md", "content": f"📄 {filename}: {summary}"}})

    if total > max_items:
        items.append({"tag": "div", "text": {"tag": "lark_md", "content": f"_...还有 {total - max_items} 个文件_"}})

    content = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📊 PDF关键信息提取结果"},
                "subtitle": {"tag": "plain_text", "content": f"文档类型: {doc_type_display} | 共 {total} 个文件"},
            },
            "elements": items + [
                {"tag": "hr"},
                {
                    "tag": "note",
                    "elements": [
                        {"tag": "plain_text", "content": f"提取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"}
                    ],
                },
            ],
        },
    }

    return content


def build_feishu_text_message(results: List[Dict[str, Any]], doc_type: str = "generic") -> Dict[str, str]:
    """
    Build a simple Feishu text message for extraction results.

    Args:
        results: List of result dictionaries.
        doc_type: Document type.

    Returns:
        Dictionary with msg_type and content.
    """
    if not results:
        return {"msg_type": "text", "content": "未提取到任何数据"}

    lines = [f"✅ PDF关键信息提取完成（共 {len(results)} 个文件）\n"]

    doc_type_display = {
        "invoice": "发票",
        "contract": "合同",
        "receipt": "收据",
        "bank_statement": "银行对账单",
        "license": "营业执照",
        "id_card": "身份证/护照",
        "express": "快递单",
        "generic": "通用文档",
    }.get(doc_type, doc_type)

    lines.append(f"📋 文档类型: {doc_type_display}\n")

    for i, result in enumerate(results[:5], 1):
        filename = result.get("_filename", f"文档{i}")
        lines.append(f"\n📄 {i}. {filename}")
        for key, value in result.items():
            if key.startswith("_"):
                continue
            if value:
                value_str = str(value)[:100]
                lines.append(f"   • {key}: {value_str}")

    if len(results) > 5:
        lines.append(f"\n...还有 {len(results) - 5} 个文件")

    return {"msg_type": "text", "content": "\n".join(lines)}


def merge_results(
    results: List[Dict[str, Any]],
    source_filenames: Optional[List[str]] = None,
    doc_types: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Merge multiple extraction results with metadata.

    Args:
        results: List of field dictionaries.
        source_filenames: Optional list of source filenames.
        doc_types: Optional list of document types per result.

    Returns:
        List of enriched result dictionaries.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    merged = []

    for i, result in enumerate(results):
        enriched = dict(result)
        enriched["_filename"] = source_filenames[i] if source_filenames and i < len(source_filenames) else f"document_{i+1}"
        enriched["_timestamp"] = timestamp
        enriched["_doc_type"] = doc_types[i] if doc_types and i < len(doc_types) else "generic"
        merged.append(enriched)

    return merged
