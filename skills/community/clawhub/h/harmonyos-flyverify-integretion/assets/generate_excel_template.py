#!/usr/bin/env python3
"""
生成 FlyVerify HarmonyOS 配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_sheet_with_header(wb, sheet_name, headers, data):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 36
    ws.freeze_panes = "A2"
    return ws


def create_basic_info_sheet(wb):
    headers = ["配置项", "说明", "您的信息（必填/按需填写）"]
    data = [
        ["appKey", "MobTech FlyVerify 的 appKey，从 MobTech 控制台获取", ""],
        ["appSecret", "MobTech FlyVerify 的 appSecret，从 MobTech 控制台获取", ""],
        ["needPreVerify", "是否需要预取号：是/否", "是"],
        ["needVerify", "是否需要一键验证：是/否", "是"],
        ["needOneClickLogin", "是否需要一键登录：是/否", "否"],
        ["needCustomTheme", "是否需要授权页 UI 定制：是/否", "否"],
        ["preVerifyTimeoutMs", "预取号超时时间，单位毫秒；官方建议 3000-5000", "4000"],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)
    for row in [2, 3, 4, 5, 6, 7]:
        ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    return ws


def create_confirmed_config_sheet(wb):
    headers = ["配置项", "官方已确认内容", "你的项目值/备注"]
    data = [
        ["installCommand1", "ohpm install @zztsdk/zztcore", ""],
        ["installCommand2", "ohpm install @zztsdk/flyverify", ""],
        ["compatibleSdkVersion", "5.0.0(12)", ""],
        ["useNormalizedOHMUrl", "true", ""],
        ["permission1", "ohos.permission.INTERNET", ""],
        ["permission2", "ohos.permission.GET_NETWORK_INFO", ""],
        ["initApi", 'ZztSDK.init(context, "appKey", "appSecret")', ""],
        ["privacyApi", "ZztSDK.submitPolicyGrantResult(true)", ""],
        ["preVerifyApi", "FlyVerify.preVerify(callback)", ""],
        ["verifyApi", "FlyVerify.verify(uiContext, callback)", ""],
        ["themeApi", "FlyVerify.setTheme(theme)", ""],
    ]
    return create_sheet_with_header(wb, "官方确认项", headers, data)

def create_risk_sheet(wb):
    headers = ["项", "说明", "备注"]
    data = [
        ["文档歧义", "扩展业务文档中存在 ZztCustomController / MobCustomController 命名不一致，真正落代码前要先检查依赖导出", ""],
        ["服务端对接", "verify 成功后拿到的 token / opToken 应交服务端校验，不应直接当成本地登录态", ""],
        ["前置调用", "文档返回码 6119171 表示 verify 前未先调用 preVerify", ""],
        ["频控", "6119172 表示预取号或取号调用过于频繁", ""],
        ["网络权限", "一键登录依赖 INTERNET 和 GET_NETWORK_INFO 权限", ""],
    ]
    return create_sheet_with_header(wb, "风险提示", headers, data)


def create_instructions_sheet(wb):
    headers = ["说明项", "详细内容"]
    data = [
        [
            "填写步骤",
            "1. 在 MobTech 控制台获取 appKey 和 appSecret\n"
            "2. 如果还没有 appKey 和 appSecret，先前往 https://new.dashboard.mob.com/#/summary 申请\n"
            "3. 参考“官方确认项”核对你项目里的依赖、SDK 版本、权限和 API\n"
            "4. 填写功能开关和预取号超时\n"
            "5. 隐私弹窗“同意”回调、业务落点和 UI 定制细节由 Agent 在对话中逐步单独询问\n"
            "6. 保存后告诉 Agent“填好了”",
        ],
        [
            "已确认信息",
            "当前模板已经内置本地文档中确认过的依赖、权限、初始化、预取号、取号、主题定制入口和超时建议，不需要你再手抄正文。",
        ],
        [
            "必填项",
            "基础信息 Sheet 中的 appKey、appSecret、needPreVerify、needVerify、needOneClickLogin、needCustomTheme、preVerifyTimeoutMs 为必填；隐私回调位置、业务落点和 UI 定制细节不在 Excel 中填写，由 Agent 在对话中单独询问。",
        ],
        [
            "布尔字段规则",
            "支持填写：是/否、true/false、TRUE/FALSE；Agent 会统一转换为 true/false。",
        ],
        [
            "官方文档",
            "集成指南：https://www.mob.com/wiki/detailed?wiki=717&id=78\n"
            "秒验鸿蒙端合规指南：https://www.mob.com/wiki/detailed?wiki=754&id=78",
        ],
        [
            "重要提醒",
            "如果缺少 appKey 或 appSecret，本 skill 必须直接停止，不继续执行依赖安装、工程修改、联调或验收，并先前往 https://new.dashboard.mob.com/#/summary 申请；如果 verify 前未调用 preVerify，可能命中错误码 6119171；如果调用过于频繁，可能命中 6119172。",
        ],
    ]
    return create_sheet_with_header(wb, "填写说明", headers, data)


def main():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_confirmed_config_sheet(wb)
    create_risk_sheet(wb)
    create_instructions_sheet(wb)

    target_order = ["基础信息", "官方确认项", "风险提示", "填写说明"]
    for idx, name in enumerate(target_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    output_path = Path(__file__).resolve().parent / "FlyVerify_HarmonyOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
