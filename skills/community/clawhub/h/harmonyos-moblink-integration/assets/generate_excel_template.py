#!/usr/bin/env python3
"""
生成 MobLink 配置模板 Excel 文件（HarmonyOS NEXT 版）
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A2"
    return ws


def create_basic_info_sheet(wb):
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "MobTech 应用 Key，从 MobTech 官网注册应用获取", ""],
        ["appSecret", "MobTech 应用密钥，与 appKey 一同获取", ""],
        ["bundleName", "鸿蒙应用包名（bundleName），例如 com.example.app", ""]
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)
    for row in range(2, 6):
        ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    return ws


def create_moblink_sheet(wb):
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["uriScheme", "MobLink 后台配置的 scheme，对应 module.json5 中的 URI scheme", ""],
        ["host", "MobLink 后台配置的 host（AppLink Host），对应 module.json5 中的 URI host", ""],
        ["defaultScenePath", "默认场景路径，生成 Scene.path 示例时使用，如 /demo/a", "/demo/a"],
        ["exampleLinkPrefix", "用于拼接 mobID 的业务链接前缀，可留空，由业务分享链接决定", ""],
        ["restoreAbility", "承接场景还原的 Ability 类名，如 EntryAbility 或 com.example.EntryAbility", ""],
        ["abilityStagePath", "AbilityStage 类文件路径，用于设置全局场景还原监听器", ""],
    ]
    return create_sheet_with_header(wb, "MobLink配置", headers, data)


def create_privacy_sheet(wb):
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["privacyPolicyReady", "是否已有隐私政策并在首次冷启动弹窗征得用户同意：是/否", "否"],
        ["privacyCallbackLocation", "用户点击隐私政策同意按钮后的代码位置，如 EntryAbility.ts:onPrivacyAgreed", ""],
        ["useZztCustomController", "是否使用 App 数据采集主动控制器（ZztCustomController）：是/否", "否"],
        ["allowLocationData", "是否允许 MobLink 主动采集地理位置信息：是/否", "是"],
        ["allowDeviceIdData", "是否允许 MobLink 主动采集设备标识信息：是/否", "是"],
        ["allowAppListData", "是否允许 MobLink 主动采集应用列表信息：是/否", "是"],
        ["allowNetworkData", "是否允许 MobLink 主动采集网络状态信息：是/否", "是"],
    ]
    return create_sheet_with_header(wb, "隐私合规", headers, data)


def create_api_sheet(wb):
    headers = ["功能", "说明", "使用位置"]
    data = [
        ["ZztSDK.init", "ZztSDK 初始化，传入 context、appKey 和 appSecret，需在隐私提交之前调用", ""],
        ["ZztSDK.submitPolicyGrantResult", "提交隐私授权结果，用户同意隐私政策后调用", ""],
        ["mobLink.init", "MobLink 初始化，需在 ZztSDK 初始化和隐私提交之后调用", ""],
        ["mobLink.getMobID", "制作场景，传入 Scene.path 和 Scene.params，回调 mobID 后拼接到业务链接", ""],
        ["mobLink.setRestoreSceneListener", "设置全局场景还原监听器，官方建议放在 AbilityStage 中", ""],
        ["mobLink.updateNewWant", "在 Ability.onNewWant 中调用，防止 MobLink 无法还原", ""],
        ["mobLink.getSceneFromWant", "在 Ability 的 onCreate 或 onNewWant 中获取场景还原内容", ""],
    ]
    return create_sheet_with_header(wb, "API使用点", headers, data)


def create_permissions_sheet(wb):
    headers = ["权限名称", "类型", "说明", "您的信息"]
    data = [
        ["ohos.permission.INTERNET", "必需", "网络访问权限，MobLink 正常工作必需", ""],
        ["ohos.permission.APP_TRACKING_CONSENT", "建议", "用于产生设备 ID", ""],
        ["ohos.permission.GET_NETWORK_INFO", "建议", "用于判断网络类型，进行连接与数据传输优化", ""],
    ]
    return create_sheet_with_header(wb, "权限配置", headers, data)


def create_instructions_sheet(wb):
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 在 MobTech 后台创建应用，获取 appKey 和 appSecret\n2. 在 MobLink 后台完成 scheme、AppLink Host 等配置\n3. 填写基础信息与 MobLink 配置\n4. 根据业务需要填写场景还原 Ability 和隐私合规配置"],
        ["必填项", "基础信息 Sheet 中的 appKey、appSecret、bundleName 必填；uriScheme 和 host 建议从 MobLink 后台配置后填写"],
        ["ohpm集成", "官方文档说明使用 ohpm 安装 @zztsdk/zztcore 和 @zztsdk/moblink，最低 API version 12"],
        ["隐私合规", "ZztSDK.init() 内部会做隐私授权判断，在提交隐私授权同意前不会初始化业务；用户同意隐私政策后调用 ZztSDK.submitPolicyGrantResult(true)"],
        ["URI scheme", "需要在 module.json5 的 Ability skills 中配置 URI scheme 和 host，actions 不能为空"],
        ["场景还原", "全局监听器建议在 AbilityStage 中设置；Ability 需在 onNewWant 中调用 mobLink.updateNewWant"],
        ["官方文档", "入口：https://www.mob.com/wiki/detailed?wiki=661&id=34\nHarmonyOS NEXT 集成指南：https://mob.com/wiki/detailed?wiki=731&id=34\nMobLink 后台配置：https://mob.com/wiki/detailed?wiki=527&id=34\n鸿蒙端合规：https://mob.com/wiki/detailed?wiki=758&id=34\n常见问题：https://mob.com/wiki/detailed?wiki=530&id=34\n扩展业务功能：https://www.mob.com/wiki/detailed?wiki=730&id=34"],
    ]

    ws = wb.create_sheet(title="填写说明")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 70
    return ws


def main():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_moblink_sheet(wb)
    create_privacy_sheet(wb)
    create_api_sheet(wb)
    create_permissions_sheet(wb)
    create_instructions_sheet(wb)

    target_order = ["基础信息", "MobLink配置", "隐私合规", "API使用点", "权限配置", "填写说明"]
    for idx, name in enumerate(target_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    output_path = Path(__file__).resolve().parent / "MobLink_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
