"""
银行对账单转换主程序
将银行对账单文件转换为目标系统的标准导入模板
"""

import os
import sys
import argparse
import json
import logging
from typing import List, Optional, Dict, Any
from pathlib import Path
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.parser_factory import ParserFactory
from core.mapper_factory import MapperFactory
from core.data_structures import (
    BankStatementData, MappingResult, OutputConfig, BIPV5_FIELDS
)
from openpyxl.styles import Font, Alignment, PatternFill


# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class BankStatementConverter:
    """银行对账单转换器"""

    def __init__(self):
        self.parser = None
        self.mapper = None

    def convert(
        self,
        input_file: str,
        output_file: str,
        target_system: str = 'BIPV5',
        **kwargs
    ) -> Dict[str, Any]:
        """
        转换银行对账单文件

        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径
            target_system: 目标系统 (BIPV5, NCC, NC等)
            **kwargs:
                - account_code: 账户使用组织编码
                - bank_name: 银行名称
                - template_file: 模板文件路径

        Returns:
            转换结果字典
        """
        result = {
            'success': False,
            'input_file': input_file,
            'output_file': output_file,
            'target_system': target_system,
            'message': '',
            'details': {}
        }

        try:
            # Step 1: 检测文件格式并解析
            logger.info(f"正在解析文件: {input_file}")
            parser = ParserFactory.get_parser(input_file)
            source_data = parser.parse(input_file)

            result['details']['format'] = parser.format_name
            result['details']['account_number'] = source_data.header.account_number
            result['details']['transaction_count'] = len(source_data.transactions)

            if source_data.parse_errors:
                logger.warning(f"解析警告: {source_data.parse_errors}")

            # Step 2: 映射到目标系统
            logger.info(f"正在映射到 {target_system} 系统")
            mapper = MapperFactory.get_mapper(target_system)
            mapping_result = mapper.map(
                source_data,
                account_code=kwargs.get('account_code'),
                bank_name=kwargs.get('bank_name'),
                closing_balance=kwargs.get('closing_balance')
            )

            result['details']['mapped_count'] = mapping_result.successful_count
            result['details']['failed_count'] = mapping_result.failed_count

            # Step 3: 验证映射结果
            validation_result = mapper.validate(mapping_result)
            result['details']['validation'] = {
                'is_valid': validation_result.is_valid,
                'warnings': validation_result.warnings
            }

            # Step 4: 生成输出文件
            logger.info(f"正在生成输出文件: {output_file}")
            self._write_output(
                mapping_result,
                output_file,
                template_file=kwargs.get('template_file'),
                target_system=target_system,
                kwargs=kwargs
            )

            result['success'] = True
            result['message'] = f"成功转换 {mapping_result.successful_count} 条记录"
            logger.info(result['message'])

        except Exception as e:
            result['message'] = f"转换失败: {str(e)}"
            logger.error(result['message'])
            if hasattr(e, '__traceback__'):
                logger.exception(e)

        return result

    def _write_output(
        self,
        mapping_result: MappingResult,
        output_file: str,
        template_file: Optional[str] = None,
        target_system: str = 'BIPV5',
        kwargs: Optional[dict] = None,
    ):
        """写入输出文件

        委托给具体的 Mapper 写入。
        - BIPV5: 使用模板（assets/template/YYBIPV5_banktransaction.xlsx）
        - EAS_YXH/FINGARD/NSTC/YYNCC: 使用各自模板

        模板路径解析策略（兼容 clawhub 等不支持上传二进制模板的市场）：
        1. 优先使用显式传入的 template_file
        2. 其次通过 template_manager 在 assets/template 下查找
        3. 仍不存在则从互联网地址自动下载到本地
        """
        kwargs = kwargs or {}
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 解析模板路径：未显式传入时，通过 template_manager 解析（本地查找 + 自动下载）
        if not template_file:
            from core.template_manager import resolve_template
            template_file = resolve_template(target_system)

        # 获取 mapper（已经在 convert() 中获取过）
        from core.mapper_factory import MapperFactory
        try:
            mapper = MapperFactory.get_mapper(target_system)
        except ValueError:
            mapper = None

        if mapper is None or not hasattr(mapper, 'write_output'):
            # 兼容：未注册 mapper 时的兜底（生成空白模板）
            self._write_output_default(mapping_result, output_file, target_system)
            return

        # 调用 mapper 自带的 write_output
        if target_system.upper() == 'BIPV5':
            # BIPV5: 使用现有 BIPV5 模板
            try:
                mapper.write_output(
                    mapping_result,
                    output_file,
                    template_path=template_file
                )
            except Exception as e:
                logger.warning(f"BIPV5 模板写入失败，回退到默认: {e}")
                self._write_output_default(mapping_result, output_file, target_system)
        else:
            # EAS_YXH / FINGARD / NSTC / YYNCC: 调用 mapper.write_output
            mapper.write_output(
                mapping_result,
                output_file,
                template_path=template_file
            )

    def _write_output_default(
        self,
        mapping_result: MappingResult,
        output_file: str,
        target_system: str = 'BIPV5',
    ):
        """兜底输出：生成简单模板（不依赖外部模板）"""
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from decimal import Decimal

        wb = Workbook()
        all_fields = list(BIPV5_FIELDS.keys())
        chinese_names = [BIPV5_FIELDS[f]['cn_name'] for f in all_fields]
        ws = wb.active
        ws.title = "银行流水处理"

        for col_idx, field in enumerate(all_fields, 1):
            cell = ws.cell(row=4, column=col_idx, value=field)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for col_idx, cn_name in enumerate(chinese_names, 1):
            cell = ws.cell(row=6, column=col_idx, value=cn_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            if cn_name.startswith('*'):
                cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

        data_start_row = 10
        for row_idx, mapped_record in enumerate(mapping_result.mapped_records):
            excel_row = data_start_row + row_idx
            record = mapped_record.record
            for col_idx, field in enumerate(all_fields, 1):
                value = record.get(field, '')
                if isinstance(value, Decimal):
                    value = float(value)
                if value is None:
                    value = ''
                ws.cell(row=excel_row, column=col_idx, value=value)

        for col_idx in range(1, len(all_fields) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        wb.save(output_file)
        logger.info(f"文件已保存: {output_file}")


def batch_convert(
    input_dir: str,
    output_dir: str,
    target_system: str = 'BIPV5',
    pattern: str = '*.*',
    **kwargs
) -> List[Dict[str, Any]]:
    """
    批量转换目录中的银行对账单文件

    Args:
        input_dir: 输入目录
        output_dir: 输出目录
        target_system: 目标系统
        pattern: 文件匹配模式
        **kwargs: 额外参数

    Returns:
        转换结果列表
    """
    converter = BankStatementConverter()
    results = []

    input_path = Path(input_dir)
    output_path = Path(output_dir)

    # 创建输出目录
    output_path.mkdir(parents=True, exist_ok=True)

    # 查找输入文件
    for file_path in input_path.glob(pattern):
        if file_path.is_file():
            # 生成输出文件名
            output_file = output_path / f"{file_path.stem}_{target_system}.xlsx"

            result = converter.convert(
                str(file_path),
                str(output_file),
                target_system,
                **kwargs
            )
            results.append(result)

    return results


def batch_convert_merge(
    input_dir: str,
    output_file: str,
    target_system: str = 'BIPV5',
    pattern: str = '*',
    **kwargs
) -> Dict[str, Any]:
    """
    批量转换多个银行对账单文件并合并到单个Excel文件（基于模板）

    Args:
        input_dir: 输入目录
        output_file: 输出文件路径（未指定时默认输出到 assets/result/ 目录）
        target_system: 目标系统
        pattern: 文件匹配模式
        **kwargs: 额外参数

    Returns:
        合并转换结果
    """
    import openpyxl
    from openpyxl import load_workbook
    from decimal import Decimal
    import shutil

    converter = BankStatementConverter()
    input_path = Path(input_dir)

    # 获取模板文件路径 - 通过 template_manager 解析（本地查找 + 自动下载）
    # 这样可以兼容 clawhub 等不支持上传二进制模板的市场
    from core.template_manager import resolve_template
    template_file = resolve_template(target_system)
    if not template_file:
        # 兜底：直接拼接路径（仅在模板解析失败时使用，文件可能不存在）
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_file = os.path.join(base_dir, 'assets', 'template', 'YYBIPV5_banktransaction.xlsx')

    # 确保输出目录存在
    # 【重要修改】默认输出到输入文件所在目录，保持与原始数据同位置
    output_dir = os.path.dirname(output_file)
    if not output_dir:
        # 默认使用输入目录作为输出目录
        output_dir = os.path.abspath(input_dir)
        output_file = os.path.join(output_dir, os.path.basename(output_file) or f"merged_{target_system}.xlsx")
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 收集所有文件的数据
    all_mapping_results = []
    file_count = 0

    for file_path in sorted(input_path.iterdir()):
        if file_path.is_file() and not file_path.name.startswith('.'):
            try:
                logger.info(f"正在处理文件: {file_path.name}")
                parser = ParserFactory.get_parser(str(file_path))
                source_data = parser.parse(str(file_path))
                mapper = MapperFactory.get_mapper(target_system)
                mapping_result = mapper.map(
                    source_data,
                    account_code=kwargs.get('account_code'),
                    bank_name=kwargs.get('bank_name'),
                    closing_balance=kwargs.get('closing_balance')
                )
                all_mapping_results.append({
                    'file': str(file_path),
                    'mapping_result': mapping_result,
                    'count': len(mapping_result.mapped_records)
                })
                file_count += 1
                logger.info(f"  -> 提取 {len(mapping_result.mapped_records)} 条记录")
            except Exception as e:
                logger.warning(f"  处理失败: {file_path.name}, 错误: {str(e)}")

    if not all_mapping_results:
        return {
            'success': False,
            'message': '没有找到可处理的文件',
            'output_file': output_file
        }

    # 【重要修改】默认输出到输入文件所在目录，保持与原始数据同位置
    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if not output_dir:
        # 默认使用输入目录作为输出目录
        output_dir = os.path.abspath(input_dir)
        output_file = os.path.join(output_dir, os.path.basename(output_file) or f"merged_{target_system}.xlsx")
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 严格基于模板文件进行输出
    wb = None
    ws = None
    template_fields = []

    # 银行流水处理 sheet 的 UTF-8 字节（避免终端编码问题）
    TARGET_SHEET_BYTES = 'e993b6e8a18ce6b581e6b0b4e5a484e79086'

    if os.path.exists(template_file):
        try:
            # 复制模板文件到输出路径
            shutil.copy2(template_file, output_file)
            # 加载模板文件
            wb = load_workbook(output_file)

            # 找到银行流水处理 sheet（通过UTF-8字节比较）
            target_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name.encode('utf-8').hex() == TARGET_SHEET_BYTES:
                    target_sheet = sheet_name
                    break

            if target_sheet is None:
                # 如果找不到，尝试使用第二个sheet
                if len(wb.sheetnames) > 1:
                    target_sheet = wb.sheetnames[1]
                else:
                    raise Exception("找不到 银行流水处理 sheet")

            ws = wb[target_sheet]
            logger.info(f"已基于模板文件生成输出: {template_file}")

            # 从模板第4行获取字段顺序
            for col_idx in range(1, ws.max_column + 1):
                field_name = ws.cell(row=4, column=col_idx).value
                if field_name:
                    template_fields.append(field_name)

        except Exception as e:
            logger.error(f"模板文件处理失败: {e}")
            return {
                'success': False,
                'message': f'模板文件处理失败: {str(e)}',
                'output_file': output_file
            }
    else:
        logger.error(f"模板文件不存在: {template_file}")
        return {
            'success': False,
            'message': f'模板文件不存在: {template_file}',
            'output_file': output_file
        }

    # 写入数据（第10行开始）- 保留模板原有内容
    data_start_row = 10
    current_row = data_start_row
    total_records = 0
    global_seq_counter = 1  # 全局流水号计数器

    # 找到流水号字段的列索引
    seq_field_idx = None
    for col_idx, field_name in enumerate(template_fields, 1):
        if field_name == 'bank_seq_no':
            seq_field_idx = col_idx
            break

    for item in all_mapping_results:
        mapping_result = item['mapping_result']
        file_name = os.path.basename(item['file'])
        for mapped_record in mapping_result.mapped_records:
            record = mapped_record.record

            # 生成全局唯一的流水号 - 保留原始流水号，只在缺失时自动生成
            if seq_field_idx is not None:
                existing_seq = record.get('bank_seq_no', '')
                # 只有当原始流水号为空或无效时才自动生成
                if not existing_seq or str(existing_seq).strip().upper() in ['', 'NONREF', 'NONE', 'NULL']:
                    from datetime import datetime
                    timestamp = datetime.now().strftime('%y%m%d%H%M%S')
                    unique_seq = f"{timestamp}{global_seq_counter:06d}"
                    record['bank_seq_no'] = unique_seq
                global_seq_counter += 1

            # 按照模板的列顺序写入数据
            for col_idx, field_name in enumerate(template_fields, 1):
                value = record.get(field_name, '')
                if isinstance(value, Decimal):
                    value = float(value)
                if value is None:
                    value = ''
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
            total_records += 1

    # 保存文件
    wb.save(output_file)
    logger.info(f"合并文件已保存: {output_file}")

    return {
        'success': True,
        'message': f"成功合并 {file_count} 个文件，共 {total_records} 条记录",
        'output_file': output_file,
        'file_count': file_count,
        'record_count': total_records
    }


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='银行对账单转换工具 - 将银行对账单转换为BIP系统导入模板'
    )

    parser.add_argument(
        'input_file',
        nargs='?',
        help='输入文件路径 (支持 MT940 格式)'
    )

    parser.add_argument(
        '-o', '--output',
        dest='output_file',
        help='输出文件路径'
    )

    parser.add_argument(
        '-t', '--target',
        dest='target_system',
        default='BIPV5',
        choices=['BIPV5', 'EAS_YXH', 'FINGARD', 'NSTC', 'YYNCC', 'NCC', 'NC'],
        help='目标系统 (默认: BIPV5，可选: EAS_YXH=金蝶云星瀚, FINGARD=保融ATS, NSTC=九恒星司库, YYNCC=用友NCC)'
    )

    parser.add_argument(
        '-c', '--account-code',
        dest='account_code',
        help='账户使用组织编码'
    )

    parser.add_argument(
        '-b', '--bank-name',
        dest='bank_name',
        help='银行名称'
    )

    parser.add_argument(
        '--batch',
        action='store_true',
        help='批量转换模式'
    )

    parser.add_argument(
        '--merge',
        action='store_true',
        help='批量转换并合并到单个文件（不重复表头）'
    )

    parser.add_argument(
        '-i', '--input-dir',
        dest='input_dir',
        help='批量转换输入目录'
    )

    parser.add_argument(
        '--config',
        dest='config_file',
        help='映射配置文件路径'
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='详细输出'
    )

    args = parser.parse_args()

    # 设置日志级别
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # 检查必要参数
    if not args.input_file and not args.input_dir:
        parser.print_help()
        print("\n示例用法:")
        print("  python main.py bank_statement.sta -o output.xlsx")
        print("  python main.py bank_statement.sta -o output.xlsx -c ORG001 -b \"花旗银行\"")
        print("  python main.py --batch -i ./input -o ./output")
        print("  python main.py --merge -i ./input -o ./merged.xlsx  # 合并多个文件到单个Excel")
        sys.exit(1)

    converter = BankStatementConverter()
    kwargs = {
        'account_code': args.account_code,
        'bank_name': args.bank_name,
        'config_path': args.config_file
    }

    if args.batch:
        if not args.input_dir:
            print("错误: 批量模式需要指定输入目录 (-i/--input-dir)")
            sys.exit(1)

        # 【重要修改】当没有指定输出路径时，默认使用输入目录
        # 这样可以避免skill目录只读导致的问题
        if not args.output_file:
            # 默认使用输入目录作为输出目录
            output_dir = os.path.abspath(args.input_dir)
        else:
            output_dir = args.output_file
        results = batch_convert(
            args.input_dir,
            output_dir,
            args.target_system,
            **kwargs
        )

        print(f"\n批量转换完成，共处理 {len(results)} 个文件")
        for result in results:
            status = "成功" if result['success'] else "失败"
            print(f"  [{status}] {result['input_file']} -> {result['output_file']}")
            if result['message']:
                print(f"         {result['message']}")

    elif args.merge:
        if not args.input_dir:
            print("错误: 合并模式需要指定输入目录 (-i/--input-dir)")
            sys.exit(1)

        # 【重要修改】当没有指定输出路径时，默认使用输入目录
        # 这样可以避免skill目录只读导致的问题
        if not args.output_file:
            # 默认使用输入目录作为输出目录
            result_dir = os.path.abspath(args.input_dir)
            args.output_file = os.path.join(result_dir, f"merged_{args.target_system}.xlsx")
        else:
            # 如果用户只指定了文件名没有目录，使用输入目录
            if os.path.dirname(args.output_file) == '':
                result_dir = os.path.abspath(args.input_dir)
                args.output_file = os.path.join(result_dir, args.output_file)

        print(f"\n批量转换并合并模式:")
        print(f"  输入目录: {args.input_dir}")
        print(f"  输出文件: {args.output_file}")

        result = batch_convert_merge(
            args.input_dir,
            args.output_file,
            args.target_system,
            **kwargs
        )

        print(f"\n合并转换结果:")
        print(f"  状态: {'成功' if result['success'] else '失败'}")
        print(f"  消息: {result['message']}")
        if result.get('file_count'):
            print(f"  处理文件数: {result['file_count']}")
        if result.get('record_count'):
            print(f"  总记录数: {result['record_count']}")

    else:
        # 【重要修改】当没有指定输出路径时，默认使用源文件所在目录
        # 这样可以避免skill目录只读导致的问题
        if not args.output_file:
            input_path = Path(args.input_file)
            input_dir = os.path.dirname(os.path.abspath(args.input_file))
            args.output_file = os.path.join(input_dir, f"{input_path.stem}_{args.target_system}.xlsx")
        else:
            # 如果用户只指定了文件名没有目录，使用输入目录
            if os.path.dirname(args.output_file) == '':
                input_dir = os.path.dirname(os.path.abspath(args.input_file))
                args.output_file = os.path.join(input_dir, args.output_file)

        result = converter.convert(
            args.input_file,
            args.output_file,
            args.target_system,
            **kwargs
        )

        print(f"\n转换结果:")
        print(f"  状态: {'成功' if result['success'] else '失败'}")
        print(f"  输入: {result['input_file']}")
        print(f"  输出: {result['output_file']}")
        print(f"  消息: {result['message']}")

        if result['details']:
            print(f"\n详细信息:")
            for key, value in result['details'].items():
                print(f"  {key}: {value}")


if __name__ == '__main__':
    main()
