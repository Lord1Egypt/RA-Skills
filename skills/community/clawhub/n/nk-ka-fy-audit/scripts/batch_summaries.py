#!/usr/bin/env python3
"""KA返佣审计重算 - 综合报告脚本"""
import json, os, glob, gc
import openpyxl
import sqlite3

DATA_DIR = '/workspace/data'
JSON_PATH = '/tmp/month_summaries.json'
DB_PATH = '/workspace/data/ka_commission_audit.db'

# Load existing Excel data
existing = {}
if os.path.exists(JSON_PATH):
    with open(JSON_PATH) as f:
        existing = json.load(f)

all_months = sorted([d for d in os.listdir(DATA_DIR) if d.isdigit() and len(d) == 6])

# Process any remaining months
remaining = [m for m in all_months if m not in existing]
for month in remaining:
    patterns = [os.path.join(DATA_DIR, month, f'KA返佣{month}*.xlsx'),
                os.path.join(DATA_DIR, month, f'KA{month}*.xlsx')]
    found = None
    for p in patterns:
        fs = glob.glob(p)
        if fs: found = fs[0]; break
    if not found:
        existing[month] = {'month': month, 'error': 'file_not_found'}; continue
    
    wb = openpyxl.load_workbook(found, data_only=True)
    result = {'month': month, 'summary_total': 0.0, 'expansion_307_recalc': None}
    try:
        ws = wb['汇总']
        total = 0.0
        for ri in range(2, ws.max_row + 1):
            b = ws.cell(row=ri, column=2).value
            i = ws.cell(row=ri, column=9).value
            if b is not None and i is not None:
                try:
                    float(str(b).strip())
                    total += float(i)
                except: pass
        result['summary_total'] = round(total, 2)
    except: pass
    
    # expansion_307
    if month == '202307' and '拓展' in wb.sheetnames:
        ws = wb['拓展']
        total = 0.0
        for ri in range(4, ws.max_row + 1):
            b2 = ws.cell(row=ri, column=2).value
            f6 = ws.cell(row=ri, column=6).value
            if not b2 and not f6: continue
            if not str(b2).strip() if b2 else '': continue
            def fv(cn):
                v = ws.cell(row=ri, column=cn).value
                try: return float(v) if v is not None else 0
                except: return 0
            n = round(fv(9) * (fv(10) - fv(11) - fv(12)) * fv(13), 2)
            t = round(fv(15) * (fv(16) - fv(17) - fv(18)) * fv(19), 2)
            z = round(fv(21) * (fv(22) - fv(23) - fv(24)) * fv(25), 2)
            af = round(fv(27) * (fv(28) - fv(29) - fv(30)) * fv(31), 2)
            total += n + t + z + af
        result['expansion_307_recalc'] = round(total, 2)
    
    wb.close(); del wb; gc.collect()
    existing[month] = result
    with open(JSON_PATH, 'w') as f: json.dump(existing, f)

# ========== DB计算 ==========
print('\n=== 数据库审计重算 ===')
conn = sqlite3.connect(DB_PATH, timeout=60)
conn.row_factory = sqlite3.Row
c = conn.cursor()

db_results = {}
for month in all_months:
    total = 0.0
    
    # 1. 常规
    for row in c.execute("SELECT alipay_txn_amount, alipay_rebate_ratio, wechat_txn_amount, wechat_rebate_ratio FROM raw_records WHERE month=? AND sheet_name='常规'", (month,)):
        at, ar, wt, wr = row[0] or 0, row[1] or 0, row[2] or 0, row[3] or 0
        total += round(at * ar, 2) + round(wt * wr, 2)
    
    # 2. 中快系列 (LIKE匹配，覆盖中快7月中快8月中快等)
    total += c.execute("SELECT COALESCE(SUM(COALESCE(total_rebate,0)),0) FROM raw_records WHERE month=? AND sheet_name LIKE '中快%'", (month,)).fetchone()[0]
    
    # 3. sql_recalc (202602的SQL表不应计入)
    if month != '202602':
        total += c.execute("SELECT COALESCE(SUM(COALESCE(total_rebate,0)),0) FROM raw_records WHERE month=? AND calc_method='sql_recalc'", (month,)).fetchone()[0]
    
    # 4. 历史差额调整
    total += c.execute("SELECT COALESCE(SUM(COALESCE(total_rebate,0)),0) FROM raw_records WHERE month=? AND audit_note='历史差额调整'", (month,)).fetchone()[0]
    
    # 注意: 志华/拓展人志华的数据通过calc_method='expansion_308'在下一步中计入，不在这里重复计算
    # 202307拓展数据通过expansion_307_recals从Excel读取，不使用DB内字段(缺少渠道分润比例列)
    
    # 5. expansion data
    if month == '202307':
        if month in existing and existing[month].get('expansion_307_recalc'):
            total += existing[month]['expansion_307_recalc']
        else:
            total += c.execute("SELECT COALESCE(SUM(COALESCE(total_rebate,0)),0) FROM raw_records WHERE month=? AND calc_method='expansion_307'", (month,)).fetchone()[0]
    else:
        total += c.execute("SELECT COALESCE(SUM(COALESCE(total_rebate,0)),0) FROM raw_records WHERE month=? AND calc_method IN ('expansion_307','expansion_308')", (month,)).fetchone()[0]
    
    # 注意: 不重复加sheet_name='拓展'(202307)，因为expansion_307已通过步骤6处理
    
    db_results[month] = round(total, 2)

conn.close()

# ====== Report ======
print('\n' + '='*90)
print('  KA返佣审计重算综合报告')
print('='*90)
print(f'{"月份":<8} {"汇总表":>12} {"审计重算":>10} {"差额":>10} {"状态":<8}')
print('-'*90)

total_s = total_r = 0.0
pc = wc = fc = 0
failed = []

for month in all_months:
    info = existing.get(month, {})
    if 'error' in info:
        print(f'{month:<8} {"N/A":>12} {"N/A":>10} {"N/A":>10} {"❌":<8}')
        fc += 1; continue
    
    sm = info.get('summary_total', 0)
    rc = db_results.get(month, 0)
    df = round(sm - rc, 2)
    total_s += sm; total_r += rc
    
    if abs(df) <= 1:
        label = '✅通过'; pc += 1
    elif abs(df) <= 100:
        label = '⚠️偏差'; wc += 1
    else:
        label = '❌差异'; fc += 1; failed.append((month, df))
    
    print(f'{month:<8} {sm:>12.2f} {rc:>10.2f} {df:>+10.2f} {label:<8}')

print('-'*90)
print(f'{"总计":<8} {total_s:>12.2f} {total_r:>10.2f} {total_s-total_r:>+10.2f}')
print(f'通过: {pc}月 | 偏差(<100): {wc}月 | 差异(>100): {fc}月')

if failed:
    print(f'\n⚠️ 需关注月份:')
    for m, d in sorted(failed, key=lambda x: abs(x[1]), reverse=True)[:10]:
        print(f'  {m}: 差{d:+.2f}')

# Save report
report = {
    'total_summary': round(total_s, 2), 'total_recalc': round(total_r, 2),
    'total_diff': round(total_s - total_r, 2),
    'pass': pc, 'warn': wc, 'fail': fc,
    'months': {m: {'summary': existing.get(m,{}).get('summary_total',0),
                    'recalc': db_results.get(m,0),
                    'diff': round(existing.get(m,{}).get('summary_total',0) - db_results.get(m,0), 2)}
               for m in all_months}
}
with open('/tmp/audit_report.json', 'w') as f:
    json.dump(report, f, ensure_ascii=False)
