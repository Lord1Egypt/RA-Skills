#!/usr/bin/env python3
"""
Zabbix 监控报告：采集数据 → XLSX/CSV → 飞书消息 → 邮件
定时任务只运行这个脚本即可
"""
import os
import sys
import csv
import json
from dotenv import load_dotenv
load_dotenv()  # 加载 ~/.hermes/.env
import smtplib
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.error import URLError
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from collections import defaultdict

# ========== Zabbix 配置 ==========
ZABBIX_URL = os.environ.get("ZABBIX_URL", "")
ZABBIX_USER = os.environ.get("ZABBIX_USER", "")
ZABBIX_PASSWORD = os.environ.get("ZABBIX_PASSWORD", "")

if not ZABBIX_URL:
    print("[ERROR] ZABBIX_URL is not set. Please configure it in ~/.hermes/.env", file=sys.stderr)
    sys.exit(1)

if not ZABBIX_USER or not ZABBIX_PASSWORD:
    print("[ERROR] ZABBIX_USER / ZABBIX_PASSWORD are not set. Please configure them in ~/.hermes/.env", file=sys.stderr)
    sys.exit(1)

ITEMS_KEY = {
    "memory_avail": "vm.memory.size[available]",
    "memory_total": "vm.memory.size[total]",
    "cpu": "system.cpu.util",
}

EXCLUDE_GROUPS = {"Templates", "Templates/Applications", "Templates/Databases",
                  "Templates/Modules", "Templates/Network devices",
                  "Templates/Operating systems", "Templates/Server hardware",
                  "Templates/Virtualization", "Discovered hosts"}

HERMES_DIR = os.environ.get("HERMES_DIR", os.path.expanduser("~/.hermes"))
CSV_PATH  = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.csv")
XLSX_PATH = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.xlsx")
FEISHU_CHAT_ID = os.environ.get("FEISHU_CHAT_ID", "")


# ========== Zabbix API ==========

def api_call(method, params, auth=None):
    payload = {"jsonrpc": "2.0", "method": method, "params": params, "id": 1}
    if auth:
        payload["auth"] = auth
    data = json.dumps(payload).encode("utf-8")
    req = Request(ZABBIX_URL, data=data, headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except URLError as e:
        print(f"API 请求失败: {e}")
        sys.exit(1)
    if "error" in result:
        print(f"API 错误: {result['error']}")
        sys.exit(1)
    return result.get("result", [])


# ========== 数据采集 ==========

def fetch_all(auth):
    groups = api_call("hostgroup.get", {"output": ["groupid", "name"]}, auth=auth)
    groups = [g for g in groups if g["name"] not in EXCLUDE_GROUPS]
    group_ids = [g["groupid"] for g in groups]

    hosts = api_call("host.get", {
        "output": ["hostid", "name", "host"],
        "groupids": group_ids,
        "selectGroups": ["groupid", "name"],
    }, auth=auth)

    all_items = []
    for i in range(0, len(hosts), 100):
        batch_ids = [h["hostid"] for h in hosts[i:i+100]]
        items = api_call("item.get", {
            "output": ["itemid", "hostid", "key_", "lastvalue"],
            "hostids": batch_ids,
            "filter": {"key_": list(ITEMS_KEY.values())},
        }, auth=auth)
        all_items.extend(items)

    item_map = {(it["hostid"], it["key_"]): it.get("lastvalue", "")
                for it in all_items}

    rows = []
    for host in hosts:
        hid = host["hostid"]
        host_groups = host.get("groups", [])
        gnames = [g["name"] for g in host_groups]
        valid_gnames = [n for n in gnames if n not in EXCLUDE_GROUPS]
        if not valid_gnames:
            continue
        gname = valid_gnames[0]

        mem_total = item_map.get((hid, ITEMS_KEY["memory_total"]), "")
        mem_avail = item_map.get((hid, ITEMS_KEY["memory_avail"]), "")
        cpu = item_map.get((hid, ITEMS_KEY["cpu"]), "")

        mem_total_gb = float(mem_total) / (1024**3) if mem_total else None
        mem_avail_gb = float(mem_avail) / (1024**3) if mem_avail else None
        cpu_pct = float(cpu) if cpu else None
        mem_used_pct = (1 - float(mem_avail) / float(mem_total)) * 100 \
            if mem_avail and mem_total else None

        rows.append({
            "group": gname,
            "name": host["name"],
            "ip": host["host"],
            "mem_total_gb": mem_total_gb,
            "mem_avail_gb": mem_avail_gb,
            "mem_used_pct": mem_used_pct,
            "cpu_pct": cpu_pct,
        })

    return rows


# ========== XLSX 生成 ==========

def generate_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, text):
        cell.value = text
        cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

    def pct_color(pct, bg_base):
        if pct is None: return bg_base, "000000"
        if pct >= 80:   return "FF4444", "FFFFFF"
        if pct >= 60:   return "FFAA44", "000000"
        if pct >= 40:   return "FFEE88", "000000"
        return bg_base, "000000"

    group_rows = defaultdict(list)
    for r in rows:
        group_rows[r["group"]].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 总览 Sheet
    ws_ov = wb.create_sheet(title="总览")
    ws_ov.cell(row=1, column=1, value="服务器监控总览").font = Font(name="微软雅黑", bold=True, size=14)
    ws_ov.cell(row=1, column=1).alignment = Alignment(horizontal="left")
    ws_ov.row_dimensions[1].height = 24
    ws_ov.cell(row=2, column=1, value=f"采集时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws_ov.cell(row=2, column=1).font = Font(name="微软雅黑", size=10, color="666666")
    ws_ov.cell(row=3, column=1, value=f"共 {len(rows)} 台主机，{len(group_rows)} 个主机组")
    ws_ov.cell(row=3, column=1).font = Font(name="微软雅黑", size=10, color="666666")

    for col_idx, h in enumerate(["主机组", "主机数", "内存告警(≥80%)", "CPU告警(≥80%)"], 1):
        hdr(ws_ov.cell(row=5, column=col_idx), h)
    ws_ov.row_dimensions[5].height = 20

    for row_idx, (gname, gdata) in enumerate(sorted(group_rows.items()), start=6):
        mem_alarm = sum(1 for r in gdata
                        if r["mem_used_pct"] is not None and r["mem_used_pct"] >= 80)
        cpu_alarm = sum(1 for r in gdata
                        if r["cpu_pct"] is not None and r["cpu_pct"] >= 80)
        for col_idx, val in enumerate([gname, len(gdata), mem_alarm, cpu_alarm], 1):
            cell = ws_ov.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
            if col_idx == 3 and mem_alarm > 0:
                cell.fill = PatternFill("solid", fgColor="FF4444")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            elif col_idx == 4 and cpu_alarm > 0:
                cell.fill = PatternFill("solid", fgColor="FF4444")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")

    for col_idx, w in enumerate([24, 10, 16, 16], 1):
        ws_ov.column_dimensions[get_column_letter(col_idx)].width = w

    # 各主机组 Sheet
    col_defs = [("主机名", 32), ("IP", 18), ("内存总量(GB)", 14),
                ("内存可用(GB)", 14), ("内存占用率(%)", 14), ("CPU占用率(%)", 13)]

    for gname, gdata in sorted(group_rows.items()):
        ws = wb.create_sheet(title=gname[:31])
        ws.row_dimensions[1].height = 20
        for col_idx, (hdr_text, _) in enumerate(col_defs, 1):
            hdr(ws.cell(row=1, column=col_idx), hdr_text)

        gdata.sort(key=lambda x: (-(x["mem_used_pct"] or 0), -(x["cpu_pct"] or 0)))

        for row_idx, r in enumerate(gdata, start=2):
            bg = "EEF2FF" if row_idx % 2 == 0 else "FFFFFF"
            mem_bg, mem_fc = pct_color(r.get("mem_used_pct"), bg)
            cpu_bg, cpu_fc = pct_color(r.get("cpu_pct"), bg)

            vals = [
                (r["name"],           bg, "000000", None),
                (r["ip"],             bg, "000000", None),
                (r["mem_total_gb"],   bg, "000000", "0.0"),
                (r["mem_avail_gb"],   bg, "000000", "0.0"),
                (r["mem_used_pct"],   mem_bg, mem_fc, "0.0"),
                (r["cpu_pct"],        cpu_bg, cpu_fc, "0.0"),
            ]
            for col_idx, (val, cbg, cfc, fmt) in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if val is None:
                    cell.value = "N/A"
                else:
                    cell.value = val
                    if fmt:
                        cell.number_format = fmt
                cell.font = Font(name="微软雅黑", size=10, color=cfc)
                cell.fill = PatternFill("solid", fgColor=cbg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()

        for col_idx, (_, width) in enumerate(col_defs, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

    # ========== TOPN Sheet ==========
    topn = int(os.environ.get("TOPN", "50"))
    if topn > 0:
        ws_top = wb.create_sheet(title=f"TOP{topn}(内存+CPU)")
        ws_top.row_dimensions[1].height = 20
        for col_idx, (col_hdr, _) in enumerate(col_defs, start=1):
            hdr(ws_top.cell(row=1, column=col_idx), col_hdr)
        # 合并所有数据，按内存+CPU综合降序
        all_data = list(rows)
        all_data.sort(key=lambda x: (-(x["mem_used_pct"] or 0), -(x["cpu_pct"] or 0)))
        top_data = all_data[:topn]
        for row_idx, r in enumerate(top_data, start=2):
            bg = "EEF2FF" if row_idx % 2 == 0 else "FFFFFF"
            mem_bg, mem_fc = pct_color(r.get("mem_used_pct"), bg)
            cpu_bg, cpu_fc = pct_color(r.get("cpu_pct"), bg)
            row_vals = [
                (r["name"],         bg, "000000"),
                (r["ip"],           bg, "000000"),
                (r["mem_total_gb"],bg, "000000"),
                (r["mem_avail_gb"],bg, "000000"),
                (r["mem_used_pct"],mem_bg, mem_fc),
                (r["cpu_pct"],     cpu_bg, cpu_fc),
            ]
            for col_idx, (val, cbg, cfc) in enumerate(row_vals, start=1):
                if val is None:
                    display_val, fmt = "N/A", None
                elif col_idx in (3, 4):
                    display_val, fmt = f"{val:.1f}", '0.0'
                elif col_idx in (5, 6):
                    display_val, fmt = val, '0.0'
                else:
                    display_val, fmt = val, None
                cell = ws_top.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = Font(name="微软雅黑", size=10, color=cfc)
                cell.fill = PatternFill("solid", fgColor=cbg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="CCCCCC")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if fmt:
                    cell.number_format = fmt
        for col_idx, (_, width) in enumerate(col_defs, start=1):
            ws_top.column_dimensions[get_column_letter(col_idx)].width = width
        ws_top.column_dimensions["A"].width = 36
        ws_top.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"XLSX: {XLSX_PATH}")


# ========== CSV 生成（UTF-8-BOM，兼容 Windows Excel）==========

def generate_csv(rows):
    os.makedirs(os.path.dirname(CSV_PATH), exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["主机组", "主机名", "IP", "内存总量(GB)",
                         "内存可用(GB)", "内存占用率(%)", "CPU占用率(%)"])
        for r in rows:
            writer.writerow([
                r["group"],
                r["name"],
                r["ip"],
                f"{r['mem_total_gb']:.1f}" if r['mem_total_gb'] is not None else "N/A",
                f"{r['mem_avail_gb']:.1f}" if r['mem_avail_gb'] is not None else "N/A",
                f"{r['mem_used_pct']:.1f}" if r['mem_used_pct'] is not None else "N/A",
                f"{r['cpu_pct']:.1f}"      if r['cpu_pct'] is not None else "N/A",
            ])
    print(f"CSV: {CSV_PATH}")


# ========== 飞书消息 ==========

def build_feishu_summary(rows):
    """构建飞书摘要消息（Markdown格式）"""
    from collections import defaultdict
    group_rows = defaultdict(list)
    for r in rows:
        group_rows[r["group"]].append(r)

    # 重点关注：内存占用≥60% 或 CPU≥60%
    warnings = [r for r in rows
                if (r["mem_used_pct"] or 0) >= 60 or (r["cpu_pct"] or 0) >= 60]
    warnings.sort(key=lambda x: (-(x["mem_used_pct"] or 0), -(x["cpu_pct"] or 0)))

    lines = ["## 服务器监控报告", ""]
    lines.append(f"**采集时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"共 **{len(rows)}** 台主机，覆盖 **{len(group_rows)}** 个主机组")
    lines.append("")

    if warnings:
        lines.append("### ⚠ 重点关注（内存占用≥60% 或 CPU≥60%）")
        lines.append("")
        lines.append("| 主机名 | 主机组 | 内存占用率(%) | CPU占用率(%) |")
        lines.append("|---|---|---|---|")
        for r in warnings[:20]:   # 最多显示20条
            lines.append(f"| {r['name']} | {r['group']} | "
                         f"{r['mem_used_pct']:.1f} | {r['cpu_pct']:.1f} |")
        if len(warnings) > 20:
            lines.append(f"...（共 {len(warnings)} 台，详见附件）")
        lines.append("")
    else:
        lines.append("### ✅ 全部正常（无告警主机）")
        lines.append("")

    lines.append(f"完整数据：`{CSV_PATH}`")
    return "\n".join(lines)


# ========== 邮件发送 ==========

def load_env():
    env_path = os.path.join(os.environ.get("HERMES_DIR", os.path.expanduser("~/.hermes")), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    os.environ[k] = v.strip()


def send_email(subject, html_body, attachments=None):
    load_env()
    smtp_host  = os.environ.get("SMTP_HOST", "")
    smtp_port  = os.environ.get("SMTP_PORT", "465")
    smtp_from  = os.environ.get("SMTP_FROM", "")
    smtp_token = os.environ.get("SMTP_TOKEN", "")
    target     = os.environ.get("TARGET_EMAIL", "")

    if not all([smtp_host, smtp_from, smtp_token, target]):
        print("邮件配置不完整，跳过")
        return

    msg = MIMEMultipart()
    msg["From"]    = smtp_from
    msg["To"]      = target
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html", "utf-8"))

    for fpath in (attachments or []):
        if os.path.exists(fpath):
            with open(fpath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part["Content-Disposition"] = f"attachment; filename={os.path.basename(fpath)}"
                msg.attach(part)

    try:
        if smtp_port == "465":
            with smtplib.SMTP_SSL(smtp_host, int(smtp_port)) as s:
                s.login(smtp_from, smtp_token)
                s.sendmail(smtp_from, target, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, int(smtp_port)) as s:
                s.starttls()
                s.login(smtp_from, smtp_token)
                s.sendmail(smtp_from, target, msg.as_string())
        print(f"邮件已发送: {target}")
    except Exception as e:
        print(f"邮件发送失败: {e}")


def build_html_body(rows):
    group_rows = defaultdict(list)
    for r in rows:
        group_rows[r["group"]].append(r)

    html = f"""<html><body>
    <h2>服务器监控报告</h2>
    <p><b>采集时间：</b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    <p><b>共 {len(rows)} 台主机，{len(group_rows)} 个主机组</b></p>"""

    for gname, gdata in sorted(group_rows.items()):
        html += f"<h3>{gname} ({len(gdata)} 台)</h3>"
        html += ("<table border='1' cellpadding='4' cellspacing='0' "
                 "style='border-collapse:collapse;font-size:12px;'>")
        html += ("<tr bgcolor='#4472C4' style='color:white;'>"
                 "<th>主机名</th><th>IP</th>"
                 "<th>内存总量(GB)</th><th>内存可用(GB)</th>"
                 "<th>内存占用率(%)</th><th>CPU占用率(%)</th></tr>")

        for i, r in enumerate(gdata):
            bg = "#EEF2FF" if i % 2 == 0 else "#FFFFFF"
            mp = r["mem_used_pct"] or 0
            cp = r["cpu_pct"] or 0

            ms = ("background:#FF4444;color:white;" if mp >= 80 else
                  "background:#FFAA44;" if mp >= 60 else
                  "background:#FFEE88;" if mp >= 40 else "")
            cs = ("background:#FF4444;color:white;" if cp >= 80 else
                  "background:#FFAA44;" if cp >= 60 else
                  "background:#FFEE88;" if cp >= 40 else "")

            html += f"<tr bgcolor='{bg}'>"
            html += f"<td>{r['name']}</td><td>{r['ip']}</td>"
            html += f"<td>{r['mem_total_gb']:.1f}</td>" if r['mem_total_gb'] else "<td>N/A</td>"
            html += f"<td>{r['mem_avail_gb']:.1f}</td>" if r['mem_avail_gb'] else "<td>N/A</td>"
            html += f"<td style='{ms}'>{r['mem_used_pct']:.1f}</td>" if r['mem_used_pct'] else "<td>N/A</td>"
            html += f"<td style='{cs}'>{r['cpu_pct']:.1f}</td>" if r['cpu_pct'] else "<td>N/A</td>"
            html += "</tr>"
        html += "</table><br/>"

    html += "</body></html>"
    return html


# ========== 主流程 ==========

def main():
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始巡检...")

    # 1. Zabbix 登录
    auth = api_call("user.login", {
        "user": ZABBIX_USER,
        "password": ZABBIX_PASSWORD,
    })
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 登录成功")

    # 2. 采集数据
    rows = fetch_all(auth)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 采集完成: {len(rows)} 台主机")

    # 3. 生成文件
    generate_csv(rows)
    generate_xlsx(rows)

    # 4. 飞书消息（通过 Hermes send_message API 发送）
    summary = build_feishu_summary(rows)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 飞书摘要:\n{summary[:500]}")

    # 5. 邮件
    subject = f"【监控报告】服务器巡检 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    html = build_html_body(rows)
    atts = [f for f in [XLSX_PATH, CSV_PATH] if os.path.exists(f)]
    send_email(subject, html, atts)

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 全部完成!")


if __name__ == "__main__":
    main()
