"""
银行流水 Excel 格式转换器
将不同银行导出的流水 Excel 统一转换为监管流水导入模板格式。
"""

import sys
import os
import re
import glob
import io
import csv
import zipfile
import traceback
import tempfile
import shutil
import xml.etree.ElementTree as ET
from collections import OrderedDict
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from field_mapping import (
    FIELD_ALIASES, HEADER_KEYWORDS, DATE_COLS, AMOUNT_COLS, TEMPLATE_COLUMNS
)


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk", "utf-16", "utf-16le", "utf-16be")
SUPPORTED_INPUT_EXTENSIONS = {".xlsx", ".xls", ".csv"}


def clean_header(header):
    """去掉银行导出表头中的英文说明，如 交易日期[ Transaction Date ]。"""
    text = re.sub(r'\[.*?\]', '', str(header or "")).strip()
    return text.replace("（", "(").replace("）", ")")


def is_standard_template_header(row):
    """判断一行是否已经是监管 9 列标准模板表头。"""
    normalized = [clean_header(c) for c in row]
    return normalized[:len(TEMPLATE_COLUMNS)] == TEMPLATE_COLUMNS


def find_header_row(rows, max_scan=25):
    """
    在前 max_scan 行中寻找包含表头关键词的行。
    表头行的特征：包含关键词，且同一行至少有 4 个非空单元格。
    """
    best_idx = None
    best_score = -1
    for i, row in enumerate(rows[:max_scan]):
        non_empty = sum(1 for c in row if c and c != "None")
        if non_empty < 4:
            continue
        if is_standard_template_header(row):
            return i
        cleaned = [clean_header(c) for c in row]
        keyword_hits = sum(1 for c in cleaned if any(kw in c for kw in HEADER_KEYWORDS))
        field_hits = sum(
            1
            for c in cleaned
            for aliases in FIELD_ALIASES.values()
            if any(alias == c for alias in aliases)
        )
        score = keyword_hits * 3 + field_hits
        if score > best_score and keyword_hits:
            best_idx = i
            best_score = score
    return best_idx


def find_mapped_column(headers, target_field):
    """根据 FIELD_ALIASES 找到某个模板字段对应的源列。"""
    aliases = FIELD_ALIASES.get(target_field, [])
    for idx, header in enumerate(headers):
        stripped = clean_header(header)
        if any(alias == stripped or alias in stripped for alias in aliases):
            return idx
    return None


def is_data_row(row, headers):
    """判断是否为有效数据行（非空行、非汇总行）"""
    non_empty = sum(1 for c in row if c and c != "None")
    if non_empty < 3:
        return False
    if is_standard_template_header(row):
        return False
    first_cell = str(row[0]).strip() if row else ""
    if is_standard_template_header(headers) and first_cell and not extract_date(first_cell):
        return False
    text = "".join(str(c) for c in row)
    if any(kw in text for kw in ["合计", "小计", "累计", "笔数"]):
        return False
    date_col = find_mapped_column(headers, "日期")
    if date_col is not None:
        date_value = row[date_col] if date_col < len(row) else ""
        if not extract_date(date_value):
            return False
    return True


def _format_date_parts(year, month, day):
    """格式化日期部件，非法日期返回空字符串。"""
    try:
        parsed = date(int(year), int(month), int(day))
    except (TypeError, ValueError):
        return ""
    return parsed.strftime("%Y-%m-%d")


def extract_date(value):
    """从日期值中提取日期部分 YYYY-MM-DD"""
    if not value or value == "None":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        normalized = _format_date_parts(m.group(1), m.group(2), m.group(3))
        return normalized or s
    m = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", s)
    if m:
        normalized = _format_date_parts(m.group(1), m.group(2), m.group(3))
        return normalized or s
    m = re.match(r"^(\d{4})(\d{2})(\d{2})(?:\d{6})?(?:\D|$)", s)
    if m:
        normalized = _format_date_parts(m.group(1), m.group(2), m.group(3))
        return normalized or s
    return ""


def extract_amount(value):
    """清理金额值（去除逗号、提取数字）"""
    if not value or value == "None":
        return ""
    s = str(value).strip()
    # 银行导出常用 "-" 表示无发生额，视作空
    if s in ("-", "—", "--"):
        return ""
    # 从中文大写金额中提取数字：如 "贰拾捌元整,￥28.00元" → "28.00"
    m = re.search(r"[￥¥]([\d,]+\.?\d*)", s)
    if m:
        return m.group(1).replace(",", "")
    s = s.replace(",", "")
    if not s:
        return ""
    try:
        float(s)
        return s
    except ValueError:
        return s


def positive_amount(value):
    """收入/支出已由方向列判断时，输出金额使用正数。"""
    amount = extract_amount(value)
    if amount.startswith("-"):
        return amount[1:]
    return amount


def date_sort_key(value):
    """
    日期排序键：银行导出流水通常为倒序（新→旧），需转成正序（旧→新）。
    空值排到末尾；其余按原始字符串升序（同文件日期列格式一致，字符串序即时间序）。
    Python sort 稳定，同日期内保持原序。
    """
    s = str(value).strip() if value and value != "None" else ""
    return (s == "", s)


def is_reverse_ordered(values):
    """判断序列是否整体倒序（新→旧）：倒序对数严格多于正序对数时为真。"""
    vs = [str(v).strip() for v in values if v and v != "None"]
    if len(vs) < 2:
        return False
    lt = sum(1 for i in range(len(vs) - 1) if vs[i] < vs[i + 1])
    gt = sum(1 for i in range(len(vs) - 1) if vs[i] > vs[i + 1])
    return gt > lt


def sort_rows_ascending(rows, key_func):
    """
    按 key_func 升序排列行。
    银行导出常为整体倒序，且同一日期/时间戳内也是倒序：仅靠稳定排序无法纠正同键内的顺序。
    因此先检测整体是否倒序，若是则先整体翻转——翻转后同键内的顺序即恢复为正序，
    随后的稳定排序既保证跨日期升序，又保持同日期内的正确顺序。
    """
    if is_reverse_ordered([key_func(r) for r in rows]):
        rows = list(reversed(rows))
    return sorted(rows, key=key_func)


def decode_csv_text(filepath):
    """按常见银行导出编码读取 CSV 文本。"""
    with open(filepath, "rb") as f:
        raw = f.read()

    if not raw:
        return "", None

    for encoding in CSV_ENCODINGS:
        try:
            return raw.decode(encoding), None
        except UnicodeDecodeError:
            continue

    return "", "CSV 编码无法识别，已尝试 utf-8 / gb18030 / gbk / utf-16"


def detect_csv_dialect(sample):
    """识别 CSV 分隔符；失败时按逗号处理。"""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel()
        dialect.delimiter = ","
        return dialect


def load_xlsx(filepath):
    """加载 .xlsx 文件，返回 (headers, data_rows, error)"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        all_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            all_rows.append([str(c).strip() if c is not None else "" for c in row])

        if not all_rows:
            return [], [], "空文件"

        header_idx = find_header_row(all_rows)
        if header_idx is None:
            return [], [], "未找到表头行"

        headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]
        data_rows = [r for r in data_rows if is_data_row(r, headers)]

        return headers, data_rows, None
    except Exception as e:
        return None, [], str(e)


def load_csv(filepath):
    """加载 .csv 文件，返回 (headers, data_rows, error)"""
    try:
        content, error = decode_csv_text(filepath)
        if error:
            return [], [], error

        if not content.strip():
            return [], [], "空文件"

        dialect = detect_csv_dialect(content[:4096])
        reader = csv.reader(io.StringIO(content), dialect)
        all_rows = []
        for row in reader:
            cells = [str(c).replace("\ufeff", "").strip() if c is not None else "" for c in row]
            if any(cells):
                all_rows.append(cells)

        if not all_rows:
            return [], [], "空文件"

        header_idx = find_header_row(all_rows)
        if header_idx is None:
            return [], [], "CSV 中未找到表头行"

        headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]
        data_rows = [r for r in data_rows if is_data_row(r, headers)]

        return headers, data_rows, None
    except Exception as e:
        return None, [], f"CSV 解析失败: {e}"


def load_xls(filepath):
    """加载 .xls 二进制格式文件，返回 (headers, data_rows, error)"""
    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)

        all_rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(val.strftime("%Y-%m-%d %H:%M:%S") if val else "")
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    if cell.value == int(cell.value):
                        row.append(str(int(cell.value)))
                    else:
                        row.append(str(cell.value))
                else:
                    row.append(str(cell.value).strip() if cell.value else "")
            all_rows.append(row)

        if not all_rows:
            return [], [], "空文件"

        header_idx = find_header_row(all_rows)
        if header_idx is None:
            return [], [], "未找到表头行"

        headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]
        data_rows = [r for r in data_rows if is_data_row(r, headers)]

        return headers, data_rows, None
    except Exception as e:
        return None, [], str(e)


def load_xml_xls(filepath):
    """解析 XML Spreadsheet 格式的 .xls 文件（工行电子回单等）"""
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
        root = ET.fromstring(content)
        worksheet = root.find(".//ss:Worksheet", ns)
        if worksheet is None:
            return [], [], "XML 中未找到 Worksheet"
        table = worksheet.find("ss:Table", ns)
        if table is None:
            return [], [], "XML 中未找到 Table"

        all_rows = []
        for row_elem in table.findall("ss:Row", ns):
            cells = []
            for cell_elem in row_elem.findall("ss:Cell", ns):
                data_elem = cell_elem.find("ss:Data", ns)
                cells.append(
                    data_elem.text.strip()
                    if data_elem is not None and data_elem.text
                    else ""
                )
            all_rows.append(cells)

        if not all_rows:
            return [], [], "空文件"

        header_idx = find_header_row(all_rows)
        if header_idx is None:
            return [], [], "XML 中未找到表头行"

        headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]
        data_rows = [r for r in data_rows if is_data_row(r, headers)]

        return headers, data_rows, None
    except Exception as e:
        return None, [], f"XML 解析失败: {e}"


def load_html_xls(filepath):
    """解析 HTML 表格格式的 .xls 文件（银行导出常见）"""
    try:
        from html.parser import HTMLParser

        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        class TableParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.rows = []
                self.current_row = []
                self.current_cell = ""
                self.in_cell = False

            def handle_starttag(self, tag, attrs):
                if tag == "tr":
                    self.current_row = []
                elif tag in ("td", "th"):
                    self.in_cell = True
                    self.current_cell = ""

            def handle_endtag(self, tag):
                if tag in ("td", "th"):
                    self.in_cell = False
                    self.current_row.append(self.current_cell.strip())
                elif tag == "tr":
                    if self.current_row:
                        self.rows.append(self.current_row)

            def handle_data(self, data):
                if self.in_cell:
                    self.current_cell += data

            def handle_entityref(self, name):
                if self.in_cell:
                    import html
                    self.current_cell += " " if name == "nbsp" else html.unescape(f"&{name};")

        parser = TableParser()
        parser.feed(content)

        if not parser.rows:
            return [], [], "HTML 中未找到表格数据"

        header_idx = find_header_row(parser.rows)
        if header_idx is None:
            return [], [], "HTML 中未找到表头行"

        headers = parser.rows[header_idx]
        data_rows = parser.rows[header_idx + 1:]
        data_rows = [r for r in data_rows if is_data_row(r, headers)]

        return headers, data_rows, None
    except Exception as e:
        return None, [], f"HTML 解析失败: {e}"


def load_workbook_auto(filepath):
    """自动识别文件实际格式并加载，返回 (headers, data_rows, error)"""
    ext = os.path.splitext(filepath)[1].lower()

    with open(filepath, "rb") as f:
        head = f.read(20)

    # PK = ZIP (xlsx)，但有些文件扩展名是 .xls 实际是 xlsx
    if head[:2] == b"PK":
        if ext == ".xls":
            # openpyxl 会因扩展名拒绝，临时复制为 .xlsx
            tmp = os.path.join(tempfile.gettempdir(), "_claw_convert.xlsx")
            shutil.copy2(filepath, tmp)
            result = load_xlsx(tmp)
            os.unlink(tmp)
            return result
        return load_xlsx(filepath)
    elif ext == ".xls":
        if head[:5] == b"<?xml":
            return load_xml_xls(filepath)
        elif b"<html" in head[:20].lower():
            return load_html_xls(filepath)
        elif HAS_XLRD:
            headers, data_rows, error = load_xls(filepath)
            if error is None:
                return headers, data_rows, None
            return load_html_xls(filepath)
        else:
            return load_html_xls(filepath)
    elif ext == ".csv":
        return load_csv(filepath)
    elif ext == ".xlsx":
        return load_xlsx(filepath)
    else:
        return None, [], f"不支持的文件格式: {ext}"


def build_column_mapping(headers):
    """
    将源文件的列名映射到模板字段。
    返回 (mapping_dict, ccyc_col_idx, amount_col_idx, boc_debit_idx, boc_credit_idx) 元组。
    """
    mapping = {}
    used_targets = set()
    ccyc_col = None
    ccyc_mode = None
    amount_col = None
    boc_debit_col = None
    boc_credit_col = None
    boc_type_col = None

    cleaned_headers = [clean_header(header) for header in headers]
    special_cols = set()

    for col_idx, stripped in enumerate(cleaned_headers):
        if not stripped or stripped == "None":
            continue

        if stripped == "收支状况":
            ccyc_col = col_idx
            ccyc_mode = "收支"
            special_cols.add(col_idx)
            continue
        if stripped == "借贷标记":
            ccyc_col = col_idx
            ccyc_mode = "借贷"
            special_cols.add(col_idx)
            continue
        if ccyc_col is not None and stripped == "交易金额" and amount_col is None:
            amount_col = col_idx
            special_cols.add(col_idx)
            continue

        if stripped in ("交易类型", "业务方向"):
            boc_type_col = col_idx
            special_cols.add(col_idx)
            continue

        if stripped == "往账金额":
            boc_debit_col = col_idx
            special_cols.add(col_idx)
            continue
        if stripped == "来账金额":
            boc_credit_col = col_idx
            special_cols.add(col_idx)
            continue

    # Pass 1: 全表精确匹配，避免后续模糊匹配先占用更准确的字段。
    for col_idx, stripped in enumerate(cleaned_headers):
        if col_idx in special_cols or not stripped or stripped == "None":
            continue
        for target_field, aliases in FIELD_ALIASES.items():
            if target_field in used_targets:
                continue
            for alias in aliases:
                if alias == stripped:
                    mapping[col_idx] = target_field
                    used_targets.add(target_field)
                    break
            if col_idx in mapping:
                break

    # Pass 2: 模糊匹配（只允许 alias in header），处理带括号说明等银行表头。
    for col_idx, stripped in enumerate(cleaned_headers):
        if col_idx in special_cols or col_idx in mapping or not stripped or stripped == "None":
            continue
        for target_field, aliases in FIELD_ALIASES.items():
            if target_field in used_targets:
                continue
            if target_field == "对方银行" and any(kw in stripped for kw in ("行号", "地址")):
                continue
            for alias in aliases:
                if alias in stripped:
                    mapping[col_idx] = target_field
                    used_targets.add(target_field)
                    break
            if col_idx in mapping:
                break

    # 中国银行: "交易类型"列 + "交易金额"列（正负金额）
    # 仅当没有"收支状况"列时才使用交易类型作为收支判断
    if boc_type_col is not None and ccyc_col is None:
        for col_idx, header in enumerate(headers):
            h = clean_header(header)
            if h == "交易金额" and col_idx not in mapping:
                amount_col = col_idx
                break
        if amount_col is not None:
            ccyc_col = boc_type_col
            ccyc_mode = "收支"

    return mapping, ccyc_col, amount_col, boc_debit_col, boc_credit_col, ccyc_mode


def convert_row(row, mapping, ccyc_col=None, amount_col=None, boc_debit_col=None, boc_credit_col=None, ccyc_mode=None):
    """将源文件的一行数据按映射规则转换为模板格式"""
    result = OrderedDict()
    for field in TEMPLATE_COLUMNS:
        result[field] = ""

    for col_idx, target_field in mapping.items():
        if col_idx < len(row):
            val = row[col_idx]
            if target_field in DATE_COLS:
                result[target_field] = extract_date(val)
            elif target_field in AMOUNT_COLS:
                result[target_field] = extract_amount(val)
            else:
                result[target_field] = str(val).strip() if val and val != "None" else ""

    # 处理"收支状况"+"交易金额"组合列
    if ccyc_col is not None and amount_col is not None:
        ccyc_val = row[ccyc_col].strip() if ccyc_col < len(row) else ""
        amount_val = row[amount_col] if amount_col < len(row) else ""
        amount_clean = positive_amount(amount_val)
        if amount_clean:
            if ccyc_mode == "借贷":
                if "贷" in ccyc_val:
                    result["收入"] = amount_clean
                elif "借" in ccyc_val:
                    result["支出"] = amount_clean
            else:
                if "收入" in ccyc_val or "来账" in ccyc_val:
                    result["收入"] = amount_clean
                elif "支出" in ccyc_val or "往账" in ccyc_val:
                    result["支出"] = amount_clean

    # 处理中国银行"往账金额/来账金额"双列格式
    if boc_credit_col is not None and boc_debit_col is not None:
        debit_val = positive_amount(row[boc_debit_col]) if boc_debit_col < len(row) else ""
        credit_val = positive_amount(row[boc_credit_col]) if boc_credit_col < len(row) else ""
        if credit_val:
            result["收入"] = credit_val
        if debit_val:
            result["支出"] = debit_val

    return result


def is_electronic_receipt(headers):
    """判断是否为电子回单格式（工行等）"""
    header_text = "".join(h for h in headers if h)
    return "电子回单号码" in header_text


def get_value_by_headers(row, headers, names):
    """按清理后的表头精确取值。"""
    wanted = set(names)
    for idx, header in enumerate(headers):
        if clean_header(header) in wanted and idx < len(row):
            return str(row[idx]).strip()
    return ""


def fill_counterparty_from_party_columns(result, row, headers, company_name):
    """处理中国银行 CSV/XLS 常见的付款人/收款人列。"""
    payer_name = get_value_by_headers(row, headers, [
        "付款人名称",
        "付款方名称",
        "付款户名",
        "付款人户名",
        "付款方户名",
        "付款人账户名称",
        "付款方账户名称",
        "付方账户名称",
        "付方名称",
        "借方户名",
        "借方名称",
        "名义付款人名称",
    ])
    payer_bank = get_value_by_headers(row, headers, [
        "付款人开户行名",
        "付款人开户行",
        "付款人开户机构",
        "付款方开户行名",
        "付款方开户行",
        "付款方开户机构",
        "付款账户开户行",
        "付款账户开户机构",
        "付方开户银行名称",
        "付方开户行",
        "付方开户机构",
        "借方开户行",
        "借方开户机构",
        "名义付款人开户行名",
        "名义付款人开户行",
    ])
    payer_account = get_value_by_headers(row, headers, [
        "付款人账号",
        "付款方账号",
        "付款账号",
        "付款人账户",
        "付款方账户",
        "付方账号",
        "借方账号",
        "名义付款人账号",
    ])
    payee_name = get_value_by_headers(row, headers, [
        "收款人名称",
        "收款方名称",
        "收款户名",
        "收款人户名",
        "收款方户名",
        "收款人账户名称",
        "收款方账户名称",
        "收方账户名称",
        "收方名称",
        "贷方户名",
        "贷方名称",
        "名义收款人名称",
    ])
    payee_bank = get_value_by_headers(row, headers, [
        "收款人开户行名",
        "收款人开户行",
        "收款人开户机构",
        "收款方开户行名",
        "收款方开户行",
        "收款方开户机构",
        "收款账户开户行",
        "收款账户开户机构",
        "收方开户银行名称",
        "收方开户行",
        "收方开户机构",
        "贷方开户行",
        "贷方开户机构",
        "名义收款人开户行名",
        "名义收款人开户行",
    ])
    payee_account = get_value_by_headers(row, headers, [
        "收款人账号",
        "收款方账号",
        "收款账号",
        "收款人账户",
        "收款方账户",
        "收方账号",
        "贷方账号",
        "名义收款人账号",
    ])
    txn_type = get_value_by_headers(row, headers, ["交易类型", "业务方向", "收支状况", "借贷标记"])

    if not (payer_name or payee_name or payer_account or payee_account):
        return

    direction = ""
    if result.get("支出") and not result.get("收入"):
        direction = "out"
    elif result.get("收入") and not result.get("支出"):
        direction = "in"
    elif company_name and payer_name and company_name in payer_name:
        direction = "out"
    elif company_name and payee_name and company_name in payee_name:
        direction = "in"
    elif "往账" in txn_type or "支出" in txn_type or "借" in txn_type:
        direction = "out"
    elif "来账" in txn_type or "收入" in txn_type or "贷" in txn_type:
        direction = "in"

    if direction == "out":
        counterparty = (payee_name, payee_bank, payee_account)
    elif direction == "in":
        counterparty = (payer_name, payer_bank, payer_account)
    else:
        return

    for field, value in zip(("对方户名", "对方银行", "对方账号"), counterparty):
        if value and not result.get(field):
            result[field] = value


def fill_counterparty_from_candidates(result, row, headers):
    """
    对方户名/对方银行/对方账号 为空时，扫描所有别名匹配的候选列，取第一个非空值。
    解决招行等格式中"分行名"为空但"开户行名"有值、单列映射却取到空值的问题。
    """
    for field in ("对方户名", "对方银行", "对方账号"):
        if result.get(field):
            continue
        aliases = FIELD_ALIASES.get(field, [])
        for idx, header in enumerate(headers):
            stripped = clean_header(header)
            if not stripped or stripped == "None":
                continue
            # 与 build_column_mappping 保持一致：对方银行排除行号/地址列
            if field == "对方银行" and any(kw in stripped for kw in ("行号", "地址")):
                continue
            if any(alias == stripped or alias in stripped for alias in aliases):
                if idx < len(row):
                    val = str(row[idx]).strip() if row[idx] and row[idx] != "None" else ""
                    if val:
                        result[field] = val
                        break


def convert_receipt_row(row, headers, company_name):
    """将电子回单格式的一行转换为模板格式"""
    result = OrderedDict()
    for field in TEMPLATE_COLUMNS:
        result[field] = ""

    def get(col_name):
        for i, h in enumerate(headers):
            if h == col_name and i < len(row):
                return row[i].strip()
        return ""

    # 日期：优先用记账日期（格式: 2026年04月20日），否则用时间戳
    accounting_date = get("记账日期")
    if accounting_date:
        m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", accounting_date)
        if m:
            result["日期"] = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    if not result["日期"]:
        result["日期"] = extract_date(get("时间戳"))

    amount = extract_amount(get("金额"))
    payer_name = get("付方账户名称")
    payee_name = get("收方账户名称")

    # 判断收入/支出：如果公司名匹配付方，则是支出
    if company_name and payer_name and company_name in payer_name:
        result["支出"] = amount
        result["对方户名"] = payee_name
        result["对方银行"] = get("收方开户银行名称")
        result["对方账号"] = get("收方账号")
    else:
        result["收入"] = amount
        result["对方户名"] = payer_name
        result["对方银行"] = get("付方开户银行名称")
        result["对方账号"] = get("付方账号")

    result["摘要"] = get("摘要")
    result["备注"] = get("用途") or get("备注")

    return result


def process_file(filepath):
    """处理单个 Excel/CSV 文件，返回 (company_name, converted_rows, error)"""
    company_name = os.path.splitext(os.path.basename(filepath))[0]
    # 清理文件名中的日期后缀
    company_name = re.sub(r"\s*\d{4}.*$", "", company_name).strip()
    company_name = re.sub(r"\s+新$", "", company_name).strip()
    if not company_name:
        company_name = os.path.splitext(os.path.basename(filepath))[0][:30]

    headers, data_rows, error = load_workbook_auto(filepath)
    if error:
        return company_name, [], error

    converted = []
    if is_electronic_receipt(headers):
        for row in data_rows:
            converted.append(convert_receipt_row(row, headers, company_name))
        # 电子回单按转换后日期正序排列（倒序文件会先翻转）
        converted = sort_rows_ascending(converted, key_func=lambda r: date_sort_key(r.get("日期", "")))
    else:
        mapping, ccyc_col, amount_col, boc_debit_col, boc_credit_col, ccyc_mode = build_column_mapping(headers)
        if not mapping and boc_debit_col is None and boc_credit_col is None:
            return company_name, [], f"无法映射任何列 (表头: {headers[:8]})"
        # 按原始日期列升序排列（银行导出流水通常为倒序，需先翻转再排序）
        date_idx = next((idx for idx, field in mapping.items() if field == "日期"), None)
        if date_idx is not None:
            data_rows = sort_rows_ascending(
                data_rows,
                key_func=lambda r: date_sort_key(r[date_idx] if date_idx < len(r) else ""),
            )
        for row in data_rows:
            converted_row = convert_row(row, mapping, ccyc_col, amount_col, boc_debit_col, boc_credit_col, ccyc_mode)
            fill_counterparty_from_party_columns(converted_row, row, headers, company_name)
            fill_counterparty_from_candidates(converted_row, row, headers)
            converted.append(converted_row)

    # 摘要为空时用备注回填（部分银行无摘要列，用途/附言仅出现在备注列）
    for row_data in converted:
        if not row_data.get("摘要") and row_data.get("备注"):
            row_data["摘要"] = row_data["备注"]

    return company_name, converted, None


def _safe_zip_name(name):
    """清理用于 zip 内文件名的非法字符。"""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip(" .") or "未命名"


def _utf8_zip_info(name):
    """ZipInfo 带 UTF-8 文件名标志位，避免 Linux/Mac 解压中文乱码。"""
    zinfo = zipfile.ZipInfo(name)
    zinfo.flag_bits |= 0x800
    zinfo.compress_type = zipfile.ZIP_DEFLATED
    return zinfo


def _recover_mojibake(name):
    """
    修复 GBK 文件名被当作 Latin-1/CP437 误读形成的 Unicode 乱码。
    场景：Windows 创建的 ZIP（GBK 字节、无 UTF-8 flag）在 Linux 用 unzip
    解压后，文件名字节被按 CP437 映射成 Unicode 码点写到磁盘，glob 读到
    后是 b"╗▌╓▌..." 之类的乱码字符串。
    若 name 能用 cp437 编码且解出 GBK 后包含 CJK 字符，视为乱码并返回还原值。
    """
    if not name or all(ord(c) < 128 for c in name):
        return name
    try:
        recovered = name.encode("cp437").decode("gbk")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name
    has_cjk = lambda s: any("\u4e00" <= c <= "\u9fff" for c in s)
    if has_cjk(recovered) and not has_cjk(name):
        return recovered
    return name


def _workbook_to_bytes(wb):
    """openpyxl Workbook 序列化到内存 bytes。"""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _template_cell_value(row_data, field):
    """模板写出前的最终规范化。"""
    value = row_data.get(field, "")
    if field == "日期":
        return extract_date(value)
    return value


def _build_single_workbook(rows, sheet_title):
    """单个源文件 → 单 Sheet Excel（9 列）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Sheet1"
    for col, field in enumerate(TEMPLATE_COLUMNS, 1):
        ws.cell(row=1, column=col, value=field)
    for r, row_data in enumerate(rows, 2):
        for col, field in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=r, column=col, value=_template_cell_value(row_data, field))
            if field == "日期":
                cell.number_format = "@"
    return wb


def _build_summary_workbook(all_rows):
    """汇总 Excel：所有数据 + _来源公司 列。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    headers = TEMPLATE_COLUMNS + ["_来源公司"]
    for col, field in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=field)
    for r, row_data in enumerate(all_rows, 2):
        for col, field in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=_template_cell_value(row_data, field))
            if field == "日期":
                cell.number_format = "@"
    return wb


def process_folder(input_dir, output_path, template_path=None):
    """
    扫描文件夹，逐文件转换，输出 zip 压缩包。
    zip 内含：汇总.xlsx + 每个源文件一份独立 Excel。
    返回 (total_files, success_count, error_details)。
    """
    # 收集所有支持的流水文件
    files = []
    for root, _, filenames in os.walk(input_dir):
        for filename in filenames:
            ext = os.path.splitext(filename)[1].lower()
            if ext in SUPPORTED_INPUT_EXTENSIONS:
                files.append(os.path.join(root, filename))
    files = [
        f for f in files
        if not os.path.basename(f).startswith("~")
        and "说明" not in os.path.basename(f)
        and os.path.splitext(os.path.basename(f))[0] != "汇总"
    ]

    if not files:
        return 0, 0, ["未找到任何 Excel/CSV 文件"]

    success = 0
    errors = []
    all_summary_rows = []
    used_names = set()

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for filepath in sorted(files):
            company, rows, error = process_file(filepath)
            rel_path = os.path.relpath(filepath, input_dir)

            if error:
                errors.append(f"[失败] {rel_path}: {error}")
                continue

            if not rows:
                errors.append(f"[跳过] {rel_path}: 无数据行")
                continue

            # 用原文件名命名，去掉路径和扩展名，统一 .xlsx
            base = os.path.splitext(os.path.basename(filepath))[0]
            base = _recover_mojibake(base)
            zip_name = _safe_zip_name(base) + ".xlsx"
            if zip_name == "汇总.xlsx":
                zip_name = "汇总_来源文件.xlsx"

            # 防重名（不同子目录下同名文件）
            n, candidate = 1, zip_name
            while candidate in used_names:
                n += 1
                candidate = f"{_safe_zip_name(base)}_{n}.xlsx"
            zip_name = candidate
            used_names.add(zip_name)

            wb = _build_single_workbook(rows, company or base)
            zf.writestr(_utf8_zip_info(zip_name), _workbook_to_bytes(wb))

            for row_data in rows:
                row_data["_来源公司"] = company
                all_summary_rows.append(row_data)

            success += 1

        if all_summary_rows:
            summary_wb = _build_summary_workbook(all_summary_rows)
            zf.writestr(_utf8_zip_info("汇总.xlsx"), _workbook_to_bytes(summary_wb))

    return len(files), success, errors


def main():
    if len(sys.argv) < 2:
        print("用法: python converter.py <输入文件夹> [输出 zip 路径]")
        print("示例: python converter.py ./流水数据 ./输出/监管流水_汇总.zip")
        sys.exit(1)

    input_dir = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(os.path.dirname(input_dir), "输出")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"监管流水_{timestamp}.zip")

    if not os.path.isdir(input_dir):
        print(f"错误: 文件夹不存在: {input_dir}")
        sys.exit(1)

    print(f"扫描文件夹: {input_dir}")
    total, success, errors = process_folder(input_dir, output_path)

    print(f"\n处理完成:")
    print(f"  总文件数: {total}")
    print(f"  成功转换: {success}")
    print(f"  失败/跳过: {total - success}")

    if errors:
        print(f"\n失败详情:")
        for err in errors:
            print(f"  {err}")

    print(f"\n输出 zip: {output_path}")


if __name__ == "__main__":
    main()
