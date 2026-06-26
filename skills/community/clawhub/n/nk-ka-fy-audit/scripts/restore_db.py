#!/usr/bin/env python3
"""安全恢复数据库 - 每次只处理一个月，不累积内存
步骤：
1. 创建raw_records表（从data sheets逐行解析）
2. 创建brand_audit表（从汇总表按品牌汇总，从raw_records计算重算）
3. 创建monthly_audit表

不删除已有数据，只用INSERT OR REPLACE处理缺失月份。
"""
import sqlite3, os, glob, re, gc, sys, json

DATA_DIR = '/workspace/data'
DB_PATH = '/workspace/data/ka_commission_audit.db'

def S(v): return str(v).strip() if v is not None else ''
def F(v):
    try: return float(v) if v is not None else 0
    except: return 0

def find_hdr(ws):
    for r in range(1, min(ws.max_row+1, 30)):
        txt = '|'.join(str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, ws.max_column+1))
        if '品牌' in txt and '返佣合计' in txt: return r
    for r in range(1, min(ws.max_row+1, 30)):
        txt = '|'.join(str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, ws.max_column+1))
        if '品牌' in txt and '交易金额' in txt: return r
    for r in range(1, min(ws.max_row+1, 30)):
        txt = '|'.join(str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, ws.max_column+1))
        if '品牌' in txt: return r
    return 1

def parse_summary_brands(ws):
    """解析汇总表，按品牌汇总返回 {brand: total_amount}"""
    hr = find_hdr(ws)
    hdrs = [str(ws.cell(row=hr, column=c).value or '') for c in range(1, min(ws.max_column+1, 20))]
    
    brand_col = None; amt_col = None
    for i, h in enumerate(hdrs):
        hl = h.lower()
        if hl == '品牌': brand_col = i
        elif '应返金额' in hl or '入账' in hl: amt_col = i
    
    if brand_col is None:
        brand_col, amt_col = 3, 8
    if amt_col is None:
        amt_col = 8
    
    result = {}
    for r in range(hr+1, ws.max_row+1):
        b = S(ws.cell(row=r, column=brand_col+1).value)
        if not b or b == '品牌': continue
        amt = F(ws.cell(row=r, column=amt_col+1).value)
        if b not in result:
            result[b] = 0.0
        result[b] += amt
    return result

def parse_data_sheet(ws, hr, hdrs, sheet_name):
    """解析数据sheet，返回记录列表"""
    idx = {}
    for i, h in enumerate(hdrs):
        hl = h.lower().strip()
        if hl in ('品牌', '品牌名称'): idx['brand'] = i
        elif hl in ('收钱吧商户名称', '商户名称', '商户名'): idx['merchant'] = i
        elif hl in ('收钱吧商户号', '商户号'): idx['merchant_no'] = i
        elif hl in ('支付宝交易金额', '支付宝交易额'): idx['ali_txn'] = i
        elif hl in ('支付宝返佣比例', '支付宝分润比例'): idx['ali_ratio'] = i
        elif hl in ('支付宝返佣金额', '支付宝分润金额'): idx['ali_amt'] = i
        elif hl in ('微信交易金额', '微信交易额'): idx['wx_txn'] = i
        elif hl in ('微信返佣比例', '微信分润比例'): idx['wx_ratio'] = i
        elif hl in ('微信返佣金额', '微信分润金额'): idx['wx_amt'] = i
        elif '返佣合计' in hl or '返佣金额' in hl: idx['total'] = i
        elif hl in ('交易期日', '交易期', '月份', 'month'): idx['occ_date'] = i
        elif '备注' in hl: idx['note'] = i
    
    # level fields
    for i, h in enumerate(hdrs):
        hl = h.lower().strip()
        if not idx.get('level1') and ('一级组织' in hl or '一级' == hl): idx['level1'] = i
        if not idx.get('level2') and ('二级组织' in hl or '二级' == hl): idx['level2'] = i
        if not idx.get('level3') and ('三级组织' in hl or '三级' == hl): idx['level3'] = i
    
    if 'brand' not in idx:
        if 'merchant' in idx: idx['brand'] = idx['merchant']
        else: return []
    
    brand_col = idx['brand']
    records = []
    
    for r in range(hr+1, ws.max_row+1):
        fc = S(ws.cell(row=r, column=1).value)
        b = S(ws.cell(row=r, column=brand_col+1).value)
        if not b or b.lower() in ('品牌', '品牌名称', '合计', '汇总', '小计', '总计', '商户号', ''): continue
        if not fc and not b: continue
        
        ali_txn = F(ws.cell(row=r, column=idx.get('ali_txn', -1)+1).value) if 'ali_txn' in idx else 0
        ali_ratio = F(ws.cell(row=r, column=idx.get('ali_ratio', -1)+1).value) if 'ali_ratio' in idx else 0
        ali_amt = F(ws.cell(row=r, column=idx.get('ali_amt', -1)+1).value) if 'ali_amt' in idx else 0
        wx_txn = F(ws.cell(row=r, column=idx.get('wx_txn', -1)+1).value) if 'wx_txn' in idx else 0
        wx_ratio = F(ws.cell(row=r, column=idx.get('wx_ratio', -1)+1).value) if 'wx_ratio' in idx else 0
        wx_amt = F(ws.cell(row=r, column=idx.get('wx_amt', -1)+1).value) if 'wx_amt' in idx else 0
        total = F(ws.cell(row=r, column=idx.get('total', -1)+1).value) if 'total' in idx else 0
        occ = S(ws.cell(row=r, column=idx.get('occ_date', -1)+1).value) if 'occ_date' in idx else ''
        note = S(ws.cell(row=r, column=idx.get('note', -1)+1).value) if 'note' in idx else ''
        l1 = S(ws.cell(row=r, column=idx.get('level1', -1)+1).value) if 'level1' in idx else ''
        l2 = S(ws.cell(row=r, column=idx.get('level2', -1)+1).value) if 'level2' in idx else ''
        l3 = S(ws.cell(row=r, column=idx.get('level3', -1)+1).value) if 'level3' in idx else ''
        
        records.append({
            'brand': b,
            'ali_txn': ali_txn, 'ali_ratio': ali_ratio, 'ali_amt': ali_amt,
            'wx_txn': wx_txn, 'wx_ratio': wx_ratio, 'wx_amt': wx_amt,
            'total_rebate': total,
            'occurrence_date': occ, 'original_note': note,
            'level1': l1, 'level2': l2, 'level3': l3,
        })
    
    return records

def get_snote(ws):
    return S(ws.cell(row=1, column=1).value) or ''

def should_inc(snote, row_note):
    if '视同无差异' in row_note: return False
    if '调账' in row_note: return False
    return True

def process_month(month):
    """处理单个月份"""
    import openpyxl as _pxl
    found = None
    for p in [os.path.join(DATA_DIR, month, 'KA返佣{}*.xlsx'.format(month)),
              os.path.join(DATA_DIR, month, 'KA{}*.xlsx'.format(month))]:
        fs = glob.glob(p)
        if fs: found = fs[0]; break
    if not found: return None, None, None, None, '文件未找到: {}'.format(month)
    
    wb = _pxl.load_workbook(found, data_only=True)
    snames = wb.sheetnames
    
    # 1. 解析汇总表
    brand_summary = {}
    if '汇总' in snames:
        brand_summary = parse_summary_brands(wb['汇总'])
    
    # 2. 解析数据sheet
    raw_records = []
    calc_by_brand = {}
    raw_by_brand = {}
    adj_by_brand = {}
    
    for sn in snames:
        if sn == '汇总': continue
        sn_lower = sn.lower().strip()
        if sn_lower == 'sql': continue
        
        ws = wb[sn]
        hr = find_hdr(ws)
        hdrs = [str(ws.cell(row=hr, column=c).value or '') for c in range(1, ws.max_column+1)]
        txt = '|'.join(v.lower() for v in hdrs)
        
        is_problem_sheet = '问题' in sn or '异常' in sn
        is_zhongkuai = '中快' in sn
        is_english = 'bill_email' in txt or 'email' in txt or '分润汇总' in txt
        is_zhihua = '志华' in sn or '拓展' in sn
        
        if is_problem_sheet:
            rows = parse_data_sheet(ws, hr, hdrs, sn)
            snote = get_snote(ws)
            for row in rows:
                b = row['brand']
                recalc = round(row.get('ali_amt', 0), 2) + round(row.get('wx_amt', 0), 2)
                if not recalc:
                    recalc = round(row['ali_txn'] * row['ali_ratio'], 2) + round(row['wx_txn'] * row['wx_ratio'], 2)
                if not recalc:
                    recalc = row['total_rebate']
                
                orig_note = row.get('original_note', '')
                is_adj = '调账' in (orig_note + ' ' + snote)
                
                if is_adj:
                    adj_by_brand[b] = adj_by_brand.get(b, 0) + recalc
                    raw_by_brand[b] = raw_by_brand.get(b, 0) + row['total_rebate']
                    audit_note = '历史差额调整'
                elif should_inc(snote, orig_note):
                    calc_by_brand[b] = calc_by_brand.get(b, 0) + recalc
                    raw_by_brand[b] = raw_by_brand.get(b, 0) + row['total_rebate']
                    audit_note = '有问题，暂时不调整'
                else:
                    audit_note = '有问题，暂时不调整'
                
                raw_records.append({
                    'brand': b, 'calc_method': 'abnormal', 'sheet_name': sn,
                    'ali_txn': row['ali_txn'], 'ali_ratio': row['ali_ratio'],
                    'wx_txn': row['wx_txn'], 'wx_ratio': row['wx_ratio'],
                    'total_rebate': row['total_rebate'],
                    'audit_note': audit_note,
                    'original_note': orig_note,
                    'occurrence_date': row.get('occurrence_date', ''),
                    'level1': row.get('level1', ''), 'level2': row.get('level2', ''), 'level3': row.get('level3', ''),
                })
            continue
        
        if is_zhongkuai:
            rows = parse_data_sheet(ws, hr, hdrs, sn)
            for row in rows:
                b = row['brand']
                calc_by_brand[b] = calc_by_brand.get(b, 0) + row['total_rebate']
                raw_by_brand[b] = raw_by_brand.get(b, 0) + row['total_rebate']
                raw_records.append({
                    'brand': b, 'calc_method': 'zhongkuai', 'sheet_name': sn,
                    'ali_txn': row['ali_txn'], 'ali_ratio': row['ali_ratio'],
                    'wx_txn': row['wx_txn'], 'wx_ratio': row['wx_ratio'],
                    'total_rebate': row['total_rebate'],
                    'audit_note': '', 'original_note': row.get('original_note', ''),
                    'occurrence_date': row.get('occurrence_date', ''),
                    'level1': row.get('level1', ''), 'level2': row.get('level2', ''), 'level3': row.get('level3', ''),
                })
            continue
        
        rows = parse_data_sheet(ws, hr, hdrs, sn)
        for row in rows:
            b = row['brand']
            if b == '大客户' and row.get('level1'):
                row['brand'] = row['level1']
                b = row['brand']
            
            recalc = round(row.get('ali_amt', 0), 2) + round(row.get('wx_amt', 0), 2)
            if not recalc:
                recalc = round(row['ali_txn'] * row['ali_ratio'], 2) + round(row['wx_txn'] * row['wx_ratio'], 2)
            if not recalc:
                recalc = row['total_rebate']
            
            calc_by_brand[b] = calc_by_brand.get(b, 0) + recalc
            raw_by_brand[b] = raw_by_brand.get(b, 0) + row['total_rebate']
            
            calc_method = 'regular'
            if is_zhihua:
                if month == '202307' and '拓展' in sn:
                    calc_method = 'expansion_307'
                elif '志华' in sn:
                    calc_method = 'expansion_308'
                else:
                    calc_method = 'expansion_307'
            
            raw_records.append({
                'brand': b, 'calc_method': calc_method, 'sheet_name': sn,
                'ali_txn': row['ali_txn'], 'ali_ratio': row['ali_ratio'],
                'wx_txn': row['wx_txn'], 'wx_ratio': row['wx_ratio'],
                'total_rebate': row['total_rebate'],
                'audit_note': '', 'original_note': row.get('original_note', ''),
                'occurrence_date': row.get('occurrence_date', ''),
                'level1': row.get('level1', ''), 'level2': row.get('level2', ''), 'level3': row.get('level3', ''),
            })
    
    # 计入调整
    for b, v in adj_by_brand.items():
        calc_by_brand[b] = calc_by_brand.get(b, 0) + v
    
    wb.close()
    del wb; gc.collect()
    
    return raw_records, brand_summary, calc_by_brand, raw_by_brand, None

def main():
    all_months = sorted([d for d in os.listdir(DATA_DIR) if d.isdigit() and len(d) == 6])
    if not all_months:
        print("没有找到数据目录"); return
    
    conn = sqlite3.connect(DB_PATH, timeout=120)
    
    # 创建表
    conn.execute("""CREATE TABLE IF NOT EXISTS raw_records(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        month TEXT, brand TEXT, calc_method TEXT, sheet_name TEXT,
        alipay_txn_amount REAL, alipay_rebate_ratio REAL,
        wechat_txn_amount REAL, wechat_rebate_ratio REAL,
        total_rebate REAL, audit_note TEXT, original_note TEXT,
        occurrence_date TEXT, level1_name TEXT, level2_name TEXT, level3_name TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS brand_audit(
        month TEXT, brand TEXT,
        raw_amount REAL, recalc_amount REAL, summary_amount REAL,
        diff REAL, diff_note TEXT, is_real_diff INTEGER
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS monthly_audit(
        month TEXT PRIMARY KEY,
        total_recalc REAL, total_summary REAL,
        total_diff REAL, real_diff REAL, brand_count INTEGER
    )""")
    conn.commit()
    
    existing = set(r[0] for r in conn.execute("SELECT DISTINCT month FROM brand_audit").fetchall())
    to_process = [m for m in all_months if m not in existing]
    
    if not to_process:
        print("所有 {} 个月份已在数据库中".format(len(all_months)))
        show_stats(conn); conn.close(); return
    
    print("需要处理 {} 个月份: {} ~ {}".format(len(to_process), to_process[0], to_process[-1]))
    
    for i, month in enumerate(to_process):
        print("  [{}/] {}...".format(i+1, len(to_process), month), end=' ', flush=True)
        
        result = process_month(month)
        if result[4] is not None:
            print("ERROR: {}".format(result[4]))
            continue
        
        raw_records, brand_summary, calc_by_brand, raw_by_brand = result[:4]
        
        if not raw_records:
            print('无数据'); continue
        
        insert_sql = """INSERT INTO raw_records(
            month, brand, calc_method, sheet_name,
            alipay_txn_amount, alipay_rebate_ratio,
            wechat_txn_amount, wechat_rebate_ratio,
            total_rebate, audit_note, original_note,
            occurrence_date, level1_name, level2_name, level3_name
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
        
        batch = [(month, rec['brand'], rec['calc_method'], rec['sheet_name'],
                  rec['ali_txn'], rec['ali_ratio'], rec['wx_txn'], rec['wx_ratio'],
                  rec['total_rebate'], rec['audit_note'], rec['original_note'],
                  rec['occurrence_date'], rec['level1'], rec['level2'], rec['level3'])
                 for rec in raw_records]
        conn.executemany(insert_sql, batch)
        
        all_brands = set(list(calc_by_brand.keys()) + list(brand_summary.keys()))
        for b in sorted(all_brands):
            recalc = round(calc_by_brand.get(b, 0), 2)
            sm = round(brand_summary.get(b, 0), 2)
            ra = round(raw_by_brand.get(b, 0), 2)
            diff = round(sm - recalc, 2)
            
            if diff < 0:
                is_real, note = 0, '少发，视同无差异'
            elif abs(diff) < 1:
                is_real, note = 0, '小数舍入'
            else:
                is_real, note = 1, ''
            
            conn.execute("""INSERT INTO brand_audit(month,brand,raw_amount,recalc_amount,summary_amount,diff,diff_note,is_real_diff)
                VALUES(?,?,?,?,?,?,?,?)""", (month, b, ra, recalc, sm, diff, note, is_real))
        
        conn.commit()
        print("{}条,品牌".format(len(raw_records), len(all_brands)))
        
        del raw_records, brand_summary, calc_by_brand, raw_by_brand, batch, result
        gc.collect()
    
    # 更新monthly_audit
    print("\n更新月度汇总...")
    months = set(r[0] for r in conn.execute("SELECT DISTINCT month FROM brand_audit").fetchall())
    for month in months:
        rows = conn.execute("SELECT recalc_amount, summary_amount, is_real_diff, diff FROM brand_audit WHERE month=?", (month,)).fetchall()
        tr = round(sum(r[0] for r in rows), 2)
        ts = round(sum(r[1] for r in rows), 2)
        rd = round(sum(r[3] for r in rows if r[2]), 2)
        bc = len(rows)
        conn.execute("""INSERT OR REPLACE INTO monthly_audit(month,total_recalc,total_summary,total_diff,real_diff,brand_count)
            VALUES(?,?,?,?,?,?)""", (month, tr, ts, round(ts-tr, 2), rd, bc))
    conn.commit()
    
    show_stats(conn)
    conn.close()

def show_stats(conn):
    ma = list(conn.execute("SELECT * FROM monthly_audit ORDER BY month").fetchall())
    print("\n总计: {}个月份".format(len(ma)))
    for m in ma:
        st = 'PASS' if m[4] == 0 else ('WARN' if abs(m[4]) < 1000 else 'DIFF')
        print("  {}: {}品牌 汇总={:.2f} 重算={:.2f} 真实差异={:+.2f} {}".format(
            m[0], m[5], m[2], m[1], m[4], st))
    total_rd = sum(r[4] for r in ma)
    perfect = sum(1 for r in ma if r[4] == 0)
    print("通过: {}/{} | 真实差异合计: {:+,.2f}".format(perfect, len(ma), total_rd))

if __name__ == '__main__':
    main()
