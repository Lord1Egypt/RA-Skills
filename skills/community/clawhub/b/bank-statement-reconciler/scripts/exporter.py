"""
Reconciliation Results Exporter
Exports to Excel format
"""
import os
from datetime import datetime
from typing import Dict, List

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ReconciliationExporter:
    """Export reconciliation results to Excel."""
    
    def export(self, result: Dict, output_dir: str = "/tmp") -> str:
        """
        Export reconciliation results to Excel.
        
        Args:
            result: Reconciliation result dict
            output_dir: Output directory
        
        Returns:
            Path to exported Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback: create CSV
            return self._export_csv(result, output_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reconciliation_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        self._write_summary(wb, result["summary"])
        
        # Matched transactions
        if result["matched"]:
            self._write_matched(wb, result["matched"])
        
        # Differences
        if result["differences"]:
            self._write_differences(wb, result["differences"])
        
        # Unclaimed (money without order)
        if result["unclaimed"]:
            self._write_unclaimed(wb, result["unclaimed"])
        
        # Unmatched orders (order without payment)
        if result["unmatched_orders"]:
            self._write_unmatched(wb, result["unmatched_orders"])
        
        wb.save(filepath)
        return filepath
    
    def _write_summary(self, wb, summary: Dict):
        """Write summary sheet."""
        ws = wb.active
        ws.title = "汇总"
        
        # Styles
        label_font = Font(bold=True)
        
        # Title
        ws["A1"] = "对账汇总"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")
        
        # Summary data
        row = 3
        summary_items = [
            ("总交易笔数", summary.get("total_transactions", 0)),
            ("总订单笔数", summary.get("total_orders", 0)),
            ("已匹配笔数", summary.get("matched_count", 0)),
            ("差异笔数", summary.get("difference_count", 0)),
            ("未认领笔数（有钱没订单）", summary.get("unclaimed_count", 0)),
            ("未核销笔数（有订单没收钱）", summary.get("unmatched_count", 0)),
            ("匹配率", f"{summary.get('match_rate', 0):.2f}%"),
            ("认账率", f"{summary.get('recognition_rate', 0):.2f}%"),
            ("已匹配金额", f"¥{summary.get('matched_amount', 0):,.2f}"),
            ("差异金额", f"¥{summary.get('difference_amount', 0):,.2f}"),
            ("未认领金额", f"¥{summary.get('unclaimed_amount', 0):,.2f}"),
            ("未核销金额", f"¥{summary.get('unmatched_amount', 0):,.2f}"),
        ]
        
        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
    
    def _write_matched(self, wb, matched: List[Dict]):
        """Write matched transactions sheet."""
        ws = wb.create_sheet("匹配结果")
        
        headers = ["日期", "金额", "对方账户", "摘要", "订单日期", "订单金额", "订单对方", "匹配方式"]
        
        # Write headers
        header_fill = PatternFill("solid", fgColor="70AD47")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, item in enumerate(matched, 2):
            order = item.get("matched_order", {})
            
            ws.cell(row=row_idx, column=1, value=item.get("date", ""))
            ws.cell(row=row_idx, column=2, value=item.get("amount", 0))
            ws.cell(row=row_idx, column=3, value=item.get("counterparty", ""))
            ws.cell(row=row_idx, column=4, value=item.get("summary", ""))
            ws.cell(row=row_idx, column=5, value=order.get("date", ""))
            ws.cell(row=row_idx, column=6, value=order.get("amount", 0))
            ws.cell(row=row_idx, column=7, value=order.get("counterparty", ""))
            ws.cell(row=row_idx, column=8, value=item.get("match_type", ""))
        
        # Column widths
        for col, width in enumerate([15, 15, 25, 30, 15, 15, 25, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _write_differences(self, wb, differences: List[Dict]):
        """Write differences sheet."""
        ws = wb.create_sheet("差异")
        
        headers = ["交易日期", "交易金额", "对方账户", "摘要", 
                   "订单日期", "订单金额", "差异金额", "处理状态"]
        
        header_fill = PatternFill("solid", fgColor="ED7D31")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(differences, 2):
            order = item.get("matched_order", {})
            
            ws.cell(row=row_idx, column=1, value=item.get("date", ""))
            ws.cell(row=row_idx, column=2, value=item.get("trans_amount", item.get("amount", 0)))
            ws.cell(row=row_idx, column=3, value=item.get("counterparty", ""))
            ws.cell(row=row_idx, column=4, value=item.get("summary", ""))
            ws.cell(row=row_idx, column=5, value=order.get("date", ""))
            ws.cell(row=row_idx, column=6, value=order.get("order_amount", order.get("amount", 0)))
            ws.cell(row=row_idx, column=7, value=item.get("difference", 0))
            ws.cell(row=row_idx, column=8, value=item.get("status", "待处理"))
        
        for col, width in enumerate([15, 15, 25, 30, 15, 15, 15, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _write_unclaimed(self, wb, unclaimed: List[Dict]):
        """Write unclaimed transactions sheet."""
        ws = wb.create_sheet("未认领")
        
        headers = ["日期", "金额", "对方账户", "摘要", "余额", "处理状态"]
        
        header_fill = PatternFill("solid", fgColor="FFC000")
        header_font = Font(bold=True, color="000000")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(unclaimed, 2):
            ws.cell(row=row_idx, column=1, value=item.get("date", ""))
            ws.cell(row=row_idx, column=2, value=item.get("amount", 0))
            ws.cell(row=row_idx, column=3, value=item.get("counterparty", ""))
            ws.cell(row=row_idx, column=4, value=item.get("summary", ""))
            ws.cell(row=row_idx, column=5, value=item.get("balance", ""))
            ws.cell(row=row_idx, column=6, value=item.get("status", "待处理"))
        
        for col, width in enumerate([15, 15, 25, 30, 15, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _write_unmatched(self, wb, unmatched: List[Dict]):
        """Write unmatched orders sheet."""
        ws = wb.create_sheet("未核销")
        
        headers = ["订单日期", "订单金额", "客户", "订单号", "商品", "处理状态"]
        
        header_fill = PatternFill("solid", fgColor="9E480E")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(unmatched, 2):
            ws.cell(row=row_idx, column=1, value=item.get("date", ""))
            ws.cell(row=row_idx, column=2, value=item.get("amount", 0))
            ws.cell(row=row_idx, column=3, value=item.get("counterparty", ""))
            ws.cell(row=row_idx, column=4, value=item.get("order_no", ""))
            ws.cell(row=row_idx, column=5, value=item.get("summary", ""))
            ws.cell(row=row_idx, column=6, value=item.get("status", "待处理"))
        
        for col, width in enumerate([15, 15, 25, 20, 30, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _export_csv(self, result: Dict, output_dir: str) -> str:
        """Fallback CSV export."""
        import csv
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export matched
        if result["matched"]:
            filepath = os.path.join(output_dir, f"reconciliation_matched_{timestamp}.csv")
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["日期", "金额", "对方账户", "摘要", "订单日期", "订单金额"])
                for item in result["matched"]:
                    order = item.get("matched_order", {})
                    writer.writerow([
                        item.get("date", ""),
                        item.get("amount", 0),
                        item.get("counterparty", ""),
                        item.get("summary", ""),
                        order.get("date", ""),
                        order.get("amount", 0),
                    ])
            return filepath
        
        return ""
