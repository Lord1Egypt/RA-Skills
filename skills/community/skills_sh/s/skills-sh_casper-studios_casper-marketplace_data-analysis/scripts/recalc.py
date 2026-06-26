#!/usr/bin/env python3
"""
Excel Formula Recalculation Script

Recalculates all formulas in an Excel file using LibreOffice and checks for errors.

Adapted from Anthropic's xlsx skill: https://github.com/anthropics/skills/tree/main/skills/xlsx

Usage:
    python recalc.py <excel_file> [timeout_seconds]

Examples:
    python recalc.py output.xlsx
    python recalc.py model.xlsx 60

Returns JSON with:
    - status: 'success' or 'errors_found'
    - total_errors: Count of formula errors
    - total_formulas: Count of formulas in file
    - error_summary: Breakdown by error type with locations

Requirements:
    - LibreOffice installed (automatically configured on first run)
    - openpyxl: pip install openpyxl
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured."""
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser(
            '~/Library/Application Support/LibreOffice/4/user/basic/Standard'
        )
    else:
        macro_dir = os.path.expanduser(
            '~/.config/libreoffice/4/user/basic/Standard'
        )

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    # Check if macro already exists
    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True

    # Initialize LibreOffice config if needed
    if not os.path.exists(macro_dir):
        subprocess.run(
            ['soffice', '--headless', '--terminate_after_init'],
            capture_output=True,
            timeout=10
        )
        os.makedirs(macro_dir, exist_ok=True)

    # Create the macro
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors.

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)

    Returns:
        dict with error locations and counts
    """
    if not OPENPYXL_AVAILABLE:
        return {'error': 'openpyxl is required. Install with: pip install openpyxl'}

    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    # Setup LibreOffice macro
    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}

    # Build command
    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    # Add timeout wrapper
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # Check for gtimeout on macOS
            try:
                subprocess.run(
                    ['gtimeout', '--version'],
                    capture_output=True,
                    timeout=1,
                    check=False
                )
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd

    # Run recalculation
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0 and result.returncode != 124:  # 124 = timeout
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly'}
        else:
            return {'error': error_msg}

    # Check for Excel errors in the recalculated file
    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            '#VALUE!', '#DIV/0!', '#REF!',
            '#NAME?', '#NULL!', '#NUM!', '#N/A'
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        # Scan all cells
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        # Build result
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }

        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }

        # Count formulas
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()

        result['total_formulas'] = formula_count

        return result

    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Excel Formula Recalculation Script")
        print("")
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("")
        print("Recalculates all formulas in an Excel file using LibreOffice")
        print("")
        print("Returns JSON with:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        print("")
        print("Example:")
        print("  python recalc.py output.xlsx")
        print("  python recalc.py model.xlsx 60")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
