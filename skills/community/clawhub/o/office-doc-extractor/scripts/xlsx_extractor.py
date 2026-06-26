#!/usr/bin/env python3
"""
XLSX to Markdown extractor using bundled openpyxl (pure Python).
"""

import sys
import os
from pathlib import Path

# Add bundled openpyxl to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

def extract_xlsx(input_path: str, output_path: str = None) -> str:
    """Extract text from XLSX and convert to Markdown tables."""
    input_file = Path(input_path)
    if not input_file.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    if output_path is None:
        output_path = input_file.with_suffix('.md')
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    md_lines = []
    md_lines.append(f"# {input_file.stem}")
    md_lines.append('')
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        md_lines.append(f"## Sheet: {sheet_name}")
        md_lines.append('')
        
        # Build table
        rows = []
        max_cols = 0
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                val = cell.value if cell.value is not None else ''
                row_data.append(str(val))
            if row_data:
                rows.append(row_data)
                max_cols = max(max_cols, len(row_data))
        
        if not rows:
            md_lines.append('*Empty sheet*')
            md_lines.append('')
            continue
        
        # Pad rows to equal length
        for row in rows:
            while len(row) < max_cols:
                row.append('')
        
        # Write markdown table
        if rows:
            header = rows[0]
            md_lines.append('| ' + ' | '.join(header) + ' |')
            md_lines.append('|' + '|'.join(['---'] * max_cols) + '|')
            for row in rows[1:]:
                md_lines.append('| ' + ' | '.join(row) + ' |')
        
        md_lines.append('')
    
    markdown = '\n'.join(md_lines)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(markdown)
    
    return str(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 xlsx_extractor.py <input.xlsx> [-o output.md]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output = None
    if '-o' in sys.argv:
        idx = sys.argv.index('-o')
        output = sys.argv[idx + 1]
    
    result = extract_xlsx(input_file, output)
    print(f"✅ Converted: {input_file} → {result}")
