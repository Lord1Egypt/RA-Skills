#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

EXPECTED_CORE = ["时刻", "统调负荷", "省间联络线", "总出力", "非现货机组出力", "新能源", "水电", "抽蓄"]


def find_source_sheet(wb, preferred: Optional[str]):
    if preferred:
        if preferred not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {preferred}")
        return wb[preferred]

    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if all(h in headers for h in EXPECTED_CORE):
            data_rows = 0
            for r in range(2, ws.max_row + 1):
                vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in vals):
                    data_rows += 1
            if data_rows == 96:
                return ws
    raise ValueError("No 96-point source sheet found. Use --source-sheet to specify one.")


def read_headers(ws) -> List[str]:
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def read_rows(ws) -> List[List[float]]:
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in vals):
            rows.append(vals)
    return rows


def infer_time_labels(rows: List[List[object]]) -> List[int]:
    return list(range(1, 25))


def average(values):
    return sum(values) / len(values)


def copy_sheet_data(src_ws, dst_ws):
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            dst_ws.cell(r, c, src_ws.cell(r, c).value)


def set_widths(ws, columns: int):
    for i in range(1, columns + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['A'].width = 12


def build_output(src_ws, output_path: Path):
    headers = read_headers(src_ws)
    rows = read_rows(src_ws)
    if len(rows) != 96:
        raise ValueError(f"Expected 96 non-empty data rows, got {len(rows)}")

    wb = Workbook()
    ws96 = wb.active
    ws96.title = "source96"
    ws24 = wb.create_sheet("result24")
    wsf = wb.create_sheet("formula24")

    copy_sheet_data(src_ws, ws96)

    for c, h in enumerate(headers, start=1):
        ws24.cell(1, c, h)
        wsf.cell(1, c, h)

    time_labels = infer_time_labels(rows)
    for hour in range(24):
        chunk = rows[hour * 4:(hour + 1) * 4]
        row_idx = hour + 2
        ws24.cell(row_idx, 1, time_labels[hour])
        wsf.cell(row_idx, 1, time_labels[hour])
        for col_idx in range(2, len(headers) + 1):
            vals = [chunk[i][col_idx - 1] for i in range(4)]
            ws24.cell(row_idx, col_idx, average(vals))
            col_letter = get_column_letter(col_idx)
            wsf.cell(row_idx, col_idx, f"=SUM(OFFSET(source96!${col_letter}$2,(ROW()-2)*4,0,4,1))/4")

    for ws in (ws96, ws24, wsf):
        set_widths(ws, len(headers))

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert electricity-trading 96-point data to 24-point hourly averages.")
    parser.add_argument("input", help="Path to source xlsx")
    parser.add_argument("output", help="Path to output xlsx")
    parser.add_argument("--source-sheet", help="Explicit source sheet name")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    wb = load_workbook(input_path, data_only=True)
    src_ws = find_source_sheet(wb, args.source_sheet)
    build_output(src_ws, output_path)
    print(output_path)


if __name__ == "__main__":
    main()
