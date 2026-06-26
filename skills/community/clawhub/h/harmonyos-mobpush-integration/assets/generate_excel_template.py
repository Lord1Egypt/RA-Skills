#!/usr/bin/env python3
"""
生成 MobPush 鸿蒙端配置模板 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 40

    ws.freeze_panes = 'A2'
    return ws


def create_basic_info_sheet(wb):
    """创建基础信息 Sheet"""
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "MobPush应用Key，从MobTech官网(https://www.mob.com/)注册应用获取", ""],
        ["appSecret", "MobPush应用密钥，与appKey一同获取", ""],
        ["包名", "鸿蒙应用包名，如：com.example.app", ""],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)

    for row in range(2, 5):
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, color="FF0000")

    return ws


def create_huawei_sheet(wb):
    """创建华为推送 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用华为推送通道（填写是或否）", "否"],
        ["Client ID", "华为 Client ID，从 AppGallery Connect 获取", ""],
        ["", "", ""],
        ["获取地址", "https://developer.huawei.com/consumer/cn/service/josp/agc/index.html", ""],
        ["注意事项", "1. 登录 AppGallery Connect -> 我的项目 -> 项目设置 -> 常规 -> 应用\n2. 获取 Client ID 后填写到上方\n3. 需同时在 MobTech 后台配置鸿蒙厂商参数", ""],
    ]
    return create_sheet_with_header(wb, "华为推送", headers, data)


def create_privacy_sheet(wb):
    """创建隐私合规 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["合规要求", "根据 MobTech 隐私合规要求和工信部相关规范，使用 MobPush 需要在用户同意隐私政策后才能初始化 SDK"],
        ["代码位置", "在用户点击隐私政策\"同意\"按钮的回调中添加授权代码"],
        ["调用时机", "必须先展示隐私政策弹窗，用户点击\"同意\"后才能调用"],
        ["授权代码", "ZztSDK.submitPolicyGrantResult(true);\n\n如使用主动控制器方案：\nZztSDK.submitPolicyGrantResult(true, new MyCustomController());"],
        ["自定义控制器", "继承 ZztCustomController 类，重写数据采集开关方法\n详见：https://www.mob.com/wiki/detailed?wiki=745&id=136"],
        ["隐私政策内容", "在 App 隐私政策中应包含：\n- SDK名称：MobPush\n- 第三方主体：上海掌之淘信息技术有限公司\n- 使用目的：提供消息推送服务\n- 处理的个人信息类型：系统运行信息、网络状态信息、设备标识信息（OAID）、地理位置信息\n- 官网：https://www.mob.com/\n- 隐私政策：https://policy.zztfly.com/sdk/mobpush/privacy"],
    ]

    ws = wb.create_sheet(title="隐私合规")

    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 85
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 55

    return ws


def create_instructions_sheet(wb):
    """创建填写说明 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 先在 MobTech 官网(https://www.mob.com/) 注册应用，获取 appKey 和 appSecret\n2. 填写基础信息（必填）\n3. 如需使用华为推送，在\"华为推送\"Sheet 中填写参数\n4. 不需要华为推送可留空或填写\"否\"\n5. 填写完成后告诉我\"填好了\"，我将继续下一步"],
        ["必填项", "基础信息 Sheet 中的 appKey、appSecret、包名是必填项"],
        ["华为推送说明", "华为推送可提升鸿蒙设备的推送到达率：\n- 需要在 AppGallery Connect 注册应用并获取 Client ID\n- 在 MobTech 后台配置鸿蒙厂商参数\n- 在 module.json5 中配置 client_id metadata"],
        ["权限说明", "MobPush 鸿蒙端需要以下权限：\n- ohos.permission.INTERNET（必选）：网络连接\n- ohos.permission.GET_NETWORK_INFO（必选）：网络状态检测\n- ohos.permission.APP_TRACKING_CONSENT（可选）：设备标识\n- ohos.permission.APPROXIMATELY_LOCATION（可选）：地理位置\n- ohos.permission.GET_WIFI_INFO（可选）：WiFi信息"],
        ["隐私合规", "必须在用户同意隐私政策后才能初始化 SDK\n详见\"隐私合规\"Sheet"],
        ["常见问题", "收不到推送：检查包名是否与后台一致\n华为推送失败：检查 Client ID 是否正确\n无点击回调：检查是否在 UIAbility 中调用了 notificationClickAck()"],
    ]

    ws = wb.create_sheet(title="填写说明")

    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 85
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 65

    return ws


def main():
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_huawei_sheet(wb)
    create_instructions_sheet(wb)

    target_order = ["基础信息", "隐私合规", "华为推送", "填写说明"]
    for idx, name in enumerate(target_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    output_path = Path(__file__).resolve().parent / "MobPush_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
