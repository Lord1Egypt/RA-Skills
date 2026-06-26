#!/usr/bin/env python3
"""
生成 iOS SMSSDK 最小配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def style_header(ws, headers, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.freeze_panes = "A2"


def auto_width(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


def create_basic_info_sheet(wb):
    ws = wb.create_sheet(title="基础信息")
    headers = ["配置项", "说明", "您的填写"]
    style_header(ws, headers, "4472C4")

    rows = [
        ["appKey", "MobTech 后台申请得到的 AppKey，必填", ""],
        ["appSecret", "MobTech 后台申请得到的 AppSecret，必填", ""],
        ["短信签名是否已申请", "上线前必须申请自定义短信签名；默认签名仅用于测试。只填 是 或 否", "否"],
        ["是否需要语音验证码", "本次是否同时接入语音验证码。只填 是 或 否", "否"],
        ["是否需要本机号码认证", "本次是否同时接入 getMobileAuthTokenWith 和 verifyMobileWithPhone。只填 是 或 否", "否"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, 7):
        ws.cell(row=row, column=1).font = Font(bold=True, color="C00000")

    auto_width(ws, {"A": 28, "B": 82, "C": 22})
    return ws


def create_privacy_sheet(wb):
    ws = wb.create_sheet(title="隐私与上架")
    headers = ["事项", "说明"]
    style_header(ws, headers, "ED7D31")

    rows = [
        ["隐私政策", "首次冷启动时应向用户展示《隐私政策》，并在同意后才能调用 SMSSDK 相关能力。"],
        ["隐私回传", "用户同意后，必须调用 MobSDK 的 uploadPrivacyPermissionStatus 接口，再使用验证码能力。"],
        ["严格隐私配置", "默认 plist 需配置 MOBNetLater = 2。"],
        ["App Store 隐私标签", "需参考 MobTech 提供的 App Store Connect 后台隐私数据项配置文档进行配置。"],
        ["短信签名", "默认签名仅用于测试，不保证到达率；上线前必须申请自定义签名。"],
        ["扩展业务主动控制器", "本轮不主动收集该配置；如未来需要控制地理位置、WiFi、IP 等扩展业务数据采集，可参考最终文档说明。"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 24, "B": 95})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 38
    return ws


def create_instructions_sheet(wb):
    ws = wb.create_sheet(title="填写说明")
    headers = ["说明项", "详细内容"]
    style_header(ws, headers, "70AD47")

    rows = [
        ["填写顺序", "1. 在 MobTech 后台申请 appKey / appSecret；2. 确认是否已申请自定义短信签名；3. 只填写基础信息 Sheet；4. 填完后告诉 Agent “填好了”。"],
        ["最小原则", "不要在表里填写 Bundle ID、Target、Info.plist 路径、入口类名、手机号输入框路径，这些应由 Agent 扫描工程自动推断。"],
        ["语言支持", "Agent 扫描项目时会记录 Objective-C、Swift 或混编状态，并按项目语言生成对应接入代码。"],
        ["可选能力", "语音验证码、本机号码认证、扩展业务主动控制器默认不作为前置项；需要时再按官方文档追加。"],
        ["参考链接", "创建应用：https://www.mob.com/wiki/detailed?wiki=539&id=23\n集成指南：https://www.mob.com/wiki/detailed?wiki=110&id=23\nSDK API：https://www.mob.com/wiki/detailed?wiki=467&id=23\n合规指南：https://www.mob.com/wiki/detailed?wiki=211&id=23"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 18, "B": 105})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 55
    return ws


def main():
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    output_path = Path(__file__).resolve().parent / "SMSSDK_iOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
