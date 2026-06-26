#!/usr/bin/env python3
"""
生成 iOS FlyVerify 最小配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def style_header(ws, headers, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.freeze_panes = "A2"


def auto_width(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


def create_basic_info_sheet(wb):
    ws = wb.create_sheet(title="基础信息")
    headers = ["配置项", "说明", "您的填写"]
    style_header(ws, headers, "4472C4")

    rows = [
        ["appKey", "MobTech 后台申请得到的 AppKey，必填", ""],
        ["appSecret", "MobTech 后台申请得到的 AppSecret，必填", ""],
        ["秒验审核", "是否已在 MobTech 后台提交秒验审核并通过，只填 是 或 否", "否"],
        ["授权页自定义", "是否需要本次同时改 FlyVerify 授权页 UI，只填 是 或 否", "否"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, 6):
        ws.cell(row=row, column=1).font = Font(bold=True, color="C00000")

    auto_width(ws, {"A": 24, "B": 70, "C": 20})
    return ws


def create_privacy_sheet(wb):
    ws = wb.create_sheet(title="隐私与上架")
    headers = ["事项", "说明"]
    style_header(ws, headers, "ED7D31")

    rows = [
        ["隐私政策", "首次冷启动时应向用户展示《隐私政策》，并在同意后才能调用 FlyVerify 相关能力。"],
        ["隐私回传", "用户同意后，必须先调用 FlyVerifyC 的 agreePrivacy 接口，再使用 preLogin 或授权页。"],
        ["严格隐私模式", "默认 plist 需配置 FlyVerifyPLevel = 2。"],
        ["App Store 隐私标签", "需参考 MobTech 提供的 App Store Connect 后台隐私数据项配置文档进行配置。"],
        ["服务端职责", "客户端拿到 token 后，手机号置换属于后续服务端步骤，本轮客户端集成只保留注释占位。"],
        ["网络要求", "一键登录依赖运营商网关取号，测试时必须开启移动蜂窝网络。"],
        ["扩展业务主动控制器", "本轮不主动收集该配置；如未来需要控制地理位置、WiFi、IP 等扩展业务数据采集，可参考最终文档说明。"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 22, "B": 90})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 38
    return ws


def create_instructions_sheet(wb):
    ws = wb.create_sheet(title="填写说明")
    headers = ["说明项", "详细内容"]
    style_header(ws, headers, "70AD47")

    rows = [
        ["填写顺序", "1. 在 MobTech 后台申请 appKey / appSecret；2. 提交秒验审核；3. 只填写基础信息 Sheet；4. 填完后告诉 Agent “填好了”。"],
        ["最小原则", "不要在表里填写 Bundle ID、Target、Info.plist 路径、入口类名，这些应由 Agent 扫描工程自动推断。"],
        ["语言支持", "Agent 扫描项目时会记录 Objective-C、Swift 或混编状态，并按项目语言生成对应接入代码。"],
        ["授权页自定义", "如果本次只想先跑通一键登录，建议保持“否”；后续需要时再改 UI。"],
        ["参考链接", "创建应用流程：https://www.mob.com/wiki/detailed?wiki=538&id=78\n秒验审核流程：https://www.mob.com/wiki/detailed?wiki=159&id=78\n服务端接入文档：https://www.mob.com/wiki/detailed?wiki=157&id=78"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 18, "B": 95})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 52
    return ws


def main():
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    output_path = Path(__file__).resolve().parent / "FlyVerify_iOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
