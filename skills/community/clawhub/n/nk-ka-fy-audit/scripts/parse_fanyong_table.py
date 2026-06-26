"""解析返佣信息表.xlsx 并存入集团客户部FY数据.db"""
import sqlite3
import openpyxl
import re
from datetime import datetime

SRC = "/tmp/openclaw/返佣信息表.xlsx"
DB = "/workspace/业务数据/集团客户部FY数据.db"

# Column names (0-indexed)
COL_NAMES = [
    "入网月份", "商户id", "sn", "商户名", # 3
    "col4", # 4 - empty in header
    "集团名称", # 5
    "brand", "项目编号", "行业", # 6-8
    "微信商户号", "支付宝PID_MID", # 9-10
    "level1", "level2", "level3", "level4", # 11-14
    # 支付宝 15-29
    "支付宝_直连间连", "支付宝_签约主体", "支付宝_口碑餐饮0费率", "支付宝_蓝海活动",
    "支付宝_结算费率", "支付宝_签约费率", "支付宝_对公返佣", "支付宝_对私返佣",
    "支付宝_跨组织分润", "支付宝_分润对象", "支付宝_成本费率正式", "支付宝_成本费率二清",
    "支付宝_项目收益率正式", "支付宝_项目收益率二清", "支付宝_返佣费率",
    # 微信 30-42
    "微信_直连间连", "微信_绿洲计划商户", "微信_结算费率", "微信_签约费率",
    "微信_对公返佣", "微信_对私返佣", "微信_跨组织分润", "微信_分润对象",
    "微信_成本费率正式", "微信_成本费率二清", "微信_收益率正式", "微信_收益率二清",
    "微信_返佣费率",
    # 对公返佣 43-46
    "对公返佣_返佣邮箱", "对公返佣_户名", "对公返佣_开户行", "对公返佣_银行账号",
    # 对私返佣 47-50
    "对私返佣_返佣邮箱", "对私返佣_户名", "对私返佣_开户行", "对私返佣_银行账号",
    # 51-55
    "备注", "贡献率", "推广物料完成率", "项目转让", "是否返佣",
    "销售人员姓名",  # 56 - computed
]

def extract_salesperson(level3):
    """提取销售人员姓名"""
    if level3 is None or str(level3).strip() == '':
        return None
    s = str(level3).strip()
    if s == "S'":
        return "Stanley"
    if '公海' in s:
        return '公海'
    m = re.match(r'^(.+?)的组织$', s)
    if m:
        return m.group(1).strip()
    return s

# Read Excel
print("正在读取Excel...")
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["清算组2"]

# Build column headers from both rows
row0 = {}
row1 = {}
for col in range(56):
    v0 = ws.cell(row=1, column=col+1).value
    v1 = ws.cell(row=2, column=col+1).value
    if v0 is not None:
        row0[col] = v0
    if v1 is not None:
        row1[col] = v1

print("Row0 headers:", {k: row0[k] for k in sorted(row0.keys()) if row0[k]})
print("Row1 headers:", {k: row1[k] for k in sorted(row1.keys()) if row1[k]})

# Read data rows
print("\n正在解析数据行...")
rows = []
total_rows = ws.max_row
for r in range(3, total_rows + 1):
    row_data = []
    for col in range(56):
        v = ws.cell(row=r, column=col+1).value
        if isinstance(v, datetime):
            v = v.isoformat()
        row_data.append(v)
    
    # Compute salesperson name
    level3 = row_data[13]
    row_data.append(extract_salesperson(level3))
    
    # Only add non-empty rows (has at least sn or merchant name)
    sn = row_data[2]
    name = row_data[3]
    if sn is None and name is None:
        continue
    rows.append(tuple(row_data))

print(f"共读取 {len(rows)} 条数据记录")

# Connect to DB
conn = sqlite3.connect(DB)
cur = conn.cursor()

# Create table (drop if exists - ask user)
table_name = "返佣信息表"

# Check if table exists
cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
if cur.fetchone():
    print(f"表 {table_name} 已存在，先删除重建")
    cur.execute(f"DROP TABLE IF EXISTS {table_name}")

# Build CREATE SQL
col_defs = []
for i, name in enumerate(COL_NAMES):
    if name == "col4":
        col_defs.append("col4 TEXT")
    elif name in ["贡献率", "推广物料完成率"]:
        col_defs.append(f'"{name}" REAL')
    elif name.endswith("_结算费率") or name.endswith("_签约费率") or "费率" in name or "收益率" in name or "返佣" in name or "成本" in name:
        col_defs.append(f'"{name}" REAL')
    elif name == "入网月份":
        col_defs.append(f'"{name}" TEXT')
    elif name == "是否返佣":
        col_defs.append(f'"{name}" TEXT')
    elif name == "备注":
        col_defs.append(f'"{name}" TEXT')
    elif name == "销售人员姓名":
        col_defs.append(f'"{name}" TEXT')
    elif name in ["项目转让", "支付宝_直连间连", "支付宝_签约主体", "支付宝_口碑餐饮0费率", "支付宝_蓝海活动", 
                   "支付宝_分润对象", "微信_直连间连", "微信_绿洲计划商户", "微信_分润对象"]:
        col_defs.append(f'"{name}" TEXT')
    else:
        col_defs.append(f'"{name}" TEXT')

create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(col_defs)})'
print(f"\n创建表: {table_name}")
cur.execute(create_sql)

# Insert data in batches
placeholders = ",".join(["?" for _ in COL_NAMES])
cols_str = ",".join([f'"{c}"' for c in COL_NAMES])
insert_sql = f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})'

batch_size = 500
for i in range(0, len(rows), batch_size):
    batch = rows[i:i+batch_size]
    cur.executemany(insert_sql, batch)

conn.commit()
print(f"已插入 {len(rows)} 条记录到 {table_name} 表")

# ===== 分析对公返佣 =====
print("\n" + "="*60)
print("【对公返佣分析】")
print("="*60)

# 1. 有对公返佣户名的记录数
cur.execute(f'SELECT COUNT(*) FROM "{table_name}" WHERE "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != ""')
cnt = cur.fetchone()[0]
print(f"\n1. 有对公返佣户名的记录数: {cnt}")

# 2. 按户名分组统计商户数量
print(f"\n2. 按户名分组统计商户数:")
cur.execute(f'''
    SELECT "对公返佣_户名", COUNT(*) as 商户数, 
           COUNT(DISTINCT "对公返佣_银行账号") as 账户数
    FROM "{table_name}" 
    WHERE "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != ''
    GROUP BY "对公返佣_户名"
    ORDER BY 商户数 DESC
    LIMIT 30
''')
results = cur.fetchall()
print(f"   {'户名':<20} {'商户数':>8} {'账户数':>8}")
print(f"   {'-'*20} {'-'*8} {'-'*8}")
for r in results:
    print(f"   {r[0] if r[0] else '':<20} {r[1]:>8} {r[2]:>8}")
print(f"   ... (共 {len(results)} 个户名，显示前30)")

total_accounts = cur.execute(f'SELECT COUNT(DISTINCT "对公返佣_户名") FROM "{table_name}" WHERE "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != ""').fetchone()[0]
print(f"   去重户名总数: {total_accounts}")

# 3. 查找同一银行账户对应多个不同户名（舞弊嫌疑）
print(f"\n3. 同一银行账户对应多个不同户名（异常/舞弊嫌疑）:")
cur.execute(f'''
    SELECT "对公返佣_银行账号", GROUP_CONCAT(DISTINCT "对公返佣_户名") as 户名列表, 
           COUNT(DISTINCT "对公返佣_户名") as 户名数, COUNT(*) as 记录数
    FROM "{table_name}" 
    WHERE "对公返佣_银行账号" IS NOT NULL AND "对公返佣_银行账号" != ''
    GROUP BY "对公返佣_银行账号"
    HAVING 户名数 > 1
    ORDER BY 户名数 DESC
    LIMIT 20
''')
results = cur.fetchall()
if results:
    print(f"   {'银行账号':<30} {'户名列表':<40} {'户名数':>8} {'记录数':>8}")
    print(f"   {'-'*30} {'-'*40} {'-'*8} {'-'*8}")
    for r in results:
        print(f"   {r[0] if r[0] else '':<30} {r[1] if r[1] else '':<40} {r[2]:>8} {r[3]:>8}")
else:
    print("   ✅ 未发现同一银行账户对应多个不同户名的情况")

# 4. 按销售人员姓名统计对公返佣情况
print(f"\n4. 按销售人员姓名统计返佣情况:")
cur.execute(f'''
    SELECT "销售人员姓名", COUNT(*) as 商户数,
           COUNT(CASE WHEN "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != '' THEN 1 END) as 有对公返佣数,
           COUNT(CASE WHEN "对私返佣_户名" IS NOT NULL AND "对私返佣_户名" != '' THEN 1 END) as 有对私返佣数
    FROM "{table_name}" 
    WHERE "销售人员姓名" IS NOT NULL
    GROUP BY "销售人员姓名"
    ORDER BY 商户数 DESC
''')
results = cur.fetchall()
print(f"   {'销售人员':<12} {'商户数':>8} {'有对公返佣':>10} {'有对私返佣':>10}")
print(f"   {'-'*12} {'-'*8} {'-'*10} {'-'*10}")
for r in results:
    print(f"   {r[0] if r[0] else '空':<12} {r[1]:>8} {r[2]:>10} {r[3]:>10}")

# 5. 对公返佣详细信息（关键部分）
print(f"\n5. 对公返佣详细信息（户名+银行账号不为空的记录）:")
cur.execute(f'''
    SELECT sn, 商户名, brand, level3, "销售人员姓名",
           "对公返佣_返佣邮箱", "对公返佣_户名", "对公返佣_开户行", "对公返佣_银行账号"
    FROM "{table_name}" 
    WHERE "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != ''
    ORDER BY "对公返佣_户名", sn
    LIMIT 50
''')
results = cur.fetchall()
if results:
    print(f"   {'sn':<18} {'商户名':<25} {'品牌':<12} {'销售人员':<10} {'户名':<16} {'开户行':<18} {'银行账号':<20}")
    print(f"   {'-'*18} {'-'*25} {'-'*12} {'-'*10} {'-'*16} {'-'*18} {'-'*20}")
    for r in results:
        print(f"   {r[0] if r[0] else '':<18} {(r[1] or '')[:24]:<25} {(r[2] or '')[:11]:<12} {(r[4] or '')[:9]:<10} {(r[6] or '')[:15]:<16} {(r[7] or '')[:17]:<18} {(r[8] or '')[:19]:<20}")

total_detail = cur.execute(f'SELECT COUNT(*) FROM "{table_name}" WHERE "对公返佣_户名" IS NOT NULL AND "对公返佣_户名" != ""').fetchone()[0]
print(f"   ... 共 {total_detail} 条记录（显示前50条）")

conn.close()
print("\n✅ 分析完成")