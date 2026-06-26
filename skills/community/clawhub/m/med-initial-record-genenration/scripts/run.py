#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence
from xml.etree import ElementTree as ET

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from gen_initial_record import generate_initial_record

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


class PreprocessError(ValueError):
    pass


SPEAKER_ALIASES = {
    "patient": "患者",
    "pt": "患者",
    "user": "患者",
    "client": "患者",
    "患者": "患者",
    "病人": "患者",
    "患方": "患者",
    "病友": "患者",
    "家属": "患者",
    "doctor": "医生",
    "dr": "医生",
    "physician": "医生",
    "assistant": "医生",
    "医生": "医生",
    "医师": "医生",
    "大夫": "医生",
    "问": "医生",
    "答": "患者",
}

SECTION_HINTS = ("主诉", "现病史", "既往史", "查体", "辅助检查", "诊断", "处理")
TABLE_SPEAKER_HEADERS = ("speaker", "role", "who", "name", "label", "说话人", "角色", "发言人")
TABLE_TEXT_HEADERS = ("text", "content", "utterance", "message", "dialogue", "内容", "文本", "发言")
TEXT_FILE_TYPES = {"txt", "md"}
IMAGE_FILE_TYPES = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
OFFICE_FILE_TYPES = {"doc", "docx"}
TABLE_FILE_TYPES = {"csv", "xlsx", "xls"}
SUPPORTED_FILE_TYPES = TEXT_FILE_TYPES | IMAGE_FILE_TYPES | OFFICE_FILE_TYPES | TABLE_FILE_TYPES | {
    "json",
    "pdf",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preprocess pdf/doc/docx/xls/xlsx/csv/txt/json input into dialogue text, then generate an initial outpatient record."
    )
    parser.add_argument("--input", required=True, help="Input file path.")
    parser.add_argument(
        "--input-type",
        default="auto",
        choices=["auto", *sorted(SUPPORTED_FILE_TYPES)],
        help="Explicit input type. Default: auto.",
    )
    parser.add_argument("--sheet", default="", help="Optional sheet name for xlsx/xls inputs.")
    parser.add_argument("--encoding", default="utf-8", help="Text encoding for txt/csv inputs. Default: utf-8.")
    parser.add_argument("--output", default="", help="Output path for generated record.")
    parser.add_argument("--diag-id", default="skill-diag", help="Dialogue ID used for backend services.")
    parser.add_argument("--timeout", type=int, default=0, help="HTTP request timeout seconds. 0 means wait forever.")
    parser.add_argument(
        "--save-prepared",
        action="store_true",
        help="Save the normalized dialogue text next to the output for debugging.",
    )
    return parser.parse_args()


def detect_input_type(path: Path, explicit: str) -> str:
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def extract_with_soffice(path: Path) -> str:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for office document conversion.")
    with tempfile.TemporaryDirectory(prefix="med-initial-record-") as tmp_dir:
        proc = subprocess.run(
            [
                office_bin,
                "--headless",
                "--convert-to",
                "txt:Text",
                "--outdir",
                tmp_dir,
                str(path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(f"Failed to convert office document to text: {proc.stderr.strip() or proc.stdout.strip()}")
        return out_path.read_text(encoding="utf-8", errors="replace")


def extract_pdf_text(path: Path) -> str:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            text = "\n\n".join(page for page in pages if page)
            if text.strip():
                return text
        except Exception:
            pass

    pdf_to_text = shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout
    raise PreprocessError("Unable to extract text from pdf.")


def extract_image_text(path: Path) -> str:
    tesseract_bin = shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError("tesseract not found for image OCR.")
    langs = detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


def detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


def shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def load_input_artifact(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "text", "text": read_text_file(path, encoding)}
    if input_type == "json":
        return {"kind": "json", "data": read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        return {"kind": "text", "text": extract_with_soffice(path)}
    if input_type == "docx":
        return {"kind": "text", "text": extract_docx_text(path)}
    if input_type == "doc":
        return {"kind": "text", "text": extract_with_soffice(path)}
    if input_type == "pdf":
        return {"kind": "text", "text": extract_pdf_text(path)}
    if input_type in IMAGE_FILE_TYPES:
        return {"kind": "text", "text": extract_image_text(path)}
    raise PreprocessError(f"Unsupported input type: {input_type}")


def normalize_speaker(raw: str) -> Optional[str]:
    key = raw.strip().lower()
    if not key:
        return None
    return SPEAKER_ALIASES.get(key)


def looks_like_record_text(text: str) -> bool:
    hits = sum(1 for marker in SECTION_HINTS if marker in text)
    return hits >= 2


def dialogue_text_from_messages(items: Iterable[Any]) -> str:
    lines: List[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        speaker = item.get("speaker") or item.get("role") or item.get("from") or item.get("name") or item.get("label")
        content = item.get("text") or item.get("content") or item.get("message") or item.get("utterance") or item.get("value")
        speaker_norm = normalize_speaker(str(speaker or ""))
        if speaker_norm and isinstance(content, str) and content.strip():
            lines.append(f"{speaker_norm}：{content.strip()}")
    if len(lines) < 2:
        raise PreprocessError("JSON conversation did not contain at least two recognizable dialogue turns.")
    return "\n".join(lines)


def dialogue_text_from_json(data: Any) -> str:
    if isinstance(data, str):
        return normalize_dialogue_text(data)
    if isinstance(data, list):
        return dialogue_text_from_messages(data)
    if not isinstance(data, dict):
        raise PreprocessError("JSON input must be a string, list, or object.")

    for key in ("dialogue", "diag", "text", "content", "record"):
        value = data.get(key)
        if isinstance(value, str) and value.strip():
            return normalize_dialogue_text(value)

    for key in ("messages", "conversation", "turns", "dialogues", "items"):
        value = data.get(key)
        if isinstance(value, list):
            return dialogue_text_from_messages(value)

    if "speaker" in data and "text" in data:
        return dialogue_text_from_messages([data])

    raise PreprocessError("Unsupported JSON schema for dialogue generation.")


def table_to_dialogue(table: Dict[str, Any]) -> Optional[str]:
    rows = table.get("rows") or []
    if not rows:
        return None

    header_row = rows[0]
    header_map = {normalize_header(cell): idx for idx, cell in enumerate(header_row) if cell.strip()}
    speaker_idx = first_matching_index(header_map, TABLE_SPEAKER_HEADERS)
    text_idx = first_matching_index(header_map, TABLE_TEXT_HEADERS)

    candidate_rows = rows[1:] if speaker_idx is not None and text_idx is not None else rows
    if speaker_idx is None or text_idx is None:
        speaker_idx, text_idx = 0, 1

    lines: List[str] = []
    for row in candidate_rows:
        if max(speaker_idx, text_idx) >= len(row):
            continue
        speaker = normalize_speaker(row[speaker_idx])
        content = row[text_idx].strip()
        if speaker and content:
            lines.append(f"{speaker}：{content}")
    if len(lines) >= 2:
        return "\n".join(lines)
    return None


def first_matching_index(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(normalize_header(alias))
        if idx is not None:
            return idx
    return None


def normalize_dialogue_text(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        raise PreprocessError("Input text is empty.")
    if looks_like_record_text(text):
        raise PreprocessError("Input looks like a medical record rather than a doctor-patient dialogue.")

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    normalized: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    pattern = re.compile(r"^\s*([A-Za-z\u4e00-\u9fff]{1,12})\s*[:：]\s*(.+?)\s*$")

    for line in lines:
        match = pattern.match(line)
        if match:
            speaker = normalize_speaker(match.group(1))
            if speaker:
                if current:
                    normalized.append(current)
                current = {"speaker": speaker, "parts": [match.group(2).strip()]}
                continue
        if current is not None:
            current["parts"].append(line)

    if current:
        normalized.append(current)

    if len(normalized) < 2:
        raise PreprocessError("Did not find enough recognizable dialogue turns in the input.")

    result = []
    for item in normalized:
        content = " ".join(part for part in item["parts"] if part).strip()
        if content:
            result.append(f"{item['speaker']}：{content}")

    if len(result) < 2:
        raise PreprocessError("Did not find enough recognizable dialogue turns after normalization.")
    return "\n".join(result)


def preprocess_to_dialogue(path: Path, input_type: str, encoding: str, sheet_name: str) -> str:
    artifact = load_input_artifact(path, input_type, encoding, sheet_name)
    kind = artifact["kind"]
    if kind == "text":
        return normalize_dialogue_text(artifact["text"])
    if kind == "json":
        return dialogue_text_from_json(artifact["data"])
    if kind == "tables":
        for table in artifact["tables"]:
            dialogue = table_to_dialogue(table)
            if dialogue:
                return dialogue
        raise PreprocessError("Could not find speaker/text columns or recognizable dialogue rows in the table input.")
    raise PreprocessError(f"Unsupported artifact kind: {kind}")


def build_prepared_path(output_path: Path) -> Path:
    suffix = output_path.suffix or ".txt"
    return output_path.with_name(f"{output_path.stem}.prepared{suffix}")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"✗ Error: Input file not found: {input_path}", file=sys.stderr)
        return 1

    output_path = Path(args.output) if args.output else Path("..") / "runs" / "med-initial-record-gen" / "record.txt"

    try:
        input_type = detect_input_type(input_path, args.input_type)
        dialogue_text = preprocess_to_dialogue(input_path, input_type, args.encoding, args.sheet)
    except Exception as exc:
        print(f"✗ Preprocess Error: {exc}", file=sys.stderr)
        return 1

    if args.save_prepared:
        prepared_path = build_prepared_path(output_path)
        prepared_path.parent.mkdir(parents=True, exist_ok=True)
        prepared_path.write_text(dialogue_text, encoding="utf-8")
        print(f"✓ Prepared dialogue saved to: {prepared_path}")

    temp_path: Optional[Path] = None
    try:
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".txt", delete=False) as tmp:
            tmp.write(dialogue_text)
            temp_path = Path(tmp.name)
        generate_initial_record(
            input_path=str(temp_path),
            output_path=str(output_path),
            timeout=args.timeout,
            diag_id=args.diag_id,
        )
        print("\n✓ Initial visit record generated successfully!")
        return 0
    except Exception as exc:
        print(f"✗ Unexpected Error: {exc}", file=sys.stderr)
        return 1
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    sys.exit(main())
