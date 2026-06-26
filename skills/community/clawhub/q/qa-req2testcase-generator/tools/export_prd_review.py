#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PRD 质量审查报告 Excel 生成脚本 — 将 P0.5 输出 JSON 转换为可读 Excel 报告。

Sheet 结构：
  Sheet1: 审查概览（综合评分、等级、各维度分）
  Sheet2: 问题清单（阻塞/警告/建议三类问题，含位置/描述/建议）

依赖: openpyxl
用法: python3 export_prd_review.py --input p0.5_output.json --output prd_review_report.xlsx

退出码:
  0 = 成功
  1 = 输入文件不存在
  2 = JSON 解析失败
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少 openpyxl，请执行 pip install openpyxl", file=sys.stderr)
    sys.exit(3)

# ── 颜色常量 ──────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"   # 深蓝，表头背景
COLOR_HEADER_FONT = "FFFFFF"   # 白色，表头字体
COLOR_BLOCKING    = "FFE0E0"   # 浅红，阻塞问题行
COLOR_WARNING     = "FFF3CD"   # 浅黄，警告问题行
COLOR_INFO        = "E8F4FD"   # 浅蓝，建议优化行
COLOR_SCORE_A     = "C6EFCE"   # 绿，A级
COLOR_SCORE_B     = "FFEB9C"   # 黄，B级
COLOR_SCORE_C     = "FFCC99"   # 橙，C级
COLOR_SCORE_D     = "FFC7CE"   # 红，D级
COLOR_SECTION_BG  = "D6E4F0"   # 浅蓝，分区标题背景

GRADE_COLOR = {"A": COLOR_SCORE_A, "B": COLOR_SCORE_B, "C": COLOR_SCORE_C, "D": COLOR_SCORE_D}

DIMENSION_LABELS = {
    "completeness":          "完整性",
    "consistency":           "一致性",
    "clarity":               "清晰度",
    "testability":           "可测试性",
    "performance_security":  "性能/安全",
}

SEVERITY_LABELS = {
    "blocking": "🔴 阻塞",
    "warning":  "🟡 警告",
    "info":     "ℹ️ 建议",
}

ACTION_LABELS = {
    "PROCEED":          "✅ 继续执行",
    "REVIEW_REQUIRED":  "⚠️ 建议补充后再执行",
    "BLOCKED":          "🚫 暂停，必须澄清阻塞问题",
}


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_style(ws, cell_ref, text, bg=COLOR_HEADER_BG, font_size=11, bold=True):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(bold=bold, color=COLOR_HEADER_FONT, size=font_size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell


def data_cell(ws, cell_ref, text, bg=None, bold=False, wrap=True, align="left"):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(bold=bold, size=10)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()
    return cell


def build_overview_sheet(ws, report):
    ws.title = "审查概览"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20

    grade = report.get("grade", "?")
    grade_color = GRADE_COLOR.get(grade, "FFFFFF")
    action = report.get("recommended_action", "")
    action_label = ACTION_LABELS.get(action, action)

    # ── 标题行 ──
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = "📋 PRD 质量审查报告"
    cell.font = Font(bold=True, size=14, color=COLOR_HEADER_FONT)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── 基本信息 ──
    rows = [
        ("需求ID",    report.get("requirement_id", "-")),
        ("综合评分",  f"{report.get('overall_score', '-')} 分"),
        ("质量等级",  f"{grade}  {report.get('grade_label', '')}"),
        ("建议动作",  action_label),
        ("生成时间",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        data_cell(ws, f"A{i}", label, bold=True, align="right")
        cell = data_cell(ws, f"B{i}", value)
        if label == "质量等级":
            cell.fill = PatternFill("solid", fgColor=grade_color)
            cell.font = Font(bold=True, size=11)
        ws.merge_cells(f"B{i}:C{i}")
        ws.row_dimensions[i].height = 22

    # ── 各维度评分 ──
    row = len(rows) + 3
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "各维度评分"
    cell.font = Font(bold=True, size=11, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    header_style(ws, f"A{row}", "维度", bg="2E75B6")
    header_style(ws, f"B{row}", "得分（满分100）", bg="2E75B6")
    header_style(ws, f"C{row}", "权重", bg="2E75B6")
    ws.row_dimensions[row].height = 22
    row += 1

    dimensions = report.get("dimensions", {})
    for key, label in DIMENSION_LABELS.items():
        dim = dimensions.get(key, {})
        score = dim.get("score", "-")
        weight = dim.get("weight", "-")
        score_bg = COLOR_SCORE_A if isinstance(score, (int, float)) and score >= 80 else \
                   COLOR_SCORE_B if isinstance(score, (int, float)) and score >= 60 else COLOR_SCORE_D
        data_cell(ws, f"A{row}", label, bold=True)
        data_cell(ws, f"B{row}", f"{score} 分", bg=score_bg, align="center")
        data_cell(ws, f"C{row}", f"{int(weight * 100)}%" if isinstance(weight, float) else str(weight), align="center")
        ws.row_dimensions[row].height = 20
        row += 1

    # ── 问题汇总 ──
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "问题汇总"
    cell.font = Font(bold=True, size=11, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    summary_rows = [
        ("🔴 阻塞问题", len(report.get("blocking_issues", [])), COLOR_BLOCKING),
        ("🟡 警告问题", len(report.get("warning_issues", [])),  COLOR_WARNING),
        ("ℹ️ 建议优化", len(report.get("info_issues", [])),     COLOR_INFO),
    ]
    for label, count, bg in summary_rows:
        data_cell(ws, f"A{row}", label, bold=True)
        data_cell(ws, f"B{row}", f"{count} 条", bg=bg, align="center")
        ws.merge_cells(f"B{row}:C{row}")
        ws.row_dimensions[row].height = 20
        row += 1


def build_issues_sheet(ws, report):
    ws.title = "问题清单"

    col_widths = [8, 12, 20, 50, 40, 30]
    col_labels = ["序号", "严重等级", "位置", "问题描述", "建议/补充内容", "影响范围"]
    for i, (w, label) in enumerate(zip(col_widths, col_labels), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        header_style(ws, f"{get_column_letter(i)}1", label)
    ws.row_dimensions[1].height = 24

    row = 2
    all_issues = []

    for issue in report.get("blocking_issues", []):
        all_issues.append(("blocking", issue))
    for issue in report.get("warning_issues", []):
        all_issues.append(("warning", issue))
    for issue in report.get("info_issues", []):
        all_issues.append(("info", issue))

    # 也从 dimensions.*.issues 收集（如果 blocking/warning/info_issues 为空）
    if not all_issues:
        for key in DIMENSION_LABELS:
            dim = report.get("dimensions", {}).get(key, {})
            for issue in dim.get("issues", []):
                sev = issue.get("severity", "info")
                all_issues.append((sev, issue))

    if not all_issues:
        ws["A2"].value = "无问题"
        return

    for idx, (severity, issue) in enumerate(all_issues, start=1):
        bg = COLOR_BLOCKING if severity == "blocking" else \
             COLOR_WARNING  if severity == "warning"  else COLOR_INFO
        sev_label = SEVERITY_LABELS.get(severity, severity)

        data_cell(ws, f"A{row}", idx, align="center")
        data_cell(ws, f"B{row}", sev_label, bg=bg, bold=(severity == "blocking"), align="center")
        data_cell(ws, f"C{row}", issue.get("location", "-"), bg=bg)
        data_cell(ws, f"D{row}", issue.get("description", "-"), bg=bg)
        data_cell(ws, f"E{row}", issue.get("suggestion", issue.get("question", "-")), bg=bg)
        data_cell(ws, f"F{row}", issue.get("impact", "-"), bg=bg)
        ws.row_dimensions[row].height = 40
        row += 1


def main():
    parser = argparse.ArgumentParser(description="PRD 质量审查报告 Excel 生成")
    parser.add_argument("--input",  required=True, help="p0.5_output.json 路径")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 输入文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        with open(input_path, encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"错误: JSON 解析失败: {e}", file=sys.stderr)
        sys.exit(2)

    # 兼容两种结构：直接是 report，或包在 prd_quality_report 字段里
    report = data.get("prd_quality_report", data)

    wb = Workbook()
    ws1 = wb.active
    build_overview_sheet(ws1, report)

    ws2 = wb.create_sheet()
    build_issues_sheet(ws2, report)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"OK: {output_path}")


if __name__ == "__main__":
    main()
