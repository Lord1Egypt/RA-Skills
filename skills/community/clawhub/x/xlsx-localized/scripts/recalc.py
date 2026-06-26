"""
Excel Formula Recalculation Script (Windows-adapted)
Recalculates all formulas in an Excel file using LibreOffice.
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from office.soffice import get_soffice_env, _find_soffice

from openpyxl import load_workbook

# LibreOffice macro directory on Windows
MACRO_DIR_WINDOWS = os.path.join(
    os.environ.get("APPDATA", ""),
    "LibreOffice", "4", "user", "basic", "Standard"
)
MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _get_macro_dir():
    """Return the LibreOffice macro directory for the current platform."""
    if platform.system() == "Windows":
        return os.path.expandvars(MACRO_DIR_WINDOWS)
    elif platform.system() == "Darwin":
        return os.path.expanduser(MACRO_DIR_MACOS)
    else:
        return os.path.expanduser(MACRO_DIR_LINUX)


def _has_timeout_cmd():
    """Check if 'timeout' command is available (Windows has it natively)."""
    if platform.system() == "Windows":
        return True
    try:
        subprocess.run(["timeout", "--version"], capture_output=True, timeout=1, check=False, shell=True)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro():
    """Set up the LibreOffice recalc macro if not already present."""
    macro_dir = _get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text(encoding="utf-8")
    ):
        return True

    soffice = _find_soffice()

    if not os.path.exists(macro_dir):
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """Recalculate all formulas in the given Excel file."""
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    soffice = _find_soffice()

    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if _has_timeout_cmd():
        if platform.system() == "Windows":
            cmd = ["timeout", "/t", str(timeout)] + cmd
        else:
            cmd = ["timeout", str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    # timeout returns 124 on Linux; on Windows it's different
    if result.returncode != 0:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result_out = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result_out["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result_out["total_formulas"] = formula_count

        return result_out

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
