#!/usr/bin/env python3
"""Read table headers from XLSX or CSV with metadata extraction."""

import sys
import json
from pathlib import Path


def infer_domain(headers):
    """Infer academic domain from header patterns."""
    header_str = ' '.join(headers).lower()
    
    psych_markers = ['participants', '被试', 'experimental design', '实验设计', 'measures', '测量', 'manipulation', '操纵', 'effect size', '效应量', 'preregistration', '预注册']
    neuro_markers = ['imaging', '成像', 'preprocessing', '预处理', 'task paradigm', '任务范式', 'brain regions', '脑区', 'analysis type', '分析类型', 'behavioral-neural', '行为-神经']
    cs_markers = ['algorithm', '算法', 'model architecture', '模型架构', 'dataset', '数据集', 'training', '训练', 'metrics', '指标', 'code availability', '代码', 'ablation', '消融']
    brain_markers = ['species', '物种', 'recording method', '记录方法', 'stimulation', '刺激', 'cell type', '细胞类型', 'behavioral task', '行为任务']
    
    scores = {
        'psychology': sum(1 for m in psych_markers if m in header_str),
        'cognitive_neuroscience': sum(1 for m in neuro_markers if m in header_str),
        'computer_science': sum(1 for m in cs_markers if m in header_str),
        'brain_science': sum(1 for m in brain_markers if m in header_str),
    }
    
    max_score = max(scores.values())
    if max_score > 0:
        domain = max(scores, key=scores.get)
        return domain
    return 'general'


def infer_header_type(header):
    """Infer data type and constraints for a header."""
    h = header.lower()
    
    # Text fields
    if any(k in h for k in ['title', '标题', 'author', '作者', 'journal', '期刊', 'objective', '目标', 'conclusion', '结论', 'method', '方法', 'finding', '发现']):
        return {'type': 'text', 'constraints': 'free text, preserve original language'}
    
    # Numeric fields
    if any(k in h for k in ['year', '年份', 'size', '样本量', 'number', '数量']):
        return {'type': 'number', 'constraints': 'integer, 4 digits for year'}
    
    # Measurement fields
    if any(k in h for k in ['effect size', '效应量', 'correlation', '相关', 'metric', '指标', 'accuracy', '精确度']):
        return {'type': 'measurement', 'constraints': 'numeric with units/statistics'}
    
    # URL fields
    if any(k in h for k in ['doi', 'code', '代码', 'link', '链接', 'url']):
        return {'type': 'url', 'constraints': 'valid URL format'}
    
    # List fields
    if any(k in h for k in ['authors', 'author', '作者', 'region', '脑区', 'measure', '测量']):
        return {'type': 'list', 'constraints': 'semicolon-separated values'}
    
    return {'type': 'text', 'constraints': 'free text'}


def read_xlsx_headers(path):
    """Read headers from XLSX with metadata."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Check for second row with descriptions
        descriptions = {}
        if ws.max_row >= 2:
            for col_idx, header in enumerate(headers, 1):
                desc_cell = ws.cell(row=2, column=col_idx)
                if desc_cell.value and desc_cell.value != header:
                    descriptions[header] = str(desc_cell.value).strip()
        
        return headers, descriptions
    except ImportError:
        return ["[ERROR: openpyxl not installed]"], {}
    except Exception as e:
        return [f"[ERROR: Could not read XLSX: {e}]"], {}


def read_csv_headers(path):
    """Read headers from CSV with metadata."""
    import csv
    encodings = ['utf-8', 'gbk', 'latin-1']
    
    for encoding in encodings:
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                headers = [h.strip() for h in headers if h.strip()]
                
                # Check for second row with descriptions
                descriptions = {}
                try:
                    second_row = next(reader, [])
                    for i, header in enumerate(headers):
                        if i < len(second_row) and second_row[i] and second_row[i] != header:
                            descriptions[header] = second_row[i].strip()
                except StopIteration:
                    pass
                
                return headers, descriptions
        except UnicodeDecodeError:
            continue
        except Exception as e:
            return [f"[ERROR: Could not read CSV: {e}]"], {}
    
    return ["[ERROR: Could not decode CSV file]"], {}


def read_headers(table_path):
    """Route to appropriate reader based on file extension."""
    path = Path(table_path)
    ext = path.suffix.lower()
    
    if ext == '.xlsx':
        headers, descriptions = read_xlsx_headers(path)
    elif ext == '.csv':
        headers, descriptions = read_csv_headers(path)
    else:
        return [f"[ERROR: Unsupported table format: {ext}]"]
    
    if headers[0].startswith("[ERROR:"):
        return headers
    
    # Build metadata
    domain = infer_domain(headers)
    header_metadata = {}
    for header in headers:
        header_metadata[header] = {
            'type': infer_header_type(header)['type'],
            'constraints': infer_header_type(header)['constraints'],
            'description': descriptions.get(header, '')
        }
    
    result = {
        'headers': headers,
        'domain': domain,
        'header_metadata': header_metadata,
        'file': str(path),
        'format': ext.lstrip('.')
    }
    
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: read_table.py <table_path>", file=sys.stderr)
        sys.exit(1)
    
    result = read_headers(sys.argv[1])
    
    if isinstance(result, list):
        # Error case
        print("\n".join(result))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
