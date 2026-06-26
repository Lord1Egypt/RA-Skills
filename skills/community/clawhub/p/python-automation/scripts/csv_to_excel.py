#!/usr/bin/env python3
"""Convert CSV files to Excel (.xlsx) workbook."""

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl is required. Install: pip install openpyxl")
    sys.exit(1)


def csv_to_excel(input_path, output_path, sheet_name="Sheet1", delimiter=","):
    """Convert a single CSV to an Excel sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(input_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)


def merge_csvs(input_files, output_path, sheet_names=None, delimiter=","):
    """Merge multiple CSVs into one workbook, each as a separate sheet."""
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    for i, input_file in enumerate(input_files):
        name = (sheet_names[i] if sheet_names and i < len(sheet_names)
                else Path(input_file).stem[:31])  # Excel max sheet name = 31 chars
        ws = wb.create_sheet(title=name)

        with open(input_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert CSV files to Excel")
    parser.add_argument("input", nargs="+", help="CSV file(s) to convert")
    parser.add_argument("-o", "--output", default="output.xlsx", help="Output Excel file (default: output.xlsx)")
    parser.add_argument("-s", "--sheet", nargs="+", help="Sheet names (one per input file)")
    parser.add_argument("-d", "--delimiter", default=",", help="CSV delimiter (default: comma)")
    args = parser.parse_args()

    if len(args.input) == 1:
        csv_to_excel(args.input[0], args.output, args.sheet[0] if args.sheet else "Sheet1", args.delimiter)
    else:
        merge_csvs(args.input, args.output, args.sheet, args.delimiter)

    print(f"Created: {args.output}")


if __name__ == "__main__":
    main()
