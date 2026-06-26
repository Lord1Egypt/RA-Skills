"""
server/excel_export.py — Excel export for Friday Budgeting Pro.

Exports ledger data to a .xlsx file using openpyxl.

Public API
----------
build_workbook(conn, years=None) -> openpyxl.Workbook
    Build an in-memory workbook from the database.

export_to_file(conn, path, years=None) -> Path
    Build the workbook and save it atomically to *path* with mode 0o600.
"""

from __future__ import annotations

import os
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from server.sync_lock import acquire_sync_lock


class ExportBusy(Exception):
    """Raised when export_to_file cannot acquire the sync lock within the timeout."""


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_INCOME_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
_EXPENSE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
_BOLD_FONT = Font(bold=True)

_MONTHS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _fetch_ledgers(conn: sqlite3.Connection) -> list[dict]:
    """Return all ledgers as dicts with keys id, name."""
    rows = conn.execute("SELECT id, name FROM ledgers ORDER BY name").fetchall()
    return [dict(r) for r in rows]


def _fetch_line_items(conn: sqlite3.Connection, ledger_id: str) -> list[dict]:
    """Return all line items for a ledger."""
    rows = conn.execute(
        "SELECT id, name, item_type FROM line_items WHERE ledger_id = ? ORDER BY item_type, name",
        (ledger_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def _fetch_years(conn: sqlite3.Connection) -> list[int]:
    """Return all distinct years present in the transactions table."""
    rows = conn.execute(
        "SELECT DISTINCT CAST(strftime('%Y', date) AS INTEGER) AS yr FROM transactions ORDER BY yr"
    ).fetchall()
    return [r[0] for r in rows]


def _fetch_monthly_totals(
    conn: sqlite3.Connection,
    line_item_id: str,
    year: int,
) -> dict[int, float]:
    """Return {month: total_amount} for a given line_item and year.

    month is 1-indexed (1=Jan … 12=Dec).
    """
    rows = conn.execute(
        """
        SELECT CAST(strftime('%m', t.date) AS INTEGER) AS month,
               SUM(te.amount) AS total
        FROM transaction_entries te
        JOIN transactions t ON t.id = te.transaction_id
        WHERE te.line_item_id = ?
          AND strftime('%Y', t.date) = ?
        GROUP BY month
        """,
        (line_item_id, str(year)),
    ).fetchall()
    return {r[0]: r[1] for r in rows}


def _fetch_raw_transactions(
    conn: sqlite3.Connection,
    years: list[int] | None,
) -> list[dict]:
    """Return every classified transaction entry with full context."""
    if years:
        placeholders = ",".join("?" * len(years))
        year_strs = [str(y) for y in years]
        rows = conn.execute(
            f"""
            SELECT t.date, t.merchant, te.amount, l.name AS ledger,
                   li.name AS line_item
            FROM transaction_entries te
            JOIN transactions t ON t.id = te.transaction_id
            JOIN line_items li ON li.id = te.line_item_id
            JOIN ledgers l ON l.id = te.ledger_id
            WHERE strftime('%Y', t.date) IN ({placeholders})
            ORDER BY t.date, t.merchant
            """,
            year_strs,
        ).fetchall()
    else:
        rows = conn.execute("""
            SELECT t.date, t.merchant, te.amount, l.name AS ledger,
                   li.name AS line_item
            FROM transaction_entries te
            JOIN transactions t ON t.id = te.transaction_id
            JOIN line_items li ON li.id = te.line_item_id
            JOIN ledgers l ON l.id = te.ledger_id
            ORDER BY t.date, t.merchant
            """).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _write_ledger_sheet(
    ws,
    conn: sqlite3.Connection,
    ledger: dict,
    years: list[int],
) -> None:
    """Populate a ledger worksheet with one section per year."""
    line_items = _fetch_line_items(conn, ledger["id"])
    income_items = [li for li in line_items if li["item_type"] == "income"]
    expense_items = [li for li in line_items if li["item_type"] == "expense"]

    row = 1  # 1-indexed openpyxl rows

    for year in years:
        # --- Year header ---
        if len(years) > 1:
            ws.cell(row=row, column=1, value=str(year)).font = _BOLD_FONT
            row += 1

        # --- Column header row ---
        ws.cell(row=row, column=1, value="Line Item")
        for i, m in enumerate(_MONTHS, start=2):
            ws.cell(row=row, column=i, value=m)
        ws.cell(row=row, column=14, value="YTD")
        for col in range(1, 15):
            ws.cell(row=row, column=col).font = _BOLD_FONT
        row += 1

        # --- Income section ---
        _income_section_row = row  # reserved for future section anchoring
        for li in income_items:
            totals = _fetch_monthly_totals(conn, li["id"], year)
            ws.cell(row=row, column=1, value=li["name"]).fill = _INCOME_FILL
            ws.cell(row=row, column=1).font = _BOLD_FONT
            ytd = 0.0
            for m in range(1, 13):
                val = totals.get(m, 0.0)
                ws.cell(row=row, column=m + 1, value=val)
                ytd += val
            ws.cell(row=row, column=14, value=ytd)
            row += 1

        # --- Expense section ---
        for li in expense_items:
            totals = _fetch_monthly_totals(conn, li["id"], year)
            ws.cell(row=row, column=1, value=li["name"]).fill = _EXPENSE_FILL
            ws.cell(row=row, column=1).font = _BOLD_FONT
            ytd = 0.0
            for m in range(1, 13):
                val = totals.get(m, 0.0)
                ws.cell(row=row, column=m + 1, value=val)
                ytd += val
            ws.cell(row=row, column=14, value=ytd)
            row += 1

        # --- Net row ---
        # Compute net: income - expenses for each month
        income_by_month: dict[int, float] = {}
        for li in income_items:
            for m, v in _fetch_monthly_totals(conn, li["id"], year).items():
                income_by_month[m] = income_by_month.get(m, 0.0) + v

        expense_by_month: dict[int, float] = {}
        for li in expense_items:
            for m, v in _fetch_monthly_totals(conn, li["id"], year).items():
                expense_by_month[m] = expense_by_month.get(m, 0.0) + v

        net_cell = ws.cell(row=row, column=1, value="Net")
        net_cell.font = _BOLD_FONT
        net_ytd = 0.0
        for m in range(1, 13):
            net = income_by_month.get(m, 0.0) - expense_by_month.get(m, 0.0)
            ws.cell(row=row, column=m + 1, value=net)
            net_ytd += net
        ws.cell(row=row, column=14, value=net_ytd)
        row += 1

        # Blank separator row between years
        row += 1


def _write_summary_sheet(
    ws,
    conn: sqlite3.Connection,
    ledgers: list[dict],
    years: list[int],
) -> None:
    """Populate the Summary sheet: ledger × year income/expense/net."""
    # Header
    ws.cell(row=1, column=1, value="Ledger").font = _BOLD_FONT
    ws.cell(row=1, column=2, value="Year").font = _BOLD_FONT
    ws.cell(row=1, column=3, value="Income").font = _BOLD_FONT
    ws.cell(row=1, column=4, value="Expenses").font = _BOLD_FONT
    ws.cell(row=1, column=5, value="Net").font = _BOLD_FONT

    row = 2
    for ledger in ledgers:
        line_items = _fetch_line_items(conn, ledger["id"])
        income_items = [li for li in line_items if li["item_type"] == "income"]
        expense_items = [li for li in line_items if li["item_type"] == "expense"]

        for year in years:
            total_income = 0.0
            for li in income_items:
                totals = _fetch_monthly_totals(conn, li["id"], year)
                total_income += sum(totals.values())

            total_expense = 0.0
            for li in expense_items:
                totals = _fetch_monthly_totals(conn, li["id"], year)
                total_expense += sum(totals.values())

            net = total_income - total_expense

            ws.cell(row=row, column=1, value=ledger["name"])
            ws.cell(row=row, column=2, value=year)
            ws.cell(row=row, column=3, value=total_income)
            ws.cell(row=row, column=4, value=total_expense)
            ws.cell(row=row, column=5, value=net)
            row += 1


def _write_raw_sheet(ws, conn: sqlite3.Connection, years: list[int] | None) -> None:
    """Populate the Raw Transactions sheet."""
    headers = ["Date", "Merchant", "Amount", "Ledger", "Line Item"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = _BOLD_FONT

    rows = _fetch_raw_transactions(conn, years)
    for r_idx, row_data in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=row_data["date"])
        ws.cell(row=r_idx, column=2, value=row_data["merchant"])
        ws.cell(row=r_idx, column=3, value=row_data["amount"])
        ws.cell(row=r_idx, column=4, value=row_data["ledger"])
        ws.cell(row=r_idx, column=5, value=row_data["line_item"])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_workbook(
    conn: sqlite3.Connection,
    years: list[int] | None = None,
) -> Workbook:
    """Build and return an openpyxl Workbook from the database.

    Parameters
    ----------
    conn:
        An open sqlite3 connection.
    years:
        List of years to include.  If None, all years present in the data
        are included.

    Returns
    -------
    openpyxl.Workbook
    """
    effective_years: list[int] = years if years is not None else _fetch_years(conn)
    ledgers = _fetch_ledgers(conn)

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    # One sheet per ledger
    for ledger in ledgers:
        ws = wb.create_sheet(title=ledger["name"])
        _write_ledger_sheet(ws, conn, ledger, effective_years)

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    _write_summary_sheet(ws_summary, conn, ledgers, effective_years)

    # Raw Transactions sheet
    ws_raw = wb.create_sheet(title="Raw Transactions")
    _write_raw_sheet(ws_raw, conn, years)

    return wb


def export_excel_bytes(
    conn: sqlite3.Connection,
    years: list[int] | None = None,
) -> bytes:
    """Generate an Excel export and return as bytes (for browser streaming).

    Parameters
    ----------
    conn:
        An open sqlite3 connection.
    years:
        Passed through to build_workbook.  If None, all years are included.

    Returns
    -------
    Raw bytes of the .xlsx file (suitable for streaming to a browser).
    """
    import io

    wb = build_workbook(conn, years)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_to_file(
    conn: sqlite3.Connection,
    path: str | Path,
    years: list[int] | None = None,
    lock_timeout: float = 30.0,
) -> Path:
    """Build the workbook and save it atomically to *path* with mode 0o600.

    Uses a write-to-tmp-then-rename pattern so that the target file is never
    in a partially-written state (atomic on macOS/Linux).

    Acquires the sync lock before writing so that an in-progress sync cannot
    race with the export.  The lock is held for the duration of the write and
    rename, then released.

    Parameters
    ----------
    conn:
        An open sqlite3 connection.
    path:
        Destination file path.
    years:
        Passed through to build_workbook.
    lock_timeout:
        How long (in seconds) to wait for the sync lock before raising
        :class:`ExportBusy`.  Default is 30 seconds.

    Returns
    -------
    Path to the written file.

    Raises
    ------
    ExportBusy
        When the sync lock cannot be acquired within *lock_timeout* seconds.
    """
    path = Path(path)
    pid = os.getpid()
    tmp_path = path.with_suffix(path.suffix + f".tmp.{pid}")

    # Acquire the sync lock before touching any files so that a running sync
    # cannot observe a partial write.
    lock = acquire_sync_lock(timeout=lock_timeout)
    if lock is None:
        raise ExportBusy(
            f"Could not acquire sync lock within {lock_timeout}s — "
            "a sync may be in progress; try again shortly."
        )

    try:
        wb = build_workbook(conn, years)

        try:
            wb.save(str(tmp_path))
            os.chmod(tmp_path, 0o600)
            os.replace(str(tmp_path), str(path))
        except Exception:
            # Clean up temp file on failure; original file (if any) is untouched
            # because we only call os.replace on success.
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                pass
            raise
    finally:
        lock.close()

    return path
