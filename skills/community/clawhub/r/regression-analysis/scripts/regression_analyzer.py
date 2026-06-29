#!/usr/bin/env python3
"""
投流回归分析脚本 - 使用模板填充，保留格式
生成的Excel包含所有模型×所有时段的Sheet
"""
import numpy as np
import pandas as pd
import openpyxl
import statsmodels.api as sm
from scipy import stats as scipy_stats
from datetime import datetime, timedelta, date
import sys, os, argparse
from copy import copy
from openpyxl.cell.cell import MergedCell

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
DEFAULT_TEMPLATE = os.path.join(SKILL_DIR, "assets", "template.xlsx")

parser = argparse.ArgumentParser(description="投流回归分析 - 生成Excel结果")
parser.add_argument("source", help="原始数据文件路径")
parser.add_argument("-o", "--output", default="投流回归结果.xlsx")
parser.add_argument("-t", "--template", default=DEFAULT_TEMPLATE)
args = parser.parse_args()

TEMPLATE_PATH, OUTPUT_PATH, SRC_PATH = args.template, args.output, args.source

# ── 1. Load raw data ──────────────────────────────────────────────
print("读取原始数据...")
wb_in = openpyxl.load_workbook(SRC_PATH, data_only=True)
ws_in = wb_in["Sheet1"]
records = []
for row_idx in range(2, ws_in.max_row + 1):
    aq = ws_in.cell(row=row_idx, column=43).value
    ar = ws_in.cell(row=row_idx, column=44).value
    a_s = ws_in.cell(row=row_idx, column=45).value
    at = ws_in.cell(row=row_idx, column=46).value
    an = ws_in.cell(row=row_idx, column=40).value
    if any(v is None for v in [aq, ar, a_s, at, an]): continue
    try:
        if isinstance(an, datetime): dt = an.date()
        elif isinstance(an, date): dt = an
        elif isinstance(an, str): dt = datetime.strptime(an[:10], '%Y-%m-%d').date()
        elif isinstance(an, (int, float)): dt = (datetime(1899,12,30) + timedelta(days=int(an))).date()
        else: continue
        records.append({"date": dt, "aq": float(aq), "ar": float(ar), "as_val": float(a_s), "at": float(at)})
    except: continue
records.sort(key=lambda r: r["date"])
unique_dates = sorted(set(r["date"] for r in records))

daily_buckets = {d: {"aq":[],"ar":[],"as_val":[],"at":[]} for d in unique_dates}
for r in records:
    daily_buckets[r["date"]]["aq"].append(r["aq"])
    daily_buckets[r["date"]]["ar"].append(r["ar"])
    daily_buckets[r["date"]]["as_val"].append(r["as_val"])
    daily_buckets[r["date"]]["at"].append(r["at"])

daily_list = []
for d in unique_dates:
    b = daily_buckets[d]
    daily_list.append({"累积金额": np.sum(b["aq"]),
                       "累计短视频数量": np.sum(b["ar"]),
                       "累计短视频播放量": np.sum(b["as_val"]),
                       "累计总支付金额": np.sum(b["at"])})
df_daily = pd.DataFrame(daily_list)
n_daily = len(df_daily)
print(f"日度数据: {n_daily} 天 ({unique_dates[0]} ~ {unique_dates[-1]})")

periods_info = []
for i in range(0, n_daily, 30):
    batch = df_daily.iloc[i:i+30]
    if len(batch) < 30: continue
    periods_info.append({"label": f"P{i//30+1}", "df": batch,
                         "start": unique_dates[i], "end": unique_dates[min(i+29, n_daily-1)]})

# ── 2. Regression ─────────────────────────────────────────────────
def do_reg(df, x_vars, y_var):
    X = df[x_vars].copy()
    Xc = sm.add_constant(X)
    y = df[y_var]
    model = sm.OLS(y, Xc).fit()
    n = len(df); k = len(x_vars)
    ssr = model.ess; sse = model.ssr; sst = ssr + sse
    df_reg = k; df_res = n - k - 1; df_tot = n - 1
    msr = ssr/df_reg if df_reg>0 else 0
    mse = sse/df_res if df_res>0 else 0
    f_stat = msr/mse if mse>0 else 0
    f_pval = 1 - scipy_stats.f.cdf(f_stat, df_reg, df_res)
    ci = model.conf_int(alpha=0.05)
    return {"mr": np.sqrt(model.rsquared) if model.rsquared>=0 else 0,
            "rsq": model.rsquared, "adj_rsq": model.rsquared_adj,
            "se": np.sqrt(model.mse_resid), "n": n,
            "ssr": ssr, "sse": sse, "sst": sst,
            "df_reg": df_reg, "df_res": df_res, "df_tot": df_tot,
            "msr": msr, "mse": mse, "f": f_stat, "fpval": f_pval,
            "params": model.params, "stderr": model.bse,
            "tvals": model.tvalues, "pvals": model.pvalues,
            "ci_low": ci[0], "ci_high": ci[1]}

# ── 3. Template fill helpers ──────────────────────────────────────
def clear_cell(ws, r, c):
    cell = ws.cell(row=r, column=c)
    if not isinstance(cell, MergedCell): cell.value = None

def set_cell(ws, r, c, val):
    cell = ws.cell(row=r, column=c)
    if not isinstance(cell, MergedCell): cell.value = val

def fill_sheet(ws, reg, x_vars, y_label, date_start, date_end):
    k = len(x_vars)
    set_cell(ws, 1, 1, f"Y变量：{y_label}     X变量：{'、'.join(x_vars)}    样本量：{reg['n']}")
    set_cell(ws, 4, 2, reg["mr"]); set_cell(ws, 5, 2, reg["rsq"])
    set_cell(ws, 6, 2, reg["adj_rsq"]); set_cell(ws, 7, 2, reg["se"])
    set_cell(ws, 8, 2, reg["n"])

    anova = [("回归分析", reg["df_reg"], reg["ssr"], reg["msr"], reg["f"], reg["fpval"]),
             ("残差", reg["df_res"], reg["sse"], reg["mse"], None, None),
             ("总计", reg["df_tot"], reg["sst"], None, None, None)]
    for i, (label, *vals) in enumerate(anova):
        set_cell(ws, 12+i, 1, label)
        for j, v in enumerate(vals):
            if v is not None: set_cell(ws, 12+i, 2+j, v)

    var_order = ["const"] + x_vars
    for i, var_name in enumerate(var_order):
        r = 18 + i
        set_cell(ws, r, 1, "Intercept" if var_name == "const" else var_name)
        set_cell(ws, r, 2, reg["params"][var_name])
        set_cell(ws, r, 3, reg["stderr"][var_name])
        set_cell(ws, r, 4, reg["tvals"][var_name])
        set_cell(ws, r, 5, reg["pvals"][var_name])
        set_cell(ws, r, 6, reg["ci_low"][var_name])
        set_cell(ws, r, 7, reg["ci_high"][var_name])
    for r in range(18+len(var_order), 24):
        for c in range(1, 8): clear_cell(ws, r, c)

    if k <= 2: eq_row, period_row = 21, 22
    else: eq_row, period_row = 22, 23

    eq_parts = [f"{reg['params']['const']:.4f}"]
    for vn in x_vars:
        c = reg["params"][vn]
        eq_parts.append(f" + {c:.4f}{vn}" if c >= 0 else f" - {abs(c):.4f}{vn}")
    set_cell(ws, eq_row, 1, f"回归方程：Y = {''.join(eq_parts)}")
    set_cell(ws, period_row, 1, f"期间: {date_start} ~ {date_end}, {reg['n']}天")
    coeff_end = 18 + len(var_order)
    for r in range(coeff_end, 25):
        if r != eq_row and r != period_row: clear_cell(ws, r, 1)

# ── 4. Model definitions ──────────────────────────────────────────
# Each model: (sheet_name, x_vars, y_var, y_label)
ALL_MODELS = [
    # Y=累计总支付金额 (6 models)
    ("一元-累积金额→支付", ["累积金额"], "累计总支付金额"),
    ("一元-数量→支付", ["累计短视频数量"], "累计总支付金额"),
    ("一元-播放量→支付", ["累计短视频播放量"], "累计总支付金额"),
    ("二元-金额+数量→支付", ["累积金额", "累计短视频数量"], "累计总支付金额"),
    ("二元-金额+播放量→支付", ["累积金额", "累计短视频播放量"], "累计总支付金额"),
    ("三元-金额+播放量+数量→支付", ["累积金额", "累计短视频播放量", "累计短视频数量"], "累计总支付金额"),
    # New: X→Y cross-variable (3 models)
    ("一元-数量→播放量", ["累计短视频数量"], "累计短视频播放量"),
    ("一元-金额→数量", ["累积金额"], "累计短视频数量"),
    ("一元-金额→播放量", ["累积金额"], "累计短视频播放量"),
]

# Full sheet name → period suffix
FULL_TO_PERIOD_SUFFIX = {
    "一元-累积金额→支付": "一元",
    "一元-数量→支付": "数量→支付",
    "一元-播放量→支付": "播放量→支付",
    "二元-金额+数量→支付": "金额+数量",
    "二元-金额+播放量→支付": "金额+播放量",
    "三元-金额+播放量+数量→支付": "金额+播放量+数量",
    "一元-数量→播放量": "数量→播放量",
    "一元-金额→数量": "金额→数量",
    "一元-金额→播放量": "金额→播放量",
}

def build_sheet_map():
    m = {}
    for sn, x_vars, y_var in ALL_MODELS:
        suffix = FULL_TO_PERIOD_SUFFIX.get(sn, sn)
        # Full sample
        m[sn] = (x_vars, y_var, df_daily, unique_dates[0], unique_dates[-1])
        # Periods
        for p in periods_info:
            ws = f"{p['label']}-{suffix}"
            m[ws] = (x_vars, y_var, p["df"], p["start"], p["end"])
    return m

# ── 5. Open template, add new sheets, fill ────────────────────────
wb = openpyxl.load_workbook(TEMPLATE_PATH)

# Add missing sheets to template (copy from 一元-累积金额→支付)
src_ws = wb["一元-累积金额→支付"]
all_sheet_names = set(wb.sheetnames)
for sn in build_sheet_map():
    if sn not in all_sheet_names:
        ws_new = wb.copy_worksheet(src_ws)
        ws_new.title = sn
        all_sheet_names.add(sn)
        print(f"  新建Sheet: {sn}")

# Fill all sheets
sheet_map = build_sheet_map()
for sn, (x_vars, y_var, df, d_start, d_end) in sheet_map.items():
    if sn not in wb.sheetnames:
        print(f"  ⚠️ 跳过不存在: {sn}")
        continue
    reg = do_reg(df, x_vars, y_var)
    y_label = y_var
    fill_sheet(wb[sn], reg, x_vars, y_label, d_start, d_end)
    print(f"  ✅ {sn}: R²={reg['rsq']:.6f} (Y={y_var})")

# ── 6. Reorder sheets ─────────────────────────────────────────────
desired = []
full_order = [m[0] for m in ALL_MODELS]
period_suffixes = [FULL_TO_PERIOD_SUFFIX[sn] for sn in full_order]
for sn in full_order:
    desired.append(sn)
for p in ['P1','P2','P3','P4','P5']:
    for suffix in period_suffixes:
        desired.append(f"{p}-{suffix}")

for target_idx, sn in enumerate(desired):
    if sn not in wb.sheetnames: continue
    current_idx = wb.sheetnames.index(sn)
    if current_idx != target_idx:
        wb.move_sheet(sn, offset=target_idx - current_idx)

wb.save(OUTPUT_PATH)
print(f"\n✅ 保存: {OUTPUT_PATH}")
print(f"Sheet数: {len(wb.sheetnames)}")
