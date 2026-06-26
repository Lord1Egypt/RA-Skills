#!/usr/bin/env python3
"""
翔云发票 Excel 导出模块
支持 5 种标准模版：
  deduction  — 增值税发票勾选抵扣表
  transport  — 国内旅客运输服务抵扣表
  goods      — 增值税发票货物明细表
  ledger     — 增值税发票台账表
  booking    — 发票入账表
"""

from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import sys
    print("缺少依赖，请先执行：pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────
# 模版定义：每个模版包含（标题、列定义列表）
# 列定义：(列头显示名, 数据字段key, 列宽)
# ──────────────────────────────────────────────

TEMPLATES = {

    "deduction": {
        "title": "增值税发票勾选抵扣表",
        "columns": [
            ("是否勾选",       "isCheck",               10),
            ("发票代码",       "invoiceCode",            18),
            ("发票号码",       "invoiceNumber",          18),
            ("开票日期",       "billingDate",            14),
            ("税额",           "totalTax",               14),
            ("可抵扣税额",     "totalTax",               14),  # 同字段，两列
            ("销方名称",       "salesName",              24),
            ("销方税号",       "salesTaxNo",             20),
            ("金额",           "totalAmount",            14),
            ("用途",           "uses",                   14),
            ("发票类型",       "invoiceType",            18),
            ("管理状态",       "manageState",            12),
            ("数据来源",       "dataSource",             12),
            ("差额扣除标识",   "varianceDeductionFlag",  14),
        ],
    },

    "transport": {
        "title": "国内旅客运输服务抵扣表",
        "columns": [
            ("乘客姓名",       "name",              16),
            ("开票日期",       "billingDate",      14),
            ("发票类型",       "invoiceType",      18),
            ("票面金额",       "amountTax",        14),
            ("不含税金额",     "_taxableAmount",   14),
            ("税率",           "_taxRate",         10),
            ("抵扣进项税额",   "_deductionTax",   16),
        ],
    },

    "goods": {
        "title": "增值税发票货物明细表",
        "columns": [
            ("序号",                       "SERIAL",            6),
            ("凭证种类（必录）",           "voucherType",       18),
            ("数电票号码（全电票必录）",   "fullInvoiceNumber", 24),
            ("发票代码",                   "invoiceCode",       18),
            ("发票号码",                   "invoiceNumber",     18),
            ("开票日期（必录）",           "billingDate",       14),
            ("海关缴款书号码",             "ccpNumber",         18),
            ("代扣代缴完税凭证号码",       "wtcNumber",         20),
            ("销售方纳税人识别号",         "salesTaxNo",        20),
            ("销售方纳税人名称",           "salesName",         24),
            ("金额（必录）",               "totalAmount",       14),
            ("税额（必录）",               "totalTax",          14),
            ("入账用途（必录）",           "purpose",           16),
            ("入账时间",                   "postingDate",       14),
        ],
    },

    "ledger": {
        "title": "增值税发票台账表",
        "columns": [
            ("发票类型",           "invoiceType",          18),
            ("发票代码",           "invoiceCode",          18),
            ("发票号码",           "invoiceNumber",        18),
            ("开票日期",           "billingDate",          14),
            ("不含税金额",         "totalAmount",          14),
            ("校验码",             "checkCode",            18),
            ("购方名称",           "purchaserName",        24),
            ("购方税号",           "purchaserTaxNo",       20),
            ("购方地址电话",       "purchaserAddressPhone",22),
            ("购方开户行账户",     "purchaserBank",        22),
            ("销方名称",           "salesName",            24),
            ("销方税号",           "salesTaxNo",           20),
            ("销方地址及电话",     "salesAddressPhone",    22),
            ("销方开户行及账户",   "salesBankAndNo",       22),
            ("合计税额",           "totalTax",             14),
            ("票面金额",           "amountTax",            14),
            ("票面金额(大写)",     "amountTaxCN",          18),
            ("明细列表",           "invoiceLists",         30),
            ("备注",               "remarks",              20),
        ],
    },

    "booking": {
        "title": "发票入账表",
        "columns": [
            ("发票类型",                   "invoiceType",          18),
            ("发票代码",                   "invoiceCode",          18),
            ("发票号码",                   "invoiceNumber",        18),
            ("开票日期",                   "billingDate",          14),
            ("不含税金额",                 "totalAmount",          14),
            ("校验码",                     "checkCode",            18),
            ("购方名称",                   "purchaserName",        24),
            ("购方税号",                   "purchaserTaxNo",       20),
            ("购方地址电话",               "purchaserAddressPhone",22),
            ("购方开户行账户",             "purchaserBank",        22),
            ("销方名称",                   "salesName",            24),
            ("销方税号",                   "salesTaxNo",           20),
            ("销方地址及电话",             "salesAddressPhone",    22),
            ("销方开户行及账户",           "salesBankAndNo",       22),
            ("合计税额",                   "totalTax",             14),
            ("票面金额",                   "amountTax",            14),
            ("票面金额(大写)",             "amountTaxCN",          18),
            ("备注",                       "remarks",              20),
            ("身份证号码/组织机构代码",    "idCardNo",             20),
            ("车辆类型",                   "vehicleType",          14),
            ("厂牌型号",                   "brandModel",           14),
            ("发动机号",                   "engineNo",             16),
            ("车辆识别代号/车架号码",      "vehicleNo",            20),
            ("完税凭证号码",               "paymentVoucherNo",     18),
            ("车牌照号",                   "carNumber",            12),
            ("经营/拍卖单位名称",          "auctionName",          20),
            ("经营/拍卖单位纳税人识别号",  "auctionTaxNo",         22),
            ("二手车市场名称",             "usedCarName",          18),
            ("二手车市场纳税人识别号",     "usedCarTaxNo",         22),
        ],
    },
}

# ──────────────────────────────────────────────
# 样式常量
# ──────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
TITLE_FONT    = Font(name="微软雅黑", bold=True, size=12)
DATA_FONT     = Font(name="微软雅黑", size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

# ──────────────────────────────────────────────
# 辅助：从 fields 取值（处理 invoiceLists 序列化）
# ──────────────────────────────────────────────
def _get_cell_value(fields, key):
    val = fields.get(key, "")
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        try:
            import json
            return json.dumps(val, ensure_ascii=False)
        except Exception:
            return str(val)
    return str(val)


# ──────────────────────────────────────────────
# 国内旅客运输服务抵扣表预处理
# ──────────────────────────────────────────────

def _to_float(val):
    """安全转浮点数，失败返回 0.0"""
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def preprocess_transport_rows(rows):
    """
    对旅客运输服务抵扣表数据进行预处理：
      1. 补全：统一姓名字段（name/NAME/userCardNo）
      2. 过滤：10/83/92 类型必须明细含"客运服务"才保留
      3. 计算：22（船票）/26（客运汽车）税额 = 票面/1.03*0.03
      4. 计算：20（火车票）/27（航空行程单-改签）税额 = 票面/1.09*0.09
      5. 计算：27（正常机票）税额 = (票面-民航发展基金)/1.09*0.09
      6. 补全 _taxableAmount / _taxRate / _deductionTax 字段

    返回预处理后的 rows（深拷贝，不修改原数据）。
    """
    result = []
    for fields in rows:
        invoice_type = str(fields.get("invoiceTypeCode", "")).lstrip("0")

        # ── 姓名字段统一处理 ────────────────────────────────────────
        # 优先使用 name，兼容 NAME 和 userCardNo
        passenger_name = fields.get("name") or fields.get("NAME") or ""
        fields = dict(fields)  # 深拷贝
        fields["name"] = passenger_name  # 统一为小写 name
        amount_tax   = _to_float(fields.get("amountTax", 0))   # 票面金额（价税合计）
        total_tax   = _to_float(fields.get("totalTax", 0))    # OCR 识别的税额
        total_amount= _to_float(fields.get("totalAmount", 0)) # 不含税金额
        invoice_lists = fields.get("invoiceLists", [])

        # ── 10/83/92：过滤不含"客运服务"的行 ──────────────────
        if invoice_type in ("10", "83", "92"):
            has_passenger = False
            if isinstance(invoice_lists, list):
                for item in invoice_lists:
                    name = str(item.get("commodityName", ""))
                    if "客运服务" in name:
                        has_passenger = True
                        break
            if not has_passenger:
                continue  # 跳过

        # ── 基数：若无 amountTax，用 totalAmount+totalTax 推算 ──
        if amount_tax == 0 and total_amount > 0 and total_tax > 0:
            amount_tax = total_amount + total_tax

        # ── 船票（22）/ 客运汽车（26）：税额=票面/1.03*0.03 ─────
        if invoice_type in ("22", "26"):
            tax_rate       = 0.03
            deduction_tax  = round(amount_tax / 1.03 * tax_rate, 2)
            taxable_amount = round(amount_tax - deduction_tax, 2)
            proportion     = "3%"

        # ── 铁路（20）：税额=票面/1.09*0.09 ────────────────────
        elif invoice_type == "20":
            tax_rate       = 0.09
            deduction_tax  = round(amount_tax / 1.09 * tax_rate, 2)
            taxable_amount = round(amount_tax - deduction_tax, 2)
            proportion     = "9%"

        # ── 航空行程单（27）：判断改签 ─────────────────────────
        elif invoice_type == "27":
            # 票价（原票价）和民航发展基金（基建费）
            ticket_price = _to_float(fields.get("ticketPrice") or
                                     fields.get(" fare ") or
                                     fields.get("ticket_price") or 0)
            fuel_fee     = _to_float(fields.get("fuelFee") or
                                      fields.get("燃油附加费") or
                                      fields.get("fuel_fee") or 0)
            airport_fee  = _to_float(fields.get("airportFee") or
                                      fields.get("民航发展基金") or
                                      fields.get("airport_fee") or 0)
            # 尝试从 invoiceLists 找
            if ticket_price == 0 and isinstance(invoice_lists, list):
                for item in invoice_lists:
                    name = str(item.get("commodityName", ""))
                    if "票价" in name or "机票" in name:
                        ticket_price = _to_float(item.get("amount", 0))
                    if "燃油" in name or "附加费" in name:
                        fuel_fee = _to_float(item.get("amount", 0))
                    if "民航" in name or "基建" in name:
                        airport_fee = _to_float(item.get("amount", 0))

            # 改签判断：票价+燃油 > 票面金额 → 改签
            is_rebooking = (ticket_price + fuel_fee) > amount_tax if ticket_price > 0 else False

            if is_rebooking:
                # 改签票：与火车票相同算法
                tax_rate       = 0.09
                deduction_tax  = round(amount_tax / 1.09 * tax_rate, 2)
                taxable_amount = round(amount_tax - deduction_tax, 2)
                proportion     = "9%（改签）"
            else:
                # 正常机票：民航发展基金不计入基数
                tax_base       = amount_tax - airport_fee
                tax_rate       = 0.09
                deduction_tax  = round(tax_base / 1.09 * tax_rate, 2) if tax_base > 0 else 0.0
                taxable_amount = round(amount_tax - airport_fee - deduction_tax, 2) if tax_base > 0 else 0.0
                proportion     = "9%（正常）"

        # ── 其他类型（61/62 等）：用 OCR 识别的税额 ─────────────
        else:
            deduction_tax  = total_tax if total_tax > 0 else round(amount_tax / 1.09 * 0.09, 2)
            taxable_amount = round(amount_tax - deduction_tax, 2)
            tax_rate       = 0.09
            proportion     = "9%"

        # 浅拷贝字段，追加虚拟计算字段
        row = dict(fields)
        row["_taxableAmount"] = taxable_amount  # 不含税金额
        row["_taxRate"]       = tax_rate         # 税率
        row["_deductionTax"] = deduction_tax    # 抵扣进项税额
        row["_proportion"]    = proportion      # 抵扣比例（含标注）
        result.append(row)

    return result


# ──────────────────────────────────────────────
# 核心导出函数
# ──────────────────────────────────────────────

# ──────────────────────────────────────────────
# 核心导出函数
# ──────────────────────────────────────────────
def export_to_excel(rows, template_name, save_path):
    """
    rows          : list of dict，每个 dict 是一张发票的字段集合
    template_name : 模版名称（deduction/transport/goods/ledger/booking）
    save_path     : 输出文件路径（.xlsx）
    返回 (True, path) 或 (False, error_message)
    """
    if template_name not in TEMPLATES:
        return False, f"未知模版：{template_name}"

    # 旅客运输抵扣表需预处理（过滤/税额计算）
    if template_name == "transport":
        rows = preprocess_transport_rows(rows)
        # 若全部过滤掉，rows 为空仍生成空表头文件（按预期）

    tmpl   = TEMPLATES[template_name]
    title  = tmpl["title"]
    cols   = tmpl["columns"]  # [(header, field_key, width), ...]
    n_cols = len(cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    # ── 第1行：大标题（合并单元格）
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # ── 第2行：列标题
    border = _thin_border()
    for col_idx, (header, _, width) in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # ── 数据行（从第3行开始）
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx, fields in enumerate(rows, start=3):
        fill = alt_fill if row_idx % 2 == 1 else None
        for col_idx, (_, field_key, _) in enumerate(cols, start=1):
            value = _get_cell_value(fields, field_key)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.alignment = LEFT_ALIGN
            cell.border    = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # ── 冻结标题行
    ws.freeze_panes = "A3"

    try:
        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
        return True, save_path
    except Exception as e:
        return False, str(e)


# ──────────────────────────────────────────────
# 命令行独立使用（从 JSON 文件导出）
# ──────────────────────────────────────────────
if __name__ == "__main__":
    import argparse, json, sys

    parser = argparse.ArgumentParser(description="从 JSON 文件导出发票 Excel")
    parser.add_argument("--json",     required=True, help="识别/查验结果 JSON 文件路径")
    parser.add_argument("--template", default="ledger",
                        choices=list(TEMPLATES.keys()),
                        help="导出模版（默认：ledger）")
    parser.add_argument("--output",   default=None, help="输出 xlsx 路径")
    args = parser.parse_args()

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    # 支持单个 dict 或 list
    rows = data if isinstance(data, list) else [data]

    out = args.output or str(Path(args.json).with_suffix(f".{args.template}.xlsx"))
    ok, msg = export_to_excel(rows, args.template, out)
    if ok:
        print(f"[OK] 已导出：{out}")
    else:
        print(f"[FAIL] {msg}")
        sys.exit(1)
