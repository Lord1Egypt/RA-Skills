"""
Excel 报告生成脚本
用法: python generate_xlsx.py --output 输出路径
从 STDIN 读取 JSON 数据，生成 8 Sheet 结构化 Excel 报告
"""

import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime


# 样式定义
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1d28", end_color="1a1d28", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

UP_FONT = Font(name="Microsoft YaHei", size=10, color="E74C3C", bold=True)
DOWN_FONT = Font(name="Microsoft YaHei", size=10, color="27AE60", bold=True)

TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="4DA6FF")

THIN_BORDER = Border(
    left=Side(style="thin", color="2A2D3A"),
    right=Side(style="thin", color="2A2D3A"),
    top=Side(style="thin", color="2A2D3A"),
    bottom=Side(style="thin", color="2A2D3A"),
)

BG_FILL = PatternFill(start_color="0F1117", end_color="0F1117", fill_type="solid")
ALT_FILL = PatternFill(start_color="1A1D28", end_color="1A1D28", fill_type="solid")


def style_header(ws, row, max_col):
    """设置表头样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(cell, is_up=None):
    """设置数据单元格样式"""
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGNMENT
    cell.border = THIN_BORDER
    if is_up is True:
        cell.font = UP_FONT
    elif is_up is False:
        cell.font = DOWN_FONT


def set_col_widths(ws, widths):
    """设置列宽"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_sheet_macro_overview(wb, data):
    """Sheet 1: 宏观概览"""
    ws = wb.active
    ws.title = "宏观概览"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = f"大类资产配置建议 — 宏观概览"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"报告日期: {data.get('report_date', datetime.now().strftime('%Y-%m-%d'))} | 数据源: AKShare (stats.gov.cn)"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="8B8FA3")
    ws["A2"].alignment = Alignment(horizontal="center")

    # 中国宏观
    ws["A4"] = "中国宏观指标"
    ws["A4"].font = Font(name="Microsoft YaHei", bold=True, size=12, color="F0B90B")

    headers_cn = ["指标", "最新值", "时间", "趋势", "说明"]
    for i, h in enumerate(headers_cn, 1):
        ws.cell(row=5, column=i, value=h)
    style_header(ws, 5, len(headers_cn))

    cn_data = data.get("china_macro", [])
    for r, row_data in enumerate(cn_data, 6):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:  # 趋势列
                is_up = "↑" in str(val) or "上行" in str(val) or "加速" in str(val)
                is_down = "↓" in str(val) or "下行" in str(val) or "放缓" in str(val)
                style_data_cell(cell, is_up=True if is_up else (False if is_down else None))
            else:
                style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c != 5 else DATA_ALIGNMENT

    # 全球宏观
    global_start = 6 + len(cn_data) + 1
    ws.cell(row=global_start, column=1, value="全球宏观指标").font = Font(name="Microsoft YaHei", bold=True, size=12, color="F0B90B")

    headers_gl = ["指标", "最新值", "说明"]
    for i, h in enumerate(headers_gl, 1):
        ws.cell(row=global_start + 1, column=i, value=h)
    style_header(ws, global_start + 1, len(headers_gl))

    gl_data = data.get("global_macro", [])
    for r, row_data in enumerate(gl_data, global_start + 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)

    set_col_widths(ws, [22, 16, 14, 8, 50, 30])


def create_sheet_industry(wb, data):
    """Sheet 2: 行业分析"""
    ws = wb.create_sheet("行业分析")

    ws.merge_cells("A1:G1")
    ws["A1"] = "申万一级行业分析"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["行业名称", "近20日涨跌幅", "资金评分", "政策评分", "综合评分", "趋势", "产业链简析"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    ind_data = data.get("industries", [])
    for r, row_data in enumerate(ind_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:  # 涨跌幅
                is_up = isinstance(val, (int, float)) and val > 0
                style_data_cell(cell, is_up=is_up if val != 0 else None)
            elif c == 6:  # 趋势
                is_up = "上升" in str(val)
                style_data_cell(cell, is_up=is_up if "上升" in str(val) else ("下降" in str(val)))
            else:
                style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 6 else DATA_ALIGNMENT

    set_col_widths(ws, [16, 14, 10, 10, 10, 8, 60])


def create_sheet_strategy(wb, data):
    """Sheet 3: 战略配置"""
    ws = wb.create_sheet("战略配置")

    ws.merge_cells("A1:G1")
    ws["A1"] = "大类资产战略配置"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["资产类别", "建议权重", "较上期", "趋势", "置信度", "核心逻辑"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    strategy_data = data.get("strategy", [])
    for r, row_data in enumerate(strategy_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 3:
                is_up = "↑" in str(val)
                style_data_cell(cell, is_up=is_up)
            else:
                style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 5 else DATA_ALIGNMENT

    set_col_widths(ws, [16, 12, 10, 8, 10, 70])


def create_sheet_bond(wb, data):
    """Sheet 4: 债券配置"""
    ws = wb.create_sheet("债券配置")

    ws.merge_cells("A1:E1")
    ws["A1"] = "债券配置建议"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["行业/类型", "品种", "代码", "操作建议", "理由"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    bond_data = data.get("bonds", [])
    for r, row_data in enumerate(bond_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 3 else DATA_ALIGNMENT

    set_col_widths(ws, [14, 16, 14, 12, 70])


def create_sheet_stock(wb, data):
    """Sheet 5: 股票基金"""
    ws = wb.create_sheet("股票基金")

    ws.merge_cells("A1:G1")
    ws["A1"] = "股票及基金配置建议"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["市场", "行业", "代码", "名称", "类型", "操作建议", "理由"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    stock_data = data.get("stocks", [])
    for r, row_data in enumerate(stock_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 5 else DATA_ALIGNMENT

    set_col_widths(ws, [10, 14, 14, 20, 12, 12, 70])


def create_sheet_commodity(wb, data):
    """Sheet 6: 商品配置"""
    ws = wb.create_sheet("商品配置")

    ws.merge_cells("A1:F1")
    ws["A1"] = "商品配置建议"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["行业", "商品品种", "相关基金/ETF", "方向", "逻辑"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    comm_data = data.get("commodities", [])
    for r, row_data in enumerate(comm_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:
                is_up = "看多" in str(val)
                style_data_cell(cell, is_up=is_up)
            else:
                style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 4 else DATA_ALIGNMENT

    set_col_widths(ws, [14, 16, 24, 8, 70])


def create_sheet_derivatives(wb, data):
    """Sheet 7: 期货期权"""
    ws = wb.create_sheet("期货期权")

    ws.merge_cells("A1:H1")
    ws["A1"] = "期货期权策略建议"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["类别", "品种", "方向", "策略", "合约/行权价", "建议仓位", "止损参考", "逻辑"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    deriv_data = data.get("derivatives", [])
    for r, row_data in enumerate(deriv_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 3:
                is_up = "做多" in str(val) or "买看涨" in str(val)
                is_down = "做空" in str(val) or "买看跌" in str(val)
                style_data_cell(cell, is_up=is_up if is_up else (False if is_down else None))
            else:
                style_data_cell(cell)
            cell.alignment = CENTER_ALIGNMENT if c <= 6 else DATA_ALIGNMENT

    set_col_widths(ws, [10, 14, 8, 14, 16, 10, 12, 60])


def create_sheet_risk(wb, data):
    """Sheet 8: 风险提示"""
    ws = wb.create_sheet("风险提示")

    ws.merge_cells("A1:B1")
    ws["A1"] = "风险提示与免责声明"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["风险类别", "内容"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers))

    risk_data = data.get("risks", [])
    for r, row_data in enumerate(risk_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)
        ws.row_dimensions[r].height = 40

    # 免责声明
    notice_row = 3 + len(risk_data) + 1
    ws.merge_cells(start_row=notice_row, start_column=1, end_row=notice_row, end_column=2)
    ws.cell(row=notice_row, column=1, value="免责声明").font = Font(name="Microsoft YaHei", bold=True, size=11, color="E74C3C")

    disclaimer = ws.merge_cells(start_row=notice_row + 1, start_column=1, end_row=notice_row + 1, end_column=2)
    ws.cell(row=notice_row + 1, column=1, value=(
        "本报告由 AI 基于公开市场数据和宏观经济指标自动生成，仅供投资研究参考，不构成任何投资建议。"
        "宏观经济数据可能存在滞后性，行业趋势分析基于历史数据，过去表现不代表未来收益。"
        "期货期权属于杠杆交易工具，存在较大亏损风险。投资者应审慎决策，自行承担投资风险。"
    ))
    ws.cell(row=notice_row + 1, column=1).font = Font(name="Microsoft YaHei", size=9, color="8B8FA3")
    ws.cell(row=notice_row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[notice_row + 1].height = 60

    set_col_widths(ws, [16, 100])


def generate_xlsx(data, output_path):
    """生成完整的 Excel 报告"""
    wb = openpyxl.Workbook()

    create_sheet_macro_overview(wb, data)
    create_sheet_industry(wb, data)
    create_sheet_strategy(wb, data)
    create_sheet_bond(wb, data)
    create_sheet_stock(wb, data)
    create_sheet_commodity(wb, data)
    create_sheet_derivatives(wb, data)
    create_sheet_risk(wb, data)

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="asset_config_report.xlsx")
    parser.add_argument("--data", help="JSON data file path")
    args = parser.parse_args()

    if args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        # 从 STDIN 读取
        data = json.loads(sys.stdin.read())

    generate_xlsx(data, args.output)
