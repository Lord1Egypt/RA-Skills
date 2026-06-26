"""
tests/test_excel_export.py — Tests for server/excel_export.py and the
export_excel MCP tool in server/main.py.

Fixtures
--------
- tmp_path: pytest built-in temp directory
- monkeypatch: redirect server.paths.DB_PATH, APP_DIR, EXPORTS_DIR
- Seeded DB: Personal ledger with Groceries (expense) + Salary (income),
  4 transactions in Jan/Feb 2026.
"""

from __future__ import annotations

import sqlite3
import stat
import uuid
from pathlib import Path

import openpyxl
import pytest

import server.paths
from server.db import get_db, init_db
from server.excel_export import build_workbook, export_to_file

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def patched_paths(tmp_path: Path, monkeypatch):
    """Redirect all server.paths path constants to tmp_path."""
    app_dir = tmp_path / ".friday-bp"
    exports_dir = app_dir / "exports"
    db_path = app_dir / "data.db"

    monkeypatch.setattr(server.paths, "APP_DIR", app_dir)
    monkeypatch.setattr(server.paths, "EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(server.paths, "DB_PATH", db_path)

    return {
        "app_dir": app_dir,
        "exports_dir": exports_dir,
        "db_path": db_path,
    }


@pytest.fixture()
def seeded_db(patched_paths) -> sqlite3.Connection:
    """Initialise DB and seed with minimal data for testing.

    Ledger: Personal
      Line items:
        - Groceries (expense)
        - Salary (income)
      Transactions (4 total):
        - 2026-01-05, Superstore, $100 → Groceries
        - 2026-01-25, Employer,   $2000 → Salary
        - 2026-02-03, Superstore, $120 → Groceries
        - 2026-02-15, Employer,   $2000 → Salary
    """
    db_path = patched_paths["db_path"]
    db_path.parent.mkdir(parents=True, exist_ok=True)
    init_db(db_path)
    conn = get_db(db_path)

    ledger_id = str(uuid.uuid4())
    groceries_id = str(uuid.uuid4())
    salary_id = str(uuid.uuid4())

    # Ledger
    conn.execute("INSERT INTO ledgers (id, name) VALUES (?, ?)", (ledger_id, "Personal"))

    # Line items
    conn.execute(
        "INSERT INTO line_items (id, ledger_id, name, item_type) VALUES (?, ?, ?, ?)",
        (groceries_id, ledger_id, "Groceries", "expense"),
    )
    conn.execute(
        "INSERT INTO line_items (id, ledger_id, name, item_type) VALUES (?, ?, ?, ?)",
        (salary_id, ledger_id, "Salary", "income"),
    )

    # Bank connection + account (required by FK chain)
    conn_id = str(uuid.uuid4())
    acct_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO bank_connections (id, plaid_item_id, plaid_access_token_encrypted) VALUES (?, ?, ?)",
        (conn_id, "plaid-item-1", "enc-token"),
    )
    conn.execute(
        "INSERT INTO bank_accounts (id, connection_id, plaid_account_id, name) VALUES (?, ?, ?, ?)",
        (acct_id, conn_id, "plaid-acct-1", "Chequing"),
    )

    # Transactions + entries
    txns = [
        (str(uuid.uuid4()), "2026-01-05", "Superstore", 100.0, groceries_id, ledger_id),
        (str(uuid.uuid4()), "2026-01-25", "Employer", 2000.0, salary_id, ledger_id),
        (str(uuid.uuid4()), "2026-02-03", "Superstore", 120.0, groceries_id, ledger_id),
        (str(uuid.uuid4()), "2026-02-15", "Employer", 2000.0, salary_id, ledger_id),
    ]

    for txn_id, date, merchant, amount, li_id, led_id in txns:
        conn.execute(
            "INSERT INTO transactions (id, bank_account_id, plaid_transaction_id, date, merchant, amount) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (txn_id, acct_id, f"plaid-{txn_id}", date, merchant, amount),
        )
        entry_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO transaction_entries (id, transaction_id, ledger_id, line_item_id, amount, source) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (entry_id, txn_id, led_id, li_id, amount, "manual"),
        )

    conn.commit()
    yield conn
    conn.close()


# ---------------------------------------------------------------------------
# build_workbook tests
# ---------------------------------------------------------------------------


def test_build_workbook_sheets(seeded_db):
    """Workbook has Personal, Summary, and Raw Transactions sheets."""
    wb = build_workbook(seeded_db)
    assert "Personal" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Raw Transactions" in wb.sheetnames


def test_personal_sheet_has_month_headers(seeded_db):
    """Personal sheet column 2 header is 'Jan'."""
    wb = build_workbook(seeded_db)
    ws = wb["Personal"]

    # Find the header row (the one containing "Jan")
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if "Jan" in row:
            header_row = row
            break

    assert header_row is not None, "Could not find a header row with 'Jan'"
    assert "Feb" in header_row
    assert "YTD" in header_row


def test_personal_sheet_has_net_row(seeded_db):
    """Personal sheet contains a 'Net' row."""
    wb = build_workbook(seeded_db)
    ws = wb["Personal"]

    net_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Net" in net_values, "Personal sheet is missing a 'Net' row"


def test_personal_sheet_income_row(seeded_db):
    """Personal sheet Salary row has correct Jan total (2000.0)."""
    wb = build_workbook(seeded_db)
    ws = wb["Personal"]

    # Find the Salary row
    salary_row = None
    for row in ws.iter_rows():
        if row[0].value == "Salary":
            salary_row = row
            break

    assert salary_row is not None, "Salary row not found"
    # Column B = Jan (index 1 in 0-based row tuple)
    assert salary_row[1].value == 2000.0


def test_personal_sheet_expense_row(seeded_db):
    """Personal sheet Groceries row has correct Jan total (100.0)."""
    wb = build_workbook(seeded_db)
    ws = wb["Personal"]

    groceries_row = None
    for row in ws.iter_rows():
        if row[0].value == "Groceries":
            groceries_row = row
            break

    assert groceries_row is not None, "Groceries row not found"
    assert groceries_row[1].value == 100.0


def test_raw_transactions_sheet_row_count(seeded_db):
    """Raw Transactions sheet has 4 data rows (1 header + 4 data = 5 rows)."""
    wb = build_workbook(seeded_db)
    ws = wb["Raw Transactions"]

    # Count non-empty rows (skip header)
    data_rows = [
        r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)
    ]
    assert len(data_rows) == 4


def test_summary_sheet_has_personal(seeded_db):
    """Summary sheet contains 'Personal' ledger entries."""
    wb = build_workbook(seeded_db)
    ws = wb["Summary"]

    ledger_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "Personal" in ledger_names


# ---------------------------------------------------------------------------
# export_to_file tests
# ---------------------------------------------------------------------------


def test_export_to_file_creates_file(seeded_db, tmp_path):
    """export_to_file writes a file at the given path."""
    dest = tmp_path / "test_export.xlsx"
    result = export_to_file(seeded_db, dest)
    assert result == dest
    assert dest.exists()


def test_export_to_file_mode_0600(seeded_db, tmp_path):
    """Exported file has mode 0o600."""
    dest = tmp_path / "test_export.xlsx"
    export_to_file(seeded_db, dest)
    file_mode = stat.S_IMODE(dest.stat().st_mode)
    assert file_mode == 0o600


def test_export_to_file_valid_xlsx(seeded_db, tmp_path):
    """openpyxl can open the exported file."""
    dest = tmp_path / "test_export.xlsx"
    export_to_file(seeded_db, dest)
    wb = openpyxl.load_workbook(str(dest))
    assert "Personal" in wb.sheetnames


# ---------------------------------------------------------------------------
# export_excel MCP tool tests
# ---------------------------------------------------------------------------


def test_export_excel_tool(patched_paths, seeded_db):
    """export_excel() returns status=ok with a valid path and size > 0."""
    import server.main as main_mod

    # Patch the DB path used by the tool
    result = main_mod.export_excel()

    assert result["status"] == "ok"
    assert "path" in result
    assert result["size_bytes"] > 0

    out_path = Path(result["path"])
    assert out_path.exists()
    assert out_path.parent == patched_paths["exports_dir"]


def test_export_excel_tool_with_years(patched_paths, seeded_db):
    """export_excel(years=[2026]) also returns status=ok."""
    import server.main as main_mod

    result = main_mod.export_excel(years=[2026])
    assert result["status"] == "ok"
    assert result["size_bytes"] > 0
