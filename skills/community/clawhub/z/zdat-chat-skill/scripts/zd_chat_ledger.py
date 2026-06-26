"""
ZDAT 线索台账查看脚本 v1.0
读取 clue_ledger.xlsx 展示高意向线索
"""
import sys, os, datetime
from pathlib import Path
from openpyxl import load_workbook

WORKDIR = Path(os.getenv("WORKDIR", os.path.expanduser("~/.molili/workspaces/default")))
LEDGER_PATH = WORKDIR / "clue_ledger.xlsx"

def show_ledger(limit=20):
    if not LEDGER_PATH.exists():
        print("📋 暂无线索台账")
        return
    
    wb = load_workbook(str(LEDGER_PATH))
    ws = wb.active
    
    print(f"\n📋 ZDAT 线索台账 — {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"   最近 {limit} 条:\n")
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= limit:
            break
        print(f"   {' | '.join(str(c) for c in row if c)}")

if __name__ == "__main__":
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 20
    show_ledger(limit)
