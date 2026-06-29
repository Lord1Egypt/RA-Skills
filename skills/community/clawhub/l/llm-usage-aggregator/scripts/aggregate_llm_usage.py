#!/usr/bin/env python3
"""
LLM Usage Aggregator - 汇总LLM使用流水数据
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os
import json


def load_pricing_config(config_path=None):
    """加载定价配置"""
    if config_path is None:
        config_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'references', 'pricing_config.json'
        )
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


class PricingError:
    """记录计价异常"""
    def __init__(self, model, gen_type, reason, row_index=None):
        self.model = model
        self.gen_type = gen_type
        self.reason = reason
        self.row_index = row_index

    def __repr__(self):
        loc = f" (行{self.row_index})" if self.row_index else ""
        return f"[{self.model}|gen_type={self.gen_type or '空'}] {self.reason}{loc}"


def normalize_model_name(model, pricing):
    """将原始模型名归一化为用于计价和汇总的标准简称。"""
    if pd.isna(model) or model == '':
        return model
    model_aliases = pricing.get('model_aliases', {})
    return model_aliases.get(model, model)


def normalize_model_column(df, pricing):
    """保留原始模型名，并将 llm_model 归一化为标准简称。"""
    if 'llm_model' not in df.columns:
        return df
    df = df.copy()
    df['llm_model_raw'] = df['llm_model']
    df['llm_model'] = df['llm_model'].apply(lambda model: normalize_model_name(model, pricing))
    return df


def collect_unmatched_models(df, pricing_models):
    """汇总归一化后仍未匹配到定价配置的模型及记录数。"""
    unmatched = {}
    for model in df['llm_model'].unique():
        if pd.notna(model) and model != '' and model not in pricing_models:
            count = len(df[df['llm_model'] == model])
            unmatched[model] = count
    return unmatched


def resolve_model_type(config, gen_type):
    """解析模型类型，优先使用配置中的 model_type，其次回退到 generation_type。"""
    model_type = config.get('model_type', '')
    if model_type:
        return model_type
    if gen_type == 'image':
        return '图片'
    if gen_type == 'video':
        return '视频'
    return '文本'


def calc_row_cost(row, pricing_models, errors, row_index=None):
    """计算单行数据的成本，返回 (cost_usd, pricing_detail_dict)"""
    model = row['llm_model']
    gen_type = row.get('generation_type', '')

    # 默认明细
    detail = {
        'pricing_type': '',
        'model_type': '',
        'unit_price_prompt': None,
        'unit_price_completion': None,
        'unit_price_request': None
    }

    # 模型名为空
    if pd.isna(model) or model == '':
        errors.append(PricingError('NaN', gen_type, 'llm_model为空', row_index))
        return 0.0, detail

    config = pricing_models.get(model)
    if config is None:
        errors.append(PricingError(model, gen_type, '模型未在pricing_config.json中配置', row_index))
        return 0.0, detail

    detail['pricing_type'] = config.get('pricing_type', '')
    detail['model_type'] = resolve_model_type(config, gen_type)

    # 文本类型：始终按输入/输出 token 单价计价
    if detail['model_type'] == '文本':
        if config.get('pricing_type') != 'per_million_tokens':
            errors.append(PricingError(
                model, gen_type,
                f"文本类型需要per_million_tokens计价，但配置为{config.get('pricing_type')}",
                row_index
            ))
            return 0.0, detail
        if 'prompt_price' not in config:
            errors.append(PricingError(model, gen_type, '文本计价缺prompt_price', row_index))
            return 0.0, detail
        if 'completion_price' not in config:
            errors.append(PricingError(model, gen_type, '文本计价缺completion_price', row_index))
            return 0.0, detail
        prompt = row.get('prompt_tokens', 0) or 0
        completion = row.get('completion_tokens', 0) or 0
        prompt_price = config['prompt_price']
        completion_price = config['completion_price']
        detail['unit_price_prompt'] = prompt_price
        detail['unit_price_completion'] = completion_price
        prompt_cost = prompt / 1_000_000 * prompt_price
        completion_cost = completion / 1_000_000 * completion_price
        return prompt_cost + completion_cost, detail

    # 图片类型：按张计价，单价 × generated_image_count
    if detail['model_type'] == '图片':
        if config.get('pricing_type') != 'per_request':
            errors.append(PricingError(model, gen_type,
                f"图片类型需要per_request计价，但配置为{config.get('pricing_type')}", row_index))
            return 0.0, detail
        img_count = row.get('generated_image_count', float('nan'))
        if pd.isna(img_count) or img_count == 0:
            img_count = 1
        unit_price = config['price_per_request']
        detail['unit_price_request'] = unit_price
        return unit_price * img_count, detail

    # 视频类型：按秒计价，判断有声/无声，单价 × duration_seconds
    if detail['model_type'] == '视频' or gen_type == 'video':
        # 视频生成场景：generation_type=video 时强制按秒计价，不受 model pricing_type 影响
        has_audio = row.get('generate_audio', False)
        if pd.isna(has_audio):
            has_audio = False
        duration = row.get('duration_seconds', float('nan'))
        if pd.isna(duration) or duration == 0:
            duration = 5

        # 优先用 audio_price（有声/无声分档），否则用 price_per_request
        if has_audio and 'audio_price' in config:
            unit_price = config['audio_price'].get('with_audio')
            if unit_price is None:
                unit_price = config['price_per_request']
        elif not has_audio and 'audio_price' in config:
            unit_price = config['audio_price'].get('without_audio')
            if unit_price is None:
                unit_price = config['price_per_request']
        else:
            unit_price = config.get('price_per_request', config.get('prompt_price', 0))

        detail['unit_price_request'] = unit_price
        return unit_price * duration, detail

    errors.append(PricingError(model, gen_type,
        f"未识别的model_type: {detail['model_type'] or '空'}", row_index))
    return 0.0, detail


def aggregate_llm_usage(csv_path, output_path=None, pricing_config_path=None):
    """
    汇总LLM使用流水CSV数据

    Args:
        csv_path: CSV文件路径
        output_path: 输出Excel路径（可选，默认同目录）
        pricing_config_path: 定价配置JSON路径（可选）

    Returns:
        输出文件路径
    """
    # 自动识别文件类型
    if csv_path.endswith('.xlsx') or csv_path.endswith('.xls'):
        df = pd.read_excel(csv_path)
    else:
        df = pd.read_csv(csv_path)

    # 仅保留 status 为 success 的记录
    if 'status' in df.columns:
        before = len(df)
        df = df[df['status'] == 'success']
        print(f"  [过滤] status=success: 保留 {len(df)}/{before} 条记录")
    else:
        print("  [警告] 数据中无 status 列，未做过滤")

    pricing = load_pricing_config(pricing_config_path)
    df = normalize_model_column(df, pricing)

    # 创建用户标识：优先 email，没有则用 phone
    # 统一用户标识：优先 email > phone > user_id
    if 'email' in df.columns and 'phone' in df.columns:
        df['user_id_combined'] = df.apply(
            lambda x: x['email'] if pd.notna(x['email']) and x['email'] != '' else x['phone'],
            axis=1
        )
    elif 'email' in df.columns:
        df['user_id_combined'] = df['email'].fillna(df.get('phone', pd.Series([''] * len(df))))
        df['user_id_combined'] = df['user_id_combined'].replace('', df.get('phone', pd.Series([''] * len(df))))
        df['user_id_combined'] = df.apply(
            lambda x: x['email'] if pd.notna(x['email']) and x['email'] != '' else (x.get('phone', '') if 'phone' in x.index else ''),
            axis=1
        )
    elif 'phone' in df.columns:
        df['user_id_combined'] = df['phone'].fillna(df.get('email', pd.Series([''] * len(df))))
    else:
        df['user_id_combined'] = df['user_id'].astype(str)

    # 判断内部/外部用户
    def is_internal(user_id):
        """根据用户标识判断是否为内部用户。"""
        if pd.isna(user_id) or user_id == '':
            return '外部用户'
        user_id_lower = str(user_id).lower()
        if 'footprint' in user_id_lower or 'maybe' in user_id_lower or 'fastest' in user_id_lower:
            return '内部用户'
        return '外部用户'

    df['user_type'] = df['user_id_combined'].apply(is_internal)

    # 汇总字段
    agg_fields = ['prompt_tokens', 'completion_tokens', 'generated_image_count', 'duration_seconds']

    # 处理数值字段中的空值
    # prompt_tokens/completion_tokens: 空值填0（token计价需要）
    # generated_image_count/duration_seconds: 保留NaN，在计价时按类型给默认值
    for field in ['prompt_tokens', 'completion_tokens']:
        if field in df.columns:
            df[field] = df[field].fillna(0)
        else:
            df[field] = 0
    for field in ['generated_image_count', 'duration_seconds']:
        if field not in df.columns:
            df[field] = float('nan')

    # 加载定价配置并计算每行成本
    pricing_models = pricing['pricing_models']

    # 预检：汇总所有未配置定价的模型及其影响行数
    unmatched = collect_unmatched_models(df, pricing_models)
    errors = []
    pricing_details = []

    def calc_and_store(row):
        """计算单行成本并同步收集计价明细。"""
        cost, detail = calc_row_cost(row, pricing_models, errors, row_index=row.name + 1)
        pricing_details.append(detail)
        return cost

    df['cost_usd'] = df.apply(calc_and_store, axis=1)

    # 将定价明细写入df
    detail_df = pd.DataFrame(pricing_details, index=df.index)
    df['pricing_type'] = detail_df['pricing_type']
    df['model_type'] = detail_df['model_type']
    df['unit_price_prompt'] = detail_df['unit_price_prompt']
    df['unit_price_completion'] = detail_df['unit_price_completion']
    df['unit_price_request'] = detail_df['unit_price_request']

    # 输出校验报告
    if unmatched:
        print("\n⚠ 未配置定价的模型（成本按0计算）：")
        for model, count in sorted(unmatched.items(), key=lambda x: -x[1]):
            print(f"  - {model}: {count} 条记录")

    if errors:
        # 按原因分组统计
        from collections import Counter
        reason_counts = Counter(f"{e.model}|{e.gen_type or '空'}|{e.reason}" for e in errors)
        print(f"\n⚠ 计价异常（共{len(errors)}条，成本按0计算）：")
        for key, count in reason_counts.most_common(20):
            parts = key.split('|')
            print(f"  - [{parts[0]}|gen_type={parts[1]}] {parts[2]}: {count}条")

    if not unmatched and not errors:
        print("✓ 定价校验通过，所有模型均匹配且计价正常")

    # 校验：文本类型的行应按 token 计价，成本为0的属于异常
    # embedding 模型无 completion token，成本为0属正常，不报警
    text_rows = df[df['model_type'] == '文本']
    zero_cost_token = text_rows[text_rows['cost_usd'] == 0]
    # 过滤掉 embedding 模型（cost=0 正常）
    zero_cost_token = zero_cost_token[~zero_cost_token['llm_model'].str.contains('embedding', na=False)]
    if len(zero_cost_token) > 0:
        zero_by_model = zero_cost_token.groupby('llm_model').size().sort_values(ascending=False)
        print(f"\n⚠ 文本模型token计价行成本为0（共{len(zero_cost_token)}条，应逐行有计价）：")
        for model, count in zero_by_model.items():
            print(f"  - {model}: {count}条")

    # 计价完成后，将 generated_image_count/duration_seconds 的NaN填0用于汇总展示
    df['generated_image_count'] = df['generated_image_count'].fillna(0)
    df['duration_seconds'] = df['duration_seconds'].fillna(0)

    # 创建Excel：单Sheet，4个区块由上至下
    wb = Workbook()
    ws = wb.active
    ws.title = '汇总报表'

    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cost_header_fill = PatternFill('solid', fgColor='548235')
    section_font = Font(bold=True, size=13, color='1F4E79')
    section_fill = PatternFill('solid', fgColor='D6E4F0')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    num_cols = len(agg_fields) + 2  # group_col + agg_fields + cost + count

    def write_section(ws, start_row, section_title, group_col, df, agg_fields, cost_field='cost_usd'):
        """写入一个汇总区块，返回下一个区块的起始行"""
        all_agg = agg_fields + [cost_field]
        grouped = df.groupby(group_col)[all_agg].sum().reset_index()
        cost_col_idx = len(agg_fields) + 2

        # 区块标题行（合并单元格）
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        cell = ws.cell(row=start_row, column=1, value=section_title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, num_cols + 1):
            ws.cell(row=start_row, column=c).border = thin_border
            ws.cell(row=start_row, column=c).fill = section_fill

        # 表头行
        header_row = start_row + 1
        col_labels = {
            'user_type': '用户类型', 'llm_provider': '模型平台', 'llm_model': '模型', 'user_id_combined': '用户'
        }
        headers = [col_labels.get(group_col, group_col)] + agg_fields + ['成本(USD)', '记录数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = cost_header_fill if header == '成本(USD)' else header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, row_data in enumerate(grouped.itertuples(), header_row + 1):
            ws.cell(row=row_idx, column=1, value=row_data[1]).border = thin_border
            for col_idx, field in enumerate(agg_fields, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=getattr(row_data, field))
                cell.border = thin_border
            cost_val = getattr(row_data, cost_field)
            cell = ws.cell(row=row_idx, column=cost_col_idx, value=round(cost_val, 4))
            cell.number_format = '#,##0.0000'
            cell.border = thin_border
            count = df[df[group_col] == row_data[1]].shape[0]
            cell = ws.cell(row=row_idx, column=cost_col_idx + 1, value=count)
            cell.border = thin_border

        # 合计行
        total_row = header_row + 1 + len(grouped)
        ws.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = thin_border
        for col_idx, field in enumerate(agg_fields, 2):
            total = grouped[field].sum()
            cell = ws.cell(row=total_row, column=col_idx, value=total)
            cell.font = Font(bold=True)
            cell.border = thin_border
        total_cost = grouped[cost_field].sum()
        cell = ws.cell(row=total_row, column=cost_col_idx, value=round(total_cost, 4))
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.0000'
        cell.border = thin_border
        cell = ws.cell(row=total_row, column=cost_col_idx + 1, value=len(df))
        cell.font = Font(bold=True)
        cell.border = thin_border

        return total_row + 2  # 空一行后接下一个区块

    # 由上至下依次写入4个区块
    cur_row = 1
    cur_row = write_section(ws, cur_row, '按内外部用户', 'user_type', df, agg_fields)
    if 'llm_provider' in df.columns:
        cur_row = write_section(ws, cur_row, '按模型平台', 'llm_provider', df, agg_fields)
    if 'llm_model' in df.columns:
        cur_row = write_section(ws, cur_row, '按模型', 'llm_model', df, agg_fields)
    cur_row = write_section(ws, cur_row, '按用户', 'user_id_combined', df, agg_fields)

    # 设置列宽
    ws.column_dimensions['A'].width = 45
    for col in range(2, num_cols + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    # === 第二个Sheet：计价明细 ===
    ws2 = wb.create_sheet('计价明细')

    detail_columns = [
        'user_id_combined', 'user_type', 'llm_provider', 'llm_model', 'llm_model_raw', 'model_type',
        'generation_type', 'prompt_tokens', 'completion_tokens',
        'generated_image_count', 'duration_seconds', 'generate_audio',
        'pricing_type', 'unit_price_prompt', 'unit_price_completion',
        'unit_price_request', 'cost_usd'
    ]
    # 只取存在的列
    export_cols = [c for c in detail_columns if c in df.columns]

    # 写入表头
    col_labels = {
        'user_id_combined': '用户', 'user_type': '用户类型',
        'llm_provider': '模型平台', 'llm_model': '模型', 'llm_model_raw': '原始模型',
        'model_type': '模型类型',
        'generation_type': '生成类型', 'generate_audio': '有声',
        'pricing_type': '计价方式',
        'unit_price_prompt': '输入单价($/M tokens)',
        'unit_price_completion': '输出单价($/M tokens)',
        'unit_price_request': '次单价($/次)',
        'cost_usd': '成本(USD)'
    }
    for col_idx, col_name in enumerate(export_cols, 1):
        label = col_labels.get(col_name, col_name)
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(export_cols, 1):
            val = row_data[col_name]
            # 处理NaN
            if pd.isna(val):
                val = ''
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # 成本列格式
            if col_name == 'cost_usd' and val != '':
                cell.number_format = '#,##0.0000'
            # 单价格式
            if col_name in ('unit_price_prompt', 'unit_price_completion', 'unit_price_request') and val != '':
                cell.number_format = '#,##0.0000'

    # 设置列宽
    col_widths = {
        'user_id_combined': 35, 'user_type': 12, 'llm_provider': 14,
        'llm_model': 42, 'generation_type': 10, 'generate_audio': 8,
        'pricing_type': 18, 'unit_price_prompt': 20,
        'unit_price_completion': 20, 'unit_price_request': 14,
        'cost_usd': 14
    }
    for col_idx, col_name in enumerate(export_cols, 1):
        from openpyxl.utils import get_column_letter
        ws2.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # 保存
    if output_path is None:
        csv_dir = os.path.dirname(csv_path)
        csv_name = os.path.splitext(os.path.basename(csv_path))[0]
        output_path = os.path.join(csv_dir, f"{csv_name}_summary.xlsx")

    wb.save(output_path)

    # 打印总成本
    total_cost = df['cost_usd'].sum()
    print(f"总成本: ${total_cost:.4f}")

    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python aggregate_llm_usage.py <csv_path> [output_path] [pricing_config_path]")
        sys.exit(1)

    csv_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    pricing_config_path = sys.argv[3] if len(sys.argv) > 3 else None

    result = aggregate_llm_usage(csv_path, output_path, pricing_config_path)
    print(f"输出文件: {result}")
