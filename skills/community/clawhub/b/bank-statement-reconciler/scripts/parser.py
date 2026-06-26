"""
Bank Statement and Order Parser
Supports CSV, Excel, PDF, JSON formats
"""
import re
import json
import csv
import io
from datetime import datetime
from typing import List, Dict, Optional, Tuple


# Column mapping for different bank formats
BANK_COLUMN_MAPPINGS = {
    # Chinese Bank (BOC)
    "boc": {
        "date": ["交易日期", "日期", "交易时间", "记账日期"],
        "amount": ["交易金额", "金额", "发生额", "支出", "收入"],
        "counterparty": ["对方账户", "对方户名", "交易对方", "对方"],
        "balance": ["余额", "账户余额", "可用余额"],
        "summary": ["摘要", "备注", "说明", "用途"],
        "type": ["收支", "交易类型", "类型"],
    },
    # ICBC
    "icbc": {
        "date": ["日期", "交易日期", "记账日期"],
        "amount": ["金额", "交易金额", "发生额"],
        "counterparty": ["对方户名", "对方账户", "收款人", "付款人"],
        "balance": ["余额", "账户余额"],
        "summary": ["摘要", "备注", "用途"],
        "type": ["交易类型", "收支"],
    },
    # CCB
    "ccb": {
        "date": ["交易时间", "日期", "交易日期"],
        "amount": ["交易金额", "金额", "支出金额", "收入金额"],
        "counterparty": ["对方账户", "对方户名", "收款人", "付款人"],
        "balance": ["余额"],
        "summary": ["备注", "摘要", "用途"],
        "type": ["交易类型"],
    },
    # ABC (Agricultural Bank)
    "abc": {
        "date": ["交易日期", "日期"],
        "amount": ["金额", "交易金额", "支出金额", "收入金额"],
        "counterparty": ["对方姓名", "对方账户", "收款人"],
        "balance": ["余额"],
        "summary": ["用途", "摘要", "备注"],
        "type": ["交易类型", "收支"],
    },
    # Alipay
    "alipay": {
        "date": ["交易时间", "日期时间", "创建时间"],
        "amount": ["金额", "交易金额", "收入", "支出"],
        "counterparty": ["对方", "交易对方", "收款人", "付款人"],
        "balance": ["余额"],
        "summary": ["说明", "商品说明", "备注"],
        "type": ["状态", "交易类型"],
    },
    # WeChat Pay
    "wechat": {
        "date": ["交易时间", "日期时间"],
        "amount": ["交易金额", "金额", "收支金额"],
        "counterparty": ["交易对方", "对手方", "商户"],
        "balance": [],
        "summary": ["备注", "备注信息"],
        "type": ["交易类型", "状态"],
    },
    # PayPal
    "paypal": {
        "date": ["Date", "Transaction Date", "日期"],
        "amount": ["Amount", "Net", "交易金额"],
        "counterparty": ["Name", "Counterparty", "From", "To", "交易对方"],
        "balance": ["Balance"],
        "summary": ["Subject", "Item", "说明", "备注"],
        "type": ["Status", "Type", "状态", "类型"],
    },
    # Stripe
    "stripe": {
        "date": ["Created", "Date", "结算日期"],
        "amount": ["Amount", "Gross", "金额"],
        "counterparty": ["Description", "Customer", "商户", "描述"],
        "balance": [],
        "summary": ["Note", "备注"],
        "type": ["Status", "状态", "Type"],
    },
    # Amazon
    "amazon": {
        "date": ["Order Date", "Date", "日期", "order date"],
        "amount": ["Item Total", "Order Total", "Amount", "金额", "order total"],
        "counterparty": ["Buyer", "Customer", "买家"],
        "balance": [],
        "summary": ["Product Name", "Item", "商品"],
        "type": ["Order Status", "Status", "状态"],
    },
    # Shopify
    "shopify": {
        "date": ["Created", "Date", "日期", "created at"],
        "amount": ["Total", "Order Total", "Amount", "金额"],
        "counterparty": ["Name", "Customer", "Customer Name", "买家"],
        "balance": [],
        "summary": ["Lineitem name", "Product", "商品"],
        "type": ["Financial Status", "Fulfillment Status", "状态"],
    },
    # Temu
    "temu": {
        "date": ["Date", "Order Date", "日期", "创建时间"],
        "amount": ["Amount", "Order Amount", "金额", "实付金额"],
        "counterparty": ["Supplier", "Merchant", "商户", "供应商"],
        "balance": [],
        "summary": ["Product", "Item", "商品", "备注"],
        "type": ["Status", "Order Status", "状态"],
    },
}


def detect_format(text: str, filename: str = "") -> str:
    """Auto-detect the statement format from content and filename."""
    # Check filename hints
    filename_lower = filename.lower()
    if "alipay" in filename_lower or "支付宝" in filename_lower:
        return "alipay"
    if "wechat" in filename_lower or "微信" in filename_lower:
        return "wechat"
    if "paypal" in filename_lower:
        return "paypal"
    if "stripe" in filename_lower:
        return "stripe"
    if "amazon" in filename_lower or "亚马逊" in filename_lower:
        return "amazon"
    if "shopify" in filename_lower:
        return "shopify"
    if "temu" in filename_lower:
        return "temu"
    if "icbc" in filename_lower or "工商银行" in filename_lower:
        return "icbc"
    if "ccb" in filename_lower or "建设银行" in filename_lower:
        return "ccb"
    if "boc" in filename_lower or "中国银行" in filename_lower:
        return "boc"
    
    # Check content patterns
    text_lower = text.lower()
    if "交易时间" in text and "交易对方" in text:
        if "支付宝" in text:
            return "alipay"
        if "微信" in text:
            return "wechat"
    if "对方户名" in text or "对方账户" in text:
        if "工商银行" in text or "icbc" in text_lower:
            return "icbc"
        if "建设银行" in text or "ccb" in text_lower:
            return "ccb"
        if "中国银行" in text or "boc" in text_lower:
            return "boc"
        if "农业银行" in text:
            return "abc"
        return "boc"  # Default for Chinese bank format
    if "transaction date" in text_lower and "amount" in text_lower:
        if "counterparty" in text_lower or "name" in text_lower:
            return "paypal"
        if "stripe" in text_lower:
            return "stripe"
    if "order date" in text_lower or "order total" in text_lower:
        if "amazon" in text_lower:
            return "amazon"
        if "shopify" in text_lower:
            return "shopify"
    if "temu" in text_lower:
        return "temu"
    
    # Default
    return "boc"


class StatementParser:
    """Parser for bank statements."""
    
    def parse_file(self, filepath: str, format_type: str = "auto") -> List[Dict]:
        """Parse a bank statement file."""
        filepath = filepath.strip()
        
        if filepath.endswith(".csv"):
            return self._parse_csv(filepath, format_type)
        elif filepath.endswith((".xlsx", ".xls")):
            return self._parse_excel(filepath, format_type)
        elif filepath.endswith(".pdf"):
            return self._parse_pdf(filepath, format_type)
        elif filepath.endswith(".json"):
            return self._parse_json(filepath, format_type)
        else:
            # Try CSV first
            try:
                return self._parse_csv(filepath, format_type)
            except Exception:
                raise ValueError(f"Unsupported file format: {filepath}")
    
    def parse_text(self, text: str, format_type: str = "auto") -> List[Dict]:
        """Parse bank statement from text content."""
        if format_type == "auto":
            format_type = detect_format(text)
        
        lines = text.strip().split("\n")
        return self._parse_lines(lines, format_type)
    
    def _parse_csv(self, filepath: str, format_type: str) -> List[Dict]:
        """Parse CSV file."""
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        if not rows:
            return []
        
        return self._map_columns(rows, format_type)
    
    def _parse_excel(self, filepath: str, format_type: str) -> List[Dict]:
        """Parse Excel file."""
        import openpyxl
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            # Try second row
            headers = [cell.value for cell in ws[2]]
        
        # Get data rows
        rows = []
        for row_idx in range(2 if headers else 1, ws.max_row + 1):
            row = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    row[header] = ws.cell(row=row_idx, column=col_idx).value
            if any(v is not None for v in row.values()):
                rows.append(row)
        
        if not rows:
            return []
        
        return self._map_columns(rows, format_type)
    
    def _parse_pdf(self, filepath: str, format_type: str) -> List[Dict]:
        """Parse PDF file using doc_parse."""
        import subprocess
        
        result = subprocess.run(
            ["miaoda-studio-cli", "doc-parse", "--file", filepath, "--output", "text"],
            capture_output=True, text=True, timeout=60
        )
        
        if result.returncode != 0:
            raise ValueError(f"PDF parsing failed: {result.stderr}")
        
        text = result.stdout
        lines = text.strip().split("\n")
        return self._parse_lines(lines, format_type)
    
    def _parse_json(self, filepath: str, format_type: str) -> List[Dict]:
        """Parse JSON file."""
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            if "transactions" in data:
                items = data["transactions"]
            elif "data" in data:
                items = data["data"]
            elif "records" in data:
                items = data["records"]
            else:
                items = [data]
        else:
            raise ValueError("Unknown JSON structure")
        
        return self._map_columns(items, format_type)
    
    def _parse_lines(self, lines: List[str], format_type: str) -> List[Dict]:
        """Parse text lines into transaction records."""
        transactions = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Try to extract date, amount, counterparty from line
            # This is a fallback for parsed PDF text
            trans = self._extract_from_line(line, format_type)
            if trans:
                transactions.append(trans)
        
        return transactions
    
    def _extract_from_line(self, line: str, format_type: str) -> Optional[Dict]:
        """Extract transaction info from a text line."""
        # Date patterns
        date_patterns = [
            r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
            r"(\d{4}年\d{1,2}月\d{1,2}日)",
            r"(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        ]
        
        # Amount patterns
        amount_patterns = [
            r"[¥$]?\s*([-+]?\d{1,3}(?:,\d{3})*(?:\.\d{2})?)",
            r"([-+]?\d+\.\d{2})",
        ]
        
        date = None
        for pattern in date_patterns:
            match = re.search(pattern, line)
            if match:
                date = match.group(1)
                break
        
        amount = None
        for pattern in amount_patterns:
            match = re.search(pattern, line)
            if match:
                amount_str = match.group(1).replace(",", "")
                try:
                    amount = float(amount_str)
                    break
                except ValueError:
                    continue
        
        if date or amount:
            return {
                "date": date,
                "amount": amount,
                "counterparty": line[:50] if len(line) > 50 else line,
                "raw": line,
                "summary": "",
            }
        
        return None
    
    def _map_columns(self, rows: List[Dict], format_type: str) -> List[Dict]:
        """Map column names to standard fields."""
        if format_type == "auto":
            # Detect from first row
            if rows:
                sample = rows[0]
                format_type = detect_format(str(sample))
        
        mapping = BANK_COLUMN_MAPPINGS.get(format_type, BANK_COLUMN_MAPPINGS["boc"])
        
        transactions = []
        for row in rows:
            trans = {
                "date": None,
                "amount": None,
                "counterparty": None,
                "balance": None,
                "summary": None,
                "type": None,
                "raw": str(row),
            }
            
            # Map each field
            for field, possible_names in mapping.items():
                for name in possible_names:
                    if name in row and row[name] is not None:
                        value = row[name]
                        
                        if field == "date":
                            trans["date"] = self._parse_date(value)
                        elif field == "amount":
                            trans["amount"] = self._parse_amount(value)
                        else:
                            trans[field] = str(value) if value else ""
                        break
            
            # Only add if has date or amount
            if trans["date"] or trans["amount"]:
                transactions.append(trans)
        
        return transactions
    
    def _parse_date(self, value) -> Optional[str]:
        """Parse date value."""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        
        value = str(value).strip()
        
        # Try common formats
        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y%m%d",
            "%Y年%m月%d日",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        return value
    
    def _parse_amount(self, value) -> Optional[float]:
        """Parse amount value."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        value = str(value).strip()
        
        # Remove currency symbols and commas
        value = re.sub(r"[¥$,，]", "", value)
        
        # Handle parentheses as negative
        if value.startswith("(") and value.endswith(")"):
            value = "-" + value[1:-1]
        
        # Remove parentheses
        value = re.sub(r"[()]", "", value)
        
        try:
            return float(value)
        except ValueError:
            return None


class OrderParser:
    """Parser for orders and invoices."""
    
    ORDER_COLUMN_MAPPINGS = {
        "invoice": {
            "date": ["开票日期", "发票日期", "日期", "invoice date"],
            "amount": ["金额", "发票金额", "税额", "total", "amount"],
            "counterparty": ["购买方", "客户", "buyer", "customer"],
            "invoice_no": ["发票号", "invoice number", "invoice_no"],
            "summary": ["商品名称", "项目", "description", "品名"],
        },
        "order": {
            "date": ["订单日期", "下单日期", "日期", "order date", "created"],
            "amount": ["订单金额", "金额", "total", "amount", "order total"],
            "counterparty": ["客户", "买家", "customer", "buyer"],
            "order_no": ["订单号", "order id", "order_no", "name"],
            "summary": ["商品", "商品名称", "product", "item"],
        },
        "auto": {
            "date": ["日期", "date", "order date", "invoice date", "订单日期", "开票日期"],
            "amount": ["金额", "amount", "total", "order total", "invoice amount", "订单金额", "发票金额"],
            "counterparty": ["客户", "customer", "buyer", "买家", "购买方"],
            "order_no": ["订单号", "order id", "invoice number", "订单ID", "invoice_no"],
            "summary": ["商品", "product", "description", "商品名称", "项目"],
        },
    }
    
    def parse_file(self, filepath: str, order_type: str = "auto") -> List[Dict]:
        """Parse orders/invoices file."""
        filepath = filepath.strip()
        
        if filepath.endswith(".csv"):
            return self._parse_csv(filepath, order_type)
        elif filepath.endswith((".xlsx", ".xls")):
            return self._parse_excel(filepath, order_type)
        elif filepath.endswith(".pdf"):
            return self._parse_pdf(filepath, order_type)
        elif filepath.endswith(".json"):
            return self._parse_json(filepath, order_type)
        else:
            try:
                return self._parse_csv(filepath, order_type)
            except Exception:
                raise ValueError(f"Unsupported order file format: {filepath}")
    
    def parse_text(self, text: str, order_type: str = "auto") -> List[Dict]:
        """Parse orders from text content."""
        lines = text.strip().split("\n")
        return self._parse_lines(lines, order_type)
    
    def _parse_csv(self, filepath: str, order_type: str) -> List[Dict]:
        """Parse CSV file."""
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        
        if not rows:
            return []
        
        return self._map_columns(rows, order_type)
    
    def _parse_excel(self, filepath: str, order_type: str) -> List[Dict]:
        """Parse Excel file."""
        import openpyxl
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            headers = [cell.value for cell in ws[2]]
        
        rows = []
        for row_idx in range(2 if headers else 1, ws.max_row + 1):
            row = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    row[header] = ws.cell(row=row_idx, column=col_idx).value
            if any(v is not None for v in row.values()):
                rows.append(row)
        
        if not rows:
            return []
        
        return self._map_columns(rows, order_type)
    
    def _parse_pdf(self, filepath: str, order_type: str) -> List[Dict]:
        """Parse PDF file."""
        import subprocess
        
        result = subprocess.run(
            ["miaoda-studio-cli", "doc-parse", "--file", filepath, "--output", "text"],
            capture_output=True, text=True, timeout=60
        )
        
        if result.returncode != 0:
            raise ValueError(f"PDF parsing failed: {result.stderr}")
        
        text = result.stdout
        lines = text.strip().split("\n")
        return self._parse_lines(lines, order_type)
    
    def _parse_json(self, filepath: str, order_type: str) -> List[Dict]:
        """Parse JSON file."""
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            if "orders" in data:
                items = data["orders"]
            elif "invoices" in data:
                items = data["invoices"]
            elif "data" in data:
                items = data["data"]
            else:
                items = [data]
        else:
            raise ValueError("Unknown JSON structure")
        
        return self._map_columns(items, order_type)
    
    def _parse_lines(self, lines: List[str], order_type: str) -> List[Dict]:
        """Parse text lines into order records."""
        orders = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            order = self._extract_from_line(line, order_type)
            if order:
                orders.append(order)
        
        return orders
    
    def _extract_from_line(self, line: str, order_type: str) -> Optional[Dict]:
        """Extract order info from text line."""
        date_patterns = [
            r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
            r"(\d{4}年\d{1,2}月\d{1,2}日)",
        ]
        
        amount_patterns = [
            r"[¥$]?\s*([-+]?\d{1,3}(?:,\d{3})+(?:\.\d{2})?)",
            r"([-+]?\d+\.\d{2})",
        ]
        
        date = None
        for pattern in date_patterns:
            match = re.search(pattern, line)
            if match:
                date = match.group(1)
                break
        
        amount = None
        for pattern in amount_patterns:
            match = re.search(pattern, line)
            if match:
                amount_str = match.group(1).replace(",", "")
                try:
                    amount = float(amount_str)
                    break
                except ValueError:
                    continue
        
        if date or amount:
            return {
                "date": date,
                "amount": amount,
                "counterparty": line[:50] if len(line) > 50 else line,
                "order_no": "",
                "raw": line,
                "summary": "",
            }
        
        return None
    
    def _map_columns(self, rows: List[Dict], order_type: str) -> List[Dict]:
        """Map column names to standard fields."""
        mapping = self.ORDER_COLUMN_MAPPINGS.get(order_type, self.ORDER_COLUMN_MAPPINGS["auto"])
        
        orders = []
        for row in rows:
            order = {
                "date": None,
                "amount": None,
                "counterparty": None,
                "order_no": None,
                "summary": None,
                "raw": str(row),
            }
            
            for field, possible_names in mapping.items():
                for name in possible_names:
                    if name in row and row[name] is not None:
                        value = row[name]
                        
                        if field == "date":
                            order["date"] = self._parse_date(value)
                        elif field == "amount":
                            order["amount"] = self._parse_amount(value)
                        else:
                            order[field] = str(value) if value else ""
                        break
            
            if order["date"] or order["amount"]:
                orders.append(order)
        
        return orders
    
    def _parse_date(self, value) -> Optional[str]:
        """Parse date value."""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        
        value = str(value).strip()
        
        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y%m%d",
            "%Y年%m月%d日",
            "%m/%d/%Y",
            "%Y-%m-%d %H:%M:%S",
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        return value
    
    def _parse_amount(self, value) -> Optional[float]:
        """Parse amount value."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        value = str(value).strip()
        value = re.sub(r"[¥$,，]", "", value)
        
        if value.startswith("(") and value.endswith(")"):
            value = "-" + value[1:-1]
        
        value = re.sub(r"[()]", "", value)
        
        try:
            return float(value)
        except ValueError:
            return None
