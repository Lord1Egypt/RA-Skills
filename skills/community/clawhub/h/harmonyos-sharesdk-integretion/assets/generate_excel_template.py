#!/usr/bin/env python3
"""
生成 ShareSDK HarmonyOS 配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_sheet_with_header(wb, sheet_name, headers, data):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 34
    ws.freeze_panes = "A2"
    return ws


def create_basic_info_sheet(wb):
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "MobTech ShareSDK 的 AppKey，从 MobTech 控制台获取", ""],
        ["appSecret", "MobTech ShareSDK 的 AppSecret，从 MobTech 控制台获取", ""],
        ["bundleName", "HarmonyOS 应用包名或项目标识，用于识别目标工程", ""],
        ["apiVersion", "HarmonyOS API 版本，官方文档要求 >= 12", "12"],
        ["enableHuaweiAuth", "是否启用华为授权：是/否", "否"],
        ["clientId", "华为授权的 Client ID；仅在启用华为授权时必填", ""],
        ["enableWeibo", "是否启用新浪微博授权/分享：是/否", "否"],
        ["weiboAppKey", "微博平台 AppKey；仅在启用微博时必填", ""],
        ["weiboRedirectUrl", "微博平台 RedirectURI；仅在启用微博时必填", ""],
        ["weiboCallbackAbilityName", "微博回呼 Ability 名称，例如 EntryAbility；仅在启用微博时必填", ""],
        ["weiboScope", "微博授权 Scope；选填", ""],
        ["enableWechat", "是否启用微信授权/分享：是/否", "否"],
        ["wechatAppKey", "微信平台 AppKey；仅在启用微信时必填", ""],
        ["wechatAppSecret", "微信平台 AppSecret；仅在启用微信时必填", ""],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)
    for row in [2, 3, 4, 5]:
        ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    return ws


def create_privacy_sheet(wb):
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["privacyPolicyReady", "是否已有隐私政策并已准备在首次冷启动时展示：是/否", "否"],
        ["privacyPopupReady", "是否已有隐私弹窗并由用户主动点击同意：是/否", "否"],
        ["discloseSdkName", "隐私政策中是否披露 ShareSDK 名称：是/否", "否"],
        ["discloseCompany", "隐私政策中是否披露第三方主体上海掌之淘信息技术有限公司：是/否", "否"],
        ["disclosePolicyLink", "隐私政策中是否披露 SDK 隐私政策链接：是/否", "否"],
        ["postIntegrationAdjustNote", "集成完成后是否需要 Agent 在项目文档中补充扩展业务功能后续调整说明：是/否", "是"],
    ]
    return create_sheet_with_header(wb, "隐私合规", headers, data)


def create_instructions_sheet(wb):
    headers = ["说明项", "详细内容"]
    data = [
        [
            "填写步骤",
            "1. 在 MobTech 控制台获取 appKey 和 appSecret\n"
            "2. 如需华为授权，到 AppGallery Connect 获取 Client ID\n"
            "3. 如需微博授权/分享，补齐微博 AppKey、RedirectURI、CallbackAbilityName\n"
            "4. 如需微信授权/分享，补齐微信 AppKey、AppSecret\n"
            "5. 基础集成阶段仅主动声明 INTERNET 权限，其余权限不在模板中单独选择\n"
            "6. 隐私回调位置、分享落点、授权落点由 Agent 在后续步骤中通过对话询问\n"
            "7. 保存后告诉 Agent“填好了”",
        ],
        [
            "必填项",
            "基础信息 Sheet 中的 appKey、appSecret、bundleName、apiVersion 为必填；当 enableHuaweiAuth=是 时，clientId 为必填；当 enableWeibo=是 时，weiboAppKey、weiboRedirectUrl、weiboCallbackAbilityName 为必填；当 enableWechat=是 时，wechatAppKey、wechatAppSecret 为必填",
        ],
        [
            "布尔字段规则",
            "支持填写：是/否、true/false、TRUE/FALSE；Agent 会统一转换为 true/false",
        ],
        [
            "默认策略",
            "本模板不提供权限和可选能力开关。基础集成阶段仅主动声明 INTERNET 权限；GET_NETWORK_INFO、APP_TRACKING_CONSENT、APPROXIMATELY_LOCATION、GET_WIFI_INFO 等权限不主动写入 module.json5，后续如有需要再补充",
        ],
        [
            "后续调整",
            "如果项目后续需要收敛数据采集能力，Agent 会在集成完成后的项目文档中补充 ShareSDK 扩展业务功能设置说明，包括 ZztCustomController、submitPolicyGrantResult(..., controller) 和 updateZztCustomController(...) 的使用方式；如启用微博/微信，也会补充 querySchemes、setPlatformDevInfoAsync(...) 和 handlerWant(...) 的平台差异说明",
        ],
        [
            "官方文档",
            "集成指南：https://www.mob.com/wiki/detailed?wiki=696&id=14\n"
            "扩展业务功能设置：https://www.mob.com/wiki/detailed?wiki=711&id=14\n"
            "鸿蒙端合规使用说明：https://www.mob.com/wiki/detailed?wiki=748&id=14\n"
            "新浪微博授权与分享：https://www.mob.com/wiki/detailed?wiki=722&id=14\n"
            "微信授权与分享：https://www.mob.com/wiki/detailed?wiki=724&id=14",
        ],
        [
            "文档未明确项",
            "authorize(params) 的 params 结构、ArkTS 的实际 import 语句、是否必须修改 oh-package.json5 / build-profile.json5 / app.json5 均需以实际工程或依赖导出为准",
        ],
    ]
    return create_sheet_with_header(wb, "填写说明", headers, data)


def main():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    target_order = ["基础信息", "隐私合规", "填写说明"]
    for idx, name in enumerate(target_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    output_path = Path(__file__).resolve().parent / "ShareSDK_HarmonyOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
