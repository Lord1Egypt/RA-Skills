import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from collections import defaultdict

# 读取Excel文件 - 使用data_only=True获取计算后的值
file_path = "user_input_files/发货清单（2026-4-24）.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

def get_cell_value(ws, row, col):
    """获取单元格值"""
    return ws.cell(row=row, column=col).value

# 只过滤：支架、膨胀、螺丝、镀锌
ignore_keywords = ['支架', '膨胀', '螺丝', '镀锌']

# 用于存储数据的字典
data_dict = defaultdict(lambda: {'count': [], 'qty': []})

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 找到表头行
    header_row = None
    for row_idx in range(1, min(15, ws.max_row + 1)):
        row_values = [get_cell_value(ws, row_idx, col) for col in range(1, 12)]
        if '序号' in str(row_values[0]) or '序号' in str(row_values):
            header_row = row_idx
            break

    if header_row is None:
        continue

    # 找到关键列索引
    name_col = None
    spec_col = None
    count_col = None
    qty_col = None

    for col in range(1, ws.max_column + 1):
        header = get_cell_value(ws, header_row, col)
        if header:
            header_str = str(header).strip()
            if '项目名称' in header_str or '部件名称' in header_str:
                name_col = col
            elif '生产规格' in header_str:
                spec_col = col
            elif header_str == '件':
                count_col = col
            elif '数量' in header_str and ('M)' in header_str or 'm)' in header_str):
                qty_col = col

    # 备用方案
    if name_col is None:
        name_col = 2
    if spec_col is None:
        spec_col = 4
    if count_col is None:
        for col in range(1, ws.max_column + 1):
            header = get_cell_value(ws, header_row, col)
            if header and str(header).strip() == '件':
                count_col = col
                break
    if qty_col is None:
        for col in range(1, ws.max_column + 1):
            header = get_cell_value(ws, header_row, col)
            if header and '数量' in str(header) and '合计' not in str(header):
                qty_col = col
                break

    # 处理数据行
    start_row = header_row + 1
    for row_idx in range(start_row, ws.max_row + 1):
        name = get_cell_value(ws, row_idx, name_col)
        spec = get_cell_value(ws, row_idx, spec_col)

        # 标准化
        name_str = str(name).strip() if name else ''
        spec_str = str(spec).strip() if spec else ''

        # 跳过名称和规格都为空或None的行
        if not name_str and not spec_str:
            continue
        if name_str == 'None' and spec_str == 'None':
            continue

        # 跳过包含忽略关键字的行
        if name_str and any(kw in name_str for kw in ignore_keywords):
            continue
        if spec_str and any(kw in spec_str for kw in ignore_keywords):
            continue

        # 跳过包含关键字的行
        skip_keywords = ['包装', '合计', '税金', '总计', '说明', '甲方', '现场', '施工', '制单人', '甲方工作人员', '现场门卫', '施工单位']
        if name_str and any(kw in name_str for kw in skip_keywords):
            continue

        # 获取件数
        count = get_cell_value(ws, row_idx, count_col) if count_col else None
        if isinstance(count, (int, float)):
            count = int(count) if count == int(count) else count
        elif count is None or count == '':
            count = 0
        else:
            count = 0

        # 获取数量
        qty = get_cell_value(ws, row_idx, qty_col) if qty_col else None
        if isinstance(qty, (int, float)):
            qty = qty
        elif qty is None or qty == '':
            qty = 0
        else:
            qty = 0

        # 只有当件数或数量有有效值时才添加
        if count > 0 or qty > 0:
            key = (name_str if name_str else spec_str, spec_str if spec_str else name_str)
            if count > 0:
                data_dict[key]['count'].append(count)
            if qty > 0:
                data_dict[key]['qty'].append(qty)

wb.close()

# 创建新的工作簿
new_wb = openpyxl.Workbook()
ws = new_wb.active
ws.title = "合并发货清单"

# 设置表头
headers_list = ['序号', '项目名称及特征', '生产规格', '合计件数', '合计数量']
for col, header in enumerate(headers_list, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 添加边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def get_sort_key(item):
    """根据关键字优先级排序"""
    name = item[0][0]
    name_str = str(name).lower()

    priority = 999
    if '梁' in name_str:
        priority = 1
    elif '柱' in name_str:
        priority = 2
    elif '档' in name_str:
        priority = 3
    elif '板' in name_str:
        priority = 4
    elif '椽' in name_str:
        priority = 5
    elif '枋' in name_str:
        priority = 6
    elif '机' in name_str:
        priority = 7
    elif '戗' in name_str:
        priority = 8
    else:
        priority = 9

    if not name_str or len(name_str) < 2:
        return (priority, name_str, item[0][1])

    return (priority, name_str, item[0][1])

# 按关键字优先级排序
sorted_items = sorted(data_dict.items(), key=get_sort_key)

# 填充数据
row_num = 2
for idx, (key, data) in enumerate(sorted_items, 1):
    total_count = sum(data['count'])
    total_qty = sum(data['qty'])

    ws.cell(row=row_num, column=1, value=idx)
    ws.cell(row=row_num, column=2, value=key[0])
    ws.cell(row=row_num, column=3, value=key[1])
    ws.cell(row=row_num, column=4, value=int(total_count) if total_count == int(total_count) else round(total_count, 2) if total_count > 0 else '')
    ws.cell(row=row_num, column=5, value=round(total_qty, 2) if total_qty > 0 else '')

    for col in range(1, 6):
        ws.cell(row=row_num, column=col).border = thin_border

    row_num += 1

# 调整列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

# 保存文件
output_file = "/workspace/合并发货清单汇总.xlsx"
new_wb.save(output_file)

print(f"文件已保存到: {output_file}")
print(f"共合并 {len(data_dict)} 个不同项目")
print(f"排序优先级: 梁 > 柱 > 档 > 板 > 椽 > 枋 > 机 > 戗 > 其他")