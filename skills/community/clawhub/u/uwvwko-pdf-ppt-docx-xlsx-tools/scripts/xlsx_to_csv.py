"""XLSX → CSV (每个 sheet 导出为单独 CSV)

用法:
    python xlsx_to_csv.py input.xlsx [output_dir]

参数:
    input.xlsx  - XLSX 文件路径
    output_dir  - 输出目录 (默认: ./xlsx_csv_output)
"""
import openpyxl
import csv
import sys
import os


def main():
    if len(sys.argv) < 2:
        print("用法: python xlsx_to_csv.py input.xlsx [output_dir]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else "./xlsx_csv_output"

    if not os.path.isfile(xlsx_path):
        print(f"错误: 文件不存在 - {xlsx_path}")
        sys.exit(1)

    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name.replace(" ", "_").replace("/", "_")
        csv_path = os.path.join(out_dir, f"{safe_name}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
        print(f"导出: {csv_path}")

    total = len(wb.sheetnames)
    wb.close()
    print(f"完成: {total} 个 sheet -> {out_dir}")


if __name__ == "__main__":
    main()