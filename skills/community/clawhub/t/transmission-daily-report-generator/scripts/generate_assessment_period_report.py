#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成日报报表（按考核周期统计）
每个考核周期一个 sheet，最后保留原始清单
"""

import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 配置
INPUT_DIR = "/Users/ahs/.openclaw/workspace/传输单边故障/output"
OUTPUT_DIR = "/Users/ahs/.openclaw/workspace/传输单边故障/output"

# 超时标准
TIMEOUT_STANDARDS = {
    '汇聚骨干单边': 4,
    '重要环': 12,
    '一般环': 24
}

def get_assessment_period(year, month):
    """
    获取考核周期的起止日期
    考核周期：上个月27日 - 本月26日

    Args:
        year: 年份（如 2026）
        month: 月份（如 2）

    Returns:
        (start_date, end_date, period_name)
    """
    # 计算上个月
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1

    # 考核周期：上个月27日 - 本月26日
    start_date = datetime(prev_year, prev_month, 27)
    end_date = datetime(year, month, 26, 23, 59, 59)

    # 周期名称：如"2026年2月（1月27日-2月26日）"
    period_name = f"{year}年{month}月（{prev_month}月27日-{month}月26日）"

    return start_date, end_date, period_name

def calculate_timeout_status(row):
    """计算是否超时"""
    fault_type = row['故障类型']
    duration = row['持续时长（小时）']

    if pd.isna(duration):
        return False

    if fault_type == '汇聚骨干单边':
        return duration > TIMEOUT_STANDARDS['汇聚骨干单边']
    elif fault_type == '接入环单边':
        is_important = row['重要环'] == '是'
        if is_important:
            return duration > TIMEOUT_STANDARDS['重要环']
        else:
            return duration > TIMEOUT_STANDARDS['一般环']
    else:
        return False

def get_fault_type(row):
    """获取故障类型（重要接入环单边、一般接入环单边）"""
    fault_type = row['故障类型']

    if fault_type == '汇聚骨干单边':
        return '汇聚骨干单边'
    elif fault_type == '接入环单边':
        is_important = row['重要环'] == '是'
        if is_important:
            return '重要接入环单边'
        else:
            return '一般接入环单边'
    else:
        return fault_type

def filter_by_period(df, start_date, end_date):
    """
    根据考核周期筛选数据

    Args:
        df: 原始数据
        start_date: 考核周期开始日期
        end_date: 考核周期结束日期

    Returns:
        筛选后的数据
    """
    # 确保告警发生时间是 datetime 类型
    df['告警发生时间'] = pd.to_datetime(df['告警发生时间'])

    # 筛选在考核周期内的数据
    filtered = df[
        (df['告警发生时间'] >= start_date) &
        (df['告警发生时间'] <= end_date)
    ]

    return filtered

def generate_period_report(df, period_name):
    """
    生成单个考核周期的报表数据

    Args:
        df: 该考核周期的数据
        period_name: 考核周期名称

    Returns:
        报表数据 DataFrame
    """
    # 计算是否超时
    df['是否超时'] = df.apply(calculate_timeout_status, axis=1)

    # 计算故障类型
    df['故障类型分类'] = df.apply(get_fault_type, axis=1)

    # 获取所有区县
    areas = sorted(df['所属片区'].unique())

    # 创建报表数据
    report_data = []

    for area in areas:
        area_data = df[df['所属片区'] == area]

        # 汇聚骨干单边
        backbone_data = area_data[area_data['故障类型分类'] == '汇聚骨干单边']
        backbone_count = len(backbone_data)
        backbone_avg_duration = backbone_data['持续时长（小时）'].mean() if backbone_count > 0 else 0
        backbone_timeout_count = backbone_data['是否超时'].sum() if backbone_count > 0 else 0
        backbone_timely_rate = ((backbone_count - backbone_timeout_count) / backbone_count * 100) if backbone_count > 0 else 0

        # 重要接入环单边
        important_data = area_data[area_data['故障类型分类'] == '重要接入环单边']
        important_count = len(important_data)
        important_avg_duration = important_data['持续时长（小时）'].mean() if important_count > 0 else 0
        important_timeout_count = important_data['是否超时'].sum() if important_count > 0 else 0
        important_timely_rate = ((important_count - important_timeout_count) / important_count * 100) if important_count > 0 else 0

        # 汇聚骨干和重要接入单边
        combined_data = area_data[
            (area_data['故障类型分类'] == '汇聚骨干单边') |
            (area_data['故障类型分类'] == '重要接入环单边')
        ]
        combined_count = len(combined_data)
        combined_avg_duration = combined_data['持续时长（小时）'].mean() if combined_count > 0 else 0
        combined_timeout_count = combined_data['是否超时'].sum() if combined_count > 0 else 0
        combined_timely_rate = ((combined_count - combined_timeout_count) / combined_count * 100) if combined_count > 0 else 0

        # 一般接入环单边（有业务）- 剔除风险评分为0的记录
        general_data = area_data[area_data['故障类型分类'] == '一般接入环单边']
        # 过滤掉风险评分为0的记录
        general_data = general_data[general_data['风险评分'] != 0]
        general_count = len(general_data)
        general_avg_duration = general_data['持续时长（小时）'].mean() if general_count > 0 else 0
        general_timeout_count = general_data['是否超时'].sum() if general_count > 0 else 0
        general_timely_rate = ((general_count - general_timeout_count) / general_count * 100) if general_count > 0 else 0

        report_data.append({
            '所属片区': area,
            '汇聚骨干单边_工单数量': backbone_count,
            '汇聚骨干单边_平均时长': round(backbone_avg_duration, 2),
            '汇聚骨干单边_超时数量': backbone_timeout_count,
            '汇聚骨干单边_及时率': round(backbone_timely_rate, 2),
            '重要接入环单边_工单数量': important_count,
            '重要接入环单边_平均时长': round(important_avg_duration, 2),
            '重要接入环单边_超时数量': important_timeout_count,
            '重要接入环单边_及时率': round(important_timely_rate, 2),
            '汇聚骨干和重要接入单边_工单数量': combined_count,
            '汇聚骨干和重要接入单边_平均时长': round(combined_avg_duration, 2),
            '汇聚骨干和重要接入单边_超时数量': combined_timeout_count,
            '汇聚骨干和重要接入单边_及时率': round(combined_timely_rate, 2),
            '一般接入环单边（有业务）_工单数量': general_count,
            '一般接入环单边（有业务）_平均时长': round(general_avg_duration, 2),
            '一般接入环单边（有业务）_超时数量': general_timeout_count,
            '一般接入环单边（有业务）_及时率': round(general_timely_rate, 2)
        })

    # 创建 DataFrame
    df_report = pd.DataFrame(report_data)

    # 添加全区汇总
    total_backbone_count = df_report['汇聚骨干单边_工单数量'].sum()
    total_backbone_avg_duration = df_report['汇聚骨干单边_平均时长'].mean()
    total_backbone_timeout_count = df_report['汇聚骨干单边_超时数量'].sum()
    total_backbone_timely_rate = ((total_backbone_count - total_backbone_timeout_count) / total_backbone_count * 100) if total_backbone_count > 0 else 0

    total_important_count = df_report['重要接入环单边_工单数量'].sum()
    total_important_avg_duration = df_report['重要接入环单边_平均时长'].mean()
    total_important_timeout_count = df_report['重要接入环单边_超时数量'].sum()
    total_important_timely_rate = ((total_important_count - total_important_timeout_count) / total_important_count * 100) if total_important_count > 0 else 0

    total_combined_count = df_report['汇聚骨干和重要接入单边_工单数量'].sum()
    total_combined_avg_duration = df_report['汇聚骨干和重要接入单边_平均时长'].mean()
    total_combined_timeout_count = df_report['汇聚骨干和重要接入单边_超时数量'].sum()
    total_combined_timely_rate = ((total_combined_count - total_combined_timeout_count) / total_combined_count * 100) if total_combined_count > 0 else 0

    total_general_count = df_report['一般接入环单边（有业务）_工单数量'].sum()
    total_general_avg_duration = df_report['一般接入环单边（有业务）_平均时长'].mean()
    total_general_timeout_count = df_report['一般接入环单边（有业务）_超时数量'].sum()
    total_general_timely_rate = ((total_general_count - total_general_timeout_count) / total_general_count * 100) if total_general_count > 0 else 0

    total_row = {
        '所属片区': '全区',
        '汇聚骨干单边_工单数量': total_backbone_count,
        '汇聚骨干单边_平均时长': round(total_backbone_avg_duration, 2),
        '汇聚骨干单边_超时数量': total_backbone_timeout_count,
        '汇聚骨干单边_及时率': round(total_backbone_timely_rate, 2),
        '重要接入环单边_工单数量': total_important_count,
        '重要接入环单边_平均时长': round(total_important_avg_duration, 2),
        '重要接入环单边_超时数量': total_important_timeout_count,
        '重要接入环单边_及时率': round(total_important_timely_rate, 2),
        '汇聚骨干和重要接入单边_工单数量': total_combined_count,
        '汇聚骨干和重要接入单边_平均时长': round(total_combined_avg_duration, 2),
        '汇聚骨干和重要接入单边_超时数量': total_combined_timeout_count,
        '汇聚骨干和重要接入单边_及时率': round(total_combined_timely_rate, 2),
        '一般接入环单边（有业务）_工单数量': total_general_count,
        '一般接入环单边（有业务）_平均时长': round(total_general_avg_duration, 2),
        '一般接入环单边（有业务）_超时数量': total_general_timeout_count,
        '一般接入环单边（有业务）_及时率': round(total_general_timely_rate, 2)
    }

    df_report = pd.concat([df_report, pd.DataFrame([total_row])], ignore_index=True)

    return df_report

def write_sheet_with_format(ws, df_report, period_name):
    """
    写入一个 sheet，包含格式化

    Args:
        ws: worksheet 对象
        df_report: 报表数据
        period_name: 考核周期名称
    """
    # 设置 sheet 标题（Excel sheet 名称限制在 31 个字符以内）
    # 保留完整的周期名称，如果超过 31 个字符则截取
    if len(period_name) > 31:
        ws.title = period_name[:31]
    else:
        ws.title = period_name

    # 写入表头
    # 第一行表头
    ws.cell(row=1, column=1, value='所属片区')
    ws.cell(row=1, column=2, value='汇聚骨干单边')
    ws.cell(row=1, column=6, value='重要接入环单边')
    ws.cell(row=1, column=10, value='汇聚骨干和重要接入单边')
    ws.cell(row=1, column=14, value='一般接入环单边（有业务）')

    # 第二行表头
    ws.cell(row=2, column=2, value='工单数量')
    ws.cell(row=2, column=3, value='平均时长')
    ws.cell(row=2, column=4, value='超时数量')
    ws.cell(row=2, column=5, value='及时率')
    ws.cell(row=2, column=6, value='工单数量')
    ws.cell(row=2, column=7, value='平均时长')
    ws.cell(row=2, column=8, value='超时数量')
    ws.cell(row=2, column=9, value='及时率')
    ws.cell(row=2, column=10, value='工单数量')
    ws.cell(row=2, column=11, value='平均时长')
    ws.cell(row=2, column=12, value='超时数量')
    ws.cell(row=2, column=13, value='及时率')
    ws.cell(row=2, column=14, value='工单数量')
    ws.cell(row=2, column=15, value='平均时长')
    ws.cell(row=2, column=16, value='超时数量')
    ws.cell(row=2, column=17, value='及时率')

    # 写入数据
    for i, (_, row) in enumerate(df_report.iterrows(), start=3):
        ws.cell(row=i, column=1, value=row['所属片区'])
        ws.cell(row=i, column=2, value=row['汇聚骨干单边_工单数量'])
        ws.cell(row=i, column=3, value=row['汇聚骨干单边_平均时长'])
        ws.cell(row=i, column=4, value=row['汇聚骨干单边_超时数量'])
        ws.cell(row=i, column=5, value=row['汇聚骨干单边_及时率'])
        ws.cell(row=i, column=6, value=row['重要接入环单边_工单数量'])
        ws.cell(row=i, column=7, value=row['重要接入环单边_平均时长'])
        ws.cell(row=i, column=8, value=row['重要接入环单边_超时数量'])
        ws.cell(row=i, column=9, value=row['重要接入环单边_及时率'])
        ws.cell(row=i, column=10, value=row['汇聚骨干和重要接入单边_工单数量'])
        ws.cell(row=i, column=11, value=row['汇聚骨干和重要接入单边_平均时长'])
        ws.cell(row=i, column=12, value=row['汇聚骨干和重要接入单边_超时数量'])
        ws.cell(row=i, column=13, value=row['汇聚骨干和重要接入单边_及时率'])
        ws.cell(row=i, column=14, value=row['一般接入环单边（有业务）_工单数量'])
        ws.cell(row=i, column=15, value=row['一般接入环单边（有业务）_平均时长'])
        ws.cell(row=i, column=16, value=row['一般接入环单边（有业务）_超时数量'])
        ws.cell(row=i, column=17, value=row['一般接入环单边（有业务）_及时率'])

    # 设置样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 应用表头样式
    for row in range(1, 3):
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    # 数据样式
    data_alignment = Alignment(horizontal='center', vertical='center')

    # 应用数据样式
    for row in range(3, len(df_report) + 3):
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border

    # 合并表头单元格
    ws.merge_cells('A1:A2')  # 所属片区
    ws.merge_cells('B1:E1')  # 汇聚骨干单边
    ws.merge_cells('F1:I1')  # 重要接入环单边
    ws.merge_cells('J1:M1')  # 汇聚骨干和重要接入单边
    ws.merge_cells('N1:Q1')  # 一般接入环单边（有业务）

    # 冻结窗格
    ws.freeze_panes = 'A3'

def generate_assessment_period_report():
    """生成按考核周期统计的日报报表"""
    print("正在生成按考核周期统计的日报报表...")

    # 读取结果 D
    result_d_file = os.path.join(INPUT_DIR, '结果D_最终数据_20260428_181530.xlsx')

    # 读取所有月份数据
    excel_file = pd.ExcelFile(result_d_file)
    all_data = []

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(result_d_file, sheet_name=sheet_name)
        all_data.append(df)

    # 合并所有数据
    all_data = pd.concat(all_data, ignore_index=True)

    # 确保告警发生时间是 datetime 类型
    all_data['告警发生时间'] = pd.to_datetime(all_data['告警发生时间'])

    # 获取数据的时间范围
    min_date = all_data['告警发生时间'].min()
    max_date = all_data['告警发生时间'].max()

    print(f"数据时间范围: {min_date} ~ {max_date}")

    # 确定需要统计的考核周期
    # 从数据的最早日期开始，到最晚日期结束
    periods = []

    # 找到第一个完整的考核周期
    # 如果数据从 2026-01-01 开始，第一个考核周期应该是 2026年1月（2025-12-27 ~ 2026-01-26）
    start_year = min_date.year
    start_month = min_date.month

    # 如果数据在 27 日之前，需要从上个月开始
    if min_date.day < 27:
        if start_month == 1:
            start_year -= 1
            start_month = 12
        else:
            start_month -= 1

    # 生成所有考核周期
    current_year = start_year
    current_month = start_month

    while True:
        start_date, end_date, period_name = get_assessment_period(current_year, current_month)

        # 如果这个考核周期的开始日期已经超过数据的最大日期，停止
        if start_date > max_date:
            break

        periods.append({
            'year': current_year,
            'month': current_month,
            'start_date': start_date,
            'end_date': end_date,
            'period_name': period_name
        })

        # 下一个考核周期
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1

    print(f"共 {len(periods)} 个考核周期:")
    for period in periods:
        print(f"  - {period['period_name']}")

    # 创建工作簿
    wb = Workbook()
    # 删除默认的 sheet
    wb.remove(wb.active)

    # 先创建全年累计 sheet（放在最前面）
    print("\n生成全年累计报表...")
    df_annual_report = generate_period_report(all_data, "全年累计")
    ws_annual = wb.create_sheet("全年累计")
    write_sheet_with_format(ws_annual, df_annual_report, "全年累计")
    print(f"  ✅ 全年累计报表已生成（{len(all_data)} 条数据）")

    # 为每个考核周期创建一个 sheet
    for period in periods:
        start_date = period['start_date']
        end_date = period['end_date']
        period_name = period['period_name']

        print(f"\n处理考核周期: {period_name}")

        # 筛选该考核周期的数据
        period_data = filter_by_period(all_data, start_date, end_date)

        if len(period_data) == 0:
            print(f"  ⚠️  该考核周期无数据")
            continue

        print(f"  数据量: {len(period_data)} 条")

        # 生成报表
        df_report = generate_period_report(period_data, period_name)

        # 创建 sheet
        ws = wb.create_sheet()

        # 写入数据
        write_sheet_with_format(ws, df_report, period_name)

        print(f"  ✅ 已生成报表")

    # 添加原始清单 sheet
    print("\n添加原始清单...")
    ws_raw = wb.create_sheet("原始清单")

    # 使用所有列
    all_columns = all_data.columns.tolist()

    # 写入表头
    for i, col in enumerate(all_columns, start=1):
        ws_raw.cell(row=1, column=i, value=col)

    # 写入数据
    for i, (_, row) in enumerate(all_data.iterrows(), start=2):
        for j, col in enumerate(all_columns, start=1):
            value = row[col]
            # 处理 NaN 值
            if pd.isna(value):
                value = ''
            ws_raw.cell(row=i, column=j, value=value)

    # 设置样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='left', vertical='center')

    # 表头样式
    for col in range(1, len(all_columns) + 1):
        cell = ws_raw.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据样式
    for row in range(2, len(all_data) + 2):
        for col in range(1, len(all_columns) + 1):
            cell = ws_raw.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border

    # 冻结窗格
    ws_raw.freeze_panes = 'A2'

    # 自动调整列宽
    for col in ws_raw.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column].width = adjusted_width

    print(f"  ✅ 原始清单已添加（{len(all_data)} 条记录）")

    # 保存报表
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_file = os.path.join(OUTPUT_DIR, f'日报报表_按考核周期_{timestamp}.xlsx')

    wb.save(report_file)

    print(f"\n✅ 报表已保存: {report_file}")
    print(f"  共 {len(wb.sheetnames)} 个 sheet:")
    for sheet_name in wb.sheetnames:
        print(f"    - {sheet_name}")

    return report_file

if __name__ == '__main__':
    generate_assessment_period_report()
