"""
腾讯文档归档台账更新脚本
用法: python update_ledger.py [--date YYYYMMDD] --url <docs.qq.com链接> [--url ...]
  --date: 下载日期，默认今天
  --url: 腾讯文档链接，可重复多次
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def get_ledger_path() -> Path:
    desktop = Path.home() / "Desktop" / "腾讯文档"
    desktop.mkdir(parents=True, exist_ok=True)
    return desktop / "腾讯文档链接台账.xlsx"


def ensure_ledger(ledger_path: Path) -> None:
    if not ledger_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "台账"
        ws.append(["下载日期", "腾讯文档链接"])
        wb.save(ledger_path)


def append_rows(ledger_path: Path, date: str, urls: list[str]) -> int:
    ensure_ledger(ledger_path)
    wb = openpyxl.load_workbook(ledger_path)
    ws = wb.active
    added = 0
    for url in urls:
        # deduplicate: skip if exact (date, url) row already exists
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == date and row[1] == url:
                break
        else:
            ws.append([date, url])
            added += 1
    wb.save(ledger_path)
    return added


def main():
    parser = argparse.ArgumentParser(description="更新腾讯文档归档台账")
    parser.add_argument("--date", default=datetime.now().strftime("%Y%m%d"), help="下载日期 YYYYMMDD")
    parser.add_argument("--url", action="append", required=True, help="腾讯文档链接（可多次指定）")
    args = parser.parse_args()

    ledger_path = get_ledger_path()
    added = append_rows(ledger_path, args.date, args.url)
    print(f"台账已更新: {ledger_path}")
    print(f"新增 {added} 条记录 (日期={args.date})")


if __name__ == "__main__":
    main()
