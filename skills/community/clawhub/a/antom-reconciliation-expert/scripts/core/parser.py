"""
Reconciliation Report Parser

Parses reconciliation report CSV/XLSX files provided by Antom (transaction details, settlement details, settlement summary), supports DSL filtering/aggregation (WHERE/SELECT/GROUP BY/ORDER BY/LIMIT).
"""

import csv
import json
import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from functools import reduce
import operator

from .constants import (
    ALL_FEE_FIELDS,
    SETTLEMENT_DETAIL_FILENAME_RE,
    SETTLEMENT_DETAIL_CORE_COLUMNS,
    SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS,
    SETTLEMENT_DETAIL_MIN_CORE_MATCH,
)


# ============================================================
# Settlement Detail report type detection (filename + lightweight content sanity)
# Filename gate is mandatory; content sanity catches obvious renaming/spoofing.
# ============================================================
ALLOWED_EXTENSIONS = (".csv", ".xlsx")


class ReportTypeError(ValueError):
    """Raised when a file is rejected by detect_report_type()."""

    def __init__(self, reason: str, file_path: str, kind: str):
        super().__init__(reason)
        self.reason = reason
        self.file_path = file_path
        # kind ∈ {"extension", "filename", "content"}
        self.kind = kind


def _read_csv_header(file_path: str) -> List[str]:
    for encoding in ("utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                return next(reader, []) or []
        except UnicodeDecodeError:
            continue
        except StopIteration:
            return []
    return []


def _read_xlsx_header(file_path: str) -> List[str]:
    from openpyxl import load_workbook

    # NOTE: read_only=True can return only the first column for xlsx files
    # exported by some Antom pipelines (sheet dimension attribute does not
    # cover the full row range). Use the default (non-read_only) mode and
    # iterate by ws.max_column to guarantee the entire header row is read.
    wb = load_workbook(filename=file_path, data_only=True)
    try:
        ws = wb.active
        max_col = ws.max_column or 0
        if max_col <= 0:
            return []
        return [
            (str(ws.cell(row=1, column=c).value).strip()
             if ws.cell(row=1, column=c).value is not None else "")
            for c in range(1, max_col + 1)
        ]
    finally:
        wb.close()


def detect_report_type(file_path: str) -> Dict[str, Any]:
    """Validate a file is a Settlement Detail report.

    Returns a dict with shape ``{"report_type": "SETTLEMENT_DETAIL", "file_path": ...}``
    on success. Raises ``ReportTypeError`` otherwise.

    Detection strategy (cheap → expensive):
      1. Extension whitelist: ``.csv`` / ``.xlsx`` only.
      2. Filename gate: must match ``SETTLEMENT_DETAIL_FILENAME_RE``.
      3. Content sanity (two-layer rule on the header row):
         a. POSITIVE: header must contain at least
            ``SETTLEMENT_DETAIL_MIN_CORE_MATCH`` columns from
            ``SETTLEMENT_DETAIL_CORE_COLUMNS`` (rejects Settlement Summary
            files renamed as Settlement Detail).
         b. NEGATIVE: header must NOT contain any column from
            ``SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS`` (rejects Transaction
            Detail files renamed as Settlement Detail).
    """
    file_name = os.path.basename(file_path)
    ext = os.path.splitext(file_name)[1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise ReportTypeError(
            reason=(
                "I can only read CSV or XLSX report files. Please re-export the "
                "report in one of these two formats and send it again."
            ),
            file_path=file_path,
            kind="extension",
        )

    if not SETTLEMENT_DETAIL_FILENAME_RE.match(file_name):
        raise ReportTypeError(
            reason=(
                "I can only analyze the Settlement Detail report. The file you "
                "provided does not look like a Settlement Detail report. Please "
                "download the Settlement Detail report from the merchant portal "
                "and send it again. (The filename should contain both "
                "`SETTLEMENT` and `DETAIL`, e.g. `SETTLEMENT_DETAIL_*.xlsx` or "
                "`Settlement_Detail_*.csv`.)"
            ),
            file_path=file_path,
            kind="filename",
        )

    try:
        header = _read_xlsx_header(file_path) if ext == ".xlsx" else _read_csv_header(file_path)
    except Exception as e:
        raise ReportTypeError(
            reason=f"Cannot read report header for type detection: {e}",
            file_path=file_path,
            kind="content",
        )

    header_set = {col for col in header if col}

    # Positive signal: must contain at least N of the canonical SD columns.
    core_matches = header_set & SETTLEMENT_DETAIL_CORE_COLUMNS
    if len(core_matches) < SETTLEMENT_DETAIL_MIN_CORE_MATCH:
        raise ReportTypeError(
            reason=(
                "I can only analyze the Settlement Detail report. The file's "
                "header does not match a Settlement Detail report structure "
                f"(only {len(core_matches)} of "
                f"{len(SETTLEMENT_DETAIL_CORE_COLUMNS)} core columns found; "
                f"at least {SETTLEMENT_DETAIL_MIN_CORE_MATCH} required). "
                "Please re-download the Settlement Detail report from the "
                "merchant portal without renaming."
            ),
            file_path=file_path,
            kind="content",
        )

    # Negative signal: must not contain any TX-only column.
    forbidden_hits = header_set & SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS
    if forbidden_hits:
        raise ReportTypeError(
            reason=(
                "This file looks like a Transaction Detail report, not a "
                "Settlement Detail report (it contains transaction-only "
                f"columns: {sorted(forbidden_hits)}). I can only analyze "
                "Settlement Detail reports. Please re-download the Settlement "
                "Detail report from the merchant portal."
            ),
            file_path=file_path,
            kind="content",
        )

    return {"report_type": "SETTLEMENT_DETAIL", "file_path": file_path}


def get_fee_summary(data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Return summary of all 17 fee fields at once, eliminating LLM's freedom to selectively choose fields."""
    fee_details = {}
    non_zero_fields = []
    total_fees = 0.0

    for field in ALL_FEE_FIELDS:
        field_sum = 0.0
        for row in data:
            val = row.get(field, "")
            if val and str(val).strip() not in ("", "null"):
                try:
                    field_sum += float(val)
                except (ValueError, TypeError):
                    pass
        fee_details[field] = round(field_sum, 2)
        total_fees += field_sum
        if abs(field_sum) > 0.001:
            non_zero_fields.append(field)

    return {
        "fee_details": fee_details,
        "total_fees": round(total_fees, 2),
        "non_zero_fields": non_zero_fields,
        "field_count": {
            "total": len(ALL_FEE_FIELDS),
            "non_zero": len(non_zero_fields),
            "zero_or_empty": len(ALL_FEE_FIELDS) - len(non_zero_fields),
        },
    }


# JSON Schema validation (optional dependency)
try:
    import jsonschema
    HAS_JSONSCHEMA = True
except ImportError:
    HAS_JSONSCHEMA = False


def get_dsl_schema() -> Dict[str, Any]:
    """Get DSL Schema definition."""
    schema_path = Path(__file__).parent / "dsl_schema.json"
    with open(schema_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def validate_dsl(filters: Dict[str, Any]) -> Optional[str]:
    """Validate DSL against Schema, return None or error message."""
    if not filters:
        return None
    
    if not HAS_JSONSCHEMA:
        # jsonschema not installed, skip validation
        return None
    
    try:
        schema = get_dsl_schema()
        jsonschema.validate(instance=filters, schema=schema)
        return None
    except jsonschema.ValidationError as e:
        return f"DSL validation failed: {e.message}"


class FilterDSL:
    """Filter DSL Executor"""
    
    def __init__(self, filters: Optional[Dict[str, Any]] = None):
        """Initialize DSL executor."""
        self.filters = filters or {}
    
    def _compare_values(self, actual: Any, op: str, expected: Any) -> bool:
        """Compare two values, supports =, !=, >, >=, <, <=, IN, NOT IN, LIKE, IS NULL, IS NOT NULL."""
        # Type conversion
        try:
            if isinstance(expected, (int, float)):
                actual = float(actual)
        except (ValueError, TypeError):
            pass
        
        if op == "=":
            return actual == expected
        elif op == "!=":
            return actual != expected
        elif op == ">":
            return float(actual) > float(expected)
        elif op == ">=":
            return float(actual) >= float(expected)
        elif op == "<":
            return float(actual) < float(expected)
        elif op == "<=":
            return float(actual) <= float(expected)
        elif op == "IN":
            return actual in expected
        elif op == "NOT IN":
            return actual not in expected
        elif op == "LIKE":
            # Support % wildcard, escape regex special chars first then restore % wildcard
            import re
            escaped = re.escape(expected)
            pattern = escaped.replace("%", ".*")
            return bool(re.match(f"^{pattern}$", str(actual)))
        elif op == "IS NULL":
            return actual is None or actual == "" or str(actual).lower() == "null"
        elif op == "IS NOT NULL":
            return not (actual is None or actual == "" or str(actual).lower() == "null")
        else:
            raise ValueError(f"Unsupported operator: {op}")
    
    def _evaluate_condition(self, row: Dict[str, Any], condition: Dict[str, Any]) -> bool:
        """Evaluate single condition (supports AND/OR nesting)."""
        # AND/OR nesting
        if "AND" in condition:
            return all(self._evaluate_condition(row, c) for c in condition["AND"])
        
        if "OR" in condition:
            return any(self._evaluate_condition(row, c) for c in condition["OR"])
        
        # Simple condition
        column = condition.get("column")
        op = condition.get("op")
        value = condition.get("value")
        
        if column is None or op is None:
            return True
        
        actual_value = row.get(column)
        return self._compare_values(actual_value, op, value)
    
    def apply_where(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply WHERE filtering."""
        where_clause = self.filters.get("WHERE")
        if not where_clause:
            return rows
        
        return [row for row in rows if self._evaluate_condition(row, where_clause)]
    
    def apply_select(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply SELECT column selection."""
        select_columns = self.filters.get("SELECT")
        if not select_columns:
            return rows
        
        return [
            {k: v for k, v in row.items() if k in select_columns}
            for row in rows
        ]
    
    def _apply_aggregation(self, rows: List[Dict[str, Any]], agg: Dict[str, Any]) -> Any:
        """Apply single aggregation function (COUNT/SUM/AVG/MIN/MAX/FIRST/LAST)."""
        column = agg.get("column")
        func = agg.get("function", "COUNT")
        
        values = [row.get(column) for row in rows]
        
        if func == "COUNT":
            return len([v for v in values if v is not None and v != ""])
        elif func == "SUM":
            try:
                return sum(float(v) for v in values if v is not None and v != "")
            except (ValueError, TypeError):
                return 0
        elif func == "AVG":
            try:
                numeric_values = [float(v) for v in values if v is not None and v != ""]
                return sum(numeric_values) / len(numeric_values) if numeric_values else 0
            except (ValueError, TypeError):
                return 0
        elif func == "MIN":
            try:
                numeric_values = [float(v) for v in values if v is not None and v != ""]
                return min(numeric_values) if numeric_values else None
            except (ValueError, TypeError):
                return None
        elif func == "MAX":
            try:
                numeric_values = [float(v) for v in values if v is not None and v != ""]
                return max(numeric_values) if numeric_values else None
            except (ValueError, TypeError):
                return None
        elif func == "FIRST":
            return values[0] if values else None
        elif func == "LAST":
            return values[-1] if values else None
        else:
            raise ValueError(f"Unsupported aggregation function: {func}")
    
    def apply_group_by(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply GROUP BY grouping aggregation."""
        group_by_clause = self.filters.get("GROUP_BY")
        if not group_by_clause:
            return rows
        
        group_columns = group_by_clause.get("columns", [])
        aggregations = group_by_clause.get("aggregations", [])
        
        if not group_columns:
            # No grouping columns, aggregate entire dataset
            result = {}
            for agg in aggregations:
                alias = agg.get("alias", agg.get("column"))
                result[alias] = self._apply_aggregation(rows, agg)
            return [result]
        
        # Grouping
        groups = {}
        for row in rows:
            key = tuple(row.get(col) for col in group_columns)
            if key not in groups:
                groups[key] = []
            groups[key].append(row)
        
        # Apply aggregation to each group
        results = []
        for key, group_rows in groups.items():
            result = {col: val for col, val in zip(group_columns, key)}
            for agg in aggregations:
                alias = agg.get("alias", agg.get("column"))
                result[alias] = self._apply_aggregation(group_rows, agg)
            results.append(result)
        
        return results
    
    def apply_order_by(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply ORDER BY sorting (supports multi-column)."""
        order_by_clause = self.filters.get("ORDER_BY")
        if not order_by_clause:
            return rows
        
        # Support multi-column sorting
        for order in reversed(order_by_clause):
            column = order.get("column")
            direction = order.get("direction", "ASC").upper()
            reverse = (direction == "DESC")
            
            try:
                rows.sort(
                    key=lambda x: float(x.get(column, 0) or 0),
                    reverse=reverse
                )
            except (ValueError, TypeError):
                # Non-numeric, sort as string
                rows.sort(
                    key=lambda x: str(x.get(column, "")),
                    reverse=reverse
                )
        
        return rows
    
    def apply_limit(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply LIMIT restriction."""
        limit = self.filters.get("LIMIT")
        if limit is None:
            return rows
        
        return rows[:limit]
    
    def execute(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Execute complete DSL (order: WHERE → GROUP BY → ORDER BY → LIMIT → SELECT)."""
        result = rows
        
        # 1. WHERE
        result = self.apply_where(result)
        
        # 2. GROUP BY
        result = self.apply_group_by(result)
        
        # 3. ORDER BY
        result = self.apply_order_by(result)
        
        # 4. LIMIT
        result = self.apply_limit(result)
        
        # 5. SELECT
        result = self.apply_select(result)
        
        return result


def detect_end_marker(row: Dict[str, Any]) -> bool:
    """Detect if it's an END marker row (characteristics: <END>, full row END, all non-empty fields are END)."""
    row_values = [str(v).strip().upper() for v in row.values() if v is not None]
    
    # Check if has <END> or END
    for val in row.values():
        if val is not None:
            val_str = str(val).strip()
            if "<END>" in val_str or val_str.upper() == "END":
                return True
    
    # Check if all non-empty fields are "END"
    non_empty = [v for v in row_values if v != ""]
    if non_empty and all(v == "END" for v in non_empty):
        return True
    
    return False


def parse_csv_file(file_path: str) -> Dict[str, Any]:
    """Parse single CSV file, return data + metadata (including empty batch detection)."""
    rows = []
    end_marker_rows = 0
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_dict = dict(row)
                # Detect if it's an END marker
                if detect_end_marker(row_dict):
                    end_marker_rows += 1
                else:
                    rows.append(row_dict)
    except UnicodeDecodeError:
        # Try other encodings
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row_dict = dict(row)
                    if detect_end_marker(row_dict):
                        end_marker_rows += 1
                    else:
                        rows.append(row_dict)
        except Exception as e:
            return {
                "success": False,
                "data": [],
                "metadata": {
                    "file_path": file_path,
                    "error": f"Encoding error: {str(e)}"
                }
            }
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "metadata": {
                "file_path": file_path,
                "error": f"Parsing failed: {str(e)}"
            }
        }
    
    # Calculate metadata
    total_rows = len(rows) + end_marker_rows
    is_empty_batch = (len(rows) == 0)
    
    metadata = {
        "file_path": file_path,
        "total_rows": total_rows,
        "data_rows": len(rows),
        "end_marker_rows": end_marker_rows,
        "is_empty_batch": is_empty_batch,
        "reason": None,
        "business_meaning": None
    }
    
    # Determine reason and business meaning for empty batch
    if is_empty_batch:
        if total_rows == 0:
            metadata["reason"] = "CSV file is completely empty (no header, no data)"
        elif total_rows == end_marker_rows:
            metadata["reason"] = "CSV contains only END marker, no data rows"
        else:
            metadata["reason"] = "CSV contains only header, no data rows"
        
        metadata["business_meaning"] = "Empty batch: Merchant's pending settlement balance at Antom has not reached the agreed settlement threshold, so there is no actual settlement content today"
    
    return {
        "success": True,
        "data": rows,
        "metadata": metadata
    }


def parse_xlsx_file(file_path: str) -> Dict[str, Any]:
    """Parse single XLSX file, return data + metadata (including empty batch detection)."""
    from openpyxl import load_workbook
    
    rows = []
    
    try:
        # Fix: cannot use read_only=True, otherwise header reading may be incomplete for some files
        wb = load_workbook(filename=file_path, read_only=False)
        ws = wb.active
        
        # Get header
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value) if cell.value is not None else "")
        
        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    value = cell.value
                    # Convert types to strings (consistent with CSV)
                    if value is None:
                        value = ""
                    elif isinstance(value, (int, float)):
                        value = str(value)
                    else:
                        value = str(value)
                    row_dict[headers[col_idx]] = value
            
            # Detect END marker
            if not detect_end_marker(row_dict):
                rows.append(row_dict)
        
        wb.close()
    
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "metadata": {
                "file_path": file_path,
                "error": f"XLSX parsing failed: {str(e)}"
            }
        }
    
    # Calculate metadata
    is_empty_batch = (len(rows) == 0)
    
    metadata = {
        "file_path": file_path,
        "total_rows": len(rows),
        "data_rows": len(rows),
        "end_marker_rows": 0,
        "is_empty_batch": is_empty_batch,
        "reason": "No data rows in XLSX file" if is_empty_batch else None,
        "business_meaning": "Empty batch: Merchant's pending settlement balance at Antom has not reached the agreed settlement threshold, so there is no actual settlement content" if is_empty_batch else None
    }
    
    return {
        "success": True,
        "data": rows,
        "metadata": metadata
    }


def parse_reports(
    input: Union[List[str], str],
    filters: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """Main reconciliation report parsing function: parse CSV/XLSX, apply DSL filtering, return data + fee_summary + metadata."""
    # Process input parameter
    if isinstance(input, str):
        input = [input]
    
    if not isinstance(input, list):
        raise ValueError("input must be a list of strings or a single string")

    # ── Report type gate (filename + lightweight content sanity) ──
    # Must run BEFORE any parsing. Fails the whole call on the first invalid
    # file — Agent then surfaces ReportTypeError.reason to the merchant.
    for file_path in input:
        detect_report_type(file_path)

    # Parse all CSV files
    all_rows = []
    parsed_files = []  # Successfully parsed file paths
    parse_errors = []  # Files with parsing errors
    
    file_metadata = {
        "file_count": len(input),
        "success_files": 0,
        "failed_files": 0,
        "total_rows": 0,
        "data_rows": 0,
        "end_marker_rows": 0,
        "empty_batch_files": [],
        "parse_errors": []
    }
    
    for file_path in input:
        # Select parser based on file extension
        if file_path.lower().endswith('.xlsx'):
            result = parse_xlsx_file(file_path)
        elif file_path.lower().endswith('.csv'):
            result = parse_csv_file(file_path)
        else:
            # Default to CSV
            result = parse_csv_file(file_path)
        
        if not result.get("success"):
            # ❌ Parsing failed: record but continue parsing other files
            parse_errors.append({
                "file_path": file_path,
                "error": result.get("metadata", {}).get("error", "Unknown parsing error")
            })
            file_metadata["failed_files"] += 1
            continue
        
        # ✅ Success: accumulate data rows
        parsed_files.append(file_path)
        all_rows.extend(result.get("data", []))
        
        # Accumulate metadata
        meta = result.get("metadata", {})
        file_metadata["total_rows"] += meta.get("total_rows", 0)
        file_metadata["data_rows"] += meta.get("data_rows", 0)
        file_metadata["end_marker_rows"] += meta.get("end_marker_rows", 0)
        
        # Record empty batch files
        if meta.get("is_empty_batch"):
            file_metadata["empty_batch_files"].append(file_path)
    
    # Update successful file count
    file_metadata["success_files"] = len(parsed_files)
    file_metadata["parse_errors"] = parse_errors
    
    # Determine if all are empty batches (only consider successfully parsed files)
    is_all_empty = (len(all_rows) == 0) and (len(file_metadata["empty_batch_files"]) == file_metadata["success_files"])
    
    # Validate filters before executing DSL
    if filters:
        error_msg = validate_dsl(filters)
        if error_msg:
            return {
                "success": False,
                "partial_success": len(parsed_files) > 0,
                "data": [],
                "metadata": {
                    **file_metadata,
                    "error": error_msg
                }
            }
    
    # Apply DSL filtering (only for non-empty batches)
    if filters and all_rows:
        dsl = FilterDSL(filters)
        all_rows = dsl.execute(all_rows)
    
    # Build return result
    partial_success = len(parse_errors) > 0 and len(parsed_files) > 0
    has_success = len(parsed_files) > 0
    
    # Auto-calculate fee summary (full coverage of 17 fee fields)
    fee_summary = get_fee_summary(all_rows) if all_rows else None
    
    result = {
        "success": has_success,
        "partial_success": partial_success,
        "data": all_rows,
        "fee_summary": fee_summary,
        "metadata": file_metadata
    }
    
    # If empty batch, add business meaning
    if is_all_empty and has_success:
        result["metadata"]["is_empty_batch"] = True
        result["metadata"]["business_meaning"] = "Empty batch: Merchant's pending settlement balance at Antom has not reached the agreed settlement threshold, so there is no actual settlement content. This is not an error, but a normal business state."
    
    return result


