#!/usr/bin/env python3
"""V3.3.1: 代码硬控全面升级 + is_smoke统一 + 动态门禁"""
"""
轻量 Excel 生成脚本 — 将 P6 输出 JSON 转换为公司标准 19 列测试用例 Excel。

依赖: openpyxl
用法: python3 export_excel.py --input p6_output.json --output testcases.xlsx

退出码:
  0 = 成功
  1 = 输入文件不存在
  2 = JSON 解析失败
"""

import argparse
import json
import sys
import hashlib
import hmac as _hmac_mod
import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 缺少 openpyxl，请执行 pip install openpyxl", file=sys.stderr)
    sys.exit(3)

# ── 公司标准 19 列定义（V4.x: 字段统一）──
# P6 JSON 字段 → Excel 列名
COLUMNS = [
    ("project",          "项目"),
    ("case_type",        "类型"),
    ("case_id",          "用例编号"),
    ("requirement",      "需求"),  # V4.x: 实际取 case["project"] 的值
    ("priority",         "优先级"),
    ("title",            "标题"),
    ("menu_path",        "用例菜单"),
    ("preconditions",    "预置条件"),
    ("steps",            "步骤"),
    ("expected_results", "期望结果"),
    ("is_smoke",         "是否冒烟用例"),
    ("creator",          "创建者"),  # V4.x: 固定返回 "AI"
    ("assignee",         "经办人"),
    ("test_case_type",   "用例类型"),
    ("test_category",    "测试类别"),
    ("status",           "执行结果"),
    ("screenshot",       "截图"),
    ("test_suite",       "测试用例集"),
    ("remarks",          "备注"),
]

# 优先级映射: P6 输出 → 公司字典(Highest/High/Medium/Low/Lowest)
PRIORITY_MAP = {
    "P0": "Highest", "core": "Highest",
    "P1": "High",   "high": "High",
    "P2": "Medium", "medium": "Medium",
    "P3": "Low",    "low": "Low",
    # 兼容中文输入
    "核心": "Highest",
    "高": "High",
    "中": "Medium",
    "低": "Low",
}

# 用例类型映射: P6 输出 → 公司字典(正例/反例)
TEST_CASE_TYPE_MAP = {
    "main_flow": "正例", "正向": "正例",
    "branch": "正例",
    "integration": "正例",
    "正向验证": "正例",
    "分支验证": "正例",
    "逻辑组合": "正例",
    "状态迁移": "正例",
    "兼容回归": "正例",
    "接口验证": "正例",
    "条件验证": "正例",
    "异常处理": "反例",
    "边界验证": "反例",
    "boundary": "反例",
    "exception": "反例",
    "集成异常": "反例",
    "歧义验证": "反例",
    "权限验证": "反例",
    "性能验证": "反例",
    "安全验证": "反例",
}

# 测试类别映射: P6 输出 → 公司字典
TEST_CATEGORY_MAP = {
    "功能": "功能测试", "功能测试": "功能测试", "main_flow": "功能测试",
    "分支": "功能测试", "branch": "功能测试",
    "边界": "功能测试", "boundary": "功能测试",
    "异常": "功能测试", "exception": "功能测试",
    "集成": "功能测试", "integration": "功能测试",
    "性能": "性能测试", "性能测试": "性能测试",
    "安全": "安全性测试",
    "兼容性": "兼容性测试",
    "可靠性": "可靠性测试",
    "可移植性": "可移植性测试",
    "可维护性": "可维护性测试",
    # 直接匹配公司字典值
    "功能测试": "功能测试",
    "性能测试": "性能测试",
    "安全性测试": "安全性测试",
    "兼容性测试": "兼容性测试",
    "可靠性测试": "可靠性测试",
    "可移植性测试": "可移植性测试",
    "可维护性测试": "可维护性测试",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _get_case_field(case, field, default=""):
    """V4.x: 从P6用例中读取字段值，兼容两种结构：
    1. 扁平结构：case[field]
    2. 嵌套结构：case["fields"][field]
    
    特殊字段处理：
    - creator: 固定返回 "AI"
    - requirement: 取 case["project"] 的值（而非 case["requirement"]，因为后者存的是task_id）
    """
    # V4.x: creator 固定返回 "AI自动生成"
    if field == "creator":
        return "AI自动生成"
    # V4.x: requirement 字段特殊处理，取 project 的值
    if field == "requirement":
        val = case.get("project")
        if val is not None and val != "":
            return val
        val = case.get("requirement")
        if val is not None and val != "":
            return val
        fields = case.get("fields")
        if isinstance(fields, dict):
            val = fields.get("project") or fields.get("requirement", "")
            if val is not None and val != "":
                return val
        return default
    # V4.8.12: project 也尝试从 requirement 取值
    if field == "project":
        val = case.get("project")
        if val is not None and val != "":
            return val
        val = case.get("requirement")
        if val is not None and val != "":
            return val
        return default
    val = case.get(field)
    if val is not None and val != "":
        return val
    fields = case.get("fields")
    if isinstance(fields, dict):
        val = fields.get(field)
        if val is not None:
            return val
    return default


def _is_smoke_check(value):
    """V3.2.8: 统一is_smoke判断逻辑，与orchestrator._is_smoke()完全一致。
    支持: True, "true", "是", "Y", "yes", 1
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    if isinstance(value, str):
        return value.lower() in ("true", "是", "y", "yes", "1")
    return False


def normalize_value(key: str, raw) -> str:
    """将 P6 JSON 值转换为 Excel 友好字符串。"""
    try:
        if raw is None:
            return ""
        if key == "priority":
            return PRIORITY_MAP.get(str(raw), str(raw))
        if key == "test_case_type" or key == "case_type":
            return TEST_CASE_TYPE_MAP.get(str(raw), str(raw))
        if key == "test_category":
            return TEST_CATEGORY_MAP.get(str(raw), str(raw))
        if key == "is_smoke":
            # V3.2.8: 统一布尔值判断，兼容所有格式
            if isinstance(raw, bool):
                return "是" if raw else "否"
            if isinstance(raw, (int, float)):
                return "是" if raw == 1 else "否"
            if isinstance(raw, str):
                return "是" if raw.lower() in ("true", "是", "y", "yes", "1") else "否"
            return "否"
        if key == "steps" and isinstance(raw, list):
            return "\n".join(f"{i+1}. {s}" for i, s in enumerate(raw))
        if key == "expected_results" and isinstance(raw, list):
            return "\n".join(f"{i+1}. {r}" for i, r in enumerate(raw))
        if key == "remarks":
            # V4.8.12: 清理内部质量标记，保留评审工具用的模块分类
            s = str(raw).strip() if raw else ""
            import re as _re_rem2
            s = _re_rem2.sub(r'\[术语不一致[^\]]*\]', '', s)
            s = _re_rem2.sub(r'\[GATE_ISSUE:[^\]]*\]', '', s)
            s = _re_rem2.sub(r'\[占位符[^\]]*\]', '', s)
            s = _re_rem2.sub(r'\[步骤[^\]]*\]', '', s)
            s = _re_rem2.sub(r'\[期望[^\]]*\]', '', s)
            return ' '.join(s.split())
        if isinstance(raw, (list, dict)):
            return json.dumps(raw, ensure_ascii=False, indent=None)
        return str(raw)
    except Exception:
        return str(raw)[:500]  # 兜底：截断过长或异常内容


def build_workbook(cases: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 写表头
    for col_idx, (_, cn_name) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cn_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写数据行（V3.2.4: 使用_get_case_field兼容fields嵌套结构）
    for row_idx, case in enumerate(cases, start=2):
        for col_idx, (field, _) in enumerate(COLUMNS, start=1):
            val = normalize_value(field, _get_case_field(case, field, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动列宽（简易估算）
    for col_idx, (_, cn_name) in enumerate(COLUMNS, start=1):
        max_len = len(cn_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value).split("\n")[0]), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    return wb


def main():
    parser = argparse.ArgumentParser(description="P6 JSON → 公司标准 19 列 Excel")
    parser.add_argument("--input", required=True, help="P6 输出 JSON 文件路径")
    parser.add_argument("--output", required=True, help="输出 Excel 文件路径")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 输入文件不存在 → {input_path}", file=sys.stderr)
        sys.exit(1)

    # === V3.2.6: gate pass前置检查 + HMAC验签（即使Agent直接调用也无法绕过） ===
    data_dir = input_path.parent
    gate_dir = data_dir / "gates"
    required_gates = ["onboarding", "P0", "P1", "P2", "P3", "P4", "P5", "P6", "P7"]
    missing_gates = []

    # V4.6.14: 使用固定HMAC_SECRET（与orchestrator._get_hmac_key一致）
    HMAC_SECRET = "xy-req2testcase-v4-hmac-2026"
    hmac_key = HMAC_SECRET.encode("utf-8")
    # V4.0.1兼容: 旧版基于文件哈希的密钥，用于验签兼容
    orch_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "orchestrator.py")
    legacy_hmac_key = None
    if os.path.exists(orch_path):
        with open(orch_path, "rb") as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
        legacy_hmac_key = f"orch-gate-{file_hash[:32]}".encode("utf-8")

    # 从 task_meta 获取 task_id
    task_id = None
    task_meta_path = data_dir / "task_meta.json"
    if task_meta_path.exists():
        try:
            with open(task_meta_path, "r", encoding="utf-8") as f:
                task_meta = json.load(f)
            task_id = task_meta.get("task_id", "")
        except Exception:
            pass

    for step in required_gates:
        gate_file = gate_dir / f"{step}.pass.json"
        if not gate_file.exists():
            missing_gates.append(f"{step}(文件缺失)")
            continue
        try:
            with open(gate_file, "r", encoding="utf-8") as f:
                gp = json.load(f)
            # 校验task_id
            if task_id and gp.get("task_id") != task_id:
                missing_gates.append(f"{step}(task_id不匹配)")
                continue
            # V4.6.14: HMAC验签（先新密钥→旧密钥兼容）
            if hmac_key:
                stored_hmac = gp.get("hmac")
                if not stored_hmac:
                    missing_gates.append(f"{step}(缺少HMAC签名)")
                    continue
                data_copy = {k: v for k, v in gp.items() if k != "hmac"}
                data_copy["_task_id"] = gp.get("task_id", "")
                payload = json.dumps(data_copy, sort_keys=True, ensure_ascii=False)
                # V4.6.14: 先用新固定密钥验签
                expected = _hmac_mod.new(hmac_key, payload.encode("utf-8"), hashlib.sha256).hexdigest()
                hmac_ok = _hmac_mod.compare_digest(stored_hmac, expected)
                # 新密钥失败时，尝试旧密钥（文件哈希派生，兼容V4.0.1之前的gate）
                if not hmac_ok and legacy_hmac_key:
                    expected_legacy = _hmac_mod.new(legacy_hmac_key, payload.encode("utf-8"), hashlib.sha256).hexdigest()
                    hmac_ok = _hmac_mod.compare_digest(stored_hmac, expected_legacy)
                if not hmac_ok:
                    missing_gates.append(f"{step}(HMAC签名无效)")
                    continue
        except Exception:
            missing_gates.append(f"{step}(读取失败)")
    
    if missing_gates:
        print(f"❌ 导出被拒绝: 以下步骤的gate pass缺失: {missing_gates}", file=sys.stderr)
        print(f"必须通过orchestrator完成全部流程后才能导出。", file=sys.stderr)
        sys.exit(10)

    # === V3.2.8: P6质量硬门禁（与orchestrator quality_check/step7_export完全一致） ===
    try:
        with open(input_path, "r", encoding="utf-8") as f:
            p6_check = json.load(f)
        check_cases = p6_check.get("testcases", []) if isinstance(p6_check, dict) else p6_check if isinstance(p6_check, list) else []
        total = len(check_cases)
        if total > 0:
            smoke = sum(1 for c in check_cases if _is_smoke_check(_get_case_field(c, "is_smoke", "")))
            p0 = sum(1 for c in check_cases if _get_case_field(c, "priority", "").upper() in ("P0", "HIGHEST"))
            quality_issues = []
            quality_warnings = []  # V4.12.2: 分级处理，WARNING不阻断
            # V3.2.8: 动态最低用例数（与orchestrator同步）
            p6_min = 15
            p5_path = data_dir / "p5_output.json"
            if p5_path.exists():
                try:
                    with open(p5_path, "r", encoding="utf-8") as f:
                        p5_data = json.load(f)
                    total_expected = p5_data.get("coverage_summary", {}).get("total_expected_cases", 0)
                    p5_count = len(p5_data.get("test_points", []))
                    if total_expected > 0:
                        p6_min = max(15, total_expected)
                    elif p5_count > 0:
                        p6_min = max(15, int(p5_count * 1.5))
                except Exception:
                    pass
            # V4.12.2: 用例数不足分级处理（P5预估值可能不准确）
            if total < p6_min:
                ratio = total / p6_min if p6_min > 0 else 1.0
                if ratio < 0.5:
                    quality_issues.append(f"用例数{total}<{p6_min}(占比{ratio:.0%}, <50%异常)")
                else:
                    quality_warnings.append(f"⚠️ 用例数{total}<P5预估{p6_min}(占比{ratio:.0%}),P7已通过则忽略")
            if smoke / total > 0.20:
                quality_issues.append(f"冒烟{smoke/total:.0%}>20%")
            if smoke / total < 0.05:
                quality_issues.append(f"冒烟{smoke/total:.0%}<5%")
            if p0 / total > 0.20:
                quality_issues.append(f"P0占{p0/total:.0%}>20%")
            if quality_warnings:
                for w in quality_warnings:
                    print(w, file=sys.stderr)
            if quality_issues:
                print(f"❌ 导出被拒绝: P6质量不达标: {quality_issues}", file=sys.stderr)
                sys.exit(11)
    except Exception:
        pass  # 质量检查失败不阻塞导出（gate检查已是主防线）

    try:
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        print(f"错误: JSON 解析失败 → {e}", file=sys.stderr)
        sys.exit(2)

    # 兼容两种格式: 直接数组 或 {"testcases": [...]}
    if isinstance(data, dict):
        cases = data.get("testcases", data.get("test_cases", []))
    elif isinstance(data, list):
        cases = data
    else:
        print("错误: JSON 顶层结构不是数组或对象", file=sys.stderr)
        sys.exit(2)

    if not cases:
        print("⚠️ 警告: 用例列表为空，将生成空 Excel", file=sys.stderr)

    # V4.8.12: 从 task_meta.json 自动填充 project/requirement/case_type
    if task_meta_path.exists():
        try:
            with open(task_meta_path, "r", encoding="utf-8") as f:
                tm = json.load(f)
            proj_name = tm.get("project", tm.get("project_name", ""))
            req_id = tm.get("task_id", "")
            if proj_name or req_id:
                for c in cases:
                    if isinstance(c, dict):
                        if not c.get("project") and proj_name:
                            c["project"] = proj_name
                        if not c.get("requirement") and req_id:
                            c["requirement"] = req_id
                        if not c.get("case_type"):
                            c["case_type"] = "测试用例"
        except Exception:
            pass

    wb = build_workbook(cases)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(output_path))
    except Exception as e:
        print(f"错误: Excel 写入失败 → {e}", file=sys.stderr)
        sys.exit(4)
    print(f"✅ 已生成 {len(cases)} 条用例 → {output_path}")


if __name__ == "__main__":
    main()
