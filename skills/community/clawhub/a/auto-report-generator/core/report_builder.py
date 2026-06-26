"""报表构建模块 - 生成多sheet Excel报表"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path
from datetime import datetime


def build_excel_report(data: dict, output_path: str) -> str:
    """构建多sheet Excel报表

    Args:
        data: 报表数据字典，包含以下键：
            - title: 报表标题
            - stats: 统计数据
            - charts: 图表路径列表
            - ai_analysis: AI分析文本
            - raw_data: 原始数据 DataFrame
            - template: 模板名称
        output_path: 输出文件路径

    Returns:
        str: 生成的报表文件路径
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_cover_sheet(wb, data)
    _build_overview_sheet(wb, data)
    _build_charts_sheet(wb, data)
    _build_ai_analysis_sheet(wb, data)
    _build_raw_data_sheet(wb, data)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return str(output_path)


def _build_cover_sheet(wb: openpyxl.Workbook, data: dict):
    """构建封面sheet"""
    ws = wb.create_sheet("封面", 0)

    title = data.get('title', '自动报表')
    template = data.get('template', '默认模板')

    title_font = Font(name='微软雅黑', size=24, bold=True, color='FFFFFF')
    subtitle_font = Font(name='微软雅黑', size=14, color='CCCCCC')

    fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    for row in range(1, 10):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill

    ws.merge_cells('A2:G7')
    title_cell = ws.cell(row=4, column=1)
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A8:G8')
    date_cell = ws.cell(row=8, column=1)
    date_cell.value = f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = subtitle_font
    date_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A9:G9')
    template_cell = ws.cell(row=9, column=1)
    template_cell.value = f"模板: {template}"
    template_cell.font = subtitle_font
    template_cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['G'].width = 3


def _build_overview_sheet(wb: openpyxl.Workbook, data: dict):
    """构建数据概览sheet"""
    ws = wb.create_sheet("数据概览", 1)

    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell_font = Font(name='微软雅ji', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    stats = data.get('stats', {})
    if isinstance(stats, dict):
        overview_data = [
            ["指标", "数值"],
            ["数据行数", stats.get('row_count', 'N/A')],
            ["数据列数", stats.get('col_count', 'N/A')],
            ["缺失值总数", sum(stats.get('missing_values', {}).values()) if isinstance(stats.get('missing_values'), dict) else 'N/A'],
        ]

        for i, row_data in enumerate(overview_data, start=1):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.font = header_font if i == 1 else cell_font
                cell.fill = header_fill if i == 1 else PatternFill(fill_type=None)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20


def _build_charts_sheet(wb: openpyxl.Workbook, data: dict):
    """构建图表sheet"""
    ws = wb.create_sheet("图表分析", 2)

    charts = data.get('charts', [])
    if charts:
        for i, chart_path in enumerate(charts[:5], start=1):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(chart_path)
            img.width = 400
            img.height = 240
            ws.add_image(img, f'A{i*10}')
    else:
        ws.cell(row=1, column=1, value="暂无图表数据")

    ws.column_dimensions['A'].width = 30


def _build_ai_analysis_sheet(wb: openpyxl.Workbook, data: dict):
    """构建AI分析sheet"""
    ws = wb.create_sheet("AI分析", 3)

    ai_analysis = data.get('ai_analysis', '暂无AI分析结果')

    header_font = Font(name='微软雅黑', size=12, bold=True)
    content_font = Font(name='微软雅黑', size=11)

    ws.cell(row=1, column=1, value="AI 数据分析报告").font = header_font
    ws.merge_cells('A1:D1')

    ws.cell(row=3, column=1, value="分析内容:").font = header_font
    ws.merge_cells('A3:D3')

    ws.cell(row=4, column=1, value=ai_analysis).font = content_font
    ws.merge_cells('A4:D10')
    ws.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 80


def _build_raw_data_sheet(wb: openpyxl.Workbook, data: dict):
    """构建原始数据sheet"""
    ws = wb.create_sheet("原始数据", 4)

    raw_data = data.get('raw_data')
    if raw_data is not None and isinstance(raw_data, pd.DataFrame):
        df = raw_data

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
