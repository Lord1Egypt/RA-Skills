#!/usr/bin/env python3
"""Convert a CSV file into a formatted Excel report using openpyxl."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NUMERIC_RE = re.compile(r"^[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^[-+]?\d+(?:\.\d+)?$")


def parse_number(value: Any) -> float | int | None:
    """Return an int/float for numeric-looking CSV values; otherwise None."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    normalized = text.replace(",", "")
    if not NUMERIC_RE.match(text):
        return None
    try:
        number = float(normalized)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def unique_headers(headers: list[str], width: int) -> list[str]:
    """Ensure headers are non-empty and unique for Excel tables."""
    result: list[str] = []
    seen: dict[str, int] = {}
    for idx in range(width):
        base = (headers[idx].strip() if idx < len(headers) else "") or f"Column {idx + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def read_csv(csv_path: Path, encoding: str = "utf-8-sig") -> tuple[list[str], list[list[str]]]:
    with csv_path.open("r", newline="", encoding=encoding) as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty")
    width = max(len(row) for row in rows)
    headers = unique_headers(rows[0], width)
    data = [row + [""] * (width - len(row)) for row in rows[1:]]
    return headers, data


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def build_report(csv_path: Path, output_path: Path, sheet_name: str = "Report") -> Path:
    headers, data = read_csv(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    numeric_flags = [False] * len(headers)
    numeric_sums = [0.0] * len(headers)

    for row in data:
        out_row = []
        for idx, value in enumerate(row):
            parsed = parse_number(value)
            if parsed is None:
                out_row.append(value)
            else:
                out_row.append(parsed)
                numeric_flags[idx] = True
                numeric_sums[idx] += float(parsed)
        ws.append(out_row)

    for col_idx, is_numeric in enumerate(numeric_flags, start=1):
        if is_numeric:
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell[0].number_format = "#,##0.00" if isinstance(cell[0].value, float) and not float(cell[0].value).is_integer() else "#,##0"

    total_row_idx = ws.max_row + 1
    if any(numeric_flags):
        ws.cell(row=total_row_idx, column=1, value="Total")
        ws.cell(row=total_row_idx, column=1).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=1).fill = PatternFill("solid", fgColor="D9EAF7")
        for col_idx, is_numeric in enumerate(numeric_flags, start=1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            if is_numeric:
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                cell.number_format = "#,##0.00" if any(isinstance(ws.cell(row=r, column=col_idx).value, float) and not float(ws.cell(row=r, column=col_idx).value).is_integer() for r in range(2, total_row_idx)) else "#,##0"
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    else:
        total_row_idx = ws.max_row

    if ws.max_row >= 2 and ws.max_column >= 1:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{total_row_idx}"
        table = Table(displayName="ReportTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    first_numeric_col = next((idx + 1 for idx, flag in enumerate(numeric_flags) if flag), None)
    if first_numeric_col and len(data) >= 1:
        chart = BarChart()
        chart.title = f"{headers[first_numeric_col - 1]} by {headers[0]}"
        chart.y_axis.title = headers[first_numeric_col - 1]
        chart.x_axis.title = headers[0]
        max_data_row = min(11, 1 + len(data))  # Header + first 10 data rows
        values = Reference(ws, min_col=first_numeric_col, min_row=1, max_row=max_data_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=max_data_row)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(total_row_idx, 1)}"
    autosize_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert CSV to a formatted Excel report (.xlsx).")
    parser.add_argument("csv_path", type=Path, help="Input CSV file path")
    parser.add_argument("output_path", type=Path, nargs="?", help="Output .xlsx path; defaults beside CSV")
    parser.add_argument("--sheet-name", default="Report", help="Worksheet name, max 31 chars")
    args = parser.parse_args()

    if not args.csv_path.exists():
        print(f"CSV not found: {args.csv_path}", file=sys.stderr)
        return 2
    output_path = args.output_path or args.csv_path.with_suffix(".xlsx")
    try:
        result = build_report(args.csv_path, output_path, args.sheet_name)
    except Exception as exc:  # keep CLI useful to agents/users
        print(f"Failed to create report: {exc}", file=sys.stderr)
        return 1
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
