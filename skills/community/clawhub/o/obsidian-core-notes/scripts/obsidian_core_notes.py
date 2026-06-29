#!/usr/bin/env python3
"""Maintain Obsidian core-file sidecar notes and folder indexes."""

from __future__ import annotations

import argparse
import ast
import hashlib
import json
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None


FILE_MARKER = "<!-- CORE_FILE_NOTE_V1 -->"
FOLDER_MARKER = "<!-- CORE_FOLDER_INDEX_V1 -->"
ROOT_MARKER = "<!-- CORE_WORKSPACE_INDEX_V1 -->"
INDEX_NAME = "\u8d44\u6599\u7d22\u5f15.md"
ROOT_INDEX_NAME = "\u6838\u5fc3\u6587\u4ef6\u7d22\u5f15.md"
TOPIC_NOTE_NAME = "\u4e13\u9898\u7efc\u5408.core.md"
TOPIC_SYNTHESIS_SCHEMA = "topic-synthesis-v3-like"
INDEX_ALIAS = "\u8d44\u6599\u7d22\u5f15"
INDEX_TITLE_PREFIX = "\u8d44\u6599\u7d22\u5f15\uff1a"
ROOT_TITLE = "\u6838\u5fc3\u6587\u4ef6\u7d22\u5f15"
WORKSPACE_LABEL = "\u5de5\u4f5c\u533a"

BUSINESS_EXTS = {
    ".md",
    ".docx",
    ".doc",
    ".pptx",
    ".ppt",
    ".xlsx",
    ".xls",
    ".pdf",
    ".txt",
    ".csv",
    ".json",
    ".yaml",
    ".yml",
}
CODE_EXTS = {
    ".md",
    ".py",
    ".js",
    ".ts",
    ".svelte",
    ".html",
    ".css",
    ".json",
    ".yaml",
    ".yml",
    ".toml",
    ".txt",
    ".cmd",
    ".ps1",
    ".sh",
}
MANIFEST_NAMES = {
    "pyproject.toml",
    "package.json",
    "requirements.txt",
    "project.config.json",
    "manifest.json",
    "AGENTS.md",
    "README.md",
    "config.yaml",
}
KEY_NAMES = {
    "readme.md",
    "readme.txt",
    "requirements.txt",
    "pyproject.toml",
    "package.json",
    "project.config.json",
    "config.yaml",
    "manifest.json",
    "agents.md",
    "changelog.md",
    "roadmap.md",
}
CODE_DIRS = {
    "src",
    "scripts",
    "tests",
    "config",
    "configs",
    "templates",
    "agents",
    "cloudfunctions",
    "opendata",
    "js",
    "rules",
    "commands",
    "skills",
}
EXCLUDED_PARTS = {
    ".git",
    "node_modules",
    ".venv",
    "venv",
    "env",
    "__pycache__",
    ".pytest_cache",
    ".mypy_cache",
    ".ruff_cache",
    ".trash",
    "dist",
    "build",
    ".next",
    ".nuxt",
    ".cache",
    ".obsidian",
    ".claude",
    ".claudian",
}
EXCLUDED_PATH_PATTERNS = [
    r"/node_modules/",
    r"/__pycache__/",
    r"/autogluon_result/",
    r"/ag_exp_",
    r"/models/",
    r"/nginx-[^/]+/",
    r"/res/uniview/",
    r"/00775802801001040/res/",
    r"/Y17C_C1_ULWL_PP-inst-G2/",
    r"/project_output/",
    r"/generated_workspace/",
]

MAX_TEXT_BYTES = 1_000_000
MAX_JSON_BYTES = 2_000_000
MAX_PREVIEW_CHARS = 1200
TOPIC_SAMPLE_BYTES = 200_000
TOPIC_CANDIDATE_KEYWORDS = [
    "AI",
    "AIGC",
    "Agent",
    "ChatBI",
    "MCP",
    "Skill",
    "Codex",
    "OpenClaw",
    "GEO",
    "harness",
    "POC",
    "\u4e2d\u79d1\u4e91\u8c37",  # 中科云谷
    "\u8f6c\u578b",  # 转型
    "\u7b97\u6cd5",  # 算法
    "\u98ce\u63a7",  # 风控
    "\u5546\u673a",  # 商机
    "\u6570\u636e\u8d28\u91cf",  # 数据质量
    "\u6570\u636e\u6cbb\u7406",  # 数据治理
    "\u8425\u8fd0",  # 营运
    "\u8425\u9500",  # 营销
    "\u4f1a\u8bae\u7eaa\u8981",  # 会议纪要
    "\u65b9\u6848",  # 方案
    "\u67b6\u6784",  # 架构
    "\u6a21\u578b",  # 模型
    "\u8c03\u7814",  # 调研
    "\u89c4\u5212",  # 规划
    "\u7ba1\u7406",  # 管理
    "\u5e73\u53f0",  # 平台
    "\u7ec4\u4ef6",  # 组件
    "\u667a\u80fd\u4f53",  # 智能体
]
TOPIC_VERSION_PATTERN = re.compile(
    r"(?i)(v\d+(?:\.\d+)?|20\d{6}|20\d{2}[-_.]?\d{1,2}[-_.]?\d{1,2}|\u5b9a\u7a3f|\u4fee\u8ba2|\u4fee\u6539|\u5468\u62a5|\u8fdb\u5c55|\u6c47\u62a5)"
)
TOPIC_MAIN_DOC_PATTERN = re.compile(
    r"(?i)(V\d+|\u5b9a\u7a3f|\u65b9\u6848|\u67b6\u6784|\u62a5\u544a|\u8c03\u7814|\u4f1a\u8bae\u7eaa\u8981|\u89c4\u5212|\u6c47\u62a5|\u5468\u62a5)"
)


@dataclass
class CoreRecord:
    path: Path
    root: Path
    kind: str
    note_path: Path

    @property
    def rel(self) -> str:
        return self.path.relative_to(self.root).as_posix()

    @property
    def note_rel(self) -> str:
        return self.note_path.relative_to(self.root).as_posix()

    @property
    def ext(self) -> str:
        return self.path.suffix.lower()

    @property
    def norm(self) -> str:
        return normalize_name(self.path)


def long_path(path: Path) -> str:
    resolved = str(path.resolve())
    if os.name == "nt" and not resolved.startswith("\\\\?\\"):
        return "\\\\?\\" + resolved
    return resolved


def read_text(path: Path) -> str:
    with open(long_path(path), "r", encoding="utf-8") as handle:
        return handle.read()


def write_text(path: Path, text: str) -> None:
    with open(long_path(path), "w", encoding="utf-8", newline="\n") as handle:
        handle.write(text)


def rel(root: Path, path: Path) -> str:
    return path.relative_to(root).as_posix()


def md_escape(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def yaml_quote(value: object) -> str:
    return json.dumps(str(value), ensure_ascii=False)


def strip_md_suffix(root: Path, path: Path) -> str:
    value = rel(root, path)
    return value[:-3] if value.lower().endswith(".md") else value


def wikilink(root: Path, path: Path, alias: str | None = None, attachment: bool = False) -> str:
    target = rel(root, path) if attachment else strip_md_suffix(root, path)
    return f"[[{target}|{alias}]]" if alias else f"[[{target}]]"


def fence(text: str, lang: str = "text") -> str:
    safe = (text or "").replace("```", "` ` `").rstrip()
    return f"```{lang}\n{safe}\n```"


def format_size(size: int) -> str:
    value = float(size)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{int(value)} {unit}" if unit == "B" else f"{value:.2f} {unit}"
        value /= 1024
    return f"{size} B"


def safe_generated_path(folder: Path, desired_name: str, marker: str) -> Path:
    path = folder / desired_name
    if not path.exists():
        return path
    try:
        if marker in read_text(path)[:300] or ROOT_MARKER in read_text(path)[:300]:
            return path
    except Exception:
        pass
    stem = Path(desired_name).stem
    suffix = Path(desired_name).suffix
    fallback = folder / f"{stem}.generated{suffix}"
    if not fallback.exists():
        return fallback
    try:
        if marker in read_text(fallback)[:300] or ROOT_MARKER in read_text(fallback)[:300]:
            return fallback
    except Exception:
        pass
    digest = hashlib.sha1(str(path).encode("utf-8")).hexdigest()[:8]
    return folder / f"{stem}.{digest}.generated{suffix}"


def has_excluded_path(root: Path, path: Path) -> bool:
    if path.name.startswith("~$") or path.name.startswith(".~"):
        return True
    parts = set(path.relative_to(root).parts)
    if parts & EXCLUDED_PARTS:
        return True
    path_text = "/" + rel(root, path).replace("\\", "/")
    return any(re.search(pattern, path_text, re.IGNORECASE) for pattern in EXCLUDED_PATH_PATTERNS)


def is_under(path: Path, base: Path) -> bool:
    try:
        path.relative_to(base)
        return True
    except ValueError:
        return False


def is_candidate_dir(root: Path, path: Path) -> bool:
    if not path.is_dir() or path.name.startswith("."):
        return False
    return not has_excluded_path(root, path)


def iter_dirs(root: Path, max_depth: int) -> Iterable[Path]:
    stack: list[tuple[Path, int]] = [(root, 0)]
    while stack:
        folder, depth = stack.pop()
        if depth >= max_depth:
            continue
        for child in folder.iterdir():
            if not is_candidate_dir(root, child):
                continue
            stack.append((child, depth + 1))
            yield child


def discover_roots(root: Path) -> tuple[list[Path], list[Path]]:
    business_roots: list[Path] = []
    project_roots = [
        folder
        for folder in iter_dirs(root, max_depth=3)
        if any((folder / name).exists() for name in MANIFEST_NAMES)
    ]

    for child in root.iterdir():
        if not is_candidate_dir(root, child):
            continue
        if any(file.is_file() and file.suffix.lower() in BUSINESS_EXTS for file in child.rglob("*")):
            business_roots.append(child)

    return sorted(set(business_roots)), sorted(set(project_roots))


def classify_file(root: Path, path: Path, business_roots: list[Path], project_roots: list[Path]) -> str | None:
    if has_excluded_path(root, path):
        return None
    if path.name in {INDEX_NAME, ROOT_INDEX_NAME} or path.name.startswith("_auto_"):
        return None
    ext = path.suffix.lower()
    name = path.name.lower()

    for project_root in project_roots:
        if is_under(path, project_root):
            relative_parts = [part.lower() for part in path.relative_to(project_root).parts]
            is_project_core = (
                name in KEY_NAMES
                or any(part in CODE_DIRS for part in relative_parts)
                or len(relative_parts) <= 2
            )
            if ext in CODE_EXTS and is_project_core:
                return "code_project_core"
            break

    if any(is_under(path, base) for base in business_roots):
        return "business_material" if ext in BUSINESS_EXTS else None

    if ext == ".md" and len(path.relative_to(root).parts) <= 2:
        return "workspace_note"
    return None


def collect_records(
    root: Path, business_roots: list[Path] | None = None, project_roots: list[Path] | None = None
) -> list[CoreRecord]:
    discovered_business, discovered_projects = discover_roots(root)
    business_roots = business_roots or discovered_business
    project_roots = project_roots or discovered_projects
    records: list[CoreRecord] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        kind = classify_file(root, path, business_roots, project_roots)
        if not kind:
            continue
        note_path = path if path.suffix.lower() == ".md" else safe_generated_path(
            path.parent, path.name + ".core.md", FILE_MARKER
        )
        records.append(CoreRecord(path=path, root=root, kind=kind, note_path=note_path))
    return sorted(records, key=lambda item: item.rel.lower())


def read_limited_bytes(path: Path, limit: int) -> tuple[bytes, bool]:
    with open(long_path(path), "rb") as handle:
        data = handle.read(limit + 1)
    return data[:limit], len(data) > limit


def decode_text(data: bytes) -> tuple[str, str, str | None]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "cp936", "utf-16", "latin-1"):
        try:
            return data.decode(encoding), encoding, None
        except Exception as error:
            last_error = error
    return data.decode("utf-8", errors="replace"), "utf-8-replace", str(last_error)


def keywords(text: str, limit: int = 10) -> list[str]:
    words = re.findall(r"[A-Za-z_][A-Za-z0-9_]{3,}|[\u4e00-\u9fff]{2,}", text or "")
    stop = {
        "from",
        "import",
        "class",
        "return",
        "true",
        "false",
        "none",
        "null",
        "this",
        "that",
        "with",
        "have",
        "will",
        "using",
        "data",
        "file",
        "page",
        "and",
        "for",
        "the",
        "not",
    }
    counts = Counter(word for word in words if word.lower() not in stop)
    return [word for word, _ in counts.most_common(limit)]


def summarize_text(path: Path, label: str, limit: int = MAX_TEXT_BYTES) -> tuple[list[str], list[str], str, str]:
    data, truncated = read_limited_bytes(path, limit)
    text, encoding, error = decode_text(data)
    lines = text.splitlines()
    nonempty = [line.strip()[:180] for line in lines if line.strip()][:8]
    summary = [f"{label} sample: {len(lines)} lines, {len(text)} chars."]
    if nonempty:
        summary.append("Opening: " + nonempty[0])
    key_terms = keywords(text)
    if key_terms:
        summary.append("Keywords: " + ", ".join(key_terms))
    notes = [f"encoding={encoding}"]
    if truncated:
        notes.append("sampled beginning only")
    if error:
        notes.append("decode fallback used")
    return summary, notes, text[:MAX_PREVIEW_CHARS], "text"


def summarize_json(path: Path) -> tuple[list[str], list[str], str, str]:
    data, truncated = read_limited_bytes(path, MAX_JSON_BYTES)
    text, encoding, error = decode_text(data)
    notes = [f"encoding={encoding}"]
    summary: list[str] = []
    preview = text[:MAX_PREVIEW_CHARS]
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            keys = list(parsed.keys())
            summary.append(f"JSON object with {len(keys)} top-level keys.")
            if keys:
                summary.append("Top keys: " + ", ".join(map(str, keys[:15])))
        elif isinstance(parsed, list):
            summary.append(f"JSON array with {len(parsed)} items.")
        else:
            summary.append(f"JSON scalar: {type(parsed).__name__}")
        preview = json.dumps(parsed, ensure_ascii=False, indent=2)[:MAX_PREVIEW_CHARS]
    except Exception as error_json:
        summary.append(f"JSON parse failed; treated as text: {type(error_json).__name__}")
    if truncated:
        notes.append("sampled beginning only")
    if error:
        notes.append("decode fallback used")
    return summary, notes, preview, "json"


def xml_texts(element: ET.Element, suffix: str) -> list[str]:
    return [node.text for node in element.iter() if node.tag.endswith(suffix) and node.text]


def summarize_docx(path: Path) -> tuple[list[str], list[str], str, str]:
    try:
        with zipfile.ZipFile(long_path(path)) as archive:
            document = ET.fromstring(archive.read("word/document.xml"))
            paragraphs = []
            for para in document.iter():
                if para.tag.endswith("}p"):
                    line = "".join(xml_texts(para, "}t")).strip()
                    if line:
                        paragraphs.append(line)
        joined = "\n".join(paragraphs)
        summary = [f"DOCX with {len(paragraphs)} non-empty paragraphs."]
        if paragraphs:
            summary.append("Opening: " + paragraphs[0][:180])
        key_terms = keywords(joined)
        if key_terms:
            summary.append("Keywords: " + ", ".join(key_terms))
        return summary, [], joined[:MAX_PREVIEW_CHARS], "text"
    except Exception as error:
        return [f"DOCX body extraction failed: {type(error).__name__}."], ["metadata only"], "", "text"


def summarize_pptx(path: Path) -> tuple[list[str], list[str], str, str]:
    try:
        with zipfile.ZipFile(long_path(path)) as archive:
            slide_names = [
                name for name in archive.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)
            ]
            slide_names.sort(key=lambda value: int(re.search(r"slide(\d+)\.xml", value).group(1)))  # type: ignore[union-attr]
            slides = []
            for slide_name in slide_names:
                slide = ET.fromstring(archive.read(slide_name))
                slides.append([text.strip() for text in xml_texts(slide, "}t") if text.strip()])
        all_text = "\n".join(" / ".join(slide) for slide in slides)
        summary = [f"PPTX with {len(slides)} slides."]
        first = next((slide[0] for slide in slides if slide), None)
        if first:
            summary.append("First text: " + first[:180])
        key_terms = keywords(all_text)
        if key_terms:
            summary.append("Keywords: " + ", ".join(key_terms))
        return summary, [], all_text[:MAX_PREVIEW_CHARS], "text"
    except Exception as error:
        return [f"PPTX body extraction failed: {type(error).__name__}."], ["metadata only"], "", "text"


def summarize_xlsx(path: Path) -> tuple[list[str], list[str], str, str]:
    if not openpyxl:
        return ["Excel parser unavailable."], ["metadata only"], "", "text"
    try:
        workbook = openpyxl.load_workbook(long_path(path), read_only=True, data_only=True)
        preview_parts: list[str] = []
        sheet_lines: list[str] = []
        for sheet in workbook.worksheets[:10]:
            sample_rows = []
            for row in sheet.iter_rows(max_row=4, values_only=True):
                values = ["" if value is None else str(value) for value in row[:8]]
                if any(values):
                    sample_rows.append(values)
            sheet_lines.append(f"{sheet.title}: {sheet.max_row or 0} rows x {sheet.max_column or 0} cols")
            if sample_rows:
                preview_parts.append(f"[{sheet.title}]\n" + "\n".join("\t".join(row) for row in sample_rows))
        sheet_names = workbook.sheetnames
        workbook.close()
        summary = [f"Excel workbook with {len(sheet_names)} sheets.", "Sheets: " + ", ".join(sheet_names[:12])]
        summary.extend(sheet_lines[:6])
        return summary, [], "\n\n".join(preview_parts)[:MAX_PREVIEW_CHARS], "text"
    except Exception as error:
        return [f"Excel extraction failed: {type(error).__name__}."], ["metadata only"], "", "text"


def summarize_pdf(path: Path) -> tuple[list[str], list[str], str, str]:
    try:
        with open(long_path(path), "rb") as handle:
            header = handle.read(8192)
            sample = header + handle.read(8_000_000)
        version_match = re.search(rb"%PDF-([0-9.]+)", header)
        version = version_match.group(1).decode() if version_match else "unknown"
        pages = len(re.findall(rb"/Type\s*/Page\b", sample))
        return [
            f"PDF metadata: version {version}.",
            f"Approximate pages from object count: {pages or 'not detected'}.",
        ], ["PDF body text parser unavailable; metadata-level note."], "", "text"
    except Exception as error:
        return [f"PDF metadata extraction failed: {type(error).__name__}."], ["metadata only"], "", "text"


def summarize_source(path: Path) -> tuple[list[str], list[str], str, str]:
    ext = path.suffix.lower()
    if ext == ".docx":
        return summarize_docx(path)
    if ext == ".pptx":
        return summarize_pptx(path)
    if ext in {".xlsx", ".xlsm"}:
        return summarize_xlsx(path)
    if ext == ".pdf":
        return summarize_pdf(path)
    if ext == ".json":
        return summarize_json(path)
    if ext == ".py":
        summary, notes, preview, lang = summarize_text(path, "PY")
        try:
            tree = ast.parse(preview)
            funcs = [node.name for node in tree.body if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))]
            classes = [node.name for node in tree.body if isinstance(node, ast.ClassDef)]
            if classes:
                summary.append("Classes: " + ", ".join(classes[:10]))
            if funcs:
                summary.append("Functions: " + ", ".join(funcs[:12]))
        except Exception:
            pass
        return summary, notes, preview, lang
    if ext in {".yaml", ".yml", ".txt", ".csv", ".toml", ".cmd", ".ps1", ".sh", ".js", ".ts", ".html", ".css", ".svelte"}:
        return summarize_text(path, ext.lstrip(".").upper())
    if ext in {".doc", ".ppt", ".xls"}:
        return [f"Legacy Office binary file ({ext}); metadata-level note."], ["legacy Office format"], "", "text"
    return [f"{ext or 'unknown'} file; metadata-level note."], ["no extractor configured"], "", "text"


def normalize_name(path: Path) -> str:
    text = path.stem.lower()
    text = re.sub(r"\.(docx|pptx|xlsx|pdf|json|csv|txt|py|js|md|core)$", "", text)
    text = re.sub(r"20\d{6,}|20\d{2}[-_]?\d{1,2}[-_]?\d{1,2}", "", text)
    text = re.sub(r"v\d+(\.\d+)?", "", text, flags=re.IGNORECASE)
    return re.sub("[^0-9A-Za-z_\u4e00-\u9fff]+", "", text)


def relation_score(left: CoreRecord, right: CoreRecord) -> float:
    base = 1.0 if left.path.parent == right.path.parent else 0.2
    if left.path.parent != right.path.parent:
        common = 0
        for left_part, right_part in zip(left.rel.split("/")[:-1], right.rel.split("/")[:-1]):
            if left_part == right_part:
                common += 1
            else:
                break
        base += 0.08 * common
    similarity = SequenceMatcher(None, left.norm, right.norm).ratio() if left.norm and right.norm else 0
    return base + similarity + (0.1 if left.kind == right.kind else 0)


def relation_reason(left: CoreRecord, right: CoreRecord) -> str:
    if left.path.parent == right.path.parent and left.norm == right.norm and left.norm:
        return "same folder and normalized name"
    if left.path.parent == right.path.parent:
        return "same folder"
    return "shared project/topic path"


def build_relations(records: list[CoreRecord]) -> dict[str, list[tuple[CoreRecord, str]]]:
    by_folder: dict[Path, list[CoreRecord]] = defaultdict(list)
    for record in records:
        by_folder[record.path.parent].append(record)

    relations: dict[str, list[tuple[CoreRecord, str]]] = defaultdict(list)
    for items in by_folder.values():
        for left in items:
            scored = [
                (relation_score(left, right), right, relation_reason(left, right))
                for right in items
                if right is not left
            ]
            for score, right, reason in sorted(scored, key=lambda item: item[0], reverse=True)[:8]:
                if score >= 1.05:
                    relations[left.rel].append((right, reason))

    by_norm: dict[str, list[CoreRecord]] = defaultdict(list)
    for record in records:
        if len(record.norm) >= 4:
            by_norm[record.norm].append(record)
    for items in by_norm.values():
        if len(items) < 2 or len(items) > 30:
            continue
        for left in items:
            for right in items:
                if left is not right and len(relations[left.rel]) < 12:
                    relations[left.rel].append((right, "same normalized title/version family"))

    for key, values in list(relations.items()):
        seen = set()
        deduped = []
        for record, reason in values:
            if record.rel in seen:
                continue
            seen.add(record.rel)
            deduped.append((record, reason))
        relations[key] = deduped[:12]
    return relations


def write_core_note(record: CoreRecord, relations: dict[str, list[tuple[CoreRecord, str]]], generated: str) -> None:
    summary, notes, preview, lang = summarize_source(record.path)
    tag_kind = "business-material" if record.kind == "business_material" else "code-project-core"
    folder_index = safe_generated_path(record.path.parent, INDEX_NAME, FOLDER_MARKER)
    lines = [
        "---",
        f"source: {yaml_quote(record.rel)}",
        f"core_kind: {yaml_quote(record.kind)}",
        f"file_type: {yaml_quote(record.ext.lstrip('.') or 'unknown')}",
        f"generated: {yaml_quote(generated)}",
        "tags:",
        "  - core-file",
        f"  - {tag_kind}",
        f"  - {record.ext.lstrip('.') or 'unknown'}",
        "---",
        FILE_MARKER,
        f"# {record.path.name}",
        "",
        f"- Source file: {wikilink(record.root, record.path, record.path.name, attachment=True)}",
        f"- Folder index: {wikilink(record.root, folder_index, INDEX_ALIAS)}",
        f"- Core kind: `{record.kind}`",
        f"- Size: {format_size(record.path.stat().st_size)}",
        f"- Modified: {datetime.fromtimestamp(record.path.stat().st_mtime).isoformat(timespec='seconds')}",
        "",
        "## Summary",
    ]
    lines.extend(f"- {item}" for item in summary)
    if relations.get(record.rel):
        lines.extend(["", "## Related files"])
        for related, reason in relations[record.rel]:
            lines.append(f"- {wikilink(record.root, related.note_path, related.path.name)} - {reason}")
    lines.extend(["", "## Graph tags", f"- #core-file/{tag_kind}", f"- #file-type/{record.ext.lstrip('.') or 'unknown'}"])
    if preview:
        lines.extend(["", "## Content preview", fence(preview, lang)])
    if notes:
        lines.extend(["", "## Processing notes"])
        lines.extend(f"- {note}" for note in notes)
    lines.append("")
    write_text(record.note_path, "\n".join(lines))


def write_folder_indexes(root: Path, records: list[CoreRecord], relations: dict[str, list[tuple[CoreRecord, str]]], generated: str) -> list[Path]:
    by_folder: dict[Path, list[CoreRecord]] = defaultdict(list)
    for record in records:
        by_folder[record.path.parent].append(record)

    folders = {root}
    for record in records:
        folder = record.path.parent
        folders.add(folder)
        current = root
        for part in folder.relative_to(root).parts:
            current = current / part
            folders.add(current)

    written: list[Path] = []
    for folder in sorted(folders, key=lambda item: rel(root, item).lower() if item != root else ""):
        descendants = [record for record in records if is_under(record.path, folder)]
        if not descendants:
            continue
        direct = sorted(by_folder.get(folder, []), key=lambda item: item.path.name.lower())
        child_counts: Counter[Path] = Counter()
        for record in descendants:
            if record.path.parent == folder:
                continue
            try:
                child_counts[folder / record.path.parent.relative_to(folder).parts[0]] += 1
            except Exception:
                pass

        index_path = safe_generated_path(folder, INDEX_NAME, FOLDER_MARKER)
        label = "workspace root" if folder == root else rel(root, folder)
        title = folder.name if folder != root else WORKSPACE_LABEL
        lines = [
            "---",
            f"generated: {yaml_quote(generated)}",
            f"source_folder: {yaml_quote(label)}",
            "tags:",
            "  - core-folder-index",
            "---",
            FOLDER_MARKER,
            f"# {INDEX_TITLE_PREFIX}{title}",
            "",
            f"- Source folder: `{label}`",
            f"- Direct core files: {len(direct)}",
            f"- Recursive core files: {len(descendants)}",
            "",
            "## Direct core files",
        ]
        if direct:
            for record in direct:
                if record.ext == ".md":
                    lines.append(f"- {wikilink(root, record.path, record.path.name)} - `{record.kind}`")
                else:
                    source_link = wikilink(root, record.path, record.path.name, attachment=True)
                    lines.append(
                        f"- {wikilink(root, record.note_path, record.path.name)} - `{record.kind}`; source {source_link}"
                    )
        else:
            lines.append("- No direct core files; see child folders.")
        if child_counts:
            lines.extend(["", "## Child folder indexes"])
            for child, count in child_counts.most_common(30):
                child_index = safe_generated_path(child, INDEX_NAME, FOLDER_MARKER)
                lines.append(f"- {wikilink(root, child_index, child.name)} - {count} recursive core files")
        local_pairs = []
        for record in direct:
            for related, reason in relations.get(record.rel, [])[:3]:
                if related.path.parent == folder:
                    local_pairs.append((record, related, reason))
        if local_pairs:
            lines.extend(["", "## Local relationship highlights"])
            seen = set()
            for left, right, reason in local_pairs[:20]:
                pair = tuple(sorted([left.rel, right.rel]))
                if pair in seen:
                    continue
                seen.add(pair)
                lines.append(
                    f"- {wikilink(root, left.note_path, left.path.name)} ↔ {wikilink(root, right.note_path, right.path.name)} - {reason}"
                )
        lines.extend(["", "## Tags", "- #core-folder-index", ""])
        write_text(index_path, "\n".join(lines))
        written.append(index_path)
    return written


def write_root_index(root: Path, records: list[CoreRecord], folder_indexes: list[Path], generated: str) -> Path:
    root_index = safe_generated_path(root, ROOT_INDEX_NAME, ROOT_MARKER)
    kind_counts = Counter(record.kind for record in records)
    ext_counts = Counter(record.ext for record in records)
    top_counts = Counter(record.rel.split("/")[0] for record in records)
    generated_non_md = sum(1 for record in records if record.ext != ".md")
    existing_md = sum(1 for record in records if record.ext == ".md")
    lines = [
        "---",
        f"generated: {yaml_quote(generated)}",
        "tags:",
        "  - core-workspace-index",
        "---",
        ROOT_MARKER,
        f"# {ROOT_TITLE}",
        "",
        f"- Generated: {generated}",
        f"- Core files: {len(records)}",
        f"- Generated non-md notes: {generated_non_md}",
        f"- Existing md referenced: {existing_md}",
        f"- Folder indexes: {len(folder_indexes)}",
        "",
        "## Selection rules",
        "- Business materials: Office/PDF/table/text/data files under discovered business roots.",
        "- Code project core files: README, manifests, configs, source, tests, scripts, templates, and command/rule files in project folders.",
        "- Excluded: dependencies, caches, logs, model outputs, generated workspaces, static vendor/resource bundles, temporary Office lock files.",
        "",
        "## Core distribution by kind",
        "| Kind | Count |",
        "|---|---:|",
    ]
    lines.extend(f"| `{kind}` | {count} |" for kind, count in kind_counts.most_common())
    lines.extend(["", "## Core distribution by extension", "| Extension | Count |", "|---|---:|"])
    lines.extend(f"| `{ext or '[no_ext]'}` | {count} |" for ext, count in ext_counts.most_common())
    lines.extend(["", "## Top-level navigation", "| Top folder | Core files | Index |", "|---|---:|---|"])
    for top, count in top_counts.most_common():
        folder = root / top
        index_path = safe_generated_path(folder, INDEX_NAME, FOLDER_MARKER) if folder.exists() else root_index
        lines.append(f"| `{md_escape(top)}` | {count} | {wikilink(root, index_path, top)} |")
    lines.append("")
    write_text(root_index, "\n".join(lines))
    return root_index


def command_scan(args: argparse.Namespace) -> int:
    root = args.root.resolve()
    business_roots = [Path(value).resolve() for value in args.business_root] if args.business_root else None
    project_roots = [Path(value).resolve() for value in args.project_root] if args.project_root else None
    records = collect_records(root, business_roots, project_roots)
    discovered_business, discovered_projects = discover_roots(root)
    payload = {
        "root": str(root),
        "core_files": len(records),
        "by_kind": Counter(record.kind for record in records),
        "by_extension": Counter(record.ext for record in records),
        "business_roots": [rel(root, value) for value in (business_roots or discovered_business)],
        "project_roots": [rel(root, value) for value in (project_roots or discovered_projects)],
    }
    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, default=dict))
    else:
        print(f"root={payload['root']}")
        print(f"core_files={payload['core_files']}")
        print("by_kind=" + json.dumps(dict(payload["by_kind"]), ensure_ascii=False))
        print("by_extension=" + json.dumps(dict(payload["by_extension"]), ensure_ascii=False))
    return 0


def is_generated_note_name(path: Path) -> bool:
    return path.name in {INDEX_NAME, ROOT_INDEX_NAME, TOPIC_NOTE_NAME} or path.name.endswith(".core.md")


def direct_topic_sources(root: Path, folder: Path) -> list[Path]:
    files: list[Path] = []
    for path in folder.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$") or path.name.startswith(".~"):
            continue
        if is_generated_note_name(path) or has_excluded_path(root, path):
            continue
        if path.suffix.lower() in BUSINESS_EXTS | CODE_EXTS:
            files.append(path)
    return sorted(files, key=lambda item: item.name.lower())


def topic_sample(path: Path) -> tuple[str, int, bool]:
    ext = path.suffix.lower()
    if ext in {".md", ".txt"}:
        data, _ = read_limited_bytes(path, TOPIC_SAMPLE_BYTES)
        text, _, _ = decode_text(data)
        return text[:12_000], len(text), True
    if ext == ".docx":
        summary, _, preview, _ = summarize_docx(path)
        text = "\n".join(summary + ([preview] if preview else []))
        return text, len(preview), bool(text.strip())
    if ext == ".pptx":
        summary, _, preview, _ = summarize_pptx(path)
        text = "\n".join(summary + ([preview] if preview else []))
        return text, len(preview), bool(text.strip())
    return "", 0, False


def keyword_hit_count(text: str) -> int:
    lower = text.lower()
    return sum(lower.count(keyword.lower()) for keyword in TOPIC_CANDIDATE_KEYWORDS)


def topic_candidate_for_folder(root: Path, folder: Path, sample_files: int) -> dict[str, object] | None:
    files = direct_topic_sources(root, folder)
    if not files:
        return None
    topic_note = folder / TOPIC_NOTE_NAME
    topic_text = ""
    if topic_note.exists():
        try:
            topic_text = read_text(topic_note)
        except Exception:
            topic_text = ""

    text_chars = 0
    readable_sources = 0
    keyword_hits = keyword_hit_count(rel(root, folder))
    heading_count = 0
    version_hits = len(TOPIC_VERSION_PATTERN.findall(rel(root, folder)))
    main_doc_count = 0
    sampled_files = []
    ext_counts = Counter(path.suffix.lower() or "[no_ext]" for path in files)

    for path in sorted(files, key=lambda item: item.stat().st_size, reverse=True)[:sample_files]:
        name_text = path.name
        keyword_hits += keyword_hit_count(name_text)
        version_hits += len(TOPIC_VERSION_PATTERN.findall(name_text))
        if TOPIC_MAIN_DOC_PATTERN.search(name_text):
            main_doc_count += 1
        sample, sample_chars, readable = topic_sample(path)
        text_chars += sample_chars
        if readable:
            readable_sources += 1
        if sample:
            keyword_hits += keyword_hit_count(sample)
            heading_count += len(re.findall(r"^#{1,4}\s+", sample, flags=re.MULTILINE))
        sampled_files.append(path.name)

    score = 0
    score += min(len(files), 30) * 3
    score += len(ext_counts) * 5
    score += readable_sources * 8
    score += min(text_chars // 1000, 40) * 2
    score += min(keyword_hits, 120)
    score += min(version_hits, 30) * 3
    score += min(heading_count, 60) // 2
    score += min(main_doc_count, 10) * 6

    reasons = []
    if len(files) >= 5:
        reasons.append("multiple direct sources")
    if readable_sources:
        reasons.append(f"{readable_sources} readable sampled sources")
    if version_hits:
        reasons.append("version/history cues")
    if keyword_hits >= 10:
        reasons.append("business/AI topic keywords")
    if main_doc_count:
        reasons.append("strategy/report/meeting/architecture source names")
    if TOPIC_SYNTHESIS_SCHEMA in topic_text:
        reasons.append("already has deep topic schema")

    return {
        "folder": rel(root, folder),
        "score": score,
        "source_count": len(files),
        "readable_sampled_sources": readable_sources,
        "sampled_text_chars": text_chars,
        "keyword_hits": keyword_hits,
        "version_hits": version_hits,
        "heading_count": heading_count,
        "main_doc_count": main_doc_count,
        "extensions": dict(ext_counts),
        "topic_note_exists": topic_note.exists(),
        "topic_note_size": topic_note.stat().st_size if topic_note.exists() else 0,
        "deep_topic_note": TOPIC_SYNTHESIS_SCHEMA in topic_text,
        "reasons": reasons,
        "sampled_files": sampled_files[:8],
    }


def command_topic_candidates(args: argparse.Namespace) -> int:
    root = args.root.resolve()
    folders = {path.parent for path in root.rglob(TOPIC_NOTE_NAME) if not has_excluded_path(root, path)}
    if args.include_without_topic_note:
        for folder in root.rglob("*"):
            if folder.is_dir() and not has_excluded_path(root, folder):
                folders.add(folder)

    rows = []
    exclude_prefixes = tuple(value.replace("\\", "/").strip("/") for value in args.exclude_prefix)
    for folder in folders:
        folder_rel = rel(root, folder)
        if exclude_prefixes and any(folder_rel == prefix or folder_rel.startswith(prefix + "/") for prefix in exclude_prefixes):
            continue
        candidate = topic_candidate_for_folder(root, folder, args.sample_files)
        if not candidate:
            continue
        if candidate["score"] < args.min_score:
            continue
        if args.only_missing_deep and candidate["deep_topic_note"]:
            continue
        rows.append(candidate)

    rows.sort(key=lambda item: (item["score"], item["sampled_text_chars"], item["source_count"]), reverse=True)
    rows = rows[: args.limit]
    if args.json:
        print(json.dumps({"root": str(root), "candidates": rows}, ensure_ascii=False, indent=2))
    else:
        print("score\tdeep\tsources\tfolder")
        for row in rows:
            print(f"{row['score']}\t{row['deep_topic_note']}\t{row['source_count']}\t{row['folder']}")
    return 0


def command_refresh(args: argparse.Namespace) -> int:
    root = args.root.resolve()
    business_roots = [Path(value).resolve() for value in args.business_root] if args.business_root else None
    project_roots = [Path(value).resolve() for value in args.project_root] if args.project_root else None
    records = collect_records(root, business_roots, project_roots)
    relations = build_relations(records)
    generated = datetime.now().isoformat(timespec="seconds")
    if args.dry_run:
        print(f"dry_run=true core_files={len(records)} non_md={sum(1 for item in records if item.ext != '.md')}")
        return 0
    for record in records:
        if record.ext != ".md":
            write_core_note(record, relations, generated)
    folder_indexes = write_folder_indexes(root, records, relations, generated)
    root_index = write_root_index(root, records, folder_indexes, generated)
    print(
        json.dumps(
            {
                "core_files": len(records),
                "generated_non_md_notes": sum(1 for item in records if item.ext != ".md"),
                "existing_md_referenced": sum(1 for item in records if item.ext == ".md"),
                "folder_indexes": len(folder_indexes),
                "root_index": rel(root, root_index),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def generated_marker(text: str) -> str | None:
    for marker in (FILE_MARKER, FOLDER_MARKER, ROOT_MARKER):
        if marker in text[:600]:
            return marker
    return None


def command_validate(args: argparse.Namespace) -> int:
    root = args.root.resolve()
    counts = Counter()
    bad_question = 0
    old_auto = 0
    non_core_named = 0
    topic_notes = 0
    deep_topic_notes = 0
    for path in root.rglob("*.md"):
        if path.name.startswith("_auto_non_md"):
            old_auto += 1
        try:
            text = read_text(path)
        except Exception:
            continue
        if path.name == TOPIC_NOTE_NAME:
            topic_notes += 1
            if TOPIC_SYNTHESIS_SCHEMA in text:
                deep_topic_notes += 1
        marker = generated_marker(text)
        if not marker:
            continue
        counts[marker] += 1
        if "????" in text:
            bad_question += 1
        if marker == FILE_MARKER and not path.name.endswith(".core.md"):
            non_core_named += 1
    payload = {
        "file_notes": counts[FILE_MARKER],
        "folder_indexes": counts[FOLDER_MARKER],
        "workspace_indexes": counts[ROOT_MARKER],
        "topic_notes": topic_notes,
        "deep_topic_notes": deep_topic_notes,
        "old_auto_non_md": old_auto,
        "question_mark_corruption": bad_question,
        "file_notes_not_core_named": non_core_named,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 1 if old_auto or bad_question or non_core_named else 0


def command_clean_generated(args: argparse.Namespace) -> int:
    root = args.root.resolve()
    targets = []
    for path in root.rglob("*.md"):
        try:
            text = read_text(path)
        except Exception:
            continue
        if generated_marker(text):
            targets.append(path)
    if args.dry_run:
        for path in targets:
            print(rel(root, path))
        print(f"dry_run=true generated_marked_files={len(targets)}")
        return 0
    for path in targets:
        path.unlink()
    print(f"deleted_generated_marked_files={len(targets)}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subcommands = parser.add_subparsers(dest="command", required=True)

    scan = subcommands.add_parser("scan", help="Scan core files without writing.")
    scan.add_argument("--root", type=Path, default=Path.cwd())
    scan.add_argument("--business-root", action="append", default=[])
    scan.add_argument("--project-root", action="append", default=[])
    scan.add_argument("--json", action="store_true")
    scan.set_defaults(func=command_scan)

    topic = subcommands.add_parser("topic-candidates", help="Rank folders that may justify V3-like topic synthesis notes.")
    topic.add_argument("--root", type=Path, default=Path.cwd())
    topic.add_argument("--json", action="store_true")
    topic.add_argument("--limit", type=int, default=50)
    topic.add_argument("--min-score", type=int, default=80)
    topic.add_argument("--sample-files", type=int, default=20)
    topic.add_argument("--exclude-prefix", action="append", default=[])
    topic.add_argument("--only-missing-deep", action="store_true")
    topic.add_argument("--include-without-topic-note", action="store_true")
    topic.set_defaults(func=command_topic_candidates)

    refresh = subcommands.add_parser("refresh", help="Generate or refresh core notes and indexes.")
    refresh.add_argument("--root", type=Path, default=Path.cwd())
    refresh.add_argument("--business-root", action="append", default=[])
    refresh.add_argument("--project-root", action="append", default=[])
    refresh.add_argument("--dry-run", action="store_true")
    refresh.set_defaults(func=command_refresh)

    validate = subcommands.add_parser("validate", help="Validate generated note markers and naming.")
    validate.add_argument("--root", type=Path, default=Path.cwd())
    validate.set_defaults(func=command_validate)

    clean = subcommands.add_parser("clean-generated", help="Delete generated Markdown files with core markers.")
    clean.add_argument("--root", type=Path, default=Path.cwd())
    clean.add_argument("--dry-run", action="store_true")
    clean.set_defaults(func=command_clean_generated)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
