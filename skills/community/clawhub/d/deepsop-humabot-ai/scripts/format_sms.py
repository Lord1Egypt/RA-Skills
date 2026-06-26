#!/usr/bin/env python3
"""
format_sms.py
将 Lisa 短信任务 getSmsResultList 返回的 JSON 数据生成 xlsx 文件
用法: python3 format_sms.py '<json_string>' '<output_path>'
输出: xlsx 文件路径
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_LABEL = {
    1: '发送成功',
    0: '发送失败',
}


def main():
    if len(sys.argv) < 3:
        print("用法: python3 format_sms.py '<json_string>' '<output_path>'", file=sys.stderr)
        sys.exit(1)

    try:
        raw = json.loads(sys.argv[1])
        if isinstance(raw, dict):
            data_root = raw.get('data', raw)
            data = data_root.get('rows', data_root) if isinstance(data_root, dict) else data_root
        else:
            data = raw
    except json.JSONDecodeError as e:
        print(f'JSON 解析失败: {e}', file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, list) or len(data) == 0:
        print('无短信数据', file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[2]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 表头定义: (显示名, 字段key, 列宽)
    headers = [
        ('序号',       'index',      8),
        ('手机号码',   'phoneNumber', 18),
        ('发送状态',   'success',    14),
        ('状态描述',   'errMsg',     20),
        ('状态码',     'errCode',    20),
        ('短信内容',   'content',    60),
        ('短信条数',   'smsSize',    10),
        ('发送时间',   'sendTime',   20),
        ('回执时间',   'reportTime', 20),
        ('业务ID',     'bizId',      32),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lisa短信数据'

    # 表头样式（紫色主题）
    header_font  = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='553C9A', end_color='553C9A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (label, key, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    fill_success = PatternFill(start_color='E9D8FD', end_color='E9D8FD', fill_type='solid')
    fill_fail    = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
    alt_fill     = PatternFill(start_color='F5F0FF', end_color='F5F0FF', fill_type='solid')
    data_align   = Alignment(vertical='center', wrap_text=False)

    for row_idx, item in enumerate(data, start=2):
        status_raw = item.get('success')

        for col_idx, (label, key, width) in enumerate(headers, start=1):
            if key == 'index':
                value = row_idx - 1
            elif key == 'success':
                value = STATUS_LABEL.get(status_raw, str(status_raw) if status_raw is not None else '')
            else:
                value = item.get(key)
                if value is None:
                    value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align

            if status_raw == 1:
                cell.fill = fill_success
            elif status_raw == 0:
                cell.fill = fill_fail
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(output_path)


if __name__ == '__main__':
    main()
