"""PDF → XLSX (提取表格)

用法:
    python pdf_to_xlsx.py input.pdf [output.xlsx]

参数:
    input.pdf   - PDF 文件路径
    output.xlsx - 输出文件路径 (默认: output.xlsx)

说明:
    每个检测到的表格会写入单独的 sheet，命名格式: p{页码}_t{表格序号}
"""
import pdfplumber
import openpyxl
import sys
import os


def main():
    if len(sys.argv) < 2:
        print("用法: python pdf_to_xlsx.py input.pdf [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"

    if not os.path.isfile(pdf_path):
        print(f"错误: 文件不存在 - {pdf_path}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws_total = wb.active
    ws_total.title = "All Tables"
    row_offset = 0
    table_count = 0

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for t_idx, table in enumerate(tables):
                if row_offset == 0 and t_idx == 0:
                    ws = ws_total
                else:
                    ws = wb.create_sheet(title=f"p{page_num + 1}_t{t_idx + 1}")
                for row in table:
                    ws.append(row)
                row_offset += len(table)
                table_count += 1
                print(f"Page {page_num + 1}, Table {t_idx + 1}: {len(table)} rows")

    # 如果 All Tables 是空的，移除它
    if ws_total.max_row == 1 and ws_total.max_column == 1:
        wb.remove(ws_total)
    else:
        pass

    wb.save(out_path)
    print(f"完成: {table_count} 个表格 -> {out_path}")


if __name__ == "__main__":
    main()