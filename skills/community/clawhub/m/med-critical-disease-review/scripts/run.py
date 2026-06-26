#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from format_assessment_nl import build_natural_language
from major_disease_assess import call_major_disease_assess, validate_payload

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


class PreprocessError(ValueError):
    pass


TEXT_FILE_TYPES = {"txt", "md"}
IMAGE_FILE_TYPES = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
TABLE_FILE_TYPES = {"csv", "xlsx", "xls"}
SUPPORTED_FILE_TYPES = TEXT_FILE_TYPES | IMAGE_FILE_TYPES | TABLE_FILE_TYPES | {"json", "pdf", "doc", "docx"}
DOC_TYPE_RULES = [
    ("住院病案首页", ("住院病案首页", "病案首页")),
    ("出院记录", ("出院记录",)),
    ("入院记录", ("入院记录",)),
    ("手术记录", ("手术记录",)),
    ("病程记录", ("病程记录",)),
    ("诊断证明", ("诊断证明", "诊断证明书")),
    ("检查记录", ("检查记录", "检查报告", "影像报告", "心电图报告")),
    ("检验记录", ("检验报告", "检验记录")),
]
METADATA_FIELD_MAP = {
    "regioncode": "regionCode",
    "hospitalid": "hospitalId",
    "hospitalname": "hospitalName",
    "hospitallevel": "hospiatLevel",
    "hospiatlevel": "hospiatLevel",
    "admissionid": "admissionId",
    "住院号": "admissionId",
    "医院名称": "hospitalName",
    "医院等级": "hospiatLevel",
    "地区编码": "regionCode",
}
MAIN_DIAG_KEYS = ("maindiagname", "主诊断", "主要诊断", "出院诊断", "入院诊断", "诊断")
OTHER_DIAG_KEYS = ("otherdiagname", "其他诊断")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preprocess pdf/doc/docx/xls/xlsx/csv/txt/json input into medicalRecord JSON, then run critical disease review."
    )
    parser.add_argument("--input", required=True, help="Input file path.")
    parser.add_argument(
        "--input-type",
        default="auto",
        choices=["auto", *sorted(SUPPORTED_FILE_TYPES)],
        help="Explicit input type. Default: auto.",
    )
    parser.add_argument("--sheet", default="", help="Optional sheet name for xlsx/xls inputs.")
    parser.add_argument("--encoding", default="utf-8", help="Text encoding for txt/csv inputs. Default: utf-8.")
    parser.add_argument("--disease", required=True, help="Disease type, e.g. aortic_surgery, heart_valve_surgery.")
    parser.add_argument("--model-type", default="qwq", help="Model type query param (default: qwq).")
    parser.add_argument("--timeout", type=int, default=60, help="HTTP timeout seconds (default: 60).")
    parser.add_argument("--output-json", default="", help="Path to save raw response JSON.")
    parser.add_argument("--output-text", default="", help="Path to save natural language summary.")
    parser.add_argument(
        "--save-prepared",
        action="store_true",
        help="Save the normalized medicalRecord JSON next to the outputs for debugging.",
    )
    return parser.parse_args()


def detect_input_type(path: Path, explicit: str) -> str:
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def infer_doc_type(text: str, fallback: str = "病历文书") -> str:
    content = text or ""
    for doc_type, keywords in DOC_TYPE_RULES:
        if any(keyword in content for keyword in keywords):
            return doc_type
    return fallback


def read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def extract_with_soffice(path: Path) -> str:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for office document conversion.")
    with tempfile.TemporaryDirectory(prefix="med-critical-review-") as tmp_dir:
        proc = subprocess.run(
            [
                office_bin,
                "--headless",
                "--convert-to",
                "txt:Text",
                "--outdir",
                tmp_dir,
                str(path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(f"Failed to convert office document to text: {proc.stderr.strip() or proc.stdout.strip()}")
        return out_path.read_text(encoding="utf-8", errors="replace")


def extract_pdf_pages(path: Path) -> List[str]:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            pages = [page for page in pages if page]
            if pages:
                return pages
        except Exception:
            pass

    pdf_to_text = shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return [proc.stdout]
    raise PreprocessError("Unable to extract text from pdf.")


def extract_image_text(path: Path) -> str:
    tesseract_bin = shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError("tesseract not found for image OCR.")
    langs = detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


def detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


def shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def load_input_artifact(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "pages", "pages": [read_text_file(path, encoding)]}
    if input_type == "json":
        return {"kind": "json", "data": read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        return {"kind": "pages", "pages": [extract_with_soffice(path)]}
    if input_type == "docx":
        return {"kind": "pages", "pages": [extract_docx_text(path)]}
    if input_type == "doc":
        return {"kind": "pages", "pages": [extract_with_soffice(path)]}
    if input_type == "pdf":
        return {"kind": "pages", "pages": extract_pdf_pages(path)}
    if input_type in IMAGE_FILE_TYPES:
        return {"kind": "pages", "pages": [extract_image_text(path)]}
    raise PreprocessError(f"Unsupported input type: {input_type}")


def extract_primary_diagnosis(text: str) -> Optional[str]:
    patterns = [
        r"(?:主要诊断|主诊断|出院诊断|入院诊断|诊断)\s*[:：]\s*(.+)",
        r"(?:术前诊断|术后诊断)\s*[:：]\s*(.+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        value = match.group(1).splitlines()[0].strip().strip("。；;")
        if value:
            return value[:400]
    return None


def build_doc(doc_type: str, text: str) -> Optional[Dict[str, Any]]:
    content = (text or "").strip()
    if not content:
        return None
    return {
        "docType": doc_type or infer_doc_type(content),
        "format_page_text": content,
    }


def build_medical_record_payload(docs: List[Dict[str, Any]], extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    if not docs:
        raise PreprocessError("No recognizable document text was extracted for medicalRecord.docs.")
    record: Dict[str, Any] = {"docs": docs}
    if extra:
        for key, value in extra.items():
            if value not in (None, ""):
                record[key] = value
    if "mainDiagName" not in record:
        for doc in docs:
            diagnosis = extract_primary_diagnosis(doc.get("format_page_text", ""))
            if diagnosis:
                record["mainDiagName"] = diagnosis
                break
    return {"medicalRecord": record}


def sanitize_existing_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    record = dict(data.get("medicalRecord") or {})
    docs = record.get("docs")
    if not isinstance(docs, list):
        raise PreprocessError("medicalRecord.docs must be a list.")
    normalized_docs: List[Dict[str, Any]] = []
    for item in docs:
        if not isinstance(item, dict):
            continue
        text = item.get("format_page_text") or item.get("text") or item.get("content")
        doc_type = item.get("docType") or infer_doc_type(str(text or ""))
        built = build_doc(str(doc_type), str(text or ""))
        if built:
            normalized_docs.append(built)
    if not normalized_docs:
        raise PreprocessError("medicalRecord.docs did not contain any usable document text.")
    record["docs"] = normalized_docs
    payload = {"medicalRecord": record}
    validate_payload(payload)
    return payload


def payload_from_json(data: Any) -> Dict[str, Any]:
    if isinstance(data, dict) and isinstance(data.get("medicalRecord"), dict):
        return sanitize_existing_payload(data)

    if isinstance(data, list):
        docs = []
        for item in data:
            if not isinstance(item, dict):
                continue
            text = item.get("ocrText") or item.get("format_page_text") or item.get("text") or item.get("content")
            doc_type = item.get("docType") or infer_doc_type(str(text or ""))
            built = build_doc(str(doc_type), str(text or ""))
            if built:
                docs.append(built)
        payload = build_medical_record_payload(docs)
        validate_payload(payload)
        return payload

    if isinstance(data, dict):
        if isinstance(data.get("docs"), list):
            docs = []
            for item in data["docs"]:
                if not isinstance(item, dict):
                    continue
                text = item.get("format_page_text") or item.get("text") or item.get("content")
                doc_type = item.get("docType") or item.get("docName") or infer_doc_type(str(text or ""))
                built = build_doc(str(doc_type), str(text or ""))
                if built:
                    docs.append(built)
            payload = build_medical_record_payload(docs)
            validate_payload(payload)
            return payload

        text = data.get("text") or data.get("content")
        if isinstance(text, str) and text.strip():
            payload = build_medical_record_payload([build_doc(infer_doc_type(text), text)])  # type: ignore[list-item]
            validate_payload(payload)
            return payload

    raise PreprocessError("Unsupported JSON schema for critical disease review.")


def first_matching_index(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(normalize_header(alias))
        if idx is not None:
            return idx
    return None


def payload_from_tables(tables: List[Dict[str, Any]]) -> Dict[str, Any]:
    docs: List[Dict[str, Any]] = []
    extra: Dict[str, Any] = {}

    for table in tables:
        rows = table.get("rows") or []
        if not rows:
            continue
        header_map = {normalize_header(cell): idx for idx, cell in enumerate(rows[0]) if cell.strip()}
        text_idx = first_matching_index(header_map, ("format_page_text", "text", "content", "正文", "内容", "文本"))
        doc_type_idx = first_matching_index(header_map, ("doctype", "type", "docname", "文书类型", "文档类型"))
        data_rows = rows[1:] if text_idx is not None else rows

        if text_idx is not None:
            for row in data_rows:
                if text_idx >= len(row):
                    continue
                text = row[text_idx].strip()
                if not text:
                    continue
                doc_type = row[doc_type_idx].strip() if doc_type_idx is not None and doc_type_idx < len(row) else infer_doc_type(text, table["name"])
                built = build_doc(doc_type, text)
                if built:
                    docs.append(built)
            continue

        fragments: List[str] = []
        for row in rows:
            if len(row) < 2:
                continue
            field = row[0].strip()
            value = row[1].strip()
            if not field or not value:
                continue
            field_key = normalize_header(field)
            mapped_key = METADATA_FIELD_MAP.get(field_key) or METADATA_FIELD_MAP.get(field)
            if mapped_key:
                extra[mapped_key] = value
                continue
            if field_key in {normalize_header(k) for k in MAIN_DIAG_KEYS}:
                extra.setdefault("mainDiagName", value)
                continue
            if field_key in {normalize_header(k) for k in OTHER_DIAG_KEYS}:
                extra.setdefault("otherDiagName", value)
                continue
            fragments.append(f"{field}：{value}")
        if fragments:
            built = build_doc(table["name"] or "病历文书", "\n".join(fragments))
            if built:
                docs.append(built)

    payload = build_medical_record_payload(docs, extra)
    validate_payload(payload)
    return payload


def payload_from_pages(pages: List[str]) -> Dict[str, Any]:
    docs: List[Dict[str, Any]] = []
    for page_text in pages:
        built = build_doc(infer_doc_type(page_text), page_text)
        if built:
            docs.append(built)
    payload = build_medical_record_payload(docs)
    validate_payload(payload)
    return payload


def preprocess_to_payload(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    artifact = load_input_artifact(path, input_type, encoding, sheet_name)
    kind = artifact["kind"]
    if kind == "pages":
        return payload_from_pages(artifact["pages"])
    if kind == "json":
        return payload_from_json(artifact["data"])
    if kind == "tables":
        return payload_from_tables(artifact["tables"])
    raise PreprocessError(f"Unsupported artifact kind: {kind}")


def build_prepared_path(output_json_path: Path) -> Path:
    return output_json_path.with_name(f"{output_json_path.stem}.prepared.json")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"✗ Error: Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        input_type = detect_input_type(input_path, args.input_type)
        payload = preprocess_to_payload(input_path, input_type, args.encoding, args.sheet)
    except Exception as exc:
        print(f"✗ Preprocess Error: {exc}", file=sys.stderr)
        return 1

    try:
        resp = call_major_disease_assess(args.disease, payload, model_type=args.model_type, timeout=args.timeout)
    except Exception as exc:
        print(f"✗ Error: {exc}", file=sys.stderr)
        return 1

    default_base = Path("../runs/med-major-disease-assess")
    out_json = Path(args.output_json) if args.output_json else (default_base / f"{args.disease}_resp.json")
    out_text = Path(args.output_text) if args.output_text else (default_base / f"{args.disease}_resp.txt")

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(resp, ensure_ascii=False, indent=2), encoding="utf-8")

    text = build_natural_language(resp)
    out_text.parent.mkdir(parents=True, exist_ok=True)
    out_text.write_text(text, encoding="utf-8")

    if args.save_prepared:
        prepared_path = build_prepared_path(out_json)
        prepared_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✓ Prepared medicalRecord saved to: {prepared_path}")

    print(f"✓ Saved raw JSON to: {out_json}")
    print(f"✓ Saved natural language to: {out_text}")
    print("\n--- Natural language preview ---")
    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
