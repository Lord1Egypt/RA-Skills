#!/usr/bin/env python3
"""
Excel hierarchical table structurer.

Structured restructuring of hierarchical Excel tables where:
  - Some columns are "fill-down" (parent categories repeated across child rows)
  - Some columns are "group-key" (define row groups)
  - Some columns are "group-header" (group-level info, first row only)
  - Some columns are "fill-down-group" (group-level info, all rows)

Usage:
    python3 scripts/restructure.py <input.xlsx> <output.xlsx> --spec '<JSON>'
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import argparse
import sys
from copy import copy


# ── Default formatting palette ──────────────────────────────────────────

DEFAULT_STYLES = {
    "header": {
        "font": {"bold": True, "size": 11, "color": "FFFFFF"},
        "fill": {"color": "4472C4"},
    },
    "fail": {
        "fill": {"color": "FFEBEE"},
        "font": {"color": "B71C1C", "bold": True},
        "left_border": {"style": "medium", "color": "FF1744"},
    },
    "pass": {
        "fill": {"color": "E8F5E9"},
        "font": {"color": "1B5E20", "bold": True},
    },
    "无法测试": {
        "fill": {"color": "FFF8E1"},
        "font": {"color": "E65100"},
    },
    "alt_row_1": {"fill": {"color": "FFFFFF"}},
    "alt_row_2": {"fill": {"color": "F2F2F2"}},
    "group_key_cell": {"font": {"bold": True}},
}

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _make_font(spec):
    if not spec:
        return None
    kw = {}
    for k in ("bold", "size"):
        if k in spec:
            kw[k] = spec[k]
    if "color" in spec:
        kw["color"] = spec["color"]
    return Font(**kw)


def _make_fill(spec):
    if not spec or "color" not in spec:
        return None
    return PatternFill(start_color=spec["color"], end_color=spec["color"], fill_type="solid")


def _make_border(left_spec=None):
    """Return a Border with optional left-side accent."""
    kw = {"left": THIN.left, "right": THIN.right,
          "top": THIN.top, "bottom": THIN.bottom}
    if left_spec:
        kw["left"] = Side(style=left_spec.get("style", "thin"),
                          color=left_spec.get("color", "000000"))
    return Border(**kw)


def _val(v):
    """Safely get string value."""
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _normalize_conclusion(v):
    s = _val(v).lower()
    return s if s else None


def restructure(input_path, output_path, spec):
    """
    spec dict keys:
      - sheet: str (default "Sheet1")
      - fill_down_columns: list[int]    → parent columns, forward-filled every row
      - group_key_column: int           → column that defines groups (0-based)
      - group_header_columns: list[int] → appear on 1st row of group only
      - fill_down_group_columns: list[int] → same as group-header but filled to ALL rows
      - skip_empty_rows: bool (default True)
      - style: dict (optional override for DEFAULT_STYLES)
      - freeze_panes: str (default "A2")
      - col_widths: dict {int: int} (optional, col_index -> width)

    The algorithm:
      1. Read all rows from the sheet, skip fully empty rows.
      2. Track fill-down columns: carry forward last non-empty value.
      3. Identify groups by group_key_column.
      4. For group-header columns: show value only on group's first row.
      5. For fill-down-group columns: show value on every row.
      6. Apply formatting and write to output.
    """

    # ── options ──────────────────────────────────────────────────────
    sheet_name = spec.get("sheet", "Sheet1")
    fd_cols = spec.get("fill_down_columns", [])
    gk_col = spec.get("group_key_column")
    gh_cols = spec.get("group_header_columns", [])
    fdg_cols = spec.get("fill_down_group_columns", [])
    skip_empty = spec.get("skip_empty_rows", True)
    styles = {**DEFAULT_STYLES, **(spec.get("style", {}))}

    # ── read ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    raw_data = list(ws.iter_rows(min_row=2, values_only=True))

    total_cols = len(header)

    # ── pass 1: forward-fill parent columns, group rows ──────────────
    fd_state = {}
    rows = []
    cur_group_key = None
    cur_group_rows = []

    for r in raw_data:
        # skip empty rows
        if skip_empty and not any(c is not None and _val(c) for c in r):
            continue

        # update fill-down state
        for ci in fd_cols:
            if ci < len(r) and r[ci] is not None and _val(r[ci]):
                fd_state[ci] = r[ci]

        # resolve fill-down values
        row_vals = list(r)
        for ci in fd_cols:
            if ci in fd_state:
                row_vals[ci] = fd_state[ci]

        # determine group
        raw_gk = _val(r[gk_col]) if gk_col is not None and gk_col < len(r) else None
        is_new_group = bool(raw_gk)

        if is_new_group and cur_group_rows:
            rows.append((cur_group_key, cur_group_rows))
            cur_group_rows = []

        if is_new_group:
            cur_group_key = raw_gk

        cur_group_rows.append(row_vals)

    if cur_group_rows:
        rows.append((cur_group_key, cur_group_rows))

    # ── pass 2: collect group-level values (H from first relevant row) ──
    group_values = {}

    for gk, group_rows in rows:
        gv = {}
        # Collect fill-down-group from the first row of the group
        given = {}
        for row_vals in group_rows:
            for ci in fdg_cols:
                v = _val(row_vals[ci]) if ci < len(row_vals) else ""
                if v:
                    given[ci] = row_vals[ci]
                    break  # found H for this case
            if given:
                break

        # Also look for values from the original data before forward-fill
        # (handles case where H is on a step row in original but not on case start row)
        if not given:
            pass

        gv = given
        group_values[gk] = gv

    # ── pass 3: build final rows ─────────────────────────────────────
    output_rows = []

    for gk, group_rows in rows:
        gv = group_values.get(gk, {})

        for idx, row_vals in enumerate(group_rows):
            is_first = (idx == 0)

            new_row = list(row_vals)

            # group-header: only on first row
            for ci in gh_cols:
                if not is_first:
                    new_row[ci] = None

            # fill-down-group: on every row
            for ci in fdg_cols:
                if ci in gv:
                    new_row[ci] = gv[ci]

            output_rows.append(new_row)

    # ── write output workbook ────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = sheet_name

    # Header row
    hdr_font = _make_font(styles["header"]["font"])
    hdr_fill = _make_fill(styles["header"]["fill"])
    for ci, name in enumerate(header):
        cell = ws2.cell(row=1, column=ci + 1, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN

    # Build palette by group for alternating row colors
    palette = {}
    toggle = False
    for gk, _ in rows:
        toggle = not toggle
        palette[gk] = toggle

    write_align = Alignment(vertical="top", wrap_text=True)

    gk_col_index = gk_col

    for ri, row_vals in enumerate(output_rows):
        # Determine grouping key for this row
        row_gk = _val(row_vals[gk_col_index]) if gk_col_index is not None else None
        # Find the group key (may be on first row only - need to detect group)
        # Simpler: use index-based alternation
        if row_gk:
            current_gk = row_gk
        # Use last seen gk

    # Simpler: use position-based alternate
    current_gk = None
    group_idx = -1
    row_gk_map = {}
    for ri, row_vals in enumerate(output_rows):
        raw_gk = _val(row_vals[gk_col_index]) if gk_col_index is not None else ""
        if raw_gk:
            current_gk = raw_gk
        row_gk_map[ri] = current_gk

    # Alternate row colors by group
    toggle = False
    last_gk = None
    for ri in range(len(output_rows)):
        gk = row_gk_map[ri]
        if gk != last_gk:
            toggle = not toggle
            last_gk = gk
        base = "alt_row_1" if toggle else "alt_row_2"
        row_vals = output_rows[ri]

        # Determine if this is a group-start row (has group key visible)
        is_start = bool(_val(row_vals[gk_col_index])) if gk_col_index is not None else False

        for ci in range(len(row_vals)):
            val = row_vals[ci] if ci < len(row_vals) else None
            cell = ws2.cell(row=ri + 2, column=ci + 1, value=val)
            cell.alignment = write_align
            cell.border = THIN

            # Conclusion coloring
            conclusion_val = None
            for fdg_ci in fdg_cols:
                if ci == fdg_ci:
                    conclusion_val = _normalize_conclusion(val)
                    break

            if conclusion_val and conclusion_val in styles:
                st = styles[conclusion_val]
                cell.fill = _make_fill(st.get("fill", {}))
                cell.font = _make_font(st.get("font", {}))
                # Left border accent for fail-like
                if "left_border" in st:
                    cell.border = _make_border(st["left_border"])
            elif is_start and ci == gk_col_index:
                # Group key column on start row
                cell.font = _make_font(styles.get("group_key_cell", {}))
                cell.fill = _make_fill(styles[base]["fill"])
            else:
                cell.fill = _make_fill(styles[base]["fill"])

    # Column widths
    for ci, w in spec.get("col_widths", {}).items():
        col_idx = int(ci)
        ws2.column_dimensions[get_column_letter(col_idx + 1)].width = w

    # Freeze & filter
    freeze = spec.get("freeze_panes", "A2")
    ws2.freeze_panes = freeze
    last_col = get_column_letter(len(header))
    ws2.auto_filter.ref = f"A1:{last_col}{len(output_rows) + 1}"

    wb2.save(output_path)

    # Return summary
    start_count = sum(1 for rv in output_rows
                      if gk_col_index is not None and _val(rv[gk_col_index]))
    return {
        "total_rows": len(output_rows),
        "group_count": start_count,
    }


def main():
    parser = argparse.ArgumentParser(description="Restructure hierarchical Excel tables")
    parser.add_argument("input", help="Input .xlsx file path")
    parser.add_argument("output", help="Output .xlsx file path")
    parser.add_argument("--spec", required=True,
                        help="JSON spec string describing column roles")
    args = parser.parse_args()

    spec = json.loads(args.spec)
    result = restructure(args.input, args.output, spec)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
