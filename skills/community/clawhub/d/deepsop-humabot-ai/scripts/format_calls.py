#!/usr/bin/env python3
"""
format_calls.py
将 Fran 电话任务 collaborationCallResult 返回的 JSON 数据生成 xlsx 文件
用法: python3 format_calls.py '<json_string>' '<output_path>'
输出: xlsx 文件路径
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_LABEL = {
    'Succeeded': '通话成功',
    'Failed':    '呼叫失败',
    'Answer':    '有回复',
}

def parse_job(row):
    """从 describeJobJson 中解析通话详情，返回扁平化字典"""
    result = {}
    raw = row.get('describeJobJson') or ''
    if not raw:
        return result
    try:
        obj = json.loads(raw)
        body = obj.get('body') or {}
        job  = body.get('job') or {}
        tasks = job.get('tasks') or []
        if not tasks:
            return result
        t = tasks[0]
        contact = t.get('contact') or {}
        result['contactName']        = contact.get('contactName', '')
        result['contactPhone']       = contact.get('phoneNumber', '')
        result['calledNumber']       = t.get('calledNumber', '')
        result['callingNumber']      = t.get('callingNumber', '')
        result['duration_s']         = round(t.get('duration', 0) / 1000, 1)
        result['realRingingDuration'] = t.get('realRingingDuration', '')
        result['endReason']          = t.get('endReason', '')
        # 取第一条 Robot 的 script 作为对话摘要
        conv = t.get('conversation') or []
        for c in conv:
            if c.get('speaker') == 'Robot' and c.get('script'):
                result['dialogSummary'] = c['script'][:120]
                break
        else:
            result['dialogSummary'] = ''
    except Exception:
        pass
    return result


def main():
    if len(sys.argv) < 3:
        print("用法: python3 format_calls.py '<json_string>' '<output_path>'", file=sys.stderr)
        sys.exit(1)

    try:
        raw = json.loads(sys.argv[1])
        if isinstance(raw, dict):
            data = raw.get('rows', raw.get('data', []))
        else:
            data = raw
    except json.JSONDecodeError as e:
        print(f'JSON 解析失败: {e}', file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, list) or len(data) == 0:
        print('无通话数据', file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[2]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 表头定义: (显示名, 字段key, 列宽)
    headers = [
        ('序号',         'index',               8),
        ('联系人',        'contactName',        14),
        ('被呼号码',      'contactPhone',       18),
        ('呼出号码',      'callingNumber',      18),
        ('公司名称',      'companyName',        30),
        ('姓名',          'personName',         14),
        ('呼叫状态',      'jobStatus',          14),
        ('任务状态',      'jobTaskStatus',      18),
        ('通话时长(s)',   'duration_s',          12),
        ('振铃时长(s)',   'realRingingDuration', 12),
        ('挂断原因',      'endReason',           16),
        ('对话摘要',      'dialogSummary',       60),
        ('创建时间',      'createTime',          20),
        ('更新时间',      'updateTime',          20),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Fran电话数据'

    # 表头样式（绿色主题区别于 AiWa 蓝色 / Frank 橙色）
    header_font  = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='276749', end_color='276749', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (label, key, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    fill_success = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
    fill_fail    = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
    alt_fill     = PatternFill(start_color='F0FFF4', end_color='F0FFF4', fill_type='solid')
    data_align   = Alignment(vertical='center', wrap_text=False)

    for row_idx, item in enumerate(data, start=2):
        job_info = parse_job(item)
        status   = item.get('jobStatus', '')

        for col_idx, (label, key, width) in enumerate(headers, start=1):
            if key == 'index':
                value = row_idx - 1
            elif key == 'jobStatus':
                value = STATUS_LABEL.get(status, status)
            elif key in job_info:
                value = job_info[key]
            else:
                value = item.get(key)
                if value is None:
                    value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align

            if status == 'Succeeded':
                cell.fill = fill_success
            elif status == 'Failed':
                cell.fill = fill_fail
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(output_path)


if __name__ == '__main__':
    main()
