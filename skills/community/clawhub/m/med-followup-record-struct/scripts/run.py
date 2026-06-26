#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from struct_followup_record import struct_followup_record

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


class PreprocessError(ValueError):
    pass


TEXT_FILE_TYPES = {"txt", "md"}
IMAGE_FILE_TYPES = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
TABLE_FILE_TYPES = {"csv", "xlsx", "xls"}
SUPPORTED_FILE_TYPES = TEXT_FILE_TYPES | IMAGE_FILE_TYPES | TABLE_FILE_TYPES | {"json", "pdf", "doc", "docx"}
SECTION_ORDER = ["主诉", "现病史", "既往史", "婚育史", "月经史", "个人史", "家族史", "查体", "辅助检查", "诊断", "处理"]
SECTION_ALIASES = {
    "主诉": ("主诉", "chiefcomplaint"),
    "现病史": ("现病史", "historyofpresentillness", "hpi"),
    "既往史": ("既往史", "pastmedicalhistory", "pmh"),
    "婚育史": ("婚育史", "婚姻史", "生育史"),
    "月经史": ("月经史", "menstrualhistory"),
    "个人史": ("个人史", "socialhistory"),
    "家族史": ("家族史", "familyhistory"),
    "查体": ("查体", "体格检查", "体检", "physicalexam"),
    "辅助检查": ("辅助检查", "辅检", "检查", "exam", "inspection"),
    "诊断": ("诊断", "初步诊断", "出院诊断", "finaldiagnosis"),
    "处理": ("处理", "处理意见", "治疗建议", "医嘱", "plan"),
}
SECTION_HEADER_NORMALIZED = {
    re.sub(r"[\s_\-]+", "", alias.strip().lower()): section
    for section, aliases in SECTION_ALIASES.items()
    for alias in aliases
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preprocess pdf/doc/docx/xls/xlsx/csv/txt/json input into a follow-up record text, then structure it."
    )
    parser.add_argument("--input", required=True, help="Input file path.")
    parser.add_argument(
        "--input-type",
        default="auto",
        choices=["auto", *sorted(SUPPORTED_FILE_TYPES)],
        help="Explicit input type. Default: auto.",
    )
    parser.add_argument("--sheet", default="", help="Optional sheet name for xlsx/xls inputs.")
    parser.add_argument("--encoding", default="utf-8", help="Text encoding for txt/csv inputs. Default: utf-8.")
    parser.add_argument("--output", default="", help="Output path for structured JSON.")
    parser.add_argument("--timeout", type=int, default=0, help="HTTP request timeout seconds. 0 means wait forever.")
    parser.add_argument("--diag-id", default="skill-diag", help="diag_id used by the backend service.")
    parser.add_argument("--department", default="", help="department (optional)")
    parser.add_argument(
        "--save-prepared",
        action="store_true",
        help="Save the normalized follow-up record text next to the output for debugging.",
    )
    return parser.parse_args()


def detect_input_type(path: Path, explicit: str) -> str:
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def normalize_section_name(value: str) -> Optional[str]:
    return SECTION_HEADER_NORMALIZED.get(normalize_header(value))


def read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def extract_with_soffice(path: Path) -> str:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for office document conversion.")
    with tempfile.TemporaryDirectory(prefix="med-followup-record-") as tmp_dir:
        proc = subprocess.run(
            [
                office_bin,
                "--headless",
                "--convert-to",
                "txt:Text",
                "--outdir",
                tmp_dir,
                str(path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(f"Failed to convert office document to text: {proc.stderr.strip() or proc.stdout.strip()}")
        return out_path.read_text(encoding="utf-8", errors="replace")


def extract_pdf_text(path: Path) -> str:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            text = "\n\n".join(page for page in pages if page)
            if text.strip():
                return text
        except Exception:
            pass

    pdf_to_text = shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout
    raise PreprocessError("Unable to extract text from pdf.")


def extract_image_text(path: Path) -> str:
    tesseract_bin = shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError("tesseract not found for image OCR.")
    langs = detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


def detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


def shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def load_input_artifact(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "text", "text": read_text_file(path, encoding)}
    if input_type == "json":
        return {"kind": "json", "data": read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        return {"kind": "text", "text": extract_with_soffice(path)}
    if input_type == "docx":
        return {"kind": "text", "text": extract_docx_text(path)}
    if input_type == "doc":
        return {"kind": "text", "text": extract_with_soffice(path)}
    if input_type == "pdf":
        return {"kind": "text", "text": extract_pdf_text(path)}
    if input_type in IMAGE_FILE_TYPES:
        return {"kind": "text", "text": extract_image_text(path)}
    raise PreprocessError(f"Unsupported input type: {input_type}")


def looks_like_dialogue_text(text: str) -> bool:
    hits = len(re.findall(r"(患者|医生|病人|医师|大夫)\s*[:：]", text))
    section_hits = sum(1 for marker in SECTION_ORDER if marker in text)
    return hits >= 2 and section_hits < 2


def extract_sections_from_text(text: str) -> Dict[str, str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    sections: Dict[str, List[str]] = {}
    current_section: Optional[str] = None
    pattern = re.compile(r"^\s*([A-Za-z\u4e00-\u9fff]{1,20})\s*[:：]\s*(.*)$")

    for line in lines:
        match = pattern.match(line)
        if match:
            section = normalize_section_name(match.group(1))
            if section:
                current_section = section
                content = match.group(2).strip()
                sections.setdefault(section, [])
                if content:
                    sections[section].append(content)
                continue
        if current_section:
            sections.setdefault(current_section, []).append(line)

    return {section: "\n".join(parts).strip() for section, parts in sections.items() if any(part.strip() for part in parts)}


def assemble_record_text(sections: Dict[str, str]) -> str:
    valid_sections = {name: value.strip() for name, value in sections.items() if value and value.strip()}
    if not valid_sections:
        raise PreprocessError("No recognizable follow-up record sections were found.")
    essential = {"主诉", "现病史", "诊断", "处理"}
    if not any(name in valid_sections for name in essential):
        raise PreprocessError("Missing essential follow-up record sections such as 主诉/现病史/诊断/处理.")
    return "\n".join(f"{name}：{valid_sections[name]}" for name in SECTION_ORDER if name in valid_sections)


def record_text_from_json(data: Any) -> str:
    if isinstance(data, str):
        return record_text_from_text(data)

    if not isinstance(data, dict):
        raise PreprocessError("JSON input must be an object or a plain record string.")

    for key in ("record", "text", "content"):
        value = data.get(key)
        if isinstance(value, str) and value.strip():
            return record_text_from_text(value)

    sections_obj = data.get("sections")
    if isinstance(sections_obj, dict):
        sections = {normalize_section_name(str(key)) or str(key): str(value).strip() for key, value in sections_obj.items() if value}
        normalized = {key: value for key, value in sections.items() if key in SECTION_ORDER and value}
        return assemble_record_text(normalized)

    flat_sections: Dict[str, str] = {}
    for key, value in data.items():
        section = normalize_section_name(str(key))
        if section and isinstance(value, str) and value.strip():
            flat_sections[section] = value.strip()
    if flat_sections:
        return assemble_record_text(flat_sections)

    raise PreprocessError("Unsupported JSON schema for follow-up record structuring.")


def table_to_sections(table: Dict[str, Any]) -> Dict[str, str]:
    rows = table.get("rows") or []
    if not rows:
        return {}

    header_map = {normalize_header(cell): idx for idx, cell in enumerate(rows[0]) if cell.strip()}
    section_idx = first_matching_index(header_map, ("section", "field", "name", "项目", "字段", "sectionname"))
    content_idx = first_matching_index(header_map, ("content", "value", "text", "内容", "值"))
    data_rows = rows[1:] if section_idx is not None and content_idx is not None else rows

    if section_idx is None or content_idx is None:
        section_idx, content_idx = 0, 1

    sections: Dict[str, List[str]] = {}
    for row in data_rows:
        if max(section_idx, content_idx) >= len(row):
            continue
        section = normalize_section_name(row[section_idx])
        content = row[content_idx].strip()
        if section and content:
            sections.setdefault(section, []).append(content)
    return {name: "\n".join(parts).strip() for name, parts in sections.items() if parts}


def first_matching_index(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(normalize_header(alias))
        if idx is not None:
            return idx
    return None


def record_text_from_text(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        raise PreprocessError("Input text is empty.")
    if looks_like_dialogue_text(text):
        raise PreprocessError("Input looks like doctor-patient dialogue rather than a follow-up record.")
    sections = extract_sections_from_text(text)
    return assemble_record_text(sections)


def preprocess_to_record_text(path: Path, input_type: str, encoding: str, sheet_name: str) -> str:
    artifact = load_input_artifact(path, input_type, encoding, sheet_name)
    kind = artifact["kind"]
    if kind == "text":
        return record_text_from_text(artifact["text"])
    if kind == "json":
        return record_text_from_json(artifact["data"])
    if kind == "tables":
        merged: Dict[str, str] = {}
        for table in artifact["tables"]:
            merged.update(table_to_sections(table))
        return assemble_record_text(merged)
    raise PreprocessError(f"Unsupported artifact kind: {kind}")


def build_prepared_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}.prepared.txt")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"✗ Error: Input file not found: {input_path}", file=sys.stderr)
        return 1

    output_path = Path(args.output) if args.output else Path("..") / "runs" / "med-followup-record-struct" / "structured.json"

    try:
        input_type = detect_input_type(input_path, args.input_type)
        record_text = preprocess_to_record_text(input_path, input_type, args.encoding, args.sheet)
    except Exception as exc:
        print(f"✗ Preprocess Error: {exc}", file=sys.stderr)
        return 1

    if args.save_prepared:
        prepared_path = build_prepared_path(output_path)
        prepared_path.parent.mkdir(parents=True, exist_ok=True)
        prepared_path.write_text(record_text, encoding="utf-8")
        print(f"✓ Prepared follow-up record saved to: {prepared_path}")

    temp_path: Optional[Path] = None
    try:
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".txt", delete=False) as tmp:
            tmp.write(record_text)
            temp_path = Path(tmp.name)
        struct_followup_record(
            input_path=str(temp_path),
            output_path=str(output_path),
            timeout=args.timeout,
            diag_id=args.diag_id,
            department=args.department,
        )
        print("\n✓ Follow-up record structured successfully!")
        return 0
    except Exception as exc:
        print(f"✗ Unexpected Error: {exc}", file=sys.stderr)
        return 1
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    sys.exit(main())
