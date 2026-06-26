#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Emby 电视剧整理 Excel 对照表生成脚本

用法（CLI）：
  python generate_excel.py --data <json_data_file> [--output <output_dir>]

  --output 优先级（从高到低）：
    1. CLI 参数 --output 显式指定
    2. JSON 数据文件中的 "output_dir" 字段
    3. 默认值 /vol2/1000/SyncFile/OpenClawSync

用法（Python 调用）：
  generate_excel(records, show_name, output_dir="/vol2/1000/SyncFile/OpenClawSync")
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl，正在安装...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── 字段定义 ──────────────────────────────────────────────────────────
HEADERS = [
    "序号", "电视剧名称", "年份", "原始路径", "原始文件名",
    "新路径", "新文件名", "季号", "集号", "集名称", "文件类型",
    "是否重命名", "是否移动", "处理状态", "说明"
]

COLUMN_WIDTHS = [6, 20, 10, 75, 55, 75, 55, 10, 10, 25, 14, 12, 12, 14, 55]

# 状态颜色映射
STATUS_COLORS = {
    "已整理": "C6EFCE",   # 绿
    "待确认": "FFEB9C",   # 黄
    "跳过":   "D9D9D9",   # 灰
    "失败":   "FFC7CE",   # 红
    "冲突":   "FFBF00",   # 橙
}


DEFAULT_OUTPUT_DIR = "/vol2/1000/SyncFile/OpenClawSync"


def sanitize_filename(name: str) -> str:
    """移除文件名中的非法字符"""
    return re.sub(r'[/\\:*?"<>|]', '_', name)


def generate_excel(records: list, show_name: str, output_dir: str = None) -> str:
    """
    生成整理对照表 Excel 文件。

    参数：
        records:    list[dict]，每个 dict 对应 HEADERS 中的字段（键名同 HEADERS）
        show_name:  电视剧名称（用于文件名）
        output_dir: Excel 保存目录。
                    传入 None 或不传时使用默认目录 DEFAULT_OUTPUT_DIR。

    返回：
        生成的 Excel 文件完整路径
    """
    if not output_dir:
        output_dir = DEFAULT_OUTPUT_DIR
    # 时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = sanitize_filename(show_name)
    filename = f"{safe_name}_{timestamp}.xlsx"
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整理对照表"

    # ── 标题行样式 ────────────────────────────────────────────────────
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # ── 数据行 ────────────────────────────────────────────────────────
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, record in enumerate(records, 2):
        row_data = [record.get(h, "") for h in HEADERS]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20

        status = record.get("处理状态", "")
        fill_color = STATUS_COLORS.get(status)
        row_fill = PatternFill("solid", fgColor=fill_color) if fill_color else None

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    # ── 列宽 ─────────────────────────────────────────────────────────
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 自动筛选 ─────────────────────────────────────────────────────
    full_range = f"A1:{get_column_letter(len(HEADERS))}{len(records) + 1}"
    ws.auto_filter.ref = full_range

    # ── 冻结窗格（冻结首行+首列，方便滚动查看） ─────────────────────
    ws.freeze_panes = "B2"

    # ── 统计汇总 Sheet ────────────────────────────────────────────────
    ws_summary = wb.create_sheet("整理统计")
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 55

    total = len(records)
    renamed = sum(1 for r in records if r.get("是否重命名") == "是")
    moved = sum(1 for r in records if r.get("是否移动") == "是")
    pending = sum(1 for r in records if r.get("处理状态") == "待确认")
    failed = sum(1 for r in records if r.get("处理状态") == "失败")
    conflict = sum(1 for r in records if r.get("处理状态") == "冲突")

    summary_data = [
        ["统计项目", "数值"],
        ["文件总数", total],
        ["已重命名", renamed],
        ["已移动", moved],
        ["待确认", pending],
        ["失败", failed],
        ["冲突", conflict],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["电视剧名称", show_name],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # 统计 Sheet 样式
    for row_idx in range(1, len(summary_data) + 1):
        for col_idx in range(1, 3):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="微软雅黑", size=10, bold=(row_idx == 1))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx == 1:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    ws_summary.row_dimensions[1].height = 24

    wb.save(filepath)
    print(f"[OK] Excel 已生成：{filepath}")
    return filepath


# ── CLI 入口 ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Emby 电视剧整理 Excel 生成工具")
    parser.add_argument("--data", required=True, help="JSON 数据文件路径（包含 records 列表和 show_name）")
    parser.add_argument(
        "--output", default=None,
        help=f"Excel 输出目录（默认：{DEFAULT_OUTPUT_DIR}，也可在 JSON 文件中用 output_dir 字段指定）"
    )
    args = parser.parse_args()

    with open(args.data, encoding="utf-8") as f:
        payload = json.load(f)

    records = payload.get("records", [])
    show_name = payload.get("show_name", "未知电视剧")

    # 优先级：CLI --output > JSON output_dir 字段 > 默认值
    output_dir = args.output or payload.get("output_dir") or DEFAULT_OUTPUT_DIR

    path = generate_excel(records, show_name, output_dir)
    print(path)


if __name__ == "__main__":
    main()
