#!/usr/bin/env python3
"""
生成 iOS MobLink 最小配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def style_header(ws, headers, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.freeze_panes = "A2"


def auto_width(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


def create_basic_info_sheet(wb):
    ws = wb.create_sheet(title="基础信息")
    headers = ["配置项", "说明", "您的填写"]
    style_header(ws, headers, "4472C4")

    rows = [
        ["appKey", "MobTech 后台申请得到的 AppKey，必填", ""],
        ["appSecret", "MobTech 后台申请得到的 AppSecret，必填", ""],
        ["needUniversalLink", "本次是否需要检查或补充 Universal Link / Associated Domains，只填 是 或 否", "否"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, 5):
        ws.cell(row=row, column=1).font = Font(bold=True, color="C00000")

    auto_width(ws, {"A": 30, "B": 78, "C": 22})
    return ws


def create_privacy_sheet(wb):
    ws = wb.create_sheet(title="隐私与场景还原")
    headers = ["事项", "说明"]
    style_header(ws, headers, "ED7D31")

    rows = [
        ["隐私政策", "首次冷启动时应向用户展示《隐私政策》，并在同意后才能调用 MobLink 相关能力。"],
        ["隐私回传", "用户同意后，必须先调用 MobSDK.uploadPrivacyPermissionStatus，再使用 MobLink 能力。"],
        ["严格合规模式", "默认 plist 需配置 MOBNetLater = 2。"],
        ["App Store 隐私标签", "需参考 MobTech 提供的 App Store Connect 后台隐私数据项配置文档进行配置。"],
        ["场景制作", "业务入口、分享按钮、path 和参数不放在最小模板里，由 Agent 扫描工程后逐步确认。"],
        ["场景还原", "恢复代理、控制器映射和路由位置不放在最小模板里，由 Agent 扫描工程后逐步确认。"],
        ["扩展业务主动控制器", "最小模板不收集该配置；如需限制地理位置、IDFA、IDFV、WiFi、IP 等采集，后续按 README 中官方文档单独确认。"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 22, "B": 96})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 40
    return ws


def create_instructions_sheet(wb):
    ws = wb.create_sheet(title="填写说明")
    headers = ["说明项", "详细内容"]
    style_header(ws, headers, "70AD47")

    rows = [
        ["填写顺序", "1. 在 MobTech 后台创建应用并获取 appKey / appSecret；2. 只填写基础信息 Sheet；3. 填完后告诉 Agent “填好了”。"],
        ["最小原则", "不要在表里填写 Bundle ID、Target、Info.plist 路径、入口类名、业务控制器、分享按钮位置、Web 落地页，这些应由 Agent 扫描工程后逐步确认。"],
        ["主动控制器", "不在 Excel 中填写。若后续需要控制位置、IDFA、IDFV、WiFi、IP 等扩展业务数据采集，由 Agent 在 README 中补充官方文档链接并单独确认。"],
        ["Universal Link", "如果后台已经生成 Universal Link 或本次要检查 Associated Domains，可填“是”；具体域名需在后续步骤单独确认。"],
        ["参考链接", "创建应用流程：https://www.mob.com/wiki/detailed?wiki=478&id=34\nMobLink 后台与项目配置：https://www.mob.com/wiki/detailed?wiki=527&id=34\niOS 集成指南：https://www.mob.com/wiki/detailed?wiki=83&id=34\niOS 合规指南：https://www.mob.com/wiki/detailed?wiki=220&id=34\niOS SDK API：https://www.mob.com/wiki/detailed?wiki=553&id=34\nMobLink 扩展业务功能设置：https://www.mob.com/wiki/detailed?wiki=673&id=34\nApp Store Connect 后台隐私数据项配置：https://www.mob.com/wiki/detailed?wiki=574&id=34"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 18, "B": 105})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 60
    return ws


def main():
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    output_path = Path(__file__).resolve().parent / "MobLink_iOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
