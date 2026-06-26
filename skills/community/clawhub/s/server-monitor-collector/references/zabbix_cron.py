#!/usr/bin/env python3
"""
Zabbix 监控报告：采集数据 → XLSX/CSV → 飞书消息 → 邮件
"""
import os, sys, csv, json, smtplib
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.error import URLError
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from collections import defaultdict

HERMES_DIR = os.path.expanduser("~/.hermes")
ZABBIX_URL      = os.environ.get("ZABBIX_URL", "")
ZABBIX_USER     = os.environ.get("ZABBIX_USER", "")
ZABBIX_PASSWORD = os.environ.get("ZABBIX_PASSWORD", "")
ZABBIX_TOKEN    = os.environ.get("ZABBIX_TOKEN", "")
EXCLUDE_GROUPS = {"Templates","Templates/Applications","Templates/Databases",
                  "Templates/Modules","Templates/Network devices",
                  "Templates/Operating systems","Templates/Server hardware",
                  "Templates/Virtualization","Discovered hosts"}
ITEMS_KEY = {
    "memory_avail": "vm.memory.size[available]",
    "memory_total": "vm.memory.size[total]",
    "cpu": "system.cpu.util",
}
CSV_PATH  = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.csv")
XLSX_PATH = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.xlsx")

def api_call(method, params, auth=None):
    payload = {"jsonrpc":"2.0","method":method,"params":params,"id":1}
    if auth: payload["auth"] = auth
    data = json.dumps(payload).encode("utf-8")
    req = Request(ZABBIX_URL, data=data, headers={"Content-Type":"application/json"})
    try:
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except URLError as e:
        print(f"API 请求失败: {e}"); sys.exit(1)
    if "error" in result:
        print(f"API 错误: {result['error']}"); sys.exit(1)
    return result.get("result",[])

def fetch_all(auth):
    groups = api_call("hostgroup.get",{"output":["groupid","name"]}, auth=auth)
    groups = [g for g in groups if g["name"] not in EXCLUDE_GROUPS]
    hosts = api_call("host.get",{
        "output":["hostid","name","host"],
        "groupids":[g["groupid"] for g in groups],
        "selectGroups":["groupid","name"],
    }, auth=auth)
    all_items = []
    for i in range(0, len(hosts), 100):
        batch = [h["hostid"] for h in hosts[i:i+100]]
        items = api_call("item.get",{
            "output":["itemid","hostid","key_","lastvalue"],
            "hostids":batch,
            "filter":{"key_":list(ITEMS_KEY.values())},
        }, auth=auth)
        all_items.extend(items)
    item_map = {(it["hostid"], it["key_"]): it.get("lastvalue","") for it in all_items}
    rows = []
    for host in hosts:
        hid = host["hostid"]
        gnames = [g["name"] for g in host.get("groups",[])]
        valid = [n for n in gnames if n not in EXCLUDE_GROUPS]
        if not valid: continue
        gname = valid[0]
        mem_total = item_map.get((hid, ITEMS_KEY["memory_total"]),"")
        mem_avail = item_map.get((hid, ITEMS_KEY["memory_avail"]),"")
        cpu = item_map.get((hid, ITEMS_KEY["cpu"]),"")
        mt = float(mem_total)/(1024**3) if mem_total else None
        ma = float(mem_avail)/(1024**3) if mem_avail else None
        cp = float(cpu) if cpu else None
        mp = (1 - float(mem_avail)/float(mem_total))*100 if mem_avail and mem_total else None
        rows.append({"group":gname,"name":host["name"],"ip":host["host"],
                     "mem_total_gb":mt,"mem_avail_gb":ma,"mem_used_pct":mp,"cpu_pct":cp})
    return rows

def generate_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    def tb(): s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)
    def hdr(cell, text):
        cell.value=text; cell.font=Font(name="微软雅黑",bold=True,size=10,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="4472C4")
        cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=tb()
    def pct_color(p, bg):
        if p is None: return bg,"000000"
        return ("FF4444","FFFFFF") if p>=80 else ("FFAA44","000000") if p>=60 else ("FFEE88","000000") if p>=40 else (bg,"000000")
    gr = defaultdict(list)
    for r in rows: gr[r["group"]].append(r)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws_ov = wb.create_sheet(title="总览")
    ws_ov.cell(row=1,column=1,value="服务器监控总览").font=Font(name="微软雅黑",bold=True,size=14)
    ws_ov.cell(row=1,column=1).alignment=Alignment(horizontal="left")
    ws_ov.row_dimensions[1].height=24
    ws_ov.cell(row=2,column=1,value=f"采集时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws_ov.cell(row=2,column=1).font=Font(name="微软雅黑",size=10,color="666666")
    ws_ov.cell(row=3,column=1,value=f"共 {len(rows)} 台主机，{len(gr)} 个主机组")
    ws_ov.cell(row=3,column=1).font=Font(name="微软雅黑",size=10,color="666666")
    for ci,h in enumerate(["主机组","主机数","内存告警(≥80%)","CPU告警(≥80%)"],1):
        hdr(ws_ov.cell(row=5,column=ci),h)
    ws_ov.row_dimensions[5].height=20
    for ri,(gn,gd) in enumerate(sorted(gr.items()),start=6):
        ma=sum(1 for r in gd if r["mem_used_pct"] is not None and r["mem_used_pct"]>=80)
        ca=sum(1 for r in gd if r["cpu_pct"] is not None and r["cpu_pct"]>=80)
        for ci,val in enumerate([gn,len(gd),ma,ca],1):
            c=ws_ov.cell(row=ri,column=ci,value=val)
            c.font=Font(name="微软雅黑",size=10); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=tb()
            if ci==3 and ma>0: c.fill=PatternFill("solid",fgColor="FF4444"); c.font=Font(name="微软雅黑",size=10,bold=True,color="FFFFFF")
            elif ci==4 and ca>0: c.fill=PatternFill("solid",fgColor="FF4444"); c.font=Font(name="微软雅黑",size=10,bold=True,color="FFFFFF")
    for ci,w in enumerate([24,10,16,16],1): ws_ov.column_dimensions[get_column_letter(ci)].width=w
    cols=[("主机名",32),("IP",18),("内存总量(GB)",14),("内存可用(GB)",14),("内存占用率(%)",14),("CPU占用率(%)",13)]
    for gn,gd in sorted(gr.items()):
        ws=wb.create_sheet(title=gn[:31]); ws.row_dimensions[1].height=20
        for ci,(ht,_) in enumerate(cols,1): hdr(ws.cell(row=1,column=ci),ht)
        gd.sort(key=lambda x:(-(x["mem_used_pct"] or 0),-(x["cpu_pct"] or 0)))
        for ri,r in enumerate(gd,start=2):
            bg="EEF2FF" if ri%2==0 else "FFFFFF"
            mb,mc=pct_color(r.get("mem_used_pct"),bg); cb,cc=pct_color(r.get("cpu_pct"),bg)
            for ci,(val,cbg,cfc,fmt) in enumerate([
                (r["name"],bg,"000000",None),(r["ip"],bg,"000000",None),
                (r["mem_total_gb"],bg,"000000","0.0"),(r["mem_avail_gb"],bg,"000000","0.0"),
                (r["mem_used_pct"],mb,mc,"0.0"),(r["cpu_pct"],cb,cc,"0.0"),
            ],1):
                c=ws.cell(row=ri,column=ci)
                if val is None: c.value="N/A"
                else:
                    c.value=val
                    if fmt: c.number_format=fmt
                c.font=Font(name="微软雅黑",size=10,color=cfc)
                c.fill=PatternFill("solid",fgColor=cbg)
                c.alignment=Alignment(horizontal="center",vertical="center"); c.border=tb()
        for ci,(_,w) in enumerate(cols,1): ws.column_dimensions[get_column_letter(ci)].width=w
        ws.freeze_panes="A2"
    wb.save(XLSX_PATH); print(f"XLSX: {XLSX_PATH}")

def generate_csv(rows):
    os.makedirs(os.path.dirname(CSV_PATH),exist_ok=True)
    with open(CSV_PATH,"w",newline="",encoding="utf-8-sig") as f:
        w=csv.writer(f); w.writerow(["主机组","主机名","IP","内存总量(GB)","内存可用(GB)","内存占用率(%)","CPU占用率(%)"])
        for r in rows:
            w.writerow([r["group"],r["name"],r["ip"],
                f"{r['mem_total_gb']:.1f}" if r['mem_total_gb'] is not None else "N/A",
                f"{r['mem_avail_gb']:.1f}" if r['mem_avail_gb'] is not None else "N/A",
                f"{r['mem_used_pct']:.1f}" if r['mem_used_pct'] is not None else "N/A",
                f"{r['cpu_pct']:.1f}" if r['cpu_pct'] is not None else "N/A",
            ])
    print(f"CSV: {CSV_PATH}")

def build_feishu_summary(rows):
    gr=defaultdict(list)
    for r in rows: gr[r["group"]].append(r)
    warn=[r for r in rows if (r["mem_used_pct"] or 0)>=60 or (r["cpu_pct"] or 0)>=60]
    warn.sort(key=lambda x:(-(x["mem_used_pct"] or 0),-(x["cpu_pct"] or 0)))
    lines=[f"## 服务器监控报告","",
           f"**采集时间**：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
           f"共 **{len(rows)}** 台主机，覆盖 **{len(gr)}** 个主机组",""]
    if warn:
        lines+=["### ⚠ 重点关注（内存占用≥60% 或 CPU≥60%）",""]
        lines+=["| 主机名 | 主机组 | 内存占用率(%) | CPU占用率(%) |","|---|---|---|---|"]
        for r in warn[:20]: lines.append(f"| {r['name']} | {r['group']} | {r['mem_used_pct']:.1f} | {r['cpu_pct']:.1f} |")
        if len(warn)>20: lines.append(f"...（共 {len(warn)} 台，详见附件）")
    else:
        lines+=["### ✅ 全部正常（无告警主机）",""]
    lines+=["",f"完整数据：`{CSV_PATH}`"]
    return "\n".join(lines)

def load_env():
    p = os.path.join(HERMES_DIR, ".env")
    if os.path.exists(p):
        with open(p) as f:
            for line in f:
                line=line.strip()
                if "=" in line and not line.startswith("#"):
                    k,v=line.split("=",1); os.environ[k]=v.strip()

def send_email(subject, html_body, attachments=None):
    load_env()
    host=os.environ.get("SMTP_HOST",""); port=os.environ.get("SMTP_PORT","465")
    sender=os.environ.get("SMTP_FROM",""); token=os.environ.get("SMTP_TOKEN","")
    target=os.environ.get("TARGET_EMAIL","")
    if not all([host,sender,token,target]): print("邮件配置不完整，跳过"); return
    msg=MIMEMultipart(); msg["From"]=sender; msg["To"]=target; msg["Subject"]=subject
    msg.attach(MIMEText(html_body,"html","utf-8"))
    for fpath in (attachments or []):
        if os.path.exists(fpath):
            with open(fpath,"rb") as f:
                part=MIMEBase("application","octet-stream"); part.set_payload(f.read())
                encoders.encode_base64(part)
                part["Content-Disposition"]=f"attachment; filename={os.path.basename(fpath)}"
                msg.attach(part)
    try:
        if port=="465":
            with smtplib.SMTP_SSL(host,int(port)) as s: s.login(sender,token); s.sendmail(sender,target,msg.as_string())
        else:
            with smtplib.SMTP(host,int(port)) as s: s.starttls(); s.login(sender,token); s.sendmail(sender,target,msg.as_string())
        print(f"邮件已发送: {target}")
    except Exception as e: print(f"邮件发送失败: {e}")

def build_html_body(rows):
    gr=defaultdict(list)
    for r in rows: gr[r["group"]].append(r)
    html=f"<html><body><h2>服务器监控报告</h2><p><b>采集时间：</b>{datetime.now().strftime('%Y-%m-%d %H:%M')}</p><p><b>共 {len(rows)} 台，{len(gr)} 组</b></p>"
    for gn,gd in sorted(gr.items()):
        html+=f"<h3>{gn} ({len(gd)} 台)</h3>"
        html+="<table border='1' cellpadding='4' cellspacing='0' style='border-collapse:collapse;font-size:12px;'>"
        html+="<tr bgcolor='#4472C4' style='color:white;'><th>主机名</th><th>IP</th><th>内存总量(GB)</th><th>内存可用(GB)</th><th>内存占用率(%)</th><th>CPU占用率(%)</th></tr>"
        for i,r in enumerate(gd):
            bg="#EEF2FF" if i%2==0 else "#FFFFFF"
            mp=r["mem_used_pct"] or 0; cp=r["cpu_pct"] or 0
            ms=("background:#FF4444;color:white;" if mp>=80 else "background:#FFAA44;" if mp>=60 else "background:#FFEE88;" if mp>=40 else "")
            cs=("background:#FF4444;color:white;" if cp>=80 else "background:#FFAA44;" if cp>=60 else "background:#FFEE88;" if cp>=40 else "")
            html+=f"<tr bgcolor='{bg}'><td>{r['name']}</td><td>{r['ip']}</td>"
            html+=f"<td>{r['mem_total_gb']:.1f}</td>" if r['mem_total_gb'] else "<td>N/A</td>"
            html+=f"<td>{r['mem_avail_gb']:.1f}</td>" if r['mem_avail_gb'] else "<td>N/A</td>"
            html+=f"<td style='{ms}'>{r['mem_used_pct']:.1f}</td>" if r['mem_used_pct'] else "<td>N/A</td>"
            html+=f"<td style='{cs}'>{r['cpu_pct']:.1f}</td>" if r['cpu_pct'] else "<td>N/A</td></tr>"
        html+="</table><br/>"
    html+="</body></html>"
    return html

def main():
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始巡检...")
    auth=api_call("user.login",{"user":ZABBIX_USER,"password":ZABBIX_PASSWORD})
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 登录成功")
    rows=fetch_all(auth)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 采集完成: {len(rows)} 台")
    generate_csv(rows); generate_xlsx(rows)
    summary=build_feishu_summary(rows)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 飞书摘要:\n{summary[:500]}")
    subject=f"【监控报告】服务器巡检 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    atts=[f for f in [XLSX_PATH,CSV_PATH] if os.path.exists(f)]
    send_email(subject, build_html_body(rows), atts)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 全部完成!")

if __name__=="__main__": main()
