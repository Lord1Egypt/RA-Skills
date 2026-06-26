#!/usr/bin/env python3
"""Append rows to table (XLSX or CSV) preserving existing content and format."""

import sys
import json
import csv
from pathlib import Path
from difflib import SequenceMatcher


def similar(a, b):
    """Calculate string similarity ratio."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()


def find_duplicate(xlsx_path, title, threshold=0.85):
    """Check if a paper with similar title already exists."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # Find title column
        headers = [cell.value for cell in ws[1]]
        title_col = None
        for i, h in enumerate(headers):
            if h and ('title' in str(h).lower() or '标题' in str(h)):
                title_col = i + 1
                break
        
        if not title_col:
            return False, None
        
        # Check all existing titles
        for row in range(2, ws.max_row + 1):
            existing_title = ws.cell(row=row, column=title_col).value
            if existing_title and similar(existing_title, title) > threshold:
                return True, existing_title
        
        return False, None
    except Exception:
        return False, None


def find_duplicate_csv(csv_path, title, threshold=0.85):
    """Check if a paper with similar title already exists in CSV."""
    encodings = ['utf-8', 'gbk', 'latin-1']
    
    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                
                # Find title column
                title_col = None
                for i, h in enumerate(headers):
                    if h and ('title' in str(h).lower() or '标题' in str(h)):
                        title_col = i
                        break
                
                if title_col is None:
                    return False, None
                
                # Check all existing titles
                for row in reader:
                    if len(row) > title_col:
                        existing_title = row[title_col]
                        if existing_title and similar(existing_title, title) > threshold:
                            return True, existing_title
            
            return False, None
        except UnicodeDecodeError:
            continue
        except Exception:
            return False, None
    
    return False, None


def validate_data(row_data, header_metadata=None):
    """Validate extracted data before writing."""
    errors = []
    warnings = []
    
    # Check for required fields
    if not row_data:
        errors.append("Empty data")
        return False, errors, warnings
    
    # Check for hallucination indicators
    hallucination_indicators = [
        '可能', 'probably', 'maybe', 'perhaps', '似乎', 'seems',
        '根据常识', 'according to common sense', '通常', 'usually'
    ]
    
    for key, value in row_data.items():
        if isinstance(value, dict):
            val_str = str(value.get('value', ''))
        else:
            val_str = str(value)
        
        # Check for hallucination language
        for indicator in hallucination_indicators:
            if indicator in val_str.lower():
                warnings.append(f"Field '{key}' may contain inference: '{val_str[:50]}...'")
        
        # Check for suspicious patterns
        if val_str.startswith('根据') or val_str.startswith('Based on'):
            warnings.append(f"Field '{key}' starts with inference language")
    
    return len(errors) == 0, errors, warnings


def append_xlsx(table_path, row_data, validate=True):
    """Append row to XLSX preserving format."""
    try:
        import openpyxl
        
        # Validate if requested
        if validate:
            valid, errors, warnings = validate_data(row_data)
            if not valid:
                print(f"[ERROR: Validation failed: {'; '.join(errors)}]", file=sys.stderr)
                return False
            if warnings:
                print(f"[WARNING: {'; '.join(warnings)}]", file=sys.stderr)
        
        wb = openpyxl.load_workbook(table_path)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Check for duplicates if title exists
        if isinstance(row_data, dict):
            title_value = row_data.get('Title') or row_data.get('标题') or row_data.get('title')
            if title_value:
                is_dup, dup_title = find_duplicate(table_path, title_value)
                if is_dup:
                    print(f"[WARNING: Duplicate detected - similar to '{dup_title}']", file=sys.stderr)
                    return False
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Write data matching headers
        for col_idx, header in enumerate(headers, 1):
            if header:
                # Handle nested dict (with confidence) or direct value
                if header in row_data:
                    value = row_data[header]
                    if isinstance(value, dict):
                        ws.cell(row=next_row, column=col_idx, value=value.get('value', ''))
                    else:
                        ws.cell(row=next_row, column=col_idx, value=value)
        
        wb.save(table_path)
        return True
    except Exception as e:
        print(f"[ERROR: Failed to write XLSX: {e}]", file=sys.stderr)
        return False


def append_csv(table_path, row_data, validate=True):
    """Append row to CSV."""
    encodings = ['utf-8', 'gbk', 'latin-1']
    
    # Validate if requested
    if validate:
        valid, errors, warnings = validate_data(row_data)
        if not valid:
            print(f"[ERROR: Validation failed: {'; '.join(errors)}]", file=sys.stderr)
            return False
        if warnings:
            print(f"[WARNING: {'; '.join(warnings)}]", file=sys.stderr)
    
    for encoding in encodings:
        try:
            # Read headers
            with open(table_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
            
            # Check for duplicates if title exists
            if isinstance(row_data, dict):
                title_value = row_data.get('Title') or row_data.get('标题') or row_data.get('title')
                if title_value:
                    is_dup, dup_title = find_duplicate_csv(table_path, title_value)
                    if is_dup:
                        print(f"[WARNING: Duplicate detected - similar to '{dup_title}']", file=sys.stderr)
                        return False
            
            # Append row
            with open(table_path, 'a', encoding=encoding, newline='') as f:
                writer = csv.writer(f)
                row = []
                for h in headers:
                    h_stripped = h.strip() if h else ''
                    if h_stripped in row_data:
                        value = row_data[h_stripped]
                        if isinstance(value, dict):
                            row.append(value.get('value', ''))
                        else:
                            row.append(value)
                    else:
                        row.append('')
                writer.writerow(row)
            return True
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"[ERROR: Failed to write CSV: {e}]", file=sys.stderr)
            return False
    
    return False


def write_row(table_path, row_data, validate=True):
    """Route to appropriate writer based on file extension."""
    path = Path(table_path)
    ext = path.suffix.lower()
    
    if ext == '.xlsx':
        return append_xlsx(path, row_data, validate=validate)
    elif ext == '.csv':
        return append_csv(path, row_data, validate=validate)
    else:
        print(f"[ERROR: Unsupported table format: {ext}]", file=sys.stderr)
        return False


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: write_table.py <table_path> '<json_data>' [--validate]", file=sys.stderr)
        sys.exit(1)
    
    table_path = sys.argv[1]
    row_data = json.loads(sys.argv[2])
    validate = '--validate' in sys.argv
    
    success = write_row(table_path, row_data, validate=validate)
    sys.exit(0 if success else 1)
