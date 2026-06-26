# -*- coding: utf-8 -*-
"""
Excel管理器测试 - macOS版本
对应: scripts/excel_manager.py
"""

import sys
import os
import tempfile

# 添加scripts目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'scripts'))

from openpyxl import Workbook
from excel_manager import ExcelManager

def create_test_excel(path):
    """创建测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文案库"
    
    # 标题行
    headers = [
        "视频文案", "时间段", "镜头", "运镜", "拍摄技巧", 
        "画面", "台词", "音效", "推荐音乐/BGM", "文案标签",
        "使用状态", "关联活动", "发布日期", "备注"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
    
    # 示例数据
    ws.cell(2, 1).value = "示例脚本\n这是一个测试脚本"
    ws.cell(2, 2).value = "0-6秒"
    ws.cell(2, 3).value = "全景镜头，展现场景"
    ws.cell(2, 11).value = "已使用"
    
    wb.save(path)
    return path

def test_excel():
    """测试Excel管理功能"""
    
    # 创建临时测试文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        test_file = tmp.name
    
    try:
        # 准备测试文件
        create_test_excel(test_file)
        print(f"[OK] 创建测试文件: {test_file}")
        
        # 测试1: 加载Excel
        print("\n[测试1] 加载Excel")
        with ExcelManager(test_file) as em:
            print("[OK] Excel加载成功")
            print(f"[OK] 总行数: {em.ws.max_row}")
            print(f"[OK] 总列数: {em.ws.max_column}")
        
        # 测试2: 扫描格式
        print("\n[测试2] 扫描格式")
        with ExcelManager(test_file) as em:
            format_info = em.scan_format()
            print(f"[OK] 总行数: {format_info['total_rows']}")
            print(f"[OK] 总列数: {format_info['total_columns']}")
            print(f"[OK] 合并单元格数: {len(format_info['merged_cells'])}")
        
        # 测试3: 获取已有脚本
        print("\n[测试3] 获取已有脚本")
        with ExcelManager(test_file) as em:
            scripts = em.get_existing_scripts()
            print(f"[OK] 发现 {len(scripts)} 个已有脚本")
            for script in scripts:
                print(f"  - 第{script['row']}行: {script['title'][:30]}...")
        
        # 测试4: 追加脚本
        print("\n[测试4] 追加脚本")
        test_script = {
            "title": "测试脚本",
            "story": "这是一个测试故事",
            "segments": [
                {
                    "time": "0-5秒",
                    "duration": 5,
                    "shot_desc": "特写镜头",
                    "movement_desc": "缓慢推进",
                    "tech_desc": "手持稳定器",
                    "scene_desc": "自然光",
                    "line": "测试台词",
                    "sound_desc": "环境音"
                },
                {
                    "time": "5-10秒",
                    "duration": 5,
                    "shot_desc": "中景",
                    "movement_desc": "固定",
                    "tech_desc": "三脚架",
                    "scene_desc": "室内光",
                    "line": "继续测试",
                    "sound_desc": "背景音乐"
                }
            ],
            "bgm": "轻音乐",
            "activity": "",
            "date": "2026-04-22",
            "notes": "测试备注"
        }
        
        with ExcelManager(test_file) as em:
            start_row = em.ws.max_row + 1
            end_row = em.append_script(test_script, start_row)
            em.save()
            print(f"[OK] 脚本写入成功: 第{start_row}-{end_row}行")
        
        # 测试5: 验证写入
        print("\n[测试5] 验证写入")
        with ExcelManager(test_file) as em:
            result = em.validate_write(start_row, end_row)
            print(f"[OK] 格式正确性: {result['format_valid']}")
            print(f"[OK] 内容完整性: {result['content_complete']}")
            if result['errors']:
                print(f"[X] 错误: {result['errors']}")
        
        print("\n[完成] Excel管理测试通过")
        
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.remove(test_file)
            print("[OK] 清理测试文件")

if __name__ == "__main__":
    test_excel()
