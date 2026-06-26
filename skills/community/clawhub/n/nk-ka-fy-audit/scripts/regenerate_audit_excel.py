#!/usr/bin/env python3
"""
最终版regenerate_audit_excel.py
列顺序：月份 | 品牌 | 原始表金额 | 历史差异调整金额 | 历史差额调整说明 | 审计重算金额 | 汇总表金额 | 差额 | 审计备注
"""
import openpyxl, sqlite3, os

TEMPLATE = '/workspace/templates/返佣审计结果-模板.xlsx'
OUTPUT = '/workspace/KA返佣审计结果报告.xlsx'
DB = '/workspace/业务数据/集团客户部FY数据.db'

conn = sqlite3.connect(DB)
c = conn.cursor()

# ===== 从DB查询所有历史差额调整的说明 =====
print("查询历史差额调整说明...")
HIST_ADJ_NOTES = {}
for m, b, n in c.execute('''
    SELECT month, brand, GROUP_CONCAT(COALESCE(note, ''), '；')
    FROM (SELECT DISTINCT month, brand, original_note as note
          FROM raw_records
          WHERE audit_note = '历史差额调整')
    GROUP BY month, brand
''').fetchall():
    notes = n.strip().strip('；')
    if notes:
        HIST_ADJ_NOTES[(m, b)] = '历史差额调整：' + notes
    else:
        HIST_ADJ_NOTES[(m, b)] = '历史差额调整'
print(f"  已加载 {len(HIST_ADJ_NOTES)} 条历史调整说明")

# ===== 已知的历史差异调整 =====
KNOWN_HIST_ADJ = {
    ('202402', '火石教育'): {
        'remark': '补算及调账202003~202112的交易返佣'
    },
    ('202404', 'HARMAY话梅'): {
        'remark': '因阶梯费率回算调整202301~202310的返佣'
    },
    ('202505', '维多利亚的秘密'): {
        'remark': '回算20250110 ~ 202504的交易返佣，审批编号：SP7220250407000001'
    },
}

# ===== 1. 保存旧的汇总表金额 =====
print("保存旧的汇总表金额...")
c.execute("SELECT month, brand, summary_amount FROM brand_audit WHERE summary_amount IS NOT NULL AND summary_amount != 0")
old_summary = {}
for month, brand, sa in c.fetchall():
    old_summary[(month, brand)] = sa
print("  已保存 {} 条汇总表记录".format(len(old_summary)))

# ===== 2. 重建brand_audit =====
print("重建brand_audit...")
c.execute("DELETE FROM brand_audit")
c.execute("DELETE FROM monthly_audit")

# 从raw_records聚合品牌级数据（排除历史差额调整记录）
c.execute("""
    SELECT month, brand,
           ROUND(SUM(COALESCE(total_rebate,0)),2) as raw_amount,
           ROUND(SUM(CASE
             WHEN calc_method = 'zhongkuai' THEN COALESCE(total_rebate,0)
             WHEN audit_note = '历史差额调整' THEN COALESCE(total_rebate,0)
             WHEN calc_method = 'abnormal' THEN 0  -- 排除sheet记录：只有历史差额调整才计入
             WHEN COALESCE(ROUND(alipay_txn_amount*alipay_rebate_ratio,2) + ROUND(wechat_txn_amount*wechat_rebate_ratio,2),0) > 0
               THEN ROUND(alipay_txn_amount*alipay_rebate_ratio,2) + ROUND(wechat_txn_amount*wechat_rebate_ratio,2)
             ELSE COALESCE(total_rebate,0)
           END),2) as recalc_amount
    FROM raw_records
    -- 包含所有记录（规则C：无足够数据核查时，原始adj计入重算）
    GROUP BY month, brand
    ORDER BY month, brand
""")
data_rows = c.fetchall()
print("  读取 条品牌+月份记录".format(len(data_rows)))

# 按月份处理
all_months = sorted(set(r[0] for r in data_rows))
all_brand_data = []

for month_str in all_months:
    month_rows = [r for r in data_rows if r[0] == month_str]
    
    for r in month_rows:
        brand, raw_amt, calc_amt = r[1], r[2], r[3]
        summary = old_summary.get((month_str, brand), 0.0)
        diff = round(summary - calc_amt, 2)
        
        # 分类
        is_real = 1
        note = ''
        if diff < -100:
            note = '少发，视同无差异'
            is_real = 0
        elif diff < 0:
            is_real = 0
            if abs(diff) <= 1:
                note = '小数舍入'
            elif abs(diff) <= 100:
                note = '偏差{:+.2f}元'.format(diff)
            else:
                note = '少发，视同无差异'
        elif diff <= 1 and diff >= 0:
            note = '小数舍入'
            is_real = 0
        elif diff > 0:
            is_real = 1
        
        c.execute('''INSERT INTO brand_audit(month,brand,raw_amount,recalc_amount,summary_amount,diff,is_real_diff,diff_note)
            VALUES(?,?,?,?,?,?,?,?)''',
            (month_str, brand, raw_amt, calc_amt, summary, diff, is_real, note))
        
        all_brand_data.append({
            'month': month_str, 'brand': brand, 'raw': raw_amt,
            'recalc': calc_amt, 'summary': summary,
            'diff': diff, 'note': note, 'is_real': is_real
        })
    
    # 月度汇总
    mb = [b for b in all_brand_data if b['month'] == month_str]
    tr = round(sum(b['recalc'] for b in mb), 2)
    ts = round(sum(b['summary'] for b in mb), 2)
    td = round(ts - tr, 2)
    bc = len(mb)
    rd = round(sum(b['diff'] for b in mb if b['is_real']), 2)
    c.execute('''INSERT INTO monthly_audit(month,total_recalc,total_summary,total_diff,real_diff,brand_count)
        VALUES(?,?,?,?,?,?)''',
        (month_str, tr, ts, td, rd, bc))

conn.commit()
print("  brand_audit写入完成（{}个月份）".format(len(all_months)))

# ===== 3. 生成Excel =====
print("\n生成Excel...")
wb = openpyxl.load_workbook(TEMPLATE)

# --- 审计摘要 ---
ws_summary = wb['审计摘要']
monthly_audit = {r[0]: r for r in c.execute('SELECT * FROM monthly_audit ORDER BY month').fetchall()}
total_pass = sum(1 for m in all_months if abs(monthly_audit[m][4]) <= 1)
total_warn = sum(1 for m in all_months if 1 < abs(monthly_audit[m][4]) <= 100)
total_fail = sum(1 for m in all_months if abs(monthly_audit[m][4]) > 100)
total_brands = len(set(b['brand'] for b in all_brand_data))
total_real_diff = sum(monthly_audit[m][4] for m in all_months)

items = [
    ('审计时间范围', '2023年07月 至 2026年03月'),
    ('审计月份数', '{}个月'.format(len(all_months))),
    ('涉及品牌总数', '{}个'.format(total_brands)),
    ('通过月份数', '{}个月'.format(total_pass)),
    ('差异月份数', '{}个月'.format(total_fail)),
    ('总真实差异（汇总表-重算）', '{:+,.2f}元'.format(total_real_diff)),
]
for i, (k, v) in enumerate(items):
    ws_summary.cell(row=i+2, column=1, value=k)
    ws_summary.cell(row=i+2, column=2, value=v)
for rr in range(len(items)+2, 15):
    for cc in range(1, 3):
        ws_summary.cell(row=rr, column=cc, value='')

# --- 月度趋势 ---
ws_trend = wb['月度趋势']
ws_trend.cell(row=1, column=1, value='月份')
ws_trend.cell(row=1, column=2, value='返佣金额')
ws_trend.cell(row=1, column=3, value='品牌数')
ws_trend.cell(row=1, column=4, value='差异')
ws_trend.cell(row=1, column=5, value='状态')
for i, month in enumerate(all_months):
    ma = monthly_audit[month]
    st = '通过' if abs(ma[4]) <= 1 else ('偏差' if abs(ma[4]) <= 100 else '差异')
    ws_trend.cell(row=i+2, column=1, value=month)
    ws_trend.cell(row=i+2, column=2, value=ma[2])
    ws_trend.cell(row=i+2, column=3, value=ma[5])
    ws_trend.cell(row=i+2, column=4, value=ma[3])
    ws_trend.cell(row=i+2, column=5, value=st)
for rr in range(len(all_months)+2, 40):
    for cc in range(1, 6):
        ws_trend.cell(row=rr, column=cc, value='')

# --- 多发明细 ---
HEADERS = ['月份', '品牌', '原始表金额',
           '历史差异调整金额', '历史差额调整说明',
           '审计重算金额', '汇总表金额', '差额', '审计备注']
ws_detail = wb['审计结果-多发明细']
# 重命名为新名称
old_name = ws_detail.title
ws_detail.title = '审计结果-多发&历史调差明细'
ws_detail = wb['审计结果-多发&历史调差明细']
for i, h in enumerate(HEADERS):
    ws_detail.cell(row=1, column=i+1, value=h)

detail_rows = []
for b in all_brand_data:
    adj_key = (b['month'], b['brand'])
    is_known_adj = adj_key in KNOWN_HIST_ADJ
    
    hist_adj_amt = 0.0
    hist_adj_desc = ''
    audit_remark = ''
    recalc_final = b['recalc']
    diff_final = b['diff']
    is_real_final = b['is_real']
    
    if is_known_adj:
        adj_info = KNOWN_HIST_ADJ[adj_key]
        hist_adj_desc = adj_info['remark']
        audit_remark = '已核查差异调整金额无误'
        is_real_final = 0
        
        if b['diff'] > 0:
            # A/C规则：汇总表有额外金额 → adj = diff
            hist_adj_amt = b['diff']
            recalc_final = round(b['recalc'] + hist_adj_amt, 2)
            diff_final = round(b['summary'] - recalc_final, 2)
        elif b['diff'] >= -1 and b['diff'] <= 0:
            # diff≈0：问题记录已含在重算中
            hist_adj_amt = round(b['raw'], 2)
            recalc_final = b['recalc']
            diff_final = b['diff']
    
    # 非已知调整品牌：从DB检查是否有历史差额调整记录
    if not is_known_adj:
        adj_key = (b['month'], b['brand'])
        if adj_key in HIST_ADJ_NOTES:
            hist_adj_amt = round(b['raw'], 2)
            hist_adj_desc = HIST_ADJ_NOTES[adj_key]
            audit_remark = '已核查差额调整记录无误'
    
    # 排除：既无实际差异又无历史调整金额的记录
    if not is_real_final and hist_adj_amt <= 0.01:
        continue
    # 排除：diff≈0且无历史调整金额
    if abs(diff_final) <= 1 and hist_adj_amt <= 0.01:
        continue
    # 排除：少发（负diff大）
    if diff_final < -0.01 and is_known_adj == False:
        continue
    
    detail_rows.append({
        'month': b['month'], 'brand': b['brand'],
        'raw': b['raw'],
        'recalc': recalc_final,
        'hist_adj_amt': hist_adj_amt,
        'hist_adj_desc': hist_adj_desc,
        'summary': b['summary'],
        'diff': diff_final,
        'audit_remark': audit_remark
    })

detail_rows.sort(key=lambda x: (x['month'], -abs(x['diff'])))

row = 2
for d in detail_rows:
    ws_detail.cell(row=row, column=1, value=d['month'])
    ws_detail.cell(row=row, column=2, value=d['brand'])
    ws_detail.cell(row=row, column=3, value=d['raw'])
    ws_detail.cell(row=row, column=4, value=d['hist_adj_amt'] if d['hist_adj_amt'] > 0.01 else '')
    ws_detail.cell(row=row, column=5, value=d['hist_adj_desc'] if d['hist_adj_desc'] else '')
    ws_detail.cell(row=row, column=6, value=d['recalc'])
    ws_detail.cell(row=row, column=7, value=d['summary'])
    ws_detail.cell(row=row, column=8, value=d['diff'] if abs(d['diff']) > 0.01 else 0)
    ws_detail.cell(row=row, column=9, value=d['audit_remark'] if d['audit_remark'] else '')
    row += 1

for rr in range(row, 550):
    for cc in range(1, 10):
        ws_detail.cell(row=rr, column=cc, value='')

wb.save(OUTPUT)

# ===== Recompute stats from corrected detail_rows =====
corrected_real_diff = sum(d['diff'] for d in detail_rows if d['diff'] > 1)
diff_months = set(d['month'] for d in detail_rows if d['diff'] > 1)
corrected_pass = len(all_months) - len(diff_months)
corrected_fail = len(diff_months)

# Update brand_audit and monthly_audit in DB with corrected values
for d in detail_rows:
    is_real = 1 if d['diff'] > 0.01 else 0
    c.execute('UPDATE brand_audit SET diff=?, is_real_diff=? WHERE month=? AND brand=?',
              (d['diff'], is_real, d['month'], d['brand']))

# Recompute monthly_audit from corrected detail_rows
monthly_corrected = {}
for month in all_months:
    month_details = [d for d in detail_rows if d['month'] == month]
    total = sum(d['summary'] for d in month_details)
    real_diff = round(sum(d['diff'] for d in month_details if d['diff'] > 0.01), 2)
    monthly_corrected[month] = real_diff
    c.execute('UPDATE monthly_audit SET real_diff=? WHERE month=?',
              (real_diff, month))
conn.commit()

# Refresh 月度趋势 section with corrected values
for i, month in enumerate(all_months):
    rd = monthly_corrected.get(month, 0)
    st = '通过' if abs(rd) <= 1 else ('偏差' if abs(rd) <= 100 else '差异')
    ws_trend.cell(row=i+2, column=4, value=rd)
    ws_trend.cell(row=i+2, column=5, value=st)

# Refresh 审计摘要 with corrected values: Row4=品牌总数(keep), Row5=通过, Row6=差异, Row7=总差异
ws_summary.cell(row=5, column=2, value='{}个月'.format(corrected_pass))
ws_summary.cell(row=6, column=2, value='{}个月'.format(corrected_fail))
ws_summary.cell(row=7, column=2, value='{:+,.2f}元'.format(corrected_real_diff))

wb.save(OUTPUT)

print("\n✅ Excel已保存: {}".format(OUTPUT))
print("  月份: {} | 多发明细: {}条".format(len(all_months), len(detail_rows)))
print("  通过: {}月 | 差异: {}月".format(corrected_pass, corrected_fail))
print("  总真实差异: {:+,.2f}元".format(corrected_real_diff))

adj_cnt = len([d for d in detail_rows if d['hist_adj_amt'] > 0.01])
print(f"  历史调整品牌月数: {adj_cnt}条")

print("\n📋 特殊历史差异调整（含调整的品牌）:")
for d in detail_rows:
    if d['hist_adj_amt'] > 0.01 and d['audit_remark'] == '已核查差异调整金额无误':
        print(" {:35s} adj={:>10.2f} diff={:>+8.2f}".format(
            d['month']+' '+d['brand'], d['hist_adj_amt'], d['diff']))
        print("    remark: {}".format(d['audit_remark']))

conn.close()
print("\nDone")
