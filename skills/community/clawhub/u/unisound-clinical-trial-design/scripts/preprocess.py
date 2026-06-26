#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""通用文件预处理工具 — 从 pdf/doc/docx/xls/xlsx/csv/txt/json/图片 提取文本或表格数据。"""

import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


class PreprocessError(ValueError):
    pass


TEXT_FILE_TYPES = {"txt", "md"}
IMAGE_FILE_TYPES = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
TABLE_FILE_TYPES = {"csv", "xlsx", "xls"}
SUPPORTED_FILE_TYPES = TEXT_FILE_TYPES | IMAGE_FILE_TYPES | TABLE_FILE_TYPES | {"json", "pdf", "doc", "docx"}


def detect_input_type(path: Path, explicit: str) -> str:
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def extract_with_soffice(path: Path) -> str:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for office document conversion.")
    with tempfile.TemporaryDirectory(prefix="med-skill-preprocess-") as tmp_dir:
        proc = subprocess.run(
            [office_bin, "--headless", "--convert-to", "txt:Text", "--outdir", tmp_dir, str(path)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(f"Failed to convert office document: {proc.stderr.strip() or proc.stdout.strip()}")
        return out_path.read_text(encoding="utf-8-sig", errors="replace")


def extract_xls_csv(path: Path) -> List[Dict[str, Any]]:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for xls conversion.")
    with tempfile.TemporaryDirectory(prefix="med-skill-preprocess-") as tmp_dir:
        proc = subprocess.run(
            [office_bin, "--headless",
             "--convert-to", "csv:Text - txt - csv (StarCalc):44,34,76,1",
             "--outdir", tmp_dir, str(path)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False,
        )
        csv_path = Path(tmp_dir) / f"{path.stem}.csv"
        if proc.returncode != 0 or not csv_path.exists():
            raise PreprocessError(f"Failed to convert xls to csv: {proc.stderr.strip() or proc.stdout.strip()}")
        rows = read_csv_rows(csv_path, "utf-8-sig")
        if not rows:
            raise PreprocessError("No data rows found after xls conversion.")
        return [{"name": path.stem, "rows": rows}]


def extract_pdf_text(path: Path) -> str:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            text = "\n\n".join(page for page in pages if page)
            if text.strip():
                return text
        except Exception:
            pass
    pdf_to_text = shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout
    raise PreprocessError("Unable to extract text from pdf.")


def extract_image_text(path: Path) -> str:
    tesseract_bin = shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError("tesseract not found for image OCR.")
    langs = detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


def detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


def shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def load_input_artifact(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "text", "text": read_text_file(path, encoding)}
    if input_type == "json":
        return {"kind": "json", "data": read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        return {"kind": "tables", "tables": extract_xls_csv(path)}
    if input_type == "docx":
        return {"kind": "text", "text": extract_docx_text(path)}
    if input_type == "doc":
        return {"kind": "text", "text": extract_with_soffice(path)}
    if input_type == "pdf":
        return {"kind": "text", "text": extract_pdf_text(path)}
    if input_type in IMAGE_FILE_TYPES:
        return {"kind": "text", "text": extract_image_text(path)}
    raise PreprocessError(f"Unsupported input type: {input_type}")


def first_matching_index(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(normalize_header(alias))
        if idx is not None:
            return idx
    return None
