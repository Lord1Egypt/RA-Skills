#!/usr/bin/env python3
"""
生成 SMSSDK 配置模板 Excel 文件
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # 冻结首行
    ws.freeze_panes = 'A2'

    return ws


def create_basic_info_sheet(wb):
    """创建基础信息 Sheet"""
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["AppKey", "SMSSDK应用Key，从MobTech官网(https://www.mob.com/)注册应用获取", ""],
        ["AppSecret", "SMSSDK应用密钥，与AppKey一同获取", ""],
        ["包名", "Android应用包名，如：com.example.app", ""],
        ["签名MD5", "Android签名MD5（去掉冒号并转小写），如：a1b2c3d4e5f6...", ""],
        ["短信签名审核", "是否已在MobTech后台提交短信签名审核并通过（填写是或否）", "否"],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)

    # 标记必填项
    for row in range(2, 7):
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, color="FF0000")

    return ws


def create_privacy_sheet(wb):
    """创建隐私合规 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["合规要求", "根据工信部合规要求和中国区 App 上架规范，使用 MobSDK 需要在用户同意隐私政策后才能初始化 SDK"],
        ["隐私政策条款", "在 App 隐私政策中应包含以下参考条款：\n\"我们使用了第三方（上海掌之淘信息技术有限公司，以下称\"MobTech\"）MobTech 短信验证服务为您提供短信验证功能。为了顺利实现该功能，您需要授权 MobTech SDK 提供对应的服务；在您授权后，MobTech 将收集您相关的个人信息。关于 MobTech 所收集的信息种类、用途、个人信息保护的规则及退出机制等，详见短信验证隐私政策。\"", ""],
        ["代码位置", "在用户点击隐私政策\"同意\"按钮的回调中添加授权代码"],
        ["调用时机", "必须先展示隐私政策弹窗，用户点击\"同意\"后才能调用"],
        ["授权代码", "MobSDK.submitPolicyGrantResult(true, null);\n\n参数说明：\n- true: 用户同意隐私政策\n- null: 使用默认的隐私控制器"],
        ["替代方案", "MobSDK.submitPolicyGrantResult(MobCustomController cont, boolean auth);\n使用 MobCustomController 可限制 MobSDK 采集的数据维度"],
        ["合规指南", "https://mob.com/wiki/detailed?wiki=210&id=23", ""],
    ]

    ws = wb.create_sheet(title="隐私合规")

    # 设置表头
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 50

    return ws


def create_instructions_sheet(wb):
    """创建填写说明 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 先在 MobTech 官网(https://www.mob.com/) 注册应用，获取 AppKey 和 AppSecret\n2. 在开发者后台完成应用登记和短信签名审核\n3. 填写基础信息（必填）\n4. 填写完成后告诉我\"填好了\"，我将继续下一步"],
        ["必填项", "基础信息 Sheet 中的 AppKey、AppSecret、包名是必填项\n短信签名审核必须通过才能正常使用"],
        ["重要提醒", "短信签名审核通过后才能正常使用短信验证功能\n企业用户需上传营业执照申请自定义签名\n个人用户默认使用【掌淘科技】签名"],
        ["短信签名要求", "字数限制：3-8个字，不能全英文，不能添加特殊字符\n企业用户需上传营业执照\n申请应用名称还需提供商标证明、软著证明或上线截图"],
        ["国际短信", "如需支持国外或港澳台短信发送，需在签名审核页面开启\"支持全球所有运营商\"开关，并配置相应的英文签名"],
        ["常见问题", "验证码发送失败：检查AppKey/AppSecret是否正确、包名是否与后台一致\n短信签名未审核：需在MobTech后台提交短信签名审核"],
    ]

    ws = wb.create_sheet(title="填写说明")

    # 设置表头
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 60

    return ws


def main():
    wb = Workbook()

    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 按顺序创建各个 Sheet
    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    # 调整 Sheet 顺序
    sheets = wb.sheetnames
    target_order = [
        "基础信息",
        "隐私合规",
        "填写说明",
    ]

    for idx, name in enumerate(target_order):
        if name in sheets:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    # 获取当前脚本的目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 拼接输出路径（脚本目录 + 文件名）
    output_path = os.path.join(script_dir, "SMSSDK_Config_Template.xlsx")
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
