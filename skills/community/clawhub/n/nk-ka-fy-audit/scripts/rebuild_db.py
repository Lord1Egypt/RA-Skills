#!/usr/bin/env python3
"""
KA返佣审计 - 原始数据重建脚本
将33个月的Excel原始数据解析为统一结构的raw_records表
"""

import sqlite3
import os
import re
import openpyxl

DB_PATH = '/workspace/data/ka_commission_audit.db'
DATA_DIR = '/workspace/data'

# ============================================================
# 表结构定义
# ============================================================

CREATE_RAW_RECORDS = """
CREATE TABLE IF NOT EXISTS raw_records (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    
    -- 身份标识
    month           TEXT NOT NULL,
    sheet_name      TEXT NOT NULL,
    merchant_sn     TEXT,
    merchant_name   TEXT,
    brand           TEXT,
    
    -- 组织层级
    level1_name     TEXT,
    level2_name     TEXT,
    level3_name     TEXT,
    
    -- 交易金额
    alipay_txn_amount       REAL,
    alipay_indirect_txn_amt REAL,
    wechat_txn_amount       REAL,
    wechat_indirect_txn_amt REAL,
    
    -- 费率参数
    alipay_rebate_ratio     REAL,       -- 常规sheet: 支付宝返佣比例
    wechat_rebate_ratio     REAL,       -- 常规sheet: 微信返佣比例
    settlement_rate         REAL,       -- 拓展202307版: 结算费率
    cost_rate               REAL,       -- 拓展202307版: 成本费率
    merchant_rebate_ratio   REAL,       -- 拓展202307版: 商户返佣比例
    -- 拓展人志华sheet（202308+版）:
    alipay_direct_yield     REAL,
    alipay_direct_ch_ratio  REAL,
    alipay_indirect_yield   REAL,
    alipay_indirect_ch_ratio REAL,
    wechat_direct_yield     REAL,
    wechat_direct_ch_ratio  REAL,
    wechat_indirect_yield   REAL,
    wechat_indirect_ch_ratio REAL,
    
    -- 返佣金额
    alipay_rebate           REAL,       -- 支付宝返佣金额(常规sheet)
    wechat_rebate           REAL,       -- 微信返佣金额(常规sheet)
    alipay_direct_rebate    REAL,
    alipay_indirect_rebate  REAL,
    wechat_direct_rebate    REAL,
    wechat_indirect_rebate  REAL,
    total_rebate            REAL,       -- 返佣总金额/分润汇总
    
    -- 备注
    original_note   TEXT,               -- 原始备注
    audit_note      TEXT,               -- 审计备注: 历史差额调整/有问题，暂时不调整
    
    -- 元数据
    rebate_method   TEXT,
    settlement_date TEXT,
    occurrence_date TEXT,
    source_row      INTEGER,
    calc_method     TEXT                -- regular / zhongkuai / expansion_307 / expansion_308 / abnormal
)
"""

# ============================================================
# 备注解析规则
# ============================================================

# 需要加入汇总的备注关键词
INCLUDE_NOTE_KEYWORDS = [
    '上月待确认商户返佣，本月确认',
    '系统未维护返佣方式sql未跑出，手工核算',
    '调整费率导致的人工核算',
    '签约信息表返佣方式未维护导致的人工核算',
    '调账',
]

# 不需要加入汇总的备注关键词
EXCLUDE_NOTE_KEYWORDS = [
    '不返佣',
    '返佣有疑问，待确认',
    '已确认，暂不返佣',
    '线上场景',
    '微信交易被罚',
    '微信费率异常',
    '无协议',
    '返佣方式需要调整',
    '应不涉及返佣',
    '费率需排查',
]


def classify_note(note_text):
    """根据备注内容判断审计备注"""
    if not note_text or note_text.strip() in ('', '#N/A', '\\N'):
        return '有问题，暂时不调整'
    
    note = note_text.strip()
    
    for kw in INCLUDE_NOTE_KEYWORDS:
        if kw in note:
            return '历史差额调整'
    
    for kw in EXCLUDE_NOTE_KEYWORDS:
        if kw in note:
            return '有问题，暂时不调整'
    
    # 未匹配任何规则，默认不调整
    return '有问题，暂时不调整'


# ============================================================
# 常规Sheet解析 (适用：常规, 中快X月, 异常, 调差, 问题, sql)
# ============================================================

def parse_standard_sheet(ws, month, sheet_name, header_row=1, global_notes=None):
    """
    解析常规结构的sheet（含支付宝/微信交易金额+返佣比例+返佣金额）
    
    header_row: 表头所在行（1-indexed）
    早期异常/调差sheet表头在第3行，常规/中快速在第1行
    """
    records = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 检测列名映射（中文 vs 英文）
    header_cells = []
    for c in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        header_cells = [(cell.column, str(cell.value or '').strip()) for cell in c]
        break
    
    col_map = {}
    for col_idx, val in header_cells:
        if val in ('收钱吧商户号', 'merchant_sn'):
            col_map['merchant_sn'] = col_idx
        elif val == '商户号':
            col_map['merchant_sn'] = col_idx
        elif val in ('收钱吧商户名称', 'merchant_name', '商户名'):
            col_map['merchant_name'] = col_idx
        elif val in ('品牌', 'brand'):
            col_map['brand'] = col_idx
        elif val in ('level1_name', 'level1_name'):
            col_map['level1_name'] = col_idx
        elif val in ('level2_name', 'level2_name'):
            col_map['level2_name'] = col_idx
        elif val in ('level3_name', 'level3_name'):
            col_map['level3_name'] = col_idx
        elif val in ('返佣方式', 'rebate_type'):
            col_map['rebate_method'] = col_idx
        elif val in ('支付宝交易金额', 'zfb_trans_amt'):
            col_map['alipay_txn'] = col_idx
        elif val in ('支付宝返佣比例', 'zfb_commission_rate'):
            col_map['alipay_rate'] = col_idx
        elif val in ('支付宝返佣金额', 'zfb_commission_amt'):
            col_map['alipay_rebate'] = col_idx
        elif val in ('微信交易金额', 'wx_trans_amt'):
            col_map['wechat_txn'] = col_idx
        elif val in ('微信返佣比例', 'wx_commission_rate'):
            col_map['wechat_rate'] = col_idx
        elif val in ('微信返佣金额', 'wx_commission_amt'):
            col_map['wechat_rebate'] = col_idx
        elif val in ('返佣合计', 'bill_amt'):
            col_map['total_rebate'] = col_idx
        elif val in ('交易期日', 'months'):
            col_map['occurrence_date'] = col_idx
    
    if 'merchant_sn' not in col_map:
        # 尝试找商户号列的其他名称
        return records  # 无匹配列，跳过
    
    # 查找备注列（通常在最后一两列，列名不定）
    # 扫描所有header找备注相关列
    note_col = None
    for col_idx, val in header_cells:
        stripped = str(val).strip()
        if stripped in ('备注', '上月', '差') or '备注' in stripped or '上月' == stripped:
            note_col = col_idx
            break
    
    if not note_col:
        # 某些late sheet备注在S列(column 19)
        if max_col >= 19:
            s_col_val = ws.cell(row=header_row, column=19).value
            if s_col_val:
                s_str = str(s_col_val).strip().lower()
                if s_str in ('差', '差异', 'diff') or '差' in str(s_str):
                    note_col = 19  # This is diff, use R column (18) for notes
        # 通用：检查R列之后是否有备注
        for col_idx in range(18, max_col + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val and str(val).strip() in ('备注', 'remark', 'note'):
                note_col = col_idx
                break
    
    # 从数据行开始解析
    data_start = header_row + 1
    for row_idx in range(data_start, max_row + 1):
        row_vals = {}
        row_empty = True
        
        # 读取基本字段
        if 'merchant_sn' in col_map:
            val = ws.cell(row=row_idx, column=col_map['merchant_sn']).value
            if val is not None and str(val).strip() not in ('', '\\N', '#N/A'):
                row_vals['merchant_sn'] = str(val).strip()
                row_empty = False
        
        if not row_vals.get('merchant_sn'):
            # 无商户号的行可能是分隔行、汇总行或空行
            # 检查是否有备注文字
            if note_col:
                note_val = ws.cell(row=row_idx, column=note_col).value
                if note_val and str(note_val).strip() not in ('', '\\N', '#N/A', '#REF!'):
                    # 这是一个全局备注行/分隔行，记录但不作为商户数据
                    pass
            continue
        
        # 读取商户名、品牌
        if 'merchant_name' in col_map:
            row_vals['merchant_name'] = str(ws.cell(row=row_idx, column=col_map['merchant_name']).value or '')
        if 'brand' in col_map:
            row_vals['brand'] = str(ws.cell(row=row_idx, column=col_map['brand']).value or '')
        
        # 层级
        for key in ('level1_name', 'level2_name', 'level3_name'):
            if key in col_map:
                row_vals[key] = str(ws.cell(row=row_idx, column=col_map[key]).value or '')
        
        # 返佣方式
        if 'rebate_method' in col_map:
            row_vals['rebate_method'] = str(ws.cell(row=row_idx, column=col_map['rebate_method']).value or '')
        
        # 金额字段（尝试转为float）
        for key in ('alipay_txn', 'alipay_rate', 'alipay_rebate',
                     'wechat_txn', 'wechat_rate', 'wechat_rebate',
                     'total_rebate'):
            if key in col_map:
                val = ws.cell(row=row_idx, column=col_map[key]).value
                try:
                    row_vals[key] = float(val) if val is not None else None
                except (ValueError, TypeError):
                    row_vals[key] = None
        
        # 发生日期/交易期日
        if 'occurrence_date' in col_map:
            val = ws.cell(row=row_idx, column=col_map['occurrence_date']).value
            row_vals['occurrence_date'] = str(val).strip() if val else ''
        
        # 备注列（查找各列中的备注文字）
        raw_note = ''
        if note_col:
            val = ws.cell(row=row_idx, column=note_col).value
            if val and str(val).strip() not in ('', '\\N', '#N/A', '#REF!'):
                raw_note = str(val).strip()
        
        # 也扫描其他列中的文字备注（比如某些表备注在R/S列）
        if not raw_note and max_col >= 18:
            for c in range(18, max_col + 1):
                val = ws.cell(row=row_idx, column=c).value
                if val and isinstance(val, str) and len(val) > 5 and val not in ('\\N', '#N/A', '#REF!'):
                    if not any(kw in val for kw in ['zfb', 'wx_', 'bill_', 'merchant_', 'level']):
                        raw_note = val.strip()
                        break
        
        # 应用全局备注
        if not raw_note and global_notes and row_idx >= data_start:
            raw_note = global_notes.get('text', '')
        
        # 分类审计备注
        audit_note = classify_note(raw_note)
        
        # 如果全局备注中有"支付宝返佣正常发放"但行备注说"暂不返佣"，以行备注为准
        # 如果只有全局备注，且全局备注说明部分渠道正常，则标记为需要排查
        
        record = {
            'month': month,
            'sheet_name': sheet_name,
            'merchant_sn': row_vals.get('merchant_sn', ''),
            'merchant_name': row_vals.get('merchant_name', ''),
            'brand': row_vals.get('brand', ''),
            'level1_name': row_vals.get('level1_name', ''),
            'level2_name': row_vals.get('level2_name', ''),
            'level3_name': row_vals.get('level3_name', ''),
            'alipay_txn_amount': row_vals.get('alipay_txn'),
            'wechat_txn_amount': row_vals.get('wechat_txn'),
            'alipay_rebate_ratio': row_vals.get('alipay_rate'),
            'wechat_rebate_ratio': row_vals.get('wechat_rate'),
            'alipay_rebate': row_vals.get('alipay_rebate'),
            'wechat_rebate': row_vals.get('wechat_rebate'),
            'total_rebate': row_vals.get('total_rebate'),
            'rebate_method': row_vals.get('rebate_method', ''),
            'occurrence_date': row_vals.get('occurrence_date', ''),
            'original_note': raw_note,
            'audit_note': audit_note,
            'source_row': row_idx,
            'calc_method': 'abnormal' if sheet_name in ('异常', '调差', '问题') else 'regular' if sheet_name == '常规' else 'zhongkuai',
        }
        records.append(record)
    
    return records


# ============================================================
# 拓展Sheet解析（202307版）- 表头在第3行
# ============================================================

def parse_expansion_307(ws, month, sheet_name='拓展'):
    """
    解析202307版"拓展"sheet
    表头在第3行，含：sn, 支付宝商户号, 微信商户号, name, brand, 
    各支付方式交易金额/结算费率/返佣比例，分润在N/T/Z/AF列，总金额AG列
    公式：返佣=交易金额×(结算费率-成本费率)×渠道分润比例
    """
    records = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    header_row = 3
    header_cells = []
    for c in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        for cell in c:
            letter = openpyxl.utils.get_column_letter(cell.column)
            val = cell.value
            if val is not None:
                header_cells.append((letter, cell.column, str(val).strip()))
        break
    
    # 构建列映射
    col_map = {}
    for letter, col_idx, val in header_cells:
        if val in ('sn', '支付宝商户号'):
            col_map['sn'] = (letter, col_idx)
        elif val in ('name',):
            col_map['name'] = (letter, col_idx)
        elif val in ('brand',):
            col_map['brand'] = (letter, col_idx)
        elif val in ('支付宝直联交易金额',):
            col_map['alipay_direct_txn'] = (letter, col_idx)
        elif val in ('结算费率',):
            col_map['settlement_rate'] = (letter, col_idx)
        elif val in ('商户返佣比例',):
            col_map['merchant_rebate_ratio'] = (letter, col_idx)
    
    # 分润列: N直联, T, Z间联, AF, AG汇总
    # 通过列名查找
    for letter, col_idx, val in header_cells:
        if '成本费率' in val or '结算费率' in val:
            if '结算' in val:
                col_map['settlement_rate'] = (letter, col_idx)
        if '返佣比例' in val:
            col_map['merchant_rebate_ratio'] = (letter, col_idx)
    
    # 预设已知列（202307 拓展固定的列布局）
    # 行3列I=支付宝直联交易金额, J=结算费率, K=商户返佣比例, L=成本费率
    # 各通道分润在N(=14), T(=20), Z(=26), AF(=32), AG(=33)
    col_map['alipay_direct_rebate'] = ('N', 14)
    col_map['channel_rebate_T'] = ('T', 20)
    col_map['channel_rebate_Z'] = ('Z', 26)
    col_map['channel_rebate_AF'] = ('AF', 32)
    col_map['total_rebate'] = ('AG', 33)
    col_map['cost_rate'] = ('L', 12)
    
    # 扫描实际表头验证
    for letter, col_idx, val in header_cells:
        if '分润' in val or '返佣' in val:
            if letter in ('N', 'T', 'Z', 'AF', 'AG'):
                pass  # 确认存在
    
    if 'sn' not in col_map and 'brand' not in col_map:
        return records
    
    # 数据从第4行开始
    for row_idx in range(header_row + 1, max_row + 1):
        # 获取sn/品牌
        sn = None
        brand = None
        if 'sn' in col_map:
            val = ws.cell(row=row_idx, column=col_map['sn'][1]).value
            if val is not None:
                sn = str(val).strip()
        if 'brand' in col_map:
            val = ws.cell(row=row_idx, column=col_map['brand'][1]).value
            if val is not None:
                brand = str(val).strip()
        
        if not sn and not brand:
            continue
        
        # 读取各字段
        def get_float(col_letter_or_idx):
            if isinstance(col_letter_or_idx, tuple):
                col_idx = col_letter_or_idx[1]
            else:
                col_idx = col_letter_or_idx
            val = ws.cell(row=row_idx, column=col_idx).value
            try:
                return float(val) if val is not None else None
            except (ValueError, TypeError):
                return None
        
        def get_str(col_letter_or_idx):
            if isinstance(col_letter_or_idx, tuple):
                col_idx = col_letter_or_idx[1]
            else:
                col_idx = col_letter_or_idx
            val = ws.cell(row=row_idx, column=col_idx).value
            return str(val).strip() if val else ''
        
        name = get_str(col_map.get('name', ('', 5))) if 'name' in col_map else ''
        
        record = {
            'month': month,
            'sheet_name': sheet_name,
            'merchant_sn': sn or '',
            'merchant_name': name,
            'brand': brand or '',
            'level1_name': '',
            'level2_name': '',
            'level3_name': '',
            'alipay_txn_amount': get_float(col_map.get('alipay_direct_txn', ('I', 9))),
            'settlement_rate': get_float(col_map.get('settlement_rate', ('J', 10))),
            'cost_rate': get_float(col_map.get('cost_rate', ('L', 12))),
            'merchant_rebate_ratio': get_float(col_map.get('merchant_rebate_ratio', ('K', 11))),
            'alipay_direct_rebate': get_float(col_map.get('alipay_direct_rebate', ('N', 14))),
            'channel_rebate_col_T': get_float(col_map.get('channel_rebate_T', ('T', 20))),
            'channel_rebate_col_Z': get_float(col_map.get('channel_rebate_Z', ('Z', 26))),
            'channel_rebate_col_AF': get_float(col_map.get('channel_rebate_AF', ('AF', 32))),
            'total_rebate': get_float(col_map.get('total_rebate', ('AG', 33))),
            'original_note': '',
            'audit_note': '',
            'rebate_method': '',
            'occurrence_date': '',
            'source_row': row_idx,
            'calc_method': 'expansion_307',
        }
        records.append(record)
    
    return records


# ============================================================
# 拓展人志华Sheet解析（202308+版）- 表头在第1行
# ============================================================

def parse_expansion_308(ws, month, sheet_name, header_row=1):
    """
    解析202308+版"志华/拓展人志华"sheet
    表头在第1行（或第2行），含：商户号, 商户名, 品牌,
    支付宝直连交易金额/收益率/渠道分润比例/分润,
    支付宝间连..., 微信直连..., 微信间连...
    分润在G/K/O/S列，总金额T列
    公式：返佣=交易金额×收益率×分润比例
    """
    records = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 扫描所有行找表头
    actual_header = header_row
    for r in range(header_row, min(header_row + 5, max_row + 1)):
        c_a = ws.cell(row=r, column=1).value
        if c_a and str(c_a).strip() == '商户号':
            actual_header = r
            break
    
    header_cells = []
    for c in ws.iter_rows(min_row=actual_header, max_row=actual_header, values_only=False):
        header_cells = [(cell.column, str(cell.value or '').strip()) for cell in c]
        break
    
    col_map = {}
    for col_idx, val in header_cells:
        if val in ('商户号',):
            col_map['merchant_sn'] = col_idx
        elif val in ('商户名',):
            col_map['merchant_name'] = col_idx
        elif val in ('品牌',):
            col_map['brand'] = col_idx
        elif val in ('支付宝直连交易金额',):
            col_map['alipay_direct_txn'] = col_idx
        elif val in ('支付宝直连收益率',):
            col_map['alipay_direct_yield'] = col_idx
        elif val in ('支付宝直连渠道分润比例',):
            col_map['alipay_direct_ch_ratio'] = col_idx
        elif val in ('支付宝直连分润',):
            col_map['alipay_direct_rebate'] = col_idx
        elif val in ('支付宝间连交易金额',):
            col_map['alipay_indirect_txn'] = col_idx
        elif val in ('支付宝间连收益率',):
            col_map['alipay_indirect_yield'] = col_idx
        elif val in ('支付宝间连渠道分润比例',):
            col_map['alipay_indirect_ch_ratio'] = col_idx
        elif val in ('支付宝间连分润',):
            col_map['alipay_indirect_rebate'] = col_idx
        elif val in ('微信直连交易金额',):
            col_map['wechat_direct_txn'] = col_idx
        elif val in ('微信直连收益率',):
            col_map['wechat_direct_yield'] = col_idx
        elif val in ('微信直连渠道分润比例',):
            col_map['wechat_direct_ch_ratio'] = col_idx
        elif val in ('微信直连分润',):
            col_map['wechat_direct_rebate'] = col_idx
        elif val in ('微信间连交易金额',):
            col_map['wechat_indirect_txn'] = col_idx
        elif val in ('微信间连收益率',):
            col_map['wechat_indirect_yield'] = col_idx
        elif val in ('微信间连渠道分润比例',):
            col_map['wechat_indirect_ch_ratio'] = col_idx
        elif val in ('微信间连分润',):
            col_map['wechat_indirect_rebate'] = col_idx
        elif val in ('分润汇总',):
            col_map['total_rebate'] = col_idx
    
    if 'merchant_sn' not in col_map and 'brand' not in col_map:
        return records
    
    def get_float(row, col):
        val = ws.cell(row=row, column=col).value
        try:
            return float(val) if val is not None else None
        except (ValueError, TypeError):
            return None
    
    def get_str(row, col):
        val = ws.cell(row=row, column=col).value
        s = str(val).strip() if val else ''
        return '' if s in ('', '\\N', '#N/A', 'None', '#REF!') else s
    
    for row_idx in range(actual_header + 1, max_row + 1):
        # 检查是否有商户号
        sn = None
        if 'merchant_sn' in col_map:
            sn = get_str(row_idx, col_map['merchant_sn'])
        
        if not sn:
            continue
        
        record = {
            'month': month,
            'sheet_name': sheet_name,
            'merchant_sn': sn,
            'merchant_name': get_str(row_idx, col_map.get('merchant_name', 1)),
            'brand': get_str(row_idx, col_map.get('brand', 1)),
            'level1_name': '',
            'level2_name': '',
            'level3_name': '',
            'alipay_txn_amount': get_float(row_idx, col_map.get('alipay_direct_txn', 1)),
            'alipay_indirect_txn_amt': get_float(row_idx, col_map.get('alipay_indirect_txn', 1)),
            'alipay_direct_yield': get_float(row_idx, col_map.get('alipay_direct_yield', 1)),
            'alipay_direct_ch_ratio': get_float(row_idx, col_map.get('alipay_direct_ch_ratio', 1)),
            'alipay_direct_rebate': get_float(row_idx, col_map.get('alipay_direct_rebate', 1)),
            'alipay_indirect_yield': get_float(row_idx, col_map.get('alipay_indirect_yield', 1)),
            'alipay_indirect_ch_ratio': get_float(row_idx, col_map.get('alipay_indirect_ch_ratio', 1)),
            'alipay_indirect_rebate': get_float(row_idx, col_map.get('alipay_indirect_rebate', 1)),
            'total_rebate': get_float(row_idx, col_map.get('total_rebate', 1)),
            'original_note': '',
            'audit_note': '',
            'rebate_method': '',
            'occurrence_date': '',
            'source_row': row_idx,
            'calc_method': 'expansion_308',
        }
        
        # 微信通道（有些月份只有支付宝）
        if 'wechat_direct_txn' in col_map:
            record['wechat_txn_amount'] = get_float(row_idx, col_map['wechat_direct_txn'])
        if 'wechat_indirect_txn' in col_map:
            record['wechat_indirect_txn_amt'] = get_float(row_idx, col_map['wechat_indirect_txn'])
        if 'wechat_direct_yield' in col_map:
            record['wechat_direct_yield'] = get_float(row_idx, col_map['wechat_direct_yield'])
        if 'wechat_direct_ch_ratio' in col_map:
            record['wechat_direct_ch_ratio'] = get_float(row_idx, col_map['wechat_direct_ch_ratio'])
        if 'wechat_direct_rebate' in col_map:
            record['wechat_direct_rebate'] = get_float(row_idx, col_map['wechat_direct_rebate'])
        if 'wechat_indirect_yield' in col_map:
            record['wechat_indirect_yield'] = get_float(row_idx, col_map['wechat_indirect_yield'])
        if 'wechat_indirect_ch_ratio' in col_map:
            record['wechat_indirect_ch_ratio'] = get_float(row_idx, col_map['wechat_indirect_ch_ratio'])
        if 'wechat_indirect_rebate' in col_map:
            record['wechat_indirect_rebate'] = get_float(row_idx, col_map['wechat_indirect_rebate'])
        
        records.append(record)
    
    return records


# ============================================================
# 主处理循环
# ============================================================

def find_excel_file(month):
    """查找月份对应的Excel文件"""
    month_dir = os.path.join(DATA_DIR, month)
    if not os.path.isdir(month_dir):
        return None
    for f in os.listdir(month_dir):
        if f.endswith('.xlsx') and 'KA返佣' in f:
            return os.path.join(month_dir, f)
    return None


def get_global_note(ws):
    """从异常/问题sheet的Row 1提取全局备注"""
    row1_val = ws.cell(row=1, column=1).value
    if row1_val and isinstance(row1_val, str) and len(row1_val) > 5:
        return {'row': 1, 'text': row1_val.strip()}
    return None


def parse_month(month, excel_path):
    """解析一个月的所有sheet"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_records = []
    
    sheet_names = wb.sheetnames
    print(f'  解析 {month}: sheets={sheet_names}')
    
    for sn in sheet_names:
        ws = wb[sn]
        
        if sn == '汇总':
            # 汇总sheet跳过（这是汇总结果，我们只需原始数据）
            continue
        
        elif sn in ('常规',):
            records = parse_standard_sheet(ws, month, sn, header_row=1)
            for r in records:
                r['calc_method'] = 'regular'
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn.startswith('中快'):
            records = parse_standard_sheet(ws, month, sn, header_row=1)
            for r in records:
                r['calc_method'] = 'zhongkuai'
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('拓展',) and month == '202307':
            records = parse_expansion_307(ws, month, sn)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('拓展人志华', '志华'):
            records = parse_expansion_308(ws, month, sn)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('异常',):
            global_note = get_global_note(ws)
            # 早期异常sheet表头在第3行
            # 检查第1行是否有数据
            row1_is_header = False
            c1_1 = ws.cell(row=1, column=1).value
            if c1_1 and str(c1_1).strip() in ('账单邮箱', 'bill_email'):
                row1_is_header = True
            
            if row1_is_header:
                # 表头在第1行（中期以后）
                records = parse_standard_sheet(ws, month, sn, header_row=1, global_notes=global_note)
            else:
                # 早期异常sheet，表头在第3行
                records = parse_standard_sheet(ws, month, sn, header_row=3, global_notes=global_note)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条 (global_note={bool(global_note)})')
        
        elif sn in ('调差', '回算'):
            global_note = get_global_note(ws)
            records = parse_standard_sheet(ws, month, sn, header_row=3, global_notes=global_note)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('问题',):
            records = parse_standard_sheet(ws, month, sn, header_row=1)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('sql',):
            records = parse_standard_sheet(ws, month, sn, header_row=1)
            for r in records:
                r['calc_method'] = 'sql_recalc'
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        elif sn in ('异常及调账', '异常及回算', '调差'):
            records = parse_standard_sheet(ws, month, sn, header_row=3)
            all_records.extend(records)
            if records:
                print(f'    {sn}: {len(records)}条')
        
        else:
            # 尝试通用解析
            pass
    
    wb.close()
    return all_records


def main():
    # 创建新表
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 删除旧表
    cursor.execute('DROP TABLE IF EXISTS raw_records')
    cursor.execute(CREATE_RAW_RECORDS)
    conn.commit()
    print('已创建 raw_records 表')
    
    # 获取所有月份目录
    months = sorted([d for d in os.listdir(DATA_DIR) 
                     if os.path.isdir(os.path.join(DATA_DIR, d)) and re.match(r'^\d{6}$', d)])
    print(f'找到 {len(months)} 个月份目录')
    
    total_records = 0
    for month in months:
        excel_path = find_excel_file(month)
        if not excel_path:
            print(f'  跳过 {month}: 未找到Excel文件')
            continue
        
        records = parse_month(month, excel_path)
        
        if not records:
            print(f'  {month}: 无解析记录')
            continue
        
        # 批量插入
        columns = list(records[0].keys())
        placeholders = ','.join(['?' for _ in columns])
        col_str = ','.join(columns)
        
        values = []
        for r in records:
            row = [r.get(c) for c in columns]
            values.append(row)
        
        cursor.executemany(f'INSERT INTO raw_records ({col_str}) VALUES ({placeholders})', values)
        conn.commit()
        total_records += len(records)
        print(f'  {month}: {len(records)}条写入完成')
    
    print(f'\n总计: {total_records} 条记录写入 raw_records 表')
    
    # 打印各月份统计
    print('\n=== 各月数据量 ===')
    cursor.execute('SELECT month, sheet_name, COUNT(*) FROM raw_records GROUP BY month, sheet_name ORDER BY month, sheet_name')
    for r in cursor.fetchall():
        print(f'  {r[0]:6s} | {r[1]:20s} | {r[2]:>5}条')
    
    conn.close()


if __name__ == '__main__':
    main()