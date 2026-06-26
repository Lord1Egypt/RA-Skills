#!/usr/bin/env python3
"""
class_reminder.py - 课程提醒核心脚本（支持教师课表格式）

功能：
1. parse_teacher_schedule(excel_path) - 解析教师课表Excel，返回结构化课程数据
2. get_todays_classes(schedule_data, date_str) - 查询指定日期的课程
3. get_tomorrows_classes(schedule_data, date_str) - 查询明日课程（用于提醒）

教师课表格式：
- 行：时间段（第1-2节、第3-4节等）
- 列：星期一到星期日
- 单元格：课程信息，如 "【实验】网络与系统安全 [1-4节][6-6周] [T2604-T2606]"
"""

import sys
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "缺少 openpyxl 库，请运行: pip install openpyxl"}))
    sys.exit(1)


def parse_time_slot(slot_str):
    """解析时间段字符串，如 '第1-2节' -> (1, 2)
    返回 (start_section, end_section)
    """
    if not slot_str:
        return (1, 2)  # 默认
    
    # 匹配 "第1-2节" 或 "第1节"
    m = re.search(r'第(\d+)(?:-(\d+))?节', str(slot_str))
    if m:
        start = int(m.group(1))
        end = int(m.group(2)) if m.group(2) else start
        return (start, end)
    
    # 匹配纯数字 "1-2" 或 "1"
    m = re.search(r'(\d+)(?:-(\d+))?', str(slot_str))
    if m:
        start = int(m.group(1))
        end = int(m.group(2)) if m.group(2) else start
        return (start, end)
    
    return (1, 2)


def section_to_time(section):
    """将节次转换为时间
    1: 08:00, 2: 08:50, 3: 10:00, 4: 10:50, 5: 14:00, 6: 14:50, 7: 16:00, 8: 16:50, 9: 19:00, 10: 19:50
    """
    time_map = {
        1: "08:00", 2: "08:50", 3: "10:00", 4: "10:50",
        5: "14:00", 6: "14:50", 7: "16:00", 8: "16:50",
        9: "19:00", 10: "19:50", 11: "20:40"
    }
    return time_map.get(section, "08:00")


def parse_course_info(course_text):
    """解析课程信息文本
    输入: "【实验】网络与系统安全 [1-4节][6-6周] [T2604-T2606]"
    返回: {
        "type": "实验",
        "name": "网络与系统安全", 
        "section_range": [1, 4],
        "week_range": [6, 6],
        "location": "T2604-T2606"
    }
    """
    if not course_text or str(course_text).strip() == "None":
        return None
        
    text = str(course_text).strip()
    
    # 提取课程类型 【实验】
    course_type = ""
    type_match = re.search(r'【([^】]+)】', text)
    if type_match:
        course_type = type_match.group(1)
        text = text.replace(type_match.group(0), "").strip()
    
    # 提取节次范围 [1-4节]
    section_range = [1, 2]
    section_match = re.search(r'\[(\d+)(?:-(\d+))?节\]', text)
    if section_match:
        start = int(section_match.group(1))
        end = int(section_match.group(2)) if section_match.group(2) else start
        section_range = [start, end]
        text = text.replace(section_match.group(0), "").strip()
    
    # 提取周次范围 [6-6周] 或 [11-12,14-14周]
    week_ranges = []
    week_match = re.search(r'\[((?:\d+(?:-\d+)?(?:,\d+(?:-\d+)?)*)?)周\]', text)
    if week_match:
        weeks_str = week_match.group(1)
        if weeks_str:
            # 处理 "11-12,14-14" 这样的格式
            for part in weeks_str.split(','):
                part = part.strip()
                if '-' in part:
                    start_w, end_w = map(int, part.split('-'))
                    week_ranges.extend(range(start_w, end_w + 1))
                else:
                    week_ranges.append(int(part))
        text = text.replace(week_match.group(0), "").strip()
    
    if not week_ranges:
        week_ranges = list(range(1, 21))  # 默认全学期
    
    # 剩余部分是课程名称和地点
    parts = [p.strip() for p in text.split('[') if p.strip()]
    course_name = ""
    location = "待定"
    
    if parts:
        # 第一部分通常是课程名称
        course_name = parts[0].rstrip(']').strip()
        # 其他部分可能是地点
        if len(parts) > 1:
            location = parts[1].rstrip(']').strip()
    
    return {
        "type": course_type,
        "name": course_name,
        "section_range": section_range,
        "week_list": sorted(set(week_ranges)),
        "location": location or "待定"
    }


def parse_teacher_schedule(excel_path, semester_start=None):
    """
    解析教师课表Excel文件。
    
    Args:
        excel_path: Excel 文件路径
        semester_start: 学期开始日期 (YYYY-MM-DD)
    
    Returns:
        dict: {
            "classes": [...],        # 课程列表
            "semester_start": str,   # 学期开始日期
            "error": str or None
        }
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return {"classes": [], "error": f"文件不存在: {excel_path}"}
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return {"classes": [], "error": f"无法打开文件: {e}"}
    
    classes = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows or len(rows) < 2:
            continue
            
        # 第一行是标题 "时间", "教师课表"
        # 第二行是星期列头：星期一、星期二...
        if len(rows) < 2:
            continue
            
        weekday_headers = rows[1]  # 第二行
        time_slots = []  # 时间段列表
        
        # 构建星期映射
        weekday_map = {}
        for col_idx, header in enumerate(weekday_headers):
            if header and str(header).strip():
                day_str = str(header).strip()
                if "星期一" in day_str:
                    weekday_map[col_idx] = 1
                elif "星期二" in day_str:
                    weekday_map[col_idx] = 2
                elif "星期三" in day_str:
                    weekday_map[col_idx] = 3
                elif "星期四" in day_str:
                    weekday_map[col_idx] = 4
                elif "星期五" in day_str:
                    weekday_map[col_idx] = 5
                elif "星期六" in day_str:
                    weekday_map[col_idx] = 6
                elif "星期日" in day_str:
                    weekday_map[col_idx] = 7
        
        # 解析课程数据行
        for row_idx in range(2, len(rows)):
            row = rows[row_idx]
            if not row or not any(row):
                continue
                
            # 第一列是时间段
            time_slot_cell = row[0] if len(row) > 0 else None
            if not time_slot_cell:
                continue
                
            time_slot = parse_time_slot(time_slot_cell)
            
            # 解析每个星期的课程
            for col_idx in range(1, len(row)):
                if col_idx not in weekday_map:
                    continue
                    
                cell_content = row[col_idx]
                if not cell_content or str(cell_content).strip() in ("None", ""):
                    continue
                
                # 一个单元格可能包含多门课程（换行分隔）
                course_entries = str(cell_content).split('\n')
                for entry in course_entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                        
                    course_info = parse_course_info(entry)
                    if not course_info:
                        continue
                        
                    # 为节次范围内的每一节课创建记录（但合并同一课程）
                    # 使用课程名+类型+星期+开始节次作为唯一键去重
                    course_key = f"{course_info['name']}_{course_info['type']}_{weekday_map[col_idx]}_{course_info['section_range'][0]}"
                    
                    # 只添加一次，保留节次范围信息
                    classes.append({
                        "name": course_info["name"],
                        "type": course_info["type"],
                        "weekday": weekday_map[col_idx],
                        "section_start": course_info["section_range"][0],
                        "section_end": course_info["section_range"][1],
                        "start_time": section_to_time(course_info["section_range"][0]),
                        "end_time": section_to_time(course_info["section_range"][1] + 1) if course_info["section_range"][1] + 1 <= 12 else "21:30",
                        "location": course_info["location"],
                        "week_list": course_info["week_list"],
                        "sheet": sheet_name,
                    })
    
    if semester_start is None:
        semester_start = _guess_semester_start()
    
    return {
        "classes": classes,
        "semester_start": semester_start,
        "error": None,
    }


def _guess_semester_start():
    """根据当前日期推测学期开始日期"""
    now = datetime.now()
    year = now.year
    
    # 春季学期从2月开始，秋季学期从9月开始
    if now.month >= 9:
        return f"{year}-09-01"
    elif now.month >= 2:
        return f"{year}-02-15"
    else:
        return f"{year - 1}-09-01"


def _get_week_number(date_str, semester_start):
    """计算指定日期是第几周"""
    date = datetime.strptime(date_str, "%Y-%m-%d")
    start = datetime.strptime(semester_start, "%Y-%m-%d")
    delta = (date - start).days
    week_num = delta // 7 + 1
    return max(1, week_num)


def get_classes_for_date(schedule_data, date_str):
    """
    获取指定日期的课程列表。
    """
    classes = schedule_data.get("classes", [])
    semester_start = schedule_data.get("semester_start", _guess_semester_start())
    
    try:
        date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return []
    
    weekday = date.isoweekday()  # 1=周一, 7=周日
    week_num = _get_week_number(date_str, semester_start)
    
    today_classes = []
    for cls in classes:
        # 检查星期是否匹配
        if cls.get("weekday") != weekday:
            continue
            
        # 检查周次是否在范围内
        week_list = cls.get("week_list", [])
        if week_list and week_num not in week_list:
            continue
            
        today_classes.append({
            "name": cls["name"],
            "type": cls.get("type", ""),
            "start_time": cls["start_time"],
            "end_time": cls.get("end_time", ""),
            "location": cls.get("location", "待定"),
            "section_start": cls.get("section_start", 1),
            "section_end": cls.get("section_end", 2),
        })
    
    # 去重：按课程名+开始时间+地点
    seen = set()
    unique_classes = []
    for cls in today_classes:
        key = f"{cls['name']}_{cls['start_time']}_{cls['location']}"
        if key not in seen:
            seen.add(key)
            unique_classes.append(cls)
    
    # 按开始节次排序
    unique_classes.sort(key=lambda x: x["section_start"])
    return unique_classes


def get_tomorrows_classes(schedule_data, date_str=None):
    """
    获取明天的课程列表。
    """
    if date_str is None:
        tomorrow = datetime.now() + timedelta(days=1)
        date_str = tomorrow.strftime("%Y-%m-%d")
    else:
        tomorrow = datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)
        date_str = tomorrow.strftime("%Y-%m-%d")
    
    weekday_names = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}
    
    classes = get_classes_for_date(schedule_data, date_str)
    tomorrow_weekday = tomorrow.isoweekday()
    
    return {
        "date": date_str,
        "weekday": weekday_names.get(tomorrow_weekday, ""),
        "classes": classes,
    }


def format_reminder_text(reminder_data):
    """
    将明日课程数据格式化为提醒文本。
    """
    date = reminder_data.get("date", "")
    weekday = reminder_data.get("weekday", "")
    classes = reminder_data.get("classes", [])
    
    if not classes:
        return f"📅 {date}（{weekday}）没有课程安排，好好休息！🍋"
    
    lines = [f"📅 明天 {date}（{weekday}）的课程提醒：\n"]
    
    for i, cls in enumerate(classes, 1):
        time_range = cls["start_time"]
        if cls.get("end_time"):
            time_range = f"{cls['start_time']}-{cls['end_time']}"
        
        course_title = cls["name"]
        if cls.get("type"):
            course_title = f"【{cls['type']}】{course_title}"
            
        line = f"  {i}. {course_title}"
        line += f"\n     ⏰ {time_range}"
        line += f"\n     📍 {cls['location']}"
        
        lines.append(line)
    
    lines.append(f"\n共 {len(classes)} 节课，加油！🍋")
    return "\n".join(lines)


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="教师课程提醒工具")
    parser.add_argument("action", choices=["parse", "today", "tomorrow", "date"], help="操作类型")
    parser.add_argument("file", help="Excel 课程表文件路径")
    parser.add_argument("--date", help="指定日期 (YYYY-MM-DD)")
    parser.add_argument("--semester-start", help="学期开始日期 (YYYY-MM-DD)")
    parser.add_argument("--output", help="输出文件路径（JSON格式）")
    
    args = parser.parse_args()
    
    semester_start = args.semester_start
    
    if args.action == "parse":
        result = parse_teacher_schedule(args.file, semester_start)
        output = {
            "status": "ok" if result["error"] is None else "error",
            "course_count": len(result["classes"]),
            "semester_start": result["semester_start"],
            "error": result["error"],
        }
    elif args.action == "today":
        date = args.date or datetime.now().strftime("%Y-%m-%d")
        schedule = parse_teacher_schedule(args.file, semester_start)
        classes = get_classes_for_date(schedule, date)
        output = {
            "date": date,
            "course_count": len(classes),
            "courses": classes,
            "error": schedule.get("error"),
        }
    elif args.action == "tomorrow":
        schedule = parse_teacher_schedule(args.file, semester_start)
        base_date = args.date or datetime.now().strftime("%Y-%m-%d")
        result = get_tomorrows_classes(schedule, base_date)
        output = {
            **result,
            "formatted": format_reminder_text(result),
            "error": schedule.get("error"),
        }
    elif args.action == "date":
        if not args.date:
            print("错误: date 操作需要 --date 参数")
            sys.exit(1)
        schedule = parse_teacher_schedule(args.file, semester_start)
        classes = get_classes_for_date(schedule, args.date)
        
        # 计算星期
        date_obj = datetime.strptime(args.date, "%Y-%m-%d")
        weekday_names = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}
        weekday = weekday_names.get(date_obj.isoweekday(), "")
        
        result = {
            "date": args.date,
            "weekday": weekday,
            "classes": classes,
            "formatted": format_reminder_text({
                "date": args.date,
                "weekday": weekday,
                "classes": classes,
            }),
        }
        output = {
            **result,
            "error": schedule.get("error"),
        }
    else:
        output = {"error": f"未知操作: {args.action}"}
    
    if args.output:
        Path(args.output).write_text(json.dumps(output, ensure_ascii=False, indent=2))
        print(f"结果已保存到: {args.output}")
    else:
        print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
