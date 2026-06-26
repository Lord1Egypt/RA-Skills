#!/usr/bin/env python3
"""
export.py
翔云护照 OCR Skill — 结果导出工具

用法:
  # 将标准输入（JSON 格式识别结果）导出为指定格式
  python recognize.py --file passport.jpg | python export.py --format csv --output result.csv
  python recognize.py --file passport.jpg | python export.py --format excel --output result.xlsx
  python recognize.py --file passport.jpg | python export.py --format json --output result.json

  # 或直接传入 JSON 文件
  python export.py --input result.json --format excel --output result.xlsx

  # 批量导出（多个识别结果 JSON 文件）
  python export.py --batch-input results/*.json --format excel --output batch_result.xlsx

  # 推荐：从图片目录读取已缓存的识别结果直接导出（零 API 消耗）
  python export.py --from-dir /path/to/images --format excel --output result.xlsx

支持格式: csv, excel, json
依赖: pip install openpyxl (仅 excel 格式需要)
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime

# ---------- 字段定义 ----------

FIELDS = [
    ("source_image",        "源文件"),
    ("document_type",       "护照类型"),
    ("passport_number",     "护照号码"),
    ("passport_number_mrz", "护照号码(MRZ)"),
    ("name",                "姓名"),
    ("name_english",        "英文姓名"),
    ("name_pinyin",         "拼音"),
    ("sex",                 "性别"),
    ("birth_date",          "出生日期"),
    ("birth_place",         "出生地点"),
    ("birth_place_pinyin",  "出生地点(拼音)"),
    ("issue_date",          "签发日期"),
    ("expiry_date",         "有效期限"),
    ("issue_place",         "签发地点"),
    ("issue_place_pinyin",  "签发地点(拼音)"),
    ("issuing_authority",   "签发机关"),
    ("issuing_authority_ocr","签发机关(OCR)"),
    ("issuing_country_code","签发国代码"),
    ("holder_country_code", "国籍代码"),
    ("id_number",           "身份证号码"),
]


# ---------- 数据读取 ----------

def read_result(input_path: str = None) -> list:
    """从文件或 stdin 读取识别结果，返回记录列表。"""
    if input_path:
        # 兼容 UTF-8 BOM（Windows 下 PowerShell/CMD 常见）
        with open(input_path, "r", encoding="utf-8-sig") as f:
            raw = json.load(f)
    else:
        raw_text = sys.stdin.buffer.read().decode("utf-8-sig")
        raw = json.loads(raw_text)

    # 支持单条记录或记录列表
    if isinstance(raw, list):
        return raw
    return [raw]


def read_batch(file_patterns: list) -> list:
    """批量读取多个 JSON 文件，合并为记录列表。"""
    import glob
    records = []
    for pattern in file_patterns:
        for path in glob.glob(pattern):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    records.extend(data)
                else:
                    records.append(data)
            except Exception as e:
                print(f"[WARN] 跳过文件 {path}: {e}", file=sys.stderr)
    return records


def read_from_dir(dir_path: str) -> list:
    """从目录中读取所有识别结果 JSON 文件（{图片名}.json），返回记录列表。"""
    import glob
    records = []
    pattern = os.path.join(dir_path, "*.json")
    for path in sorted(glob.glob(pattern)):
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
            # 跳过非识别结果文件
            if isinstance(data, list):
                continue
            if isinstance(data, dict):
                # 必须包含 success 字段才是识别结果
                if "success" not in data:
                    continue
                data.setdefault("source_image", os.path.basename(path))
                records.append(data)
        except Exception as e:
            print(f"[WARN] 跳过文件 {path}: {e}", file=sys.stderr)
    return records


# ---------- 导出函数 ----------

def to_row(record: dict) -> dict:
    """将识别结果转换为可导出的行数据。"""
    row = {}
    for key, label in FIELDS:
        row[label] = record.get(key, "")
    row["识别时间"] = record.get("recognized_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row["状态"] = "成功" if record.get("success", True) else f"失败: {record.get('error_message', '')}"
    return row


def export_json(records: list, output_path: str) -> None:
    rows = [to_row(r) for r in records]
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"[OK] 已导出 JSON: {output_path}（共 {len(rows)} 条）")


def export_csv(records: list, output_path: str) -> None:
    rows = [to_row(r) for r in records]
    if not rows:
        print("[WARN] 无数据可导出", file=sys.stderr)
        return
    fieldnames = list(rows[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[OK] 已导出 CSV: {output_path}（共 {len(rows)} 条）")


def export_excel(records: list, output_path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[ERROR] 导出 Excel 需要安装 openpyxl：pip install openpyxl", file=sys.stderr)
        sys.exit(3)

    rows = [to_row(r) for r in records]
    if not rows:
        print("[WARN] 无数据可导出", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "护照识别结果"

    # 表头样式
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    fieldnames = list(rows[0].keys())

    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(fieldnames, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name, ""))

    # 自适应列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    print(f"[OK] 已导出 Excel: {output_path}（共 {len(rows)} 条）")


# ---------- 主程序 ----------

EXPORTERS = {
    "json":  export_json,
    "csv":   export_csv,
    "excel": export_excel,
}


def main():
    parser = argparse.ArgumentParser(description="翔云护照 OCR 识别结果导出工具")
    parser.add_argument("--format",  choices=["csv", "excel", "json"], default="csv", help="导出格式（默认 csv）")
    parser.add_argument("--output",  required=True, help="输出文件路径")
    parser.add_argument("--input",   help="输入 JSON 文件路径（不指定则从 stdin 读取）")
    parser.add_argument("--batch-input", nargs="+", help="批量输入多个 JSON 文件（支持通配符）")
    parser.add_argument("--from-dir", help="从目录读取识别结果 JSON（图片同级目录中 {图片名}.json），无需重复识别")

    args = parser.parse_args()

    # 读取数据
    try:
        if args.from_dir:
            records = read_from_dir(args.from_dir)
        elif args.batch_input:
            records = read_batch(args.batch_input)
        else:
            records = read_result(args.input)
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON 解析失败: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] 读取数据失败: {e}", file=sys.stderr)
        sys.exit(1)

    if not records:
        print("[WARN] 无有效记录，跳过导出", file=sys.stderr)
        sys.exit(0)

    # 确保输出目录存在
    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)

    # 执行导出
    exporter = EXPORTERS[args.format]
    exporter(records, args.output)


if __name__ == "__main__":
    main()
