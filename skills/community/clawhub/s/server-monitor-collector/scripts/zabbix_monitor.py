#!/usr/bin/env python3
"""
Zabbix 监控数据采集 → XLSX（每主机组一个 Sheet，按内存/CPU 占用率降序）
"""

import json
import csv
import sys
import os
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()  # 加载 ~/.hermes/.env
from urllib.request import urlopen, Request
from urllib.error import URLError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 配置 ==========
ZABBIX_URL = os.environ.get("ZABBIX_URL", "")
ZABBIX_USER = os.environ.get("ZABBIX_USER", "")
ZABBIX_PASSWORD = os.environ.get("ZABBIX_PASSWORD", "")

if not ZABBIX_URL:
    print("[ERROR] ZABBIX_URL is not set. Please configure it in ~/.hermes/.env", file=sys.stderr)
    sys.exit(1)

if not ZABBIX_USER or not ZABBIX_PASSWORD:
    print("[ERROR] ZABBIX_USER / ZABBIX_PASSWORD are not set. Please configure them in ~/.hermes/.env", file=sys.stderr)
    sys.exit(1)

ITEMS_KEY = {
    "memory_avail": "vm.memory.size[available]",
    "memory_total": "vm.memory.size[total]",
    "cpu": "system.cpu.util",
}

EXCLUDE_GROUPS = {"Templates", "Templates/Applications", "Templates/Databases",
                  "Templates/Modules", "Templates/Network devices",
                  "Templates/Operating systems", "Templates/Server hardware",
                  "Templates/Virtualization", "Discovered hosts"}

HERMES_DIR = os.environ.get("HERMES_DIR", os.path.expanduser("~/.hermes"))
CSV_PATH = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.csv")
XLSX_PATH = os.path.join(HERMES_DIR, "cron", "output", "zabbix_monitor.xlsx")

# TOPN: 关注 top n 台机器（内存+CPU 综合排序），0=关闭
TOPN = int(os.environ.get("TOPN", "50"))


# ========== Zabbix API ==========

def api_call(method, params, auth=None):
    payload = {
        "jsonrpc": "2.0",
        "method": method,
        "params": params,
        "id": 1,
    }
    if auth:
        payload["auth"] = auth

    data = json.dumps(payload).encode("utf-8")
    req = Request(ZABBIX_URL, data=data, headers={"Content-Type": "application/json"})

    try:
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except URLError as e:
        print(f"API 请求失败: {e}", file=sys.stderr)
        sys.exit(1)

    if "error" in result:
        print(f"API 错误: {result['error']}", file=sys.stderr)
        sys.exit(1)
    return result.get("result", [])


def fetch_all(auth):
    """获取所有主机+监控数据"""

    # 1. 主机组
    groups = api_call("hostgroup.get", {"output": ["groupid", "name"]}, auth=auth)
    groups = [g for g in groups if g["name"] not in EXCLUDE_GROUPS]
    print(f"有效主机组 ({len(groups)} 个)")
    group_ids = [g["groupid"] for g in groups]

    # 2. 主机
    hosts = api_call("host.get", {
        "output": ["hostid", "name", "host"],
        "groupids": group_ids,
        "selectGroups": ["groupid", "name"],
    }, auth=auth)
    print(f"主机总数: {len(hosts)}")

    # 3. 监控项（分批）
    key_filters = list(ITEMS_KEY.values())
    all_items = []
    host_ids = [h["hostid"] for h in hosts]

    BATCH = 100
    for i in range(0, len(host_ids), BATCH):
        batch_ids = host_ids[i:i+BATCH]
        items = api_call("item.get", {
            "output": ["itemid", "hostid", "key_", "lastvalue"],
            "hostids": batch_ids,
            "filter": {"key_": key_filters},
        }, auth=auth)
        all_items.extend(items)

    print(f"监控项: {len(all_items)} 个")

    # 4. 组装数据
    item_map = {}
    for item in all_items:
        item_map[(item["hostid"], item["key_"])] = item.get("lastvalue", "")

    rows = []
    for host in hosts:
        hid = host["hostid"]
        host_groups = host.get("groups", [])
        gnames = [g["name"] for g in host_groups]

        # 跳过完全属于排除组的机器（同时不属于任何有效组）
        valid_gnames = [n for n in gnames if n not in EXCLUDE_GROUPS]
        if not valid_gnames:
            continue

        # 用第一个有效组名作为该主机的归属组
        gname = valid_gnames[0]

        mem_total = item_map.get((hid, ITEMS_KEY["memory_total"]), "")
        mem_avail = item_map.get((hid, ITEMS_KEY["memory_avail"]), "")
        cpu = item_map.get((hid, ITEMS_KEY["cpu"]), "")

        mem_total_gb = float(mem_total) / (1024**3) if mem_total else None
        mem_avail_gb = float(mem_avail) / (1024**3) if mem_avail else None
        cpu_pct = float(cpu) if cpu else None

        # 内存占用率 = 100 - 可用率
        if mem_avail and mem_total:
            mem_used_pct = (1 - float(mem_avail) / float(mem_total)) * 100
        else:
            mem_used_pct = None

        rows.append({
            "group": gname,
            "name": host["name"],
            "ip": host["host"],
            "mem_avail_gb": mem_avail_gb,
            "mem_total_gb": mem_total_gb,
            "mem_used_pct": mem_used_pct,
            "cpu_pct": cpu_pct,
        })

    return rows


# ========== Excel 生成 ==========

def make_style(bold=False, size=11, color=None, bg_color=None, align="center"):
    font = Font(name="微软雅黑", bold=bold, size=size, color=color or "000000")
    if bg_color:
        fill = PatternFill("solid", fgColor=bg_color)
    else:
        fill = None
    align_obj = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return font, fill, align_obj


def style_header(cell, text):
    cell.value = text
    cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell, value, bg="FFFFFF", font_color="000000", number_fmt=None):
    cell.value = value
    cell.font = Font(name="微软雅黑", size=10, color=font_color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_fmt:
        cell.number_format = number_fmt


def pct_color(pct, bg_base):
    """根据占用率百分比返回(背景色, 字体色)"""
    if pct is None:
        return bg_base, "000000"
    if pct >= 80:
        return "FF4444", "FFFFFF"
    if pct >= 60:
        return "FFAA44", "000000"
    if pct >= 40:
        return "FFEE88", "000000"
    return bg_base, "000000"


def generate_xlsx(rows):
    """生成 xlsx，按主机组分 sheet，每 sheet 按内存+CPU 占用率降序"""
    from collections import defaultdict
    group_rows = defaultdict(list)
    for r in rows:
        group_rows[r["group"]].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 列定义：(列名, 列宽)
    col_defs = [
        ("主机名", 32),
        ("IP", 18),
        ("内存总量(GB)", 14),
        ("内存可用(GB)", 14),
        ("内存占用率(%)", 14),
        ("CPU占用率(%)", 13),
    ]

    # ========== 总览 Sheet ==========
    ws_ov = wb.create_sheet(title="总览")
    ws_ov.cell(row=1, column=1, value="服务器监控总览").font = Font(name="微软雅黑", bold=True, size=14)
    ws_ov.cell(row=1, column=1).alignment = Alignment(horizontal="left")
    ws_ov.row_dimensions[1].height = 24

    ws_ov.cell(row=2, column=1, value=f"采集时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws_ov.cell(row=2, column=1).font = Font(name="微软雅黑", size=10, color="666666")
    ws_ov.cell(row=3, column=1, value=f"共 {len(rows)} 台主机，{len(group_rows)} 个有效主机组")
    ws_ov.cell(row=3, column=1).font = Font(name="微软雅黑", size=10, color="666666")

    ov_headers = ["主机组", "主机数", "内存告警(≥80%)", "CPU告警(≥80%)"]
    ov_widths   = [22, 10, 16, 16]
    for col_idx, hdr in enumerate(ov_headers, start=1):
        style_header(ws_ov.cell(row=5, column=col_idx), hdr)
    ws_ov.row_dimensions[5].height = 20

    for row_idx, (gname, gdata) in enumerate(sorted(group_rows.items()), start=6):
        valid_mem = [r for r in gdata if r["mem_used_pct"] is not None]
        valid_cpu = [r for r in gdata if r["cpu_pct"] is not None]
        mem_alarm = sum(1 for r in valid_mem if r["mem_used_pct"] >= 80)
        cpu_alarm = sum(1 for r in valid_cpu if r["cpu_pct"] >= 80)

        vals = [gname, len(gdata), mem_alarm, cpu_alarm]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws_ov.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col_idx == 3 and mem_alarm > 0:
                cell.fill = PatternFill("solid", fgColor="FF4444")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            elif col_idx == 4 and cpu_alarm > 0:
                cell.fill = PatternFill("solid", fgColor="FF4444")
                cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")

    for col_idx, width in enumerate(ov_widths, start=1):
        ws_ov.column_dimensions[get_column_letter(col_idx)].width = width
    ws_ov.column_dimensions["A"].width = 24

    # ========== 各主机组 Sheet ==========
    for gname, gdata in sorted(group_rows.items()):
        ws = wb.create_sheet(title=gname[:31])
        ws.row_dimensions[1].height = 20

        # 表头
        for col_idx, (hdr, _) in enumerate(col_defs, start=1):
            style_header(ws.cell(row=1, column=col_idx), hdr)

        # 排序：内存占用率降序，再 CPU 降序
        gdata.sort(key=lambda x: (-(x["mem_used_pct"] or 0), -(x["cpu_pct"] or 0)))

        # 数据行
        for row_idx, r in enumerate(gdata, start=2):
            bg = "EEF2FF" if row_idx % 2 == 0 else "FFFFFF"
            mem_bg, mem_fc = pct_color(r.get("mem_used_pct"), bg)
            cpu_bg, cpu_fc = pct_color(r.get("cpu_pct"), bg)

            row_vals = [
                (r["name"], bg, "000000"),
                (r["ip"], bg, "000000"),
                (r["mem_total_gb"], bg, "000000"),
                (r["mem_avail_gb"], bg, "000000"),
                (r["mem_used_pct"], mem_bg, mem_fc),
                (r["cpu_pct"], cpu_bg, cpu_fc),
            ]
            for col_idx, (val, cbg, cfc) in enumerate(row_vals, start=1):
                if val is None:
                    display_val = "N/A"
                    fmt = None
                elif col_idx in (3, 4):
                    display_val = f"{val:.1f}"
                    fmt = '0.0'
                elif col_idx in (5, 6):
                    display_val = val
                    fmt = '0.0'
                else:
                    display_val = val
                    fmt = None
                cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = Font(name="微软雅黑", size=10, color=cfc)
                cell.fill = PatternFill("solid", fgColor=cbg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="CCCCCC")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if fmt:
                    cell.number_format = fmt

        # 列宽
        for col_idx, (_, width) in enumerate(col_defs, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

    # ========== TOPN Sheet ==========
    if TOPN > 0:
        ws_top = wb.create_sheet(title=f"TOP{TOPN}(内存+CPU)")
        ws_top.row_dimensions[1].height = 20

        # 表头
        for col_idx, (hdr, _) in enumerate(col_defs, start=1):
            style_header(ws_top.cell(row=1, column=col_idx), hdr)

        # 合并所有数据，按内存占用率+CPU占用率综合降序
        all_data = []
        for gname, gdata in group_rows.items():
            for r in gdata:
                r = dict(r)  # 复制，避免跨组污染
                r["group"] = gname
                all_data.append(r)

        all_data.sort(key=lambda x: (-(x["mem_used_pct"] or 0), -(x["cpu_pct"] or 0)))
        top_data = all_data[:TOPN]

        for row_idx, r in enumerate(top_data, start=2):
            bg = "EEF2FF" if row_idx % 2 == 0 else "FFFFFF"
            mem_bg, mem_fc = pct_color(r.get("mem_used_pct"), bg)
            cpu_bg, cpu_fc = pct_color(r.get("cpu_pct"), bg)

            row_vals = [
                (r["name"], bg, "000000"),
                (r["ip"], bg, "000000"),
                (r["mem_total_gb"], bg, "000000"),
                (r["mem_avail_gb"], bg, "000000"),
                (r["mem_used_pct"], mem_bg, mem_fc),
                (r["cpu_pct"], cpu_bg, cpu_fc),
            ]
            for col_idx, (val, cbg, cfc) in enumerate(row_vals, start=1):
                if val is None:
                    display_val = "N/A"
                    fmt = None
                elif col_idx in (3, 4):
                    display_val = f"{val:.1f}"
                    fmt = '0.0'
                elif col_idx in (5, 6):
                    display_val = val
                    fmt = '0.0'
                else:
                    display_val = val
                    fmt = None
                cell = ws_top.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = Font(name="微软雅黑", size=10, color=cfc)
                cell.fill = PatternFill("solid", fgColor=cbg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="CCCCCC")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if fmt:
                    cell.number_format = fmt

        for col_idx, (_, width) in enumerate(col_defs, start=1):
            ws_top.column_dimensions[get_column_letter(col_idx)].width = width
        ws_top.column_dimensions["A"].width = 36  # 主机名列稍宽
        ws_top.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"XLSX 已写入: {XLSX_PATH}")


def main():
    # 1. 登录
    auth = api_call("user.login", {
        "user": ZABBIX_USER,
        "password": ZABBIX_PASSWORD,
    })
    print(f"登录成功")

    # 2. 采集数据
    rows = fetch_all(auth)

    # 3. 生成 xlsx
    generate_xlsx(rows)

    # 4. 同时保留 CSV（UTF-8-BOM 编码，兼容 Windows Excel）
    os.makedirs(os.path.dirname(CSV_PATH), exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["主机组", "主机名", "IP", "内存可用(GB)",
                         "内存总量(GB)", "内存占用率(%)", "CPU占用率(%)"])
        for r in rows:
            writer.writerow([
                r["group"], r["name"], r["ip"],
                f"{r['mem_avail_gb']:.1f}" if r['mem_avail_gb'] is not None else "N/A",
                f"{r['mem_total_gb']:.1f}" if r['mem_total_gb'] is not None else "N/A",
                f"{r['mem_used_pct']:.1f}" if r['mem_used_pct'] is not None else "N/A",
                f"{r['cpu_pct']:.1f}" if r['cpu_pct'] is not None else "N/A",
            ])
    print(f"CSV 已写入: {CSV_PATH}")


if __name__ == "__main__":
    main()
