#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI智能结算引擎 v2.1.0
支持：
- 数据源：CSV 和 Excel (xlsx/xls)，支持多 Sheet
- 结算模式：达标瓜分、排名/榜单奖、混合多奖池并行
- 精度：Decimal 高精度均分，禁止 round() 截断
"""

import csv
import json
import re
import os
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum

getcontext().prec = 28


# ---------------------------------------------------------------------------
# 枚举 & 数据类
# ---------------------------------------------------------------------------

class SettlementMode(Enum):
    GUARANTEED = "guaranteed"   # 达标瓜分
    RANKING    = "ranking"      # 排名/榜单奖
    WEIGHTED   = "weighted"     # 权重比例分配
    HYBRID     = "hybrid"       # 混合（多奖池并行，同一奖池内不互斥）


@dataclass
class TopicRule:
    """话题词规则"""
    topics: List[str]
    logic: str = "OR"   # AND / OR

    def check(self, title: str) -> bool:
        if not self.topics:
            return True
        matches = []
        for topic in self.topics:
            pattern = (r'(?:^|[\s,，。！!？?；;：:、])' +
                       re.escape(topic) +
                       r'(?:$|[\s,，。！!？?；;：:、])')
            matches.append(bool(re.search(pattern, title)))
        return all(matches) if self.logic == "AND" else any(matches)


@dataclass
class RankingTier:
    """榜单奖单档"""
    rank_start: int          # 起始排名（1-based）
    rank_end:   int          # 结束排名（含）
    amount:     float        # 每人金额（已知固定值）或 0 表示瓜分
    pool:       float = 0.0  # 若>0 则对 rank_start~rank_end 的人等额瓜分


@dataclass
class AwardPool:
    """单个奖池配置"""
    name:           str
    amount:         float
    mode:           SettlementMode
    # 达标瓜分
    condition:      Optional[Dict] = None
    # 排名奖
    ranking_field:  Optional[str]  = None   # 排序依据字段
    ranking_tiers:  Optional[List[RankingTier]] = None
    # 权重分配
    weight_field:   Optional[str]  = None
    # 话题词过滤
    topic_rule:     Optional[TopicRule] = None
    # 多条件（AND）：满足所有条件才达标
    conditions:     Optional[List[Dict]] = None


@dataclass
class AuthorData:
    """原始聚合数据"""
    author_id:    str
    author_name:  str
    videos:       int   = 0
    total_plays:  int   = 0
    total_likes:  int   = 0
    video_titles: List[str] = field(default_factory=list)
    # 直播数据（可选）
    live_duration:   int = 0    # 直播时长(秒)
    live_count:      int = 0    # 开播场次
    live_sales:      int = 0    # 直播销售额(分)
    # 通用扩展字段（key→value），用于承载任意业务指标
    extra: Dict[str, Any] = field(default_factory=dict)


@dataclass
class SettlementResult:
    """结算产出：每位作者的奖金明细"""
    author_id:    str
    author_name:  str
    videos:       int
    total_plays:  int
    total_likes:  int
    awards:       Dict[str, str]   # 奖池名 → 精确金额字符串（保留完整小数）
    total_amount: str              # 总奖金精确字符串


# ---------------------------------------------------------------------------
# 精度工具
# ---------------------------------------------------------------------------

def _distribute_equal(total_yuan: float, n: int) -> str:
    """
    等额瓜分，返回每人金额的精确字符串（保留刚好整除所需的小数位）。
    保证：Decimal(result) * n == Decimal(str(total_yuan))
    若无有限小数能整除，返回10位小数的字符串（无限循环分数）。
    """
    total = Decimal(str(total_yuan))
    count = Decimal(n)
    per   = total / count

    for digits in range(2, 11):
        fmt = '0.' + '0' * digits
        per_rounded = per.quantize(Decimal(fmt), rounding=ROUND_HALF_UP)
        if per_rounded * count == total:
            return str(per_rounded)

    # 无法整除：保留10位，并在调用方标注
    return str(per.quantize(Decimal('0.0000000001'), rounding=ROUND_HALF_UP))


def _fmt_amount(val: str) -> str:
    """去掉末尾多余零，但保留至少两位小数。"""
    d = Decimal(val)
    # 规范化后检查小数位
    normalized = d.normalize()
    sign, digits, exp = normalized.as_tuple()
    if exp >= 0:
        return str(d.quantize(Decimal('0.01')))
    if exp > -2:
        return str(d.quantize(Decimal('0.01')))
    return str(normalized)


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------

class DataLoader:
    """统一数据加载：CSV 和 Excel，支持多 Sheet"""

    @staticmethod
    def load(file_path: str,
             sheet_name: Optional[str] = None,
             all_sheets: bool = False
             ) -> Dict[str, List[Dict]]:
        """
        返回 {sheet_name: [row_dict, ...]}
        CSV 只有一个 sheet，key 为文件名（无扩展名）。
        all_sheets=True 时返回所有 sheet；否则只返回 sheet_name 指定的 sheet
        （sheet_name=None 则取第一个 sheet）。
        """
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return DataLoader._load_csv(file_path)
        elif ext in ('.xlsx', '.xls', '.xlsm'):
            return DataLoader._load_excel(file_path, sheet_name, all_sheets)
        else:
            raise ValueError(f"不支持的文件格式: {ext}，请使用 CSV 或 Excel")

    @staticmethod
    def _load_csv(file_path: str) -> Dict[str, List[Dict]]:
        rows = []
        for enc in ('utf-8-sig', 'gbk', 'utf-8'):
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    reader = csv.DictReader(f)
                    rows = [row for row in reader]
                break
            except (UnicodeDecodeError, Exception):
                continue
        name = os.path.splitext(os.path.basename(file_path))[0]
        return {name: rows}

    @staticmethod
    def _load_excel(file_path: str,
                    sheet_name: Optional[str],
                    all_sheets: bool) -> Dict[str, List[Dict]]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "读取 Excel 需要安装 openpyxl：pip install openpyxl"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result = {}

        if all_sheets:
            target_sheets = wb.sheetnames
        elif sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"
                )
            target_sheets = [sheet_name]
        else:
            target_sheets = [wb.sheetnames[0]]

        for sname in target_sheets:
            ws = wb[sname]
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(h).strip() if h is not None else '' for h in next(rows_iter, [])]
            rows = []
            for row in rows_iter:
                row_dict = {header[i]: (row[i] if i < len(row) else None)
                            for i in range(len(header))}
                rows.append(row_dict)
            result[sname] = rows

        wb.close()
        return result


# ---------------------------------------------------------------------------
# 数据聚合
# ---------------------------------------------------------------------------

FIELD_ALIASES = {
    'author_id':   ['作者ID', '用户ID', '账号ID', 'authorId', 'uid'],
    'author_name': ['作者名称（最新）', '作者名称', '用户名称', '昵称', 'authorName'],
    'title':       ['视频标题', '标题', 'title'],
    'plays':       ['视频累计外显播放次数', '播放量', '播放次数', 'plays'],
    'likes':       ['视频累计外显点赞次数', '点赞量', '获赞数', 'likes'],
    'live_duration': ['直播时长', '直播累计时长', 'liveDuration'],
    'live_count':    ['开播场次', '直播场次', 'liveCount'],
    'live_sales':    ['直播销售额', '销售额', 'liveSales'],
}


def _resolve_field(row: Dict, field_key: str) -> Any:
    """按别名列表查找字段值，不存在返回 None"""
    for alias in FIELD_ALIASES.get(field_key, [field_key]):
        if alias in row:
            return row[alias]
    return None


def _parse_number(val: Any) -> int:
    if val is None or val == '':
        return 0
    try:
        return int(float(str(val).replace(',', '')))
    except (ValueError, TypeError):
        return 0


def aggregate_authors(rows: List[Dict]) -> Dict[str, AuthorData]:
    """将数据行聚合为 author_id → AuthorData"""
    authors: Dict[str, AuthorData] = {}
    for row in rows:
        aid = str(_resolve_field(row, 'author_id') or '').strip()
        if not aid:
            continue
        if aid not in authors:
            authors[aid] = AuthorData(author_id=aid, author_name='')

        a = authors[aid]
        name = str(_resolve_field(row, 'author_name') or '').strip()
        if name:
            a.author_name = name
        a.videos       += 1
        a.total_plays  += _parse_number(_resolve_field(row, 'plays'))
        a.total_likes  += _parse_number(_resolve_field(row, 'likes'))
        a.live_duration+= _parse_number(_resolve_field(row, 'live_duration'))
        a.live_count   += _parse_number(_resolve_field(row, 'live_count'))
        a.live_sales   += _parse_number(_resolve_field(row, 'live_sales'))
        title = str(_resolve_field(row, 'title') or '').strip()
        if title:
            a.video_titles.append(title)

        # 透传其他字段到 extra
        for k, v in row.items():
            if k not in a.extra:
                a.extra[k] = v

    return authors


# ---------------------------------------------------------------------------
# 结算引擎
# ---------------------------------------------------------------------------

class SettlementEngine:

    def __init__(self, pools: List[AwardPool]):
        self.pools = pools
        self.authors: Dict[str, AuthorData] = {}

    def load_data(self, file_path: str,
                  sheet_name: Optional[str] = None,
                  all_sheets: bool = False,
                  merge: bool = True) -> None:
        """
        加载数据文件。
        all_sheets=True 时合并所有 Sheet 数据（适合多底表场景）。
        merge=True 时与已有数据合并（可多次调用加载不同文件）。
        """
        sheets = DataLoader.load(file_path, sheet_name, all_sheets)
        if not merge:
            self.authors.clear()

        for sname, rows in sheets.items():
            new_authors = aggregate_authors(rows)
            for aid, data in new_authors.items():
                if aid in self.authors:
                    # 合并：累加数值字段
                    existing = self.authors[aid]
                    if data.author_name:
                        existing.author_name = data.author_name
                    existing.videos        += data.videos
                    existing.total_plays   += data.total_plays
                    existing.total_likes   += data.total_likes
                    existing.live_duration += data.live_duration
                    existing.live_count    += data.live_count
                    existing.live_sales    += data.live_sales
                    existing.video_titles  += data.video_titles
                    existing.extra.update(data.extra)
                else:
                    self.authors[aid] = data

    def process(self) -> List[SettlementResult]:
        """执行所有奖池结算，返回合并结果。"""
        # 初始化结果字典
        results: Dict[str, SettlementResult] = {}
        for aid, a in self.authors.items():
            results[aid] = SettlementResult(
                author_id=aid,
                author_name=a.author_name,
                videos=a.videos,
                total_plays=a.total_plays,
                total_likes=a.total_likes,
                awards={},
                total_amount='0'
            )

        for pool in self.pools:
            if pool.mode == SettlementMode.GUARANTEED:
                self._process_guaranteed(pool, results)
            elif pool.mode == SettlementMode.RANKING:
                self._process_ranking(pool, results)
            elif pool.mode == SettlementMode.WEIGHTED:
                self._process_weighted(pool, results)

        # 计算每人总奖金
        for r in results.values():
            total = sum(Decimal(v) for v in r.awards.values())
            r.total_amount = str(total.normalize() if total != 0 else Decimal('0'))

        # 只保留有奖金的作者，按总金额降序
        winners = [r for r in results.values()
                   if Decimal(r.total_amount) > 0]
        winners.sort(key=lambda x: Decimal(x.total_amount), reverse=True)
        return winners

    # ---- 达标瓜分 ----
    def _process_guaranteed(self, pool: AwardPool,
                             results: Dict[str, SettlementResult]) -> None:
        qualified_ids = []
        for aid, a in self.authors.items():
            if pool.topic_rule and not self._check_topic(a, pool.topic_rule):
                continue
            if self._check_conditions(a, pool):
                qualified_ids.append(aid)

        if not qualified_ids:
            return

        n = len(qualified_ids)
        per_str = _distribute_equal(pool.amount, n)
        for aid in qualified_ids:
            results[aid].awards[pool.name] = per_str

    # ---- 排名/榜单奖 ----
    def _process_ranking(self, pool: AwardPool,
                         results: Dict[str, SettlementResult]) -> None:
        if not pool.ranking_field or not pool.ranking_tiers:
            return

        # 先过滤话题词
        candidates = [
            a for a in self.authors.values()
            if not pool.topic_rule or self._check_topic(a, pool.topic_rule)
        ]
        if pool.condition or pool.conditions:
            candidates = [a for a in candidates if self._check_conditions(a, pool)]

        # 按 ranking_field 降序排列
        field_map = {
            '播放量': 'total_plays',
            '播放': 'total_plays',
            '点赞': 'total_likes',
            '获赞': 'total_likes',
            '直播时长': 'live_duration',
            '开播场次': 'live_count',
            '销售额': 'live_sales',
            '作品数': 'videos',
        }
        attr = field_map.get(pool.ranking_field, pool.ranking_field)
        candidates.sort(
            key=lambda a: (getattr(a, attr, 0) or a.extra.get(attr, 0)),
            reverse=True
        )

        for tier in pool.ranking_tiers:
            start_idx = tier.rank_start - 1   # 0-based
            end_idx   = tier.rank_end          # exclusive
            tier_candidates = candidates[start_idx:end_idx]
            if not tier_candidates:
                continue

            if tier.pool > 0:
                # 该档奖金瓜分
                n = len(tier_candidates)
                per_str = _distribute_equal(tier.pool, n)
            else:
                # 固定每人金额
                per_str = _distribute_equal(tier.amount, 1)

            for a in tier_candidates:
                results[a.author_id].awards[pool.name] = per_str

    # ---- 权重分配 ----
    def _process_weighted(self, pool: AwardPool,
                          results: Dict[str, SettlementResult]) -> None:
        if not pool.weight_field:
            return

        field_map = {
            '播放量': 'total_plays',
            '点赞': 'total_likes',
            '直播时长': 'live_duration',
            '销售额': 'live_sales',
        }
        attr = field_map.get(pool.weight_field, pool.weight_field)

        candidates = [
            a for a in self.authors.values()
            if not pool.topic_rule or self._check_topic(a, pool.topic_rule)
        ]
        if pool.condition or pool.conditions:
            candidates = [a for a in candidates if self._check_conditions(a, pool)]

        total_weight = sum(getattr(a, attr, 0) or a.extra.get(attr, 0)
                           for a in candidates)
        if total_weight == 0:
            return

        pool_d = Decimal(str(pool.amount))
        tw_d   = Decimal(str(total_weight))
        for a in candidates:
            w = Decimal(str(getattr(a, attr, 0) or a.extra.get(attr, 0)))
            award = (pool_d * w / tw_d).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if award > 0:
                results[a.author_id].awards[pool.name] = str(award)

    # ---- 辅助 ----
    def _check_topic(self, a: AuthorData, rule: TopicRule) -> bool:
        return any(rule.check(t) for t in a.video_titles)

    def _check_conditions(self, a: AuthorData, pool: AwardPool) -> bool:
        conditions = pool.conditions or ([pool.condition] if pool.condition else [])
        for cond in conditions:
            field_name = cond.get('field', '')
            op         = cond.get('op', '>=')
            value      = cond.get('value', 0)
            field_map  = {
                '播放量': 'total_plays', '播放': 'total_plays',
                '获赞': 'total_likes', '点赞': 'total_likes',
                '作品数': 'videos', '作品': 'videos',
                '直播时长': 'live_duration', '开播场次': 'live_count',
                '销售额': 'live_sales',
            }
            attr = field_map.get(field_name, field_name)
            actual = getattr(a, attr, None)
            if actual is None:
                actual = a.extra.get(attr, 0)

            ok = False
            if op in ('>=', '≥'):
                ok = actual >= value
            elif op in ('>', '>'):
                ok = actual > value
            elif op in ('<=', '≤'):
                ok = actual <= value
            elif op == '<':
                ok = actual < value
            elif op == '==':
                ok = actual == value
            if not ok:
                return False
        return True


# ---------------------------------------------------------------------------
# 自然语言规则解析器（AI 辅助增强版）
# ---------------------------------------------------------------------------

class RuleParser:
    """
    支持解析：
    - 达标瓜分（单/多条件）
    - 多奖池并行
    - 排名/榜单奖
    - 话题词过滤
    AI 在处理复杂规则时可直接构造 AwardPool 列表，无需通过此解析器。
    """

    @staticmethod
    def parse_topic_rule(text: str) -> Optional[TopicRule]:
        if not any(k in text for k in ['话题', '#', '携带']):
            return None
        topics = re.findall(r'#[^#\s,，和或]+', text)
        if not topics:
            return None
        logic = "AND" if any(k in text for k in ['且', '和', '同时', 'AND', '都']) else "OR"
        return TopicRule(topics=topics, logic=logic)

    @staticmethod
    def parse(rule_text: str) -> List[AwardPool]:
        """简化解析器：处理常见单奖池达标瓜分。复杂规则由 AI 直接构造。"""
        pools: List[AwardPool] = []

        amount_match = re.search(r'总奖金(\d+(?:\.\d+)?)[万]?元?', rule_text)
        if not amount_match:
            return pools

        raw_amount = float(amount_match.group(1))
        if '万' in rule_text[amount_match.start():amount_match.end() + 2]:
            raw_amount *= 10000

        conditions = []

        videos_match = re.search(r'作品[数量]?[≥>=](\d+)', rule_text)
        if videos_match:
            conditions.append({'field': '作品数', 'op': '>=', 'value': int(videos_match.group(1))})

        plays_match = re.search(r'播放[量]?[≥>=](\d+)万', rule_text)
        if plays_match:
            conditions.append({'field': '播放量', 'op': '>=', 'value': int(plays_match.group(1)) * 10000})
        else:
            plays_match2 = re.search(r'播放[量]?[≥>=](\d+)', rule_text)
            if plays_match2:
                conditions.append({'field': '播放量', 'op': '>=', 'value': int(plays_match2.group(1))})

        topic_rule = RuleParser.parse_topic_rule(rule_text)

        pool = AwardPool(
            name="达标瓜分奖池",
            amount=raw_amount,
            mode=SettlementMode.GUARANTEED,
            conditions=conditions if conditions else None,
            topic_rule=topic_rule,
        )
        pools.append(pool)
        return pools


# ---------------------------------------------------------------------------
# 格式化 & 输出
# ---------------------------------------------------------------------------

def format_rule_understanding(pools: List[AwardPool]) -> str:
    lines = ['=' * 80, '📋 规则理解确认', '=' * 80, '']
    for idx, pool in enumerate(pools, 1):
        lines.append(f"【奖池 {idx}】{pool.name}")
        lines.append(f"  💰 奖池金额: {pool.amount:,.2f} 元")
        lines.append(f"  📊 结算模式: {pool.mode.value}")
        if pool.conditions:
            for c in pool.conditions:
                lines.append(f"  ✅ 达标条件: {c['field']} {c['op']} {c['value']:,}")
        elif pool.condition:
            c = pool.condition
            lines.append(f"  ✅ 达标条件: {c['field']} {c['op']} {c['value']:,}")
        if pool.topic_rule:
            logic_text = "且" if pool.topic_rule.logic == "AND" else "或"
            topics_text = f" {logic_text} ".join(pool.topic_rule.topics)
            lines.append(f"  🏷️  话题词要求: {topics_text} ({pool.topic_rule.logic})")
        if pool.ranking_tiers:
            lines.append(f"  🏆 排序字段: {pool.ranking_field}")
            for t in pool.ranking_tiers:
                if t.pool > 0:
                    lines.append(f"     第{t.rank_start}-{t.rank_end}名: 瓜分 {t.pool:,.2f}元")
                else:
                    lines.append(f"     第{t.rank_start}-{t.rank_end}名: 每人 {t.amount:,.2f}元")
        lines.append('')
    lines += ['=' * 80, '❓ 请确认以上规则理解是否正确？',
              '   ✅ 回复「确认」或「正确」开始结算',
              '   ✏️  提出修改意见，我会更新规则理解',
              '=' * 80]
    return '\n'.join(lines)


def export_to_csv(results: List[SettlementResult],
                  output_path: str,
                  stats: Dict) -> None:
    """导出结算结果到 CSV。金额保留完整精度，不四舍五入截断到两位小数。"""
    # 收集所有奖池列名
    pool_names: List[str] = []
    for r in results:
        for k in r.awards:
            if k not in pool_names:
                pool_names.append(k)

    header = ['序号', '作者ID', '作者名称', '发布作品数', '累计播放量', '累计获赞']
    header += [f'{p}(元)' for p in pool_names]
    header += ['总奖金(元)']

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for idx, r in enumerate(results, 1):
            row = [idx, r.author_id, r.author_name, r.videos, r.total_plays, r.total_likes]
            for p in pool_names:
                row.append(r.awards.get(p, ''))
            row.append(r.total_amount)
            writer.writerow(row)

        # 汇总行
        summary = ['汇总', '', '',
                   stats['total_videos'], stats['total_plays'], stats['total_likes']]
        for p in pool_names:
            pool_total = sum(Decimal(r.awards[p]) for r in results if p in r.awards)
            summary.append(str(pool_total.normalize()))
        summary.append(stats['total_award'])
        writer.writerow(summary)


def print_summary(results: List[SettlementResult], stats: Dict) -> None:
    print('=' * 80)
    print('结算结果摘要')
    print('=' * 80)
    print(f"\n奖池配置：")
    for pool in stats.get('pools', []):
        print(f"  - {pool['name']}: {pool['amount']:,.2f}元 ({pool['mode']})")
    print(f"\n统计信息：")
    print(f"  - 参与作者总数：{stats['total_authors']} 人")
    print(f"  - 获奖作者数：{stats['qualified_authors']} 人")
    print(f"  - 累计发布作品：{stats['total_videos']} 条")
    print(f"  - 累计播放量：{stats['total_plays']:,} 次")
    print(f"  - 总奖金发放：{stats['total_award']} 元")
    print('\n' + '=' * 80)
    print(f"\n获奖作者前10名：")
    print(f"{'序号':<5}{'作者ID':<16}{'作者名称':<16}{'作品数':<8}{'播放量':<12}{'总奖金'}")
    print('-' * 80)
    for i, r in enumerate(results[:10], 1):
        print(f"{i:<5}{r.author_id:<16}{r.author_name:<16}{r.videos:<8}{r.total_plays:<12}{r.total_amount}")
    print('\n' + '=' * 80)


# ---------------------------------------------------------------------------
# 高层 API
# ---------------------------------------------------------------------------

def process_settlement(
    file_path: str,
    rule_text: str,
    sheet_name: Optional[str] = None,
    all_sheets: bool = False,
) -> Tuple[List[SettlementResult], Dict]:
    """
    一行调用完成结算。
    - file_path: CSV 或 Excel 路径
    - rule_text: 自然语言规则（简单场景自动解析）
    - sheet_name: 指定 Sheet（Excel）
    - all_sheets: 合并所有 Sheet
    """
    pools = RuleParser.parse(rule_text)
    if not pools:
        raise ValueError("无法解析规则，请使用 build_engine() 手动构造奖池配置")
    return _run_engine(file_path, pools, sheet_name, all_sheets)


def build_engine(pools: List[AwardPool]) -> SettlementEngine:
    """AI 直接构造奖池时使用，返回 SettlementEngine 供进一步操作。"""
    return SettlementEngine(pools)


def _run_engine(
    file_path: str,
    pools: List[AwardPool],
    sheet_name: Optional[str],
    all_sheets: bool,
) -> Tuple[List[SettlementResult], Dict]:
    engine = SettlementEngine(pools)
    engine.load_data(file_path, sheet_name=sheet_name, all_sheets=all_sheets)
    results = engine.process()

    stats = {
        'total_authors':    len(engine.authors),
        'qualified_authors': len(results),
        'total_videos':     sum(r.videos for r in results),
        'total_plays':      sum(r.total_plays for r in results),
        'total_likes':      sum(r.total_likes for r in results),
        'total_award':      str(sum(Decimal(r.total_amount) for r in results).normalize()),
        'pools': [{'name': p.name, 'amount': p.amount, 'mode': p.mode.value}
                  for p in pools],
    }
    return results, stats


# ---------------------------------------------------------------------------
# 命令行入口
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("用法: python settlement_engine.py <数据文件> <规则描述> [sheet名称]")
        print("示例: python settlement_engine.py data.xlsx '总奖金2万，播放量≥3万作者瓜分' Sheet1")
        sys.exit(1)

    _file   = sys.argv[1]
    _rule   = sys.argv[2]
    _sheet  = sys.argv[3] if len(sys.argv) > 3 else None
    _all    = '--all-sheets' in sys.argv

    try:
        _results, _stats = process_settlement(_file, _rule,
                                               sheet_name=_sheet,
                                               all_sheets=_all)
        print_summary(_results, _stats)
        _out = _file.rsplit('.', 1)[0] + '_结算结果.csv'
        export_to_csv(_results, _out, _stats)
        print(f"\n✓ 结算结果已保存至: {_out}")
    except Exception as e:
        print(f"❌ 结算失败: {e}")
        import traceback
        traceback.print_exc()
