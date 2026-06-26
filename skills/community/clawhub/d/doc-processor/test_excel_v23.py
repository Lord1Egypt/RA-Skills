#!/usr/bin/env python3
"""Excel 模板支持测试 - v2.3"""

import sys
import tempfile
from pathlib import Path

# 添加 Skill 目录到 Python 路径
sys.path.insert(0, str(Path(__file__).parent))

from doc_processor import DocumentProcessor

def test_excel_fill():
    """测试 Excel 模板填充"""
    print("=" * 70)
    print("Excel 模板支持测试 - v2.3")
    print("=" * 70)
    
    processor = DocumentProcessor()
    
    # 创建测试 Excel 模板
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = '{{company_name}} 发票'
    ws['A3'] = '客户名称：{{customer_name}}'
    ws['A4'] = '发票号码：{{invoice_number}}'
    ws['A5'] = '日期：{{invoice_date}}'
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        template_path = f.name
        wb.save(template_path)
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name
    
    try:
        # 填充模板
        data = {
            'company_name': '某某公司',
            'customer_name': '张三',
            'invoice_number': 'INV2026001',
            'invoice_date': '2026-03-26'
        }
        
        result = processor.fill_template(template_path, data, output_path)
        print(f"✅ Excel 模板填充成功")
        print(f"   输出：{result}")
        
        # 验证填充结果
        from openpyxl import load_workbook
        filled_wb = load_workbook(output_path)
        filled_ws = filled_wb.active
        
        print(f"\n📊 验证填充结果:")
        print(f"   A1: {filled_ws['A1'].value}")
        print(f"   A3: {filled_ws['A3'].value}")
        print(f"   A4: {filled_ws['A4'].value}")
        print(f"   A5: {filled_ws['A5'].value}")
        
        # 断言验证
        assert '某某公司' in filled_ws['A1'].value, "公司名未填充"
        assert '张三' in filled_ws['A3'].value, "客户名未填充"
        assert 'INV2026001' in filled_ws['A4'].value, "发票号未填充"
        assert '2026-03-26' in filled_ws['A5'].value, "日期未填充"
        
        print(f"\n✅ 所有验证通过！")
        
    except Exception as e:
        print(f"❌ 测试失败：{e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        Path(template_path).unlink(missing_ok=True)
        Path(output_path).unlink(missing_ok=True)
    
    return True

if __name__ == "__main__":
    success = test_excel_fill()
    sys.exit(0 if success else 1)
