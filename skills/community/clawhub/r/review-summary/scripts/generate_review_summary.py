#!/usr/bin/env python3
"""
评审汇总表生成器 — Generate 4-sheet review summary from 广联达导表 data.

Input:  2+ Excel files in a data directory (分部分项汇总审核导表.xlsx, 总审核导表.xlsx)
Output: 4-sheet .xlsx file matching 评审汇总表模板 format.
"""

import os, re, sys, argparse
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ─── Constants ───────────────────────────────────────────────────────
CAT_LABELS = {
    1: "1重复计取审减", 2: "2漏项审增", 3: "3工程量审增减",
    4: "4定额子目审增减", 5: "5材料价格审增减", 6: "6其他", 7: "7基数调整审增减"
}
CAT_SHORT = {1: "重复计取", 2: "漏项", 3: "工程量", 4: "定额子目", 5: "材料价格", 6: "其他", 7: "基数调整"}
CAT_ORDER = [(1, "一、重复计取"), (2, "二、因漏项审增"), (3, "三、因工程量审增减"),
             (4, "四、因定额子目套用错误审增减"), (5, "五、因材料价格审增减"),
             (6, "六、因其他零星项和计算基数调整审增减")]

# Excel styling
THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
F_TITLE = Font(name='宋体', size=14, bold=True)
F_HEAD = Font(name='宋体', size=10, bold=True)
F_BODY = Font(name='宋体', size=9)
F_SMALL = Font(name='宋体', size=8)
F_BOLD = Font(name='宋体', size=10, bold=True)
A_CENTER = Alignment(horizontal='center', vertical='center')
NUM_FMT = '#,##0.00'
PCT_FMT = '0.00%'

COMPANY_SUFFIXES = ['-建基', '-晖泽', '-元晟', '-诚辉', '-公正', '-日新', '-华明']


# ─── Classification ──────────────────────────────────────────────────
def classify_item(tag, reason, name, desc):
    """Classify a change item into category 1-7 using keyword priority matching."""
    combined = f"{tag or ''}{reason or ''}{name or ''}{desc or ''}"

    if re.search(r'基数|税金|措施费|总价措施|安文|环境保护费|文明施工费|临时设施|扬尘|夜间|二次搬运|冬雨季', combined):
        return (7, CAT_LABELS[7])
    if re.search(r'重复', combined):
        return (1, CAT_LABELS[1])
    if re.search(r'定额|子目|套用|调整为|人工乘', combined):
        return (4, CAT_LABELS[4])
    if re.search(r'价格|材料|信息价|市场询价|材料价格', combined):
        return (5, CAT_LABELS[5])
    if '[增项]' in str(tag or '') or re.search(r'漏项|补项|未报送', combined):
        return (2, CAT_LABELS[2])
    if any(x in str(tag or '') for x in ('[调量]', '[调量,调价]', '[减项]')):
        return (3, CAT_LABELS[3])
    if any(x in str(tag or '') for x in ('[调价]', '[调项]')):
        # 细分：纯材料类 → 材料价格，设备/安装类 → 定额子目
        name_desc = f"{name or ''}{desc or ''}"
        if re.search(r'防静电|绝缘垫|防火涂料|防火堵料|防火封堵|油漆|抹灰|砌体|土方|回填|混凝土|钢筋|模板|门窗|涂料|防水|保温|石材|瓷砖|地板|砂$|石$|砖$|水泥|碎石|沥青|井盖', name_desc):
            return (5, CAT_LABELS[5])
        return (4, CAT_LABELS[4])
    return (6, CAT_LABELS[6])


# ─── Helpers ─────────────────────────────────────────────────────────
def safe_float(v):
    if v is None: return 0.0
    try: return float(str(v).replace("'", "").replace('"', ''))
    except (ValueError, TypeError): return 0.0


def clean_name(name):
    for sfx in COMPANY_SUFFIXES:
        name = name.replace(sfx, '')
    return name


def hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.border, cell.alignment = F_HEAD, HEADER_FILL, THIN_BORDER, A_CENTER


def cell(ws, r, c, val=None, font=F_BODY, align=None, fmt=None):
    cl = ws.cell(row=r, column=c)
    if val is not None: cl.value = val
    cl.font, cl.border = font, THIN_BORDER
    if align: cl.alignment = align
    if fmt: cl.number_format = fmt
    return cl


def detect_col_map(ws):
    """Scan header row to map column names → 0-based indices. Returns dict."""
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    mapping = {}
    for i, v in enumerate(header_row):
        v = str(v).strip() if v else ''
        mapping[v] = i
    return mapping


# ─── Data Loading ────────────────────────────────────────────────────
def load_source_data(data_dir):
    unit_projects = []
    items = []

    # 1. 总审核导表
    zs_path = os.path.join(data_dir, "总审核导表.xlsx")
    if os.path.exists(zs_path):
        wb = openpyxl.load_workbook(zs_path)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            seq, name = row[0], row[1]
            if not name: continue
            seq_s = str(seq)
            if seq_s == '合计': continue
            try:
                if float(seq_s) == int(float(seq_s)): continue  # skip integer parent rows
            except (ValueError, TypeError): continue
            send = safe_float(row[2]); review = safe_float(row[3]); diff = safe_float(row[4])
            unit_projects.append({'seq': seq_s, 'name': clean_name(str(name)),
                                  'send_amt': send, 'review_amt': review, 'diff': diff})

    # 2. 分部分项汇总审核导表 — auto-detect column layout from header
    fb_path = os.path.join(data_dir, "分部分项汇总审核导表.xlsx")
    if os.path.exists(fb_path):
        wb = openpyxl.load_workbook(fb_path)
        ws = wb[wb.sheetnames[0]]

        # Scan first few data rows to find the real column positions
        # (广联达 exports have empty columns that shift indices)
        col_map = {}
        for r in range(1, min(5, ws.max_row + 1)):
            for c_idx, cell in enumerate(list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]):
                if cell is not None:
                    col_map[str(cell).strip()] = c_idx

        # Find key column positions from detected labels
        idx = {}
        for label, col_name in [('seq', '序号'), ('unit_eng', '单位工程'), ('code', '项目编码'),
                                 ('name', '项目名称'), ('desc', '特征描述'), ('unit', '单位'),
                                 ('send_qty', '送审工程量'), ('send_price', '送审综合单价'),
                                 ('send_total', '送审综合合价'), ('review_qty', '审定工程量'),
                                 ('review_price', '审定综合单价'), ('review_total', '审定综合合价'),
                                 ('diff', '增减金额'), ('note', '增减说明'), ('sign', '增/减符'),
                                 ('analysis', '增减分析'), ('reason', '增减原因')]:
            idx[label] = col_map.get(col_name)
            if idx[label] is None:
                # Fallback: try partial match
                for k, v in col_map.items():
                    if col_name in k:
                        idx[label] = v
                        break

        # If detection failed, use known offset layout
        if idx.get('send_qty') is None:
            idx = {'seq': 1, 'unit_eng': 2, 'code': 3, 'name': 4, 'desc': 5, 'unit': 7,
                   'send_qty': 8, 'send_price': 9, 'send_total': 10, 'review_qty': 11,
                   'review_price': 12, 'review_total': 13, 'diff': 14, 'note': 15,
                   'sign': 16, 'analysis': 18, 'reason': 19}

        def getv(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        for row in ws.iter_rows(min_row=4, values_only=True):
            seq = getv(row, 'seq')
            if seq is None: continue

            # Skip subtotal rows (integer seq without detail columns)
            try:
                if float(str(seq)) == int(float(str(seq))) and getv(row, 'note') is None:
                    continue
            except (ValueError, TypeError):
                if isinstance(seq, str) and not any(c.isdigit() for c in str(seq)):
                    continue

            diff = safe_float(getv(row, 'diff'))
            if diff == 0: continue

            unit_eng = str(getv(row, 'unit_eng') or '')
            code = str(getv(row, 'code') or '')
            name = str(getv(row, 'name') or '')
            desc = str(getv(row, 'desc') or '')
            unit = str(getv(row, 'unit') or '')
            note = str(getv(row, 'note') or '')
            analysis = str(getv(row, 'analysis') or '')
            reason = str(getv(row, 'reason') or '')

            cat_id, cat_name = classify_item(note, reason, name, desc)

            items.append({
                'unit_eng': unit_eng, 'code': code, 'name': name, 'desc': desc, 'unit': unit,
                'send_qty': safe_float(getv(row, 'send_qty')),
                'send_price': safe_float(getv(row, 'send_price')),
                'send_total': safe_float(getv(row, 'send_total')),
                'review_qty': safe_float(getv(row, 'review_qty')),
                'review_price': safe_float(getv(row, 'review_price')),
                'review_total': safe_float(getv(row, 'review_total')),
                'diff': diff, 'note': note, 'analysis': analysis, 'reason': reason,
                'cat_id': cat_id, 'cat_name': cat_name,
            })

    return unit_projects, items


# ─── Sheet Generators ────────────────────────────────────────────────
def sheet1_draft(ws, project, companies, send, review, notes=""):
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value=f"{project}\n评审底稿").font = F_TITLE
    ws.cell(row=1, column=1).alignment = A_CENTER
    ws.cell(row=2, column=1, value="评审公司").font = F_HEAD
    ws.cell(row=2, column=2, value=companies).font = F_BODY
    ws.cell(row=4, column=1, value="评审过程分析").font = F_HEAD
    ws.cell(row=4, column=2, value="实际报送金额：").font = F_BODY
    cell(ws, 4, 3, send, fmt=NUM_FMT)
    if notes: ws.cell(row=4, column=5, value=notes).font = F_SMALL
    reduction = send - review
    ws.cell(row=5, column=2, value="初审金额：").font = F_BODY
    cell(ws, 5, 3, review, fmt=NUM_FMT)
    ws.cell(row=6, column=2, value="审减金额：").font = F_BODY
    cell(ws, 6, 3, reduction, fmt=NUM_FMT)
    ws.cell(row=7, column=2, value="审减率：").font = F_BODY
    cell(ws, 7, 3, reduction / send if send else 0, fmt=PCT_FMT)
    for r, c, t in [(19, 2, "初审："), (19, 4, "审核："), (19, 6, "日期：")]:
        ws.cell(row=r, column=c, value=t).font = F_BODY


def sheet2_summary(ws, project, unit_projects, items_by_unit_cat, send, bldg_area):
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value=f"{project}\n评审汇总表").font = F_TITLE
    ws.cell(row=1, column=1).alignment = A_CENTER
    cell(ws, 1, 12, send, fmt=NUM_FMT)

    headers = ['序号', '名称', '送审金额', '审定金额', '审增减金额', '审增减率',
               '建筑面积', '单方指标', '备注', '分配责任人', '互审人', '索引列']
    for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 12)

    review = sum(up['review_amt'] for up in unit_projects)
    diff = review - send
    r = 3
    ws.cell(row=r, column=1, value="一")
    ws.cell(row=r, column=2, value="电力工程合计").font = F_BOLD
    cell(ws, r, 3, send, fmt=NUM_FMT)
    cell(ws, r, 4, review, fmt=NUM_FMT)
    cell(ws, r, 5, diff, fmt=NUM_FMT)
    cell(ws, r, 6, diff / send if send else 0, fmt=PCT_FMT)
    cell(ws, r, 7, bldg_area, fmt='#,##0.00')
    cell(ws, r, 8, review / bldg_area if bldg_area else 0, fmt='#,##0.00')

    r = 4
    for i, up in enumerate(unit_projects):
        seq_main = i + 1
        up_name = clean_name(up['name'])
        ws.cell(row=r, column=1, value=float(seq_main))
        ws.cell(row=r, column=2, value=up_name)
        cell(ws, r, 3, up['send_amt'], fmt=NUM_FMT)
        cell(ws, r, 4, up['review_amt'], fmt=NUM_FMT)
        cell(ws, r, 5, up['diff'], fmt=NUM_FMT)
        cell(ws, r, 6, up['diff'] / up['send_amt'] if up['send_amt'] else 0, fmt=PCT_FMT)
        ws.cell(row=r, column=12, value=up_name)
        r += 1

        for cat_id in range(1, 8):
            cat_total = items_by_unit_cat.get((up_name, cat_id), 0)
            ws.cell(row=r, column=1, value=seq_main + cat_id * 0.1)
            w = abs(cat_total) / 10000
            if cat_id == 2:
                label = f"因漏项审增{w:.2f}万元"
            elif cat_total == 0:
                label = f"因{CAT_SHORT[cat_id]}审减0万元"
            else:
                direction = "审减" if cat_total < 0 else "审增"
                label = f"因{CAT_SHORT[cat_id]}{direction}{w:.2f}万元"
            ws.cell(row=r, column=2, value=label)
            cell(ws, r, 3, cat_total, fmt=NUM_FMT)
            cell(ws, r, 5, cat_total, fmt=NUM_FMT)
            ws.cell(row=r, column=9, value=CAT_LABELS[cat_id])
            ws.cell(row=r, column=12, value=up_name)
            r += 1
    return r


def sheet3_data_ref(ws, items):
    headers = ['L', '序号', '单位工程', '项目编码', '项目名称', '特征描述', '单位',
               '送审工程量', '送审综合单价', '送审综合合价', '审定工程量', '审定综合单价',
               '审定综合合价', '增减金额', '审增减原因（不要合并单元格）', '审增减类别',
               '分配人', '单项工程', '备注']
    ws.cell(row=1, column=2, value="审增减明细导表").font = F_HEAD
    ws.cell(row=1, column=15, value="填写区").font = F_SMALL
    for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 19)
    r = 3
    for seq, item in enumerate(items, 1):
        cell(ws, r, 1); cell(ws, r, 2, seq)
        ws.cell(row=r, column=3, value=item['unit_eng'])
        ws.cell(row=r, column=4, value=item['code'])
        ws.cell(row=r, column=5, value=item['name'])
        ws.cell(row=r, column=6, value=item['desc'][:500] if item['desc'] else '')
        ws.cell(row=r, column=7, value=item['unit'])
        for c2, k in [(8, 'send_qty'), (9, 'send_price'), (10, 'send_total'),
                       (11, 'review_qty'), (12, 'review_price'), (13, 'review_total'), (14, 'diff')]:
            cell(ws, r, c2, item[k], fmt=NUM_FMT)
        ws.cell(row=r, column=15, value=item['reason'][:500] if item['reason'] else item['analysis'])
        ws.cell(row=r, column=16, value=item['cat_name'])
        ws.cell(row=r, column=18, value='电力工程')
        r += 1
    return r


def sheet4_detail(ws, project, items_by_category, send, review):
    ws.merge_cells('A1:P1')
    ws.cell(row=1, column=1, value=f"审增减明细：{project}").font = F_TITLE

    h1 = ['序号', '单位工程', '项目编码', '名称', '特征描述', '计量单位',
          '送审（元）', '', '', '审定（元）', '', '', '审增减金额\n（-为审减）',
          '备注（审增减原因）', '审增减类别', '分配人']
    for c, h in enumerate(h1, 1): ws.cell(row=2, column=c, value=h)
    hdr(ws, 2, 16)
    for i, sh in enumerate(['数量', '单价', '合价', '数量', '单价', '合价']):
        cl = ws.cell(row=3, column=7 + i, value=sh)
        cl.font, cl.fill, cl.border, cl.alignment = F_HEAD, HEADER_FILL, THIN_BORDER, A_CENTER

    r = 4
    ws.cell(row=r, column=1, value="总合计").font = F_BOLD
    ws.cell(row=r, column=7, value="本项目送审金额：")
    cell(ws, r, 9, send, fmt=NUM_FMT)
    ws.cell(row=r, column=10, value="审定金额：")
    cell(ws, r, 12, review, fmt=NUM_FMT)
    cell(ws, r, 13, review - send, fmt=NUM_FMT)
    r += 1

    for cat_id, cat_title in CAT_ORDER:
        cat_items = items_by_category.get(cat_id, [])
        cat_send = sum(it['send_total'] for it in cat_items)
        cat_review = sum(it['review_total'] for it in cat_items)
        cat_diff = cat_review - cat_send

        ws.cell(row=r, column=1, value=cat_title).font = F_BOLD
        cell(ws, r, 9, cat_send, font=F_BOLD, fmt=NUM_FMT)
        cell(ws, r, 12, cat_review, font=F_BOLD, fmt=NUM_FMT)
        cell(ws, r, 13, cat_diff, font=F_BOLD, fmt=NUM_FMT)
        r += 1

        for idx, item in enumerate(cat_items):
            ws.cell(row=r, column=1, value=f"{cat_id}.{idx + 1}")
            ws.cell(row=r, column=2, value=item['unit_eng'])
            ws.cell(row=r, column=3, value=item['code'])
            ws.cell(row=r, column=4, value=item['name'])
            ws.cell(row=r, column=5, value=item['desc'][:500] if item['desc'] else '')
            ws.cell(row=r, column=6, value=item['unit'])
            for c2, k in [(7, 'send_qty'), (8, 'send_price'), (9, 'send_total'),
                           (10, 'review_qty'), (11, 'review_price'), (12, 'review_total'), (13, 'diff')]:
                cell(ws, r, c2, item[k], fmt=NUM_FMT)
            ws.cell(row=r, column=14, value=item['reason'][:500] if item['reason'] else item['analysis'])
            ws.cell(row=r, column=15, value=item['cat_name'])
            r += 1
        r += 1  # blank row between categories

    all_send = sum(it['send_total'] for ct in items_by_category.values() for it in ct)
    all_review = sum(it['review_total'] for ct in items_by_category.values() for it in ct)
    ws.cell(row=r, column=1, value="合计").font = F_BOLD
    cell(ws, r, 9, all_send, font=F_BOLD, fmt=NUM_FMT)
    cell(ws, r, 12, all_review, font=F_BOLD, fmt=NUM_FMT)
    cell(ws, r, 13, all_review - all_send, font=F_BOLD, fmt=NUM_FMT)

    r += 5
    ws.cell(row=r, column=9, value=f"送审总额：{send:,.2f}")
    ws.cell(row=r, column=12, value=f"审定总额：{review:,.2f}")
    ws.cell(row=r + 1, column=13, value=f"审增减：{review - send:,.2f}")


# ─── Main ────────────────────────────────────────────────────────────
def generate(data_dir, output, project=None, companies=None, notes="", bldg_area=0):
    unit_projects, items = load_source_data(data_dir)

    project = project or "固始县人民医院新院建设项目电力工程"
    companies = companies or "大成、河南晖泽、建基、公正、恒信、华明、诚辉、日新、元晟、鑫诚"

    total_send = sum(up['send_amt'] for up in unit_projects)
    total_review = sum(up['review_amt'] for up in unit_projects)

    # Classify and group
    items_by_category = defaultdict(list)
    for item in items:
        items_by_category[item['cat_id']].append(item)

    items_by_unit_cat = defaultdict(float)
    for item in items:
        # Match item's unit_eng to a project name
        up_name = item['unit_eng']
        for up in unit_projects:
            up_clean = clean_name(up['name'])
            if up_clean in up_name or up_name.replace('-建基', '').replace('-晖泽', '') in up_clean:
                items_by_unit_cat[(up_clean, item['cat_id'])] += item['diff']
                break

    # Build workbook
    wb = openpyxl.Workbook()

    ws1 = wb.active; ws1.title = "评审底稿"
    sheet1_draft(ws1, project, companies, total_send, total_review, notes)

    ws2 = wb.create_sheet("汇总表")
    sheet2_summary(ws2, project, unit_projects, items_by_unit_cat, total_send, bldg_area)

    ws3 = wb.create_sheet("导表填写")
    sheet3_data_ref(ws3, items)

    ws4 = wb.create_sheet("审增减明细生成表")
    sheet4_detail(ws4, project, items_by_category, total_send, total_review)

    os.makedirs(os.path.dirname(output) or '.', exist_ok=True)
    wb.save(output)

    # Report
    change = total_review - total_send
    rate = change / total_send * 100 if total_send else 0
    print(f"\n{'='*60}")
    print(f"评审汇总表 → {output}")
    print(f"{'='*60}")
    print(f"送审: {total_send:>15,.2f}  审定: {total_review:>15,.2f}")
    print(f"审增减: {change:>13,.2f}  审减率: {rate:>11.2f}%")
    print(f"单位工程: {len(unit_projects)}  清单条目: {len(items)}")
    print(f"\n按类别:")
    for cat_id in sorted(items_by_category):
        citems = items_by_category[cat_id]
        ctotal = sum(it['diff'] for it in citems)
        print(f"  {CAT_SHORT[cat_id]}: {len(citems):>5}条  审增减 {ctotal:>14,.2f}元")
    return wb


if __name__ == '__main__':
    p = argparse.ArgumentParser(description='Generate 评审汇总表 from 广联达导表')
    p.add_argument('data_dir', help='Directory with 总审核导表.xlsx and 分部分项汇总审核导表.xlsx')
    p.add_argument('output', help='Output .xlsx path')
    p.add_argument('--project', '-p', default='固始县人民医院新院建设项目电力工程')
    p.add_argument('--companies', '-c', default='大成、河南晖泽、建基、公正、恒信、华明、诚辉、日新、元晟、鑫诚')
    p.add_argument('--notes', '-n', default='')
    p.add_argument('--building-area', '-b', type=float, default=0, help='Building area in sqm')
    args = p.parse_args()
    generate(args.data_dir, args.output, args.project, args.companies, args.notes, args.building_area)
