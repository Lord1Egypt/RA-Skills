#!/usr/bin/env python3
"""
PDF增值税发票抽取工具 v1.0
支持：增值税电子发票（普通发票）+ 铁路电子客票（火车票）
依赖：pdfplumber, openpyxl
"""
import pdfplumber, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

COL_SPLIT = 295; COL_SPLIT_RIGHT = 302

KANGXI = {'\u2f63':'\u751f','\u2f53':'\u6d77','\u2f49':'\u6708','\u2f47':'\u65e5',
          '\u2f43':'\u4ea7','\u2f55':'\u6d88','\u2f41':'\u7518','\u2f4b':'\u76ee'}

def norm(t):
    if not t: return t
    for k,v in KANGXI.items(): t = t.replace(k, v)
    return t

def parse_float(s):
    if not s: return 0.0
    s = str(s).replace(",","").replace("¥","").strip()
    m = re.search(r'[-+]?\d*\.\d+|\d+', s)
    return float(m.group()) if m else 0.0

def normalize_date(text):
    if not text: return ""
    m = re.search(r'开票日期[：:\s]*([^\n]{6,30})', text)
    if m:
        raw = re.sub(r'\s+', '', m.group(1).strip())
        dm = re.search(r'(\d{4})年(\d{2})[\u6708\u2F49\u4e00-\u9fff]*(\d{2})[\u65e5\u2F47\u4e00-\u9fff]*', raw)
        if dm: return f"{dm.group(1)}-{dm.group(2)}-{dm.group(3)}"
    # 水印/水印页终极兜底
    dm2 = re.search(r'(\d{4})年(\d{2})月(\d{2})日', text)
    if dm2: return f"{dm2.group(1)}-{dm2.group(2)}-{dm2.group(3)}"
    return ""

def extract_amounts(text):
    heji_match = re.search(r'合\s*计\s+(.+)', text)
    if heji_match:
        heji_line = heji_match.group(1)
        nums = re.findall(r'¥\s*([0-9,]+\.\d+|\*+|\d+)', heji_line)
        if nums:
            amount_raw = nums[0].replace(',','')
            if '***' in heji_line or amount_raw == '***':
                jshj_m = re.search(r'价税合计[^¥]*¥\s*([0-9,]+\.\d+)', text)
                total = parse_float(jshj_m.group(1).replace(',','')) if jshj_m else 0.0
                return 0.0, 0.0, total
            amount = float(amount_raw) if '.' in amount_raw else float(amount_raw)
            tax_raw = nums[1].replace(',','') if len(nums) > 1 else '0'
            tax = float(tax_raw) if tax_raw != '***' else 0.0
            jshj_m = re.search(r'价税合计[^¥]*¥\s*([0-9,]+\.\d+)', text)
            total = parse_float(jshj_m.group(1).replace(',','')) if jshj_m else 0.0
            return amount, tax, total
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if line.strip().startswith('¥'):
            nums = re.findall(r'¥\s*([0-9,]+\.\d+|\*+|\d+)', line)
            if nums:
                amount_raw = nums[0].replace(',',''); amount = float(amount_raw) if '.' in amount_raw else float(amount_raw)
                tax_raw = nums[1].replace(',','') if len(nums) > 1 else '0'
                tax = float(tax_raw) if tax_raw != '***' else 0.0
                for j in range(i+1, len(lines)):
                    jm = re.search(r'¥\s*([0-9,]+\.\d+)', lines[j])
                    if jm: return amount, tax, float(jm.group(1).replace(',',''))
                return amount, tax, 0.0
    return 0.0, 0.0, 0.0

def extract_tax_rate(text):
    rates = re.findall(r'(\d+)%', text)
    unique = sorted(set(rates))
    if len(unique) == 0:
        if '不征税' in text: return '不征税'
        if '免税' in text: return '免税'
        return ""
    if len(unique) > 1: return ""
    return f"{unique[0]}%"

def extract_item_name(text, page_words):
    header_y = col_end_x = None
    for w in page_words:
        if '项目名称' in w['text']: header_y = w['top']; break
    for w in page_words:
        if header_y and abs(w['top']-header_y)<3 and '规格' in w['text']:
            col_end_x = w['x0']; break
    if header_y is None: return ""
    ITEM_END_X = col_end_x or 120
    x0_min = 200
    for w in page_words:
        if w['top'] > header_y and w['text'].startswith('*'):
            x0_min = min(x0_min, w['x0'])
    top_groups = defaultdict(list)
    for w in page_words:
        if w['top'] > header_y and x0_min <= w['x0'] <= ITEM_END_X:
            top_groups[int(w['top'])].append((w['x0'], norm(w['text'])))
    sorted_keys = sorted(top_groups.keys())
    merged = []
    for k in sorted_keys:
        if merged and (k - merged[-1][0]) <= 20:
            merged[-1] = (merged[-1][0], merged[-1][1] + top_groups[k])
        else:
            merged.append((k, top_groups[k]))

    def parse_star_word(txt):
        stars = [i for i, c in enumerate(txt) if c == '*']
        if len(stars) < 2: return txt
        double = any(stars[i]+1 == stars[i+1] for i in range(len(stars)-1))
        if double:
            sl, last = stars[-2], stars[-1]
            cat, name = txt[1:sl], txt[last+1:]
        else:
            second = stars[1]; cat, name = txt[1:second], txt[second+1:]
        return f"*{cat}*{name}" if cat and name else txt

    seen = set(); results = []
    for k, words_list in merged:
        line_text = ''.join(w[1] for w in sorted(words_list))
        if '*' not in line_text: continue
        line_text_normalized = re.sub(r'\*\*', '*', line_text)
        for m in re.finditer(r'\*([^\*]+)\*([^\s\*]+)', line_text_normalized):
            full = parse_star_word(m.group())
            if full and full not in seen: seen.add(full); results.append(full)
    return '\n'.join(results) if results else ""

def concat_name(words, start_word, max_x):
    if not start_word: return ""
    top = start_word['top']
    relevant = [w for w in words if abs(w['top']-top)<4 and start_word['x0']<=w['x0']<=max_x]
    relevant.sort(key=lambda w: w['x0'])
    raw = ''.join(norm(w['text']) for w in relevant)
    return re.sub(r'^名称[：:]\s*','',raw).strip()

def find_invoice_number(text):
    m = re.search(r'发票号码[：:\s]*(\d{10,})', text)
    if m: return m.group(1)
    m = re.search(r'(?:监\s*制\s*)(\d{10,})', text)
    if m: return m.group(1)
    m = re.search(r'([^0-9]|^)(\d{20})([^0-9]|$)', text)
    return m.group(2) if m else None

def find_tax_code_in_zone(words, zone_x_min, zone_x_max, top_hint=None, exclude_watermark=False):
    for w in words:
        if not (zone_x_min <= w['x0'] <= zone_x_max): continue
        if top_hint is not None and abs(w['top'] - top_hint) > 30: continue
        if exclude_watermark and w["top"] < 55: continue
        cm = re.search(r'([A-Z0-9]{10,})', w['text'])
        if cm: return cm.group(1)
    return None

def is_train_ticket(text, page_words):
    if not text: return False
    has_buyer_line = any('购买方名称:' in w['text'] for w in page_words)
    has_train_number = any('电子客票号:' in w['text'] for w in page_words)
    return has_buyer_line and has_train_number

def extract_train_ticket(page_words, page_num):
    text = ' '.join(w['text'] for w in page_words)
    inv = None
    for w in page_words:
        m = re.search(r'(\d{20})', w['text'])
        if m: inv = m.group(1); break
    date_str = ""
    for w in page_words:
        m = re.search(r'开票日期[：:](\d{4})年(\d{2})月(\d{2})日', w['text'])
        if m: date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"; break
    buyer = ""
    for w in page_words:
        if '购买方名称:' in w['text']:
            buyer = re.sub(r'^购买方名称[：:]\s*', '', w['text']).strip()
            break
    buyer_code = ""
    for w in page_words:
        if '统一社会信用代码:' in w['text']:
            m = re.search(r'([A-Z0-9]{10,})', w['text'])
            if m: buyer_code = m.group(1); break
    total = 0.0
    for w in page_words:
        m = re.search(r'[￥¥]\s*([0-9,]+\.?\d*)', w['text'])
        if m:
            try: total = float(m.group(1).replace(',','')); break
            except: pass
    return {
        'invoice_id': inv or '', 'date': date_str,
        'buyer': buyer, 'buyer_code': buyer_code,
        'seller': '', 'seller_code': '',
        'amount': 0.0, 'tax': 0.0, 'total': total,
        'rate': '', 'item_name': '火车票',
        'page_num': page_num, 'is_last': True,
    }

def get_page_info(page, text, page_num):
    page_words = page.extract_words() or []
    if is_train_ticket(text, page_words):
        return extract_train_ticket(page_words, page_num)
    inv = find_invoice_number(text)
    if not inv: return None
    date_str = normalize_date(text)
    name_w_buyer = name_w_seller = None
    for w in page_words:
        if ('名称' in w['text'] and '统一' not in w['text']
            and not w['text'].startswith('项目名称')):
            if w['x0'] < COL_SPLIT and not name_w_buyer: name_w_buyer = w
            elif w['x0'] > COL_SPLIT_RIGHT and not name_w_seller: name_w_seller = w
    buyer_name = concat_name(page_words, name_w_buyer, COL_SPLIT-1)
    seller_name = concat_name(page_words, name_w_seller, 600)
    buyer_code = seller_code = ""
    for w in page_words:
        if '统一社会信用代码' in w['text'] or '纳税人识别号' in w['text']:
            cm = re.search(r'([A-Z0-9]{10,})', w['text'])
            if cm:
                if w['x0'] < COL_SPLIT and not buyer_code: buyer_code = cm.group(1)
                elif w['x0'] > COL_SPLIT_RIGHT and not seller_code: seller_code = cm.group(1)
    if not buyer_code:
        bc = find_tax_code_in_zone(page_words, 0, COL_SPLIT-1)
        if bc: buyer_code = bc
    if not seller_code:
        sc = find_tax_code_in_zone(page_words, COL_SPLIT_RIGHT+1, 700, exclude_watermark=True)
        if sc: seller_code = sc
    amount, tax, total = extract_amounts(text)
    rate_str = extract_tax_rate(text)
    item_name = extract_item_name(text, page_words)
    return {
        'invoice_id': inv, 'date': date_str,
        'buyer': buyer_name, 'buyer_code': buyer_code,
        'seller': seller_name, 'seller_code': seller_code,
        'amount': amount, 'tax': tax, 'total': total,
        'rate': rate_str, 'item_name': item_name,
        'page_num': page_num, 'is_last': True,
    }

def extract_all_invoices(pdf_path):
    inv_pages = defaultdict(list)
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            try: text = page.extract_text()
            except: continue
            if not text: continue
            info = get_page_info(page, text, page_num)
            if info: inv_pages[info['invoice_id']].append(info)
    invoices = []
    for inv, pages in inv_pages.items():
        if len(pages) == 1:
            p = pages[0]
            total = p['total'] if p['total'] > 0 else p['amount']
            invoices.append({
                '发票号码': inv, '开票日期': p['date'],
                '购买方信息-名称': p['buyer'],
                '购买方信息-统一社会信用代码': p['buyer_code'],
                '销售方信息-名称': p['seller'],
                '销售方信息-统一社会信用代码': p['seller_code'],
                '项目名称': p['item_name'],
                '金额': p['amount'], '税率/征收率': p['rate'],
                '税额': p['tax'], '价税合计': total,
            })
        else:
            pages.sort(key=lambda x: x['page_num'])
            last = pages[-1]; first = pages[0]
            total = last['total'] if last['total'] > 0 else last['amount']
            all_rates = [p['rate'] for p in pages if p['rate']]
            unique = sorted(set(all_rates))
            final_rate = unique[0] if len(unique)==1 else ('' if len(unique)>1 else '')
            invoices.append({
                '发票号码': inv, '开票日期': first['date'],
                '购买方信息-名称': first['buyer'],
                '购买方信息-统一社会信用代码': first['buyer_code'],
                '销售方信息-名称': first['seller'],
                '销售方信息-统一社会信用代码': first['seller_code'],
                '项目名称': first['item_name'],
                '金额': last['amount'], '税率/征收率': final_rate,
                '税额': last['tax'], '价税合计': total,
            })
    return invoices

def detect_duplicates(invoices):
    dupes = []
    for i, inv in enumerate(invoices):
        for j, other in enumerate(invoices):
            if i >= j: continue
            if (inv['发票号码']==other['发票号码'] and inv['开票日期']==other['开票日期']
                and inv['购买方信息-名称']==other['购买方信息-名称']
                and inv['销售方信息-名称']==other['销售方信息-名称']
                and abs(inv['金额']-other['金额'])<0.01
                and abs(inv['价税合计']-other['价税合计'])<0.01):
                dupes.extend([i,j])
    return sorted(set(dupes))

def format_rate(rate_str):
    if not rate_str: return ""
    if rate_str in ('不征税','免税'): return rate_str
    if '%' in rate_str:
        try: return float(rate_str.replace('%',''))/100
        except: return rate_str
    return rate_str

def create_excel(invoices, duplicate_rows, output_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "发票统计"
    headers = ['序号','发票号码','开票日期','购买方信息-名称',
               '购买方信息-统一社会信用代码','销售方信息-名称',
               '销售方信息-统一社会信用代码','项目名称',
               '金额','税率/征收率','税额','价税合计','备注','报销人','报销日期']
    hdr_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align
    yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
    for row_idx, inv in enumerate(invoices, 1):
        row_num = row_idx + 1; is_dup = (row_idx-1) in duplicate_rows
        rate_val = format_rate(inv['税率/征收率'])
        row_data = [row_idx, inv['发票号码'], inv['开票日期'],
                    inv['购买方信息-名称'], inv['购买方信息-统一社会信用代码'],
                    inv['销售方信息-名称'], inv['销售方信息-统一社会信用代码'],
                    inv['项目名称'], inv['金额'], rate_val, inv['税额'],
                    inv['价税合计'], '', '', '']
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            if is_dup: c.fill = yellow_fill
            if col==2: c.number_format='@'
            elif col==3: c.number_format='YYYY-MM-DD'
            elif col in [5,7]: c.number_format='@'
            elif col in [9,11,12]: c.number_format='#,##0.00'
            elif col==10 and isinstance(val,float): c.number_format='0.00%'
            elif col==8: c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.freeze_panes='A2'
    for col in range(1,len(headers)+1):
        letter=get_column_letter(col)
        max_len=len(headers[col-1])
        for row in range(2,len(invoices)+2):
            v=ws.cell(row=row,column=col).value
            if v: max_len=min(max(max_len,len(str(v))),50)
        ws.column_dimensions[letter].width=max_len+3
    wb.save(output_path)

def main():
    if len(sys.argv) < 2:
        print("用法: python3 invoice_extractor.py <PDF路径> [输出Excel路径]")
        print("  <PDF路径>  : 必填，合并PDF文件路径")
        print("  [输出Excel路径]: 选填，默认在PDF同目录生成 <原文件名>_发票统计.xlsx")
        sys.exit(1)
    pdf_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        import os
        base = os.path.splitext(pdf_path)[0]
        output_path = base + "_发票统计.xlsx"

    print(f"PDF: {pdf_path}")
    print(f"输出: {output_path}")
    invoices = extract_all_invoices(pdf_path)
    empty_date = [i+1 for i,inv in enumerate(invoices) if not inv['开票日期']]
    empty_buyer = [i+1 for i,inv in enumerate(invoices) if not inv['购买方信息-名称']]
    empty_buyer_code = [i+1 for i,inv in enumerate(invoices) if not inv['购买方信息-统一社会信用代码']]
    zero_amt = [i+1 for i,inv in enumerate(invoices) if inv['金额']==0 and inv['项目名称']!='火车票']
    empty_item = [i+1 for i,inv in enumerate(invoices) if not inv['项目名称']]
    print(f"总数:{len(invoices)} | 日期空:{len(empty_date)} | 购方空:{len(empty_buyer)} | 购方税号空:{len(empty_buyer_code)} | 金额0(非火车票):{len(zero_amt)} | 项目空:{len(empty_item)}")
    if empty_date: print(f"  日期空: {[invoices[i-1]['发票号码'] for i in empty_date]}")
    if empty_buyer_code: print(f"  购方税号空: {[invoices[i-1]['发票号码'] for i in empty_buyer_code]}")
    if empty_item: print(f"  项目空: {[invoices[i-1]['发票号码'] for i in empty_item]}")
    dupes = detect_duplicates(invoices)
    print(f"重复:{len(dupes)}行")
    create_excel(invoices, dupes, output_path)
    from collections import Counter
    rates = Counter(inv['税率/征收率'] for inv in invoices)
    print(f"税率分布:{dict(sorted(rates.items(),key=lambda x:-x[1]))}")
    print(f"完成! -> {output_path}")

if __name__=='__main__':
    main()
