#!/usr/bin/env python3
"""
gRPC Test Report Generator
生成 Excel 格式的测试报告
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter


def create_header_style():
    """创建表头样式"""
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_cell_style():
    """创建单元格样式"""
    return {
        'alignment': Alignment(horizontal="left", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_style(ws, row, col, style_dict):
    """应用样式到单元格"""
    cell = ws.cell(row=row, column=col)
    for key, value in style_dict.items():
        setattr(cell, key, value)


def generate_report(test_results, output_file=None):
    """
    生成 Excel 测试报告
    
    Args:
        test_results: 测试结果列表
        output_file: 输出文件路径
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"jmeter/results/test_report_{timestamp}.xlsx"
    
    # 确保目录存在
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    header_style = create_header_style()
    cell_style = create_cell_style()
    
    # ========== Sheet 1: Summary ==========
    ws1 = wb.active
    ws1.title = "测试概览"
    
    # 标题
    ws1.merge_cells('A1:D1')
    title_cell = ws1['A1']
    title_cell.value = "VENC gRPC 测试报告"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # 统计数据
    total = len(test_results)
    passed = sum(1 for r in test_results if r.get('status') == 'PASSED')
    failed = total - passed
    pass_rate = f"{passed/total*100:.1f}%" if total > 0 else "0%"
    
    summary_data = [
        ["测试计划", "VENC gRPC 测试"],
        ["测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["测试环境", "localhost:8080"],
        ["", ""],
        ["总测试数", total],
        ["通过", passed],
        ["失败", failed],
        ["通过率", pass_rate],
    ]
    
    for i, (label, value) in enumerate(summary_data, start=3):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=value)
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    
    # ========== Sheet 2: Test Results ==========
    ws2 = wb.create_sheet("测试详情")
    
    # 表头
    headers = ["用例ID", "接口", "请求", "预期结果", "实际结果", "状态", "响应时间(ms)"]
    for col, header in enumerate(headers, start=1):
        ws2.cell(row=1, column=col, value=header)
        apply_style(ws2, 1, col, header_style)
    
    # 数据行
    for row_idx, result in enumerate(test_results, start=2):
        ws2.cell(row=row_idx, column=1, value=result.get('test_id', ''))
        ws2.cell(row=row_idx, column=2, value=result.get('interface', ''))
        ws2.cell(row=row_idx, column=3, value=result.get('request', ''))
        ws2.cell(row=row_idx, column=4, value=result.get('expected', ''))
        ws2.cell(row=row_idx, column=5, value=result.get('actual', ''))
        
        # 状态颜色
        status = result.get('status', '')
        status_cell = ws2.cell(row=row_idx, column=6, value=status)
        if status == 'PASSED':
            status_cell.font = Font(color="00B050", bold=True)
        elif status == 'FAILED':
            status_cell.font = Font(color="FF0000", bold=True)
        
        ws2.cell(row=row_idx, column=7, value=result.get('time_ms', ''))
        
        # 应用样式
        for col in range(1, 8):
            apply_style(ws2, row_idx, col, cell_style)
    
    # 设置列宽
    col_widths = [12, 20, 40, 30, 30, 10, 15]
    for i, width in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    
    # ========== Sheet 3: Performance ==========
    ws3 = wb.create_sheet("性能指标")
    
    perf_headers = ["接口", "请求数", "平均(ms)", "最小(ms)", "最大(ms)", "90%线", "错误率", "吞吐量"]
    for col, header in enumerate(perf_headers, start=1):
        ws3.cell(row=1, column=col, value=header)
        apply_style(ws3, 1, col, header_style)
    
    # 按接口统计性能数据
    interface_stats = {}
    for r in test_results:
        iface = r.get('interface', 'Unknown')
        time_ms = r.get('time_ms', 0)
        if iface not in interface_stats:
            interface_stats[iface] = []
        interface_stats[iface].append(time_ms)
    
    for row_idx, (iface, times) in enumerate(interface_stats.items(), start=2):
        if times:
            ws3.cell(row=row_idx, column=1, value=iface)
            ws3.cell(row=row_idx, column=2, value=len(times))
            ws3.cell(row=row_idx, column=3, value=round(sum(times)/len(times), 2))
            ws3.cell(row=row_idx, column=4, value=min(times))
            ws3.cell(row=row_idx, column=5, value=max(times))
            
            sorted_times = sorted(times)
            p90_idx = int(len(sorted_times) * 0.9)
            ws3.cell(row=row_idx, column=6, value=sorted_times[p90_idx] if sorted_times else 0)
            
            failed_count = sum(1 for r in test_results 
                             if r.get('interface') == iface and r.get('status') == 'FAILED')
            error_rate = f"{failed_count/len(times)*100:.1f}%"
            ws3.cell(row=row_idx, column=7, value=error_rate)
            
            throughput = f"{1000/(sum(times)/len(times)):.2f}/sec" if times else "0"
            ws3.cell(row=row_idx, column=8, value=throughput)
    
    # ========== Sheet 4: Error Details ==========
    ws4 = wb.create_sheet("错误详情")
    
    error_headers = ["用例ID", "接口", "错误信息", "请求"]
    for col, header in enumerate(error_headers, start=1):
        ws4.cell(row=1, column=col, value=header)
        apply_style(ws4, 1, col, header_style)
    
    error_row = 2
    for result in test_results:
        if result.get('status') == 'FAILED':
            ws4.cell(row=error_row, column=1, value=result.get('test_id', ''))
            ws4.cell(row=error_row, column=2, value=result.get('interface', ''))
            ws4.cell(row=error_row, column=3, value=result.get('error', ''))
            ws4.cell(row=error_row, column=4, value=result.get('request', ''))
            error_row += 1
    
    # 保存文件
    wb.save(output_file)
    print(f"✅ 报告已生成: {output_file}")
    return output_file


def load_results_from_json(json_file):
    """从 JSON 文件加载测试结果"""
    with open(json_file, 'r') as f:
        data = json.load(f)
    return data.get('tests', data if isinstance(data, list) else [])


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("Usage: python generate_report.py <results.json> [output.xlsx]")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    results = load_results_from_json(json_file)
    generate_report(results, output_file)


if __name__ == '__main__':
    main()
