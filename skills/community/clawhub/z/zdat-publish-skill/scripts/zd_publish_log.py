"""
ZDAT 发布日志查看脚本 v1.0
读取 publish_log.xlsx 展示最近发布记录
"""
import sys, os, datetime
from pathlib import Path
from openpyxl import load_workbook

WORKDIR = Path(os.getenv("WORKDIR", os.path.expanduser("~/.molili/workspaces/default")))
LOG_PATH = WORKDIR / "publish_log.xlsx"

def show_log(limit=10):
    if not LOG_PATH.exists():
        print("📋 暂无发布日志")
        return
    
    wb = load_workbook(str(LOG_PATH))
    ws = wb.active
    
    print(f"\n📋 ZDAT 发布日志 — {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"   最近 {limit} 条记录:\n")
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= limit:
            break
        print(f"   {' | '.join(str(c) for c in row if c)}")

if __name__ == "__main__":
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 10
    show_log(limit)
