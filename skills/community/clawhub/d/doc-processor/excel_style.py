#!/usr/bin/env python3
"""Excel 样式保持模块 - v2.6.0"""

from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Dict
import logging

logger = logging.getLogger(__name__)

@dataclass
class CellStyle:
    font: Dict = field(default_factory=dict)
    fill: Dict = field(default_factory=dict)
    border: Dict = field(default_factory=dict)
    alignment: Dict = field(default_factory=dict)
    number_format: str = ""

class StylePreserver:
    def __init__(self):
        self._styles: Dict[str, CellStyle] = {}
        self._formulas: Dict[str, str] = {}
    
    def _get_cell_key(self, row: int, column: int) -> str:
        return f"{get_column_letter(column)}{row}"
    
    def save_styles(self, ws):
        self._styles.clear()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value or cell.has_style:
                    key = self._get_cell_key(cell.row, cell.column)
                    self._styles[key] = CellStyle(
                        font={'name': cell.font.name, 'size': cell.font.size, 'bold': cell.font.bold, 'italic': cell.font.italic},
                        fill={'type': 'pattern', 'patternType': cell.fill.patternType if hasattr(cell.fill, 'patternType') else 'solid'},
                        border={},
                        alignment={'horizontal': cell.alignment.horizontal if hasattr(cell.alignment, 'horizontal') else None, 'vertical': cell.alignment.vertical if hasattr(cell.alignment, 'vertical') else None},
                        number_format=cell.number_format
                    )
    
    def save_formulas(self, ws):
        self._formulas.clear()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    key = self._get_cell_key(cell.row, cell.column)
                    self._formulas[key] = cell.value
    
    def restore_styles(self, ws):
        if not self._styles:
            return
        for row in ws.iter_rows():
            for cell in row:
                key = self._get_cell_key(cell.row, cell.column)
                if key in self._styles:
                    style = self._styles[key]
                    if style.font.get('name'):
                        cell.font = Font(name=style.font.get('name'), size=style.font.get('size'), bold=style.font.get('bold'), italic=style.font.get('italic'))
    
    def restore_formulas(self, ws):
        if not self._formulas:
            return
        for row in ws.iter_rows():
            for cell in row:
                key = self._get_cell_key(cell.row, cell.column)
                if key in self._formulas:
                    if not cell.value or not str(cell.value).startswith('='):
                        cell.value = self._formulas[key]
    
    def get_stats(self) -> Dict:
        return {'styles_count': len(self._styles), 'formulas_count': len(self._formulas)}
