#!/usr/bin/env python3
"""Excel 表格区域填充测试 - v2.3"""

import sys
import tempfile
from pathlib import Path

# 添加 Skill 目录到 Python 路径
sys.path.insert(0, str(Path(__file__).parent))

from doc_processor import DocumentProcessor

def test_excel_table_fill():
    """测试 Excel 表格区域填充"""
    print("=" * 70)
    print("Excel 表格区域填充测试 - v2.3")
    print("=" * 70)
    
    processor = DocumentProcessor()
    
    # 创建测试 Excel 模板（带表格）
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = '产品列表'
    ws['A3'] = '产品名称'
    ws['B3'] = '数量'
    ws['C3'] = '单价'
    ws['D3'] = '金额'
    ws['A4'] = '{{product_data}}'  # 表格数据占位符
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        template_path = f.name
        wb.save(template_path)
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name
    
    try:
        # 填充模板（包含表格数据）
        data = {
            'product_data': [
                ['产品 A', 2, 100, 200],
                ['产品 B', 1, 200, 200],
                ['产品 C', 3, 150, 450]
            ]
        }
        
        result = processor.fill_template(template_path, data, output_path)
        print(f"✅ Excel 表格填充成功")
        print(f"   输出：{result}")
        
        # 验证填充结果
        from openpyxl import load_workbook
        filled_wb = load_workbook(output_path)
        filled_ws = filled_wb.active
        
        print(f"\n📊 验证填充结果:")
        print(f"   A4: {filled_ws['A4'].value}")
        print(f"   B4: {filled_ws['B4'].value}")
        print(f"   C4: {filled_ws['C4'].value}")
        print(f"   D4: {filled_ws['D4'].value}")
        print(f"   A5: {filled_ws['A5'].value}")
        print(f"   A6: {filled_ws['A6'].value}")
        
        # 断言验证
        assert filled_ws['A4'].value == '产品 A', f"第一行产品名错误：{filled_ws['A4'].value}"
        assert filled_ws['B4'].value == 2, f"第一行数量错误：{filled_ws['B4'].value}"
        assert filled_ws['A5'].value == '产品 B', f"第二行产品名错误：{filled_ws['A5'].value}"
        assert filled_ws['A6'].value == '产品 C', f"第三行产品名错误：{filled_ws['A6'].value}"
        
        print(f"\n✅ 所有验证通过！")
        
    except Exception as e:
        print(f"❌ 测试失败：{e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        Path(template_path).unlink(missing_ok=True)
        Path(output_path).unlink(missing_ok=True)
    
    return True

if __name__ == "__main__":
    success = test_excel_table_fill()
    sys.exit(0 if success else 1)
