#!/usr/bin/env python3
"""
Attendance Data Processor for Zhuihui Branch (珠晖支行)

This script reads raw attendance Excel data, applies business rules
(exclusions, department splitting, schedule-based calculations),
and outputs structured JSON for AI report generation.

Usage:
    python3 analyze_attendance.py <input_xlsx> [--output <output_json>]

Example:
    python3 analyze_attendance.py 每日统计表_20260224_20260228.xlsx --output result.json
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIGURATION — Edit these when business rules change
# ============================================================================

# Employees to exclude from each department
EXCLUDED_EMPLOYEES = {
    "新湘支行": ["珠晖-新湘-王治国", "珠晖-新湘-魏紫兰"],
    "茶山支行": ["珠晖-茶山-陈妍"],
    "火车站支行": ["珠晖-周文娟"],
    "珠晖支行本部": ["陈喆", "刘一婧", "谭国球", "谭庆荣", "谢清林"],
}

# 珠晖支行本部 sub-department mapping
HEADQUARTERS_SUB_DEPARTMENTS = {
    "综合管理部": ["彭涵", "谢祎", "王友", "蒋蕙", "朱亚民"],
    "业务管理部": ["彭国柱", "肖冬梅", "邓意平", "王知生", "颜慧松", "李娜", "刘新恒", "李卿"],
    "公司业务部": ["陈瑶", "周鹍", "甘健生", "管巧林", "欧海滨"],
    "个人金融部": ["刘平", "廖欢", "陈杰", "吴欣钰", "费鸿平", "汤锋"],
}

# Branch schedule types
WEEKEND_ROTATION_BRANCHES = ["新湘支行", "茶山支行"]  # Saturday coverage
FULL_COVERAGE_BRANCH = "珠晖支行营业部"  # Saturday + Sunday coverage
# All other branches are standard Mon-Fri


# ============================================================================
# CORE LOGIC
# ============================================================================

def parse_date(date_val):
    """Parse date from various formats to datetime object."""
    date_str = str(date_val).strip()
    try:
        return datetime.strptime(date_str, "%Y%m%d")
    except ValueError:
        return None


def is_valid_clock(val):
    """Check if a clock-in/out value is a valid time (not '-', empty, or whitespace)."""
    if val is None:
        return False
    s = str(val).strip()
    return s != "" and s != "-" and s != " "


def get_weekday_name_zh(dt):
    """Get Chinese weekday name."""
    names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    return names[dt.weekday()]


def is_weekend(dt):
    """Check if date is Saturday(5) or Sunday(6)."""
    return dt.weekday() >= 5


def is_saturday(dt):
    return dt.weekday() == 5


def is_sunday(dt):
    return dt.weekday() == 6


def read_excel(filepath):
    """
    Read attendance Excel file and return structured records.
    Expects: Row 1 = group headers, Row 2 = column names, Row 3+ = data.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if len(row) < 11:
            continue

        name = row[0]
        emp_id = row[1]
        dept = row[2]
        date_val = row[3]
        clock_in1 = row[4]
        clock_out1 = row[5]
        # row[6-9]: additional clock pairs (usually empty)
        result = row[10]

        if not name or not dept:
            continue

        dt = parse_date(date_val)
        if dt is None:
            continue

        # Additional time stats (if available)
        late_min = row[14] if len(row) > 14 else 0
        early_min = row[15] if len(row) > 15 else 0
        actual_hours = row[13] if len(row) > 13 else 0

        records.append({
            "name": str(name).strip(),
            "emp_id": str(emp_id).strip() if emp_id else "-",
            "dept": str(dept).strip(),
            "date": dt,
            "date_str": dt.strftime("%Y-%m-%d"),
            "weekday": get_weekday_name_zh(dt),
            "clock_in": str(clock_in1).strip() if clock_in1 else "",
            "clock_out": str(clock_out1).strip() if clock_out1 else "",
            "has_clock_in": is_valid_clock(clock_in1),
            "has_clock_out": is_valid_clock(clock_out1),
            "result": str(result).strip() if result else "-",
            "late_minutes": float(late_min) if late_min and str(late_min).strip() not in ("", "-") else 0,
            "early_minutes": float(early_min) if early_min and str(early_min).strip() not in ("", "-") else 0,
            "actual_hours": float(actual_hours) if actual_hours and str(actual_hours).strip() not in ("", "-") else 0,
        })

    return records


def exclude_employees(records):
    """Remove excluded employees from records."""
    excluded_count = defaultdict(int)

    def should_exclude(rec):
        dept = rec["dept"]
        name = rec["name"]
        if dept in EXCLUDED_EMPLOYEES and name in EXCLUDED_EMPLOYEES[dept]:
            excluded_count[f"{dept}/{name}"] += 1
            return True
        return False

    filtered = [r for r in records if not should_exclude(r)]
    return filtered, dict(excluded_count)


def split_headquarters(records):
    """
    Split 珠晖支行本部 into 4 sub-departments.
    Records not matching any sub-department remain under 珠晖支行本部.
    """
    name_to_sub = {}
    for sub_dept, staff in HEADQUARTERS_SUB_DEPARTMENTS.items():
        for staff_name in staff:
            name_to_sub[staff_name] = sub_dept

    unmatched = set()
    for rec in records:
        if rec["dept"] == "珠晖支行本部":
            sub = name_to_sub.get(rec["name"])
            if sub:
                rec["dept"] = sub
            else:
                unmatched.add(rec["name"])

    return records, unmatched


def is_rotation_branch(dept):
    """Check if a department is a weekend-rotation branch (not a standard Mon-Fri branch)."""
    return dept in WEEKEND_ROTATION_BRANCHES or dept == FULL_COVERAGE_BRANCH


def identify_weekend_workers(records):
    """
    Identify employees who actually worked on weekends.
    ONLY checks branches that are configured as rotation branches.
    Standard branches (Mon-Fri) will NOT have weekend workers even if
    the Excel contains Saturday/Sunday rows for them.

    Returns a dict: {dept: {name: [weekend_dates]}}
    """
    weekend_workers = defaultdict(lambda: defaultdict(list))

    for rec in records:
        # Only look at rotation branches
        if not is_rotation_branch(rec["dept"]):
            continue

        if is_weekend(rec["date"]) and (rec["has_clock_in"] or rec["has_clock_out"]):
            if rec["result"] not in ("-",):
                weekend_workers[rec["dept"]][rec["name"]].append({
                    "date": rec["date_str"],
                    "weekday": rec["weekday"],
                    "clock_in": rec["clock_in"],
                    "clock_out": rec["clock_out"],
                    "result": rec["result"],
                })

    return {dept: dict(names) for dept, names in weekend_workers.items()}


def calculate_expected_clockins(records, weekend_workers):
    """
    Calculate expected clock-ins for each employee based on their schedule.

    Rules:
    - Standard branches (Mon-Fri): count weekday dates only, ignore weekend records
    - 新湘/茶山 weekend rotation staff: include Saturday; rest days on weekdays excluded
    - 营业部 weekend rotation staff: include Sat+Sun; rest days on weekdays excluded
    - Non-rotation staff in rotation branches: weekday dates only
    """
    # Group records by (dept, name)
    person_records = defaultdict(list)
    for rec in records:
        person_records[(rec["dept"], rec["name"])].append(rec)

    person_expected = {}

    for (dept, name), recs in person_records.items():
        is_weekend_worker = name in weekend_workers.get(dept, {})

        if is_weekend_worker:
            # Weekend worker in a rotation branch:
            # Count all dates where they have non-'-' result as working dates
            # (includes their weekend days, excludes their rest weekdays)
            working_dates = [rec["date"] for rec in recs if rec["result"] != "-"]
            expected_days = len(working_dates)
        else:
            # Standard worker OR non-rotation worker in any branch:
            # Only count weekday dates with non-'-' result
            working_dates = [
                rec["date"] for rec in recs
                if not is_weekend(rec["date"]) and rec["result"] != "-"
            ]
            expected_days = len(working_dates)

        expected_clockins = expected_days * 2

        person_expected[(dept, name)] = {
            "expected_days": expected_days,
            "expected_clockins": expected_clockins,
            "is_weekend_worker": is_weekend_worker,
        }

    return person_expected


def calculate_actual_clockins(records, weekend_workers):
    """Count actual valid clock-ins per employee, respecting branch schedules."""
    person_actual = defaultdict(lambda: {
        "actual_clockins": 0,
        "missing_clockins": 0,
        "missing_details": [],
        "late_details": [],
        "early_details": [],
        "total_late_minutes": 0,
        "total_early_minutes": 0,
    })

    for rec in records:
        key = (rec["dept"], rec["name"])
        data = person_actual[key]

        # Skip non-working records
        if rec["result"] == "-":
            continue

        # For non-rotation branches, skip weekend records entirely
        if is_weekend(rec["date"]):
            is_weekend_worker = rec["name"] in weekend_workers.get(rec["dept"], {})
            if not is_weekend_worker:
                continue

        # Count clock-ins
        if rec["has_clock_in"]:
            data["actual_clockins"] += 1
        else:
            data["missing_clockins"] += 1
            data["missing_details"].append({
                "date": rec["date_str"],
                "weekday": rec["weekday"],
                "type": "上班缺卡",
            })

        if rec["has_clock_out"]:
            data["actual_clockins"] += 1
        else:
            data["missing_clockins"] += 1
            data["missing_details"].append({
                "date": rec["date_str"],
                "weekday": rec["weekday"],
                "type": "下班缺卡",
            })

        # Track late/early
        if rec["late_minutes"] > 0:
            data["late_details"].append({
                "date": rec["date_str"],
                "weekday": rec["weekday"],
                "clock_in": rec["clock_in"],
                "minutes": rec["late_minutes"],
            })
            data["total_late_minutes"] += rec["late_minutes"]

        if rec["early_minutes"] > 0:
            data["early_details"].append({
                "date": rec["date_str"],
                "weekday": rec["weekday"],
                "clock_out": rec["clock_out"],
                "minutes": rec["early_minutes"],
            })
            data["total_early_minutes"] += rec["early_minutes"]

    return dict(person_actual)


def generate_department_summary(records, person_expected, person_actual, weekend_workers):
    """Generate per-department attendance summary."""
    # Get all departments
    departments = sorted(set(r["dept"] for r in records))

    # Get all persons per department
    dept_persons = defaultdict(set)
    for rec in records:
        if rec["result"] != "-":
            dept_persons[rec["dept"]].add(rec["name"])

    summaries = []

    for dept in departments:
        persons = sorted(dept_persons.get(dept, set()))
        if not persons:
            continue

        total_expected = 0
        total_actual = 0
        total_missing = 0
        missing_persons = []
        late_persons = []
        early_persons = []
        person_details = []

        for name in persons:
            key = (dept, name)
            exp = person_expected.get(key, {"expected_clockins": 0, "expected_days": 0, "is_weekend_worker": False})
            act = person_actual.get(key, {"actual_clockins": 0, "missing_clockins": 0, "missing_details": [], "late_details": [], "early_details": [], "total_late_minutes": 0, "total_early_minutes": 0})

            p_expected = exp["expected_clockins"]
            p_actual = act["actual_clockins"]
            p_missing = act["missing_clockins"]

            total_expected += p_expected
            total_actual += p_actual
            total_missing += p_missing

            if p_missing > 0:
                missing_persons.append({
                    "name": name,
                    "missing_count": p_missing,
                    "details": act["missing_details"],
                })

            if act["late_details"]:
                late_persons.append({
                    "name": name,
                    "late_count": len(act["late_details"]),
                    "total_minutes": act["total_late_minutes"],
                    "details": act["late_details"],
                })

            if act["early_details"]:
                early_persons.append({
                    "name": name,
                    "early_count": len(act["early_details"]),
                    "total_minutes": act["total_early_minutes"],
                    "details": act["early_details"],
                })

            person_details.append({
                "name": name,
                "is_weekend_worker": exp["is_weekend_worker"],
                "expected_days": exp["expected_days"],
                "expected_clockins": p_expected,
                "actual_clockins": p_actual,
                "missing_clockins": p_missing,
                "attendance_rate": round(p_actual / p_expected * 100, 2) if p_expected > 0 else 0,
            })

        attendance_rate = round(total_actual / total_expected * 100, 2) if total_expected > 0 else 0

        summaries.append({
            "department": dept,
            "employee_count": len(persons),
            "total_expected_clockins": total_expected,
            "total_actual_clockins": total_actual,
            "total_missing_clockins": total_missing,
            "attendance_rate": attendance_rate,
            "missing_persons": missing_persons,
            "late_persons": late_persons,
            "early_persons": early_persons,
            "weekend_workers": list(weekend_workers.get(dept, {}).keys()),
            "person_details": person_details,
        })

    # Sort by attendance rate (ascending = worst first, or descending = best first)
    summaries.sort(key=lambda x: x["attendance_rate"], reverse=True)

    return summaries


def analyze(filepath):
    """Main analysis pipeline."""
    print(f"📂 Reading file: {filepath}")
    records = read_excel(filepath)
    print(f"   Total records: {len(records)}")

    # Date range info
    all_dates = sorted(set(r["date"] for r in records))
    date_range = {
        "start": all_dates[0].strftime("%Y-%m-%d"),
        "end": all_dates[-1].strftime("%Y-%m-%d"),
        "days": len(all_dates),
        "dates": [{"date": d.strftime("%Y-%m-%d"), "weekday": get_weekday_name_zh(d)} for d in all_dates],
        "has_saturday": any(is_saturday(d) for d in all_dates),
        "has_sunday": any(is_sunday(d) for d in all_dates),
    }
    print(f"   Date range: {date_range['start']} ~ {date_range['end']} ({date_range['days']} days)")

    # Step 1: Exclude employees
    print("\n🧹 Step 1: Excluding employees...")
    records, excluded = exclude_employees(records)
    for key, count in excluded.items():
        print(f"   Removed: {key} ({count} records)")
    print(f"   Records after exclusion: {len(records)}")

    # Step 2: Split headquarters
    print("\n🏢 Step 2: Splitting 珠晖支行本部...")
    records, unmatched = split_headquarters(records)
    if unmatched:
        print(f"   ⚠ Unmatched staff (kept under 珠晖支行本部): {unmatched}")
    else:
        print("   All staff matched to sub-departments.")

    # Step 3: Identify weekend workers
    print("\n📅 Step 3: Identifying weekend workers...")
    weekend_workers = identify_weekend_workers(records)
    for dept, workers in weekend_workers.items():
        print(f"   {dept}: {list(workers.keys())}")

    # Step 4: Calculate expected clock-ins
    print("\n🔢 Step 4: Calculating expected clock-ins...")
    person_expected = calculate_expected_clockins(records, weekend_workers)

    # Step 5: Calculate actual clock-ins
    print("\n✅ Step 5: Calculating actual clock-ins...")
    person_actual = calculate_actual_clockins(records, weekend_workers)

    # Step 6: Generate department summaries
    print("\n📊 Step 6: Generating department summaries...")
    summaries = generate_department_summary(records, person_expected, person_actual, weekend_workers)

    # Build final output
    output = {
        "meta": {
            "source_file": os.path.basename(filepath),
            "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "date_range": date_range,
            "total_records_raw": len(records) + sum(excluded.values()),
            "total_records_after_exclusion": len(records),
            "excluded_employees": excluded,
        },
        "department_summaries": summaries,
    }

    # Print summary
    print("\n" + "=" * 60)
    print("📋 ATTENDANCE SUMMARY (sorted by attendance rate)")
    print("=" * 60)
    for s in summaries:
        print(f"\n  {s['department']}")
        print(f"    员工数: {s['employee_count']}")
        print(f"    出勤率: {s['attendance_rate']}%")
        print(f"    应打卡: {s['total_expected_clockins']} | 实际: {s['total_actual_clockins']} | 缺卡: {s['total_missing_clockins']}")
        if s["missing_persons"]:
            names = ", ".join(f"{p['name']}({p['missing_count']}次)" for p in s["missing_persons"])
            print(f"    缺卡人员: {names}")
        if s["late_persons"]:
            names = ", ".join(f"{p['name']}({p['total_minutes']}min)" for p in s["late_persons"])
            print(f"    迟到人员: {names}")
        if s["early_persons"]:
            names = ", ".join(f"{p['name']}({p['total_minutes']}min)" for p in s["early_persons"])
            print(f"    早退人员: {names}")
        if s["weekend_workers"]:
            print(f"    周末轮值: {', '.join(s['weekend_workers'])}")

    return output


def main():
    parser = argparse.ArgumentParser(description="Analyze attendance records for Zhuihui Branch")
    parser.add_argument("input_file", help="Input Excel file (.xlsx)")
    parser.add_argument("--output", "-o", help="Output JSON file (default: <input>_result.json)")
    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Error: File not found: {args.input_file}")
        sys.exit(1)

    output = analyze(args.input_file)

    # Write JSON output
    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(os.path.basename(args.input_file))[0]
        output_path = f"{base}_result.json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n💾 Results saved to: {output_path}")


if __name__ == "__main__":
    main()
