#!/usr/bin/env python3
"""
生成 iOS MobPush 最小配置模板 Excel 文件
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def style_header(ws, headers, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.freeze_panes = "A2"


def auto_width(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


def create_basic_info_sheet(wb):
    ws = wb.create_sheet(title="基础信息")
    headers = ["配置项", "说明", "您的填写"]
    style_header(ws, headers, "4472C4")

    rows = [
        ["appKey", "MobTech 后台申请得到的 AppKey，必填", ""],
        ["appSecret", "MobTech 后台申请得到的 AppSecret，必填", ""],
        ["apnsAuthMode", "APNs 鉴权方式，只填 p12、p8 或 已存在", "已存在"],
        ["needLiveActivity", "本次是否需要同时接入 Live Activity，只填 是 或 否", "否"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, 6):
        ws.cell(row=row, column=1).font = Font(bold=True, color="C00000")

    auto_width(ws, {"A": 24, "B": 70, "C": 20})
    return ws


def create_privacy_sheet(wb):
    ws = wb.create_sheet(title="隐私与推送")
    headers = ["事项", "说明"]
    style_header(ws, headers, "ED7D31")

    rows = [
        ["隐私政策", "首次冷启动时应向用户展示《隐私政策》，并在同意后才能调用 MobPush 相关能力。"],
        ["隐私回传", "用户同意后，必须先调用 MobSDK.uploadPrivacyPermissionStatus，再使用 MobPush 能力。"],
        ["严格合规模式", "默认 plist 需配置 MOBNetLater = 2。"],
        ["App Store 隐私标签", "需参考 MobTech 提供的 App Store Connect 后台隐私数据项配置文档进行配置。"],
        ["推送能力", "项目需开启 Push Notifications 和 Background Modes > Remote notifications。"],
        ["APNs 鉴权材料", "p12 或 p8、Key ID、Team ID、Bundle ID 等材料不放在最小模板里，由开发者在苹果后台和 MobPush 控制台准备。"],
        ["扩展业务主动控制器", "本轮不主动收集该配置；如未来需要控制地理位置、WiFi、IP 等扩展业务数据采集，可参考最终文档说明。"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 22, "B": 92})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 38
    return ws


def create_instructions_sheet(wb):
    ws = wb.create_sheet(title="填写说明")
    headers = ["说明项", "详细内容"]
    style_header(ws, headers, "70AD47")

    rows = [
        ["填写顺序", "1. 在 MobTech 后台创建应用并获取 appKey / appSecret；2. 准备 APNs 鉴权材料；3. 只填写基础信息 Sheet；4. 填完后告诉 Agent “填好了”。"],
        ["最小原则", "不要在表里填写 Bundle ID、Target、Info.plist 路径、入口类名、推送回调方法名，这些应由 Agent 扫描工程自动推断。"],
        ["鉴权方式", "如果控制台已完成 APNs 配置，可填“已存在”；如果本次还没配好，则根据计划选择 p12 或 p8。"],
        ["Live Activity", "如果本次只想先跑通基础推送，建议保持“否”；后续需要时再单独接入。"],
        ["参考链接", "创建应用流程：https://www.mob.com/wiki/detailed?wiki=494&id=136\n合规指南：https://www.mob.com/wiki/detailed?wiki=501&id=136\n集成指南：https://www.mob.com/wiki/detailed?wiki=502&id=136\nSDK API：https://www.mob.com/wiki/detailed?wiki=503&id=136\n证书配置：https://www.mob.com/wiki/detailed?wiki=504&id=136"],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, {"A": 18, "B": 98})
    for row in range(2, len(rows) + 2):
        ws.row_dimensions[row].height = 56
    return ws


def main():
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    output_path = Path(__file__).resolve().parent / "MobPush_iOS_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()

