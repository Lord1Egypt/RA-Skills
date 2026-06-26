#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from chronic_disease_review import (
    DEFAULT_BASE,
    _infer_label,
    _resolve_disease_code,
    call_review_by_ocr,
    validate_ocr_data,
)
from format_review_nl import build_natural_language

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


class PreprocessError(ValueError):
    pass


TEXT_FILE_TYPES = {"txt", "md"}
IMAGE_FILE_TYPES = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
TABLE_FILE_TYPES = {"csv", "xlsx", "xls"}
SUPPORTED_FILE_TYPES = TEXT_FILE_TYPES | IMAGE_FILE_TYPES | TABLE_FILE_TYPES | {"json", "pdf", "doc", "docx"}
DOC_TYPE_RULES = [
    ("住院病案首页", ("住院病案首页", "病案首页")),
    ("出院记录", ("出院记录",)),
    ("入院记录", ("入院记录",)),
    ("诊断证明", ("诊断证明", "诊断证明书")),
    ("检验记录", ("检验报告", "检验记录")),
    ("检查记录", ("检查记录", "检查报告", "心电图报告")),
    ("长期医嘱", ("长期医嘱", "长期医嘱单")),
    ("体温记录", ("体温记录",)),
    ("身份证", ("居民身份证", "身份证")),
    ("社会保障卡", ("社会保障卡", "医保卡")),
    ("客户截图", ("审核结果", "认定材料", "门诊慢特病购药信息")),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preprocess pdf/doc/docx/xls/xlsx/csv/txt/json input into OCR-array JSON, then run chronic disease review."
    )
    parser.add_argument("--input", required=True, help="Input file path.")
    parser.add_argument(
        "--input-type",
        default="auto",
        choices=["auto", *sorted(SUPPORTED_FILE_TYPES)],
        help="Explicit input type. Default: auto.",
    )
    parser.add_argument("--sheet", default="", help="Optional sheet name for xlsx/xls inputs.")
    parser.add_argument("--encoding", default="utf-8", help="Text encoding for txt/csv inputs. Default: utf-8.")
    parser.add_argument(
        "--disease-code",
        required=True,
        help="糖尿病/高血压 or aliases diabetes/hypertension/dm/htn. Also supports auto.",
    )
    parser.add_argument("--review-type", default="慢病审核", help="review_type (default: 慢病审核)")
    parser.add_argument("--base", default=DEFAULT_BASE, help=f"Service base URL (default: {DEFAULT_BASE})")
    parser.add_argument("--llm-model", default="", help="Optional llm_model.")
    parser.add_argument("--timeout", type=int, default=0, help="HTTP timeout seconds. 0 means wait forever.")
    parser.add_argument("--output-json", default="", help="Path to save raw response JSON.")
    parser.add_argument("--output-text", default="", help="Path to save natural language summary.")
    parser.add_argument(
        "--save-prepared",
        action="store_true",
        help="Save the normalized OCR array next to the outputs for debugging.",
    )
    return parser.parse_args()


def detect_input_type(path: Path, explicit: str) -> str:
    if explicit != "auto":
        return explicit
    suffix = path.suffix.lower().lstrip(".")
    if suffix in SUPPORTED_FILE_TYPES:
        return suffix
    raise PreprocessError(f"Unsupported input file type: {path.suffix or '(none)'}")


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def normalize_to_int(value: Any, default: int) -> int:
    try:
        if value is None or str(value).strip() == "":
            return default
        return int(float(str(value).strip()))
    except Exception:
        return default


def infer_doc_type(text: str, fallback: str = "未分类文书") -> str:
    content = text or ""
    for doc_type, keywords in DOC_TYPE_RULES:
        if any(keyword in content for keyword in keywords):
            return doc_type
    return fallback


def read_text_file(path: Path, encoding: str) -> str:
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8-sig")


def read_json_file(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        return [[str(cell).strip() for cell in row] for row in reader]


def read_xlsx_tables(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise PreprocessError("openpyxl is required to parse xlsx inputs.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    selected_names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    tables: List[Dict[str, Any]] = []
    for name in selected_names:
        if name not in workbook.sheetnames:
            raise PreprocessError(f"Sheet not found: {name}")
        sheet = workbook[name]
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"name": name, "rows": rows})
    if not tables:
        raise PreprocessError("No non-empty rows found in workbook.")
    return tables


def extract_docx_text(path: Path) -> str:
    paragraphs: List[str] = []
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    if not paragraphs:
        raise PreprocessError("No text found in docx document.")
    return "\n".join(paragraphs)


def extract_with_soffice(path: Path) -> str:
    office_bin = shutil_which("soffice") or shutil_which("libreoffice")
    if not office_bin:
        raise PreprocessError("libreoffice/soffice not found for office document conversion.")
    with tempfile.TemporaryDirectory(prefix="med-chronic-review-") as tmp_dir:
        proc = subprocess.run(
            [
                office_bin,
                "--headless",
                "--convert-to",
                "txt:Text",
                "--outdir",
                tmp_dir,
                str(path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        out_path = Path(tmp_dir) / f"{path.stem}.txt"
        if proc.returncode != 0 or not out_path.exists():
            raise PreprocessError(f"Failed to convert office document to text: {proc.stderr.strip() or proc.stdout.strip()}")
        return out_path.read_text(encoding="utf-8", errors="replace")


def extract_pdf_pages(path: Path) -> List[str]:
    if PdfReader is not None:
        try:
            reader = PdfReader(str(path))
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
            pages = [page for page in pages if page]
            if pages:
                return pages
        except Exception:
            pass

    pdf_to_text = shutil_which("pdftotext")
    if pdf_to_text:
        proc = subprocess.run(
            [pdf_to_text, "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return [proc.stdout]
    raise PreprocessError("Unable to extract text from pdf.")


def extract_image_text(path: Path) -> str:
    tesseract_bin = shutil_which("tesseract")
    if not tesseract_bin:
        raise PreprocessError("tesseract not found for image OCR.")
    langs = detect_tesseract_langs(tesseract_bin)
    lang_arg = "chi_sim+eng" if "chi_sim" in langs and "eng" in langs else None
    cmd = [tesseract_bin, str(path), "stdout"]
    if lang_arg:
        cmd.extend(["-l", lang_arg])
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    if proc.returncode != 0 or not proc.stdout.strip():
        raise PreprocessError(f"Image OCR failed: {proc.stderr.strip() or 'no text returned'}")
    return proc.stdout


def detect_tesseract_langs(tesseract_bin: str) -> Sequence[str]:
    proc = subprocess.run(
        [tesseract_bin, "--list-langs"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return []
    return [line.strip() for line in proc.stdout.splitlines()[1:] if line.strip()]


def shutil_which(name: str) -> Optional[str]:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def load_input_artifact(path: Path, input_type: str, encoding: str, sheet_name: str) -> Dict[str, Any]:
    if input_type in TEXT_FILE_TYPES:
        return {"kind": "pages", "pages": [read_text_file(path, encoding)]}
    if input_type == "json":
        return {"kind": "json", "data": read_json_file(path)}
    if input_type == "csv":
        return {"kind": "tables", "tables": [{"name": path.stem, "rows": read_csv_rows(path, encoding)}]}
    if input_type == "xlsx":
        return {"kind": "tables", "tables": read_xlsx_tables(path, sheet_name)}
    if input_type == "xls":
        return {"kind": "pages", "pages": [extract_with_soffice(path)]}
    if input_type == "docx":
        return {"kind": "pages", "pages": [extract_docx_text(path)]}
    if input_type == "doc":
        return {"kind": "pages", "pages": [extract_with_soffice(path)]}
    if input_type == "pdf":
        return {"kind": "pages", "pages": extract_pdf_pages(path)}
    if input_type in IMAGE_FILE_TYPES:
        return {"kind": "pages", "pages": [extract_image_text(path)]}
    raise PreprocessError(f"Unsupported input type: {input_type}")


def build_ocr_item(
    *,
    file_name: str,
    page: int,
    doc_type: str,
    text: str,
) -> Optional[Dict[str, Any]]:
    content = (text or "").strip()
    if not content:
        return None
    return {
        "fileName": file_name,
        "page": page,
        "docType": doc_type or infer_doc_type(content),
        "ocrText": content,
    }


def coerce_ocr_items(items: List[Any], source_name: str) -> List[Dict[str, Any]]:
    ocr_items: List[Dict[str, Any]] = []
    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        text = item.get("ocrText") or item.get("text") or item.get("content") or item.get("format_page_text")
        if not isinstance(text, str) or not text.strip():
            continue
        file_name = str(item.get("fileName") or item.get("file") or item.get("name") or source_name)
        page = normalize_to_int(item.get("page") or item.get("pageNo") or item.get("page_num"), idx)
        doc_type = str(item.get("docType") or item.get("type") or item.get("documentType") or infer_doc_type(text))
        built = build_ocr_item(file_name=file_name, page=page, doc_type=doc_type, text=text)
        if built:
            ocr_items.append(built)
    if not ocr_items:
        raise PreprocessError("JSON array did not contain recognizable OCR items.")
    return validate_ocr_data(ocr_items)


def docs_to_ocr_items(docs: List[Any], source_name: str) -> List[Dict[str, Any]]:
    ocr_items: List[Dict[str, Any]] = []
    for idx, doc in enumerate(docs, start=1):
        if not isinstance(doc, dict):
            continue
        text = doc.get("format_page_text") or doc.get("text") or doc.get("content")
        if not isinstance(text, str) or not text.strip():
            continue
        doc_type = str(doc.get("docType") or doc.get("doc_name") or doc.get("docName") or infer_doc_type(text))
        built = build_ocr_item(file_name=source_name, page=idx, doc_type=doc_type, text=text)
        if built:
            ocr_items.append(built)
    if not ocr_items:
        raise PreprocessError("medicalRecord/docs JSON did not contain recognizable document text.")
    return validate_ocr_data(ocr_items)


def ocr_array_from_json(data: Any, source_name: str) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return coerce_ocr_items(data, source_name)

    if isinstance(data, dict):
        for key in ("ocr_data", "items", "pages", "results"):
            if isinstance(data.get(key), list):
                return coerce_ocr_items(list(data[key]), source_name)
        if isinstance(data.get("medicalRecord"), dict):
            docs = data["medicalRecord"].get("docs")
            if isinstance(docs, list):
                return docs_to_ocr_items(docs, source_name)
        if isinstance(data.get("docs"), list):
            return docs_to_ocr_items(list(data["docs"]), source_name)
        if isinstance(data.get("text"), str):
            item = build_ocr_item(file_name=source_name, page=1, doc_type=infer_doc_type(data["text"]), text=data["text"])
            if item:
                return validate_ocr_data([item])

    raise PreprocessError("Unsupported JSON schema for chronic disease review.")


def ocr_array_from_tables(tables: List[Dict[str, Any]], source_name: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for table in tables:
        rows = table.get("rows") or []
        if not rows:
            continue
        header_map = {normalize_header(cell): idx for idx, cell in enumerate(rows[0]) if cell.strip()}
        text_idx = first_matching_index(header_map, ("ocrtext", "text", "content", "内容", "文本"))
        page_idx = first_matching_index(header_map, ("page", "页码", "pageno"))
        file_idx = first_matching_index(header_map, ("filename", "file", "name", "文件名"))
        doc_type_idx = first_matching_index(header_map, ("doctype", "type", "文书类型", "文档类型"))
        data_rows = rows[1:] if text_idx is not None else rows

        if text_idx is not None:
            for row_idx, row in enumerate(data_rows, start=1):
                if text_idx >= len(row):
                    continue
                text = row[text_idx].strip()
                if not text:
                    continue
                file_name = row[file_idx].strip() if file_idx is not None and file_idx < len(row) and row[file_idx].strip() else source_name
                page = normalize_to_int(row[page_idx] if page_idx is not None and page_idx < len(row) else None, row_idx)
                doc_type_value = row[doc_type_idx].strip() if doc_type_idx is not None and doc_type_idx < len(row) else infer_doc_type(text, table["name"])
                built = build_ocr_item(file_name=file_name, page=page, doc_type=doc_type_value, text=text)
                if built:
                    items.append(built)
            continue

        merged_lines = [" | ".join(cell for cell in row if cell) for row in rows if any(cell for cell in row)]
        merged_text = "\n".join(line for line in merged_lines if line.strip()).strip()
        if merged_text:
            built = build_ocr_item(
                file_name=source_name,
                page=len(items) + 1,
                doc_type=infer_doc_type(merged_text, table["name"]),
                text=merged_text,
            )
            if built:
                items.append(built)

    if not items:
        raise PreprocessError("Could not derive OCR text rows from the table input.")
    return validate_ocr_data(items)


def first_matching_index(header_map: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(normalize_header(alias))
        if idx is not None:
            return idx
    return None


def ocr_array_from_pages(pages: List[str], source_name: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for idx, page_text in enumerate(pages, start=1):
        built = build_ocr_item(
            file_name=source_name,
            page=idx,
            doc_type=infer_doc_type(page_text),
            text=page_text,
        )
        if built:
            items.append(built)
    if not items:
        raise PreprocessError("No non-empty text pages were extracted from the input.")
    return validate_ocr_data(items)


def preprocess_to_ocr_array(path: Path, input_type: str, encoding: str, sheet_name: str) -> List[Dict[str, Any]]:
    artifact = load_input_artifact(path, input_type, encoding, sheet_name)
    kind = artifact["kind"]
    if kind == "pages":
        return ocr_array_from_pages(artifact["pages"], path.name)
    if kind == "json":
        return ocr_array_from_json(artifact["data"], path.name)
    if kind == "tables":
        return ocr_array_from_tables(artifact["tables"], path.name)
    raise PreprocessError(f"Unsupported artifact kind: {kind}")


def resolve_or_infer_disease_code(raw: str, ocr_data: List[Dict[str, Any]]) -> str:
    value = (raw or "").strip()
    if value and value.lower() != "auto":
        resolved = _resolve_disease_code(value)
        if resolved is None:
            raise PreprocessError("Unable to resolve --disease-code.")
        return resolved

    content = "\n".join(item.get("ocrText", "") for item in ocr_data)
    if "糖尿病" in content:
        return "糖尿病"
    if "高血压" in content:
        return "高血压"
    raise PreprocessError("Could not infer disease code from OCR text; please pass --disease-code explicitly.")


def build_prepared_path(output_json_path: Path) -> Path:
    return output_json_path.with_name(f"{output_json_path.stem}.prepared.json")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"✗ Error: Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        input_type = detect_input_type(input_path, args.input_type)
        ocr_data = preprocess_to_ocr_array(input_path, input_type, args.encoding, args.sheet)
        disease_code = resolve_or_infer_disease_code(args.disease_code, ocr_data)
    except Exception as exc:
        print(f"✗ Preprocess Error: {exc}", file=sys.stderr)
        return 1

    req_body: Dict[str, Any] = {"review_type": args.review_type or "慢病审核", "ocr_data": ocr_data, "disease_code": disease_code}
    if args.llm_model:
        req_body["llm_model"] = args.llm_model

    try:
        resp = call_review_by_ocr(args.base, req_body, timeout=args.timeout)
    except Exception as exc:
        print(f"✗ Error: {exc}", file=sys.stderr)
        return 1

    label = _infer_label(disease_code)
    default_base = Path("../runs/med-chronic-disease-review")
    out_json = Path(args.output_json) if args.output_json else (default_base / f"{label}_resp.json")
    out_text = Path(args.output_text) if args.output_text else (default_base / f"{label}_resp.txt")

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(resp, ensure_ascii=False, indent=2), encoding="utf-8")

    text = build_natural_language(resp)
    out_text.parent.mkdir(parents=True, exist_ok=True)
    out_text.write_text(text, encoding="utf-8")

    if args.save_prepared:
        prepared_path = build_prepared_path(out_json)
        prepared_path.write_text(json.dumps(ocr_data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✓ Prepared OCR array saved to: {prepared_path}")

    print(f"✓ Saved raw JSON to: {out_json}")
    print(f"✓ Saved natural language to: {out_text}")
    print("\n--- Natural language preview ---")
    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
