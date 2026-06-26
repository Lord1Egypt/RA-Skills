#!/usr/bin/env python3
"""
快递批量查询脚本

调用快递查询API批量获取物流信息

凭证参数: PlatformID, MemberID, APIKey
"""

import sys
import json
import argparse
import hashlib
import base64
import urllib.parse
import csv
from datetime import datetime
from typing import List, Dict

# 使用标准requests库
import requests

# 导入openpyxl用于生成XLS文件
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExpressSignUtils:
    """接口签名工具类（严格遵循文档签名规则）"""

    @staticmethod
    def md5_encrypt(content: str) -> str:
        """MD5加密（返回16进制小写字符串）"""
        md5_obj = hashlib.md5()
        md5_obj.update(content.encode("utf-8"))
        return md5_obj.hexdigest()

    @staticmethod
    def generate_data_sign(request_data: str, api_key: str) -> str:
        """
        生成DataSign签名
        步骤：原始RequestData + APIKey → MD5 → Base64 → URL编码（UTF-8）
        :param request_data: 未URL编码的业务请求数据（JSON字符串）
        :param api_key: 小递查查分配的APIKey
        :return: 最终DataSign签名
        """
        # 1. 拼接原始字符串
        original_str = request_data + api_key

        # 2. MD5加密
        md5_hex = ExpressSignUtils.md5_encrypt(original_str)

        # 3. Base64编码（对MD5结果编码）
        base64_bytes = base64.b64encode(md5_hex.encode("utf-8"))
        base64_str = base64_bytes.decode("utf-8")

        # 4. URL编码（UTF-8）
        data_sign = urllib.parse.quote(base64_str, encoding="utf-8")

        return data_sign


def generate_xls_file(results: Dict, output_path: str = None) -> str:
    """
    生成XLS文件，包含批量查询结果

    参数:
        results: 查询结果字典
        output_path: 输出文件路径（可选，不指定则自动生成）

    返回:
        生成的XLS文件路径
    """
    if not HAS_OPENPYXL:
        raise ImportError("未安装openpyxl库，无法生成XLS文件。请运行: pip install openpyxl")

    # 生成文件名
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"express_query_results_{timestamp}.xlsx"

    # 获取绝对路径用于提示
    import os
    abs_path = os.path.abspath(output_path)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "快递查询结果"

    # 定义表头
    headers = [
        "序号",
        "运单号",
        "快递公司",
        "物流状态码",
        "物流状态",
        "创建时间",
        "揽收时间",
        "揽收超时",
        "首次发出时间",
        "首次发出超时",
        "首次发出耗时(分钟)",
        "首次到达时间",
        "首次到达超时",
        "首次到达耗时(分钟)",
        "离开首个转运中心耗时(分钟)",
        "半程停滞耗时(分钟)",
        "派件开始时间",
        "派件超时",
        "签收时间",
        "签收耗时(分钟)",
        "签收超时",
        "最后更新时间",
        "本次请求时间",
        "揽收快递员",
        "揽收电话",
        "揽收网点",
        "揽收城市",
        "首次到达转运中心名称",
        "派送网点",
        "派送城市",
        "派送快递员",
        "派送电话",
        "快递公司编码",
        "物流轨迹数",
        "最后轨迹描述"
    ]

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    row_num = 2
    for idx, (express_no, result) in enumerate(results.items(), 1):
        if result.get("success") and "data" in result:
            data = result["data"]

            # 提取物流轨迹最后一条描述
            trace_list = data.get("TraceList", [])
            last_trace = trace_list[0].get("desc", "") if trace_list else ""

            # 提取所有字段
            row_data = [
                idx,  # 序号
                data.get("LogisticCode", express_no),  # 运单号
                data.get("ExpressName", ""),  # 快递公司
                data.get("State", ""),  # 物流状态码
                data.get("StateDesc", ""),  # 物流状态
                data.get("Create_Time", ""),  # 创建时间
                data.get("Pickup_Time", ""),  # 揽收时间
                data.get("Pickup_TimeOut", ""),  # 揽收超时
                data.get("FirstIssue_Time", ""),  # 首次发出时间
                data.get("FirstIssue_TimeOut", ""),  # 首次发出超时
                data.get("FirstIssue_TimeSpan", ""),  # 首次发出耗时
                data.get("FirstArrival_Time", ""),  # 首次到达时间
                data.get("FirstArrival_TimeOut", ""),  # 首次到达超时
                data.get("FirstArrival_TimeSpan", ""),  # 首次到达耗时
                data.get("LeaveDistributionCenter_TimeSpan", ""),  # 离开首个转运中心耗时
                data.get("HalfwayStagnation_TimeSpan", ""),  # 半程停滞耗时
                data.get("Dispatch_Time", ""),  # 派件开始时间
                data.get("Dispatch_TimeOut", ""),  # 派件超时
                data.get("SignFor_Time", ""),  # 签收时间
                data.get("SignFor_TimeSpan", ""),  # 签收耗时
                data.get("SignFor_TimeOut", ""),  # 签收超时
                data.get("LastUpdate_Time", ""),  # 最后更新时间
                data.get("LastRequest_Time", ""),  # 本次请求时间
                data.get("Pickup_Courier", ""),  # 揽收快递员
                data.get("Pickup_Courier_Phone", ""),  # 揽收电话
                data.get("Pickup_Network", ""),  # 揽收网点
                data.get("Pickup_City", ""),  # 揽收城市
                data.get("FirstArrival_Name", ""),  # 首次到达转运中心名称
                data.get("Delivery_Network", ""),  # 派送网点
                data.get("Delivery_City", ""),  # 派送城市
                data.get("Delivery_Courier", ""),  # 派送快递员
                data.get("Delivery_Courier_Phone", ""),  # 派送电话
                data.get("ShipperCode", ""),  # 快递公司编码
                data.get("Count", 0),  # 物流轨迹数
                last_trace  # 最后轨迹描述
            ]

            # 写入数据行
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1
        else:
            # 查询失败的记录
            row_data = [
                idx,  # 序号
                express_no,  # 运单号
                "",  # 快递公司
                "",  # 物流状态码
                "查询失败",  # 物流状态
                "",  # 创建时间
                "",  # 揽收时间
                "",  # 揽收超时
                "",  # 首次发出时间
                "",  # 首次发出超时
                "",  # 首次发出耗时
                "",  # 首次到达时间
                "",  # 首次到达超时
                "",  # 首次到达耗时
                "",  # 离开首个转运中心耗时
                "",  # 半程停滞耗时
                "",  # 派件开始时间
                "",  # 派件超时
                "",  # 签收时间
                "",  # 签收耗时
                "",  # 签收超时
                "",  # 最后更新时间
                "",  # 本次请求时间
                "",  # 揽收快递员
                "",  # 揽收电话
                "",  # 揽收网点
                "",  # 揽收城市
                "",  # 首次到达转运中心名称
                "",  # 派送网点
                "",  # 派送城市
                "",  # 派送快递员
                "",  # 派送电话
                "",  # 快递公司编码
                0,  # 物流轨迹数
                result.get("error", "")  # 错误信息
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = ""
        temp = col_idx
        while temp > 0:
            temp, remainder = divmod(temp - 1, 26)
            col_letter = chr(65 + remainder) + col_letter

        # 根据列号设置宽度
        if col_idx == 1:
            width = 6
        elif col_idx == 2:
            width = 20
        elif col_idx == 3:
            width = 15
        elif col_idx in [4, 5]:
            width = 12
        elif col_idx in [6, 7, 9, 12, 17, 19, 22, 23]:
            width = 20
        elif col_idx in [8, 10, 13, 18, 21, 27, 30, 32]:
            width = 12
        elif col_idx in [11, 14, 15, 16, 20, 33]:
            width = 15
        elif col_idx in [24, 25, 31]:
            width = 12
        elif col_idx in [26, 28, 29]:
            width = 20
        elif col_idx == 34:
            width = 10
        elif col_idx == 35:
            width = 50
        else:
            width = 12

        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(output_path)

    # 验证文件是否存在
    if not os.path.exists(abs_path):
        raise Exception(f"文件生成失败，文件路径: {abs_path}")

    # 返回绝对路径
    return abs_path


def query_express(express_list: List[Dict], platform_id: str, member_id: str, api_key: str) -> Dict:
    """
    批量查询快递物流信息

    参数:
        express_list: 快递列表，每个元素包含 number 和 phone_tail
        platform_id: 平台ID
        member_id: 会员ID
        api_key: API密钥

    返回:
        查询结果字典，包含每个单号的物流轨迹
    """
    # 构建请求参数
    url = "https://api.xdccy.com/IsvApi/GetXdccTrackInfoV2"
    timeout = 8

    # 逐个查询
    results = {}

    for item in express_list:
        number = item["number"]
        phone_tail = item.get("phone_tail", "")
        number = number.strip()
        if not number:
            continue

        try:
            # 检查是否为顺丰快递且未提供手机尾号
            if number.upper().startswith('SF') and not phone_tail:
                results[number] = {
                    "success": False,
                    "error": "查询顺丰快递需要手机尾号，请按格式提供：运单号:尾号（如 SF5139226181410:6553）",
                    "tracks": []
                }
                continue

            # 1. 构建业务请求参数（RequestData内容）
            request_data_dict = {
                "MemberID": member_id,
                "No": number,
                "ReturnFormat": "JSON"
            }

            # 如果提供了手机尾号，添加到请求中
            if phone_tail:
                request_data_dict["Tel"] = phone_tail

            # 2. 序列化为未编码的JSON字符串
            original_request_data = json.dumps(request_data_dict, ensure_ascii=False)

            # 3. 生成DataSign签名
            data_sign = ExpressSignUtils.generate_data_sign(original_request_data, api_key)

            # 4. 对RequestData进行URL编码
            encoded_request_data = urllib.parse.quote(original_request_data, encoding="utf-8")

            # 5. 构建请求参数（表单格式）
            form_data = {
                "PlatformID": platform_id,
                "RequestData": encoded_request_data,
                "DataSign": data_sign
            }

            # 6. 发送POST请求
            response = requests.post(
                url=url,
                data=form_data,
                headers={"Content-Type": "application/x-www-form-urlencoded;charset=utf-8"},
                timeout=timeout
            )
            response.encoding = "utf-8"

            # 检查HTTP状态码
            if response.status_code >= 400:
                raise Exception(f"HTTP请求失败: 状态码 {response.status_code}, 响应内容: {response.text}")

            data = response.json()

            # 错误处理（API返回的Code是数字200）
            if data.get("Code") != 200:
                error_msg = data.get("Message", "未知错误")

                # 检查是否是渠道未配置的问题
                error_lower = error_msg.lower()
                if any(keyword in error_lower for keyword in ['未配置', '未开通', '无渠道', '不支持', '渠道', 'channel']):
                    error_msg = f"{error_msg}。请登录小递查查官网（www.xdccy.com）配置查询渠道"

                results[number] = {
                    "success": False,
                    "error": error_msg,
                    "tracks": []
                }
                continue

            # 提取物流信息 - 直接返回所有接口字段
            express_data = data.get("Data", {})

            # 检查API返回的result字段
            api_result = express_data.get("result", True)
            api_message = express_data.get("message", "")

            if not api_result:
                # API返回result=false，表示查询失败
                results[number] = {
                    "success": False,
                    "error": f"{api_message}",
                    "tracks": []
                }
                continue

            # 检查是否有物流轨迹
            trace_list = express_data.get("TraceList", [])
            if not trace_list or len(trace_list) == 0:
                # 检查是否有其他字段
                state = express_data.get("State", "")
                state_desc = express_data.get("StateDesc", "")
                results[number] = {
                    "success": False,
                    "error": f"暂无物流轨迹信息（状态：{state_desc}）",
                    "tracks": []
                }
                continue

            # 返回完整的接口数据，不做任何字段过滤
            results[number] = {
                "success": True,
                "data": express_data  # 直接返回原始的Data对象，包含所有字段
            }

        except requests.exceptions.Timeout:
            results[number] = {
                "success": False,
                "error": "接口请求超时（超过8秒）",
                "tracks": []
            }
        except requests.exceptions.ConnectionError:
            results[number] = {
                "success": False,
                "error": "接口连接失败，请检查地址或网络",
                "tracks": []
            }
        except json.JSONDecodeError:
            results[number] = {
                "success": False,
                "error": "响应解析失败",
                "tracks": []
            }
        except requests.exceptions.RequestException as e:
            results[number] = {
                "success": False,
                "error": f"网络请求失败: {str(e)}",
                "tracks": []
            }
        except Exception as e:
            results[number] = {
                "success": False,
                "error": f"查询异常: {str(e)}",
                "tracks": []
            }

    return results


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="批量查询快递物流信息")
    parser.add_argument("--numbers", required=True, help="快递单号列表，逗号分隔（格式：单号:手机尾号，如 SF123456:6553）")
    parser.add_argument("--platform-id", required=True, help="平台ID (PlatformID)")
    parser.add_argument("--member-id", required=True, help="会员ID (MemberID)")
    parser.add_argument("--api-key", required=True, help="API密钥 (APIKey)")
    parser.add_argument("--output", help="输出文件路径（可选，支持xlsx/csv/txt/json格式）")

    args = parser.parse_args()

    # 验证凭证参数
    if not args.platform_id or not args.member_id or not args.api_key:
        error_msg = {
            "error": "缺少必要的凭证参数",
            "message": "请提供以下凭证参数：PlatformID、MemberID、APIKey",
            "how_to_get": "请访问小递查查官网获取凭证参数：https://xdccy.com",
            "required_params": {
                "platform_id": "平台ID (PlatformID)",
                "member_id": "会员ID (MemberID)",
                "api_key": "API密钥 (APIKey)"
            }
        }
        print(json.dumps(error_msg, ensure_ascii=False, indent=2))
        sys.exit(1)

    # 解析快递单号（支持格式：单号 或 单号:手机尾号）
    express_list = []
    for item in args.numbers.split(","):
        item = item.strip()
        if not item:
            continue
        if ":" in item:
            parts = item.split(":")
            express_list.append({"number": parts[0].strip(), "phone_tail": parts[1].strip()})
        else:
            express_list.append({"number": item, "phone_tail": ""})

    if not express_list:
        print(json.dumps({"error": "快递单号不能为空"}, ensure_ascii=False))
        sys.exit(1)

    # 提取纯单号列表用于显示
    express_numbers = [item["number"] for item in express_list]

    # 执行查询
    try:
        results = query_express(
            express_list,
            args.platform_id,
            args.member_id,
            args.api_key
        )

        # 判断是否为批量查询（单号数量 > 1）
        is_batch = len(express_numbers) > 1

        # 输出结果
        if args.output:
            if is_batch:
                # 批量查询：根据输出文件扩展名决定格式
                output_path = args.output
                file_ext = output_path.split('.')[-1].lower() if '.' in output_path else 'txt'

                if file_ext in ['xls', 'xlsx']:
                    # XLS/XLSX格式导出
                    xls_path = generate_xls_file(results, output_path)
                    print(f"\nExcel文件已生成: {xls_path}")
                elif file_ext == 'csv':
                    # CSV格式导出
                    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                        writer = csv.writer(f)
                        # 写入表头
                        headers = [
                            '序号', '运单号', '快递公司', '物流状态', '状态码',
                            '最后更新时间', '最新轨迹描述', '最新轨迹时间', '查询状态'
                        ]
                        writer.writerow(headers)

                        # 写入数据
                        for idx, (number, result) in enumerate(results.items(), 1):
                            if result.get("success"):
                                data = result.get("data", {})
                                trace_list = data.get("TraceList", [])
                                last_trace = trace_list[0] if trace_list else {}

                                row = [
                                    idx,
                                    number,
                                    data.get('ExpressName', ''),
                                    data.get('StateDesc', ''),
                                    data.get('State', ''),
                                    data.get('LastUpdate_Time', ''),
                                    last_trace.get('desc', ''),
                                    last_trace.get('time', ''),
                                    '成功'
                                ]
                            else:
                                row = [
                                    idx,
                                    number,
                                    '',
                                    '',
                                    '',
                                    '',
                                    '',
                                    '',
                                    f'失败: {result.get("error", "未知错误")}'
                                ]
                            writer.writerow(row)

                    print(f"\nCSV格式结果已保存至: {output_path}")
                else:
                    # TXT列表格式
                    if output_path.endswith('.json'):
                        output_path = output_path[:-5] + '.txt'

                    with open(output_path, 'w', encoding='utf-8') as f:
                        f.write("=" * 80 + "\n")
                        f.write(f"批量查询结果（共 {len(express_numbers)} 个单号）\n")
                        f.write("=" * 80 + "\n\n")

                        for idx, (number, result) in enumerate(results.items(), 1):
                            f.write(f"[{idx}] {number}\n")
                            f.write("-" * 80 + "\n")
                            if result.get("success"):
                                data = result.get("data", {})
                                f.write(f"  快递公司: {data.get('ExpressName', '未知')}\n")
                                f.write(f"  物流状态: {data.get('StateDesc', '未知')} ({data.get('State', '')})\n")
                                f.write(f"  最后更新: {data.get('LastUpdate_Time', '')}\n")

                                trace_list = data.get("TraceList", [])
                                if trace_list:
                                    last_trace = trace_list[0]
                                    f.write(f"  最新轨迹: {last_trace.get('desc', '')} ({last_trace.get('time', '')})\n")
                                else:
                                    f.write("  最新轨迹: 暂无物流信息\n")
                            else:
                                f.write(f"  查询失败: {result.get('error', '未知错误')}\n")
                            f.write("\n")

                        f.write("=" * 80 + "\n")
                        f.write(f"查询完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                        f.write("=" * 80 + "\n")

                    print(f"\n列表格式结果已保存至: {output_path}")
            else:
                # 单个查询：保存JSON格式
                with open(args.output, 'w', encoding='utf-8') as f:
                    json.dump(results, f, ensure_ascii=False, indent=2)
                print(f"\nJSON查询结果已保存至: {args.output}")
        else:
            if is_batch:
                # 批量查询：显示列表格式
                print("=" * 80)
                print(f"批量查询结果（共 {len(express_numbers)} 个单号）")
                print("=" * 80)

                for idx, (number, result) in enumerate(results.items(), 1):
                    print(f"\n[{idx}] {number}")
                    print("-" * 80)
                    if result.get("success"):
                        data = result.get("data", {})
                        print(f"  快递公司: {data.get('ExpressName', '未知')}")
                        print(f"  物流状态: {data.get('StateDesc', '未知')} ({data.get('State', '')})")
                        print(f"  最后更新: {data.get('LastUpdate_Time', '')}")

                        trace_list = data.get("TraceList", [])
                        if trace_list:
                            last_trace = trace_list[0]
                            print(f"  最新轨迹: {last_trace.get('desc', '')} ({last_trace.get('time', '')})")
                        else:
                            print("  最新轨迹: 暂无物流信息")
                    else:
                        print(f"  查询失败: {result.get('error', '未知错误')}")

                print("\n" + "=" * 80)
                print("提示: 如需导出CSV文件，请添加 --output results.csv 参数")
                print("提示: 如需导出TXT文件，请添加 --output results.txt 参数")
                print("=" * 80)
            else:
                # 单个查询：显示完整JSON结果
                print(json.dumps(results, ensure_ascii=False, indent=2))

    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
