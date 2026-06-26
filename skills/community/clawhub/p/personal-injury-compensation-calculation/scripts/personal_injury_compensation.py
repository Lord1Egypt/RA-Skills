from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable


MONEY_QUANT = Decimal("0.01")
INJURY_INTERPRETATION_REVISION_DATE = date(2022, 5, 1)
DISCLAIMER = (
    "免责声明：本计算结果基于现行法律规定及当前可获取的统计数据，仅供参考。"
    "赔偿标准每年更新，部分数据（如省级统计年鉴）存在时效性，实际计算可能有出入。"
    "最终赔偿金额需根据证据情况及法院最终判决确定，建议咨询专业律师。"
)
SPECIAL_CITIES = {"深圳", "厦门", "珠海", "汕头", "青岛", "大连", "宁波"}
DISABILITY_LEVEL_TO_RATIO = {
    1: Decimal("1.0"),
    2: Decimal("0.9"),
    3: Decimal("0.8"),
    4: Decimal("0.7"),
    5: Decimal("0.6"),
    6: Decimal("0.5"),
    7: Decimal("0.4"),
    8: Decimal("0.3"),
    9: Decimal("0.2"),
    10: Decimal("0.1"),
}
MENTAL_DAMAGE_REFERENCE = {
    1: Decimal("50000"),
    2: Decimal("45000"),
    3: Decimal("40000"),
    4: Decimal("35000"),
    5: Decimal("30000"),
    6: Decimal("25000"),
    7: Decimal("20000"),
    8: Decimal("15000"),
    9: Decimal("10000"),
    10: Decimal("5000"),
}
LAW_BASIS_BY_ITEM = {
    "医疗费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第六条 —— 医疗费"],
    "误工费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第七条 —— 误工费"],
    "护理费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第八条 —— 护理费"],
    "交通费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第九条 —— 交通费"],
    "住院伙食补助费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第十条 —— 住院伙食补助费、住宿费"],
    "营养费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第十一条 —— 营养费"],
    "残疾赔偿金": ["第一千一百七十九条 —— 人身损害赔偿范围", "第十二条 —— 残疾赔偿金（重要！）", "第二十二条 —— 赔偿标准的适用"],
    "死亡赔偿金": ["第一千一百七十九条 —— 人身损害赔偿范围", "第十五条 —— 死亡赔偿金", "第二十二条 —— 赔偿标准的适用"],
    "丧葬费": ["第一千一百七十九条 —— 人身损害赔偿范围", "第十四条 —— 丧葬费", "第二十二条 —— 赔偿标准的适用"],
    "被扶养人生活费": ["第十六条 —— 被扶养人生活费（计入残疾/死亡赔偿金）", "第十七条 —— 被扶养人范围", "第二十二条 —— 赔偿标准的适用"],
    "精神抚慰金": ["第一千一百八十三条 —— 精神损害赔偿", "精神损害赔偿要点"],
    "住宿费": ["第十条 —— 住院伙食补助费、住宿费"],
    "鉴定费": [],
    "财产损失": [],
    "残疾辅助器具费": ["第一千一百七十九条 —— 人身损害赔偿范围"],
}


class CalculationError(Exception):
    """Raised when case input is incomplete or contradictory."""


class ReferenceDataError(Exception):
    """Raised when required statistics are not available in local references."""


def decimalize(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("，", "").strip()
        if not cleaned or cleaned == "—":
            raise ValueError("empty numeric value")
        return Decimal(cleaned)
    raise TypeError(f"Unsupported numeric value: {value!r}")


def to_money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def format_money(value: Decimal) -> str:
    return f"{to_money(value):,.2f}"


def normalize_region_name(value: str | None) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"\s+", "", value)
    for suffix in (
        "维吾尔自治区",
        "回族自治区",
        "壮族自治区",
        "特别行政区",
        "自治区",
        "自治州",
        "省",
        "市",
        "区",
    ):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    return cleaned


def parse_case_date(value: str) -> date:
    text = (value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%Y/%m"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%Y-%m", "%Y/%m"):
                return date(parsed.year, parsed.month, 1)
            return parsed.date()
        except ValueError:
            continue
    raise CalculationError(f"无法解析日期：{value!r}。支持 YYYY-MM 或 YYYY-MM-DD。")


def markdown_tables(text: str) -> list[list[dict[str, str]]]:
    lines = text.splitlines()
    tables: list[list[dict[str, str]]] = []
    index = 0
    while index < len(lines):
        if lines[index].strip().startswith("|") and index + 1 < len(lines):
            separator = lines[index + 1].strip()
            if separator.startswith("|") and re.search(r"[-:]", separator):
                header = [cell.strip() for cell in lines[index].strip().strip("|").split("|")]
                index += 2
                rows: list[dict[str, str]] = []
                while index < len(lines) and lines[index].strip().startswith("|"):
                    raw_row = [cell.strip() for cell in lines[index].strip().strip("|").split("|")]
                    if len(raw_row) < len(header):
                        raw_row += [""] * (len(header) - len(raw_row))
                    rows.append(dict(zip(header, raw_row)))
                    index += 1
                tables.append(rows)
                continue
        index += 1
    return tables


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


@dataclass
class ReferenceStore:
    formula_reference_text: str
    national_disposable_income: dict[int, Decimal]
    wage_tables: dict[str, dict[str, dict[int, Decimal]]]
    law_articles: dict[str, str]

    @classmethod
    def load(cls, references_dir: Path) -> "ReferenceStore":
        formula_text = load_text(references_dir / "formulas.md")
        income_text = load_text(references_dir / "disposable_income.md")
        wage_text = load_text(references_dir / "provincial_avg_wage.md")
        law_text = load_text(references_dir / "law_articles.md")
        return cls(
            formula_reference_text=formula_text,
            national_disposable_income=parse_disposable_income(income_text),
            wage_tables=parse_provincial_wages(wage_text),
            law_articles=parse_law_articles(law_text),
        )


def parse_disposable_income(text: str) -> dict[int, Decimal]:
    tables = markdown_tables(text)
    if not tables:
        raise ReferenceDataError("references/disposable_income.md 中未找到可解析的数据表。")

    results: dict[int, Decimal] = {}
    for row in tables[0]:
        year_raw = row.get("年份（上一年度）", "").strip()
        income_raw = row.get("人均可支配收入（元/年）", "").strip()
        if not year_raw or not income_raw:
            continue
        results[int(year_raw)] = decimalize(income_raw)
    if not results:
        raise ReferenceDataError("references/disposable_income.md 中未解析到全国可支配收入数据。")
    return results


def parse_provincial_wages(text: str) -> dict[str, dict[str, dict[int, Decimal]]]:
    caliber_map = {
        "## 一、城镇非私营单位就业人员平均工资（年薪，单位：元/年）": "non_private",
        "## 二、全口径城镇单位就业人员平均工资（月薪，单位：元/月）": "full",
        "## 三、城镇私营单位就业人员平均工资（年薪，单位：元/年）": "private",
    }
    sections = text.split("## ")
    results: dict[str, dict[str, dict[int, Decimal]]] = {"non_private": {}, "full": {}, "private": {}}

    for section in sections:
        if not section.strip():
            continue
        section_text = "## " + section
        first_line = section_text.splitlines()[0].strip()
        caliber = caliber_map.get(first_line)
        if not caliber:
            continue

        tables = markdown_tables(section_text)
        if not tables:
            continue
        for row in tables[0]:
            province = normalize_region_name(row.get("省份", ""))
            if not province:
                continue
            province_values: dict[int, Decimal] = {}
            for key, value in row.items():
                if key == "省份":
                    continue
                if not value or value == "—":
                    continue
                if re.fullmatch(r"\d{4}", key):
                    province_values[int(key)] = decimalize(value)
            if province_values:
                results[caliber][province] = province_values
    return results


def parse_law_articles(text: str) -> dict[str, str]:
    pieces = re.split(r"(?m)^(?:##|###) ", text)
    results: dict[str, str] = {}
    for piece in pieces:
        chunk = piece.strip()
        if not chunk:
            continue
        lines = chunk.splitlines()
        heading = lines[0].strip()
        body = "\n".join(line.rstrip() for line in lines[1:]).strip()
        results[heading] = body
    return results


@dataclass
class Dependent:
    age: int
    supporter_count: int = 1
    no_labor_capacity: bool = False
    annual_cap_share: Decimal = Decimal("0")

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "Dependent":
        age = int(payload.get("age"))
        supporter_count = int(payload.get("supporter_count", 1))
        if supporter_count <= 0:
            raise CalculationError("被扶养义务人数必须大于 0。")
        return cls(
            age=age,
            supporter_count=supporter_count,
            no_labor_capacity=bool(payload.get("no_labor_capacity", False)),
        )


@dataclass
class CalculationLine:
    category: str
    item: str
    formula: str
    amount: Decimal
    note: str = ""
    legal_basis: list[str] = field(default_factory=list)


@dataclass
class CaseData:
    case_type: str
    incident_date: date
    hearing_date: date
    victim_age: int
    court_province: str
    court_city: str = ""
    wage_caliber: str = "private"
    residency_type: str = "urban"
    rural_urban_eligible: bool = False
    disability_levels: list[int] = field(default_factory=list)
    dependents: list[Dependent] = field(default_factory=list)
    work_loss_days: int = 0
    lost_income_actual: Decimal | None = None
    annual_income_average: Decimal | None = None
    industry_average_annual_income: Decimal | None = None
    nursing_days: int = 0
    nursing_rate_per_day: Decimal | None = None
    nursing_annual_income: Decimal | None = None
    nutrition_days: int = 0
    nutrition_rate_per_day: Decimal = Decimal("50")
    hospital_days: int = 0
    hospital_food_rate_per_day: Decimal = Decimal("100")
    medical_expense: Decimal = Decimal("0")
    transport_expense: Decimal = Decimal("0")
    lodging_expense: Decimal = Decimal("0")
    appraisal_fee: Decimal = Decimal("0")
    property_loss: Decimal = Decimal("0")
    assistive_device_expense: Decimal = Decimal("0")
    mental_damage: Decimal | None = None
    auto_mental_damage: bool = False
    statistics_overrides: dict[str, Any] = field(default_factory=dict)

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "CaseData":
        case_type = str(payload.get("case_type", "")).strip().lower()
        if case_type not in {"injury", "death"}:
            raise CalculationError("case_type 仅支持 injury 或 death。")

        wage_caliber = str(payload.get("wage_caliber", "private")).strip().lower()
        if wage_caliber not in {"private", "full", "non_private"}:
            raise CalculationError("wage_caliber 仅支持 private、full、non_private。")

        residency_type = str(payload.get("residency_type", "urban")).strip().lower()
        if residency_type not in {"urban", "rural"}:
            raise CalculationError("residency_type 仅支持 urban 或 rural。")

        disability_levels = [int(level) for level in payload.get("disability_levels", [])]
        for level in disability_levels:
            if level not in DISABILITY_LEVEL_TO_RATIO:
                raise CalculationError(f"伤残等级必须是 1-10，收到：{level}")

        statistics_overrides = payload.get("statistics_overrides") or {}
        if not isinstance(statistics_overrides, dict):
            raise CalculationError("statistics_overrides 必须是对象。")

        def parse_decimal_or_none(key: str) -> Decimal | None:
            value = payload.get(key)
            if value in (None, ""):
                return None
            return decimalize(value)

        return cls(
            case_type=case_type,
            incident_date=parse_case_date(str(payload.get("incident_date", ""))),
            hearing_date=parse_case_date(str(payload.get("hearing_date", ""))),
            victim_age=int(payload.get("victim_age")),
            court_province=str(payload.get("court_province", "")).strip(),
            court_city=str(payload.get("court_city", "")).strip(),
            wage_caliber=wage_caliber,
            residency_type=residency_type,
            rural_urban_eligible=bool(payload.get("rural_urban_eligible", False)),
            disability_levels=disability_levels,
            dependents=[Dependent.from_dict(item) for item in payload.get("dependents", [])],
            work_loss_days=int(payload.get("work_loss_days", 0) or 0),
            lost_income_actual=parse_decimal_or_none("lost_income_actual"),
            annual_income_average=parse_decimal_or_none("annual_income_average"),
            industry_average_annual_income=parse_decimal_or_none("industry_average_annual_income"),
            nursing_days=int(payload.get("nursing_days", 0) or 0),
            nursing_rate_per_day=parse_decimal_or_none("nursing_rate_per_day"),
            nursing_annual_income=parse_decimal_or_none("nursing_annual_income"),
            nutrition_days=int(payload.get("nutrition_days", 0) or 0),
            nutrition_rate_per_day=decimalize(payload.get("nutrition_rate_per_day", "50")),
            hospital_days=int(payload.get("hospital_days", 0) or 0),
            hospital_food_rate_per_day=decimalize(payload.get("hospital_food_rate_per_day", "100")),
            medical_expense=decimalize(payload.get("medical_expense", 0)),
            transport_expense=decimalize(payload.get("transport_expense", 0)),
            lodging_expense=decimalize(payload.get("lodging_expense", 0)),
            appraisal_fee=decimalize(payload.get("appraisal_fee", 0)),
            property_loss=decimalize(payload.get("property_loss", 0)),
            assistive_device_expense=decimalize(payload.get("assistive_device_expense", 0)),
            mental_damage=parse_decimal_or_none("mental_damage"),
            auto_mental_damage=bool(payload.get("auto_mental_damage", False)),
            statistics_overrides=statistics_overrides,
        )

    @property
    def base_year(self) -> int:
        return self.hearing_date.year - 1

    @property
    def normalized_province(self) -> str:
        return normalize_region_name(self.court_province)

    @property
    def normalized_city(self) -> str:
        return normalize_region_name(self.court_city)


@dataclass
class CalculationResult:
    standards: list[str]
    lines: list[CalculationLine]
    law_sections: dict[str, str]

    @property
    def total(self) -> Decimal:
        return sum((line.amount for line in self.lines), start=Decimal("0"))

    def to_dict(self) -> dict[str, Any]:
        return {
            "standards": self.standards,
            "lines": [
                {
                    "category": line.category,
                    "item": line.item,
                    "formula": line.formula,
                    "amount": float(to_money(line.amount)),
                    "note": line.note,
                    "legal_basis": line.legal_basis,
                }
                for line in self.lines
            ],
            "law_sections": self.law_sections,
            "total": float(to_money(self.total)),
            "disclaimer": DISCLAIMER,
        }


def is_special_city(case_data: CaseData) -> bool:
    return case_data.normalized_city in SPECIAL_CITIES


def value_from_override(raw_value: Any, year: int) -> Decimal | None:
    if raw_value in (None, ""):
        return None
    if isinstance(raw_value, dict):
        for key in (str(year), year):
            if key in raw_value:
                return decimalize(raw_value[key])
        return None
    return decimalize(raw_value)


def resolve_stat_value(
    case_data: CaseData,
    *,
    override_key: str,
    reference_value: Decimal | None = None,
    require_override_if_special_city: bool = False,
    error_label: str,
) -> tuple[Decimal, str]:
    override_value = value_from_override(case_data.statistics_overrides.get(override_key), case_data.base_year)
    if override_value is not None:
        return override_value, "用户提供覆盖数据"
    if require_override_if_special_city and is_special_city(case_data):
        raise ReferenceDataError(
            f"{case_data.court_city}{case_data.base_year}年度{error_label}属于经济特区/计划单列市口径，"
            "本地 references 未提供该市数据，请通过 statistics_overrides 提供。"
        )
    if reference_value is None:
        raise ReferenceDataError(
            f"{case_data.court_province}{case_data.base_year}年度{error_label}在本地 references 中缺失。"
            "请通过 statistics_overrides 提供，或先联网查询权威统计数据。"
        )
    return reference_value, "本地 references"


def compensation_years(age: int) -> int:
    if age < 60:
        return 20
    if age <= 75:
        return max(1, 20 - (age - 60))
    return 5


def dependent_years(dependent: Dependent) -> int:
    if dependent.age < 18:
        return max(0, 18 - dependent.age)
    if not dependent.no_labor_capacity:
        raise CalculationError(
            f"成年被扶养人 {dependent.age} 岁未标记 no_labor_capacity=true，无法按规则计算被扶养人生活费。"
        )
    return compensation_years(dependent.age)


def total_disability_ratio(levels: list[int]) -> Decimal:
    if not levels:
        raise CalculationError("计算残疾赔偿金时必须提供 disability_levels。")
    ratios = sorted((DISABILITY_LEVEL_TO_RATIO[level] for level in levels), reverse=True)
    main_ratio = ratios[0]
    additional_ratio = min(sum(ratios[1:], start=Decimal("0")) * Decimal("0.1"), Decimal("0.1"))
    return min(main_ratio + additional_ratio, Decimal("1.0"))


def infer_mental_damage(levels: list[int]) -> Decimal | None:
    if not levels:
        return None
    best_level = min(levels)
    return MENTAL_DAMAGE_REFERENCE[best_level]


def law_excerpt_map(reference_store: ReferenceStore, headings: Iterable[str]) -> dict[str, str]:
    results: dict[str, str] = {}
    for heading in headings:
        if heading in reference_store.law_articles:
            results[heading] = reference_store.law_articles[heading]
    return results


def province_wage(case_data: CaseData, reference_store: ReferenceStore) -> tuple[Decimal, str]:
    province = case_data.normalized_province
    if not province:
        raise CalculationError("court_province 不能为空。")

    table = reference_store.wage_tables.get(case_data.wage_caliber, {})
    province_values = table.get(province, {})
    reference_value = province_values.get(case_data.base_year)
    override_key = f"{case_data.wage_caliber}_wage"
    return resolve_stat_value(
        case_data,
        override_key=override_key,
        reference_value=reference_value,
        require_override_if_special_city=True,
        error_label=f"{display_wage_caliber(case_data.wage_caliber)}平均工资",
    )


def display_wage_caliber(wage_caliber: str) -> str:
    return {
        "private": "城镇私营单位",
        "full": "全口径城镇单位",
        "non_private": "城镇非私营单位",
    }[wage_caliber]


def urban_disposable_income(case_data: CaseData, reference_store: ReferenceStore) -> tuple[Decimal, str]:
    reference_value = None
    if not case_data.normalized_province:
        reference_value = reference_store.national_disposable_income.get(case_data.base_year)
    return resolve_stat_value(
        case_data,
        override_key="urban_disposable_income",
        reference_value=reference_value,
        require_override_if_special_city=True,
        error_label="城镇居民人均可支配收入",
    )


def consumption_expenditure(case_data: CaseData, *, rural: bool) -> tuple[Decimal, str]:
    override_key = "rural_consumption_expenditure" if rural else "urban_consumption_expenditure"
    label = "农村居民人均年生活消费支出" if rural else "城镇居民人均消费支出"
    value = value_from_override(case_data.statistics_overrides.get(override_key), case_data.base_year)
    if value is None:
        raise ReferenceDataError(
            f"{case_data.court_province}{case_data.base_year}年度{label}在本地 references 中缺失。"
            f"请通过 statistics_overrides.{override_key} 提供，或先联网查询权威统计数据。"
        )
    return value, "用户提供覆盖数据"


def rural_net_income(case_data: CaseData) -> tuple[Decimal, str]:
    value = value_from_override(case_data.statistics_overrides.get("rural_net_income"), case_data.base_year)
    if value is None:
        raise ReferenceDataError(
            f"{case_data.court_province}{case_data.base_year}年度农村居民人均纯收入在本地 references 中缺失。"
            "请通过 statistics_overrides.rural_net_income 提供，或先联网查询权威统计数据。"
        )
    return value, "用户提供覆盖数据"


def is_urban_standard(case_data: CaseData) -> bool:
    if case_data.incident_date >= INJURY_INTERPRETATION_REVISION_DATE:
        return True
    if case_data.residency_type == "urban":
        return True
    return case_data.rural_urban_eligible


def lost_income_line(case_data: CaseData) -> CalculationLine | None:
    if case_data.work_loss_days <= 0:
        return None
    days = Decimal(case_data.work_loss_days)
    if case_data.lost_income_actual is not None:
        return CalculationLine(
            category="人身损害",
            item="误工费",
            formula=f"实际减少收入 {format_money(case_data.lost_income_actual)} 元",
            amount=case_data.lost_income_actual,
            note="按固定收入实际减少额计算",
            legal_basis=LAW_BASIS_BY_ITEM["误工费"],
        )
    if case_data.annual_income_average is not None:
        daily = case_data.annual_income_average / Decimal("365")
        amount = daily * days
        return CalculationLine(
            category="人身损害",
            item="误工费",
            formula=f"{format_money(case_data.annual_income_average)} 元/年 ÷ 365 × {case_data.work_loss_days} 天",
            amount=amount,
            note="按最近三年平均年收入折算",
            legal_basis=LAW_BASIS_BY_ITEM["误工费"],
        )
    if case_data.industry_average_annual_income is not None:
        daily = case_data.industry_average_annual_income / Decimal("365")
        amount = daily * days
        return CalculationLine(
            category="人身损害",
            item="误工费",
            formula=f"{format_money(case_data.industry_average_annual_income)} 元/年 ÷ 365 × {case_data.work_loss_days} 天",
            amount=amount,
            note="按同行业平均工资参照计算",
            legal_basis=LAW_BASIS_BY_ITEM["误工费"],
        )
    raise CalculationError(
        "提供了 work_loss_days，但未提供 lost_income_actual、annual_income_average 或 industry_average_annual_income。"
    )


def nursing_line(case_data: CaseData) -> CalculationLine | None:
    if case_data.nursing_days <= 0:
        return None
    if case_data.nursing_rate_per_day is not None:
        amount = case_data.nursing_rate_per_day * Decimal(case_data.nursing_days)
        return CalculationLine(
            category="人身损害",
            item="护理费",
            formula=f"{format_money(case_data.nursing_rate_per_day)} 元/天 × {case_data.nursing_days} 天",
            amount=amount,
            note="按明确日护理费计算",
            legal_basis=LAW_BASIS_BY_ITEM["护理费"],
        )
    if case_data.nursing_annual_income is not None:
        daily = case_data.nursing_annual_income / Decimal("365")
        amount = daily * Decimal(case_data.nursing_days)
        return CalculationLine(
            category="人身损害",
            item="护理费",
            formula=f"{format_money(case_data.nursing_annual_income)} 元/年 ÷ 365 × {case_data.nursing_days} 天",
            amount=amount,
            note="按护理人员收入折算",
            legal_basis=LAW_BASIS_BY_ITEM["护理费"],
        )
    raise CalculationError("提供了 nursing_days，但未提供 nursing_rate_per_day 或 nursing_annual_income。")


def dependent_lines(case_data: CaseData, annual_base: Decimal, *, rural: bool) -> list[CalculationLine]:
    if not case_data.dependents:
        return []

    annual_shares = [annual_base / Decimal(dep.supporter_count) for dep in case_data.dependents]
    total_annual_share = sum(annual_shares, start=Decimal("0"))
    cap_factor = Decimal("1")
    if total_annual_share > annual_base:
        cap_factor = annual_base / total_annual_share

    lines: list[CalculationLine] = []
    for dependent, annual_share in zip(case_data.dependents, annual_shares):
        years = dependent_years(dependent)
        adjusted_annual_share = annual_share * cap_factor
        amount = adjusted_annual_share * Decimal(years)
        note = "按法定上限同比例压缩" if cap_factor < 1 else ""
        label_suffix = "（并入残疾/死亡赔偿总额）" if case_data.incident_date >= INJURY_INTERPRETATION_REVISION_DATE else ""
        lines.append(
            CalculationLine(
                category="人身损害",
                item=f"被扶养人生活费{label_suffix}",
                formula=(
                    f"{format_money(annual_base)} 元/年 × {years} 年 ÷ {dependent.supporter_count}"
                    f"{' × ' + str(cap_factor.quantize(Decimal('0.0001'))) if cap_factor < 1 else ''}"
                ),
                amount=amount,
                note=f"{dependent.age}岁被扶养人；{'农村' if rural else '城镇'}消费支出口径。{note}".strip("。"),
                legal_basis=LAW_BASIS_BY_ITEM["被扶养人生活费"],
            )
        )
    return lines


def calculate_case(case_data: CaseData, reference_store: ReferenceStore) -> CalculationResult:
    if not case_data.court_province:
        raise CalculationError("court_province 不能为空。")

    standards: list[str] = [
        f"案件类型：{'人身伤害' if case_data.case_type == 'injury' else '死亡赔偿'}",
        f"案件发生时间：{case_data.incident_date.isoformat()}",
        f"一审法庭辩论终结时间：{case_data.hearing_date.isoformat()}，取上一统计年度 {case_data.base_year} 年",
        f"受诉法院所在地：{case_data.court_province}{(' / ' + case_data.court_city) if case_data.court_city else ''}",
        f"工资口径：{display_wage_caliber(case_data.wage_caliber)}",
    ]

    lines: list[CalculationLine] = []

    loss_items = [
        ("医疗费", case_data.medical_expense),
        ("交通费", case_data.transport_expense),
        ("住宿费", case_data.lodging_expense),
        ("鉴定费", case_data.appraisal_fee),
        ("财产损失", case_data.property_loss),
        ("残疾辅助器具费", case_data.assistive_device_expense),
    ]
    for item_name, amount in loss_items:
        if amount > 0:
            lines.append(
                CalculationLine(
                    category="财产损失" if item_name == "财产损失" else "人身损害",
                    item=item_name,
                    formula="按票据/实际发生额主张",
                    amount=amount,
                    legal_basis=LAW_BASIS_BY_ITEM.get(item_name, []),
                )
            )

    if case_data.hospital_days > 0:
        amount = case_data.hospital_food_rate_per_day * Decimal(case_data.hospital_days)
        lines.append(
            CalculationLine(
                category="人身损害",
                item="住院伙食补助费",
                formula=f"{format_money(case_data.hospital_food_rate_per_day)} 元/天 × {case_data.hospital_days} 天",
                amount=amount,
                legal_basis=LAW_BASIS_BY_ITEM["住院伙食补助费"],
            )
        )

    if case_data.nutrition_days > 0:
        amount = case_data.nutrition_rate_per_day * Decimal(case_data.nutrition_days)
        lines.append(
            CalculationLine(
                category="人身损害",
                item="营养费",
                formula=f"{format_money(case_data.nutrition_rate_per_day)} 元/天 × {case_data.nutrition_days} 天",
                amount=amount,
                legal_basis=LAW_BASIS_BY_ITEM["营养费"],
            )
        )

    work_loss = lost_income_line(case_data)
    if work_loss is not None:
        lines.append(work_loss)

    nursing = nursing_line(case_data)
    if nursing is not None:
        lines.append(nursing)

    urban_standard = is_urban_standard(case_data)
    if case_data.case_type == "injury" and case_data.disability_levels:
        income_base, income_source = (
            urban_disposable_income(case_data, reference_store) if urban_standard else rural_net_income(case_data)
        )
        ratio = total_disability_ratio(case_data.disability_levels)
        years = compensation_years(case_data.victim_age)
        amount = income_base * Decimal(years) * ratio
        label = "城镇居民人均可支配收入" if urban_standard else "农村居民人均纯收入"
        lines.append(
            CalculationLine(
                category="人身损害",
                item="残疾赔偿金",
                formula=f"{format_money(income_base)} 元/年 × {years} 年 × {ratio * Decimal('100')}%",
                amount=amount,
                note=f"{label}；来源：{income_source}",
                legal_basis=LAW_BASIS_BY_ITEM["残疾赔偿金"],
            )
        )
        standards.append(
            f"残疾赔偿金标准：{'统一城镇标准' if urban_standard else '2022-05-01 前农村标准'}；数据来源：{income_source}"
        )

    if case_data.case_type == "death":
        income_base, income_source = (
            urban_disposable_income(case_data, reference_store) if urban_standard else rural_net_income(case_data)
        )
        years = compensation_years(case_data.victim_age)
        amount = income_base * Decimal(years)
        label = "城镇居民人均可支配收入" if urban_standard else "农村居民人均纯收入"
        lines.append(
            CalculationLine(
                category="人身损害",
                item="死亡赔偿金",
                formula=f"{format_money(income_base)} 元/年 × {years} 年",
                amount=amount,
                note=f"{label}；来源：{income_source}",
                legal_basis=LAW_BASIS_BY_ITEM["死亡赔偿金"],
            )
        )
        standards.append(
            f"死亡赔偿金标准：{'统一城镇标准' if urban_standard else '2022-05-01 前农村标准'}；数据来源：{income_source}"
        )

    if case_data.case_type == "death":
        wage_value, wage_source = province_wage(case_data, reference_store)
        monthly = wage_value if case_data.wage_caliber == "full" else wage_value / Decimal("12")
        amount = monthly * Decimal("6")
        lines.append(
            CalculationLine(
                category="人身损害",
                item="丧葬费",
                formula=f"{format_money(monthly)} 元/月 × 6 个月",
                amount=amount,
                note=f"{display_wage_caliber(case_data.wage_caliber)}；来源：{wage_source}",
                legal_basis=LAW_BASIS_BY_ITEM["丧葬费"],
            )
        )
        standards.append(f"丧葬费工资基数来源：{wage_source}")

    if case_data.dependents:
        consumption_base, consumption_source = consumption_expenditure(
            case_data,
            rural=not urban_standard and case_data.residency_type == "rural" and not case_data.rural_urban_eligible,
        )
        dependent_result_lines = dependent_lines(
            case_data,
            consumption_base,
            rural=not urban_standard and case_data.residency_type == "rural" and not case_data.rural_urban_eligible,
        )
        lines.extend(dependent_result_lines)
        standards.append(f"被扶养人生活费基数来源：{consumption_source}")

    mental_damage = case_data.mental_damage
    if mental_damage is None and case_data.auto_mental_damage:
        mental_damage = infer_mental_damage(case_data.disability_levels)
    if mental_damage is not None and mental_damage > 0:
        lines.append(
            CalculationLine(
                category="精神损害",
                item="精神抚慰金",
                formula=f"酌定金额 {format_money(mental_damage)} 元",
                amount=mental_damage,
                note="参考伤残等级区间，最终以法院裁量为准",
                legal_basis=LAW_BASIS_BY_ITEM["精神抚慰金"],
            )
        )

    relevant_headings = sorted({heading for line in lines for heading in line.legal_basis})
    relevant_headings.append("第二十二条 —— 赔偿标准的适用")
    law_sections = law_excerpt_map(reference_store, relevant_headings)

    return CalculationResult(standards=standards, lines=lines, law_sections=law_sections)


def render_markdown(result: CalculationResult) -> str:
    rows = [
        "| 赔偿项目 | 计算过程 | 金额（元） | 备注 |",
        "|---------|---------|-----------:|------|",
    ]
    for line in result.lines:
        rows.append(
            f"| {line.item} | {line.formula} | {format_money(line.amount)} | {line.note or '-'} |"
        )
    rows.append(f"| **合计** |  | **{format_money(result.total)}** |  |")

    law_blocks = []
    for heading, body in result.law_sections.items():
        law_blocks.append(f"### {heading}\n{body}")

    sections = [
        "# 人身损害赔偿计算结果",
        "## 计算标准说明",
        "\n".join(f"- {item}" for item in result.standards),
        "## 计算明细表",
        "\n".join(rows),
        "## 法条依据",
        "\n\n".join(law_blocks) if law_blocks else "- 本次结果未匹配到内置法条摘录。",
        "## 最终赔偿金额",
        f"应赔偿总额：**{format_money(result.total)} 元**",
        "## 免责声明",
        f"> {DISCLAIMER}",
    ]
    return "\n\n".join(sections) + "\n"


def export_xlsx(result: CalculationResult, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise CalculationError("导出 xlsx 需要安装 openpyxl。") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "赔偿明细"
    ws.freeze_panes = "A2"

    headers = ["序号", "赔偿大类", "赔偿项目", "计算基数与过程", "计算金额", "备注/法律依据"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for index, line in enumerate(result.lines, start=1):
        ws.append(
            [
                index,
                line.category,
                line.item,
                line.formula,
                float(to_money(line.amount)),
                f"{line.note or ''} {'；'.join(line.legal_basis)}".strip(),
            ]
        )

    total_row = ws.max_row + 1
    ws.append(["", "", "合计", "", f"=SUM(E2:E{total_row - 1})", ""])
    for cell in ws[total_row]:
        cell.font = Font(bold=True)

    for row_index in range(2, ws.max_row + 1):
        ws[f"E{row_index}"].number_format = "#,##0.00"

    widths = {1: 8, 2: 14, 3: 18, 4: 40, 5: 14, 6: 48}
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def load_case_payload(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise CalculationError(f"输入文件不是合法 JSON：{exc}") from exc


def example_payload() -> dict[str, Any]:
    return {
        "case_type": "injury",
        "incident_date": "2024-06-12",
        "hearing_date": "2025-03-18",
        "victim_age": 35,
        "court_province": "广东省",
        "court_city": "广州市",
        "wage_caliber": "private",
        "residency_type": "urban",
        "disability_levels": [10],
        "work_loss_days": 90,
        "annual_income_average": 120000,
        "nursing_days": 30,
        "nursing_rate_per_day": 150,
        "nutrition_days": 60,
        "hospital_days": 10,
        "medical_expense": 18000,
        "transport_expense": 1200,
        "appraisal_fee": 2500,
        "auto_mental_damage": True,
        "dependents": [
            {
                "age": 10,
                "supporter_count": 2
            }
        ],
        "statistics_overrides": {
            "urban_disposable_income": {
                "2024": 65000
            },
            "urban_consumption_expenditure": {
                "2024": 42000
            }
        }
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="按 personal-injury-compensation-calculation 技能规则计算人身损害赔偿。"
    )
    parser.add_argument("--input", type=Path, help="案件输入 JSON 文件路径")
    parser.add_argument(
        "--format",
        choices=("markdown", "json"),
        default="markdown",
        help="输出格式，默认 markdown",
    )
    parser.add_argument("--output", type=Path, help="将结果写入文件；不传则打印到 stdout")
    parser.add_argument("--xlsx", type=Path, help="额外导出 xlsx 明细文件")
    parser.add_argument("--example", action="store_true", help="打印示例输入 JSON")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.example:
        print(json.dumps(example_payload(), ensure_ascii=False, indent=2))
        return 0

    if not args.input:
        parser.error("未提供 --input。")

    script_dir = Path(__file__).resolve().parent
    references_dir = script_dir.parent / "references"

    try:
        payload = load_case_payload(args.input)
        case_data = CaseData.from_dict(payload)
        reference_store = ReferenceStore.load(references_dir)
        result = calculate_case(case_data, reference_store)
    except (CalculationError, ReferenceDataError, ValueError, TypeError) as exc:
        print(f"计算失败：{exc}", file=sys.stderr)
        return 1

    rendered = (
        json.dumps(result.to_dict(), ensure_ascii=False, indent=2)
        if args.format == "json"
        else render_markdown(result)
    )

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered, encoding="utf-8")
    else:
        print(rendered)

    if args.xlsx:
        try:
            export_xlsx(result, args.xlsx)
        except CalculationError as exc:
            print(f"xlsx 导出失败：{exc}", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
