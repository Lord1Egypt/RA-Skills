#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fetch XingTu task author list via API and export to formatted Excel.
Usage:
    # Full fetch
    python fetch_xingtu_authors.py --task-id 7642279680695484426 --cookie "xxx" [--output path.xlsx]
    # Validate cookie only (makes minimal API call)
    python fetch_xingtu_authors.py --task-id 7642279680695484426 --cookie "xxx" --validate
"""

import argparse
import json
import os
import sys
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

API_URL = "https://www.xingtu.cn/gw/api/challenge/provider_get_task_author_list"

HEADERS_TEMPLATE = {
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "agw-js-conv": "str",
    "User-Agent": "Apifox/1.0.0 (https://apifox.com)",
    "Host": "www.xingtu.cn",
    "Accept-Charset": "UTF-8",
    "Accept-Encoding": "gzip, deflate",
}


def _build_headers(cookie):
    """Build request headers with cookie."""
    headers = HEADERS_TEMPLATE.copy()
    headers["Cookie"] = cookie
    return headers


def _fetch_page(cookie, task_id, page, limit=50):
    """Fetch a single page of author list. Returns parsed JSON dict."""
    headers = _build_headers(cookie)
    body = {
        "recruit_search_param": {},
        "task_id": str(task_id),
        "order_by": "",
        "page": page,
        "limit": limit,
    }
    resp = requests.post(API_URL, headers=headers, json=body, timeout=30)
    resp.encoding = "utf-8"
    return resp.json()


def validate_cookie(cookie, task_id):
    """
    Validate cookie by making a minimal API call (page 1, limit 1).
    Returns (valid: bool, message: str).
    """
    try:
        data = _fetch_page(cookie, task_id, 1, limit=1)
        base_resp = data.get("base_resp", {})
        if base_resp.get("status_code") == 0:
            pagination = data.get("pagination", {})
            total = pagination.get("total_count", 0)
            return True, f"Cookie valid, task has {total} author(s)"
        else:
            msg = base_resp.get("status_message", "unknown error")
            return False, f"API returned error: {msg}"
    except requests.exceptions.Timeout:
        return False, "Connection timed out - check network"
    except requests.exceptions.ConnectionError:
        return False, "Connection failed - check network/proxy"
    except Exception as e:
        return False, f"Cookie validation failed: {str(e)}"


def fetch_all_authors(cookie, task_id):
    """
    Fetch all authors across all pages.
    Returns (authors: list, total_count: int).
    Raises RuntimeError on API failure.
    """
    all_authors = []
    page = 1
    total_count = 0

    print("=" * 50)
    print(f"Task ID: {task_id}")
    print(f"API URL: {API_URL}")
    print("=" * 50)

    while True:
        print(f"\n>>> Fetching page {page} ...")
        data = _fetch_page(cookie, task_id, page, limit=50)

        base_resp = data.get("base_resp", {})
        if base_resp.get("status_code") != 0:
            raise RuntimeError(
                f"API error on page {page}: {base_resp.get('status_message', 'unknown')}"
            )

        author_list = data.get("author_list", [])
        pagination = data.get("pagination", {})
        total_count = pagination.get("total_count", len(author_list))
        has_more = pagination.get("has_more", False)

        batch_size = len(author_list)
        all_authors.extend(author_list)
        print(f"    Got {batch_size} authors, cumulative: {len(all_authors)}/{total_count}")

        if not has_more:
            break

        page += 1

    print(f"\nAll pages fetched. Total authors: {len(all_authors)}")
    return all_authors, total_count


def export_to_excel(authors, output_path, task_id):
    """Export author data to a well-formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "星图达人列表"

    # ── Headers ──
    headers = [
        "序号", "达人昵称", "作者ID", "达人等级", "粉丝数",
        "主推类目(30天)", "带货GMV(30天)", "视频GMV(30天)",
        "1-20s报价", "21-60s报价", "60s+报价",
        "预期CPM", "预期播放量", "完播率",
        "所在城市", "内容标签",
        "微信号", "报名状态",
        "报名时间", "推荐理由",
        "合作优势", "创作思路", "补充说明",
    ]

    # ── Styles ──
    header_font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Write header row ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Write data rows ──
    enroll_status_map = {
        11: "已报名",
        10: "待审核",
        20: "已通过",
        30: "已拒绝",
    }

    for idx, author in enumerate(authors, 1):
        row = idx + 1
        base = author.get("author_base_info", {})
        order = author.get("author_order_info", {})
        recruit = order.get("recruit_author_order_info", {})

        # Helper: join list safely
        categories = "、".join(base.get("all_ecom_top3_category_30d_desc", []))
        content_tags = "、".join(base.get("content_tags", []))

        # Recommend reasons
        reasons = recruit.get("recommend_reason", [])
        reason_lines = [
            "[{}] {}".format(r.get("reason_label", ""), r.get("reason_descrip", ""))
            for r in reasons
        ]

        # Enrollment status & time
        enroll_status = recruit.get("enroll_status", "")
        status_text = enroll_status_map.get(enroll_status, str(enroll_status))

        enroll_ts = recruit.get("time_info", {}).get("enroll_time", "")
        enroll_time = ""
        if enroll_ts and enroll_ts != "0":
            try:
                enroll_time = datetime.fromtimestamp(int(enroll_ts)).strftime("%Y-%m-%d %H:%M")
            except (ValueError, OSError):
                enroll_time = enroll_ts

        row_data = [
            idx,
            base.get("nick_name", ""),
            base.get("author_id", ""),
            base.get("ecom_author_level", ""),
            base.get("follower", ""),
            categories,
            base.get("all_ecom_gmv_30d_desc", ""),
            base.get("ecom_video_gmv_30d_desc", ""),
            recruit.get("recruit_cpt_info", {}).get("author_price", ""),
            base.get("price_21_60", ""),
            base.get("price_60", ""),
            base.get("prospective_cpm", ""),
            base.get("expected_play_num", ""),
            base.get("author_recruit_video_cpt_fulfillment_rate_desc", ""),
            base.get("author_resident_city", ""),
            content_tags,
            base.get("wechat", ""),
            status_text,
            enroll_time,
            "\n".join(reason_lines),
            recruit.get("coop_advantage", ""),
            recruit.get("creation_idea", ""),
            recruit.get("extra_note", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border

    # ── Column widths ──
    col_widths = {
        1: 6, 2: 16, 3: 22, 4: 10, 5: 10,
        6: 28, 7: 16, 8: 16,
        9: 12, 10: 12, 11: 14,
        12: 12, 13: 12, 14: 10,
        15: 12, 16: 18,
        17: 18, 18: 10,
        19: 18, 20: 42,
        21: 15, 22: 15, 23: 20,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze + Filter ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(
        get_column_letter(len(headers)), len(authors) + 1
    )

    # ── Row height ──
    ws.row_dimensions[1].height = 28
    for r in range(2, len(authors) + 2):
        ws.row_dimensions[r].height = 22

    # ── Save ──
    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")


# ────────────────────────────────────────
# Main
# ────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Fetch XingTu task author list")
    parser.add_argument("--task-id", required=True, help="Star chart task ID")
    parser.add_argument("--cookie", required=True, help="Cookie string for authentication")
    parser.add_argument("--output", default=None, help="Output Excel file path")
    parser.add_argument("--validate", action="store_true", help="Only validate cookie, do not fetch")

    args = parser.parse_args()

    # --- Validate-only mode ---
    if args.validate:
        valid, msg = validate_cookie(args.cookie, args.task_id)
        if valid:
            print(f"[OK] {msg}")
            sys.exit(0)
        else:
            print(f"[FAIL] {msg}", file=sys.stderr)
            sys.exit(1)

    # --- Determine output path ---
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.getcwd(), f"xingtu_authors_{args.task_id}_{timestamp}.xlsx"
        )

    # --- Fetch ---
    authors, total = fetch_all_authors(args.cookie, args.task_id)

    if not authors:
        print("[WARN] No authors returned from API.", file=sys.stderr)
        sys.exit(1)

    # --- Export ---
    export_to_excel(authors, output_path, args.task_id)

    summary = {
        "task_id": args.task_id,
        "total_authors": len(authors),
        "output": output_path,
    }
    print("\n[SUMMARY] " + json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
