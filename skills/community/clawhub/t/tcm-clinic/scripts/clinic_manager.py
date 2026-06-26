#!/usr/bin/env python3
"""中医诊所管理系统 - 统一命令行入口

使用方式:
    python clinic_manager.py <module> <action> [--data-dir DIR] [options...]

模块:
    patients     患者档案管理
    records      病历记录管理
    herbs        中药库存管理
    appointments 预约排班管理
    finance      财务收费管理
    init         初始化所有数据表

示例:
    python clinic_manager.py patients add --name "张三" --gender "男" --phone "13800138000"
    python clinic_manager.py patients search --name "张"
    python clinic_manager.py records add --patient-id "P20260402001" --complaint "头痛三日"
    python clinic_manager.py herbs alerts
    python clinic_manager.py finance summary --period month --month 2026-04
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 需要安装 openpyxl 库。请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_DATA_DIR = "clinic_data"

# ─── 表头定义 ───────────────────────────────────────────────

SCHEMAS = {
    "patients": {
        "filename": "patients.xlsx",
        "sheet": "patients",
        "headers": [
            "patient_id", "name", "gender", "birth_date", "age",
            "phone", "address", "constitution_type", "allergies",
            "chronic_diseases", "notes", "created_date", "last_visit_date"
        ],
    },
    "records": {
        "filename": "medical_records.xlsx",
        "sheet": "records",
        "headers": [
            "record_id", "patient_id", "patient_name", "visit_date",
            "chief_complaint", "tongue_condition", "pulse_condition",
            "observation", "listening_smelling", "inquiry",
            "diagnosis", "prescription", "advice", "visit_count", "notes"
        ],
    },
    "herbs": {
        "filename": "herbs_inventory.xlsx",
        "sheet": "herbs",
        "headers": [
            "herb_id", "name", "pinyin", "specification", "stock_quantity",
            "unit", "purchase_price", "retail_price", "supplier",
            "expiry_date", "entry_date", "minimum_stock", "category", "notes"
        ],
    },
    "appointments": {
        "filename": "appointments.xlsx",
        "sheet": "appointments",
        "headers": [
            "appointment_id", "patient_id", "patient_name",
            "appointment_date", "time_slot", "status",
            "purpose", "queue_number", "notes"
        ],
    },
    "finances": {
        "filename": "finances.xlsx",
        "sheet": "finances",
        "headers": [
            "finance_id", "record_id", "patient_id", "patient_name",
            "date", "type", "amount", "payment_method", "notes"
        ],
    },
}


# ─── 工具函数 ───────────────────────────────────────────────

def get_data_path(data_dir: str, module: str) -> str:
    schema = SCHEMAS[module]
    return os.path.join(data_dir, schema["filename"])


def ensure_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def init_workbook(filepath: str, sheet_name: str, headers: list):
    if os.path.exists(filepath):
        return
    ensure_dir(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(filepath)


def load_workbook_safe(filepath: str, sheet_name: str, headers: list):
    init_workbook(filepath, sheet_name, headers)
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    return wb, ws


def read_all_rows(data_dir: str, module: str) -> list:
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: (v if v is not None else "") for h, v in zip(schema["headers"], row)}
        rows.append(d)
    wb.close()
    return rows


def append_row(data_dir: str, module: str, row_data: dict):
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    row_values = [row_data.get(h, "") for h in schema["headers"]]
    ws.append(row_values)
    wb.save(filepath)
    wb.close()


def find_patient_by_name(data_dir: str, name: str) -> list:
    rows = read_all_rows(data_dir, "patients")
    return [r for r in rows if name in (r.get("name") or "")]


# 模块名 → 对应 ID 字段名
_ID_FIELDS = {
    "patients": "patient_id",
    "records": "record_id",
    "herbs": "herb_id",
    "appointments": "appointment_id",
    "finances": "finance_id",
}


def generate_id(prefix: str, data_dir: str = None, module: str = None) -> str:
    """生成递增ID：前缀 + YYYYMMDD + 3位序号（从现有数据中取最大序号+1）"""
    today = datetime.now().strftime("%Y%m%d")
    next_seq = 1
    if data_dir and module:
        rows = read_all_rows(data_dir, module)
        today_prefix = f"{prefix}{today}"
        id_field = _ID_FIELDS.get(module)
        for row in rows:
            rid = (row.get(id_field) or "") if id_field else ""
            if rid and rid.startswith(today_prefix):
                seq_str = rid[len(today_prefix):]
                try:
                    seq = int(seq_str)
                    if seq >= next_seq:
                        next_seq = seq + 1
                except ValueError:
                    pass
    return f"{prefix}{today}{next_seq:03d}"


def get_today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# ─── 患者模块 ───────────────────────────────────────────────

def cmd_patients_add(args, data_dir):
    name = args.name
    if not name:
        print("错误: --name 参数为必填项", file=sys.stderr)
        return 1

    existing = find_patient_by_name(data_dir, name)
    if existing:
        print(f"提示: 已存在同名患者 '{name}'（ID: {existing[0]['patient_id']}），如需新增请确认。")

    row = {
        "patient_id": generate_id("P", data_dir, "patients"),
        "name": name,
        "gender": args.gender or "",
        "birth_date": args.birth_date or "",
        "age": args.age or 0,
        "phone": args.phone or "",
        "address": args.address or "",
        "constitution_type": args.constitution or "",
        "allergies": args.allergies or "",
        "chronic_diseases": args.chronic_diseases or "",
        "notes": args.notes or "",
        "created_date": get_today(),
        "last_visit_date": "",
    }
    append_row(data_dir, "patients", row)
    print(f"已登记患者: {row['patient_id']} - {name}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_patients_search(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "")]
    if args.phone:
        rows = [r for r in rows if args.phone in (r.get("phone") or "")]
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]

    if not rows:
        print("未找到匹配的患者记录。")
        return 0

    print(f"共找到 {len(rows)} 条记录:")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | 电话: {r.get('phone','')} | 体质: {r.get('constitution_type','')} | 建档: {r.get('created_date','')}")
    return 0


def cmd_patients_list(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if not rows:
        print("暂无患者记录。")
        return 0
    print(f"共 {len(rows)} 位患者:")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | 电话: {r.get('phone','')} | 体质: {r.get('constitution_type','')}")
    return 0


# ─── 病历模块 ───────────────────────────────────────────────

def cmd_records_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"错误: 未找到患者 ID '{patient_id}'，请先登记患者。", file=sys.stderr)
        return 1

    complaint = args.complaint or ""
    diagnosis = args.diagnosis or ""
    if not complaint and not diagnosis:
        print("错误: --complaint（主诉）和 --diagnosis（诊断）至少需要填写一项。", file=sys.stderr)
        return 1

    count = len([r for r in read_all_rows(data_dir, "records") if r.get("patient_id") == patient_id])

    row = {
        "record_id": generate_id("R", data_dir, "records"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "visit_date": args.date or get_today(),
        "chief_complaint": complaint,
        "tongue_condition": args.tongue or "",
        "pulse_condition": args.pulse or "",
        "observation": args.observation or "",
        "listening_smelling": args.listening or "",
        "inquiry": args.inquiry or "",
        "diagnosis": diagnosis,
        "prescription": args.prescription or "",
        "advice": args.advice or "",
        "visit_count": count + 1,
        "notes": args.notes or "",
    }
    append_row(data_dir, "records", row)
    print(f"已添加病历: {row['record_id']}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_records_search(args, data_dir):
    rows = read_all_rows(data_dir, "records")
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]
    if args.date_from:
        rows = [r for r in rows if (r.get("visit_date") or "") >= args.date_from]
    if args.date_to:
        rows = [r for r in rows if (r.get("visit_date") or "") <= args.date_to]
    if args.diagnosis:
        rows = [r for r in rows if args.diagnosis in (r.get("diagnosis") or "")]

    if not rows:
        print("未找到匹配的病历记录。")
        return 0

    print(f"共找到 {len(rows)} 条病历记录:")
    for r in rows:
        print(f"  [{r['record_id']}] {r['visit_date']} | {r['patient_name']} | 主诉: {r.get('chief_complaint','')[:30]} | 诊断: {r.get('diagnosis','')}")
    return 0


# ─── 中药库存模块 ───────────────────────────────────────────

def cmd_herbs_add(args, data_dir):
    name = args.name
    if not name:
        print("错误: --name 参数为必填项", file=sys.stderr)
        return 1

    row = {
        "herb_id": generate_id("H", data_dir, "herbs"),
        "name": name,
        "pinyin": args.pinyin or "",
        "specification": args.spec or "",
        "stock_quantity": args.quantity or 0,
        "unit": args.unit or "g",
        "purchase_price": args.purchase_price or 0,
        "retail_price": args.retail_price or 0,
        "supplier": args.supplier or "",
        "expiry_date": args.expiry_date or "",
        "entry_date": get_today(),
        "minimum_stock": args.min_stock or 0,
        "category": args.category or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "herbs", row)
    print(f"已入库药材: {row['herb_id']} - {name} ({row['stock_quantity']}{row['unit']})")
    return 0


def cmd_herbs_update(args, data_dir):
    herb_id = args.herb_id
    if not herb_id:
        print("错误: --herb-id 参数为必填项", file=sys.stderr)
        return 1

    schema = SCHEMAS["herbs"]
    filepath = get_data_path(data_dir, "herbs")
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])

    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == herb_id:
            if args.quantity is not None:
                current = row[4].value or 0
                row[4].value = current + args.quantity
            if args.purchase_price is not None:
                row[6].value = args.purchase_price
            if args.retail_price is not None:
                row[7].value = args.retail_price
            if args.expiry_date:
                row[9].value = args.expiry_date
            if args.min_stock is not None:
                row[11].value = args.min_stock
            updated = True
            print(f"已更新药材 {herb_id}")
            break

    if not updated:
        print(f"错误: 未找到药材 ID '{herb_id}'", file=sys.stderr)
        wb.close()
        return 1

    wb.save(filepath)
    wb.close()
    return 0


def cmd_herbs_search(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "") or args.name in (r.get("pinyin") or "")]
    if args.category:
        rows = [r for r in rows if r.get("category") == args.category]

    if not rows:
        print("未找到匹配的药材记录。")
        return 0

    print(f"共找到 {len(rows)} 种药材:")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | 库存: {qty}{unit} | 分类: {r.get('category','')} | 有效期至: {r.get('expiry_date','')}")
    return 0


def cmd_herbs_alerts(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    today = datetime.now()

    # 保质期预警
    expiry_days = args.expiry_days or 30
    expiring = []
    for r in rows:
        exp_str = r.get("expiry_date", "")
        if exp_str:
            try:
                exp_date = datetime.strptime(exp_str, "%Y-%m-%d")
                diff = (exp_date - today).days
                if 0 <= diff <= expiry_days:
                    expiring.append((r, diff))
            except ValueError:
                pass

    # 库存预警
    low_stock = []
    for r in rows:
        qty = r.get("stock_quantity") or 0
        min_qty = r.get("minimum_stock") or 0
        if qty <= min_qty and min_qty > 0:
            low_stock.append(r)

    if expiring:
        print(f"⚠ 保质期预警（{expiry_days}天内过期，共 {len(expiring)} 项）:")
        for r, days in expiring:
            print(f"  {r['name']} | 剩余 {days} 天 | 库存: {r.get('stock_quantity',0)}{r.get('unit','g')}")

    if low_stock:
        print(f"\n⚠ 库存不足预警（共 {len(low_stock)} 项）:")
        for r in low_stock:
            qty = r.get("stock_quantity", 0) or 0
            min_q = r.get("minimum_stock", 0) or 0
            print(f"  {r['name']} | 当前: {qty}{r.get('unit','g')} | 最低: {min_q}{r.get('unit','g')}")

    if not expiring and not low_stock:
        print("所有药材状态正常，无需预警。")
    return 0


def cmd_herbs_list(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if not rows:
        print("暂无药材库存记录。")
        return 0
    print(f"共 {len(rows)} 种药材:")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | 库存: {qty}{unit} | 分类: {r.get('category','')} | 有效期: {r.get('expiry_date','')}")
    return 0


# ─── 预约模块 ───────────────────────────────────────────────

def cmd_appointments_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"错误: 未找到患者 ID '{patient_id}'", file=sys.stderr)
        return 1

    date = args.date or get_today()
    if not args.time_slot:
        print("错误: --time-slot 参数为必填项（如: 上午/下午/晚上）", file=sys.stderr)
        return 1

    row = {
        "appointment_id": generate_id("A", data_dir, "appointments"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "appointment_date": date,
        "time_slot": args.time_slot,
        "status": "待诊",
        "purpose": args.purpose or "",
        "queue_number": "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "appointments", row)
    print(f"已添加预约: {row['appointment_id']} | {date} {args.time_slot} | {patient['name']}")
    return 0


def cmd_appointments_search(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if args.date:
        rows = [r for r in rows if r.get("appointment_date") == args.date]
    if args.status:
        rows = [r for r in rows if r.get("status") == args.status]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]

    if not rows:
        print("未找到匹配的预约记录。")
        return 0

    print(f"共找到 {len(rows)} 条预约:")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | 状态: {r.get('status','')}")
    return 0


def cmd_appointments_today(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    today = get_today()
    today_rows = [r for r in rows if r.get("appointment_date") == today]

    if not today_rows:
        print(f"今日（{today}）暂无预约。")
        return 0

    pending = [r for r in today_rows if r.get("status") == "待诊"]
    completed = [r for r in today_rows if r.get("status") == "已诊"]
    cancelled = [r for r in today_rows if r.get("status") == "取消"]
    noshow = [r for r in today_rows if r.get("status") == "未到"]

    print(f"今日预约总览（{today}）:")
    print(f"  待诊: {len(pending)} | 已诊: {len(completed)} | 取消: {len(cancelled)} | 未到: {len(noshow)}")
    if pending:
        print("  ── 待诊队列 ──")
        for r in pending:
            print(f"    [{r['appointment_id']}] {r.get('time_slot','')} | {r['patient_name']} | {r.get('purpose','')}")
    return 0


def cmd_appointments_list(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if not rows:
        print("暂无预约记录。")
        return 0
    print(f"共 {len(rows)} 条预约记录:")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | {r.get('status','')}")
    return 0


# ─── 财务模块 ───────────────────────────────────────────────

def cmd_finance_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    patient_name = patient["name"] if patient else ""

    if not args.type:
        print("错误: --type 参数为必填项（挂号费/药费/针灸费/其他）", file=sys.stderr)
        return 1

    row = {
        "finance_id": generate_id("F", data_dir, "finances"),
        "record_id": args.record_id or "",
        "patient_id": patient_id,
        "patient_name": patient_name,
        "date": args.date or get_today(),
        "type": args.type,
        "amount": args.amount or 0,
        "payment_method": args.payment_method or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "finances", row)
    print(f"已添加收费记录: {row['finance_id']} | {args.type} | {row['amount']}元 | {patient_name}")
    return 0


def cmd_finance_summary(args, data_dir):
    rows = read_all_rows(data_dir, "finances")

    if args.period == "day":
        date = args.date or get_today()
        filtered = [r for r in rows if r.get("date") == date]
        title = f"日收入统计（{date}）"
    elif args.period == "month":
        month = args.month or datetime.now().strftime("%Y-%m")
        filtered = [r for r in rows if (r.get("date") or "")[:7] == month]
        title = f"月度收入统计（{month}）"
    elif args.period == "patient":
        if not args.patient_id:
            print("错误: --period patient 需要 --patient-id 参数", file=sys.stderr)
            return 1
        filtered = [r for r in rows if r.get("patient_id") == args.patient_id]
        title = f"患者费用汇总（{args.patient_id}）"
    else:
        filtered = rows
        title = "全部收入汇总"

    if not filtered:
        print(f"{title}: 无记录。")
        return 0

    total = 0
    type_totals = {}
    method_totals = {}
    for r in filtered:
        amt = r.get("amount") or 0
        total += amt
        t = r.get("type") or "未分类"
        type_totals[t] = type_totals.get(t, 0) + amt
        m = r.get("payment_method") or "未记录"
        method_totals[m] = method_totals.get(m, 0) + amt

    print(f"\n{title}")
    print(f"  总收入: {total:.2f} 元（{len(filtered)} 笔）")
    print(f"  ── 按费用类型 ──")
    for t, amt in sorted(type_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {t}: {amt:.2f} 元 ({pct:.1f}%)")
    print(f"  ── 按支付方式 ──")
    for m, amt in sorted(method_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {m}: {amt:.2f} 元 ({pct:.1f}%)")
    return 0


# ─── 初始化 ────────────────────────────────────────────────

def cmd_init(args, data_dir):
    os.makedirs(data_dir, exist_ok=True)
    for module, schema in SCHEMAS.items():
        filepath = os.path.join(data_dir, schema["filename"])
        init_workbook(filepath, schema["sheet"], schema["headers"])
        print(f"  ✓ {schema['filename']}")
    print(f"\n诊所数据目录 '{data_dir}' 初始化完成。")
    return 0


# ─── 参数解析 ───────────────────────────────────────────────

def build_parser():
    parser = argparse.ArgumentParser(
        description="中医诊所管理系统",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--data-dir", default=DEFAULT_DATA_DIR, help="数据目录路径（默认: clinic_data）")

    sub = parser.add_subparsers(dest="module")

    # patients
    p_pat = sub.add_parser("patients", help="患者档案管理")
    pat_sub = p_pat.add_subparsers(dest="action")
    pat_add = pat_sub.add_parser("add", help="登记新患者")
    pat_add.add_argument("--name", required=True, help="患者姓名")
    pat_add.add_argument("--gender", help="性别")
    pat_add.add_argument("--birth-date", help="出生日期 (YYYY-MM-DD)")
    pat_add.add_argument("--age", type=int, help="年龄")
    pat_add.add_argument("--phone", help="联系电话")
    pat_add.add_argument("--address", help="地址")
    pat_add.add_argument("--constitution", help="体质分型")
    pat_add.add_argument("--allergies", help="过敏史")
    pat_add.add_argument("--chronic-diseases", help="慢性病史")
    pat_add.add_argument("--notes", help="备注")

    pat_search = pat_sub.add_parser("search", help="搜索患者")
    pat_search.add_argument("--name", help="姓名（模糊匹配）")
    pat_search.add_argument("--phone", help="电话（模糊匹配）")
    pat_search.add_argument("--patient-id", help="患者 ID")

    pat_sub.add_parser("list", help="列出所有患者")

    # records
    p_rec = sub.add_parser("records", help="病历记录管理")
    rec_sub = p_rec.add_subparsers(dest="action")
    rec_add = rec_sub.add_parser("add", help="添加病历记录")
    rec_add.add_argument("--patient-id", required=True, help="患者 ID")
    rec_add.add_argument("--date", help="就诊日期 (YYYY-MM-DD)")
    rec_add.add_argument("--complaint", help="主诉")
    rec_add.add_argument("--tongue", help="舌诊")
    rec_add.add_argument("--pulse", help="脉诊")
    rec_add.add_argument("--observation", help="望诊")
    rec_add.add_argument("--listening", help="闻诊")
    rec_add.add_argument("--inquiry", help="问诊")
    rec_add.add_argument("--diagnosis", help="诊断")
    rec_add.add_argument("--prescription", help="处方")
    rec_add.add_argument("--advice", help="医嘱")
    rec_add.add_argument("--notes", help="备注")

    rec_search = rec_sub.add_parser("search", help="搜索病历")
    rec_search.add_argument("--patient-id", help="患者 ID")
    rec_search.add_argument("--patient-name", help="患者姓名（模糊匹配）")
    rec_search.add_argument("--date-from", help="起始日期")
    rec_search.add_argument("--date-to", help="截止日期")
    rec_search.add_argument("--diagnosis", help="诊断（模糊匹配）")

    # herbs
    p_herbs = sub.add_parser("herbs", help="中药库存管理")
    herbs_sub = p_herbs.add_subparsers(dest="action")
    herbs_add = herbs_sub.add_parser("add", help="药材入库")
    herbs_add.add_argument("--name", required=True, help="药材名称")
    herbs_add.add_argument("--pinyin", help="拼音")
    herbs_add.add_argument("--spec", help="规格")
    herbs_add.add_argument("--quantity", type=float, help="库存量")
    herbs_add.add_argument("--unit", help="单位")
    herbs_add.add_argument("--purchase-price", type=float, help="进货价")
    herbs_add.add_argument("--retail-price", type=float, help="零售价")
    herbs_add.add_argument("--supplier", help="供应商")
    herbs_add.add_argument("--expiry-date", help="保质期 (YYYY-MM-DD)")
    herbs_add.add_argument("--min-stock", type=float, help="最低库存")
    herbs_add.add_argument("--category", help="分类")
    herbs_add.add_argument("--notes", help="备注")

    herbs_update = herbs_sub.add_parser("update", help="更新药材信息")
    herbs_update.add_argument("--herb-id", required=True, help="药材 ID")
    herbs_update.add_argument("--quantity", type=float, help="库存变动量（正数入库，负数出库）")
    herbs_update.add_argument("--purchase-price", type=float, help="进货价")
    herbs_update.add_argument("--retail-price", type=float, help="零售价")
    herbs_update.add_argument("--expiry-date", help="保质期")
    herbs_update.add_argument("--min-stock", type=float, help="最低库存")

    herbs_search = herbs_sub.add_parser("search", help="搜索药材")
    herbs_search.add_argument("--name", help="药材名称/拼音（模糊匹配）")
    herbs_search.add_argument("--category", help="分类")

    herbs_alerts = herbs_sub.add_parser("alerts", help="库存与保质期预警")
    herbs_alerts.add_argument("--expiry-days", type=int, default=30, help="保质期预警天数（默认30）")

    herbs_sub.add_parser("list", help="列出所有药材")

    # appointments
    p_appt = sub.add_parser("appointments", help="预约排班管理")
    appt_sub = p_appt.add_subparsers(dest="action")
    appt_add = appt_sub.add_parser("add", help="添加预约")
    appt_add.add_argument("--patient-id", required=True, help="患者 ID")
    appt_add.add_argument("--date", help="预约日期 (YYYY-MM-DD)")
    appt_add.add_argument("--time-slot", required=True, help="时段")
    appt_add.add_argument("--purpose", help="就诊目的")
    appt_add.add_argument("--notes", help="备注")

    appt_search = appt_sub.add_parser("search", help="搜索预约")
    appt_search.add_argument("--date", help="日期")
    appt_search.add_argument("--status", help="状态")
    appt_search.add_argument("--patient-name", help="患者姓名")

    appt_sub.add_parser("today", help="今日预约队列")
    appt_sub.add_parser("list", help="列出所有预约")

    # finance
    p_fin = sub.add_parser("finance", help="财务收费管理")
    fin_sub = p_fin.add_subparsers(dest="action")
    fin_add = fin_sub.add_parser("add", help="添加收费记录")
    fin_add.add_argument("--patient-id", required=True, help="患者 ID")
    fin_add.add_argument("--record-id", help="关联病历 ID")
    fin_add.add_argument("--date", help="日期 (YYYY-MM-DD)")
    fin_add.add_argument("--type", required=True, help="费用类型")
    fin_add.add_argument("--amount", type=float, help="金额")
    fin_add.add_argument("--payment-method", help="支付方式")
    fin_add.add_argument("--notes", help="备注")

    fin_summary = fin_sub.add_parser("summary", help="财务统计")
    fin_summary.add_argument("--period", choices=["day", "month", "patient", "all"], default="day", help="统计周期")
    fin_summary.add_argument("--date", help="日期（日统计用）")
    fin_summary.add_argument("--month", help="月份（月统计用，如 2026-04）")
    fin_summary.add_argument("--patient-id", help="患者 ID（患者统计用）")

    # init
    sub.add_parser("init", help="初始化所有数据表")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.module:
        parser.print_help()
        return 0

    data_dir = args.data_dir

    # 路由到对应命令
    handlers = {
        ("patients", "add"): cmd_patients_add,
        ("patients", "search"): cmd_patients_search,
        ("patients", "list"): cmd_patients_list,
        ("records", "add"): cmd_records_add,
        ("records", "search"): cmd_records_search,
        ("herbs", "add"): cmd_herbs_add,
        ("herbs", "update"): cmd_herbs_update,
        ("herbs", "search"): cmd_herbs_search,
        ("herbs", "alerts"): cmd_herbs_alerts,
        ("herbs", "list"): cmd_herbs_list,
        ("appointments", "add"): cmd_appointments_add,
        ("appointments", "search"): cmd_appointments_search,
        ("appointments", "today"): cmd_appointments_today,
        ("appointments", "list"): cmd_appointments_list,
        ("finance", "add"): cmd_finance_add,
        ("finance", "summary"): cmd_finance_summary,
    }

    if args.module == "init":
        return cmd_init(args, data_dir)

    key = (args.module, getattr(args, "action", None))
    handler = handlers.get(key)
    if handler:
        return handler(args, data_dir)
    else:
        parser.parse_args([args.module, "--help"])
        return 0


if __name__ == "__main__":
    sys.exit(main())
