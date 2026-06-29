#!/usr/bin/env python3
"""
A股年报PDF合并三表提取脚本

功能：从年报PDF中提取合并资产负债表/利润表/现金流量表，输出为xlsx文件

用法：
    python extract_financial_tables.py --dir ./数据目录 --config ./pdf_config.json

config.json 格式：
{
  "2022": {
    "file": "年报文件名.pdf",
    "balance_pages": [104, 105, 106],
    "income_pages": [108, 109, 110],
    "cash_pages": [111, 112, 113],
    "ann_id": "公告ID",
    "ann_url": "公告URL",
    "raw_url": "PDF原始URL"
  },
  ...
}

注意：
- 页码为0索引（PDF第1页 = 索引0）
- 三张表所在页码需要先通过pdfplumber探测确认
- 科目名可能存在跨行拆分，脚本使用pending_subject状态机处理
"""

import re
import os
import sys
import json
import argparse
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_args():
    parser = argparse.ArgumentParser(description="年报PDF三表提取")
    parser.add_argument("--dir", required=True, help="数据目录")
    parser.add_argument("--config", required=True, help="PDF页码配置文件(JSON)")
    return parser.parse_args()


def extract_text(pdf_path, page_indices):
    """从PDF指定页提取文本（0索引）"""
    chunks = []
    with pdfplumber.open(pdf_path) as pdf:
        for idx in page_indices:
            if idx < len(pdf.pages):
                text = pdf.pages[idx].extract_text() or ""
                chunks.append(text)
    return "\n".join(chunks)


def parse_table_lines(text, table_name):
    """
    解析三张表为 [(项目, 期末/本期数, 上期末/上期数)]

    关键策略：
    - pending_subject 处理跨行科目名（如"三、营业利润（亏损以"－"号填列）"分3行）
    - 续行特征：以"号填列）"/"列）"/"（亏损以"结尾；整行"（"或"）"
    - 纯数字行且pending_subject非空 → 使用pending_subject
    - 章节标题（数字序号、分类标题）单独成行
    """
    raw_lines = text.splitlines()
    results = []
    num_pattern = r"-?[\d,]+\.\d{2}"
    in_table = False
    pending_subject = None

    def flush_pending():
        nonlocal pending_subject
        if pending_subject and pending_subject.strip():
            results.append((pending_subject.strip(), None, None))
        pending_subject = None

    for line in raw_lines:
        stripped = line.strip()
        if not stripped:
            continue

        # 表头识别
        if table_name in stripped and "母公司" not in stripped:
            in_table = True
            continue

        if not in_table:
            continue

        # 结束条件
        if "母公司" in stripped and any(k in stripped for k in ["资产负债表", "利润表", "现金流量表"]):
            flush_pending()
            break
        if "公司负责人" in stripped:
            flush_pending()
            break
        if "本期发生同一控制下" in stripped:
            flush_pending()
            break

        # 跳过无关行
        if re.match(r"^项目\s+附注\s+\d{4}", stripped):
            continue
        if any(k in stripped for k in ["编制单位", "单位：元", "币种：人民币"]):
            continue
        if re.match(r"^\d{4}年\d{1,2}月\d{1,2}日?$", stripped):
            continue
        if re.match(r"^202\d年[1-9]?[-—]1?[0-2]?月$", stripped):
            continue
        if re.match(r"^\d+\s*/\s*\d+$", stripped):
            continue

        numbers = re.findall(num_pattern, stripped)
        has_number = bool(numbers)

        # 清理行：去掉末尾数字和附注引用
        subject = re.sub(r"\s+", " ", stripped)
        for _ in range(5):
            subject = re.sub(r"\s*-?[\d,]+\.\d{2}\s*$", "", subject)
        subject = re.sub(r"\s*[一二三四五六七八九十]+[、，]?\d+\s*$", "", subject)
        subject = re.sub(r"\s*附注\s*[一二三四五六七八九十0-9]+\s*$", "", subject)
        subject = subject.strip()

        # 续行检测
        is_continuation = False
        if not has_number and subject:
            if (subject.endswith(("号填", "（亏损以", "（净亏损以", "号填列）", "（亏损总额以"))
                or subject in ("列）", "（", "）", "号填列）")):
                is_continuation = True
            elif re.search(r"[一二三四五六七八九十]、.+?（.*?[\"\"][-－][\"\"]$", subject):
                is_continuation = True
            elif re.match(r"^[\"\"']?[-－][\"\"']?号填列[）\)]?$", subject):
                is_continuation = True

        if is_continuation and pending_subject is not None:
            pending_subject = (pending_subject + subject).strip()
            continue

        # 数字行
        if has_number:
            final_subject = subject
            if pending_subject is not None:
                if pending_subject.endswith(("（", "（亏损以", "（净亏损以", "（亏损总额以")):
                    final_subject = pending_subject + subject
                elif pending_subject.endswith("号填"):
                    if subject in ("列）", "）") or subject.startswith("列）"):
                        final_subject = pending_subject + subject
                    else:
                        final_subject = subject
                else:
                    if pending_subject.endswith(("：", ":")):
                        results.append((pending_subject.strip(), None, None))
                    final_subject = subject
                pending_subject = None

            if not final_subject:
                if pending_subject:
                    final_subject = pending_subject
                    pending_subject = None
                else:
                    continue

            if len(numbers) == 1:
                results.append((final_subject, numbers[0], None))
            else:
                cur, prev = numbers[-2], numbers[-1]
                results.append((final_subject, cur, prev))
            continue

        # 无数字行：章节标题
        if re.match(r"^[一二三四五六七八九十]+、", subject):
            flush_pending()
            results.append((subject, None, None))
            continue
        if subject.endswith(("：", ":")) and not subject.startswith("（"):
            flush_pending()
            results.append((subject, None, None))
            continue
        if re.match(r"^[加减]：", subject) or subject.startswith("其中："):
            pending_subject = subject
            continue
        if any(c >= "\u4e00" for c in subject):
            pending_subject = (pending_subject + subject) if pending_subject else subject
            continue

    flush_pending()
    return results


def write_excel(year, info, balance, income, cashflow, base_dir, stock_name="上市公司"):
    """写入xlsx文件，包含3张sheet"""
    wb = Workbook()
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A365D")
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1A365D")
    subject_font = Font(name="微软雅黑", size=10, bold=True)
    normal_font = Font(name="微软雅黑", size=10)
    total_fill = PatternFill("solid", fgColor="E0E7FF")
    section_fill = PatternFill("solid", fgColor="F1F5F9")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def write_sheet(sheet, table_name, title_period, rows):
        sheet.title = table_name

        sheet["A1"] = f"{stock_name}"
        sheet["A1"].font = title_font
        sheet.merge_cells("A1:D1")
        sheet["A1"].alignment = center

        sheet["A2"] = table_name
        sheet["A2"].font = Font(name="微软雅黑", size=13, bold=True, color="0EA5E9")
        sheet.merge_cells("A2:D2")
        sheet["A2"].alignment = center

        sheet["A3"] = title_period
        sheet["A3"].font = Font(name="微软雅黑", size=10, italic=True, color="64748B")
        sheet.merge_cells("A3:D3")
        sheet["A3"].alignment = center

        sheet["A4"] = f"编制单位：{stock_name}"
        sheet["A4"].font = Font(name="微软雅黑", size=9, color="64748B")
        sheet.merge_cells("A4:D4")
        sheet["A4"].alignment = center

        # 表头
        sheet["A7"] = "项目"
        sheet["B7"] = "附注"
        if "资产负债" in table_name:
            sheet["C7"] = f"{year}年12月31日"
            sheet["D7"] = f"{int(year)-1}年12月31日"
        else:
            sheet["C7"] = f"{year}年度"
            sheet["D7"] = f"{int(year)-1}年度"
        for col in ("A7", "B7", "C7", "D7"):
            sheet[col].font = header_font
            sheet[col].fill = header_fill
            sheet[col].alignment = center
            sheet[col].border = border

        row_idx = 8
        for subject, cur, prev in rows:
            is_section = (cur is None and prev is None)
            if is_section:
                for c in range(1, 5):
                    cell = sheet.cell(row=row_idx, column=c)
                    cell.fill = section_fill
                    cell.font = subject_font
                    cell.border = border
                sheet.cell(row=row_idx, column=1, value=subject)
            else:
                sheet.cell(row=row_idx, column=1, value=subject).font = normal_font
                sheet.cell(row=row_idx, column=1).alignment = left_align
                sheet.cell(row=row_idx, column=1).border = border
                sheet.cell(row=row_idx, column=2, value="").font = normal_font
                sheet.cell(row=row_idx, column=2).alignment = center
                sheet.cell(row=row_idx, column=2).border = border
                sheet.cell(row=row_idx, column=3, value=cur).font = normal_font
                sheet.cell(row=row_idx, column=3).alignment = right_align
                sheet.cell(row=row_idx, column=3).border = border
                sheet.cell(row=row_idx, column=3).number_format = "#,##0.00"
                sheet.cell(row=row_idx, column=4, value=prev).font = normal_font
                sheet.cell(row=row_idx, column=4).alignment = right_align
                sheet.cell(row=row_idx, column=4).border = border
                sheet.cell(row=row_idx, column=4).number_format = "#,##0.00"

                if any(kw in subject for kw in ["合计", "总计"]):
                    for c in range(1, 5):
                        sheet.cell(row=row_idx, column=c).fill = total_fill
                        sheet.cell(row=row_idx, column=c).font = subject_font
            row_idx += 1

        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 22

    if balance:
        write_sheet(wb.active, "合并资产负债表", f"{year}年12月31日", balance)
    if income:
        wb.create_sheet("合并利润表")
        write_sheet(wb["合并利润表"], "合并利润表", f"{year}年1-12月", income)
    if cashflow:
        wb.create_sheet("合并现金流量表")
        write_sheet(wb["合并现金流量表"], "合并现金流量表", f"{year}年1-12月", cashflow)

    output_path = os.path.join(base_dir, f"{stock_name}_合并三表_{year}年.xlsx")
    wb.save(output_path)
    return output_path


def main():
    args = parse_args()
    base_dir = args.dir

    with open(args.config, "r", encoding="utf-8") as f:
        pdf_config = json.load(f)

    stock_name = os.path.basename(base_dir.rstrip("/\\")) or "上市公司"
    print("=" * 70)
    print(f"合并三张表提取任务 | 数据目录: {base_dir}")
    print("=" * 70)

    summary = []
    for year, info in pdf_config.items():
        pdf_path = os.path.join(base_dir, info["file"])
        print(f"\n>>> {year}年报: {info['file']}")

        balance_text = extract_text(pdf_path, info["balance_pages"])
        income_text = extract_text(pdf_path, info["income_pages"])
        cashflow_text = extract_text(pdf_path, info["cash_pages"])

        balance = parse_table_lines(balance_text, "合并资产负债表")
        income = parse_table_lines(income_text, "合并利润表")
        cashflow = parse_table_lines(cashflow_text, "合并现金流量表")

        print(f"    解析: 资产负债表{len(balance)}行 | 利润表{len(income)}行 | 现金流量表{len(cashflow)}行")

        xlsx_path = write_excel(year, info, balance, income, cashflow, base_dir, stock_name)
        file_size = os.path.getsize(xlsx_path) / 1024
        print(f"    输出: {xlsx_path} ({file_size:.1f} KB)")

        summary.append({"year": year, "xlsx": xlsx_path, "rows": (len(balance), len(income), len(cashflow))})

    print("\n" + "=" * 70)
    print(f"完成！共生成 {len(summary)} 份xlsx文件")
    for s in summary:
        print(f"  {s['year']}年: {os.path.basename(s['xlsx'])}  三表行数={s['rows']}")
    print("=" * 70)


if __name__ == "__main__":
    main()
