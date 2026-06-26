#!/usr/bin/env python3
"""单月数据解析+写入DB - 用openpyxl read_only+iter_rows避免OOM"""
import sys, os, glob, gc, sqlite3

DATA_DIR = '/workspace/data'
DB_PATH = '/workspace/业务数据/集团客户部FY数据.db'
month = sys.argv[1] if len(sys.argv) > 1 else ''
if not month: print("Usage: parse_one_month.py YYYYMM"); sys.exit(1)

found = None
for p in [os.path.join(DATA_DIR, month, 'KA返佣{}*.xlsx'.format(month)),
          os.path.join(DATA_DIR, month, 'KA{}*.xlsx'.format(month))]:
    fs = glob.glob(p)
    if fs: found = fs[0]; break
if not found: print("ERROR: file not found for", month); sys.exit(1)

import openpyxl
wb = openpyxl.load_workbook(found, data_only=True, read_only=True)

def S(v): return str(v).strip() if v is not None else ''
def F(v):
    try: return float(v) if v is not None else 0
    except: return 0

# --- Parse summary sheet ---
brand_summary = {}
if '汇总' in wb.sheetnames:
    ws = wb['汇总']
    header_row = None
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip().lower() if v else '' for v in row]
        txt = '|'.join(vals)
        if '品牌' in txt and '应返金额' in txt:
            header_row = row_idx
            headers = [str(v or '') for v in row]
            break
        if '品牌' in txt and '入账' in txt:
            header_row = row_idx
            headers = [str(v or '') for v in row]
            break
    
    if header_row is not None:
        brand_col = next((i for i, h in enumerate(headers) if h.strip().lower() == '品牌'), 3)
        amt_col = next((i for i, h in enumerate(headers) if '应返金额' in h or '入账' in h), 8)
        
        for row in ws.iter_rows(min_row=header_row+2, values_only=True):
            if len(row) <= max(brand_col, amt_col): continue
            b = S(row[brand_col])
            if not b or b == '品牌': continue
            amt = F(row[amt_col])
            brand_summary[b] = brand_summary.get(b, 0) + amt

# --- Parse data sheets ---
conn = sqlite3.connect(DB_PATH, timeout=120)
conn.execute("DELETE FROM raw_records WHERE month=?", (month,))
conn.execute("DELETE FROM brand_audit WHERE month=?", (month,))

insert_sql = """INSERT INTO raw_records(
    month, brand, calc_method, sheet_name,
    alipay_txn_amount, alipay_rebate_ratio,
    wechat_txn_amount, wechat_rebate_ratio,
    total_rebate, audit_note, original_note,
    occurrence_date, level1_name, level2_name, level3_name
) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""

raw_batch = []
calc_by_brand = {}
raw_by_brand = {}
adj_by_brand = {}

for sn in wb.sheetnames:
    if sn == '汇总' or sn.lower().strip() == 'sql': continue
    ws = wb[sn]
    
    is_problem = '问题' in sn or '异常' in sn
    is_adj_sheet = '调账' in sn or '调整' in sn
    is_zk = '中快' in sn
    is_zh = '志华' in sn or '拓展' in sn
    
    # Find header row
    header_row = None
    hdrs = []
    total_col_hint = None  # merged cell like '汇总' in rows above header
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Track merged cell headers for total column (志华/拓展 sheets)
        if total_col_hint is None:
            for ci, cv in enumerate(row):
                if cv is not None and str(cv).strip() in ('汇总', '合计', '分润汇总'):
                    total_col_hint = ci
                    break
        vals = [str(v).strip().lower() if v else '' for v in row]
        txt = '|'.join(vals)
        if '品牌' in txt and ('返佣' in txt or '交易金额' in txt or '交易额' in txt):
            header_row = row_idx
            hdrs = [str(v or '') for v in row]
            break
        if '品牌' in txt:
            header_row = row_idx
            hdrs = [str(v or '') for v in row]
            break
        # 拓展/志华sheet表头（英文brand列+中文交易金额列）
        if 'brand' in txt and ('交易金额' in txt or '渠道分润' in txt):
            header_row = row_idx
            hdrs = [str(v or '') for v in row]
            break
        # 英文表头
        if 'brand' in vals and any(k in vals for k in ['zfb_trans_amt', 'wx_trans_amt', 'zfb_commission_amt', 'wx_commission_amt', 'bill_amt']):
            header_row = row_idx
            hdrs = [str(v or '') for v in row]
            break
        if 'brand' in vals and 'rebate_type' in vals:
            header_row = row_idx
            hdrs = [str(v or '') for v in row]
            break
    
    if header_row is None: continue
    
    # Find column indices
    idx = {}
    for i, h in enumerate(hdrs):
        hl = h.lower().strip()
        if hl in ('品牌', '品牌名称', 'brand'): idx['brand'] = i
        elif hl in ('收钱吧商户名称', '商户名称', '商户名', 'merchant_name'): idx['merchant'] = i
        elif hl in ('支付宝交易金额', '支付宝交易额', 'zfb_trans_amt'): idx['ali_txn'] = i
        elif hl in ('支付宝返佣比例', '支付宝分润比例', 'zfb_commission_rate'): idx['ali_ratio'] = i
        elif hl in ('支付宝返佣金额', '支付宝分润金额', 'zfb_commission_amt'): idx['ali_amt'] = i
        elif hl in ('微信交易金额', '微信交易额', 'wx_trans_amt'): idx['wx_txn'] = i
        elif hl in ('微信返佣比例', '微信分润比例', 'wx_commission_rate'): idx['wx_ratio'] = i
        elif hl in ('微信返佣金额', '微信分润金额', 'wx_commission_amt'): idx['wx_amt'] = i
        elif '返佣合计' in hl or '返佣金额' in hl or 'total_commission' in hl or 'bill_amt' in hl or '分润汇总' in hl or hl == '汇总': idx['total'] = i
        elif hl in ('交易期日', '交易期', '月份', 'month', 'months'): idx['occ_date'] = i
        elif '备注' in hl or hl == 'remark': idx['note'] = i
    for i, h in enumerate(hdrs):
        hl = h.lower().strip()
        if not idx.get('level1') and ('一级组织' in hl or '一级' == hl or hl == 'level1_name' or hl == 'level1'): idx['level1'] = i
        if not idx.get('level2') and ('二级组织' in hl or '二级' == hl or hl == 'level2_name' or hl == 'level2'): idx['level2'] = i
        if not idx.get('level3') and ('三级组织' in hl or '三级' == hl or hl == 'level3_name' or hl == 'level3'): idx['level3'] = i
    
    # Set total from merged cell hint if not found in hdrs (志华/拓展 summary column)
    if 'total' not in idx and total_col_hint is not None:
        idx['total'] = total_col_hint
    # Fallback: English-header sheets often have total at col[15] with empty header
    if 'total' not in idx and 'ali_txn' in idx and len(hdrs) > 15:
        idx['total'] = 15

    if 'brand' not in idx:
        if 'merchant' in idx: idx['brand'] = idx['merchant']
        else: continue
    
    # Get sheet note (first row)
    snote = ''
    for row_idx, row in enumerate(ws.iter_rows(max_row=header_row, values_only=True)):
        if row_idx == 0:
            snote = ' '.join(str(v or '') for v in row if v)
            break
    
    bc = idx['brand']
    
    # Compute max known column index for fallback note detection
    max_hdr_col = max(idx.values()) if idx else 0
    
    section_note = ''  # Track section headers from first column (e.g., '回算', '补充返佣')
    
    for row in ws.iter_rows(min_row=header_row+2, values_only=True):
        vals = list(row)
        
        # Check first column for section header (multi-block sheets like 异常及回算)
        fc_raw = S(vals[0]) if len(vals) > 0 else ''
        if fc_raw and fc_raw not in ('\\N', '0', '-'):
            b_check = S(vals[bc]) if len(vals) > bc else ''
            if b_check and b_check.lower() in ('品牌', '品牌名称', 'brand'):
                pass  # This is a second header row, not a section header
            elif not b_check:
                # Section header row — update section_note
                fcl = fc_raw.lower().strip()
                if fcl in ('异常', '问题'):
                    section_note = ''
                else:
                    section_note = fc_raw
        
        if len(vals) <= bc: continue
        b = S(vals[bc])
        if not b or b.lower() in ('品牌', '品牌名称', '合计', '汇总', '小计', '总计', '商户号', ''): continue
        
        # Skip empty first column rows (totals usually)
        fc = S(vals[0]) if len(vals) > 0 else ''
        if not fc and not b: continue
        
        ali_txn = F(vals[idx.get('ali_txn', -1)]) if len(vals) > idx.get('ali_txn', -1) and 'ali_txn' in idx else 0
        ali_ratio = F(vals[idx.get('ali_ratio', -1)]) if len(vals) > idx.get('ali_ratio', -1) and 'ali_ratio' in idx else 0
        ali_amt = F(vals[idx.get('ali_amt', -1)]) if len(vals) > idx.get('ali_amt', -1) and 'ali_amt' in idx else 0
        wx_txn = F(vals[idx.get('wx_txn', -1)]) if len(vals) > idx.get('wx_txn', -1) and 'wx_txn' in idx else 0
        wx_ratio = F(vals[idx.get('wx_ratio', -1)]) if len(vals) > idx.get('wx_ratio', -1) and 'wx_ratio' in idx else 0
        wx_amt = F(vals[idx.get('wx_amt', -1)]) if len(vals) > idx.get('wx_amt', -1) and 'wx_amt' in idx else 0
        total = F(vals[idx.get('total', -1)]) if len(vals) > idx.get('total', -1) and 'total' in idx else 0
        occ = S(vals[idx.get('occ_date', -1)]) if len(vals) > idx.get('occ_date', -1) and 'occ_date' in idx else ''
        note = S(vals[idx.get('note', -1)]) if len(vals) > idx.get('note', -1) and 'note' in idx else ''
        if not note and len(vals) > max_hdr_col + 1:
            # Scan rightmost non-empty column for note (handles empty last col)
            for ci2 in range(len(vals)-1, max_hdr_col, -1):
                tv = str(vals[ci2] or '').strip()
                if tv:
                    note = S(tv)
                    break
        # Use section_note from first-column header if no individual note
        if not note and section_note:
            note = section_note
        l1 = S(vals[idx.get('level1', -1)]) if len(vals) > idx.get('level1', -1) and 'level1' in idx else ''
        l2 = S(vals[idx.get('level2', -1)]) if len(vals) > idx.get('level2', -1) and 'level2' in idx else ''
        l3 = S(vals[idx.get('level3', -1)]) if len(vals) > idx.get('level3', -1) and 'level3' in idx else ''
        
        # Brand mapping for 大客户
        if b == '大客户' and l1:
            b = l1
        
        # Compute recalc
        recalc = round(ali_amt, 2) + round(wx_amt, 2)
        if not recalc:
            recalc = round(ali_txn * ali_ratio, 2) + round(wx_txn * wx_ratio, 2)
        if not recalc:
            recalc = total
        
        if is_zk:
            recalc = total
        
        audit_note = ''
        if is_problem:
            # Priority-based keyword classification
            _nl = (note or '').strip().lower()
            
            # Priority 1: EXCLUDE (checked first - highest priority)
            is_exclude = '不返佣' in _nl or ('有疑问' in _nl and '待确认' in _nl)
            
            # Priority 2: INCLUDE (补发/回算/补充返佣/调账)
            is_include_p2 = any(kw in _nl for kw in ['补发', '回算', '补充返佣', '调账'])
            
            # Priority 3: INCLUDE (本月确认/手工核算/人工核算)
            is_include_p3 = any(kw in _nl for kw in ['本月确认', '手工核算', '人工核算'])
            
            # Priority 4: INCLUDE (other known include patterns)
            is_include_p4 = any(kw in _nl for kw in ['上月待确认', '系统未维护', '手工修改', '遗留', '补算', '已补协议', '人工补算', '发放历史分润', '返佣流程已补', 'sp722'])
            
            if is_exclude:
                audit_note = '有问题，暂时不调整'
            elif is_include_p2 or is_include_p3 or is_include_p4:
                audit_note = '历史差额调整'
                adj_by_brand[b] = adj_by_brand.get(b, 0) + recalc
                raw_by_brand[b] = raw_by_brand.get(b, 0) + total
            else:
                # Rule #3: no matching note → exclude
                audit_note = '有问题，暂时不调整'
        else:
            calc_by_brand[b] = calc_by_brand.get(b, 0) + recalc
            raw_by_brand[b] = raw_by_brand.get(b, 0) + total
        
        calc_method = 'regular'
        if is_zk: calc_method = 'zhongkuai'
        elif is_problem: calc_method = 'abnormal'
        elif is_zh:
            if month == '202307' and '拓展' in sn:
                calc_method = 'expansion_307'
            elif '志华' in sn:
                calc_method = 'expansion_308'
            else:
                calc_method = 'expansion_307'
        
        raw_batch.append((month, b, calc_method, sn,
                          ali_txn, ali_ratio, wx_txn, wx_ratio,
                          total, audit_note, note, occ, l1, l2, l3))
    
    gc.collect()  # Clean up between sheets

wb.close()

# Add adj to calc
for b, v in adj_by_brand.items():
    calc_by_brand[b] = calc_by_brand.get(b, 0) + v

# Write to DB
conn.executemany(insert_sql, raw_batch)

all_brands = set(list(calc_by_brand.keys()) + list(brand_summary.keys()))
for b in sorted(all_brands):
    recalc = round(calc_by_brand.get(b, 0), 2)
    sm = round(brand_summary.get(b, 0), 2)
    ra = round(raw_by_brand.get(b, 0), 2)
    diff = round(sm - recalc, 2)
    if diff < 0:
        is_real, note = 0, '少发，视同无差异'
    elif abs(diff) < 1:
        is_real, note = 0, '小数舍入'
    else:
        is_real, note = 1, ''
    conn.execute("""INSERT INTO brand_audit(month,brand,raw_amount,recalc_amount,summary_amount,diff,diff_note,is_real_diff)
        VALUES(?,?,?,?,?,?,?,?)""", (month, b, ra, recalc, sm, diff, note, is_real))

conn.commit()

real_diff = sum(r[0] for r in conn.execute("SELECT diff FROM brand_audit WHERE month=? AND is_real_diff=1", (month,)).fetchall())
print('{}: {}条, {}品牌, diff={:+.2f}'.format(month, len(raw_batch), len(all_brands), real_diff))

conn.close()
print('DONE')
