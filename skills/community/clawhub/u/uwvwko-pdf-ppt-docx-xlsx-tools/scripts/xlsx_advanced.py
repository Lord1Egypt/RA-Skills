"""XLSX 高级操作: 合并 / 信息 / 摘要

用法:
    合并:  python xlsx_advanced.py merge file1.xlsx file2.xlsx ... [output.xlsx]
    信息:  python xlsx_advanced.py info input.xlsx
    摘要:  python xlsx_advanced.py summary input.xlsx [output.txt]
"""
import openpyxl
import sys
import os


def merge(files, out_path):
    if len(files) < 2:
        print("错误: 合并至少需要 2 个 XLSX 文件")
        sys.exit(1)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for f in files:
        if not os.path.isfile(f):
            print(f"错误: 文件不存在 - {f}")
            sys.exit(1)
        src = openpyxl.load_workbook(f, read_only=True)
        base_name = os.path.splitext(os.path.basename(f))[0]
        for sheet_name in src.sheetnames:
            safe_name = f"{base_name}_{sheet_name}"[:31]
            if safe_name in wb.sheetnames:
                safe_name = safe_name + "_2"
            ws = wb.create_sheet(title=safe_name)
            for row in src[sheet_name].iter_rows(values_only=True):
                ws.append(list(row))
        src.close()
        print(f"Merged: {f}")
    wb.save(out_path)
    print(f"完成: {len(files)} 个文件合并 -> {out_path}")


def info(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print(f"文件: {xlsx_path}")
    print(f"Sheet 数量: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row
        cols = ws.max_column
        print(f"  {name}: {rows} 行 x {cols} 列")
    wb.close()


def summary(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"XLSX 摘要: {xlsx_path}\n")
        f.write(f"Sheet 数量: {len(wb.sheetnames)}\n\n")
        for name in wb.sheetnames:
            ws = wb[name]
            rows = ws.max_row
            cols = ws.max_column
            f.write(f"=== {name} ({rows} 行 x {cols} 列) ===\n")
            # 写入前 5 行作为预览
            preview_count = 0
            for row in ws.iter_rows(values_only=True):
                if preview_count >= 5:
                    f.write("  ...\n")
                    break
                f.write("  " + " | ".join(str(c) if c else "" for c in row) + "\n")
                preview_count += 1
            f.write("\n")
    wb.close()
    print(f"完成: 摘要 -> {out_path}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    action = sys.argv[1].lower()

    if action == "merge":
        out = "merged.xlsx"
        files = []
        for arg in sys.argv[2:]:
            if arg.endswith(".xlsx") and os.path.isfile(arg):
                files.append(arg)
            else:
                out = arg
        if not files:
            print("错误: 未找到有效的 XLSX 文件")
            sys.exit(1)
        else:
            merge(files, out)

    elif action == "info":
        info(sys.argv[2])

    elif action == "summary":
        xlsx_path = sys.argv[2]
        out_path = sys.argv[3] if len(sys.argv) > 3 else "xlsx_summary.txt"
        summary(xlsx_path, out_path)

    else:
        print(f"未知操作: {action}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()