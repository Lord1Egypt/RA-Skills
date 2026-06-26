"""
软件开发工时评估 Excel 生成器
输入：需求描述和拆分后的工作项
输出：多 Sheet 的 Excel 评估报告
"""

import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

# 中国法定节假日（示例，可扩展）
HOLIDAYS = [
    # 2026年
    datetime(2026, 1, 1),   # 元旦
    datetime(2026, 1, 28), datetime(2026, 1, 29), datetime(2026, 1, 30),  # 春节
    datetime(2026, 2, 1), datetime(2026, 2, 2), datetime(2026, 2, 3), datetime(2026, 2, 4),
    datetime(2026, 4, 4), datetime(2026, 4, 5), datetime(2026, 4, 6),  # 清明
    datetime(2026, 5, 1), datetime(2026, 5, 2), datetime(2026, 5, 3),  # 劳动节
    datetime(2026, 6, 1),  # 端午
    datetime(2026, 10, 1), datetime(2026, 10, 2), datetime(2026, 10, 3),  # 国庆
    datetime(2026, 10, 4), datetime(2026, 10, 5), datetime(2026, 10, 6), datetime(2026, 10, 7),
]

# 样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
SUBTITLE_FONT = Font(size=11, bold=True)
MONEY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def is_working_day(date):
    """判断是否为工作日（跳过周末和节假日）"""
    if date.weekday() >= 5:  # 0=周一, 5=周六, 6=周日
        return False
    if date in HOLIDAYS:
        return False
    return True

def add_working_days(start_date, days):
    """添加工作日后返回结束日期（跳过周末和节假日）"""
    current = start_date
    remaining = days
    while remaining > 0:
        current += timedelta(days=1)
        if is_working_day(current):
            remaining -= 1
    return current

def get_working_days_between(start_date, end_date):
    """计算两个日期之间的工作日数"""
    count = 0
    current = start_date
    while current <= end_date:
        if is_working_day(current):
            count += 1
        current += timedelta(days=1)
    return count

def set_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER_THIN
    return cell

def set_cell(ws, row, col, value, bold=False, align='left', fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill
    return cell

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_estimation_excel(requirements: str, modules: list, output_path: str = None):
    """
    生成工时评估 Excel
    
    Args:
        requirements: 需求描述
        modules: 工作模块列表，每项包含:
            {
                "name": "模块名称",
                "desc": "模块描述",
                "items": [
                    {
                        "name": "工作项名称",
                        "analysis": 1.0,  # 需求分析人天
                        "design": 2.0,     # 设计人天
                        "frontend": 3.0,   # 前端人天
                        "backend": 5.0,    # 后台人天
                        "algorithm": 0.0,  # 算法人天
                        "test": 2.0,       # 测试人天
                        "complexity": "中",
                        "risk": "低",
                        "parallel": True,
                        "prerequisite": "",
                        "coordination": ""
                    }
                ]
            }
        output_path: 输出路径
    """
    wb = Workbook()
    
    # Sheet 1: 工时总览
    create_overview_sheet(wb, modules)
    
    # Sheet 2-7: 各维度详情
    create_dimensions_sheets(wb, modules)
    
    # Sheet 8: 甘特图
    create_gantt_sheet(wb, modules)
    
    # Sheet 9: 重点评估
    create_key_risks_sheet(wb, modules)
    
    # Sheet 10: 关系协调
    create_coordination_sheet(wb, modules)
    
    # Sheet 11: 成本估算
    create_cost_sheet(wb, modules)
    
    # 保存
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"工时评估_{timestamp}.xlsx"
    
    wb.save(output_path)
    return output_path

def create_overview_sheet(wb, modules):
    ws = wb.active
    ws.title = "工时总览"
    
    # 标题
    ws.cell(row=1, column=1, value="软件开发工时评估总览").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # 表头
    headers = ["工作模块", "工作项", "需求分析", "设计", "前端", "后台", "算法", "测试", "小计", "复杂度", "风险", "并行"]
    for i, h in enumerate(headers, 1):
        set_header(ws, 4, i, h)
    
    row = 5
    total = {"analysis": 0, "design": 0, "frontend": 0, "backend": 0, "algorithm": 0, "test": 0}
    module_starts = []
    
    for module in modules:
        module_start = row
        for item in module.get("items", []):
            subtotal = item.get("analysis", 0) + item.get("design", 0) + item.get("frontend", 0) + \
                      item.get("backend", 0) + item.get("algorithm", 0) + item.get("test", 0)
            
            set_cell(ws, row, 1, module["name"])
            set_cell(ws, row, 2, item["name"])
            set_cell(ws, row, 3, item.get("analysis", 0))
            set_cell(ws, row, 4, item.get("design", 0))
            set_cell(ws, row, 5, item.get("frontend", 0))
            set_cell(ws, row, 6, item.get("backend", 0))
            set_cell(ws, row, 7, item.get("algorithm", 0))
            set_cell(ws, row, 8, item.get("test", 0))
            set_cell(ws, row, 9, subtotal, bold=True, align='center')
            set_cell(ws, row, 10, item.get("complexity", "中"))
            set_cell(ws, row, 11, item.get("risk", "低"))
            set_cell(ws, row, 12, "✓" if item.get("parallel", True) else "×")
            
            total["analysis"] += item.get("analysis", 0)
            total["design"] += item.get("design", 0)
            total["frontend"] += item.get("frontend", 0)
            total["backend"] += item.get("backend", 0)
            total["algorithm"] += item.get("algorithm", 0)
            total["test"] += item.get("test", 0)
            row += 1
        
        module_starts.append((module["name"], module_start, row - 1))
    
    # 合计行
    row += 1
    set_header(ws, row, 1, "合计")
    set_cell(ws, row, 2, "", bold=True)
    set_cell(ws, row, 3, total["analysis"], bold=True, align='center')
    set_cell(ws, row, 4, total["design"], bold=True, align='center')
    set_cell(ws, row, 5, total["frontend"], bold=True, align='center')
    set_cell(ws, row, 6, total["backend"], bold=True, align='center')
    set_cell(ws, row, 7, total["algorithm"], bold=True, align='center')
    set_cell(ws, row, 8, total["test"], bold=True, align='center')
    grand_total = sum(total.values())
    set_cell(ws, row, 9, grand_total, bold=True, align='center')
    
    # 维度统计
    row += 2
    ws.cell(row=row, column=1, value="维度工时统计").font = SUBTITLE_FONT
    row += 1
    dim_headers = ["维度", "工时(人天)", "占比"]
    for i, h in enumerate(dim_headers, 1):
        set_header(ws, row, i, h)
    row += 1
    
    dimensions = [
        ("需求分析", total["analysis"]),
        ("设计", total["design"]),
        ("前端", total["frontend"]),
        ("后台", total["backend"]),
        ("算法", total["algorithm"]),
        ("测试", total["test"]),
    ]
    
    for dim, hours in dimensions:
        if hours > 0:
            pct = f"{hours/grand_total*100:.1f}%" if grand_total > 0 else "0%"
            set_cell(ws, row, 1, dim)
            set_cell(ws, row, 2, hours, align='center')
            set_cell(ws, row, 3, pct, align='center')
            row += 1
    
    auto_width(ws)
    
    # 添加工时分布图表
    create_distribution_charts(ws, modules)

def create_dimensions_sheets(wb, modules):
    dimension_map = {
        "需求分析": "analysis",
        "设计": "design",
        "前端": "frontend",
        "后台": "backend",
        "算法": "algorithm",
        "测试": "test"
    }
    
    for sheet_name, key in dimension_map.items():
        ws = wb.create_sheet(title=sheet_name)
        
        ws.cell(row=1, column=1, value=f"{sheet_name}详情").font = TITLE_FONT
        
        headers = ["工作模块", "工作项", "工作内容", "评估工时(人天)", "评估依据", "复杂度", "备注"]
        for i, h in enumerate(headers, 1):
            set_header(ws, 3, i, h)
        
        row = 4
        for module in modules:
            for item in module.get("items", []):
                hours = item.get(key, 0)
                if hours > 0:
                    set_cell(ws, row, 1, module["name"])
                    set_cell(ws, row, 2, item["name"])
                    set_cell(ws, row, 3, item.get("desc", ""))
                    set_cell(ws, row, 4, hours, align='center')
                    set_cell(ws, row, 5, item.get("basis", f"基于{sheet_name}标准"))
                    set_cell(ws, row, 6, item.get("complexity", "中"))
                    set_cell(ws, row, 7, item.get("note", ""))
                    row += 1
        
        auto_width(ws)

def create_gantt_sheet(wb, modules):
    ws = wb.create_sheet(title="甘特图")
    
    ws.cell(row=1, column=1, value="项目进度甘特图（跳过周末和节假日）").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d')}")
    
    headers = ["任务ID", "任务名称", "执行人", "开始日期", "结束日期", "工作日(天)", "日历日(天)", "前置任务", "状态", "里程碑"]
    for i, h in enumerate(headers, 1):
        set_header(ws, 3, i, h)
    
    # 从今天开始，跳过周末和节假日
    start_date = datetime.now()
    # 确保从工作日开始
    while not is_working_day(start_date):
        start_date += timedelta(days=1)
    
    row = 4
    task_id = 1
    milestones = ["需求确认", "设计完成", "开发完成", "测试完成", "上线部署"]
    milestone_idx = 0
    
    for module in modules:
        for item in module.get("items", []):
            total_hours = item.get("analysis", 0) + item.get("design", 0) + \
                         item.get("frontend", 0) + item.get("backend", 0) + \
                         item.get("algorithm", 0) + item.get("test", 0)
            working_days = max(1, int(total_hours))
            
            # 计算工作日结束日期
            end_date = add_working_days(start_date, working_days)
            
            # 计算日历天数（含休息日）
            calendar_days = (end_date - start_date).days + 1
            
            # 判断里程碑
            is_milestone = ""
            if milestone_idx < len(milestones) and working_days >= 5:
                is_milestone = milestones[milestone_idx]
                milestone_idx += 1
            
            set_cell(ws, row, 1, f"T{task_id:03d}")
            set_cell(ws, row, 2, f"{module['name']}-{item['name']}")
            set_cell(ws, row, 3, item.get("assignee", "待分配"))
            set_cell(ws, row, 4, start_date.strftime("%Y-%m-%d"))
            set_cell(ws, row, 5, end_date.strftime("%Y-%m-%d"))
            set_cell(ws, row, 6, working_days, align='center')
            set_cell(ws, row, 7, calendar_days, align='center')
            set_cell(ws, row, 8, item.get("prerequisite", "-"))
            set_cell(ws, row, 9, "待开始")
            set_cell(ws, row, 10, is_milestone)
            
            # 下一个任务从休息日后开始（跳过周末和节假日）
            next_start = end_date + timedelta(days=1)
            while not is_working_day(next_start):
                next_start += timedelta(days=1)
            start_date = next_start
            
            task_id += 1
            row += 1
        
        # 模块间休息1天（跳过周末和节假日）
        start_date += timedelta(days=1)
        while not is_working_day(start_date):
            start_date += timedelta(days=1)
    
    # 项目总工期
    row += 2
    if task_id > 1:
        ws.cell(row=row, column=1, value="项目总工期（工作日）").font = SUBTITLE_FONT
        # 重新计算总工期
        total_start = datetime.now()
        while not is_working_day(total_start):
            total_start += timedelta(days=1)
        final_end = datetime.now()
        for m in modules:
            for it in m.get("items", []):
                days = int(it.get("analysis", 0) + it.get("design", 0) + it.get("frontend", 0) + \
                       it.get("backend", 0) + it.get("algorithm", 0) + it.get("test", 0))
                final_end = add_working_days(total_start, days)
                total_start = final_end + timedelta(days=1)
                while not is_working_day(total_start):
                    total_start += timedelta(days=1)
        total_workdays = get_working_days_between(datetime.now(), final_end)
        ws.cell(row=row, column=3, value=f"约 {total_workdays} 个工作日")
    
    auto_width(ws)
    
    # 添加甘特图条形图
    create_gantt_chart(ws, modules)

def create_cost_sheet(wb, modules):
    """创建成本估算表"""
    ws = wb.create_sheet(title="成本估算")
    
    ws.cell(row=1, column=1, value="项目成本估算").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d')}")
    
    # ========== 人力成本 ==========
    row = 4
    ws.cell(row=row, column=1, value="一、人力成本").font = SUBTITLE_FONT
    row += 1
    
    headers = ["角色", "工时(人天)", "人数", "日均成本(元)", "小计(元)", "备注"]
    for i, h in enumerate(headers, 1):
        set_header(ws, row, i, h)
    row += 1
    
    # 计算各角色总工时
    total = {"analysis": 0, "design": 0, "frontend": 0, "backend": 0, "algorithm": 0, "test": 0}
    for module in modules:
        for item in module.get("items", []):
            total["analysis"] += item.get("analysis", 0)
            total["design"] += item.get("design", 0)
            total["frontend"] += item.get("frontend", 0)
            total["backend"] += item.get("backend", 0)
            total["algorithm"] += item.get("algorithm", 0)
            total["test"] += item.get("test", 0)
    
    # 角色映射和日均成本（可配置）
    role_rates = [
        ("需求分析师", total["analysis"], 1, 1500, "需求分析"),
        ("UI/UX设计师", total["design"], 1, 1200, "设计"),
        ("前端工程师", total["frontend"], 1, 1500, "前端"),
        ("后端工程师", total["backend"], 1, 1800, "后台"),
        ("算法工程师", total["algorithm"], 1, 2000, "算法"),
        ("测试工程师", total["test"], 1, 1200, "测试"),
    ]
    
    total_labor = 0
    for role, days, count, daily_rate, _ in role_rates:
        if days > 0:
            subtotal = days * count * daily_rate
            total_labor += subtotal
            set_cell(ws, row, 1, role)
            set_cell(ws, row, 2, days, align='center')
            set_cell(ws, row, 3, count, align='center')
            set_cell(ws, row, 4, daily_rate, align='center')
            set_cell(ws, row, 5, subtotal, align='center', fill=MONEY_FILL)
            set_cell(ws, row, 6, "")
            row += 1
    
    # 人力成本合计
    set_header(ws, row, 1, "人力成本合计")
    set_cell(ws, row, 5, total_labor, bold=True, align='center', fill=MONEY_FILL)
    row += 2
    
    # ========== 软硬件成本 ==========
    ws.cell(row=row, column=1, value="二、软硬件成本").font = SUBTITLE_FONT
    row += 1
    
    headers = ["类别", "项目", "规格/数量", "单次成本(元)", "周期(月)", "小计(元)", "备注"]
    for i, h in enumerate(headers, 1):
        set_header(ws, row, i, h)
    row += 1
    
    # 软硬件成本项目
    hw_items = [
        ("服务器", "云服务器（ECS）", "2核4G", 500, 3, "部署、后端服务"),
        ("服务器", "数据库服务（RDS）", "基础版", 300, 3, "MySQL数据库"),
        ("服务器", "对象存储（OSS）", "100GB", 50, 3, "文件存储"),
        ("域名", "域名注册", "1个", 50, 12, "域名费用"),
        ("SSL证书", "HTTPS证书", "1个/年", 200, 12, "安全证书"),
        ("第三方服务", "短信服务", "按量付费", 100, 3, "验证码短信"),
        ("第三方服务", "支付通道", "按交易收费", 0, 3, "支付宝/微信"),
        ("第三方服务", "CDN加速", "基础套餐", 100, 3, "静态资源加速"),
        ("软件", "开发工具", "IDE许可证", 0, 0, "免费工具"),
        ("软件", "设计软件", "设计工具", 0, 0, "免费/Figma"),
    ]
    
    total_hw = 0
    for cat, item, spec, unit_cost, months, note in hw_items:
        subtotal = unit_cost * months
        total_hw += subtotal
        set_cell(ws, row, 1, cat)
        set_cell(ws, row, 2, item)
        set_cell(ws, row, 3, spec)
        set_cell(ws, row, 4, unit_cost if unit_cost > 0 else "-", align='center')
        set_cell(ws, row, 5, f"{months}月" if months > 0 else "-", align='center')
        set_cell(ws, row, 6, subtotal if subtotal > 0 else "-", align='center', fill=MONEY_FILL)
        set_cell(ws, row, 7, note)
        row += 1
    
    # 软硬件成本合计
    set_header(ws, row, 1, "软硬件成本合计")
    set_cell(ws, row, 6, total_hw, bold=True, align='center', fill=MONEY_FILL)
    row += 2
    
    # ========== 项目总成本 ==========
    ws.cell(row=row, column=1, value="三、项目总成本").font = SUBTITLE_FONT
    row += 1
    
    total_project = total_labor + total_hw
    set_header(ws, row, 1, "项目总预算")
    set_cell(ws, row, 2, total_project, bold=True, align='center', fill=MONEY_FILL)
    ws.cell(row=row, column=3, value=f"（人力{total_labor}元 + 软硬件{total_hw}元）")
    
    row += 2
    
    # ========== 成本说明 ==========
    ws.cell(row=row, column=1, value="四、成本说明").font = SUBTITLE_FONT
    row += 1
    notes = [
        "1. 人力成本按每天8小时工作制计算",
        "2. 日均成本为参考价，可根据实际情况调整",
        "3. 软硬件成本按最低配置估算，流量费用另计",
        "4. 第三方服务（支付、短信）通常有交易手续费",
        "5. 未包含项目管理和沟通成本",
        "6. 预留10-20%应急预算",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note)
        row += 1
    
    # 建议预算
    row += 1
    recommended = int(total_project * 1.15)  # 15% buffer
    set_cell(ws, row, 1, f"建议项目预算（含15%应急）: ", bold=True)
    set_cell(ws, row, 2, recommended, bold=True, align='center', fill=MONEY_FILL)
    ws.cell(row=row, column=3, value="元")
    
    auto_width(ws)

def create_key_risks_sheet(wb, modules):
    ws = wb.create_sheet(title="重点评估")
    
    ws.cell(row=1, column=1, value="重点评估与风险项").font = TITLE_FONT
    ws.cell(row=2, column=1, value="以下列出高风险、不确定性大或技术难点明显的工作项")
    
    headers = ["工作模块", "工作项", "风险类型", "风险描述", "影响评估", "建议措施", "优先级"]
    for i, h in enumerate(headers, 1):
        set_header(ws, 4, i, h)
    
    row = 5
    risk_types = {
        "高": "高风险",
        "中": "中等风险",
        "低": "低风险"
    }
    
    for module in modules:
        for item in module.get("items", []):
            risk = item.get("risk", "低")
            if risk in ["高", "中"]:
                # 评估不确定性
                if "algorithm" in item and item.get("algorithm", 0) > 3:
                    risk_type = "技术难点"
                elif not item.get("basis"):
                    risk_type = "需求不明确"
                else:
                    risk_type = risk_types.get(risk, "其他")
                
                set_cell(ws, row, 1, module["name"])
                set_cell(ws, row, 2, item["name"])
                set_cell(ws, row, 3, risk_type)
                set_cell(ws, row, 4, item.get("risk_desc", f"该工作项复杂度{item.get('complexity', '中')}，存在一定不确定性"))
                set_cell(ws, row, 5, item.get("impact", "可能导致进度延误或需要额外资源"))
                set_cell(ws, row, 6, item.get("suggestion", "建议预留buffer时间，提前技术验证"))
                set_cell(ws, row, 7, "高" if risk == "高" else "中", align='center')
                row += 1
    
    if row == 5:
        set_cell(ws, row, 1, "暂无高风险项")
    
    auto_width(ws)

def create_coordination_sheet(wb, modules):
    ws = wb.create_sheet(title="关系协调")
    
    ws.cell(row=1, column=1, value="工作关系与协调事项").font = TITLE_FONT
    
    headers = ["工作模块", "工作项", "前置依赖", "协调事项", "协调对象", "协调时间点", "备注"]
    for i, h in enumerate(headers, 1):
        set_header(ws, 3, i, h)
    
    row = 4
    for module in modules:
        for item in module.get("items", []):
            # 检查是否有协调事项
            has_coordination = item.get("coordination") or item.get("prerequisite")
            
            set_cell(ws, row, 1, module["name"])
            set_cell(ws, row, 2, item["name"])
            set_cell(ws, row, 3, item.get("prerequisite", "-"))
            set_cell(ws, row, 4, item.get("coordination", "-"))
            set_cell(ws, row, 5, item.get("coord_target", "待确认"))
            set_cell(ws, row, 6, item.get("coord_time", "开发前"))
            set_cell(ws, row, 7, item.get("note", ""))
            row += 1
    
    # 添加协调关系说明
    row += 2
    ws.cell(row=row, column=1, value="协调关系类型说明:").font = SUBTITLE_FONT
    row += 1
    coord_types = [
        ("前置依赖", "某工作项必须在其他工作项完成后才能开始"),
        ("接口协调", "前后端需协调接口定义和数据格式"),
        ("资源协调", "需要申请特定资源（服务器、第三方服务等）"),
        ("评审协调", "需要安排评审会议（设计评审、代码评审等）"),
    ]
    for coord_type, desc in coord_types:
        set_cell(ws, row, 1, coord_type, bold=True)
        set_cell(ws, row, 2, desc)
        row += 1
    
    auto_width(ws)

def create_gantt_chart(ws, modules):
    """在甘特图Sheet中创建条形图"""
    # 准备图表数据区域（在甘特图数据下方）
    chart_start_row = ws.max_row + 3
    
    # 写入图表数据：任务名、开始日期、时长
    ws.cell(row=chart_start_row, column=1, value="任务名称").font = Font(bold=True)
    ws.cell(row=chart_start_row, column=2, value="开始日期").font = Font(bold=True)
    ws.cell(row=chart_start_row, column=3, value="时长(天)").font = Font(bold=True)
    
    row = chart_start_row + 1
    chart_data_start = row
    
    start_date = datetime.now()
    while not is_working_day(start_date):
        start_date += timedelta(days=1)
    
    for module in modules:
        for item in module.get("items", []):
            total = item.get("analysis", 0) + item.get("design", 0) + \
                   item.get("frontend", 0) + item.get("backend", 0) + \
                   item.get("algorithm", 0) + item.get("test", 0)
            days = max(1, int(total))
            end_date = add_working_days(start_date, days)
            
            ws.cell(row=row, column=1, value=f"{module['name']}-{item['name']}")
            ws.cell(row=row, column=2, value=start_date)
            ws.cell(row=row, column=3, value=days)
            
            # 格式化日期
            ws.cell(row=row, column=2).number_format = 'YYYY-MM-DD'
            
            next_start = end_date + timedelta(days=1)
            while not is_working_day(next_start):
                next_start += timedelta(days=1)
            start_date = next_start
            row += 1
    
    chart_data_end = row - 1
    
    # 创建甘特图
    chart = BarChart()
    chart.type = "bar"  # 横向条形图
    chart.title = "项目进度甘特图"
    chart.y_axis.title = "任务"
    chart.x_axis.title = "日期"
    chart.style = 10
    
    # 数据系列
    data = Reference(ws, min_col=3, min_row=chart_start_row, max_row=chart_data_end)
    cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_data_end)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 12
    
    # 放置图表
    ws.add_chart(chart, f"H{chart_start_row}")

def create_distribution_charts(ws, modules):
    """在工作总览Sheet中创建工时分布图表"""
    # 计算各维度工时
    total = {"analysis": 0, "design": 0, "frontend": 0, "backend": 0, "algorithm": 0, "test": 0}
    module_totals = {}
    
    for module in modules:
        module_total = 0
        for item in module.get("items", []):
            item_total = item.get("analysis", 0) + item.get("design", 0) + \
                        item.get("frontend", 0) + item.get("backend", 0) + \
                        item.get("algorithm", 0) + item.get("test", 0)
            total["analysis"] += item.get("analysis", 0)
            total["design"] += item.get("design", 0)
            total["frontend"] += item.get("frontend", 0)
            total["backend"] += item.get("backend", 0)
            total["algorithm"] += item.get("algorithm", 0)
            total["test"] += item.get("test", 0)
            module_total += item_total
        module_totals[module["name"]] = module_total
    
    grand_total = sum(total.values())
    
    # 找到总览Sheet的最后一行
    chart_row = ws.max_row + 3
    
    # ========== 维度占比饼图 ==========
    ws.cell(row=chart_row, column=1, value="工时维度占比").font = Font(bold=True, size=12)
    chart_row += 1
    
    # 写入饼图数据
    ws.cell(row=chart_row, column=1, value="维度")
    ws.cell(row=chart_row, column=2, value="工时(人天)")
    pie_data_row = chart_row + 1
    
    dimensions = [("需求分析", total["analysis"]),
                  ("设计", total["design"]),
                  ("前端", total["frontend"]),
                  ("后台", total["backend"]),
                  ("算法", total["algorithm"]),
                  ("测试", total["test"])]
    
    row = pie_data_row
    for dim, hours in dimensions:
        if hours > 0:
            ws.cell(row=row, column=1, value=dim)
            ws.cell(row=row, column=2, value=hours)
            row += 1
    pie_data_end = row - 1
    
    # 创建饼图
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=pie_data_row, max_row=pie_data_end)
    data = Reference(ws, min_col=2, min_row=pie_data_row - 1, max_row=pie_data_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "各维度工时占比"
    pie.style = 10
    pie.width = 12
    pie.height = 10
    
    # 添加数据标签
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showCatName = True
    
    ws.add_chart(pie, f"D{chart_row}")
    
    # ========== 模块占比柱状图 ==========
    chart_row = pie_data_end + 3
    ws.cell(row=chart_row, column=1, value="各模块工时对比").font = Font(bold=True, size=12)
    chart_row += 1
    
    # 写入柱状图数据
    ws.cell(row=chart_row, column=1, value="模块")
    ws.cell(row=chart_row, column=2, value="工时(人天)")
    bar_data_row = chart_row + 1
    
    row = bar_data_row
    for module_name, hours in module_totals.items():
        ws.cell(row=row, column=1, value=module_name)
        ws.cell(row=row, column=2, value=hours)
        row += 1
    bar_data_end = row - 1
    
    # 创建柱状图
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "各模块工时对比"
    bar.y_axis.title = "工时(人天)"
    bar.x_axis.title = "模块"
    
    labels = Reference(ws, min_col=1, min_row=bar_data_row, max_row=bar_data_end)
    data = Reference(ws, min_col=2, min_row=bar_data_row - 1, max_row=bar_data_end)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.width = 14
    bar.height = 10
    
    ws.add_chart(bar, f"D{chart_row}")

def parse_requirements(requirements_text: str) -> list:
    """
    解析需求文本，生成模块结构
    这是一个简化的解析，实际使用时可能需要更复杂的处理
    """
    # 简单的模块拆分逻辑
    modules = []
    current_module = None
    
    lines = requirements_text.split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检测是否是模块标题（通常是 ## 或 ### 开头，或者是 "XX模块" 格式）
        if line.startswith("#"):
            if current_module:
                modules.append(current_module)
            current_module = {
                "name": line.lstrip("#").strip(),
                "desc": "",
                "items": []
            }
        elif "模块" in line and ":" in line:
            if current_module:
                modules.append(current_module)
            module_name = line.split(":")[0].strip()
            module_desc = line.split(":")[1].strip() if ":" in line else ""
            current_module = {
                "name": module_name,
                "desc": module_desc,
                "items": []
            }
    
    if current_module:
        modules.append(current_module)
    
    return modules

if __name__ == "__main__":
    # 测试
    test_modules = [
        {
            "name": "用户系统",
            "desc": "用户登录注册相关功能",
            "items": [
                {
                    "name": "登录注册",
                    "desc": "手机号+验证码登录",
                    "analysis": 1.0,
                    "design": 1.0,
                    "frontend": 2.0,
                    "backend": 3.0,
                    "algorithm": 0,
                    "test": 1.0,
                    "complexity": "低",
                    "risk": "低",
                    "parallel": True,
                    "prerequisite": "",
                    "coordination": "需与短信服务商协调"
                }
            ]
        }
    ]
    
    output = generate_estimation_excel("测试需求", test_modules)
    print(f"已生成: {output}")
