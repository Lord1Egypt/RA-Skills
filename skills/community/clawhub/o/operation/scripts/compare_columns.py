#!/usr/bin/env python3
"""
Excel Column Compare - Compare two columns in an Excel file and find differences.

Usage:
    python compare_columns.py <input_file> <column_a> <column_b> [--sheet <sheet_name>] [--output <output_file>] [--mode <mode>]

Arguments:
    input_file       Path to the Excel file (.xlsx, .xls, .csv)
    column_a         Name or letter (e.g., A, B) of the first column to compare
    column_b         Name or letter (e.g., A, B) of the second column to compare

Options:
    --sheet          Sheet name or index (0-based). Default: first sheet
    --output         Output Excel file path. Default: <input>_comparison_result.xlsx
    --mode           Comparison mode:
                       full    - Show all categories (default)
                       diff    - Show only differences (not in both)
                       unique_a - Show only items unique to column A
                       unique_b - Show only items unique to column B
                       common  - Show only items common to both columns

Examples:
    python compare_columns.py data.xlsx "姓名" "名字"
    python compare_columns.py data.xlsx A B --sheet "Sheet1"
    python compare_columns.py data.xlsx "Email" "邮箱" --mode diff
    python compare_columns.py data.xlsx C D --output result.xlsx --mode unique_a
"""

import argparse
import sys
import os

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Required packages not found. Installing pandas and openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "-q"])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def resolve_column(df, col_ref):
    """Resolve a column reference (name, letter, or 0-based index) to a column name."""
    # If it's already a column name in the dataframe
    if col_ref in df.columns:
        return col_ref

    # Try as a column letter (A=0, B=1, ..., Z=25, AA=26, etc.)
    col_upper = col_ref.upper()
    if col_upper.isalpha() and len(col_upper) <= 3:
        col_idx = 0
        for ch in col_upper:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        col_idx -= 1  # Convert to 0-based
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]

    # Try as a numeric index
    try:
        col_idx = int(col_ref)
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]
    except ValueError:
        pass

    return None


def compare_columns(input_file, col_a_ref, col_b_ref, sheet=None, output=None, mode="full"):
    """Compare two columns in an Excel/CSV file and generate a comparison report."""

    # Read the file
    ext = os.path.splitext(input_file)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(input_file)
    else:
        if sheet is not None:
            try:
                sheet = int(sheet)
            except ValueError:
                pass
        df = pd.read_excel(input_file, sheet_name=sheet if sheet is not None else 0)

    # Resolve column references
    col_a_name = resolve_column(df, col_a_ref)
    col_b_name = resolve_column(df, col_b_ref)

    if col_a_name is None:
        print(f"Error: Column '{col_a_ref}' not found. Available columns: {list(df.columns)}")
        sys.exit(1)
    if col_b_name is None:
        print(f"Error: Column '{col_b_ref}' not found. Available columns: {list(df.columns)}")
        sys.exit(1)

    if col_a_name == col_b_name:
        print("Error: Both column references resolve to the same column. Please choose two different columns.")
        sys.exit(1)

    # Extract column values, drop NaN
    series_a = df[col_a_name].dropna().astype(str).str.strip()
    series_b = df[col_b_name].dropna().astype(str).str.strip()

    set_a = set(series_a)
    set_b = set(series_b)

    # Compute comparison sets
    only_in_a = sorted(set_a - set_b)
    only_in_b = sorted(set_b - set_a)
    in_both = sorted(set_a & set_b)

    # Determine output path
    if output is None:
        base, ext = os.path.splitext(input_file)
        output = f"{base}_comparison_result.xlsx"

    # Print summary to console
    print(f"\n{'='*50}")
    print(f"  Excel Column Comparison Report")
    print(f"{'='*50}")
    print(f"  File:     {input_file}")
    print(f"  Column A: {col_a_name} ({len(set_a)} unique values)")
    print(f"  Column B: {col_b_name} ({len(set_b)} unique values)")
    print(f"{'─'*50}")
    print(f"  Only in '{col_a_name}': {len(only_in_a)} items")
    print(f"  Only in '{col_b_name}': {len(only_in_b)} items")
    print(f"  In both columns:        {len(in_both)} items")
    print(f"{'='*50}\n")

    # Create output workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_a = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_b = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_fill_common = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_fill_summary = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, row, col, fill, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = cell_alignment
        cell.border = thin_border

    def style_cell(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_data = [
        ("Metric", "Value"),
        ("Source File", input_file),
        ("Column A", col_a_name),
        ("Column B", col_b_name),
        ("Unique values in Column A", len(set_a)),
        ("Unique values in Column B", len(set_b)),
        ("Only in Column A", len(only_in_a)),
        ("Only in Column B", len(only_in_b)),
        ("Common to both", len(in_both)),
    ]
    for r, (label, value) in enumerate(summary_data, 1):
        if r == 1:
            style_header(ws_summary, r, 1, header_fill_summary, label)
            style_header(ws_summary, r, 2, header_fill_summary, value)
        else:
            style_cell(ws_summary, r, 1, label)
            style_cell(ws_summary, r, 2, value)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40

    # Sheet 2: Only in Column A
    if mode in ("full", "diff", "unique_a"):
        ws_a = wb.create_sheet(title=f"Only in {col_a_name[:28]}")
        style_header(ws_a, 1, 1, header_fill_a, f"Only in '{col_a_name}'")
        style_header(ws_a, 1, 2, header_fill_a, "Row Numbers (in original)")
        for i, val in enumerate(only_in_a, 2):
            # Find row numbers where this value appears
            rows = series_a[series_a == val].index.tolist()
            rows_str = ", ".join(str(r + 2) for r in rows[:10])  # +2 for 1-indexed + header
            if len(rows) > 10:
                rows_str += f" ... (+{len(rows)-10} more)"
            style_cell(ws_a, i, 1, val)
            style_cell(ws_a, i, 2, rows_str)
        ws_a.column_dimensions['A'].width = 40
        ws_a.column_dimensions['B'].width = 50

    # Sheet 3: Only in Column B
    if mode in ("full", "diff", "unique_b"):
        ws_b = wb.create_sheet(title=f"Only in {col_b_name[:28]}")
        style_header(ws_b, 1, 1, header_fill_b, f"Only in '{col_b_name}'")
        style_header(ws_b, 1, 2, header_fill_b, "Row Numbers (in original)")
        for i, val in enumerate(only_in_b, 2):
            rows = series_b[series_b == val].index.tolist()
            rows_str = ", ".join(str(r + 2) for r in rows[:10])
            if len(rows) > 10:
                rows_str += f" ... (+{len(rows)-10} more)"
            style_cell(ws_b, i, 1, val)
            style_cell(ws_b, i, 2, rows_str)
        ws_b.column_dimensions['A'].width = 40
        ws_b.column_dimensions['B'].width = 50

    # Sheet 4: Common values
    if mode in ("full", "common"):
        ws_common = wb.create_sheet(title="Common Values")
        style_header(ws_common, 1, 1, header_fill_common, "Common to both columns")
        for i, val in enumerate(in_both, 2):
            style_cell(ws_common, i, 1, val)
        ws_common.column_dimensions['A'].width = 40

    wb.save(output)
    print(f"Comparison result saved to: {output}")

    return {
        "only_in_a": only_in_a,
        "only_in_b": only_in_b,
        "common": in_both,
        "output_file": output,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Compare two columns in an Excel file and find differences."
    )
    parser.add_argument("input_file", help="Path to the Excel/CSV file")
    parser.add_argument("column_a", help="Name, letter (A-Z), or index of the first column")
    parser.add_argument("column_b", help="Name, letter (A-Z), or index of the second column")
    parser.add_argument("--sheet", default=None, help="Sheet name or index (default: first sheet)")
    parser.add_argument("--output", default=None, help="Output file path (default: <input>_comparison_result.xlsx)")
    parser.add_argument("--mode", choices=["full", "diff", "unique_a", "unique_b", "common"],
                        default="full", help="Comparison mode (default: full)")

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Error: File not found: {args.input_file}")
        sys.exit(1)

    compare_columns(
        input_file=args.input_file,
        col_a_ref=args.column_a,
        col_b_ref=args.column_b,
        sheet=args.sheet,
        output=args.output,
        mode=args.mode,
    )


if __name__ == "__main__":
    main()
