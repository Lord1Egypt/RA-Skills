#!/usr/bin/env python3
"""
Generate a ShareSDK iOS configuration workbook with per-platform sheets.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


PLATFORMS = [
    {
        "title": "抖音",
        "category": "国内",
        "open_url": "https://www.douyin.com/platform/apply/mobile",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "抖音开放平台 appId"),
            ("appSecret", "", "抖音开放平台 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
            ("universalLink", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "新浪微博",
        "category": "国内",
        "open_url": "http://open.weibo.com",
        "apply_tips": "链接；微博开放平台接入tips",
        "fields": [
            ("appKey", "", "微博开放平台 appKey"),
            ("appSecret", "", "微博开放平台 appSecret"),
            ("redirectUrl", "", "微博开放平台 redirectUrl"),
            ("universalLink", "", "官方示例要求配置"),
        ],
    },
    {
        "title": "QQ",
        "category": "国内",
        "open_url": "http://open.qq.com/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "QQ 开放平台 appId，纯数字"),
            ("appKey", "", "QQ 开放平台 appKey"),
            ("enableUniversalLink", "YES", "YES / NO"),
            ("universalLink", "", "官方示例要求配置"),
        ],
    },
    {
        "title": "微信",
        "category": "国内",
        "open_url": "http://open.weixin.qq.com",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "微信开放平台 appId，以 wx 开头"),
            ("appSecret", "", "微信开放平台 appSecret"),
            ("universalLink", "", "官方示例要求配置"),
            ("useWeChatFull", "", "YES / NO，YES 时使用 WeChatFull"),
        ],
    },
    {
        "title": "企业微信",
        "category": "国内",
        "open_url": "https://work.weixin.qq.com/",
        "apply_tips": "链接",
        "fields": [
            ("corpId", "", "企业微信 corpId"),
            ("agentId", "", "企业微信 agentId"),
            ("secret", "", "企业微信应用 secret"),
            ("schema", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "有道云笔记",
        "category": "国内",
        "open_url": "http://note.youdao.com/open/developguide.html#app",
        "apply_tips": "链接",
        "fields": [
            ("appKey", "", "有道云笔记 appKey"),
            ("appSecret", "", "有道云笔记 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "印象笔记",
        "category": "国内",
        "open_url": "https://dev.evernote.com/",
        "apply_tips": "链接",
        "fields": [
            ("consumerKey", "", "印象笔记 consumerKey"),
            ("consumerSecret", "", "印象笔记 consumerSecret"),
            ("sandbox", "", "YES / NO，是否使用测试环境"),
        ],
    },
    {
        "title": "易信好友",
        "category": "国内",
        "open_url": "http://open.yixin.im/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "易信 appId"),
            ("appSecret", "", "易信 appSecret"),
        ],
    },
    {
        "title": "明道",
        "category": "国内",
        "open_url": "http://open.mingdao.com/",
        "apply_tips": "链接",
        "fields": [
            ("appKey", "", "明道 appKey"),
            ("appSecret", "", "明道 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "人人网",
        "category": "国内",
        "open_url": "http://dev.renren.com",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "人人网 appId"),
            ("appKey", "", "人人网 appKey"),
            ("appSecret", "", "人人网 appSecret"),
        ],
    },
    {
        "title": "开心网",
        "category": "国内",
        "open_url": "http://open.kaixin001.com",
        "apply_tips": "链接",
        "fields": [
            ("appKey", "", "开心网 appKey"),
            ("appSecret", "", "开心网 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "钉钉",
        "category": "国内",
        "open_url": "http://open-dev.dingtalk.com/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "钉钉 appId"),
            ("appSecret", "", "钉钉 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "美拍",
        "category": "国内",
        "open_url": "http://open.meipai.com/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "美拍 appId"),
            ("appSecret", "", "美拍 appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "中国移动",
        "category": "国内",
        "open_url": "http://dev.10086.cn/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "中国移动 appId"),
            ("appKey", "", "中国移动 appKey"),
            ("appSecret", "", "中国移动 appSecret"),
        ],
    },
    {
        "title": "中国电信",
        "category": "国内",
        "open_url": "http://id.189.cn/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "中国电信 appId"),
            ("appKey", "", "中国电信 appKey"),
            ("appSecret", "", "中国电信 appSecret"),
        ],
    },
    {
        "title": "快手",
        "category": "国内",
        "open_url": "https://open.kuaishou.com/platform",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "快手 appId"),
            ("appKey", "", "快手 appKey"),
            ("appSecret", "", "快手 appSecret"),
            ("universalLink", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "支付宝",
        "category": "国内",
        "open_url": "https://open.kuaishou.com/platform",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "支付宝 appId"),
            ("pid", "", "支付宝 pid"),
            ("appScheme", "", "支付宝回调 Scheme"),
        ],
    },
    {
        "title": "Line",
        "category": "国外",
        "open_url": "https://developers.line.me/",
        "apply_tips": "链接",
        "fields": [
            ("channelId", "", "Line channelId"),
            ("channelSecret", "", "Line channelSecret"),
            ("universalLink", "", "iOS 13+ 建议校验"),
        ],
    },
    {
        "title": "Facebook",
        "category": "国外",
        "open_url": "https://developers.facebook.com",
        "apply_tips": "iOS链接 / Android链接",
        "fields": [
            ("appId", "", "Facebook appId"),
            ("appSecret", "", "Facebook appSecret"),
            ("displayName", "", "Facebook 应用名"),
        ],
    },
    {
        "title": "Twitter",
        "category": "国外",
        "open_url": "https://dev.twitter.com",
        "apply_tips": "链接",
        "fields": [
            ("consumerKey", "", "Twitter consumerKey"),
            ("consumerSecret", "", "Twitter consumerSecret"),
            ("callbackUrl", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "GooglePlus",
        "category": "国外",
        "open_url": "https://console.developers.google.com/",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "Google clientId"),
            ("clientSecret", "", "Google clientSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "LinkedIn",
        "category": "国外",
        "open_url": "https://www.linkedin.com/secure/developer?newapp=",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "LinkedIn clientId"),
            ("clientSecret", "", "LinkedIn clientSecret"),
            ("redirectUri", "", "LinkedIn redirectUri"),
        ],
    },
    {
        "title": "Flickr",
        "category": "国外",
        "open_url": "http://www.flickr.com/services/",
        "apply_tips": "链接",
        "fields": [
            ("apiKey", "", "Flickr apiKey"),
            ("apiSecret", "", "Flickr apiSecret"),
        ],
    },
    {
        "title": "Pinterest",
        "category": "国外",
        "open_url": "http://developers.pinterest.com/",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "Pinterest clientId"),
            ("appScheme", "", "Pinterest 回调 Scheme"),
        ],
    },
    {
        "title": "Tumblr",
        "category": "国外",
        "open_url": "http://www.tumblr.com/developers",
        "apply_tips": "链接",
        "fields": [
            ("consumerKey", "", "Tumblr consumerKey"),
            ("consumerSecret", "", "Tumblr consumerSecret"),
            ("callbackUrl", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "Dropbox",
        "category": "国外",
        "open_url": "https://www.dropbox.com/developers",
        "apply_tips": "链接",
        "fields": [
            ("appKey", "", "Dropbox appKey"),
            ("appSecret", "", "Dropbox appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "Instagram",
        "category": "国外",
        "open_url": "https://developers.facebook.com/",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "Instagram / Meta appId"),
            ("appSecret", "", "Instagram / Meta appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "VKontakte",
        "category": "国外",
        "open_url": "http://vk.com/dev",
        "apply_tips": "链接",
        "fields": [
            ("appId", "", "VKontakte appId"),
            ("appSecret", "", "VKontakte appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "Kakao",
        "category": "国外",
        "open_url": "https://developers.kakao.com/",
        "apply_tips": "链接",
        "fields": [
            ("appKey", "", "Kakao appKey"),
            ("appSecret", "", "Kakao appSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "YouTube",
        "category": "国外",
        "open_url": "https://console.developers.google.com",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "YouTube / Google clientId"),
            ("clientSecret", "", "YouTube / Google clientSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
    {
        "title": "Telegram",
        "category": "国外",
        "open_url": "可到App Store下载 Telegram Messenger",
        "apply_tips": "链接",
        "fields": [
            ("botName", "", "如业务接入要求则填写"),
            ("appId", "", "如平台提供则填写"),
            ("appSecret", "", "如平台提供则填写"),
        ],
    },
    {
        "title": "Reddit",
        "category": "国外",
        "open_url": "https://www.reddit.com/prefs/apps",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "Reddit clientId"),
            ("clientSecret", "", "Reddit clientSecret"),
            ("redirectUri", "", "Reddit redirectUri"),
        ],
    },
    {
        "title": "TikTok",
        "category": "国外",
        "open_url": "https://developers.tiktok.com/?refer=tiktok_web",
        "apply_tips": "链接",
        "fields": [
            ("clientKey", "", "TikTok clientKey"),
            ("clientSecret", "", "TikTok clientSecret"),
            ("redirectUri", "", "TikTok redirectUri"),
        ],
    },
    {
        "title": "SnapChat",
        "category": "国外",
        "open_url": "https://developers.tiktok.com/?refer=tiktok_web",
        "apply_tips": "链接",
        "fields": [
            ("clientId", "", "SnapChat clientId"),
            ("clientSecret", "", "SnapChat clientSecret"),
            ("redirectUri", "", "如平台要求则填写"),
        ],
    },
]


def style_header(ws, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_rows(ws, rows):
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def create_sheet(wb, title, headers, rows, widths):
    ws = wb.create_sheet(title=title)
    style_header(ws, headers)
    write_rows(ws, rows)
    set_widths(ws, widths)
    ws.freeze_panes = "A2"
    return ws


def build_platform_rows():
    rows = []
    for platform in PLATFORMS:
        rows.append(
            [
                platform["category"],
                platform["title"],
                platform["open_url"],
                platform["apply_tips"],
            ]
        )
    return rows


def create_platform_sheet(wb, platform):
    rows = list(platform["fields"])
    create_sheet(
        wb,
        platform["title"],
        ["字段", "值", "说明"],
        rows,
        {"A": 24, "B": 34, "C": 72},
    )


def main():
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    create_sheet(
        wb,
        "基础信息",
        ["字段", "值", "说明"],
        [
            ["mobAppKey", "", "MobTech 后台申请得到的 AppKey"],
            ["mobAppSecret", "", "MobTech 后台申请得到的 AppSecret"],
            ["needShareUI", "", "YES / NO，是否需要官方分享面板 UI"],
            ["needAuth", "", "YES / NO，是否需要第三方登录授权"],
        ],
        {"A": 24, "B": 34, "C": 64},
    )

    create_sheet(
        wb,
        "平台申请地址",
        ["平台类别", "平台", "开放平台地址", "APPkey 申请流程"],
        build_platform_rows(),
        {"A": 14, "B": 20, "C": 64, "D": 28},
    )

    for platform in PLATFORMS:
        create_platform_sheet(wb, platform)

    output_path = Path(__file__).resolve().parent / "ShareSDK_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
