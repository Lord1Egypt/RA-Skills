#!/usr/bin/env python3
"""从Excel重新构建raw_records表 + brand_audit表 + monthly_audit表
每次只处理一个月，关闭文件后强制GC，避免OOM。
"""
import sqlite3, os, glob, re, gc, json

DATA_DIR = '/workspace/data'
DB_PATH = '/workspace/data/ka_commission_audit.db'
SUMMARY_JSON = '/tmp/month_summaries.json'

# ===== 工具函数 =====
def S(v): return str(v).strip() if v is not None else ''
def F(v):
    try: return float(v) if v is not None else 0
    except: return 0

def find_hdr(ws):
    """找表头行"""
    for r in range(1, min(ws.max_row+1, 30)):
        row_vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column+1)]
        txt = '|'.join(v.lower() for v in row_vals)
        if '品牌' in txt or '返佣合计' in txt or '支付宝交易金额' in txt:
            return r
    return 1

def parse_std(ws, hr, hdrs):
    """解析标准数据sheet，返回记录列表"""
    from collections import OrderedDict
    # 查找各列索引
    idx = {}
    for i, h in enumerate(hdrs):
        hl = h.lower()
        if '品牌' in hl: idx['brand'] = i
        elif '收钱吧商户名' in hl: idx['merchant'] = i
        elif '收钱吧商户号' in hl: idx['merchant_no'] = i
        elif '支付宝交易金额' in hl or '支付宝交易额' in hl: idx['ali_txn'] = i
        elif '支付宝返佣比例' in hl: idx['ali_ratio'] = i
        elif '支付宝返佣金额' in hl: idx['ali_amt'] = i
        elif '微信交易金额' in hl or '微信交易额' in hl: idx['wx_txn'] = i
        elif '微信返佣比例' in hl: idx['wx_ratio'] = i
        elif '微信返佣金额' in hl: idx['wx_amt'] = i
        elif '返佣合计' in hl: idx['total'] = i
        elif '交易期日' in hl or '交易期' in hl: idx['occ_date'] = i
        elif '备注' in hl and '品牌' not in hl: idx['note'] = i
        
        # level fields
        if not idx.get('level1') and ('一级组织' in hl or '一级' in hl.lower()):
            idx['level1'] = i
        if not idx.get('level2') and ('二级组织' in hl or '二级' in hl.lower()):
            idx['level2'] = i
        if not idx.get('level3') and ('三级组织' in hl or '三级' in hl.lower()):
            idx['level3'] = i
    
    # 支付宝间连返佣列
    for i, h in enumerate(hdrs):
        hl = h.lower()
        if '支付宝间连返佣' in hl or '间连支付宝' in hl or '支付宝间连' in hl:
            idx['ali_indirect'] = i
            break
    
    if 'brand' not in idx:
        # 尝试用商户名做品牌
        if 'merchant' in idx:
            idx['brand'] = idx['merchant']
        else:
            return []
    
    records = []
    brand_col = idx['brand']
    
    for r in range(hr+1, ws.max_row+1):
        b = S(ws.cell(row=r, column=brand_col+1).value)
        if not b:
            continue
        # 跳过汇总行（金额列有汇总字样或品牌列就是'品牌'）
        if b.lower() in ('品牌', '品牌名称', '合计', '汇总', '商户号', ''):
            continue
        
        ali_txn = F(ws.cell(row=r, column=idx.get('ali_txn', -1)+1).value) if 'ali_txn' in idx else 0
        ali_ratio = F(ws.cell(row=r, column=idx.get('ali_ratio', -1)+1).value) if 'ali_ratio' in idx else 0
        ali_amt = F(ws.cell(row=r, column=idx.get('ali_amt', -1)+1).value) if 'ali_amt' in idx else 0
        wx_txn = F(ws.cell(row=r, column=idx.get('wx_txn', -1)+1).value) if 'wx_txn' in idx else 0
        wx_ratio = F(ws.cell(row=r, column=idx.get('wx_ratio', -1)+1).value) if 'wx_ratio' in idx else 0
        wx_amt = F(ws.cell(row=r, column=idx.get('wx_amt', -1)+1).value) if 'wx_amt' in idx else 0
        total = F(ws.cell(row=r, column=idx.get('total', -1)+1).value) if 'total' in idx else 0
        occ = S(ws.cell(row=r, column=idx.get('occ_date', -1)+1).value) if 'occ_date' in idx else ''
        note = S(ws.cell(row=r, column=idx.get('note', -1)+1).value) if 'note' in idx else ''
        
        ali_ind = F(ws.cell(row=r, column=idx.get('ali_indirect', -1)+1).value) if 'ali_indirect' in idx else 0
        
        l1 = S(ws.cell(row=r, column=idx.get('level1', -1)+1).value) if 'level1' in idx else ''
        l2 = S(ws.cell(row=r, column=idx.get('level2', -1)+1).value) if 'level2' in idx else ''
        l3 = S(ws.cell(row=r, column=idx.get('level3', -1)+1).value) if 'level3' in idx else ''
        
        records.append({
            'brand': b,
            'ali_txn': ali_txn, 'ali_ratio': ali_ratio, 'ali_amt': ali_amt,
            'wx_txn': wx_txn, 'wx_ratio': wx_ratio, 'wx_amt': wx_amt,
            'total_rebate': total,
            'ali_indirect': ali_ind,
            'occurrence_date': occ,
            'original_note': note,
            'level1': l1, 'level2': l2, 'level3': l3,
        })
    
    return records

def parse_summary(ws):
    """解析汇总表，返回 {merchant_no: amount}"""
    data = {}
    headers = [str(ws.cell(row=1, column=c).value or '') for c in range(1, ws.max_column+1)]
    
    # Find merchant_no column (column B, index 1) and amount column (column I, index 8)
    mn_col = None; amt_col = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if '商户号' in hl: mn_col = i
        elif '应返金额' in hl or '入账' in hl: amt_col = i
    
    if mn_col is None and amt_col is None:
        # Default: col B (index 1) = merchant_no, col I (index 8) = amount
        mn_col, amt_col = 1, 8
    
    for r in range(2, ws.max_row+1):
        mn = S(ws.cell(row=r, column=mn_col+1).value)
        amt = F(ws.cell(row=r, column=amt_col+1).value)
        if mn and len(mn) >= 13:  # 商户号至少13位
            data[mn] = amt
    
    return data

def process_month(month):
    """处理单个月份，返回数据"""
    # 找Excel文件
    found = None
    for p in [os.path.join(DATA_DIR, month, f'KA返佣{month}*.xlsx'),
              os.path.join(DATA_DIR, month, f'KA{month}*.xlsx')]:
        fs = glob.glob(p)
        if fs: found = fs[0]; break
    
    if not found:
        return None, None, f'文件未找到'
    
    import openpyxl
    wb = openpyxl.load_workbook(found, data_only=True)
    snames = wb.sheetnames
    
    summary = {}
    if '汇总' in snames:
        summary = parse_summary(wb['汇总'])
    
    # 解析数据sheet
    raw_records = []
    
    for sn in snames:
        if sn == '汇总' or sn.strip().lower() == 'sql':
            continue
        ws = wb[sn]
        hr = find_hdr(ws)
        hdrs = [str(ws.cell(row=hr, column=c).value or '') for c in range(1, ws.max_column+1)]
        txt = '|'.join(v.lower() for v in hdrs)
        
        sheet_type = 'regular'
        if '问题' in sn or '异常' in sn:
            sheet_type = 'abnormal'
        elif '中快' in sn:
            sheet_type = 'zhongkuai'
        elif '志华' in sn or '拓展' in sn:
            sheet_type = 'expansion'
        elif '分润汇总' in txt or 'bill_email' in txt:
            # 英文sheet
            hdrs_en = [str(ws.cell(row=hr, column=c).value or '') for c in range(1, ws.max_column+1)]
            en_txt = '|'.join(v.lower() for v in hdrs_en)
            if 'bill_email' in en_txt or 'email' in en_txt:
                sheet_type = 'english'
        
        records = parse_std(ws, hr, hdrs)
        for rec in records:
            rec['sheet_name'] = sn
            rec['sheet_type'] = sheet_type
        raw_records.extend(records)
    
    wb.close()
    import gc as _gc
    _gc.collect()
    
    return raw_records, summary, None

# ===== 主流程 =====
def main():
    all_months = sorted([d for d in os.listdir(DATA_DIR) if d.isdigit() and len(d) == 6])
    
    # 创建/重建SQLite
    conn = sqlite3.connect(DB_PATH, timeout=120)
    conn.execute("PRAGMA journal_mode=WAL")
    
    # 删除旧表
    for t in ['raw_records', 'brand_audit', 'monthly_audit']:
        conn.execute(f'DROP TABLE IF EXISTS {t}')
    
    # 创建表
    conn.execute("""CREATE TABLE raw_records(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        month TEXT, brand TEXT, calc_method TEXT, sheet_name TEXT,
        alipay_txn_amount REAL, alipay_rebate_ratio REAL,
        wechat_txn_amount REAL, wechat_rebate_ratio REAL,
        total_rebate REAL, audit_note TEXT, original_note TEXT,
        occurrence_date TEXT, level1_name TEXT, level2_name TEXT, level3_name TEXT
    )""")
    conn.execute("""CREATE TABLE brand_audit(
        month TEXT, brand TEXT,
        raw_amount REAL, recalc_amount REAL, summary_amount REAL,
        diff REAL, diff_note TEXT, is_real_diff INTEGER
    )""")
    conn.execute("""CREATE TABLE monthly_audit(
        month TEXT PRIMARY KEY,
        total_recalc REAL, total_summary REAL,
        total_diff REAL, real_diff REAL, brand_count INTEGER
    )""")
    conn.commit()
    
    print(f"处理 {len(all_months)} 个月份...")
    
    for i, month in enumerate(all_months):
        print(f"  [{i+1}/{len(all_months)}] {month}...", end=' ', flush=True)
        
        raw_records, summary, err = process_month(month)
        if err:
            print(f'❌ {err}')
            continue
        
        if not raw_records:
            print(f'⚠️ 无数据记录')
            continue
        
        # 插入raw_records
        insert_sql = """INSERT INTO raw_records(
            month, brand, calc_method, sheet_name,
            alipay_txn_amount, alipay_rebate_ratio,
            wechat_txn_amount, wechat_rebate_ratio,
            total_rebate, audit_note, original_note,
            occurrence_date, level1_name, level2_name, level3_name
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
        
        batch = []
        for rec in raw_records:
            sheet_type = rec.get('sheet_type', 'regular')
            
            # 计算calc_method
            calc = sheet_type
            if sheet_type == 'expansion':
                if month == '202307' and '拓展' in rec.get('sheet_name', ''):
                    calc = 'expansion_307'
                elif '志华' in rec.get('sheet_name', ''):
                    calc = 'expansion_308'
                else:
                    calc = 'expansion_307'
            elif sheet_type == 'regular':
                calc = 'regular'
            elif sheet_type == 'zhongkuai':
                calc = 'zhongkuai'
            elif sheet_type == 'abnormal':
                calc = 'abnormal'
            
            # audit_note
            audit_note = ''
            orig_note = rec.get('original_note', '')
            if '调账' in orig_note or '调差' in orig_note:
                audit_note = '历史差额调整'
            elif sheet_type == 'abnormal':
                audit_note = '有问题，暂时不调整'
            
            batch.append((
                month, rec['brand'], calc, rec.get('sheet_name', ''),
                rec['ali_txn'], rec['ali_ratio'],
                rec['wx_txn'], rec['wx_ratio'],
                rec['total_rebate'], audit_note, orig_note,
                rec.get('occurrence_date', ''),
                rec.get('level1', ''), rec.get('level2', ''), rec.get('level3', '')
            ))
        
        conn.executemany(insert_sql, batch)
        
        # 构建品牌级汇总表数据（从per-merchant汇总表匹配到品牌）
        if summary:
            # 将商户号→品牌映射：从raw_records中找每个商户号对应的品牌
            merchant_brand = {}
            c = conn.cursor()
            for mn in summary.keys():
                # 查找此商户在哪个品牌下
                rows = c.execute(
                    "SELECT DISTINCT brand FROM raw_records WHERE month=? AND (alipay_txn_amount + wechat_txn_amount > 0) LIMIT 1000",
                    (month,)).fetchall()
                if rows:
                    merchant_brand[mn] = rows[0][0]
                else:
                    merchant_brand[mn] = '未知'
            
            # 按品牌汇总
            brand_summary = {}
            for mn, amt in summary.items():
                if mn in merchant_brand:
                    b = merchant_brand[mn]
                    brand_summary[b] = brand_summary.get(b, 0) + amt
            
            # 计算品牌级audit数据
            from collections import defaultdict
            brand_recalc = defaultdict(float)
            brand_raw = defaultdict(float)
            
            for rec in raw_records:
                b = rec['brand']
                brand_raw[b] += rec['total_rebate']
                
                if rec.get('design_type', '') == 'recalc':
                    brand_recalc[b] += round(rec['ali_txn'] * rec['ali_ratio'], 2) + round(rec['wx_txn'] * rec['wx_ratio'], 2)
                elif sheet_type == 'zhongkuai':
                    brand_recalc[b] += rec['total_rebate']
                else:
                    brand_recalc[b] += rec['total_rebate']
            
            all_brands = set(list(brand_recalc.keys()) + list(brand_summary.keys()))
            for b in sorted(all_brands):
                recalc = round(brand_recalc.get(b, 0), 2)
                sm = round(brand_summary.get(b, 0), 2)
                ra = round(brand_raw.get(b, 0), 2)
                diff = round(sm - recalc, 2)
                if diff < 0: is_real, note = 0, '少发，视同无差异'
                elif abs(diff) < 1: is_real, note = 0, '小数舍入'
                else: is_real, note = 1, ''
                conn.execute("""INSERT INTO brand_audit(month,brand,raw_amount,recalc_amount,summary_amount,diff,diff_note,is_real_diff)
                    VALUES(?,?,?,?,?,?,?,?)""", (month,b,ra,recalc,sm,diff,note,is_real))
        
        conn.commit()
        print(f'✅ {len(raw_records)}条')
        
        # 清理
        del raw_records, summary, batch
        gc.collect()
    
    # ===== 生成monthly_audit =====
    print("\n生成月度汇总...")
    months_in_db = sorted(set(r[0] for r in conn.execute("SELECT DISTINCT month FROM brand_audit").fetchall()))
    for month in months_in_db:
        rows = conn.execute("SELECT recalc_amount, summary_amount, is_real_diff, diff FROM brand_audit WHERE month=?", (month,)).fetchall()
        tr = round(sum(r[0] for r in rows), 2)
        ts = round(sum(r[1] for r in rows), 2)
        rd = round(sum(r[3] for r in rows if r[2]), 2)
        bc = len(rows)
        conn.execute("""INSERT OR REPLACE INTO monthly_audit(month,total_recalc,total_summary,total_diff,real_diff,brand_count)
            VALUES(?,?,?,?,?,?)""", (month, tr, ts, round(ts-tr, 2), rd, bc))
    
    conn.commit()
    
    # 统计
    ma = list(conn.execute("SELECT * FROM monthly_audit ORDER BY month").fetchall())
    print(f"\n完成! {len(ma)}个月份")
    total_rd = sum(r[4] for r in ma)
    perfect = sum(1 for r in ma if r[4] == 0)
    print(f"完美月份: {perfect}/{len(ma)}")
    print(f"真实差异合计: {total_rd:+,.2f}")
    
    conn.close()
    print("Done")

if __name__ == '__main__':
    main()
