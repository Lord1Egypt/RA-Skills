#!/usr/bin/env python3
"""
配置模板Excel解析脚本 v1.0
自动解析包含5个子表的配置模板Excel文件，并写入SQLite数据库
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime

def parse_config_excel(file_path, db_path=None):
    """
    解析配置模板Excel文件
    
    Args:
        file_path: Excel文件路径
        db_path: SQLite数据库路径（可选，默认使用项目数据库）
    
    Returns:
        dict: 解析结果 {'success': bool, 'records': dict, 'errors': list}
    """
    try:
        import openpyxl
    except ImportError:
        return {'success': False, 'error': '缺少 openpyxl 库，请运行：pip install openpyxl'}
    
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        return {'success': False, 'error': f'读取Excel失败：{e}'}
    
    results = {
        'success': True,
        'records': {},
        'errors': []
    }
    
    # 定义子表名称到表格类型的映射
    sheet_mapping = {
        '房源清单': '房源销控表',
        '客户跟进表': '客户跟进记录',
        '渠道跟进表': '渠道跟进记录',
        '竞品情报库': '竞品情报库',
        '项目推介材料': '项目推介材料'
    }
    
    # 逐个解析子表
    for sheet_name, table_name in sheet_mapping.items():
        if sheet_name not in wb.sheetnames:
            results['errors'].append(f'⚠️ 未找到子表：{sheet_name}')
            continue
        
        ws = wb[sheet_name]
        
        # 读取表头（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        # 读取数据行
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳过空行
                continue
            
            record = {}
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    record[headers[idx]] = cell if cell is not None else ''
            
            records.append(record)
        
        results['records'][table_name] = {
            'count': len(records),
            'data': records
        }
    
    return results

def preview_results(results):
    """预览解析结果"""
    if not results['success']:
        print(f"❌ 解析失败：{results.get('error', '未知错误')}")
        return
    
    print("\n📊 解析结果预览：")
    print("=" * 60)
    
    for table_name, data in results['records'].items():
        print(f"\n【{table_name}】 共 {data['count']} 条记录")
        if data['count'] > 0:
            print("  预览前3条：")
            for i, record in enumerate(data['data'][:3]):
                print(f"  {i+1}. {record}")
    
    if results['errors']:
        print("\n⚠️ 警告：")
        for err in results['errors']:
            print(f"  {err}")

def save_to_sqlite(results, db_path='~/.workbuddy/workspace/investment-assistant/local_db.sqlite'):
    """
    将解析结果写入SQLite数据库
    
    Args:
        results: 解析结果
        db_path: 数据库路径
    """
    import sqlite3
    from pathlib import Path
    
    db_path = Path(db_path).expanduser()
    
    # 确保数据库目录存在
    db_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 连接数据库
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    
    # 创建表（如果不存在）
    tables_schema = {
        '房源销控表': '''
            CREATE TABLE IF NOT EXISTS `房源销控表` (
                `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                `房号` TEXT,
                `楼层` TEXT,
                `面积(㎡)` REAL,
                `租金单价(元/㎡/天)` REAL,
                `物业费(元/㎡/月)` REAL,
                `总价(元/月)` REAL,
                `状态` TEXT,
                `租客名称` TEXT,
                `租期到期日` TEXT,
                `装修标准` TEXT,
                `朝向` TEXT,
                `备注` TEXT,
                `入库时间` TEXT
            )
        ''',
        '客户跟进记录': '''
            CREATE TABLE IF NOT EXISTS `客户跟进记录` (
                `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                `客户ID` TEXT,
                `客户名称` TEXT,
                `联系人` TEXT,
                `联系电话` TEXT,
                `需求面积` TEXT,
                `需求楼层` TEXT,
                `预算租金` TEXT,
                `跟进状态` TEXT,
                `上次跟进时间` TEXT,
                `下次跟进时间` TEXT,
                `需求描述` TEXT,
                `备注` TEXT,
                `入库时间` TEXT
            )
        ''',
        '渠道跟进记录': '''
            CREATE TABLE IF NOT EXISTS `渠道跟进记录` (
                `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                `渠道ID` TEXT,
                `渠道名称` TEXT,
                `对接人` TEXT,
                `联系电话` TEXT,
                `推荐客户数` INTEGER,
                `成交数` INTEGER,
                `佣金比例` TEXT,
                `合作状态` TEXT,
                `上次沟通时间` TEXT,
                `备注` TEXT,
                `入库时间` TEXT
            )
        ''',
        '竞品情报库': '''
            CREATE TABLE IF NOT EXISTS `竞品情报库` (
                `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                `竞品ID` TEXT,
                `项目名称` TEXT,
                `位置` TEXT,
                `租金单价(元/㎡/天)` REAL,
                `物业费(元/㎡/月)` REAL,
                `免租期` TEXT,
                `装修补贴` TEXT,
                `空置率` TEXT,
                `主要租户` TEXT,
                `备注` TEXT,
                `入库时间` TEXT
            )
        ''',
        '项目推介材料': '''
            CREATE TABLE IF NOT EXISTS `项目推介材料` (
                `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                `项目ID` TEXT,
                `项目名称` TEXT,
                `位置` TEXT,
                `总建筑面积(㎡)` REAL,
                `可租面积(㎡)` REAL,
                `楼层数` INTEGER,
                `配套设施` TEXT,
                `交通描述` TEXT,
                `项目亮点` TEXT,
                `备注` TEXT,
                `入库时间` TEXT
            )
        '''
    }
    
    # 创建所有表
    for table_name, create_sql in tables_schema.items():
        try:
            cursor.execute(create_sql)
        except Exception as e:
            print(f"⚠️ 创建表 {table_name} 失败：{e}")
    
    # 插入数据
    inserted_count = 0
    
    for table_name, data in results['records'].items():
        if data['count'] == 0:
            continue
        
        # 清空原有数据（可选，这里选择追加）
        # cursor.execute(f'DELETE FROM {table_name}')
        
        # 插入新数据
        for record in data['data']:
            # 构建INSERT语句（列名需要用反引号括起来）
            columns = []
            placeholders = []
            values = []
            
            for key, val in record.items():
                columns.append(key)
                placeholders.append('?')
                values.append(val)
            
            # 添加入库时间
            columns.append('入库时间')
            placeholders.append('?')
            values.append(datetime.now().isoformat())
            
            # 构建SQL语句，列名用反引号括起来
            columns_sql = ', '.join([f'`{c}`' for c in columns])
            placeholders_sql = ', '.join(placeholders)
            sql = f"INSERT INTO `{table_name}` ({columns_sql}) VALUES ({placeholders_sql})"
            
            try:
                cursor.execute(sql, values)
                inserted_count += 1
            except Exception as e:
                print(f"⚠️ 插入记录失败：{e}")
                print(f"   SQL：{sql}")
                print(f"   记录：{record}")
    
    conn.commit()
    conn.close()
    
    return inserted_count

def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法：python3 parse_config_excel.py <Excel文件路径> [--db path] [--preview]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        sys.exit(1)
    
    print(f"📂 解析配置模板：{file_path}")
    
    # 解析Excel
    results = parse_config_excel(file_path)
    
    # 预览结果
    preview_results(results)
    
    if not results['success']:
        sys.exit(1)
    
    # 询问是否入库
    if '--preview' not in sys.argv:
        confirm = input("\n确认入库吗？(y/n)：")
        if confirm.lower() != 'y':
            print("已取消入库")
            sys.exit(0)
    
    # 保存到SQLite
    db_path = '~/.workbuddy/workspace/investment-assistant/local_db.sqlite'
    if '--db' in sys.argv:
        idx = sys.argv.index('--db')
        if idx + 1 < len(sys.argv):
            db_path = sys.argv[idx + 1]
    
    print(f"\n💾 写入数据库：{db_path}")
    inserted = save_to_sqlite(results, db_path)
    
    print(f"\n✅ 入库完成！共写入 {inserted} 条记录")
    print(f"   数据库路径：{db_path}")

if __name__ == '__main__':
    main()
