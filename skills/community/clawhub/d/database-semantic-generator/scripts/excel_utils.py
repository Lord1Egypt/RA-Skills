#!/usr/bin/env python3
"""Excel helpers for generate_yaml_semantic skill."""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

import requests
from openpyxl import Workbook, load_workbook


def list_excel_sheets(excel_file: str) -> List[str]:
    """Return sorted sheet names from an Excel file."""
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
    try:
        return sorted(workbook.sheetnames)
    finally:
        workbook.close()


def create_subset_excel(excel_file: str, selected_sheets: List[str]) -> Tuple[str, List[str]]:
    """
    Create a temporary Excel file that keeps only selected sheets.

    Returns:
        (temp_file_path, missing_sheets)
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    source = load_workbook(filename=excel_file, read_only=True, data_only=True)
    source_sheets = set(source.sheetnames)
    missing = [name for name in selected_sheets if name not in source_sheets]
    valid = [name for name in selected_sheets if name in source_sheets]

    if not valid:
        source.close()
        return "", missing

    target = Workbook()
    default_sheet = target.active
    target.remove(default_sheet)

    for sheet_name in valid:
        ws_src = source[sheet_name]
        ws_dst = target.create_sheet(title=sheet_name)
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row) if row else [])

    original_stem = Path(excel_file).stem
    fd, tmp_path = tempfile.mkstemp(prefix=f"{original_stem}_subset_", suffix=".xlsx")
    os.close(fd)
    target.save(tmp_path)

    source.close()
    target.close()
    return tmp_path, missing


def upload_excel_for_knowledge(
    excel_file_path: str,
    api_url: str = "https://asksql.ai/ask/api/generate_database_knowledge",
    timeout: int = 30,
) -> Dict[str, Any]:
    """
    Upload excel to async API and parse `content` to dict for in-memory processing.
    """
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Temp excel file not found: {excel_file_path}")

    with open(excel_file_path, "rb") as f:
        files = {"excel_file": (os.path.basename(excel_file_path), f)}
        data = {"format": "json", "generate_knowledge": "true"}
        response = requests.post(api_url, files=files, data=data, timeout=timeout)

    return _parse_knowledge_response(response)


def upload_database_for_knowledge(
    table_data: Dict[str, Any],
    api_url: str = "https://asksql.ai/ask/api/generate_database_knowledge",
    timeout: int = 30,
) -> Dict[str, Any]:
    """
    Upload database table payload to generate_database_knowledge API.
    """
    print(table_data)
    response = requests.post(
        api_url,
        json={"data": table_data, "format": "json"},
        timeout=timeout,
    )
    # print(response.text)
    return _parse_knowledge_response(response)


def _parse_knowledge_response(response: requests.Response) -> Dict[str, Any]:
    if response.status_code != 200:
        raise RuntimeError(f"Upstream error status={response.status_code}, body={response.text[:400]}")

    payload = response.json()
    if not isinstance(payload, dict):
        raise RuntimeError("Invalid upstream payload: expected dict")
    if not payload.get("success"):
        raise RuntimeError(f"Upstream returned success=false, message={payload.get('message', '')}")

    content = payload.get("content", {})
    if isinstance(content, str):
        try:
            content = json.loads(content)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"Failed to parse upstream content JSON: {exc}") from exc
    if not isinstance(content, dict):
        raise RuntimeError("Invalid upstream content: expected dict")
    # print(content)
    return {"raw": payload, "content_data": content}
