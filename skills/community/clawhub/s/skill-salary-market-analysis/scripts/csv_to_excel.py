#!/usr/bin/env python3
"""
CSV 转 Excel 工具
================

功能：
1. 将清洗后的 CSV 数据转换为格式化的 Excel 文件
2. 自动调整列宽
3. 添加冻结窗格和筛选
4. 可选添加数据验证

使用方法：
    python3 csv_to_excel.py -i clean_data.csv -o report.xlsx
"""

import argparse
import csv
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("❌ openpyxl 未安装，请运行：pip install openpyxl")
    sys.exit(1)


class CSVToExcel:
    """CSV 转 Excel 转换器"""

    # 列宽配置（16 字段模板）
    COLUMN_WIDTHS = {
        'id': 6,
        'position': 18,
        'company_name': 25,
        'company_type': 12,
        'salary_range': 20,
        'salary_monthly_low': 15,
        'salary_monthly_high': 15,
        'months': 8,
        'annual_salary_low': 15,
        'annual_salary_high': 15,
        'annual_salary_avg': 15,
        'location': 12,
        'experience': 12,
        'education': 10,
        'source': 12,
        'collect_date': 12
    }

    # 表头样式
    HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据样式
    DATA_FONT = Font(name='微软雅黑', size=10)
    DATA_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
    ALT_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # 边框
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file

    def convert(self):
        """执行转换"""
        # 读取 CSV
        with open(self.input_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            rows = list(reader)

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "薪酬数据"

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

            # 设置列宽
            col_letter = get_column_letter(col_idx)
            width = self.COLUMN_WIDTHS.get(header, 15)
            ws.column_dimensions[col_letter].width = width

        # 写入数据
        for row_idx, row in enumerate(rows, 2):
            is_alt = row_idx % 2 == 0
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGNMENT
                cell.border = self.THIN_BORDER
                if is_alt:
                    cell.fill = self.ALT_FILL

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 添加自动筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        # 保存
        wb.save(self.output_file)
        print(f"✅ Excel 文件已保存：{self.output_file}")
        print(f"   - {len(rows)} 行数据")
        print(f"   - {len(headers)} 列")


def main():
    parser = argparse.ArgumentParser(description='CSV 转 Excel 工具')
    parser.add_argument('-i', '--input', required=True, help='输入 CSV 文件')
    parser.add_argument('-o', '--output', required=True, help='输出 Excel 文件')

    args = parser.parse_args()

    converter = CSVToExcel(args.input, args.output)
    converter.convert()


if __name__ == '__main__':
    main()
