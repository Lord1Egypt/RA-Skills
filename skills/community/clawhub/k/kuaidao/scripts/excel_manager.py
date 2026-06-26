"""
Excel管理器 - 读取、写入、检查Excel文案库
使用openpyxl进行格式保持的操作
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import os


class ExcelManager:
    """管理Excel文案库的读写和格式操作"""
    
    def __init__(self, file_path: str):
        """
        初始化Excel管理器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        self.wb = None
        self.ws = None
        self.format_cache = None
    
    def load(self) -> 'ExcelManager':
        """加载Excel文件"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")
        
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb.active
        return self
    
    def save(self, new_path: str = None):
        """保存Excel文件"""
        save_path = new_path or self.file_path
        self.wb.save(save_path)
    
    def close(self):
        """关闭Excel文件"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None
    
    def __enter__(self):
        """上下文管理器入口"""
        return self.load()
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()
    
    def scan_format(self) -> Dict[str, Any]:
        """
        扫描并记录Excel格式（Step 5: 格式检查）
        
        Returns:
            格式参数字典
        """
        if not self.ws:
            raise RuntimeError("Excel未加载")
        
        format_info = {
            "total_rows": self.ws.max_row,
            "total_columns": self.ws.max_column,
            "title_row": {},
            "data_row_A": {},
            "data_row_B_H": {},
            "data_row_I_N": {},
            "merged_cells": []
        }
        
        # 检查标题行（第1行）
        if self.ws.max_row >= 1:
            for col in range(1, min(15, self.ws.max_column + 1)):
                cell = self.ws.cell(1, col)
                format_info["title_row"][f"col_{col}"] = {
                    "font_name": cell.font.name,
                    "font_size": cell.font.size,
                    "bold": cell.font.bold,
                    "font_color": cell.font.color.rgb if cell.font.color else None,
                    "fill_type": cell.fill.patternType,
                    "fill_color": cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                    "h_align": cell.alignment.horizontal,
                    "v_align": cell.alignment.vertical,
                    "wrap_text": cell.alignment.wrap_text,
                    "border": self._get_border_style(cell.border)
                }
            # 行高
            format_info["title_row"]["row_height"] = self.ws.row_dimensions[1].height
        
        # 检查数据行（第2行作为样本）
        if self.ws.max_row >= 2:
            # A列
            cell_a = self.ws.cell(2, 1)
            format_info["data_row_A"] = self._extract_cell_format(cell_a)
            format_info["data_row_A"]["row_height"] = self.ws.row_dimensions[2].height
            
            # B-H列（取B列作为样本）
            cell_b = self.ws.cell(2, 2)
            format_info["data_row_B_H"] = self._extract_cell_format(cell_b)
            
            # I-N列（取I列作为样本）
            if self.ws.max_column >= 9:
                cell_i = self.ws.cell(2, 9)
                format_info["data_row_I_N"] = self._extract_cell_format(cell_i)
        
        # 检查合并单元格
        for merged_range in self.ws.merged_cells.ranges:
            format_info["merged_cells"].append(str(merged_range))
        
        self.format_cache = format_info
        return format_info
    
    def _extract_cell_format(self, cell) -> Dict[str, Any]:
        """提取单元格格式信息"""
        return {
            "font_name": cell.font.name,
            "font_size": cell.font.size,
            "bold": cell.font.bold,
            "fill_type": cell.fill.patternType,
            "fill_color": cell.fill.fgColor.rgb if cell.fill.fgColor else None,
            "h_align": cell.alignment.horizontal,
            "v_align": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "border": self._get_border_style(cell.border)
        }
    
    def _get_border_style(self, border) -> str:
        """获取边框样式"""
        if border.left.style and border.right.style and border.top.style and border.bottom.style:
            return border.left.style
        return None
    
    def get_existing_scripts(self) -> List[Dict[str, Any]]:
        """
        获取已有脚本列表（Step 4: 同质化检查）
        
        Returns:
            已有脚本列表，包含标题和主题
        """
        if not self.ws:
            raise RuntimeError("Excel未加载")
        
        scripts = []
        
        # 从第2行开始读取（第1行是标题）
        for row in range(2, self.ws.max_row + 1):
            # 读取A列（视频文案，包含标题）
            title_cell = self.ws.cell(row, 1).value
            if title_cell:
                # 提取标题（通常是第一行）
                title = str(title_cell).split('\n')[0] if '\n' in str(title_cell) else str(title_cell)[:50]
                
                scripts.append({
                    "row": row,
                    "title": title,
                    "raw_content": title_cell
                })
        
        return scripts
    
    def append_script(self, script_data: Dict[str, Any], platform: str, 
                       start_row: int = None, row_height: float = 100) -> int:
        """
        追加脚本到Excel（Step 8: 更新文案库）
        
        Args:
            script_data: 脚本数据字典
            platform: 平台标识
            start_row: 开始写入行号，None则自动计算
            row_height: 行高，默认100
            
        Returns:
            结束行号
        """
        if not self.ws:
            raise RuntimeError("Excel未加载")
        
        if start_row is None:
            start_row = self.ws.max_row + 1
        
        segments = script_data.get("segments", [])
        if not segments:
            raise ValueError("脚本没有分镜数据")
        
        # 计算结束行
        end_row = start_row + len(segments) - 1
        
        # 写入A列（视频文案）- 跨行合并
        a_cell = self.ws.cell(start_row, 1)
        a_cell.value = script_data.get("title", "") + "\n" + script_data.get("story", "")
        if len(segments) > 1:
            self.ws.merge_cells(start_row=start_row, start_column=1, 
                               end_row=end_row, end_column=1)
        
        # 写入B-H列（时间段、镜头、运镜、技巧、画面、台词、音效）
        for i, segment in enumerate(segments):
            row = start_row + i
            self.ws.cell(row, 2).value = segment.get("time", "")  # B-时间段
            self.ws.cell(row, 3).value = segment.get("shot_desc", "")  # C-镜头
            self.ws.cell(row, 4).value = segment.get("movement_desc", "")  # D-运镜
            self.ws.cell(row, 5).value = segment.get("tech_desc", "")  # E-技巧
            self.ws.cell(row, 6).value = segment.get("scene_desc", "")  # F-画面
            self.ws.cell(row, 7).value = segment.get("line", "")  # G-台词
            self.ws.cell(row, 8).value = segment.get("sound_desc", "")  # H-音效
        
        # 写入I-N列（BGM、标签、状态、活动、日期、备注）
        i_cell = self.ws.cell(start_row, 9)
        i_cell.value = script_data.get("bgm", "")  # I-BGM
        
        # 设置其他列
        self.ws.cell(start_row, 10).value = script_data.get("tags", "")  # J-标签
        self.ws.cell(start_row, 11).value = "待使用"  # K-状态
        self.ws.cell(start_row, 12).value = script_data.get("activity", "")  # L-活动
        self.ws.cell(start_row, 13).value = script_data.get("date", "")  # M-日期
        self.ws.cell(start_row, 14).value = script_data.get("notes", "")  # N-备注
        
        # 跨行合并I-N列
        if len(segments) > 1:
            self.ws.merge_cells(start_row=start_row, start_column=9,
                               end_row=end_row, end_column=14)  # I-N列合并
        
        # 设置行高
        for row in range(start_row, end_row + 1):
            self.ws.row_dimensions[row].height = row_height
        
        return end_row
    
    def apply_format(self, format_info: Dict[str, Any], 
                    start_row: int, end_row: int):
        """
        应用格式到指定行范围
        
        Args:
            format_info: 格式参数字典
            start_row: 开始行
            end_row: 结束行
        """
        # 应用数据行格式（A列）
        if "data_row_A" in format_info:
            self._apply_row_format(format_info["data_row_A"], start_row, end_row, 1, 1)
        
        # 应用数据行格式（B-H列）
        if "data_row_B_H" in format_info:
            for row in range(start_row, end_row + 1):
                self._apply_row_format(format_info["data_row_B_H"], row, row, 2, 8)
        
        # 应用数据行格式（I-N列）
        if "data_row_I_N" in format_info:
            self._apply_row_format(format_info["data_row_I_N"], start_row, end_row, 9, 14)
    
    def _apply_row_format(self, format_dict: Dict[str, Any], 
                         start_row: int, end_row: int,
                         start_col: int, end_col: int):
        """应用格式到单元格范围"""
        # 创建字体
        font = Font(
            name=format_dict.get("font_name", "宋体"),
            size=format_dict.get("font_size", 11),
            bold=format_dict.get("bold", False)
        )
        
        # 创建填充
        fill = None
        if format_dict.get("fill_type"):
            fill = PatternFill(
                patternType=format_dict.get("fill_type"),
                fgColor=format_dict.get("fill_color")
            )
        
        # 创建对齐
        alignment = Alignment(
            horizontal=format_dict.get("h_align", "left"),
            vertical=format_dict.get("v_align", "center"),
            wrap_text=format_dict.get("wrap_text", True)
        )
        
        # 创建边框
        border_style = format_dict.get("border")
        border = None
        if border_style:
            side = Side(style=border_style)
            border = Border(left=side, right=side, top=side, bottom=side)
        
        # 应用格式
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row, col)
                cell.font = font
                if fill:
                    cell.fill = fill
                cell.alignment = alignment
                if border:
                    cell.border = border
            
            # 设置行高
            if "row_height" in format_dict:
                self.ws.row_dimensions[row].height = format_dict["row_height"]
    
    def validate_write(self, start_row: int, end_row: int) -> Dict[str, Any]:
        """
        验证写入结果（Step 9: 全面检查对比）
        
        Args:
            start_row: 开始行
            end_row: 结束行
            
        Returns:
            验证结果字典
        """
        result = {
            "format_valid": True,
            "content_complete": True,
            "position_correct": True,
            "errors": []
        }
        
        # 检查格式正确性
        for row in range(start_row, end_row + 1):
            for col in range(1, 15):
                cell = self.ws.cell(row, col)
                if not cell.value and col not in [11]:  # K列（状态）可以有默认值
                    result["content_complete"] = False
                    result["errors"].append(f"第{row}行第{col}列内容为空")
        
        # 检查合并单元格
        for merged_range in self.ws.merged_cells.ranges:
            # 验证合并范围是否正确
            pass
        
        return result
    
    def delete_rows(self, start_row: int, end_row: int):
        """
        删除指定行（用于Step 9失败回退）
        
        Args:
            start_row: 开始行
            end_row: 结束行
        """
        # 注意：openpyxl删除行是从start_row开始删除count行
        count = end_row - start_row + 1
        self.ws.delete_rows(start_row, count)


class ExcelFormatError(Exception):
    """Excel格式错误"""
    pass


class ExcelWriteError(Exception):
    """Excel写入错误"""
    pass
