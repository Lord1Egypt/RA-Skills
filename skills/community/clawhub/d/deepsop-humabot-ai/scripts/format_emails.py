#!/usr/bin/env python3
"""
format_emails.py
将 Frank 返回的邮件任务 JSON 数据生成 xlsx 文件
用法: python3 format_emails.py '<json_string>' '<output_path>'
输出: xlsx 文件路径
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# emailStatus 映射
STATUS_MAP = {
    0: '未发送',
    1: '发送失败',
    2: '发送成功',
}

def main():
    if len(sys.argv) < 3:
        print('用法: python3 format_emails.py \'<json_string>\' \'<output_path>\'', file=sys.stderr)
        sys.exit(1)

    # 解析 JSON
    try:
        raw = json.loads(sys.argv[1])
        # 支持 {rows: [...]} 或直接数组
        if isinstance(raw, dict):
            data = raw.get('rows', raw.get('data', []))
        else:
            data = raw
    except json.JSONDecodeError as e:
        print(f'JSON 解析失败: {e}', file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, list) or len(data) == 0:
        print('无邮件数据', file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[2]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 表头定义: (显示名, 字段key或特殊标记, 列宽)
    headers = [
        ('序号',       'index',                  8),
        ('收件人邮箱', 'recipientEmailAddress',  30),
        ('公司名称',   'companyName',            36),
        ('公司官网',   'companyUrl',             28),
        ('公司电话',   'companyPhone',           16),
        ('公司规模',   'companySize',            12),
        ('公司营收',   'companyRevenue',         14),
        ('公司简介',   'companyIntroduction',    50),
        ('联系人',     'personName',             14),
        ('职位',       'position',               14),
        ('联系电话',   'phone',                  16),
        ('WhatsApp',   'whatsapp',               18),
        ('LinkedIn',   'linkedin',               24),
        ('邮件主题',   'emailSubject',           40),
        ('发送状态',   'emailStatus',            12),
        ('发送时间',   'sendTime',               20),
        ('轮次',       'round',                   8),
        ('错误信息',   'errMsg',                 30),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Frank邮件数据'

    # 表头样式（橙色主题区别于 AiWa 蓝色）
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C05621', end_color='C05621', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (label, key, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # 状态颜色
    fill_success = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
    fill_fail    = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
    fill_pending = PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid')
    alt_fill     = PatternFill(start_color='FFF5EB', end_color='FFF5EB', fill_type='solid')
    data_align   = Alignment(vertical='center', wrap_text=False)

    for row_idx, item in enumerate(data, start=2):
        email_status = item.get('emailStatus')

        for col_idx, (label, key, width) in enumerate(headers, start=1):
            if key == 'index':
                value = row_idx - 1
            elif key == 'emailStatus':
                value = STATUS_MAP.get(email_status, str(email_status) if email_status is not None else '')
            elif key == 'personName':
                # 优先取顶层，再取 taskCustomer
                value = item.get('personName') or (item.get('taskCustomer') or {}).get('personName') or ''
            elif key == 'position':
                value = item.get('position') or (item.get('taskCustomer') or {}).get('position') or ''
            elif key == 'companyName':
                value = item.get('companyName') or (item.get('taskCustomer') or {}).get('companyName') or ''
            elif key in ('companyUrl', 'companyPhone', 'companySize', 'companyRevenue', 'companyIntroduction'):
                value = item.get(key) or (item.get('taskCustomer') or {}).get(key) or ''
            elif key in ('phone', 'whatsapp', 'linkedin'):
                value = item.get(key) or (item.get('taskCustomer') or {}).get(key) or ''
            else:
                value = item.get(key)
                if value is None:
                    value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align

            # 按状态着色整行
            if email_status == 2:
                cell.fill = fill_success
            elif email_status == 1:
                cell.fill = fill_fail
            elif email_status == 0:
                cell.fill = fill_pending
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 20

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(output_path)

if __name__ == '__main__':
    main()
